from __future__ import annotations

import os
import re
import math
import hashlib
import tempfile
import unicodedata
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROTOCOL_VERSION = 1
_ruled_grid_request_cache: dict[
    tuple[tuple[int, ...], str, bool, bytes],
    tuple[tuple[int, ...], tuple[int, ...], np.ndarray] | None,
] | None = None


def begin_ruled_grid_request_cache() -> None:
    global _ruled_grid_request_cache
    _ruled_grid_request_cache = {}


def end_ruled_grid_request_cache() -> None:
    global _ruled_grid_request_cache
    _ruled_grid_request_cache = None


def _display_width(value: str) -> int:
    return max(
        (sum(2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1 for character in line)
         for line in str(value).splitlines() or [""]),
        default=0,
    )


def _ordered_corners(points: np.ndarray) -> np.ndarray:
    points = np.asarray(points, dtype=np.float32).reshape(4, 2)
    ordered = np.zeros((4, 2), dtype=np.float32)
    coordinate_sum = points.sum(axis=1)
    coordinate_difference = np.diff(points, axis=1).reshape(-1)
    ordered[0] = points[np.argmin(coordinate_sum)]
    ordered[2] = points[np.argmax(coordinate_sum)]
    ordered[1] = points[np.argmin(coordinate_difference)]
    ordered[3] = points[np.argmax(coordinate_difference)]
    return ordered


def _refine_vertical_table_sides(image: np.ndarray, corners: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
    edges = cv2.Canny(gray, 30, 100)
    height, width = gray.shape[:2]
    lines = cv2.HoughLinesP(
        edges,
        1,
        np.pi / 720,
        threshold=max(50, height // 10),
        minLineLength=max(100, int(height * 0.5)),
        maxLineGap=max(20, height // 30),
    )
    if lines is None:
        return corners

    top_y = float((corners[0, 1] + corners[1, 1]) / 2.0)
    bottom_y = float((corners[2, 1] + corners[3, 1]) / 2.0)
    side_lines: list[tuple[float, float, float]] = []
    for x1, y1, x2, y2 in lines.reshape(-1, 4):
        if y1 == y2:
            continue
        angle = abs(float(np.degrees(np.arctan2(y2 - y1, x2 - x1))))
        length = float(np.hypot(x2 - x1, y2 - y1))
        if angle < 75 or length < height * 0.5:
            continue
        slope = (x2 - x1) / float(y2 - y1)
        top_x = x1 + (top_y - y1) * slope
        bottom_x = x1 + (bottom_y - y1) * slope
        side_lines.append((top_x, bottom_x, length))
    if not side_lines:
        return corners

    refined = corners.copy()
    center_x = width / 2.0
    top_span = abs(float(corners[1, 0] - corners[0, 0]))
    bottom_span = abs(float(corners[2, 0] - corners[3, 0]))
    maximum_inward_shift = max(8.0, min(top_span, bottom_span) * 0.020)
    left_lines = [line for line in side_lines if (line[0] + line[1]) / 2.0 < center_x]
    right_lines = [line for line in side_lines if (line[0] + line[1]) / 2.0 >= center_x]
    if left_lines:
        top_x, bottom_x, _ = min(left_lines, key=lambda line: abs(line[0] - corners[0, 0]))
        inward_shift = max(top_x - corners[0, 0], bottom_x - corners[3, 0])
        if abs(top_x - corners[0, 0]) <= width * 0.08 and inward_shift <= maximum_inward_shift:
            refined[0, 0] = top_x
            refined[3, 0] = bottom_x
    if right_lines:
        top_x, bottom_x, _ = min(right_lines, key=lambda line: abs(line[0] - corners[1, 0]))
        inward_shift = max(corners[1, 0] - top_x, corners[2, 0] - bottom_x)
        if abs(top_x - corners[1, 0]) <= width * 0.08 and inward_shift <= maximum_inward_shift:
            refined[1, 0] = top_x
            refined[2, 0] = bottom_x
    return refined


def _cluster_low_contrast_horizontal_rules(
    image: np.ndarray,
) -> list[list[dict[str, float]]]:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
    height, width = gray.shape[:2]
    normalized = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8)).apply(gray)
    binary = cv2.adaptiveThreshold(
        normalized,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        31,
        9,
    )
    horizontal = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (max(32, width // 30), 1)),
    )
    detected = cv2.HoughLinesP(
        horizontal,
        1,
        np.pi / 720,
        threshold=max(60, width // 16),
        minLineLength=max(80, int(round(width * 0.25))),
        maxLineGap=max(12, int(round(width * 0.06))),
    )
    if detected is None:
        return []

    segments: list[dict[str, float]] = []
    for x1, y1, x2, y2 in detected.reshape(-1, 4):
        if x1 > x2:
            x1, y1, x2, y2 = x2, y2, x1, y1
        slope = (float(y2) - float(y1)) / max(1.0, float(x2) - float(x1))
        angle = math.degrees(math.atan(slope))
        center_y = float(y1) + (width * 0.5 - float(x1)) * slope
        if abs(angle) > 6.0 or not height * 0.08 <= center_y <= height * 0.92:
            continue
        segments.append(
            {
                "x1": float(x1),
                "y1": float(y1),
                "x2": float(x2),
                "y2": float(y2),
                "slope": slope,
                "center_y": center_y,
                "length": float(math.hypot(x2 - x1, y2 - y1)),
            }
        )
    segments.sort(key=lambda item: item["center_y"])
    groups: list[list[dict[str, float]]] = []
    for segment in segments:
        if not groups:
            groups.append([segment])
            continue
        center = float(np.median([item["center_y"] for item in groups[-1]]))
        if abs(segment["center_y"] - center) <= 5.0:
            groups[-1].append(segment)
        else:
            groups.append([segment])
    groups = [
        group
        for group in groups
        if max(item["x2"] for item in group) - min(item["x1"] for item in group)
        >= width * 0.25
    ]
    if len(groups) < 8:
        return []
    centers = np.asarray(
        [np.median([item["center_y"] for item in group]) for group in groups],
        dtype=float,
    )
    gaps = np.diff(centers)
    ordinary = gaps[(gaps >= 12.0) & (gaps <= 70.0)]
    if ordinary.size < 6:
        return []
    typical = float(np.median(ordinary))
    regularity = float(np.mean((gaps >= typical * 0.62) & (gaps <= typical * 1.42)))
    return groups if regularity >= 0.80 else []


def _low_contrast_vertical_side(
    image: np.ndarray,
    groups: list[list[dict[str, float]]],
    *,
    right: bool,
) -> dict[str, float] | None:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
    height, width = gray.shape[:2]
    normalized = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8)).apply(gray)
    binary = cv2.adaptiveThreshold(
        normalized,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        31,
        9,
    )
    vertical = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(24, height // 30))),
    )
    detected = cv2.HoughLinesP(
        vertical,
        1,
        np.pi / 720,
        threshold=max(40, height // 12),
        minLineLength=max(60, int(round(height * 0.22))),
        maxLineGap=max(12, int(round(height * 0.08))),
    )
    if detected is None:
        return None
    centers = np.asarray(
        [np.median([item["center_y"] for item in group]) for group in groups],
        dtype=float,
    )
    endpoints = np.asarray(
        [
            max(item["x2"] for item in group)
            if right
            else min(item["x1"] for item in group)
            for group in groups
        ],
        dtype=float,
    )
    candidates: list[dict[str, float]] = []
    for x1, y1, x2, y2 in detected.reshape(-1, 4):
        if y1 > y2:
            x1, y1, x2, y2 = x2, y2, x1, y1
        delta_y = float(y2) - float(y1)
        if delta_y <= 0:
            continue
        angle = abs(math.degrees(math.atan2(delta_y, float(x2) - float(x1))))
        if not 75.0 <= angle <= 105.0:
            continue
        slope = (float(x2) - float(x1)) / delta_y
        intercept = float(x1) - slope * float(y1)
        middle_x = slope * float(np.median(centers)) + intercept
        if (right and middle_x <= width * 0.60) or (not right and middle_x >= width * 0.40):
            continue
        residual = float(np.median(np.abs(slope * centers + intercept - endpoints)))
        length = float(math.hypot(x2 - x1, y2 - y1))
        candidates.append(
            {
                "slope": slope,
                "intercept": intercept,
                "score": residual - min(length, height) * 0.01,
            }
        )
    return min(candidates, key=lambda item: item["score"]) if candidates else None


def _detect_low_contrast_ruled_quad(image: np.ndarray) -> np.ndarray | None:
    groups = _cluster_low_contrast_horizontal_rules(image)
    if not groups:
        return None
    left = _low_contrast_vertical_side(image, groups, right=False)
    right = _low_contrast_vertical_side(image, groups, right=True)
    if left is None or right is None:
        return None

    def row_model(group: list[dict[str, float]]) -> tuple[float, float]:
        segment = max(group, key=lambda item: item["length"])
        return segment["slope"], segment["y1"] - segment["slope"] * segment["x1"]

    def intersection(side: dict[str, float], row: tuple[float, float]) -> list[float]:
        row_slope, row_intercept = row
        y = (
            row_slope * side["intercept"] + row_intercept
        ) / (1.0 - row_slope * side["slope"])
        return [side["slope"] * y + side["intercept"], y]

    top = row_model(groups[0])
    bottom = row_model(groups[-1])
    quad = np.asarray(
        [
            intersection(left, top),
            intersection(right, top),
            intersection(right, bottom),
            intersection(left, bottom),
        ],
        dtype=np.float32,
    )
    if not cv2.isContourConvex(quad.astype(np.int32)):
        return None
    if abs(float(cv2.contourArea(quad))) < image.shape[0] * image.shape[1] * 0.20:
        return None
    return quad


def _expand_one_proven_trailing_grid_row(
    image: np.ndarray,
    table_corners: np.ndarray,
    document_corners: np.ndarray | None,
) -> np.ndarray | None:
    """Extend a premature bottom crop only when another complete row is physical."""
    if document_corners is None or image.size == 0:
        return None
    table = _ordered_corners(table_corners)
    document = _ordered_corners(document_corners)
    table_area = abs(float(cv2.contourArea(table)))
    document_area = abs(float(cv2.contourArea(document)))
    if (
        table_area <= 0.0
        or document_area < table_area * 1.02
        or document_area > table_area * 1.40
    ):
        return None

    def warp(corners: np.ndarray) -> np.ndarray | None:
        top_left, top_right, bottom_right, bottom_left = corners
        width = int(
            round(
                max(
                    np.linalg.norm(top_right - top_left),
                    np.linalg.norm(bottom_right - bottom_left),
                )
            )
        )
        height = int(
            round(
                max(
                    np.linalg.norm(bottom_left - top_left),
                    np.linalg.norm(bottom_right - top_right),
                )
            )
        )
        if width < 160 or height < 160:
            return None
        destination = np.asarray(
            [[0, 0], [width - 1, 0], [width - 1, height - 1], [0, height - 1]],
            dtype=np.float32,
        )
        return cv2.warpPerspective(
            image,
            cv2.getPerspectiveTransform(corners, destination),
            (width, height),
            flags=cv2.INTER_CUBIC,
            borderValue=(255, 255, 255),
        )

    preview = warp(table)
    if preview is None:
        return None
    preview_grid = extract_ruled_grid(preview, prefer_adaptive=True)
    if preview_grid is None:
        return None
    preview_columns, preview_rows, _ = preview_grid
    if len(preview_columns) < 5 or len(preview_rows) < 12:
        return None
    preview_height, preview_width = preview.shape[:2]
    edge_tolerance = max(4, int(round(preview_height * 0.01)))
    if (
        preview_rows[-1] < preview_height - 1 - edge_tolerance
        or preview_columns[0] > preview_width * 0.04
        or preview_columns[-1] < preview_width * 0.96
    ):
        return None
    recent_gaps = np.diff(preview_rows).astype(float)[-12:]
    typical_gap = float(np.median(recent_gaps)) if recent_gaps.size else 0.0
    factor = typical_gap / max(1.0, float(preview_height - 1))
    if typical_gap < 12.0 or not 0.012 <= factor <= 0.080:
        return None

    expanded = table.copy()
    expanded[2] = table[2] + (table[2] - table[1]) * factor
    expanded[3] = table[3] + (table[3] - table[0]) * factor
    image_height, image_width = image.shape[:2]
    containment_tolerance = max(8.0, min(image_height, image_width) * 0.006)
    document_polygon = document.astype(np.float32)
    for point in expanded[2:]:
        if (
            point[0] < 0
            or point[0] > image_width - 1
            or point[1] < 0
            or point[1] > image_height - 1
            or cv2.pointPolygonTest(
                document_polygon,
                (float(point[0]), float(point[1])),
                True,
            )
            < -containment_tolerance
        ):
            return None

    expanded_preview = warp(expanded)
    if expanded_preview is None:
        return None
    expanded_grid = extract_ruled_grid(expanded_preview, prefer_adaptive=True)
    if expanded_grid is None:
        return None
    expanded_columns, expanded_rows, _ = expanded_grid
    column_tolerance = max(8.0, expanded_preview.shape[1] * 0.01)
    same_columns = bool(
        len(expanded_columns) == len(preview_columns)
        and all(
            abs(float(left) - float(right)) <= column_tolerance
            for left, right in zip(preview_columns, expanded_columns)
        )
    )
    leading_outer_column_recovered = bool(
        len(expanded_columns) == len(preview_columns) + 1
        and expanded_columns[0] <= expanded_preview.shape[1] * 0.01
        and all(
            abs(float(left) - float(right)) <= column_tolerance
            for left, right in zip(preview_columns, expanded_columns[1:])
        )
    )
    trailing_outer_column_recovered = bool(
        len(expanded_columns) == len(preview_columns) + 1
        and expanded_columns[-1] >= expanded_preview.shape[1] * 0.99
        and all(
            abs(float(left) - float(right)) <= column_tolerance
            for left, right in zip(preview_columns, expanded_columns[:-1])
        )
    )
    if (
        not (
            same_columns
            or leading_outer_column_recovered
            or trailing_outer_column_recovered
        )
        or len(expanded_rows) != len(preview_rows) + 1
    ):
        return None
    expanded_height = expanded_preview.shape[0]
    if expanded_rows[-1] < expanded_height - 1 - max(4, int(expanded_height * 0.01)):
        return None
    ordinary_gaps = np.diff(expanded_rows).astype(float)[:-1]
    ordinary_gap = float(np.median(ordinary_gaps[-12:])) if ordinary_gaps.size else 0.0
    trailing_gap = float(expanded_rows[-1] - expanded_rows[-2])
    if (
        ordinary_gap < 12.0
        or not 0.65 * ordinary_gap <= trailing_gap <= 1.40 * ordinary_gap
    ):
        return None

    horizontal, vertical, _ = _grid_maps(expanded_preview)
    row_top = int(expanded_rows[-2])
    row_bottom = int(expanded_rows[-1])
    vertical_support = []
    for column in expanded_columns[1:-1]:
        band = vertical[
            row_top:row_bottom,
            max(0, column - 3) : min(vertical.shape[1], column + 4),
        ]
        vertical_support.append(
            float(np.mean(np.any(band > 0, axis=1))) if band.size else 0.0
        )
    required_columns = max(2, int(math.ceil(len(vertical_support) * 0.75)))
    if sum(support >= 0.65 for support in vertical_support) < required_columns:
        return None
    boundary_band = horizontal[
        max(0, row_top - 3) : min(horizontal.shape[0], row_top + 4),
        expanded_columns[0] : expanded_columns[-1],
    ]
    horizontal_support = (
        float(np.mean(np.any(boundary_band > 0, axis=0)))
        if boundary_band.size
        else 0.0
    )
    if horizontal_support < 0.65:
        return None
    return expanded


def _warp_perspective_table(
    image: np.ndarray,
    *,
    expand_to_document: bool = True,
) -> tuple[np.ndarray, np.ndarray | None, np.ndarray | None, bool]:
    horizontal, vertical, grid = _grid_maps(image)
    height, width = image.shape[:2]
    image_area = float(height * width)
    contours, _ = cv2.findContours(grid, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    best_corners: np.ndarray | None = None
    best_score = 0.0
    for contour in contours:
        hull = cv2.convexHull(contour)
        perimeter = cv2.arcLength(hull, True)
        approximation = cv2.approxPolyDP(hull, 0.02 * perimeter, True)
        if len(approximation) != 4 or not cv2.isContourConvex(approximation):
            continue
        corners = _ordered_corners(approximation.reshape(4, 2))
        area = abs(float(cv2.contourArea(corners)))
        if area < image_area * 0.025:
            continue
        x, y, candidate_width, candidate_height = cv2.boundingRect(corners.astype(np.int32))
        aspect = max(candidate_width, candidate_height) / max(1.0, min(candidate_width, candidate_height))
        horizontal_density = cv2.countNonZero(
            horizontal[y : y + candidate_height, x : x + candidate_width]
        ) / max(1.0, float(candidate_width * candidate_height))
        vertical_density = cv2.countNonZero(
            vertical[y : y + candidate_height, x : x + candidate_width]
        ) / max(1.0, float(candidate_width * candidate_height))
        score = _table_candidate_score(area, aspect, horizontal_density, vertical_density)
        if score > best_score:
            best_score = score
            best_corners = corners

    document_corners = _detect_light_document_corners(image) if expand_to_document else None
    low_contrast_table = _detect_low_contrast_ruled_quad(image)
    document_confirms_grid_corners = False
    if best_corners is not None and document_corners is not None:
        ordered_grid = _ordered_corners(best_corners)
        ordered_document = _ordered_corners(document_corners)
        grid_area = abs(float(cv2.contourArea(ordered_grid)))
        document_area = abs(float(cv2.contourArea(ordered_document)))
        maximum_corner_delta = max(
            float(np.linalg.norm(grid_corner - document_corner))
            for grid_corner, document_corner in zip(
                ordered_grid,
                ordered_document,
            )
        )
        document_confirms_grid_corners = bool(
            grid_area > 0.0
            and 0.97 <= document_area / grid_area <= 1.03
            and maximum_corner_delta
            <= max(10.0, min(height, width) * 0.012)
        )
    if low_contrast_table is not None and (
        best_corners is None
        or (
            not (
                document_confirms_grid_corners
                and abs(float(cv2.contourArea(low_contrast_table)))
                >= abs(float(cv2.contourArea(best_corners))) * 0.97
            )
            and abs(float(cv2.contourArea(low_contrast_table)))
            >= abs(float(cv2.contourArea(best_corners))) * 0.85
        )
    ):
        best_corners = low_contrast_table
    used_document_corners = False
    if best_corners is None:
        # Borderless Excel printouts have no ruled contour to seed the normal
        # table candidate path.  The visible sheet is still a reliable crop:
        # rectify it first so page OCR sees large, horizontal text instead of a
        # small skewed block surrounded by desk texture.
        if document_corners is None:
            return image, None, None, False
        best_corners = document_corners
        used_document_corners = True
    elif document_corners is not None:
        # Photographing a spreadsheet often leaves the outer grid line faint or
        # interrupted by a shadow.  A grid-only contour can then end one column
        # early.  If a larger bright sheet fully encloses that contour, rectify
        # the whole sheet and locate the grid afterwards instead of discarding
        # data at the paper edge.
        table_area = abs(float(cv2.contourArea(best_corners)))
        document_area = abs(float(cv2.contourArea(document_corners)))
        table_x, table_y, table_w, table_h = cv2.boundingRect(best_corners.astype(np.int32))
        document_x, document_y, document_w, document_h = cv2.boundingRect(
            document_corners.astype(np.int32)
        )
        contains_table = (
            document_x <= table_x + max(8, table_w * 0.03)
            and document_y <= table_y + max(8, table_h * 0.03)
            and document_x + document_w >= table_x + table_w - max(8, table_w * 0.03)
            and document_y + document_h >= table_y + table_h - max(8, table_h * 0.03)
        )
        contains_document = (
            table_x <= document_x + max(8, document_w * 0.03)
            and table_y <= document_y + max(8, document_h * 0.03)
            and table_x + table_w >= document_x + document_w - max(8, document_w * 0.03)
            and table_y + table_h >= document_y + document_h - max(8, document_h * 0.03)
        )
        nested_complete_sheet = bool(
            contains_document
            and table_area >= document_area * 2.5
        )
        document_to_table_ratio = document_area / max(1.0, table_area)
        near_matching_complete_sheet = bool(
            contains_table
            and 0.95 <= document_to_table_ratio <= 1.10
            and (
                document_x < table_x - max(12, int(round(table_w * 0.029)))
                or document_y < table_y - max(12, int(round(table_h * 0.029)))
                or document_x + document_w
                > table_x + table_w + max(12, int(round(table_w * 0.029)))
                or document_y + document_h
                > table_y + table_h + max(12, int(round(table_h * 0.029)))
            )
        )
        substantially_larger_complete_sheet = bool(
            contains_table and document_to_table_ratio >= 1.50
        )
        if nested_complete_sheet or (
            low_contrast_table is None
            and contains_table
            and document_area >= table_area * 1.12
        ) or near_matching_complete_sheet or substantially_larger_complete_sheet:
            best_corners = document_corners
            used_document_corners = True

    # A screenshot may start or end exactly at the table frame.  An incomplete
    # outer contour can then look like a valid axis-aligned quadrilateral that
    # touches three image sides while ending one regular row/column early.  A
    # perspective warp would irreversibly discard the still-visible edge band;
    # keep the source extent and let clipped-frame recovery restore its missing
    # boundary from continuing row/column rules.
    ordered = _ordered_corners(best_corners)
    top_left, top_right, bottom_right, bottom_left = ordered
    edge_slack_x = max(4.0, width * 0.004)
    edge_slack_y = max(4.0, height * 0.004)
    touched_sides = sum(
        (
            min(top_left[0], bottom_left[0]) <= edge_slack_x,
            max(top_right[0], bottom_right[0]) >= width - 1 - edge_slack_x,
            min(top_left[1], top_right[1]) <= edge_slack_y,
            max(bottom_left[1], bottom_right[1]) >= height - 1 - edge_slack_y,
        )
    )
    axis_aligned = bool(
        abs(float(top_right[1] - top_left[1])) <= edge_slack_y
        and abs(float(bottom_right[1] - bottom_left[1])) <= edge_slack_y
        and abs(float(bottom_left[0] - top_left[0])) <= edge_slack_x
        and abs(float(bottom_right[0] - top_right[0])) <= edge_slack_x
    )
    candidate_area = abs(float(cv2.contourArea(ordered)))
    if touched_sides >= 3 and axis_aligned and candidate_area >= image_area * 0.70:
        return image, None, None, False

    unrefined_corners = best_corners.copy()
    best_corners = _refine_vertical_table_sides(image, best_corners)
    if not used_document_corners:
        expanded_corners = _expand_one_proven_trailing_grid_row(
            image,
            best_corners,
            document_corners,
        )
        if expanded_corners is None and not np.array_equal(
            best_corners,
            unrefined_corners,
        ):
            expanded_corners = _expand_one_proven_trailing_grid_row(
                image,
                unrefined_corners,
                document_corners,
            )
        if expanded_corners is not None:
            best_corners = expanded_corners

    top_left, top_right, bottom_right, bottom_left = best_corners
    target_width = int(
        round(max(np.linalg.norm(top_right - top_left), np.linalg.norm(bottom_right - bottom_left)))
    )
    target_height = int(
        round(max(np.linalg.norm(bottom_left - top_left), np.linalg.norm(bottom_right - top_right)))
    )
    # Camera photos may contain a portrait sheet with a landscape table on it.
    # Reject only unusably small candidates; forcing a landscape crop discards
    # the real paper and can make a shadow edge look like a very wide table.
    if target_width < 80 or target_height < 50:
        return image, None, None, False
    destination = np.array(
        [[0, 0], [target_width - 1, 0], [target_width - 1, target_height - 1], [0, target_height - 1]],
        dtype=np.float32,
    )
    transform = cv2.getPerspectiveTransform(best_corners, destination)
    warped = cv2.warpPerspective(
        image,
        transform,
        (target_width, target_height),
        flags=cv2.INTER_CUBIC,
        borderValue=(255, 255, 255),
    )
    return warped, transform, best_corners, used_document_corners


def _detect_light_document_corners(image: np.ndarray) -> np.ndarray | None:
    """Return a large, low-texture paper rectangle when a photo contains one.

    This intentionally does not participate in screenshot handling.  It is a
    conservative expansion candidate for the grid contour used by photographed
    sheets, where an incomplete outer line must never crop a real data column.
    """
    height, width = image.shape[:2]
    if height < 80 or width < 80:
        return None
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image.copy()
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    _, mask = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    close_size = max(21, min(height, width) // 24)
    if close_size % 2 == 0:
        close_size += 1
    mask = cv2.morphologyEx(
        mask,
        cv2.MORPH_CLOSE,
        cv2.getStructuringElement(cv2.MORPH_RECT, (close_size, close_size)),
        iterations=2,
    )
    mask = cv2.morphologyEx(
        mask,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (7, 7)),
        iterations=1,
    )
    document_masks = [mask]
    if image.ndim == 3:
        hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        saturation = hsv[:, :, 1]
        value = hsv[:, :, 2]
        value_floor = max(145.0, min(210.0, float(np.median(value)) + 25.0))
        low_saturation_mask = np.where(
            (saturation <= 40) & (value >= value_floor),
            255,
            0,
        ).astype(np.uint8)
        low_close_size = max(15, min(height, width) // 40)
        if low_close_size % 2 == 0:
            low_close_size += 1
        low_saturation_mask = cv2.morphologyEx(
            low_saturation_mask,
            cv2.MORPH_CLOSE,
            cv2.getStructuringElement(
                cv2.MORPH_RECT,
                (low_close_size, low_close_size),
            ),
            iterations=2,
        )
        low_saturation_mask = cv2.morphologyEx(
            low_saturation_mask,
            cv2.MORPH_OPEN,
            cv2.getStructuringElement(cv2.MORPH_RECT, (7, 7)),
            iterations=1,
        )
        document_masks.append(low_saturation_mask)
    # RETR_EXTERNAL treats a bright decorative frame as one solid document and
    # hides the real sheet nested inside it. Inspect all contour levels, then
    # require the candidate interior itself to be predominantly light.
    image_area = float(height * width)
    best: np.ndarray | None = None
    best_area = 0.0
    for support_mask in document_masks:
        contours, _ = cv2.findContours(
            support_mask,
            cv2.RETR_LIST,
            cv2.CHAIN_APPROX_SIMPLE,
        )
        for contour in contours:
            area = float(cv2.contourArea(contour))
            # A phone photo can place the full sheet far from the camera. Four
            # percent still leaves hundreds of pixels on the short side at the
            # supported 8 MP input size; the interior-lightness check below
            # keeps small bright clutter from becoming a document.
            if area < image_area * 0.04:
                continue
            rectangle = cv2.minAreaRect(contour)
            candidate = _ordered_corners(cv2.boxPoints(rectangle))
            candidate_area = abs(float(cv2.contourArea(candidate)))
            if candidate_area < image_area * 0.04:
                continue
            candidate_mask = np.zeros(support_mask.shape, dtype=np.uint8)
            cv2.fillConvexPoly(candidate_mask, candidate.astype(np.int32), 255)
            candidate_pixels = cv2.countNonZero(candidate_mask)
            light_pixels = cv2.countNonZero(
                cv2.bitwise_and(support_mask, candidate_mask)
            )
            if light_pixels / max(1.0, float(candidate_pixels)) < 0.55:
                continue
            # A uniformly light screenshot/background can be segmented as one
            # image-sized rectangle. It is not a photographed sheet and would
            # suppress the denser table-grid candidate below.
            if candidate_area > image_area * 0.94:
                continue
            candidate_width = max(
                float(np.linalg.norm(candidate[1] - candidate[0])),
                float(np.linalg.norm(candidate[2] - candidate[3])),
            )
            candidate_height = max(
                float(np.linalg.norm(candidate[3] - candidate[0])),
                float(np.linalg.norm(candidate[2] - candidate[1])),
            )
            minor_side = min(candidate_width, candidate_height)
            major_side = max(candidate_width, candidate_height)
            if minor_side / max(1.0, major_side) < 0.45:
                continue
            edge_margin = max(3.0, min(width, height) * 0.008)
            edge_sides: set[str] = set()
            if float(np.min(candidate[:, 0])) <= edge_margin:
                edge_sides.add("left")
            if float(np.max(candidate[:, 0])) >= width - 1 - edge_margin:
                edge_sides.add("right")
            if float(np.min(candidate[:, 1])) <= edge_margin:
                edge_sides.add("top")
            if float(np.max(candidate[:, 1])) >= height - 1 - edge_margin:
                edge_sides.add("bottom")
            # A true sheet detected on two source borders is already incomplete;
            # more often this is a bright wall/desk region merged with paper.
            if len(edge_sides) >= 2:
                continue
            if candidate_area > best_area:
                best = candidate
                best_area = candidate_area
    return best


def _enhance_for_ocr(image: np.ndarray) -> np.ndarray:
    def enhance_lightness(lightness: np.ndarray) -> np.ndarray:
        original_median = float(np.median(lightness))
        original_p90 = float(np.percentile(lightness, 90))
        if original_median < 105.0 or original_p90 < 145.0:
            # Lift genuinely underexposed captures before illumination
            # normalization.  A bounded gamma curve preserves relative stroke
            # darkness; generative super-resolution is deliberately avoided
            # because it can invent digits that were not present in the photo.
            reference = max(8.0, min(245.0, original_median)) / 255.0
            gamma = max(0.58, min(0.92, math.log(0.62) / math.log(reference)))
            lookup = np.array(
                [round(255.0 * ((value / 255.0) ** gamma)) for value in range(256)],
                dtype=np.uint8,
            )
            lightness = cv2.LUT(lightness, lookup)
        height, width = lightness.shape[:2]
        longest = max(height, width)
        scale = min(1.0, 320.0 / float(max(1, longest)))
        if scale < 1.0:
            small = cv2.resize(
                lightness,
                (max(32, int(round(width * scale))), max(32, int(round(height * scale)))),
                interpolation=cv2.INTER_AREA,
            )
        else:
            small = lightness
        illumination = cv2.GaussianBlur(small, (0, 0), sigmaX=13.0, sigmaY=13.0)
        if illumination.shape != lightness.shape:
            illumination = cv2.resize(
                illumination,
                (width, height),
                interpolation=cv2.INTER_LINEAR,
            )
        illumination = np.maximum(illumination, 12).astype(np.uint8)
        normalized = cv2.divide(lightness, illumination, scale=238)
        balanced = cv2.addWeighted(lightness, 0.28, normalized, 0.72, 0)
        balanced = cv2.createCLAHE(clipLimit=2.1, tileGridSize=(8, 8)).apply(balanced)
        softened = cv2.GaussianBlur(balanced, (0, 0), sigmaX=0.75)
        return cv2.addWeighted(balanced, 1.18, softened, -0.18, 0)

    if image.ndim == 2:
        return enhance_lightness(image)
    lab = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
    lightness, channel_a, channel_b = cv2.split(lab)
    lightness = enhance_lightness(lightness)
    return cv2.cvtColor(cv2.merge((lightness, channel_a, channel_b)), cv2.COLOR_LAB2BGR)


def assess_image_quality(image: np.ndarray) -> dict[str, Any]:
    """Measure capture defects without guessing missing document content."""
    if image is None or image.size == 0:
        raise ValueError("image is empty")
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image.copy()
    height, width = gray.shape[:2]
    scale = min(1.0, 1280.0 / float(max(height, width)))
    if scale < 1.0:
        gray = cv2.resize(
            gray,
            (max(32, int(round(width * scale))), max(32, int(round(height * scale)))),
            interpolation=cv2.INTER_AREA,
        )
    roi_height, roi_width = gray.shape[:2]
    margin_y = int(round(roi_height * 0.04))
    margin_x = int(round(roi_width * 0.04))
    roi = gray[
        margin_y : max(margin_y + 1, roi_height - margin_y),
        margin_x : max(margin_x + 1, roi_width - margin_x),
    ]
    p05, p10, median, p90, p95 = [float(value) for value in np.percentile(roi, [5, 10, 50, 90, 95])]
    sharpness = float(cv2.Laplacian(roi, cv2.CV_32F).var())
    dark_ratio = float(np.mean(roi <= 45))
    highlight_ratio = float(np.mean(roi >= 253))

    tile_medians = []
    for row in range(4):
        for column in range(4):
            top = row * roi.shape[0] // 4
            bottom = (row + 1) * roi.shape[0] // 4
            left = column * roi.shape[1] // 4
            right = (column + 1) * roi.shape[1] // 4
            tile = roi[top:bottom, left:right]
            if tile.size:
                tile_medians.append(float(np.median(tile)))
    illumination_span = max(tile_medians, default=median) - min(tile_medians, default=median)
    contrast_span = p90 - p10

    issues: list[str] = []
    if median < 82.0 or p90 < 118.0 or dark_ratio > 0.32:
        issues.append("dark")
    if contrast_span < 48.0:
        issues.append("low_contrast")
    if sharpness < 42.0:
        issues.append("blur")
    if illumination_span > 105.0:
        issues.append("uneven_lighting")
    if highlight_ratio > 0.12 and median < 230.0:
        issues.append("clipped_highlights")

    labels = {
        "dark": "光线偏暗",
        "low_contrast": "文字对比度偏低",
        "blur": "清晰度偏低",
        "uneven_lighting": "光照不均",
        "clipped_highlights": "局部反光过曝",
    }
    return {
        "width": int(width),
        "height": int(height),
        "brightness_median": round(median, 2),
        "contrast_span": round(contrast_span, 2),
        "sharpness": round(sharpness, 2),
        "dark_ratio": round(dark_ratio, 4),
        "highlight_ratio": round(highlight_ratio, 4),
        "illumination_span": round(illumination_span, 2),
        "issues": issues,
        "issue_labels": [labels[issue] for issue in issues],
        "needs_recapture": bool("blur" in issues or len(issues) >= 3),
    }


def _table_candidate_score(area: float, aspect: float, horizontal_density: float, vertical_density: float) -> float:
    if area <= 0 or aspect < 1.0 or horizontal_density <= 0 or vertical_density <= 0:
        return 0.0
    # A dense photographed table can be almost square after perspective
    # distortion (for example a 12-column by 25-row sheet).  Rejecting every
    # candidate below 1.15 aspect leaves the complete camera frame for OCR,
    # which wastes detector work on the desk/background and makes the text
    # effectively tiny.  Only admit a near-square candidate when both line
    # directions are unusually dense; broad page/frame contours do not meet
    # these thresholds.
    if aspect < 1.15 and (horizontal_density < 0.055 or vertical_density < 0.025):
        return 0.0
    return area * horizontal_density * vertical_density * min(max(aspect, 1.15), 4.0)


def _grid_maps(image: np.ndarray) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    height, width = image.shape[:2]
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image.copy()
    gray = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8)).apply(gray)
    binary = cv2.adaptiveThreshold(
        gray,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        31,
        11,
    )
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(18, width // 30), 1))
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(18, height // 30)))
    horizontal = cv2.morphologyEx(binary, cv2.MORPH_OPEN, horizontal_kernel)
    vertical = cv2.morphologyEx(binary, cv2.MORPH_OPEN, vertical_kernel)
    grid = cv2.bitwise_or(horizontal, vertical)
    grid = cv2.morphologyEx(
        grid,
        cv2.MORPH_CLOSE,
        cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5)),
        iterations=2,
    )
    return horizontal, vertical, grid


def _line_centers(projection: np.ndarray, minimum_length: int) -> list[int]:
    if projection.size == 0 or float(projection.max()) <= 0:
        return []
    threshold = max(float(minimum_length) * 0.35, float(projection.max()) * 0.45)
    indices = np.flatnonzero(projection >= threshold)
    if indices.size == 0:
        return []
    groups = np.split(indices, np.where(np.diff(indices) > 1)[0] + 1)
    return [int(round(float(group.mean()))) for group in groups if group.size]


def _repeated_transition_centers(gray: np.ndarray, *, rows: bool) -> list[int]:
    source = gray.astype(np.int16, copy=False)
    transitions = (
        np.abs(source[2:, :] - source[:-2, :])
        if rows
        else np.abs(source[:, 2:] - source[:, :-2])
    )
    cross_axis_length = gray.shape[1] if rows else gray.shape[0]
    axis_length = transitions.shape[0] if rows else transitions.shape[1]
    candidates: list[int] = []
    for index in range(axis_length):
        values = transitions[index, :] if rows else transitions[:, index]
        histogram = np.bincount(values, minlength=256)
        histogram[:3] = 0
        if int(histogram.max()) >= cross_axis_length * 0.2:
            candidates.append(index + 1)
    if not candidates:
        return []

    groups: list[list[int]] = []
    for position in candidates:
        if not groups or position - groups[-1][-1] > 6:
            groups.append([])
        groups[-1].append(position)
    centers = [int(round(float(np.mean(group)))) for group in groups]
    if len(centers) > 1 and centers[0] <= max(4, int(axis_length * 0.005)):
        centers.pop(0)
    return centers


def _column_median_and_upper_quartile(
    values: np.ndarray,
) -> tuple[np.ndarray, np.ndarray]:
    """Compute both column quantiles with one partition pass."""
    sample_count = values.shape[0]
    quantiles = (0.5, 0.75)
    indices = sorted(
        {
            int(math.floor((sample_count - 1) * quantile))
            for quantile in quantiles
        }
        | {
            int(math.ceil((sample_count - 1) * quantile))
            for quantile in quantiles
        }
    )
    partitioned = np.partition(values, indices, axis=0)

    def quantile_value(quantile: float) -> np.ndarray:
        index = (sample_count - 1) * quantile
        lower = int(math.floor(index))
        upper = int(math.ceil(index))
        if lower == upper:
            return partitioned[lower].astype(np.float64)
        weight = index - lower
        return (
            partitioned[lower].astype(np.float64) * (1.0 - weight)
            + partitioned[upper].astype(np.float64) * weight
        )

    return quantile_value(0.5), quantile_value(0.75)


def _vertical_edge_statistics(
    gray: np.ndarray,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    gradient = np.abs(cv2.Sobel(gray.astype(np.float32), cv2.CV_32F, 1, 0, ksize=3))
    gradient = cv2.dilate(gradient, np.ones((1, 5), dtype=np.uint8))
    median, upper_quartile = _column_median_and_upper_quartile(gradient)
    coverage = np.mean(gradient > 12, axis=0)
    return gradient, median, upper_quartile, coverage


def _consistent_vertical_edge_centers(
    gray: np.ndarray,
    minimum_median: float = 40.0,
    statistics: tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray] | None = None,
) -> list[int]:
    gradient, median, upper_quartile, coverage = (
        statistics if statistics is not None else _vertical_edge_statistics(gray)
    )
    consistency = median / (upper_quartile + 1.0)
    score = median * consistency * coverage
    valid = (median >= minimum_median) & (consistency >= 0.35) & (coverage >= 0.5)

    radius = 5
    peaks = [
        position
        for position in range(radius, gray.shape[1] - radius)
        if valid[position]
        and score[position] >= float(score[position - radius : position + radius + 1].max())
    ]
    minimum_distance = max(12, gray.shape[1] // 100)
    selected: list[int] = []
    for position in sorted(peaks, key=lambda item: float(score[item]), reverse=True):
        if all(abs(position - existing) >= minimum_distance for existing in selected):
            selected.append(position)
    return sorted(selected)


def _high_coverage_vertical_edge_centers(
    gray: np.ndarray,
    statistics: tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray] | None = None,
) -> list[int]:
    """Find faint rules that remain visible through most spreadsheet rows.

    Excel/WPS screenshots can use very low-contrast vertical rules whose edge
    strength varies with alternating row fills.  Their median-to-quartile
    consistency is therefore weak even though the edge covers nearly the full
    sheet height.  Repeated text edges do not normally reach this coverage.
    """
    gradient, median, _, coverage = (
        statistics if statistics is not None else _vertical_edge_statistics(gray)
    )
    score = median * coverage
    valid = (median >= 30.0) & (coverage >= 0.75)

    radius = 5
    peaks = [
        position
        for position in range(radius, gray.shape[1] - radius)
        if valid[position]
        and score[position] >= float(score[position - radius : position + radius + 1].max())
    ]
    minimum_distance = max(12, gray.shape[1] // 100)
    selected: list[int] = []
    for position in sorted(peaks, key=lambda item: float(score[item]), reverse=True):
        if all(abs(position - existing) >= minimum_distance for existing in selected):
            selected.append(position)
    return sorted(selected)


def _merge_nearby_centers(centers: list[int], maximum_gap: int = 6) -> list[int]:
    groups: list[list[int]] = []
    for position in sorted(centers):
        if not groups or position - groups[-1][-1] > maximum_gap:
            groups.append([])
        groups[-1].append(position)
    return [int(round(float(np.mean(group)))) for group in groups]


def _recover_dense_spreadsheet_columns(
    gray: np.ndarray,
    transition_columns: list[int],
    transition_rows: list[int],
    statistics: tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray] | None = None,
) -> list[int]:
    """Recover many faint column rules without guessing on ordinary images."""
    if len(transition_rows) < 20:
        return transition_columns
    row_gaps = np.diff(transition_rows)
    typical_row_gap = float(np.median(row_gaps)) if row_gaps.size else 0.0
    if typical_row_gap < 8:
        return transition_columns
    row_tolerance = max(3.0, typical_row_gap * 0.18)
    if float(np.mean(np.abs(row_gaps - typical_row_gap) <= row_tolerance)) < 0.85:
        return transition_columns

    low_contrast = _consistent_vertical_edge_centers(
        gray,
        minimum_median=10.0,
        statistics=statistics,
    )
    high_coverage = _high_coverage_vertical_edge_centers(
        gray,
        statistics=statistics,
    )
    recovered = _merge_nearby_centers(
        transition_columns + low_contrast + high_coverage,
        maximum_gap=6,
    )
    if len(recovered) < max(8, len(transition_columns) + 4):
        return transition_columns
    if len(recovered) > 64:
        return transition_columns

    minimum_gap = max(12, gray.shape[1] // 120)
    if min(np.diff(recovered), default=minimum_gap) < minimum_gap:
        return transition_columns
    if recovered[-1] - recovered[0] < gray.shape[1] * 0.85:
        return transition_columns
    return recovered


def _select_screen_columns(
    transition_columns: list[int],
    consistent_columns: list[int],
    width: int,
) -> list[int]:
    edge_columns = [
        column
        for column in transition_columns
        if column <= width * 0.04 or column >= width * 0.96
    ]
    dense_columns = _merge_nearby_centers(consistent_columns + edge_columns)
    if len(dense_columns) >= 8 and len(transition_columns) >= len(dense_columns) * 1.55:
        return dense_columns

    if len(dense_columns) >= 4 and len(transition_columns) > len(dense_columns):
        supported = sum(
            any(abs(column - transition) <= 8 for transition in transition_columns)
            for column in dense_columns
        )
        minimum_gap = max(12, width // 120)
        dense_gaps = np.diff(dense_columns)
        typical_gap = float(np.median(dense_gaps)) if dense_gaps.size else 0.0
        if (
            supported >= max(3, int(np.ceil(len(dense_columns) * 0.8)))
            and min(dense_gaps, default=minimum_gap) >= minimum_gap
            and max(dense_gaps, default=0) <= typical_gap * 1.65
        ):
            return dense_columns

    return transition_columns


def _filter_transition_columns_by_vertical_support(
    transition_columns: list[int],
    vertical_projection: np.ndarray,
    image_height: int,
    image_width: int,
) -> list[int]:
    supported = [
        column
        for column in transition_columns
        if column <= image_width * 0.04
        or column >= image_width * 0.96
        or int(
            vertical_projection[
                max(0, column - 3) : min(image_width, column + 4)
            ].max(initial=0)
        )
        >= image_height * 0.55
    ]
    minimum_supported = max(3, int(np.ceil(len(transition_columns) * 0.6)))
    return supported if len(supported) >= minimum_supported else transition_columns


def _remove_false_transition_rows(
    centers: list[int],
    horizontal: np.ndarray,
    image_width: int,
) -> list[int]:
    if len(centers) < 5:
        return centers
    recovered = list(centers)
    while len(recovered) >= 5:
        gaps = np.diff(recovered)
        substantial = gaps[gaps >= 10]
        typical = float(np.median(substantial)) if substantial.size else 0.0
        if typical <= 0:
            break
        candidate_index = next(
            (index for index, gap in enumerate(gaps) if gap < typical * 0.45),
            None,
        )
        if candidate_index is None:
            break
        left = recovered[candidate_index]
        right = recovered[candidate_index + 1]

        def support(position: int) -> int:
            top = max(0, position - 2)
            bottom = min(horizontal.shape[0], position + 3)
            return int(np.count_nonzero(horizontal[top:bottom], axis=1).max(initial=0))

        left_support = support(left)
        right_support = support(right)
        strongest = max(left_support, right_support)
        weakest = min(left_support, right_support)
        if strongest < image_width * 0.5 or weakest > strongest * 0.25:
            break
        recovered.pop(candidate_index if left_support < right_support else candidate_index + 1)
    return recovered


def _recover_merged_header_rows(rows: list[int]) -> list[int]:
    if not 8 <= len(rows) <= 20:
        return rows
    gaps = np.diff(rows)
    typical = float(np.median(gaps))
    recovered = list(rows)
    largest_index = int(np.argmax(gaps))
    largest_gap = float(gaps[largest_index])
    other_gaps = np.delete(gaps, largest_index)
    second_largest = float(other_gaps.max()) if other_gaps.size else 0.0
    if (
        typical * 2.2 < largest_gap < typical * 3.2
        and largest_gap > second_largest * 1.35
    ):
        recovered.insert(
            largest_index + 1,
            int(round((rows[largest_index] + rows[largest_index + 1]) / 2.0)),
        )
    return recovered


def _band_profile_vertical_centers(gray: np.ndarray, rows: list[int]) -> list[int]:
    if len(rows) < 8:
        return []
    bands = list(zip(rows[:-1], rows[1:]))[-12:]
    profiles = [
        np.median(gray[top + 2 : bottom - 2], axis=0)
        for top, bottom in bands
        if bottom - top >= 8
    ]
    if len(profiles) < 6:
        return []
    profile_array = np.asarray(profiles)
    transitions = np.abs(profile_array[:, 2:] - profile_array[:, :-2])
    median = np.median(transitions, axis=0)
    upper_quartile = np.percentile(transitions, 75, axis=0)
    score = median + upper_quartile * 0.2
    radius = 5
    peaks = [
        position + 1
        for position in range(radius, score.size - radius)
        if score[position] >= 3
        and score[position] >= float(score[position - radius : position + radius + 1].max())
    ]
    minimum_distance = max(20, gray.shape[1] // 60)
    selected: list[int] = []
    for position in sorted(peaks, key=lambda item: float(score[item - 1]), reverse=True):
        if all(abs(position - existing) >= minimum_distance for existing in selected):
            selected.append(position)
    return sorted(selected)


def _header_transition_columns(gray: np.ndarray, top: int, bottom: int) -> list[int]:
    if bottom - top < 12:
        return []
    header = gray[top:bottom].astype(np.int16, copy=False)
    transitions = np.abs(header[:, 2:] - header[:, :-2])
    candidates = _repeated_transition_centers(gray[top:bottom], rows=False)
    return [
        column
        for column in candidates
        if 1 <= column < gray.shape[1] - 1
        and float(np.mean(transitions[:, column - 1] > 3)) >= 0.4
    ]


def _confirmed_spreadsheet_ruler_columns(
    gray: np.ndarray,
    rows: list[int],
) -> list[int]:
    """Recover faint spreadsheet boundaries confirmed in ruler and body bands."""
    if len(rows) < 8:
        return []
    first_gap = rows[1] - rows[0]
    body_gaps = np.diff(rows[1:])
    typical_body_gap = float(np.median(body_gaps)) if body_gaps.size else 0.0
    if typical_body_gap < 8 or first_gap >= typical_body_gap * 0.85:
        return []
    ruler_candidates = _repeated_transition_centers(
        gray[rows[0] : rows[1]],
        rows=False,
    )
    body_candidates = _band_profile_vertical_centers(gray, rows)
    if len(ruler_candidates) < 3 or len(body_candidates) < 3:
        return []
    confirmed = [
        column
        for column in ruler_candidates
        if min(abs(column - body_column) for body_column in body_candidates) <= 6
    ]
    confirmed = _merge_nearby_centers(confirmed)
    if (
        len(confirmed) < 3
        or confirmed[-1] - confirmed[0] < gray.shape[1] * 0.85
    ):
        return []
    return confirmed


def _recover_spreadsheet_ruler_grid(
    gray: np.ndarray,
    columns: list[int],
    rows: list[int],
    *,
    require_change: bool = True,
    vertical_statistics: tuple[
        np.ndarray,
        np.ndarray,
        np.ndarray,
        np.ndarray,
    ]
    | None = None,
) -> tuple[list[int], list[int]] | None:
    """依据独立可见的电子表格标尺恢复淡化的网格结构。"""
    if gray.size == 0 or len(columns) < 3 or len(rows) < 8:
        return None
    height, width = gray.shape[:2]
    if rows[0] < 12:
        return None

    body_columns = _band_profile_vertical_centers(gray, rows)
    row_gaps = np.diff(rows).astype(float)
    body_row_gaps = np.diff(rows[1:]).astype(float)
    typical_body_gap = (
        float(np.median(body_row_gaps)) if body_row_gaps.size else 0.0
    )
    embedded_ruler = bool(
        row_gaps.size
        and typical_body_gap >= 8.0
        and row_gaps[0] < typical_body_gap * 0.72
    )
    ruler_band = (
        gray[rows[0] : rows[1]]
        if embedded_ruler
        else gray[: rows[0]]
    )
    ruler_columns = _repeated_transition_centers(ruler_band, rows=False)
    confirmed_columns = _merge_nearby_centers(
        [
            column
            for column in ruler_columns
            if body_columns
            and min(abs(column - body_column) for body_column in body_columns) <= 6
        ]
    )
    minimum_column_gap = max(12, int(round(width * 0.008)))
    # 密集截图的表体文字可能制造大量假竖线，body profile 本身也会受这些笔画干扰。
    # 此时只保留“普通网格提取结果”和“顶部 Excel 标尺”同时支持的边界；两路证据
    # 相互独立，既能剔除行号栏和文字竖画，也不会依据均匀列宽猜测真实边界。
    ruler_supported_grid_columns = _merge_nearby_centers(
        [
            column
            for column in columns
            if ruler_columns
            and min(abs(column - ruler_column) for ruler_column in ruler_columns) <= 6
        ]
    )
    persistent_vertical_edges = _consistent_vertical_edge_centers(
        gray,
        minimum_median=30.0,
        statistics=vertical_statistics,
    )
    persistent_ruler_supported_columns = _merge_nearby_centers(
        [
            column
            for column in ruler_supported_grid_columns
            if persistent_vertical_edges
            and min(abs(column - edge) for edge in persistent_vertical_edges) <= 8
        ]
    )
    if len(ruler_supported_grid_columns) >= 5:
        supported_gaps = np.diff(ruler_supported_grid_columns).astype(float)
        typical_supported_gap = (
            float(np.median(supported_gaps)) if supported_gaps.size else 0.0
        )
        trailing_gap = (width - 1) - ruler_supported_grid_columns[-1]
        if (
            typical_supported_gap >= minimum_column_gap
            and ruler_supported_grid_columns[0] <= width * 0.08
            and ruler_supported_grid_columns[-1] >= width * 0.72
            and typical_supported_gap * 0.55 <= trailing_gap <= typical_supported_gap * 1.85
        ):
            ruler_supported_grid_columns.append(width - 1)
        if len(confirmed_columns) < 5 and (
            ruler_supported_grid_columns[-1] - ruler_supported_grid_columns[0]
            >= width * 0.85
            and min(np.diff(ruler_supported_grid_columns), default=0)
            >= minimum_column_gap
        ):
            confirmed_columns = ruler_supported_grid_columns
    if len(persistent_ruler_supported_columns) >= 5:
        persistent_gaps = np.diff(persistent_ruler_supported_columns).astype(float)
        typical_persistent_gap = (
            float(np.median(persistent_gaps)) if persistent_gaps.size else 0.0
        )
        trailing_gap = (width - 1) - persistent_ruler_supported_columns[-1]
        if (
            typical_persistent_gap >= minimum_column_gap
            and persistent_ruler_supported_columns[0] <= width * 0.08
            and persistent_ruler_supported_columns[-1] >= width * 0.72
            and typical_persistent_gap * 0.55
            <= trailing_gap
            <= typical_persistent_gap * 1.85
        ):
            persistent_ruler_supported_columns.append(width - 1)
    if len(confirmed_columns) >= 5:
        confirmed_gaps = np.diff(confirmed_columns).astype(float)
        typical_column_gap = (
            float(np.median(confirmed_gaps)) if confirmed_gaps.size else 0.0
        )
        trailing_gap = (width - 1) - confirmed_columns[-1]
        if (
            typical_column_gap >= minimum_column_gap
            and confirmed_columns[-1] >= width * 0.72
            and typical_column_gap * 0.55 <= trailing_gap <= typical_column_gap * 1.85
        ):
            # 电子表格截图经常恰好裁在最后一列（通常较宽）的右边缘，因此转换检测
            # 看不到第二条边。只有标尺和表体已经独立给出此前全部边界时，才允许
            # 使用截图边缘闭合最后一列。
            confirmed_columns.append(width - 1)
    if (
        len(confirmed_columns) < 5
        or confirmed_columns[-1] - confirmed_columns[0] < width * 0.85
        or min(np.diff(confirmed_columns), default=0) < minimum_column_gap
    ):
        return None

    # 若表体交叉证据因密集文字产生少量假边界，则用“网格候选∩顶部标尺”作为
    # 保守替代；仅允许比交叉结果多极少数边界，防止重新引入大量文字竖画。
    if (
        len(ruler_supported_grid_columns) >= len(confirmed_columns)
        and len(ruler_supported_grid_columns) <= len(confirmed_columns) + 2
        and ruler_supported_grid_columns[0] <= width * 0.08
        and ruler_supported_grid_columns[-1] >= width * 0.95
        and min(np.diff(ruler_supported_grid_columns), default=0) >= minimum_column_gap
    ):
        confirmed_columns = ruler_supported_grid_columns
    elif (
        len(persistent_ruler_supported_columns) >= len(confirmed_columns)
        and persistent_ruler_supported_columns[0] <= width * 0.08
        and persistent_ruler_supported_columns[-1] >= width * 0.95
        and min(np.diff(persistent_ruler_supported_columns), default=0)
        >= minimum_column_gap
    ):
        # 密集小字会让最后几列在表体中只有偏移的投影峰，旧的“最多多两条”限制
        # 因而会把一个完整 A-V 网格降成 18 列。这里要求每条新增边界同时满足：
        # 初始网格、顶部标尺和近乎全高竖边三路独立证据。正文里的文字竖画不穿过
        # 顶部标尺，标尺字母笔画也不贯穿表体，因此不能进入这个保守候选集。
        confirmed_columns = persistent_ruler_supported_columns

    # 转置后对行边界复用同一种表体投影证据。候选必须限制在已检测表格范围内，
    # 避免把电子表格左上角或列字母标尺导出成数据行。
    ordinary_gap_values = np.diff(rows)
    ordinary_gap_values = ordinary_gap_values[ordinary_gap_values >= 8]
    ordinary_gap = (
        float(np.median(ordinary_gap_values))
        if ordinary_gap_values.size
        else 0.0
    )
    row_candidates = _band_profile_vertical_centers(gray.T, confirmed_columns)
    if (
        row_candidates
        and ordinary_gap > 0.0
        and 8 <= rows[-1] - row_candidates[-1] <= ordinary_gap * 1.25
    ):
        # 裁剪边缘之外没有像素，转换投影可能漏掉该边；若普通网格检测已经证明
        # 它存在，则允许补回这个边缘行边界。
        row_candidates.append(rows[-1])
    extent_tolerance = max(6, int(round(ordinary_gap * 0.18)))
    first_body_row = rows[1] if embedded_ruler else rows[0]
    confirmed_rows = _merge_nearby_centers(
        [
            row
            for row in row_candidates
            if first_body_row - extent_tolerance <= row <= rows[-1] + extent_tolerance
        ]
    )
    if len(confirmed_rows) < 8:
        return None
    # 截图黑边可能在工作表下方制造额外转换。这里只保留从正文表头开始的规则
    # 边界序列，不能把黑边导出成一个异常巨大的数据行。
    row_gaps = np.diff(confirmed_rows).astype(float)
    ordinary_candidates = row_gaps[
        row_gaps <= np.percentile(row_gaps, 70)
    ]
    run_gap = (
        float(np.median(ordinary_candidates))
        if ordinary_candidates.size
        else 0.0
    )
    if run_gap >= 8.0:
        regular_rows = [confirmed_rows[0]]
        for boundary in confirmed_rows[1:]:
            gap = boundary - regular_rows[-1]
            if not run_gap * 0.68 <= gap <= run_gap * 1.38:
                break
            regular_rows.append(boundary)
        if len(regular_rows) >= 8:
            confirmed_rows = regular_rows
    row_gaps = np.diff(confirmed_rows).astype(float)
    typical_row_gap = float(np.median(row_gaps)) if row_gaps.size else 0.0
    if typical_row_gap < 8.0:
        return None
    regular_ratio = float(
        np.mean(np.abs(row_gaps - typical_row_gap) <= max(6.0, typical_row_gap * 0.18))
    )
    if (
        regular_ratio < 0.72
        or abs(confirmed_rows[0] - first_body_row) > extent_tolerance
        or (
            not embedded_ruler
            and abs(confirmed_rows[-1] - rows[-1]) > extent_tolerance
        )
    ):
        return None
    # 普通网格提取器为了找回淡线会采用较宽松的条件，但在密集表格中也可能把
    # 重复文字竖画误判成贯穿整表的列线。标尺与表体证据独立于该结果；当它们覆盖
    # 同一张表且差异明显时，允许用独立证据替换过度分列的网格，这与补回缺线同样重要。
    removes_two_ruler_rejected_columns = bool(
        len(columns) == len(confirmed_columns) + 2
        and len(confirmed_columns) >= 8
        and all(
            any(abs(candidate - original) <= 8 for original in columns)
            for candidate in confirmed_columns
        )
        and sum(
            not any(abs(original - candidate) <= 8 for candidate in confirmed_columns)
            for original in columns
        )
        == 2
    )
    replaces_over_split_columns = bool(
        len(confirmed_columns) >= 5
        and (
            len(columns) >= len(confirmed_columns) + 3
            or removes_two_ruler_rejected_columns
        )
        # 原始候选的最左边可能是工作表外框或行号栏，不能要求修复后的首边界仍贴着它。
        # 顶部标尺必须独立覆盖近乎整个截图，才允许删除这些标尺不支持的假边界。
        and confirmed_columns[0] <= width * 0.08
        and confirmed_columns[-1] >= width * 0.95
    )
    adds_missing_boundaries = bool(
        len(confirmed_columns) > len(columns)
        or len(confirmed_rows) > len(rows)
    )
    if require_change and not replaces_over_split_columns and not adds_missing_boundaries:
        return None
    return confirmed_columns, confirmed_rows


def spreadsheet_ruler_confirms_columns(
    image: np.ndarray,
    columns: list[int],
    rows: list[int],
) -> bool:
    """使用独立行带证据确认已提取电子表格的精确列边界。

    ``extract_screen_grid`` 已使用上面的标尺恢复逻辑替换过度分列的网格。
    这里第二次校验只确认列：工作表标尺或标题附近的行转换可以不同，但不能因此
    把已经修正的列边界误判为不可信。本函数不会生成文字，也不会修改传入网格。
    """
    if image.size == 0 or len(columns) < 5 or len(rows) < 8:
        return False
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
    height, width = gray.shape[:2]
    if rows[0] < 1 or rows[-1] >= height:
        return False
    body_columns = _band_profile_vertical_centers(gray, rows)
    if len(body_columns) < 4:
        return False
    row_gaps = np.diff(rows).astype(float)
    body_row_gaps = np.diff(rows[1:]).astype(float)
    typical_body_gap = float(np.median(body_row_gaps)) if body_row_gaps.size else 0.0
    embedded_ruler = bool(
        row_gaps.size
        and typical_body_gap >= 8.0
        and row_gaps[0] < typical_body_gap * 0.72
    )
    ruler_band = gray[rows[0] : rows[1]] if embedded_ruler else gray[: rows[0]]
    ruler_columns = _repeated_transition_centers(ruler_band, rows=False)
    if not ruler_columns:
        return False
    minimum_gap = max(12, int(round(width * 0.008)))
    if (
        columns[-1] - columns[0] < width * 0.85
        or min(np.diff(columns), default=0) < minimum_gap
    ):
        return False
    ruler_tolerance = max(6, int(round(width * 0.004)))
    body_tolerance = max(8, int(round(width * 0.007)))
    ruler_supported = sum(
        min(abs(column - candidate) for candidate in ruler_columns) <= ruler_tolerance
        for column in columns[:-1]
    )
    body_supported = sum(
        min(abs(column - candidate) for candidate in body_columns) <= body_tolerance
        for column in columns[:-1]
    )
    right_edge_closed = bool(
        columns[-1] >= width - max(2, ruler_tolerance)
        or min(abs(columns[-1] - candidate) for candidate in ruler_columns)
        <= ruler_tolerance
    )
    return bool(
        right_edge_closed
        and ruler_supported >= len(columns) - 2
        and body_supported >= max(4, int(np.ceil((len(columns) - 1) * 0.75)))
    )


def extract_spreadsheet_ruler_columns(
    image: np.ndarray,
    expected_columns: int,
    band_bottom: int,
) -> list[int]:
    """Return document column bounds confirmed by an Excel/WPS ruler band.

    The OCR layer supplies the expected A..H count.  This helper supplies only
    geometry; it never creates cell text.  A cropped right edge is accepted as
    the final boundary only when every preceding ruler transition is present.
    """
    if expected_columns < 3 or expected_columns > 32 or image.size == 0:
        return []
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
    height, width = gray.shape[:2]
    bottom = min(height, max(12, int(band_bottom)))
    candidates = _repeated_transition_centers(gray[:bottom], rows=False)
    candidates = [
        int(column)
        for column in candidates
        if 0 <= int(column) < width - 1
    ]
    if len(candidates) == expected_columns:
        candidates.append(width - 1)
    elif len(candidates) != expected_columns + 1:
        return []
    if (
        candidates[0] > width * 0.12
        or candidates[-1] < width * 0.95
        or candidates[-1] - candidates[0] < width * 0.85
    ):
        return []
    minimum_gap = max(8, int(round(width * 0.008)))
    if min(np.diff(candidates), default=0) < minimum_gap:
        return []
    return candidates


def _recover_single_missing_boundary(
    centers: list[int],
    vertical_projection: np.ndarray | None = None,
) -> list[int]:
    if len(centers) < 5:
        return centers
    gaps = np.diff(centers)
    largest_index = int(np.argmax(gaps))
    largest_gap = float(gaps[largest_index])
    other_gaps = np.delete(gaps, largest_index)
    typical = float(np.median(other_gaps)) if other_gaps.size else 0.0
    if typical * 1.7 < largest_gap < typical * 2.4:
        recovered = list(centers)
        left = centers[largest_index]
        right = centers[largest_index + 1]
        recovered_position = int(round((left + right) / 2.0))
        if vertical_projection is not None:
            search_left = max(0, left + 12)
            search_right = min(vertical_projection.size, right - 11)
            search = vertical_projection[search_left:search_right]
            if search.size:
                peak_index = int(np.argmax(search))
                peak_value = int(search[peak_index])
                global_peak = int(vertical_projection.max(initial=0))
                if global_peak and peak_value >= global_peak * 0.45:
                    recovered_position = search_left + peak_index
                else:
                    return centers
        recovered.insert(largest_index + 1, recovered_position)
        return recovered
    return centers


def _recover_regular_missing_boundaries(
    centers: list[int],
    projection: np.ndarray,
    support_extent: int,
    *,
    maximum_multiple: int = 4,
    minimum_support_ratio: float = 0.14,
    minimum_global_peak_ratio: float = 0.15,
) -> list[int]:
    """Recover faint rules only when a regular gap and real line evidence agree.

    A photographed table can lose most of one horizontal or vertical rule under
    a shadow.  The missing rule must not make a header and the first data row
    share one cell.  Conversely, a deliberately merged cell must stay merged,
    so regular spacing alone is never sufficient: a local projection peak is
    required for every inserted boundary.
    """
    if len(centers) < 5 or projection.size == 0:
        return centers

    recovered = sorted(set(int(value) for value in centers))
    gaps = np.diff(recovered)
    if gaps.size < 4:
        return recovered
    ordinary = gaps[gaps <= np.percentile(gaps, 70)]
    typical = float(np.median(ordinary)) if ordinary.size else float(np.median(gaps))
    global_peak = int(projection.max(initial=0))
    if typical < 6 or global_peak <= 0:
        return recovered

    insertions: list[int] = []
    for left, right in zip(recovered[:-1], recovered[1:]):
        gap = right - left
        multiple = int(round(gap / typical))
        if (
            not 2 <= multiple <= maximum_multiple
            or abs(gap / multiple - typical) > typical * 0.22
        ):
            continue
        for part in range(1, multiple):
            predicted = int(round(left + gap * part / multiple))
            search_left = max(0, predicted - 4)
            search_right = min(projection.size, predicted + 5)
            local = projection[search_left:search_right]
            if local.size == 0:
                continue
            peak_index = int(np.argmax(local))
            peak_value = int(local[peak_index])
            # The absolute threshold rejects text strokes; the relative one
            # accepts a genuine rule weakened by illumination or a fold.
            if (
                peak_value >= support_extent * minimum_support_ratio
                and peak_value >= global_peak * minimum_global_peak_ratio
            ):
                insertions.append(search_left + peak_index)

    return _merge_nearby_centers(recovered + insertions, maximum_gap=5)


def _recover_visible_double_row_boundary(
    image: np.ndarray,
    rows: list[int],
) -> list[int]:
    """Restore one faint row rule only when spacing and a wide transition agree."""
    if len(rows) < 10 or image.size == 0:
        return rows
    recovered = sorted(set(int(value) for value in rows))
    gaps = np.diff(recovered)
    ordinary = gaps[gaps <= np.percentile(gaps, 70)]
    typical = float(np.median(ordinary)) if ordinary.size else 0.0
    if typical < 8:
        return recovered
    candidates = [
        index
        for index, gap in enumerate(gaps)
        if typical * 1.80 <= gap <= typical * 2.60
    ]
    if len(candidates) != 1:
        return recovered
    candidate_index = candidates[0]
    other_gaps = np.delete(gaps, candidate_index)
    if other_gaps.size and float(other_gaps.max()) > typical * 1.55:
        return recovered

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
    left = max(0, int(round(gray.shape[1] * 0.02)))
    right = min(gray.shape[1], int(round(gray.shape[1] * 0.98)))
    if right - left < 100 or gray.shape[0] < 3:
        return recovered
    transitions = np.abs(
        gray[2:, left:right].astype(np.int16)
        - gray[:-2, left:right].astype(np.int16)
    )
    coverage = np.mean(transitions > 8, axis=1)
    predicted = int(round((recovered[candidate_index] + recovered[candidate_index + 1]) / 2.0))
    radius = max(4, int(round(typical * 0.18)))
    search_left = max(0, predicted - radius - 1)
    search_right = min(coverage.size, predicted + radius)
    search = coverage[search_left:search_right]
    if search.size == 0:
        return recovered
    peak_offset = int(np.argmax(search))
    peak_coverage = float(search[peak_offset])
    if peak_coverage < 0.12:
        return recovered
    recovered.insert(candidate_index + 1, search_left + peak_offset + 1)
    return recovered


def _prefer_body_consensus_columns(
    columns: list[int],
    body_columns: list[int],
) -> list[int]:
    """Prefer the body-wide column grid only when it preserves the current evidence."""
    if not columns or not body_columns:
        return columns
    matched_existing = sum(
        any(abs(column - candidate) <= 7 for candidate in body_columns)
        for column in columns
    )
    if (
        len(body_columns) >= len(columns) + 2
        and matched_existing >= max(2, len(columns) - 2)
    ):
        return body_columns
    return columns


def _recover_supported_irregular_boundaries(
    centers: list[int],
    candidates: list[int],
    projection: np.ndarray,
    support_extent: int,
    *,
    vertical: np.ndarray | None = None,
    rows: list[int] | None = None,
) -> list[int]:
    """Restore a faint rule inside a genuinely wide, non-uniform column.

    Column widths in business tables are often intentionally different, so a
    missing boundary cannot be reconstructed from equal spacing.  A boundary
    is inserted only when it is a consistent full-height edge and the line map
    independently retains substantial support at the same position.
    """
    if len(centers) < 4 or not candidates or projection.size == 0:
        return centers
    recovered = sorted(set(int(value) for value in centers))
    gaps = np.diff(recovered)
    ordinary = gaps[gaps <= np.percentile(gaps, 70)]
    typical = float(np.median(ordinary)) if ordinary.size else float(np.median(gaps))
    global_peak = int(projection.max(initial=0))
    if typical < 8 or global_peak <= 0:
        return recovered

    insertions: list[int] = []
    edge_margin = max(10, int(round(typical * 0.12)))

    def follows_rows(candidate: int) -> bool:
        if (
            vertical is None
            or vertical.size == 0
            or rows is None
            or len(rows) < 4
        ):
            return False
        radius = max(6, int(round(vertical.shape[1] * 0.01)))
        left = max(0, candidate - radius)
        right = min(vertical.shape[1], candidate + radius + 1)
        supported_rows = 0
        tested_rows = 0
        for top, bottom in zip(rows, rows[1:]):
            top = max(0, int(top) + 2)
            bottom = min(vertical.shape[0], int(bottom) - 2)
            if bottom - top < 8:
                continue
            tested_rows += 1
            coverage = float(
                np.count_nonzero(np.any(vertical[top:bottom, left:right] > 0, axis=1))
            ) / float(bottom - top)
            if coverage >= 0.72:
                supported_rows += 1
        return bool(
            tested_rows >= 3
            and supported_rows >= int(math.ceil(tested_rows * 0.70))
        )

    for left, right in zip(recovered[:-1], recovered[1:]):
        gap = right - left
        if gap < typical * 1.10:
            continue
        for candidate in candidates:
            if not left + edge_margin < candidate < right - edge_margin:
                continue
            local = projection[max(0, candidate - 3) : min(projection.size, candidate + 4)]
            peak = int(local.max(initial=0))
            ordinary_support = bool(
                gap >= typical * 1.75
                and peak >= support_extent * 0.20
                and peak >= global_peak * 0.35
            )
            exceptional_support = bool(
                peak >= support_extent * 0.28
                and peak >= global_peak * 0.55
            )
            if ordinary_support or exceptional_support or follows_rows(candidate):
                insertions.append(int(candidate))
    return _merge_nearby_centers(recovered + insertions, maximum_gap=5)


def _extend_regular_boundaries_with_evidence(
    centers: list[int],
    projection: np.ndarray,
    support_extent: int,
    lower_limit: int,
    upper_limit: int,
    maximum_steps: int = 8,
) -> list[int]:
    """Extend a fading row sequence while every next rule is still visible."""
    recovered = _merge_nearby_centers(centers, maximum_gap=5)
    if len(recovered) < 5 or projection.size == 0:
        return recovered
    gaps = np.diff(recovered)
    ordinary = gaps[gaps <= np.percentile(gaps, 70)]
    typical = float(np.median(ordinary)) if ordinary.size else float(np.median(gaps))
    global_peak = int(projection.max(initial=0))
    if typical < 6 or global_peak <= 0:
        return recovered

    def extend(direction: int) -> None:
        for _ in range(maximum_steps):
            edge = recovered[0] if direction < 0 else recovered[-1]
            predicted = int(round(edge + direction * typical))
            if predicted < lower_limit or predicted > upper_limit:
                break
            search_radius = max(5, int(round(typical * 0.22)))
            local_left = max(lower_limit, predicted - search_radius)
            local_right = min(upper_limit + 1, predicted + search_radius + 1)
            local = projection[local_left:local_right]
            if local.size == 0:
                break
            peak_offset = int(np.argmax(local))
            peak = int(local[peak_offset])
            if (
                peak < support_extent * 0.10
                or peak < global_peak * 0.10
            ):
                break
            position = local_left + peak_offset
            if abs(abs(position - edge) - typical) > typical * 0.36:
                break
            recovered.insert(0, position) if direction < 0 else recovered.append(position)

    extend(-1)
    extend(1)
    return _merge_nearby_centers(recovered, maximum_gap=5)


def _dominant_supported_span(mask: np.ndarray, maximum_hole: int) -> tuple[int, int] | None:
    indices = np.flatnonzero(mask)
    if indices.size < 2:
        return None
    groups = [
        group
        for group in np.split(indices, np.where(np.diff(indices) > maximum_hole)[0] + 1)
        if group.size
    ]
    group = max(groups, key=lambda item: (int(item[-1] - item[0]), int(item.size)))
    return int(group[0]), int(group[-1])


def _recover_sparse_low_contrast_grid(
    image: np.ndarray,
    horizontal: np.ndarray,
    vertical: np.ndarray,
    columns: list[int],
    rows: list[int],
) -> tuple[list[int], list[int]]:
    """Recover a ruled photo whose faint body lines were reduced to page edges.

    Strong perspective and illumination falloff can leave only two nearby page
    edges plus one far edge in the morphology result.  Recovery requires three
    independent facts: a long horizontal table rule, consistent vertical edge
    gradients, and substantial line-map support.  Text alignment alone cannot
    create a grid here.
    """
    height, width = image.shape[:2]
    gaps = np.diff(columns)
    sparse_columns = bool(
        len(columns) < 3
        or (
            len(columns) <= 3
            and gaps.size
            and float(gaps.max()) >= width * 0.75
        )
    )
    if len(rows) < 5:
        return columns, rows

    horizontal_projection = np.count_nonzero(horizontal, axis=1)
    usable_rows = [row for row in rows if height * 0.04 <= row <= height * 0.96]
    if not usable_rows:
        return columns, rows
    strongest_row = max(
        usable_rows,
        key=lambda row: int(
            horizontal_projection[max(0, row - 3) : min(height, row + 4)].max(initial=0)
        ),
    )
    horizontal_band = np.any(
        horizontal[max(0, strongest_row - 3) : min(height, strongest_row + 4)] > 0,
        axis=0,
    )
    horizontal_span = _dominant_supported_span(
        horizontal_band,
        maximum_hole=max(12, width // 80),
    )
    if horizontal_span is None or horizontal_span[1] - horizontal_span[0] < width * 0.35:
        return columns, rows

    left, right = horizontal_span
    horizontal_width = max(1, right - left)
    current_width = max(0, columns[-1] - columns[0]) if len(columns) >= 2 else 0
    incomplete_span = bool(
        len(columns) <= 8
        and (
            not columns
            or columns[0] > left + horizontal_width * 0.18
            or columns[-1] < right - horizontal_width * 0.18
            or current_width < horizontal_width * 0.72
        )
    )
    if not sparse_columns and not incomplete_span:
        return columns, rows

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
    vertical_projection = np.count_nonzero(vertical, axis=0)
    vertical_peak = int(vertical_projection.max(initial=0))
    relaxed_partial_recovery = incomplete_span and not sparse_columns
    minimum_vertical_support = max(
        height * (0.11 if relaxed_partial_recovery else 0.18),
        vertical_peak * (0.20 if relaxed_partial_recovery else 0.30),
    )
    supported_indices = np.flatnonzero(vertical_projection >= minimum_vertical_support)
    projection_centers: list[int] = []
    if relaxed_partial_recovery and supported_indices.size:
        maximum_gap = max(5, width // 100)
        groups = np.split(
            supported_indices,
            np.where(np.diff(supported_indices) > maximum_gap)[0] + 1,
        )
        for group in groups:
            if group.size:
                projection_centers.append(
                    int(group[int(np.argmax(vertical_projection[group]))])
                )
    recovered_columns = [
        column
        for column in (
            _consistent_vertical_edge_centers(gray, minimum_median=20.0)
            + projection_centers
            + ([left, right] if relaxed_partial_recovery else [])
        )
        if left - 12 <= column <= right + 12
        and int(
            vertical_projection[
                max(0, column - 3) : min(width, column + 4)
            ].max(initial=0)
        )
        >= minimum_vertical_support
    ]
    recovered_columns = _merge_nearby_centers(
        recovered_columns,
        maximum_gap=max(5, width // 100) if relaxed_partial_recovery else 5,
    )
    if not 4 <= len(recovered_columns) <= 20:
        return columns, rows
    if relaxed_partial_recovery:
        # Motion blur can weaken only one side of the vertical rules while the
        # horizontal row sequence remains trustworthy.  Rebuild the columns,
        # but keep those observed rows; deriving their span from one fragmented
        # vertical rule can drop the header or append a paper-margin row.
        return recovered_columns, rows

    strongest_column = max(
        recovered_columns,
        key=lambda column: int(
            vertical_projection[max(0, column - 3) : min(width, column + 4)].max(initial=0)
        ),
    )
    vertical_band = np.any(
        vertical[:, max(0, strongest_column - 3) : min(width, strongest_column + 4)] > 0,
        axis=1,
    )
    vertical_span = _dominant_supported_span(
        vertical_band,
        maximum_hole=max(50, height // 10),
    )
    if vertical_span is None or vertical_span[1] - vertical_span[0] < height * 0.30:
        return columns, rows

    top, bottom = vertical_span
    horizontal_peak = int(horizontal_projection.max(initial=0))
    consistent_rows = [
        row
        for row in _consistent_vertical_edge_centers(gray.T, minimum_median=20.0)
        if top - 12 <= row <= bottom + 12
        and int(
            horizontal_projection[
                max(0, row - 3) : min(height, row + 4)
            ].max(initial=0)
        )
        >= max(width * 0.15, horizontal_peak * 0.20)
    ]
    recovered_rows = _merge_nearby_centers(
        consistent_rows + [row for row in rows if top - 12 <= row <= bottom + 12],
        maximum_gap=5,
    )
    recovered_rows = _recover_regular_missing_boundaries(
        recovered_rows,
        horizontal_projection,
        width,
        maximum_multiple=8,
        minimum_support_ratio=0.10,
        minimum_global_peak_ratio=0.10,
    )
    recovered_rows = _extend_regular_boundaries_with_evidence(
        recovered_rows,
        horizontal_projection,
        width,
        0,
        height - 1,
    )
    if len(recovered_rows) < 4:
        return columns, rows
    return recovered_columns, recovered_rows


def grid_has_excluded_supported_rows(
    image: np.ndarray,
    columns: list[int],
    rows: list[int],
) -> bool:
    """Return true when a table-width rule sits just outside the accepted grid."""
    if image.size == 0 or len(columns) < 3 or len(rows) < 5:
        return False
    height, width = image.shape[:2]
    left = max(0, min(width - 1, int(columns[0])))
    right = max(left + 1, min(width, int(columns[-1]) + 1))
    table_width = right - left
    if table_width < width * 0.30:
        return False

    horizontal, vertical, _ = _grid_maps(image)
    projection = np.count_nonzero(horizontal[:, left:right], axis=1)
    centers = _line_centers(projection, table_width)
    gaps = np.diff(rows).astype(float)
    ordinary = gaps[gaps <= np.percentile(gaps, 75)] if gaps.size else gaps
    typical = float(np.median(ordinary)) if ordinary.size else 0.0
    if typical < 6.0:
        return False

    tolerance = max(4.0, typical * 0.18)

    def disconnected_empty_frame_band(center: int) -> bool:
        if center < rows[0]:
            start, stop = int(center), int(rows[0])
        else:
            start, stop = int(rows[-1]), int(center)
        band_margin = max(3, int(round(typical * 0.10)))
        inner_start = start + band_margin
        inner_stop = stop - band_margin
        if inner_stop <= inner_start:
            return False
        inner_columns = [int(value) for value in columns[1:-1]]
        if not inner_columns or _edge_rule_continuation(
            vertical,
            inner_columns,
            inner_start,
            inner_stop,
            horizontal=False,
        ) >= 0.25:
            return False
        horizontal_margin = max(4, int(round(table_width * 0.01)))
        interior = image[
            inner_start:inner_stop,
            min(right, left + horizontal_margin) : max(left, right - horizontal_margin),
        ]
        if interior.size == 0:
            return False
        gray = (
            cv2.cvtColor(interior, cv2.COLOR_BGR2GRAY)
            if interior.ndim == 3
            else interior
        )
        cleaned = gray.copy()
        background_level = int(np.percentile(gray, 75))
        cleaned[: min(2, cleaned.shape[0]), :] = background_level
        cleaned[max(0, cleaned.shape[0] - 2) :, :] = background_level
        x_origin = min(right, left + horizontal_margin)
        for column in columns:
            local = int(column) - x_origin
            if -4 <= local < cleaned.shape[1] + 4:
                cleaned[
                    :,
                    max(0, local - 4) : min(cleaned.shape[1], local + 5),
                ] = background_level

        kernel_side = max(9, int(round(max(typical, 12.0) * 0.45)))
        if kernel_side % 2 == 0:
            kernel_side += 1
        background = cv2.morphologyEx(
            cleaned,
            cv2.MORPH_CLOSE,
            cv2.getStructuringElement(
                cv2.MORPH_RECT,
                (kernel_side, kernel_side),
            ),
        )
        contrast = cv2.subtract(background, cleaned)
        adaptive = cv2.adaptiveThreshold(
            cleaned,
            255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY_INV,
            31,
            9,
        )
        ink = np.where((contrast >= 8) & (adaptive > 0), 255, 0).astype(np.uint8)
        count, labels, stats, _ = cv2.connectedComponentsWithStats(ink, 8)
        band_height, band_width = cleaned.shape[:2]
        min_height = max(3, int(round(max(typical, band_height) * 0.10)))
        max_height = max(min_height, int(round(band_height * 0.88)))
        for index in range(1, count):
            x, y, component_width, component_height, area = [
                int(value) for value in stats[index]
            ]
            touches_edge = bool(
                x <= 1
                or y <= 1
                or x + component_width >= band_width - 1
                or y + component_height >= band_height - 1
            )
            fill = float(area / max(1, component_width * component_height))
            aspect = float(component_width / max(1, component_height))
            mean_contrast = float(np.mean(contrast[labels == index]))
            if (
                not touches_edge
                and min_height <= component_height <= max_height
                and component_width >= 2
                and area >= max(6, int(round(typical * 0.25)))
                and aspect <= 12.0
                and 0.06 <= fill <= 0.90
                and not (
                    aspect >= 3.5
                    and fill <= 0.10
                    and area <= component_width * 1.20
                )
                # 低对比纸纹/装饰底纹会在表格外框旁形成窄碎片；可见文字在
                # 回归样本中的局部暗差明显更高。保留充足余量，避免把这些
                # 背景碎片误判成被裁掉的数据行。
                and mean_contrast >= 16.0
            ):
                return False
        return True

    for center in centers:
        if center < rows[0] - tolerance:
            distance = float(rows[0] - center)
        elif center > rows[-1] + tolerance:
            distance = float(center - rows[-1])
        else:
            continue
        if (
            typical * 0.55 <= distance <= typical * 1.50
            and not disconnected_empty_frame_band(center)
        ):
            return True
    return False


def _remove_weak_split_columns(
    columns: list[int],
    vertical: np.ndarray,
) -> list[int]:
    """Remove a repeated text stroke that falsely divides one real column.

    In compact screenshots, identical text at the same horizontal position in
    every row can survive the vertical morphology pass.  It differs from a
    faint rule in two useful ways: it splits one ordinary-width column into
    two short bands, and its vertical support is materially weaker than both
    neighbouring rules.  Requiring both conditions keeps genuine narrow
    columns and shadowed rules intact.
    """
    if len(columns) < 5 or vertical.size == 0:
        return columns

    filtered = sorted(set(int(value) for value in columns))
    projection = np.count_nonzero(vertical, axis=0)

    def support(position: int) -> int:
        left = max(0, position - 3)
        right = min(projection.size, position + 4)
        return int(projection[left:right].max(initial=0))

    while len(filtered) >= 5:
        gaps = np.diff(filtered)
        removed = False
        ordinary_gaps = gaps[gaps >= 8]
        ordinary_typical = (
            float(np.median(ordinary_gaps)) if ordinary_gaps.size else 0.0
        )
        for gap_index, gap in enumerate(gaps):
            if (
                ordinary_typical < 12.0
                or gap_index == 0
                or gap_index + 1 >= len(gaps)
                or float(gap) > max(7.0, ordinary_typical * 0.22)
            ):
                continue
            left_outer = float(gaps[gap_index - 1])
            right_outer = float(gaps[gap_index + 1])
            if not (
                ordinary_typical * 0.75
                <= left_outer
                <= ordinary_typical * 1.60
                and ordinary_typical * 0.75
                <= right_outer
                <= ordinary_typical * 1.60
            ):
                continue
            left_support = support(filtered[gap_index])
            right_support = support(filtered[gap_index + 1])
            stronger_support = max(left_support, right_support)
            weaker_support = min(left_support, right_support)
            if (
                stronger_support > 0
                and weaker_support <= stronger_support * 0.15
                and weaker_support < vertical.shape[0] * 0.08
            ):
                filtered.pop(
                    gap_index if left_support <= right_support else gap_index + 1
                )
                removed = True
                break
        if removed:
            continue
        for index in range(1, len(filtered) - 1):
            left_gap = float(gaps[index - 1])
            right_gap = float(gaps[index])
            reference_gaps = np.delete(gaps, [index - 1, index])
            if reference_gaps.size < 2:
                continue
            typical = float(np.median(reference_gaps))
            combined = left_gap + right_gap
            if typical < 12:
                continue
            if not (
                left_gap <= typical * 0.72
                and right_gap <= typical * 0.72
                and typical * 0.75 <= combined <= typical * 1.40
            ):
                continue
            candidate_support = support(filtered[index])
            neighbour_support = min(
                support(filtered[index - 1]),
                support(filtered[index + 1]),
            )
            if (
                neighbour_support > 0
                and candidate_support < neighbour_support * 0.70
                and candidate_support < vertical.shape[0] * 0.65
            ):
                filtered.pop(index)
                removed = True
                break
        if not removed:
            break
    return filtered


def _vertical_rule_window_coverage(vertical_band: np.ndarray) -> np.ndarray:
    """Return per-column row coverage for the same centered seven-pixel window."""
    if vertical_band.size == 0:
        width = vertical_band.shape[1] if vertical_band.ndim == 2 else 0
        return np.zeros(width, dtype=float)
    expanded = cv2.dilate(
        (vertical_band > 0).astype(np.uint8),
        np.ones((1, 7), dtype=np.uint8),
        borderType=cv2.BORDER_CONSTANT,
        borderValue=0,
    )
    return np.count_nonzero(expanded, axis=0) / float(expanded.shape[0])


def _body_band_consensus_columns(
    vertical: np.ndarray,
    horizontal: np.ndarray,
    rows: list[int],
) -> list[int]:
    """Find non-uniform column rules repeated through ordinary data rows."""
    if vertical.size == 0 or horizontal.size == 0 or len(rows) < 9:
        return []
    height, width = vertical.shape[:2]
    gaps = np.diff(rows).astype(float)
    ordinary = gaps[gaps <= np.percentile(gaps, 80)]
    typical = float(np.median(ordinary)) if ordinary.size else 0.0
    if typical < 8:
        return []
    bands = [
        (int(top) + 2, int(bottom) - 2)
        for index, (top, bottom) in enumerate(zip(rows, rows[1:]))
        if index >= 3
        and 8 <= bottom - top <= typical * 1.5
    ]
    if len(bands) < 6:
        return []
    support = np.zeros(width, dtype=np.int16)
    for top, bottom in bands:
        top = max(0, min(height, top))
        bottom = max(top + 1, min(height, bottom))
        coverage = _vertical_rule_window_coverage(vertical[top:bottom])
        support[coverage >= 0.72] += 1
    required = int(math.ceil(len(bands) * 0.65))
    supported = np.flatnonzero(support >= required)
    groups = (
        np.split(supported, np.where(np.diff(supported) > 1)[0] + 1)
        if supported.size
        else []
    )
    centers = [
        int(group[int(np.argmax(support[group]))])
        for group in groups
        if group.size
    ]
    centers = _merge_nearby_centers(centers, maximum_gap=7)
    endpoint_spans: list[tuple[int, int]] = []
    for row in rows:
        band = np.any(
            horizontal[
                max(0, int(row) - 3) : min(height, int(row) + 4)
            ] > 0,
            axis=0,
        )
        span = _dominant_supported_span(
            band,
            maximum_hole=max(12, width // 80),
        )
        if span is not None and span[1] - span[0] >= width * 0.35:
            endpoint_spans.append(span)
    if len(endpoint_spans) < 5:
        return []
    table_left = int(round(float(np.median([span[0] for span in endpoint_spans]))))
    table_right = int(round(float(np.median([span[1] for span in endpoint_spans]))))
    tolerance = max(8, int(round(width * 0.012)))
    centers = [
        center
        for center in centers
        if table_left - tolerance <= center <= table_right + tolerance
    ]
    centers = _merge_nearby_centers(
        [table_left, *centers, table_right],
        maximum_gap=7,
    )
    if not 4 <= len(centers) <= 32:
        return []
    if min(np.diff(centers), default=0) < max(8, width // 100):
        return []
    return centers


def _recover_crop_edge_boundaries(
    centers: list[int],
    extent: int,
) -> list[int]:
    """Restore an outer rule clipped exactly by perspective rectification.

    A tight crop can place the table's top or left rule just outside the
    returned image.  When the distance to the crop edge is one ordinary cell
    interval, the edge itself is the missing boundary.  Sparse paper margins
    are removed by the existing density guards immediately afterwards.
    """
    if len(centers) < 4 or extent < 2:
        return centers

    recovered = sorted(set(int(value) for value in centers))
    gaps = np.diff(recovered)
    ordinary = gaps[gaps >= 8]
    if ordinary.size < 3:
        return recovered
    typical = float(np.median(ordinary))
    if typical < 8:
        return recovered

    edge_slack = max(3.0, extent * 0.003)
    leading = float(recovered[0])
    if typical * 0.65 - edge_slack <= leading <= typical * 1.35 + edge_slack:
        recovered.insert(0, 0)

    trailing = float((extent - 1) - recovered[-1])
    if typical * 0.65 - edge_slack <= trailing <= typical * 1.35 + edge_slack:
        recovered.append(extent - 1)
    return recovered


def _recover_outer_columns_from_horizontal_endpoints(
    image: np.ndarray,
    horizontal: np.ndarray,
    columns: list[int],
    rows: list[int],
) -> list[int]:
    """Restore an outer column proved by repeated row-line endpoints."""
    if len(columns) < 4 or len(rows) < 6 or horizontal.size == 0:
        return columns
    height, width = image.shape[:2]
    endpoints: list[tuple[int, int]] = []
    for row in rows:
        if not 0 <= row < height:
            continue
        band = np.any(
            horizontal[max(0, row - 3) : min(height, row + 4)] > 0,
            axis=0,
        )
        span = _dominant_supported_span(
            band,
            maximum_hole=max(12, width // 80),
        )
        if span is not None and span[1] - span[0] >= width * 0.35:
            endpoints.append(span)
    if len(endpoints) < max(5, int(math.ceil(len(rows) * 0.50))):
        return columns

    left_values = np.asarray([item[0] for item in endpoints], dtype=float)
    right_values = np.asarray([item[1] for item in endpoints], dtype=float)
    left = int(round(float(np.median(left_values))))
    right = int(round(float(np.median(right_values))))
    endpoint_tolerance = max(6, int(round(width * 0.018)))
    left_support = int(np.count_nonzero(np.abs(left_values - left) <= endpoint_tolerance))
    right_support = int(np.count_nonzero(np.abs(right_values - right) <= endpoint_tolerance))

    recovered = sorted(set(int(value) for value in columns))
    gaps = np.diff(recovered)
    ordinary = gaps[gaps <= np.percentile(gaps, 70)]
    typical = float(np.median(ordinary)) if ordinary.size else float(np.median(gaps))
    if typical < 8:
        return recovered
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
    consistent = _consistent_vertical_edge_centers(gray, minimum_median=20.0)
    evidence_tolerance = max(10, int(round(width * 0.03)))
    minimum_support = max(5, int(math.ceil(len(endpoints) * 0.60)))

    if (
        recovered[0] - left >= typical * 1.20
        and left_support >= minimum_support
        and any(abs(candidate - left) <= evidence_tolerance for candidate in consistent)
    ):
        recovered.insert(0, left)
    if (
        right - recovered[-1] >= typical * 1.20
        and right_support >= minimum_support
        and any(abs(candidate - right) <= evidence_tolerance for candidate in consistent)
    ):
        recovered.append(right)
    return _merge_nearby_centers(recovered, maximum_gap=5)


def _edge_rule_continuation(
    line_map: np.ndarray,
    fixed_centers: list[int],
    start: int,
    stop: int,
    *,
    horizontal: bool,
) -> float:
    if stop <= start:
        return 0.0
    height, width = line_map.shape[:2]
    supported = []
    for center in fixed_centers:
        if horizontal:
            band = line_map[
                max(0, center - 2) : min(height, center + 3),
                start:stop,
            ]
            coverage = float(np.mean(np.any(band > 0, axis=0))) if band.size else 0.0
        else:
            band = line_map[
                start:stop,
                max(0, center - 2) : min(width, center + 3),
            ]
            coverage = float(np.mean(np.any(band > 0, axis=1))) if band.size else 0.0
        supported.append(coverage >= 0.65)
    return float(np.mean(supported)) if supported else 0.0


def _recover_clipped_frame_boundaries(
    image: np.ndarray,
    horizontal: np.ndarray,
    vertical: np.ndarray,
    columns: list[int],
    rows: list[int],
) -> tuple[list[int], list[int]]:
    """Restore one crop-edge boundary only when continuing rules prove it."""
    height, width = image.shape[:2]
    output_columns = sorted(set(int(value) for value in columns))
    output_rows = sorted(set(int(value) for value in rows))

    row_gaps = np.diff(output_rows).astype(float)
    if row_gaps.size >= 6:
        typical = float(np.median(row_gaps))
        regularity = float(np.mean(np.abs(row_gaps - typical) <= max(2.0, typical * 0.12)))
        if typical >= 8.0 and regularity >= 0.80:
            leading = float(output_rows[0])
            if typical * 0.55 <= leading <= typical * 1.45 and _edge_rule_continuation(
                vertical,
                output_columns,
                0,
                output_rows[0],
                horizontal=False,
            ) >= 0.70:
                output_rows.insert(0, 0)
            trailing = float((height - 1) - output_rows[-1])
            if typical * 0.55 <= trailing <= typical * 1.45 and _edge_rule_continuation(
                vertical,
                output_columns,
                output_rows[-1],
                height,
                horizontal=False,
            ) >= 0.70:
                output_rows.append(height - 1)

    column_gaps = np.diff(output_columns).astype(float)
    if column_gaps.size >= 5:
        typical = float(np.median(column_gaps))
        if typical >= 12.0:
            leading = float(output_columns[0])
            if typical * 0.40 <= leading <= typical * 1.45 and _edge_rule_continuation(
                horizontal,
                output_rows,
                0,
                output_columns[0],
                horizontal=True,
            ) >= 0.70:
                output_columns.insert(0, 0)
            trailing = float((width - 1) - output_columns[-1])
            if typical * 0.55 <= trailing <= typical * 1.45 and _edge_rule_continuation(
                horizontal,
                output_rows,
                output_columns[-1],
                width,
                horizontal=True,
            ) >= 0.70:
                output_columns.append(width - 1)
    if (
        len(output_columns) >= 4
        and len(output_rows) >= 8
        and width - 1 - output_columns[-1] > 3
    ):
        gaps = np.diff(output_columns).astype(float)
        typical = float(np.median(gaps)) if gaps.size else 0.0
        trailing = float((width - 1) - output_columns[-1])
        if (
            typical >= 12.0
            and typical * 0.55 <= trailing <= typical * 3.0
            and _edge_rule_continuation(
                horizontal,
                output_rows,
                output_columns[-1],
                width,
                horizontal=True,
            )
            >= 0.85
        ):
            output_columns.append(width - 1)
    return output_columns, output_rows


def _trim_proven_empty_crop_edge_column(
    image: np.ndarray,
    horizontal: np.ndarray,
    columns: list[int],
    rows: list[int],
) -> list[int]:
    """Remove a narrow outer warp sliver only when row endpoints prove emptiness."""
    if len(columns) < 6 or len(rows) < 8:
        return columns
    height, width = image.shape[:2]
    if width - 1 - columns[-1] > 3:
        return columns
    gaps = np.diff(columns).astype(float)
    typical = float(np.median(gaps[:-1])) if gaps.size > 1 else 0.0
    if typical < 12.0 or gaps[-1] > typical * 0.42:
        return columns
    endpoints: list[int] = []
    for row in rows:
        band = np.any(
            horizontal[max(0, row - 3) : min(height, row + 4)] > 0,
            axis=0,
        )
        span = _dominant_supported_span(band, maximum_hole=max(12, width // 80))
        if span is not None:
            endpoints.append(int(span[1]))
    if len(endpoints) < max(6, int(math.ceil(len(rows) * 0.60))):
        return columns
    tolerance = max(6, int(round(width * 0.02)))
    support = sum(abs(endpoint - columns[-2]) <= tolerance for endpoint in endpoints)
    return columns[:-1] if support >= math.ceil(len(endpoints) * 0.60) else columns


def _trim_disconnected_outer_frame_cells(
    image: np.ndarray,
    horizontal: np.ndarray,
    vertical: np.ndarray,
    columns: list[int],
    rows: list[int],
) -> tuple[list[int], list[int]]:
    """Remove narrow decorative frame cells proved disconnected from the table body."""
    if len(columns) < 6 or len(rows) < 8:
        return columns, rows

    height, width = image.shape[:2]
    output_columns = sorted(set(int(value) for value in columns))
    output_rows = sorted(set(int(value) for value in rows))
    column_gaps = np.diff(output_columns).astype(float)
    ordinary_column_gaps = column_gaps[column_gaps >= 8.0]
    typical_column = (
        float(np.median(ordinary_column_gaps)) if ordinary_column_gaps.size else 0.0
    )
    if typical_column < 12.0:
        return output_columns, output_rows

    edge_limit = max(6, int(round(width * 0.03)))
    body_rows = output_rows[1:-1]

    def disconnected_column_interval(index: int) -> bool:
        left = output_columns[index]
        right = output_columns[index + 1]
        return bool(
            right - left <= typical_column * 0.28
            and _edge_rule_continuation(
                horizontal,
                body_rows,
                left + 2,
                right - 2,
                horizontal=True,
            )
            < 0.25
        )

    trim_left = bool(
        output_columns[0] <= edge_limit and disconnected_column_interval(0)
    )
    trim_right = bool(
        output_columns[-1] >= width - 1 - edge_limit
        and disconnected_column_interval(len(output_columns) - 2)
    )
    if trim_left:
        output_columns.pop(0)
    if trim_right and len(output_columns) >= 5:
        output_columns.pop()

    body_columns = output_columns[1:-1]
    if len(body_columns) < 3:
        return output_columns, output_rows

    row_support = [
        _edge_rule_continuation(
            vertical,
            body_columns,
            top + 2,
            bottom - 2,
            horizontal=False,
        )
        for top, bottom in zip(output_rows[:-1], output_rows[1:])
    ]

    leading_run = 0
    for support in row_support[:4]:
        if support >= 0.25:
            break
        leading_run += 1
    if leading_run > 3 or (
        leading_run == 3 and len(row_support) > 3 and row_support[3] < 0.25
    ):
        leading_run = 0

    trailing_run = 0
    for support in reversed(row_support[-3:]):
        if support >= 0.25:
            break
        trailing_run += 1
    if trailing_run > 2 or (
        trailing_run == 2 and len(row_support) > 2 and row_support[-3] < 0.25
    ):
        trailing_run = 0

    edge_row_limit = max(6, int(round(height * 0.03)))

    def outer_row_visual_stats(index: int) -> tuple[float, float, float, float] | None:
        top = output_rows[index]
        bottom = output_rows[index + 1]
        vertical_margin = max(3, int(round((bottom - top) * 0.18)))
        horizontal_margin = max(
            3,
            int(round((output_columns[-1] - output_columns[0]) * 0.01)),
        )
        interior = image[
            top + vertical_margin : bottom - vertical_margin,
            output_columns[0] + horizontal_margin : output_columns[-1] - horizontal_margin,
        ]
        if interior.size == 0:
            return None
        gray = (
            cv2.cvtColor(interior, cv2.COLOR_BGR2GRAY)
            if interior.ndim == 3
            else interior
        )
        contrast = float(np.percentile(gray, 90) - np.percentile(gray, 10))
        edge_ratio = float(np.mean(cv2.Canny(gray, 50, 120) > 0))
        return float(np.median(gray)), float(np.std(gray)), contrast, edge_ratio

    def visually_empty_outer_row(index: int, neighbor_index: int) -> bool:
        stats = outer_row_visual_stats(index)
        neighbor_stats = outer_row_visual_stats(neighbor_index)
        if stats is None or neighbor_stats is None:
            return False
        median, deviation, contrast, edge_ratio = stats
        neighbor_median, neighbor_deviation, neighbor_contrast, neighbor_edges = (
            neighbor_stats
        )
        visually_separated = bool(
            abs(median - neighbor_median) >= 12.0
            or neighbor_deviation >= 16.0
            or neighbor_contrast >= 30.0
            or neighbor_edges >= 0.03
        )
        return bool(
            deviation <= 8.0
            and contrast <= 18.0
            and edge_ratio <= 0.012
            and visually_separated
        )

    flat_leading_row = bool(
        output_rows[0] <= edge_row_limit and visually_empty_outer_row(0, 1)
    )
    flat_trailing_row = bool(
        output_rows[-1] >= height - 1 - edge_row_limit
        and visually_empty_outer_row(len(output_rows) - 2, len(output_rows) - 3)
    )
    paired_outer_rows = bool(
        output_rows[0] <= edge_row_limit
        and output_rows[-1] >= height - 1 - edge_row_limit
        and 2 <= leading_run <= 3
        and 1 <= trailing_run <= 2
    )
    if not (
        trim_left
        or trim_right
        or paired_outer_rows
        or flat_leading_row
        or flat_trailing_row
    ):
        return output_columns, output_rows

    # A merged or dark table header may legitimately omit internal vertical
    # rules. Keep the last leading low-support interval as that header; every
    # preceding interval and all trailing disconnected intervals belong to the
    # surrounding decorative frame.
    leading_remove = max(int(flat_leading_row), max(0, leading_run - 1))
    trailing_remove = max(int(flat_trailing_row), trailing_run)
    end = len(output_rows) - trailing_remove if trailing_remove else len(output_rows)
    candidate_rows = output_rows[leading_remove:end]
    if len(candidate_rows) >= 4:
        output_rows = candidate_rows
    return output_columns, output_rows


def _regularize_ruled_rows(
    rows: list[int],
    horizontal: np.ndarray,
    vertical: np.ndarray,
    image_height: int,
    image_width: int,
    image: np.ndarray | None = None,
) -> list[int]:
    """Recover obscured ruled-table boundaries without splitting genuine merged rows."""
    if len(rows) < 6:
        return rows

    recovered = sorted(set(rows))
    gaps = np.diff(recovered)
    typical = float(np.median(gaps)) if gaps.size else 0.0
    if typical < 8:
        return recovered

    projection = np.count_nonzero(horizontal, axis=1)
    missing_candidates: list[tuple[int, int]] = []
    for index, gap in enumerate(gaps):
        if not typical * 1.65 <= gap <= typical * 2.35:
            continue
        predicted = int(round((recovered[index] + recovered[index + 1]) / 2.0))
        search_top = max(0, predicted - 3)
        search_bottom = min(image_height, predicted + 4)
        local = projection[search_top:search_bottom]
        if local.size == 0:
            continue
        local_index = int(np.argmax(local))
        candidate = search_top + local_index
        # A real but shadowed rule may be shorter than the global projection
        # threshold. Requiring at least 30% table-width support avoids dividing
        # deliberately merged/tall rows based only on text strokes.
        has_line_support = int(local[local_index]) >= image_width * 0.30
        transition_candidate: int | None = None
        if not has_line_support and image is not None:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
            scan_top = max(1, predicted - 8)
            scan_bottom = min(image_height - 1, predicted + 9)
            table_left = max(0, int(round(image_width * 0.04)))
            table_right = min(image_width, int(round(image_width * 0.96)))
            band = gray[scan_top - 1 : scan_bottom + 1, table_left:table_right]
            if band.shape[0] >= 3 and band.shape[1] >= 20:
                # A white rule between similarly coloured title/header bands
                # disappears from the inverse binary line map.  Recover it
                # from a broad luminance transition, while requiring the
                # change to cover much of the table width so text strokes do
                # not split a deliberately merged cell.
                changes = np.mean(
                    np.abs(band[2:].astype(np.int16) - band[:-2].astype(np.int16)),
                    axis=1,
                )
                changed_fraction = np.mean(
                    np.abs(band[2:].astype(np.int16) - band[:-2].astype(np.int16)) >= 24,
                    axis=1,
                )
                best = int(np.argmax(changes))
                if float(changes[best]) >= 16.0 and float(changed_fraction[best]) >= 0.35:
                    transition_candidate = scan_top + best
        if has_line_support:
            missing_candidates.append((index + 1, candidate))
        elif transition_candidate is not None:
            missing_candidates.append((index + 1, transition_candidate))

    for offset, (index, candidate) in enumerate(missing_candidates):
        recovered.insert(index + offset, candidate)

    gaps = np.diff(recovered)
    typical = float(np.median(gaps)) if gaps.size else typical
    remaining = image_height - 1 - recovered[-1]
    starts_at_crop_edge = recovered[0] <= typical * 0.35
    previous_band = vertical[
        max(0, int(round(recovered[-1] - typical)) + 2) : max(0, recovered[-1] - 1)
    ]
    trailing_band = vertical[recovered[-1] + 2 : max(0, image_height - 1)]
    vertical_continues = (
        np.count_nonzero(trailing_band)
        >= np.count_nonzero(previous_band) * 0.35
    )
    if (
        starts_at_crop_edge
        and typical * 0.65 <= remaining <= typical * 1.35
        and vertical_continues
    ):
        recovered.append(image_height - 1)

    return recovered


def _recover_leading_spreadsheet_gutter_boundary(
    columns: list[int],
    rows: list[int],
    vertical: np.ndarray,
) -> list[int]:
    """Restore a faint row-number gutter rule beside the first data column."""
    if len(columns) < 6 or len(rows) < 12 or vertical.size == 0:
        return columns
    recovered = sorted(set(int(value) for value in columns))
    gaps = np.diff(recovered).astype(float)
    ordinary = gaps[1:] if gaps.size > 2 else gaps
    typical = float(np.median(ordinary)) if ordinary.size else 0.0
    row_gaps = np.diff(rows).astype(float)
    body_row_gaps = row_gaps[1:] if row_gaps.size > 2 else row_gaps
    typical_row = float(np.median(body_row_gaps)) if body_row_gaps.size else 0.0
    if not (
        typical >= 24.0
        and typical_row >= 8.0
        and recovered[0] <= max(3, int(round(vertical.shape[1] * 0.01)))
        and gaps[0] >= typical * 1.15
        and row_gaps[0] <= typical_row * 0.85
    ):
        return recovered
    projection = np.count_nonzero(vertical, axis=0)
    table_span = max(1, rows[-1] - rows[0])
    interior = np.flatnonzero(
        projection[recovered[0] + 4 : recovered[1] - 3]
        >= table_span * 0.15
    )
    interior = interior + recovered[0] + 4 if interior.size else interior
    groups = (
        np.split(interior, np.where(np.diff(interior) > 3)[0] + 1)
        if interior.size
        else []
    )
    strong_positions: list[int] = []
    for group in groups:
        if not group.size:
            continue
        group_projection = projection[group]
        peak_value = int(group_projection.max(initial=0))
        peak_group = group[group_projection == peak_value]
        position = int(round(float(np.median(peak_group))))
        if (
            recovered[0] + typical * 0.55
            <= position
            <= recovered[1] - typical * 0.55
        ):
            strong_positions.append(position)
    regular_suffix: list[int] = []
    cursor = recovered[1]
    for position in reversed(strong_positions):
        gap = cursor - position
        if typical * 0.65 <= gap <= typical * 1.35:
            regular_suffix.append(position)
            cursor = position
        elif position < cursor - typical * 1.35:
            break
    regular_suffix.reverse()
    if len(regular_suffix) >= 2:
        first_regular = regular_suffix[0]
        gutter_left = recovered[0] + max(8, int(round(typical * 0.15)))
        gutter_right = min(
            recovered[0] + int(round(typical * 0.50)),
            first_regular - int(round(typical * 0.65)),
        )
        if gutter_right > gutter_left:
            gutter_projection = projection[gutter_left : gutter_right + 1]
            gutter_support = int(gutter_projection.max(initial=0))
            if gutter_support >= table_span * 0.06:
                gutter_indices = np.flatnonzero(
                    gutter_projection == gutter_support
                )
                gutter = gutter_left + int(
                    round(float(np.median(gutter_indices)))
                )
                if (
                    typical * 0.15
                    <= gutter - recovered[0]
                    <= typical * 0.50
                    and typical * 0.65
                    <= first_regular - gutter
                    <= typical * 1.35
                ):
                    return [
                        recovered[0],
                        gutter,
                        *regular_suffix,
                        *recovered[1:],
                    ]
    search_left = recovered[0] + max(8, int(round(typical * 0.12)))
    search_right = min(
        recovered[1] - max(8, int(round(typical * 0.65))),
        recovered[0] + int(round(typical * 0.48)),
    )
    if search_right <= search_left:
        return recovered
    local = projection[search_left : search_right + 1]
    if local.size == 0:
        return recovered
    peak_value = int(local.max(initial=0))
    peak_indices = np.flatnonzero(local == peak_value)
    candidate = search_left + int(round(float(np.median(peak_indices))))
    leading_gap = candidate - recovered[0]
    following_gap = recovered[1] - candidate
    if not (
        peak_value >= table_span * 0.15
        and typical * 0.18 <= leading_gap <= typical * 0.48
        and typical * 0.65 <= following_gap <= typical * 1.35
    ):
        return recovered
    recovered.insert(1, candidate)
    return recovered


def _recover_partial_header_row_boundaries(
    rows: list[int],
    horizontal: np.ndarray,
    image_width: int,
    image: np.ndarray | None = None,
) -> list[int]:
    """Recover partial rules inside one oversized multi-level header band."""
    if len(rows) < 12 or horizontal.size == 0 or image_width < 100:
        return rows
    recovered = sorted(set(int(value) for value in rows))
    gaps = np.diff(recovered).astype(float)
    if gaps.size < 10:
        return recovered
    body_gaps = gaps[4:] if gaps.size > 8 else gaps
    ordinary = body_gaps[body_gaps <= np.percentile(body_gaps, 75)]
    typical = float(np.median(ordinary)) if ordinary.size else 0.0
    if typical < 8.0:
        return recovered
    candidates = [
        index
        for index, gap in enumerate(gaps[:4])
        if typical * 2.25 <= gap <= typical * 4.30
    ]
    if len(candidates) != 1:
        return recovered
    index = candidates[0]
    left = recovered[index]
    right = recovered[index + 1]
    projection = np.count_nonzero(horizontal, axis=1)
    global_peak = int(projection.max(initial=0))
    minimum_support = max(image_width * 0.12, global_peak * 0.10)
    radius = max(5, int(round(typical * 0.30)))
    best: tuple[float, list[int]] | None = None
    for interval_count in (2, 3, 4):
        expected_gap = (right - left) / float(interval_count)
        if not typical * 0.85 <= expected_gap <= typical * 1.35:
            continue
        found: list[int] = []
        support_total = 0.0
        valid = True
        for part in range(1, interval_count):
            predicted = int(round(left + (right - left) * part / interval_count))
            search_left = max(left + 4, predicted - radius)
            search_right = min(right - 3, predicted + radius + 1)
            local = projection[search_left:search_right]
            if local.size == 0:
                valid = False
                break
            peak_value = int(local.max(initial=0))
            candidate: int | None = None
            if peak_value >= minimum_support:
                peak_indices = np.flatnonzero(local == peak_value)
                peak_index = int(round(float(np.median(peak_indices))))
                candidate = search_left + peak_index
            elif interval_count == 2 and image is not None:
                gray = (
                    cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
                    if image.ndim == 3
                    else image
                )
                tone_left = max(left + 3, predicted - radius)
                tone_right = min(right - 2, predicted + radius + 1)
                x_left = max(0, int(round(image_width * 0.02)))
                x_right = min(image_width, int(round(image_width * 0.98)))
                tone_band = gray[tone_left:tone_right, x_left:x_right]
                if tone_band.shape[0] >= 5 and tone_band.shape[1] >= 100:
                    background = np.percentile(tone_band, 75, axis=0)
                    coverage = np.mean(
                        tone_band <= background[np.newaxis, :] - 3,
                        axis=1,
                    )
                    tone_peak = float(coverage.max(initial=0.0))
                    tone_indices = np.flatnonzero(coverage == tone_peak)
                    tone_index = int(round(float(np.median(tone_indices))))
                    if tone_peak >= 0.65:
                        candidate = tone_left + tone_index
                        peak_value = int(round(tone_peak * image_width))
            if candidate is None:
                valid = False
                break
            found.append(candidate)
            support_total += peak_value
        if not valid:
            continue
        segments = np.diff([left, *found, right]).astype(float)
        if (
            segments.size != interval_count
            or float(segments.min()) < typical * 0.75
            or float(segments.max()) > typical * 1.45
            or float(segments.max()) / max(1.0, float(segments.min())) > 1.30
        ):
            continue
        score = support_total / float(interval_count - 1)
        if best is None or score > best[0]:
            best = (score, found)
    if best is None:
        return recovered
    return _merge_nearby_centers(recovered + best[1], maximum_gap=5)


def _trim_decorative_screen_top_frame(
    image: np.ndarray,
    columns: list[int],
    rows: list[int],
) -> list[int]:
    """Remove one full-width decorative frame above a regular dark header."""
    if len(columns) < 5 or len(rows) < 10:
        return rows
    height, width = image.shape[:2]
    if columns[-1] - columns[0] < width * 0.85:
        return rows

    gaps = np.diff(rows).astype(float)
    body_gaps = gaps[1:]
    typical = float(np.median(body_gaps)) if body_gaps.size else 0.0
    tolerance = max(4.0, typical * 0.18)
    regular_ratio = float(
        np.mean(np.abs(body_gaps - typical) <= tolerance)
    ) if body_gaps.size else 0.0
    if not (
        typical >= 8.0
        and rows[0] <= height * 0.015
        and height * 0.025 <= rows[1] <= height * 0.08
        and typical * 1.30 <= gaps[0] <= typical * 1.70
        and regular_ratio >= 0.88
    ):
        return rows

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
    horizontal_padding = max(3, int(round((columns[-1] - columns[0]) * 0.01)))

    def band_stats(top: int, bottom: int) -> tuple[float, float] | None:
        vertical_padding = max(2, int(round((bottom - top) * 0.15)))
        region = gray[
            top + vertical_padding : bottom - vertical_padding,
            columns[0] + horizontal_padding : columns[-1] - horizontal_padding,
        ]
        if region.size == 0:
            return None
        return float(np.median(region)), float(np.mean(cv2.Canny(region, 50, 120) > 0))

    frame_stats = band_stats(rows[0], rows[1])
    header_stats = band_stats(rows[1], rows[2])
    if frame_stats is None or header_stats is None:
        return rows
    frame_median, frame_edges = frame_stats
    header_median, header_edges = header_stats
    if not (
        frame_median - header_median >= 35.0
        and header_edges >= 0.02
        and frame_edges <= header_edges * 0.60
    ):
        return rows
    return rows[1:]


def _trim_sparse_page_edge_columns(
    image: np.ndarray,
    columns: list[int],
    horizontal: np.ndarray | None = None,
) -> list[int]:
    """Discard photographed page edges that were mistaken for table borders.

    Perspective rectification can leave the left and right paper edges exactly
    on the crop boundary.  They are long enough to look like vertical rules,
    but the bands between them and the real table are almost empty.  Requiring
    both an edge-aligned boundary and a strongly sparse adjacent band keeps a
    genuine table that was cropped tightly to its outer rules intact.
    """
    if len(columns) < 5:
        return columns

    height, width = image.shape[:2]
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
    threshold = int(max(120, min(210, float(np.percentile(gray, 85)) - 18)))

    def band_density(left: int, right: int) -> float:
        margin = max(2, min(5, (right - left) // 12))
        band = gray[:, left + margin : right - margin]
        if band.size == 0:
            return 1.0
        return float(np.mean(band < threshold))

    trimmed = sorted(set(int(value) for value in columns))
    densities = [band_density(left, right) for left, right in zip(trimmed[:-1], trimmed[1:])]
    if len(densities) < 3:
        return trimmed
    interior = densities[1:-1] or densities
    typical_density = float(np.median(interior))
    sparse_limit = min(0.035, typical_density * 0.30)
    horizontal_densities: list[float] = []
    if horizontal is not None and horizontal.shape[:2] == gray.shape[:2]:
        for left, right in zip(trimmed[:-1], trimmed[1:]):
            margin = max(2, min(5, (right - left) // 12))
            band = horizontal[:, left + margin : right - margin]
            horizontal_densities.append(
                1.0 if band.size == 0 else float(np.mean(band > 0))
            )
    horizontal_interior = (
        horizontal_densities[1:-1]
        if len(horizontal_densities) > 3
        else horizontal_densities
    )
    typical_horizontal_density = (
        float(np.median(horizontal_interior)) if horizontal_interior else 0.0
    )

    def is_sparse_page_band(index: int) -> bool:
        gray_sparse = densities[index] < sparse_limit
        grid_sparse = bool(
            horizontal_densities
            and typical_horizontal_density > 0
            and horizontal_densities[index] < typical_horizontal_density * 0.30
            and densities[index] < min(0.08, typical_density * 0.45)
        )
        return gray_sparse or grid_sparse

    edge_limit = max(3, int(round(width * 0.01)))
    paired_edge_limit = max(edge_limit, int(round(width * 0.03)))

    paired_sparse_page_edges = bool(
        trimmed[0] <= paired_edge_limit
        and trimmed[-1] >= width - 1 - paired_edge_limit
        and is_sparse_page_band(0)
        and is_sparse_page_band(-1)
    )
    if paired_sparse_page_edges:
        trimmed = trimmed[1:-1]
        densities = densities[1:-1]
        if horizontal_densities:
            horizontal_densities = horizontal_densities[1:-1]

    if trimmed and trimmed[0] <= edge_limit and densities and is_sparse_page_band(0):
        trimmed.pop(0)
        densities.pop(0)
        if horizontal_densities:
            horizontal_densities.pop(0)
    if (
        len(trimmed) >= 5
        and trimmed[-1] >= width - 1 - edge_limit
        and densities
        and is_sparse_page_band(-1)
    ):
        trimmed.pop()
    return trimmed


def _collapse_double_row_boundaries(
    rows: list[int],
    horizontal: np.ndarray,
) -> list[int]:
    """Collapse a very narrow false row to its better-supported rule."""
    if len(rows) < 5:
        return rows
    collapsed = sorted(set(int(value) for value in rows))
    while len(collapsed) >= 5:
        gaps = np.diff(collapsed)
        ordinary = gaps[gaps >= 8]
        if ordinary.size == 0:
            break
        typical = float(np.median(ordinary))
        narrow_indices = np.flatnonzero(gaps <= max(7.0, typical * 0.22))
        if narrow_indices.size == 0:
            break
        index = int(narrow_indices[0])

        def support(position: int) -> int:
            top = max(0, position - 2)
            bottom = min(horizontal.shape[0], position + 3)
            return int(np.count_nonzero(horizontal[top:bottom, :]))

        left = collapsed[index]
        right = collapsed[index + 1]
        # Two close detections describe one thick/duplicated separator.  Keep
        # the rule with greater horizontal evidence; ties prefer the latter so
        # the following data row does not absorb header text.
        remove_index = index if support(left) <= support(right) else index + 1
        collapsed.pop(remove_index)
    return collapsed


def _trim_sparse_trailing_page_row(
    image: np.ndarray,
    rows: list[int],
    vertical: np.ndarray | None = None,
) -> list[int]:
    """Remove a blank paper margin below a photographed table."""
    if len(rows) < 5:
        return rows
    height, width = image.shape[:2]
    ordered = sorted(set(int(value) for value in rows))
    edge_limit = max(3, int(round(height * 0.01)))
    near_image_edge = ordered[-1] >= height - 1 - edge_limit

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
    threshold = int(max(120, min(210, float(np.percentile(gray, 85)) - 18)))

    def band_density(top: int, bottom: int) -> float:
        margin = max(2, min(5, (bottom - top) // 12))
        band = gray[top + margin : bottom - margin, :]
        if band.size == 0:
            return 1.0
        return float(np.mean(band < threshold))

    densities = [band_density(top, bottom) for top, bottom in zip(ordered[:-1], ordered[1:])]
    body_densities = densities[1:-1] if len(densities) > 3 else densities[:-1]
    typical_density = float(np.median(body_densities)) if body_densities else 0.0
    grid_sparse = False
    if vertical is not None and vertical.shape[:2] == gray.shape[:2]:
        vertical_densities: list[float] = []
        for top, bottom in zip(ordered[:-1], ordered[1:]):
            margin = max(2, min(5, (bottom - top) // 12))
            band = vertical[top + margin : bottom - margin, :]
            vertical_densities.append(
                1.0 if band.size == 0 else float(np.mean(band > 0))
            )
        vertical_body = (
            vertical_densities[1:-1]
            if len(vertical_densities) > 3
            else vertical_densities[:-1]
        )
        typical_vertical_density = (
            float(np.median(vertical_body)) if vertical_body else 0.0
        )
        gaps = np.diff(ordered)
        ordinary_gaps = gaps[:-1] if gaps.size > 1 else gaps
        typical_gap = float(np.median(ordinary_gaps)) if ordinary_gaps.size else 0.0
        grid_sparse = bool(
            typical_vertical_density > 0
            and vertical_densities[-1] < typical_vertical_density * 0.30
            and gaps[-1] >= typical_gap * 1.45
            and densities[-1] < min(0.08, typical_density * 0.45)
        )
        line_proven_page_margin = bool(
            near_image_edge
            and typical_vertical_density > 0
            and vertical_densities[-1] < typical_vertical_density * 0.15
            and gaps[-1] >= typical_gap * 1.60
        )
    else:
        line_proven_page_margin = False
    if (
        (near_image_edge and densities[-1] < min(0.035, typical_density * 0.30))
        or grid_sparse
        or line_proven_page_margin
    ):
        ordered.pop()
    return ordered


def _leading_interval_has_text_components(
    image: np.ndarray,
    top: int,
    bottom: int,
    horizontal: np.ndarray | None,
    vertical: np.ndarray | None,
) -> bool:
    """Distinguish a sparse merged heading from an empty photographed margin."""
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
    margin = max(3, min(7, (bottom - top) // 10))
    band_top = max(0, top + margin)
    band_bottom = min(gray.shape[0], bottom - margin)
    if band_bottom - band_top < 8:
        return False
    band = gray[band_top:band_bottom].copy()
    local_reference = float(np.percentile(band, 78))
    threshold = int(max(55, min(205, local_reference - 24)))
    ink = (band < threshold).astype(np.uint8) * 255
    if horizontal is not None and horizontal.shape[:2] == gray.shape[:2]:
        ink[horizontal[band_top:band_bottom] > 0] = 0
    if vertical is not None and vertical.shape[:2] == gray.shape[:2]:
        ink[vertical[band_top:band_bottom] > 0] = 0
    ink = cv2.morphologyEx(
        ink,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2)),
    )
    count, _, stats, _ = cv2.connectedComponentsWithStats(ink, connectivity=8)
    band_height, band_width = ink.shape[:2]
    components: list[tuple[int, int, int, int, int]] = []
    for index in range(1, count):
        x, y, width, height, area = [int(value) for value in stats[index]]
        if (
            area >= max(5, int(round(band_height * 0.08)))
            and height >= max(4, int(round(band_height * 0.14)))
            and height <= band_height * 0.90
            and width >= 2
            and area / float(max(1, width * height)) >= 0.08
        ):
            components.append((x, y, width, height, area))
    if not components:
        return False
    ink_left = min(component[0] for component in components)
    ink_right = max(component[0] + component[2] for component in components)
    occupied_rows = float(np.mean(np.any(ink > 0, axis=1)))
    return bool(
        occupied_rows >= 0.12
        and (
            len(components) >= 2 and ink_right - ink_left >= band_width * 0.05
            or any(
                component[2] >= band_width * 0.08
                for component in components
            )
        )
    )


def _trim_sparse_leading_page_row(
    image: np.ndarray,
    rows: list[int],
    horizontal: np.ndarray | None = None,
    vertical: np.ndarray | None = None,
) -> list[int]:
    """Remove a blank paper margin above a photographed table."""
    if len(rows) < 5:
        return rows
    height, _ = image.shape[:2]
    ordered = sorted(set(int(value) for value in rows))
    edge_limit = max(3, int(round(height * 0.01)))
    leading_has_text = _leading_interval_has_text_components(
        image,
        ordered[0],
        ordered[1],
        horizontal,
        vertical,
    )
    if ordered[0] > edge_limit and horizontal is not None and len(ordered) >= 6:
        gaps = np.diff(ordered)
        ordinary = gaps[gaps >= 8]
        typical = float(np.median(ordinary)) if ordinary.size else 0.0
        if typical >= 8 and gaps[0] >= typical * 1.65:
            first_band = np.any(
                horizontal[
                    max(0, ordered[0] - 3) : min(height, ordered[0] + 4)
                ] > 0,
                axis=0,
            )
            next_band = np.any(
                horizontal[
                    max(0, ordered[1] - 3) : min(height, ordered[1] + 4)
                ] > 0,
                axis=0,
            )
            next_span = _dominant_supported_span(
                next_band,
                maximum_hole=max(12, image.shape[1] // 80),
            )
            if next_span is not None and next_span[1] - next_span[0] >= image.shape[1] * 0.35:
                overlap = float(np.mean(first_band[next_span[0] : next_span[1] + 1]))
                if overlap < 0.65 and not leading_has_text:
                    ordered.pop(0)
                    return ordered
    if ordered[0] > edge_limit:
        return ordered

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image
    threshold = int(max(120, min(210, float(np.percentile(gray, 85)) - 18)))

    def band_density(top: int, bottom: int) -> float:
        margin = max(2, min(5, (bottom - top) // 12))
        band = gray[top + margin : bottom - margin, :]
        if band.size == 0:
            return 1.0
        return float(np.mean(band < threshold))

    densities = [band_density(top, bottom) for top, bottom in zip(ordered[:-1], ordered[1:])]
    body_densities = densities[1:-1] if len(densities) > 3 else densities[1:]
    typical_density = float(np.median(body_densities)) if body_densities else 0.0
    grid_sparse = False
    if (
        vertical is not None
        and horizontal is not None
        and vertical.shape[:2] == gray.shape[:2]
        and len(ordered) >= 6
    ):
        vertical_densities: list[float] = []
        for top, bottom in zip(ordered[:-1], ordered[1:]):
            margin = max(2, min(5, (bottom - top) // 12))
            band = vertical[top + margin : bottom - margin, :]
            vertical_densities.append(
                1.0 if band.size == 0 else float(np.mean(band > 0))
            )
        body_vertical = vertical_densities[2:-1] or vertical_densities[1:]
        typical_vertical = float(np.median(body_vertical)) if body_vertical else 0.0
        gaps = np.diff(ordered)
        ordinary_gaps = gaps[1:] if gaps.size > 1 else gaps
        typical_gap = float(np.median(ordinary_gaps)) if ordinary_gaps.size else 0.0
        next_band = np.any(
            horizontal[
                max(0, ordered[1] - 3) : min(height, ordered[1] + 4)
            ] > 0,
            axis=0,
        )
        next_span = _dominant_supported_span(
            next_band,
            maximum_hole=max(12, image.shape[1] // 80),
        )
        grid_sparse = bool(
            typical_vertical > 0
            and vertical_densities[0] < typical_vertical * 0.30
            and typical_gap * 1.20 <= gaps[0] <= typical_gap * 1.50
            and densities[0] < min(0.08, typical_density * 0.45)
            and next_span is not None
            and next_span[1] - next_span[0] >= image.shape[1] * 0.35
        )
    if (
        densities[0] < min(0.035, typical_density * 0.30)
        or grid_sparse
    ) and not leading_has_text:
        ordered.pop(0)
    return ordered


def _trim_thin_leading_frame_row(
    image: np.ndarray,
    rows: list[int],
    horizontal: np.ndarray | None,
    vertical: np.ndarray | None,
) -> list[int]:
    """Remove a narrow paper-frame strip above the real first table row."""
    if len(rows) < 6:
        return rows
    ordered = sorted(set(int(value) for value in rows))
    gaps = np.diff(ordered)
    ordinary = gaps[1:]
    ordinary = ordinary[ordinary >= 8]
    if ordinary.size == 0:
        return ordered
    typical = float(np.median(ordinary))
    height = image.shape[0]
    near_edge = ordered[0] <= max(3, int(round(height * 0.01)))
    if (
        near_edge
        and typical * 0.20 <= gaps[0] <= typical * 0.55
        and not _leading_interval_has_text_components(
            image,
            ordered[0],
            ordered[1],
            horizontal,
            vertical,
        )
    ):
        ordered.pop(0)
    return ordered


def extract_embedded_spreadsheet_grid(
    image: np.ndarray,
) -> tuple[list[int], list[int], np.ndarray] | None:
    """Find a dense worksheet table below application chrome.

    Direct Excel/WPS captures can contain more strong edges in the ribbon than
    in the worksheet.  This path accepts only a long, regular run of horizontal
    rules and vertical rules that persist through that run.  It therefore
    supplies geometry without inferring any cell text.
    """
    height, width = image.shape[:2]
    if width < 800 or height < 400:
        return None
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image.copy()
    gray_float = gray.astype(np.float32)

    left = max(1, int(round(width * 0.02)))
    right = min(width - 1, int(round(width * 0.98)))
    horizontal_difference = np.abs(np.diff(gray_float[:, left:right], axis=0))
    horizontal_coverage = np.mean(horizontal_difference > 8.0, axis=1)
    row_candidates = _merge_nearby_centers(
        np.flatnonzero(horizontal_coverage >= 0.70).astype(int).tolist(),
        maximum_gap=6,
    )
    row_candidates = [
        row for row in row_candidates if height * 0.08 < row < height * 0.98
    ]
    if len(row_candidates) < 8:
        return None

    best_rows: list[int] = []
    best_typical_gap = 0.0
    candidate_gaps = [
        right_row - left_row
        for left_row, right_row in zip(row_candidates, row_candidates[1:])
        if 8 <= right_row - left_row <= 80
    ]
    for start in range(len(row_candidates) - 7):
        for seed_gap in candidate_gaps:
            run = [row_candidates[start]]
            typical_gap = float(seed_gap)
            for row in row_candidates[start + 1 :]:
                gap = row - run[-1]
                if typical_gap * 0.65 <= gap <= typical_gap * 1.40:
                    run.append(row)
                    typical_gap = float(np.median(np.diff(run)))
                else:
                    break
            if len(run) > len(best_rows):
                best_rows = run
                best_typical_gap = typical_gap
    if (
        len(best_rows) < 8
        or best_rows[-1] - best_rows[0] < height * 0.25
        or best_typical_gap < 8.0
    ):
        return None

    preceding_rows = [row for row in row_candidates if row < best_rows[0]]
    if preceding_rows:
        title_gap = best_rows[0] - preceding_rows[-1]
        if best_typical_gap * 1.5 <= title_gap <= best_typical_gap * 2.5:
            best_rows.insert(0, preceding_rows[-1])

    body_top = best_rows[0]
    if len(best_rows) >= 3 and best_rows[1] - best_rows[0] > best_typical_gap * 1.5:
        body_top = best_rows[1]
    vertical_difference = np.abs(np.diff(gray_float[body_top : best_rows[-1]], axis=1))
    vertical_coverage = np.mean(vertical_difference > 8.0, axis=0)
    columns = _merge_nearby_centers(
        np.flatnonzero(vertical_coverage >= 0.70).astype(int).tolist(),
        maximum_gap=6,
    )
    columns = [
        column for column in columns if width * 0.01 < column < width * 0.99
    ]
    if (
        not 4 <= len(columns) <= 33
        or columns[-1] - columns[0] < width * 0.75
        or min(np.diff(columns), default=0) < max(12, int(round(width * 0.012)))
    ):
        return None

    mask = np.zeros((height, width), dtype=np.uint8)
    for row in best_rows:
        cv2.line(mask, (columns[0], row), (columns[-1], row), 255, 3)
    for index, column in enumerate(columns):
        thickness = 3 if index in (0, len(columns) - 1) else 7
        cv2.line(mask, (column, best_rows[0]), (column, best_rows[-1]), 255, thickness)
    cleaned = _enhance_for_ocr(image)
    cleaned[mask > 0] = 255
    return columns, best_rows, cleaned


def extract_screen_grid(
    image: np.ndarray,
    *,
    vertical_statistics: tuple[
        np.ndarray,
        np.ndarray,
        np.ndarray,
        np.ndarray,
    ]
    | None = None,
) -> tuple[list[int], list[int], np.ndarray] | None:
    height, width = image.shape[:2]
    if width < 800 or height < 250:
        return None

    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image.copy()
    transition_rows = _repeated_transition_centers(gray, rows=True)
    transition_columns = _repeated_transition_centers(gray, rows=False)
    transition_rows = _recover_merged_header_rows(transition_rows)

    background = float(np.percentile(gray, 90))
    threshold = int(max(180, min(245, background - 20)))
    binary = cv2.threshold(gray, threshold, 255, cv2.THRESH_BINARY_INV)[1]
    horizontal = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (max(18, width // 30), 1)),
    )
    vertical = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(18, height // 30))),
    )

    adaptive_horizontal, adaptive_vertical, _ = _grid_maps(image)
    transition_rows = _remove_false_transition_rows(transition_rows, adaptive_horizontal, width)
    vertical_projection = np.maximum(
        np.count_nonzero(vertical, axis=0),
        np.count_nonzero(adaptive_vertical, axis=0),
    )
    transition_columns = _filter_transition_columns_by_vertical_support(
        transition_columns,
        vertical_projection,
        height,
        width,
    )
    vertical_edge_statistics = (
        vertical_statistics
        if vertical_statistics is not None
        else _vertical_edge_statistics(gray)
    )
    raw_consistent_columns = _consistent_vertical_edge_centers(
        gray,
        statistics=vertical_edge_statistics,
    )
    if len(raw_consistent_columns) >= len(transition_columns) * 0.9:
        selected_columns = transition_columns
    else:
        consistent_columns = [
            column
            for column in raw_consistent_columns
            if int(vertical_projection[max(0, column - 3) : min(width, column + 4)].max(initial=0))
            >= height * 0.5
        ]
        selected_columns = _select_screen_columns(transition_columns, consistent_columns, width)
    if selected_columns != transition_columns:
        transition_columns = selected_columns
    elif len(transition_columns) >= 4:
        gaps = np.diff(transition_columns)
        typical_gap = float(np.median(gaps))
        largest_index = int(np.argmax(gaps))
        if gaps[largest_index] > typical_gap * 1.7:
            left = transition_columns[largest_index]
            right = transition_columns[largest_index + 1]
            supplemental = _consistent_vertical_edge_centers(
                gray,
                minimum_median=30.0,
                statistics=vertical_edge_statistics,
            )
            inside = [column for column in supplemental if left + 12 < column < right - 12]
            if inside:
                midpoint = (left + right) / 2.0
                transition_columns = _merge_nearby_centers(
                    transition_columns + [min(inside, key=lambda column: abs(column - midpoint))]
                )
    if len(transition_columns) >= 10:
        transition_columns = _recover_single_missing_boundary(
            transition_columns,
            vertical_projection,
        )
    transition_columns = _recover_dense_spreadsheet_columns(
        gray,
        transition_columns,
        transition_rows,
        statistics=vertical_edge_statistics,
    )

    row_indices = np.flatnonzero(np.count_nonzero(horizontal, axis=1) >= width * 0.7)
    column_indices = np.flatnonzero(np.count_nonzero(vertical, axis=0) >= height * 0.06)
    rows = [
        int(round(float(group.mean())))
        for group in np.split(row_indices, np.where(np.diff(row_indices) > 1)[0] + 1)
        if group.size
    ]
    columns = [
        int(round(float(group.mean())))
        for group in np.split(column_indices, np.where(np.diff(column_indices) > 1)[0] + 1)
        if group.size
    ]
    recovered_body_grid = False
    if len(columns) < 3 and len(rows) >= 8 and len(transition_columns) < 3:
        adaptive_horizontal, adaptive_vertical, _ = _grid_maps(image)
        adaptive_rows = _line_centers(np.count_nonzero(adaptive_horizontal, axis=1), width)
        adaptive_columns = _line_centers(
            np.count_nonzero(adaptive_vertical, axis=0), height
        )
        if len(adaptive_columns) >= 3 and len(transition_rows) < 4:
            return None
        if len(adaptive_rows) >= 8:
            rows = adaptive_rows
        consistent = _consistent_vertical_edge_centers(
            gray,
            minimum_median=20.0,
            statistics=vertical_edge_statistics,
        )
        profiled = _band_profile_vertical_centers(gray, rows)
        minimum_distance = max(20, width // 60)
        combined = list(consistent)
        for column in profiled:
            if all(abs(column - existing) >= minimum_distance for existing in consistent):
                combined.append(column)
        combined = _merge_nearby_centers(combined)
        if len(combined) >= 3:
            columns = [0] + combined + [width - 1]
            rows = _recover_merged_header_rows(rows)
            recovered_body_grid = True

    if not recovered_body_grid:
        rows = [row for row in rows if height * 0.01 <= row <= height * 0.98]
        columns = [column for column in columns if width * 0.01 <= column <= width * 0.99]
    if (
        len(transition_rows) >= max(4, len(rows))
        and len(transition_columns) >= 3
        and (
            len(transition_columns) >= len(columns)
            or len(columns) > len(transition_columns) * 1.5
        )
    ):
        rows = transition_rows
        columns = transition_columns
    if len(rows) < 4 or len(columns) < 3:
        return None

    row_gaps = np.diff(rows)
    typical_row_height = float(np.median(row_gaps))
    row_tolerance = 5 if recovered_body_grid else 2
    if typical_row_height < 8 or np.mean(np.abs(row_gaps - typical_row_height) <= row_tolerance) < 0.65:
        return None

    # 返回屏幕网格快捷结果前，先执行电子表格标尺交叉校验。否则 extract_ruled_grid()
    # 会立即接受屏幕网格，后续过度分列修复永远没有机会执行。密集文字包含大量重复竖画，
    # 标尺与表体的组合证据用于区分文字笔画和真实工作表边界。
    spreadsheet_grid = _recover_spreadsheet_ruler_grid(
        gray,
        columns,
        rows,
        vertical_statistics=vertical_edge_statistics,
    )
    if spreadsheet_grid is not None:
        columns, rows = spreadsheet_grid

    # Screen captures can contain a full-width decorative page frame and one
    # faded worksheet rule at the same time.  Repair the physically supported
    # double-height gap first, then remove only the uniquely evidenced frame;
    # otherwise the apparent row count stays unchanged while every value is
    # shifted and one pair of data rows is fused.
    rows = _regularize_ruled_rows(
        rows,
        adaptive_horizontal,
        adaptive_vertical,
        height,
        width,
        image,
    )
    rows = _trim_decorative_screen_top_frame(image, columns, rows)

    mask = np.zeros((height, width), dtype=np.uint8)
    for row in rows:
        cv2.line(mask, (columns[0], row), (columns[-1], row), 255, 3)
    for index, column in enumerate(columns):
        thickness = 3 if index in (0, len(columns) - 1) else 7
        cv2.line(mask, (column, rows[0]), (column, rows[-1]), 255, thickness)
    cleaned = _enhance_for_ocr(image)
    cleaned[mask > 0] = 255
    return columns, rows, cleaned


def _low_contrast_grid_supersedes_screen_grid(
    candidate_columns: list[int],
    candidate_rows: list[int],
    screen_columns: list[int],
    screen_rows: list[int],
    *,
    width: int,
) -> bool:
    """Prefer physical low-contrast rules over a text-stroke screen candidate."""
    candidate_column_count = len(candidate_columns) - 1
    candidate_row_count = len(candidate_rows) - 1
    screen_column_count = len(screen_columns) - 1
    screen_row_count = len(screen_rows) - 1
    if not (
        3 <= candidate_column_count <= 32
        and 12 <= candidate_row_count <= 128
        and min(np.diff(candidate_columns), default=0) >= 6
        and min(np.diff(candidate_rows), default=0) >= 6
    ):
        return False

    row_gaps = np.diff(candidate_rows).astype(float)
    ordinary_row_gap = float(np.median(row_gaps)) if row_gaps.size else 0.0
    regular_row_ratio = float(
        np.mean(
            np.abs(row_gaps - ordinary_row_gap)
            <= max(6.0, ordinary_row_gap * 0.30)
        )
    )
    if ordinary_row_gap < 8.0 or regular_row_ratio < 0.70:
        return False

    screen_is_text_stroke_over_split = bool(
        screen_column_count >= max(candidate_column_count + 8, 2 * candidate_column_count)
        and candidate_row_count >= screen_row_count
    )
    if screen_is_text_stroke_over_split:
        return True

    if not (
        candidate_column_count == screen_column_count + 1
        and candidate_row_count == screen_row_count
        and len(candidate_columns) >= 3
    ):
        return False
    tolerance = max(8.0, float(width) * 0.012)
    typical_column_gap = float(np.median(np.diff(candidate_columns)))
    restores_left_outer_column = bool(
        float(screen_columns[0]) - float(candidate_columns[0])
        >= typical_column_gap * 0.65
        and all(
            any(abs(float(screen) - float(candidate)) <= tolerance for candidate in candidate_columns[1:])
            for screen in screen_columns
        )
    )
    return restores_left_outer_column


def _extract_ruled_grid_uncached(
    image: np.ndarray,
    *,
    prefer_adaptive: bool = False,
) -> tuple[list[int], list[int], np.ndarray] | None:
    height, width = image.shape[:2]
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image.copy()
    vertical_edge_statistics = (
        _vertical_edge_statistics(gray) if width >= 800 and height >= 250 else None
    )
    screen_grid = extract_screen_grid(
        image,
        vertical_statistics=vertical_edge_statistics,
    )
    if screen_grid is not None and not prefer_adaptive:
        return screen_grid
    horizontal, vertical, _ = _grid_maps(image)
    columns = _line_centers(np.count_nonzero(vertical, axis=0), height)
    rows = _line_centers(np.count_nonzero(horizontal, axis=1), width)
    sparse_columns_recovered = False
    low_contrast_grid_selected = False
    if screen_grid is not None and prefer_adaptive:
        recovered_columns, recovered_rows = _recover_sparse_low_contrast_grid(
            image,
            horizontal,
            vertical,
            list(columns),
            list(rows),
        )
        screen_columns, screen_rows, _ = screen_grid
        if _low_contrast_grid_supersedes_screen_grid(
            recovered_columns,
            recovered_rows,
            screen_columns,
            screen_rows,
            width=width,
        ):
            columns = recovered_columns
            rows = recovered_rows
            screen_grid = None
            sparse_columns_recovered = True
            low_contrast_grid_selected = True
    if screen_grid is not None:
        screen_columns, screen_rows, _ = screen_grid
        adaptive_is_usable = len(columns) >= 3 and len(rows) >= 4
        same_shape = (
            len(columns) == len(screen_columns)
            and len(rows) == len(screen_rows)
        )
        comparison_screen_rows = _collapse_double_row_boundaries(
            screen_rows,
            horizontal,
        )
        adaptive_recovers_rows_and_rejects_false_columns = (
            adaptive_is_usable
            and len(rows) >= len(comparison_screen_rows)
            and len(columns) < len(screen_columns)
        )
        boundary_tolerance = max(4, int(round(min(height, width) * 0.006)))
        adaptive_preserves_screen_interior = bool(
            adaptive_is_usable
            and all(
                any(abs(candidate - value) <= boundary_tolerance for candidate in columns)
                for value in screen_columns[1:-1]
            )
            and all(
                any(abs(candidate - value) <= boundary_tolerance for candidate in rows)
                for value in comparison_screen_rows[1:-1]
            )
        )
        adaptive_restores_clipped_frame = bool(
            adaptive_preserves_screen_interior
            and len(columns) >= len(screen_columns)
            and len(rows) >= len(comparison_screen_rows)
            and (
                columns[0] < screen_columns[0] - boundary_tolerance
                or columns[-1] > screen_columns[-1] + boundary_tolerance
                or rows[0] < comparison_screen_rows[0] - boundary_tolerance
                or rows[-1] > comparison_screen_rows[-1] + boundary_tolerance
            )
        )
        # A clipped, high-contrast spreadsheet can make character strokes look
        # like full-height screen separators.  In that case the adaptive grid
        # contains fewer columns, but every one of its interior boundaries is
        # still supported by the denser screen candidate, while its row lines
        # preserve the screen rhythm and restore the missing outer frame.  The
        # adaptive candidate is therefore the complete physical grid; the
        # additional screen columns are text-stroke false positives.
        adaptive_removes_false_screen_splits = bool(
            adaptive_is_usable
            and len(rows) >= len(comparison_screen_rows) + 2
            and len(columns) >= 3
            and len(columns) <= len(screen_columns) - 2
            and all(
                any(abs(candidate - value) <= boundary_tolerance for candidate in screen_columns)
                for value in columns[1:-1]
            )
            and all(
                any(abs(candidate - value) <= boundary_tolerance for candidate in rows)
                for value in comparison_screen_rows
            )
            and columns[0] < screen_columns[0] - boundary_tolerance
            and columns[-1] >= screen_columns[-1] - boundary_tolerance
            and rows[0] < comparison_screen_rows[0] - boundary_tolerance
            and rows[-1] > comparison_screen_rows[-1] + boundary_tolerance
        )
        outer_page_frame_candidate = bool(
            len(screen_columns) >= 5
            and len(screen_rows) >= 5
            and screen_columns[0] <= max(3, int(round(width * 0.01)))
            and screen_columns[-1] >= width - 1 - max(3, int(round(width * 0.01)))
            and screen_rows[-1] >= height * 0.95
        )
        if (
            not same_shape
            and not adaptive_recovers_rows_and_rejects_false_columns
            and not adaptive_restores_clipped_frame
            and not adaptive_removes_false_screen_splits
        ):
            if not outer_page_frame_candidate:
                return screen_grid
            columns = list(screen_columns)
            rows = list(screen_rows)
    if screen_grid is None:
        original_columns = list(columns)
        if not low_contrast_grid_selected:
            columns, rows = _recover_sparse_low_contrast_grid(
                image,
                horizontal,
                vertical,
                columns,
                rows,
            )
            sparse_columns_recovered = columns != original_columns
    if not prefer_adaptive and len(columns) >= 4 and len(rows) >= 4:
        column_gaps = np.diff(columns)
        row_gaps = np.diff(rows)
        if (
            column_gaps[0] < float(np.median(column_gaps[1:])) * 0.5
            and row_gaps[0] < float(np.median(row_gaps[1:])) * 0.8
        ):
            header_columns = _header_transition_columns(gray, rows[0], rows[1])
            ruler_columns = _confirmed_spreadsheet_ruler_columns(gray, rows)
            if len(ruler_columns) > len(header_columns):
                header_columns = ruler_columns
            if len(header_columns) >= 3:
                columns = header_columns
                rows = _recover_single_missing_boundary(rows[1:])
    # Thick rules and illumination gradients can produce two detections only a
    # few pixels apart.  Collapse them before estimating the ordinary row
    # height; otherwise the false short gap can hide a missing title/header
    # separator from the regularizer.
    rows = _collapse_double_row_boundaries(rows, horizontal)
    rows = _regularize_ruled_rows(rows, horizontal, vertical, height, width, image)
    rows = _recover_partial_header_row_boundaries(
        rows,
        horizontal,
        width,
        image,
    )
    rows = _recover_regular_missing_boundaries(
        rows,
        np.count_nonzero(horizontal, axis=1),
        width,
        maximum_multiple=16,
        minimum_support_ratio=0.10,
        minimum_global_peak_ratio=0.10,
    )
    rows = _recover_visible_double_row_boundary(image, rows)
    vertical_projection = np.count_nonzero(vertical, axis=0)
    if columns and vertical_projection.size:
        strongest_column = max(
            columns,
            key=lambda column: int(
                vertical_projection[
                    max(0, column - 3) : min(width, column + 4)
                ].max(initial=0)
            ),
        )
        vertical_band = np.any(
            vertical[:, max(0, strongest_column - 3) : min(width, strongest_column + 4)] > 0,
            axis=1,
        )
        table_span = _dominant_supported_span(
            vertical_band,
            maximum_hole=max(50, height // 10),
        )
        if table_span is not None:
            rows = _extend_regular_boundaries_with_evidence(
                rows,
                np.count_nonzero(horizontal, axis=1),
                width,
                max(0, table_span[0] - 12),
                min(height - 1, table_span[1] + 12),
            )
    columns = _recover_regular_missing_boundaries(
        columns,
        np.count_nonzero(vertical, axis=0),
        height,
    )
    body_columns = _body_band_consensus_columns(vertical, horizontal, rows)
    columns = _prefer_body_consensus_columns(columns, body_columns)
    columns = _remove_weak_split_columns(columns, vertical)
    if not low_contrast_grid_selected:
        columns = _recover_supported_irregular_boundaries(
            columns,
            _consistent_vertical_edge_centers(
                gray,
                minimum_median=20.0,
                statistics=vertical_edge_statistics,
            ),
            np.count_nonzero(vertical, axis=0),
            height,
            vertical=vertical,
            rows=rows,
        )
    columns = _remove_weak_split_columns(columns, vertical)
    columns = _recover_outer_columns_from_horizontal_endpoints(
        image,
        horizontal,
        columns,
        rows,
    )
    columns = _recover_leading_spreadsheet_gutter_boundary(
        columns,
        rows,
        vertical,
    )
    if not sparse_columns_recovered or low_contrast_grid_selected:
        columns = _recover_crop_edge_boundaries(columns, width)
    columns = _trim_sparse_page_edge_columns(image, columns, horizontal)
    rows = _recover_crop_edge_boundaries(rows, height)
    rows = _collapse_double_row_boundaries(rows, horizontal)
    rows = _trim_thin_leading_frame_row(image, rows, horizontal, vertical)
    rows = _trim_sparse_leading_page_row(image, rows, horizontal, vertical)
    rows = _trim_sparse_trailing_page_row(image, rows, vertical)
    columns, rows = _recover_clipped_frame_boundaries(
        image,
        horizontal,
        vertical,
        columns,
        rows,
    )
    columns = _trim_proven_empty_crop_edge_column(
        image,
        horizontal,
        columns,
        rows,
    )
    columns, rows = _trim_disconnected_outer_frame_cells(
        image,
        horizontal,
        vertical,
        columns,
        rows,
    )
    spreadsheet_grid = _recover_spreadsheet_ruler_grid(
        gray,
        columns,
        rows,
        vertical_statistics=vertical_edge_statistics,
    )
    if spreadsheet_grid is not None:
        columns, rows = spreadsheet_grid
    if len(columns) < 3 or len(rows) < 3:
        return None
    if min(np.diff(columns), default=0) < 6 or min(np.diff(rows), default=0) < 6:
        return None

    mask = cv2.bitwise_or(horizontal, vertical)
    mask = cv2.dilate(mask, cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3)))
    cleaned = image.copy()
    cleaned[mask > 0] = 255
    return columns, rows, cleaned


def extract_ruled_grid(
    image: np.ndarray,
    *,
    prefer_adaptive: bool = False,
) -> tuple[list[int], list[int], np.ndarray] | None:
    """Extract a ruled grid, reusing exact duplicate work within one request."""
    cache = _ruled_grid_request_cache
    if cache is None:
        return _extract_ruled_grid_uncached(
            image,
            prefer_adaptive=prefer_adaptive,
        )
    contiguous = image if image.flags.c_contiguous else np.ascontiguousarray(image)
    digest = hashlib.blake2b(memoryview(contiguous).cast("B"), digest_size=16).digest()
    key = (tuple(image.shape), str(image.dtype), bool(prefer_adaptive), digest)
    if key not in cache:
        result = _extract_ruled_grid_uncached(
            image,
            prefer_adaptive=prefer_adaptive,
        )
        cache[key] = (
            None
            if result is None
            else (tuple(result[0]), tuple(result[1]), result[2].copy())
        )
        return result
    cached = cache[key]
    if cached is None:
        return None
    columns, rows, cleaned = cached
    return list(columns), list(rows), cleaned.copy()


def assign_ocr_to_grid(
    columns: list[int],
    rows: list[int],
    boxes: Any,
    texts: Any,
    scores: Any,
    *,
    preserve_geometry: bool = False,
) -> tuple[list[list[str]], list[list[float]]]:
    row_count = max(0, len(rows) - 1)
    column_count = max(0, len(columns) - 1)
    entries: list[list[list[tuple[float, float, str, float]]]] = [
        [[] for _ in range(column_count)] for _ in range(row_count)
    ]
    if boxes is not None and texts is not None:
        score_values = list(scores or [])
        for index, (box, text) in enumerate(zip(boxes, texts)):
            points = np.asarray(box, dtype=float).reshape(-1, 2)
            center_x = float(points[:, 0].mean())
            center_y = float(points[:, 1].mean())
            row = int(np.searchsorted(rows, center_y, side="right") - 1)
            score = float(score_values[index]) if index < len(score_values) else 0.0
            value = str(text).strip()
            left_x = float(points[:, 0].min())
            right_x = float(points[:, 0].max())
            top_y = float(points[:, 1].min())
            bottom_y = float(points[:, 1].max())
            box_width = max(1.0, right_x - left_x)
            box_height = max(1.0, bottom_y - top_y)
            frame_overlap_width = max(
                0.0,
                min(right_x, float(columns[-1]))
                - max(left_x, float(columns[0])),
            )
            frame_overlap_height = max(
                0.0,
                min(bottom_y, float(rows[-1]))
                - max(top_y, float(rows[0])),
            )
            if (
                frame_overlap_width / box_width < 0.70
                or frame_overlap_height / box_height < 0.70
            ):
                continue
            crossed = [
                boundary
                for boundary in columns[1:-1]
                if left_x < boundary < right_x
                and 0.12 <= (boundary - left_x) / box_width <= 0.88
            ]
            if len(crossed) == 1:
                overlap_ratio = (crossed[0] - left_x) / box_width
                if min(overlap_ratio, 1.0 - overlap_ratio) < 0.22:
                    crossed = []
            horizontal_split: tuple[float, int] | None = None
            if len(crossed) == 1 and value and right_x - left_x > 1:
                boundary = crossed[0]
                ratio = (boundary - left_x) / (right_x - left_x)
                whitespace = [
                    position
                    for position, character in enumerate(value)
                    if character.isspace()
                ]
                if 0.2 <= ratio <= 0.8 and whitespace:
                    target = ratio * len(value)
                    split_at = min(
                        whitespace, key=lambda position: abs(position - target)
                    )
                    if abs(split_at - target) <= max(2.0, len(value) * 0.2):
                        horizontal_split = (boundary, split_at)
            crossed_rows = [
                boundary_index
                for boundary_index, boundary in enumerate(rows[1:-1], start=1)
                if top_y < boundary < bottom_y
                and 0.12 <= (boundary - top_y) / box_height <= 0.88
            ]
            if crossed_rows and value and horizontal_split is None:
                first_row = crossed_rows[0] - 1
                last_row = crossed_rows[-1]
                covered_rows = list(range(first_row, last_row + 1))
                consecutive_boundaries = crossed_rows == list(
                    range(crossed_rows[0], crossed_rows[-1] + 1)
                )
                covered_heights = [rows[item + 1] - rows[item] for item in covered_rows]
                vertical_parts: list[str] = []
                whitespace_parts = value.split()
                if len(whitespace_parts) == len(covered_rows):
                    vertical_parts = whitespace_parts
                elif (
                    re.fullmatch(r"[\u3400-\u9fff]{2,8}", value)
                    and len(value) % len(covered_rows) == 0
                    and all(
                        abs(
                            (rows[boundary] - top_y) / box_height
                            - split_index / len(covered_rows)
                        )
                        <= 0.25
                        for split_index, boundary in enumerate(crossed_rows, start=1)
                    )
                ):
                    part_length = len(value) // len(covered_rows)
                    vertical_parts = [
                        value[offset : offset + part_length]
                        for offset in range(0, len(value), part_length)
                    ]
                column = int(np.searchsorted(columns, center_x, side="right") - 1)
                if (
                    consecutive_boundaries
                    and len(vertical_parts) == len(covered_rows)
                    and box_height >= min(covered_heights) * (len(covered_rows) - 0.75)
                    and 0 <= first_row
                    and last_row < row_count
                    and 0 <= column < column_count
                ):
                    for target_row, part in zip(covered_rows, vertical_parts):
                        entries[target_row][column].append(
                            ((rows[target_row] + rows[target_row + 1]) * 0.5, center_x, part, score)
                        )
                    continue
            if horizontal_split is not None:
                boundary, split_at = horizontal_split
                left_value = value[:split_at].strip()
                right_value = value[split_at:].strip()
                left_column = int(np.searchsorted(columns, boundary, side="left") - 1)
                right_column = left_column + 1
                if (
                    left_value
                    and right_value
                    and 0 <= row < row_count
                    and 0 <= left_column < column_count
                    and 0 <= right_column < column_count
                ):
                    entries[row][left_column].append(
                        (center_y, (left_x + boundary) / 2.0, left_value, score)
                    )
                    entries[row][right_column].append(
                        (center_y, (boundary + right_x) / 2.0, right_value, score)
                    )
                    continue

            if crossed and value and 0 <= row < row_count:
                # A recognizer may return one box spanning merged cells or two
                # neighbouring cells.  If there is no trustworthy whitespace
                # split, assigning by the box centre silently shifts the value
                # into a later column.  Keep the text at the left edge of the
                # covered range and require review instead.  A full-width first
                # row is the common table-title case and is safe to keep.
                covered_columns = [
                    column_index
                    for column_index in range(column_count)
                    if right_x > columns[column_index]
                    and left_x < columns[column_index + 1]
                ]
                if covered_columns:
                    target_column = covered_columns[0]
                    review_score = score if row == 0 and len(crossed) >= 2 else -1.0
                    entries[row][target_column].append(
                        (center_y, left_x, value, review_score)
                    )
                    continue

            column = int(np.searchsorted(columns, center_x, side="right") - 1)
            if not (0 <= row < row_count and 0 <= column < column_count):
                continue
            numeric_unit_pair = re.fullmatch(
                r"([+-]?\d+\.\d{2})(\d+(?:\.\d+)?\s*[kMGT]?S/s)",
                value,
                flags=re.IGNORECASE,
            )
            if numeric_unit_pair and column > 0 and not entries[row][column - 1]:
                entries[row][column - 1].append(
                    (center_y, columns[column] - 1.0, numeric_unit_pair.group(1), score)
                )
                entries[row][column].append(
                    (center_y, center_x, numeric_unit_pair.group(2), score)
                )
                continue
            entries[row][column].append((center_y, center_x, value, score))

    grid: list[list[str]] = []
    confidence: list[list[float]] = []
    for row_index, row_entries in enumerate(entries):
        grid_row: list[str] = []
        confidence_row: list[float] = []
        for cell_entries in row_entries:
            ordered_by_y = sorted(cell_entries)
            line_threshold = max(
                2.0,
                min(8.0, float(rows[row_index + 1] - rows[row_index]) * 0.18),
            )
            grouped_lines: list[list[tuple[float, float, str, float]]] = []
            line_centers: list[float] = []
            for entry in ordered_by_y:
                if not grouped_lines or abs(entry[0] - line_centers[-1]) > line_threshold:
                    grouped_lines.append([entry])
                    line_centers.append(entry[0])
                    continue
                grouped_lines[-1].append(entry)
                line_centers[-1] = float(
                    np.median([value[0] for value in grouped_lines[-1]])
                )
            ordered = [
                entry
                for line in grouped_lines
                for entry in sorted(line, key=lambda value: value[1])
            ]
            values = [entry[2] for entry in ordered if entry[2]]
            values_scores = [entry[3] for entry in ordered if entry[2]]
            grid_row.append(" ".join(values))
            confidence_row.append(float(np.mean(values_scores)) if values_scores else 0.0)
        grid.append(grid_row)
        confidence.append(confidence_row)
    # 已锁定的物理网格只能填充文字，不能根据 OCR 文本再次增删行。
    return (grid, confidence) if preserve_geometry else _split_fused_header_data_rows(
        grid,
        confidence,
    )


def _split_fused_header_data_rows(
    grid: list[list[str]],
    confidence: list[list[float]],
) -> tuple[list[list[str]], list[list[float]]]:
    """Split a header/data band that OCR returned as ``"编号 1"`` cells.

    This is deliberately a narrow structural guard.  It activates only when a
    majority of the populated cells in one of the first three rows share a
    label/value pattern and at least two suffixes look like data.  Ordinary
    multi-word cells therefore remain untouched.
    """
    output_grid = [list(row) for row in grid]
    output_confidence = [list(row) for row in confidence]
    scan_limit = min(3, len(output_grid))
    row_index = 0
    while row_index < scan_limit:
        row = output_grid[row_index]
        populated = [index for index, value in enumerate(row) if str(value).strip()]
        if len(populated) < 3:
            row_index += 1
            continue

        splits: dict[int, tuple[str, str]] = {}
        data_like = 0
        for column_index in populated:
            value = str(row[column_index]).strip()
            match = re.fullmatch(r"([^\s：:]{1,24})[\s：:]+(.+)", value)
            if not match:
                continue
            label, data = match.group(1).strip(), match.group(2).strip()
            if not label or not data or re.search(r"\d", label):
                continue
            splits[column_index] = (label, data)
            if re.search(r"[\d%+\-−—.]", data):
                data_like += 1

        required = max(3, int(np.ceil(len(populated) * 0.6)))
        if len(splits) < required or data_like < 2:
            row_index += 1
            continue

        header_row: list[str] = []
        data_row: list[str] = []
        header_scores: list[float] = []
        data_scores: list[float] = []
        source_scores = output_confidence[row_index]
        for column_index, value in enumerate(row):
            score = source_scores[column_index] if column_index < len(source_scores) else 0.0
            if column_index in splits:
                label, data = splits[column_index]
                header_row.append(label)
                data_row.append(data)
                header_scores.append(score)
                data_scores.append(score)
            else:
                header_row.append(value)
                data_row.append("")
                header_scores.append(score)
                data_scores.append(0.0)

        output_grid[row_index : row_index + 1] = [header_row, data_row]
        output_confidence[row_index : row_index + 1] = [header_scores, data_scores]
        scan_limit = min(3, len(output_grid))
        row_index += 2

    return output_grid, output_confidence


def _deskew(image: np.ndarray) -> tuple[np.ndarray, np.ndarray, float]:
    height, width = image.shape[:2]
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if image.ndim == 3 else image.copy()
    binary = cv2.adaptiveThreshold(
        gray,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        31,
        11,
    )
    lines = cv2.HoughLinesP(
        binary,
        1,
        np.pi / 360,
        max(40, width // 16),
        minLineLength=max(80, width // 7),
        maxLineGap=max(12, width // 35),
    )
    weighted_angles: list[float] = []
    if lines is not None:
        for x1, y1, x2, y2 in lines.reshape(-1, 4):
            angle = float(np.degrees(np.arctan2(y2 - y1, x2 - x1)))
            if abs(angle) <= 25:
                length = float(np.hypot(x2 - x1, y2 - y1))
                weighted_angles.extend([angle] * max(1, int(length // 40)))
    angle = float(np.median(weighted_angles)) if weighted_angles else 0.0
    if abs(angle) < 0.15:
        angle = 0.0

    center = (width / 2.0, height / 2.0)
    transform = cv2.getRotationMatrix2D(center, angle, 1.0)
    cosine = abs(transform[0, 0])
    sine = abs(transform[0, 1])
    target_width = int(height * sine + width * cosine)
    target_height = int(height * cosine + width * sine)
    transform[0, 2] += target_width / 2.0 - center[0]
    transform[1, 2] += target_height / 2.0 - center[1]
    rotated = cv2.warpAffine(
        image,
        transform,
        (target_width, target_height),
        flags=cv2.INTER_CUBIC,
        borderValue=(255, 255, 255),
    )
    return rotated, transform, angle


def _extend_grid_bounds_along_vertical_rules(
    vertical: np.ndarray,
    bounds: tuple[int, int, int, int],
) -> tuple[int, int, int, int]:
    """Extend a partial grid crop when its vertical rules continue beyond it."""
    x, y, width, height = bounds
    image_height, image_width = vertical.shape[:2]
    left = max(0, x)
    right = min(image_width, x + width)
    if right - left < 20:
        return bounds
    support = np.count_nonzero(vertical[:, left:right], axis=1)
    threshold = max(3, int(round((right - left) * 0.003)))
    supported = np.flatnonzero(support >= threshold)
    if supported.size < 12:
        return bounds

    runs: list[tuple[int, int]] = []
    start = previous = int(supported[0])
    for value in supported[1:]:
        current = int(value)
        if current - previous > 5:
            runs.append((start, previous))
            start = current
        previous = current
    runs.append((start, previous))
    candidate_top = y
    candidate_bottom = y + height - 1
    minimum_overlap = max(12, int(round(height * 0.20)))
    matching = [
        (run_start, run_end)
        for run_start, run_end in runs
        if max(0, min(candidate_bottom, run_end) - max(candidate_top, run_start) + 1)
        >= minimum_overlap
    ]
    if not matching:
        return bounds
    run_start, run_end = max(
        matching,
        key=lambda run: min(candidate_bottom, run[1]) - max(candidate_top, run[0]),
    )
    extended_top = min(candidate_top, run_start)
    extended_bottom = max(candidate_bottom, run_end)
    if extended_bottom - extended_top + 1 <= height * 1.03:
        return bounds
    return x, extended_top, width, extended_bottom - extended_top + 1


def _full_frame_perspective_grid(
    image: np.ndarray,
) -> tuple[list[int], list[int], np.ndarray] | None:
    """Return a dense physical grid only when it reaches all four warp edges."""
    candidate = extract_ruled_grid(image, prefer_adaptive=True)
    if candidate is None:
        return None
    columns, rows, _ = candidate
    row_count = len(rows) - 1
    column_count = len(columns) - 1
    if not (
        8 <= row_count <= 120
        and 6 <= column_count <= 26
        and row_count * column_count <= 1280
    ):
        return None

    height, width = image.shape[:2]
    horizontal_tolerance = max(4, int(round(width * 0.012)))
    vertical_tolerance = max(4, int(round(height * 0.012)))
    if (
        columns[0] > horizontal_tolerance
        or columns[-1] < width - 1 - horizontal_tolerance
        or rows[0] > vertical_tolerance
        or rows[-1] < height - 1 - vertical_tolerance
    ):
        return None

    column_gaps = np.diff(columns).astype(float)
    row_gaps = np.diff(rows).astype(float)
    if not column_gaps.size or not row_gaps.size:
        return None
    typical_column_gap = float(np.median(column_gaps))
    typical_row_gap = float(np.median(row_gaps))
    if (
        typical_column_gap < max(8.0, width * 0.012)
        or typical_row_gap < max(6.0, height * 0.006)
        or float(column_gaps.min()) < typical_column_gap * 0.35
        or float(row_gaps.min()) < typical_row_gap * 0.35
        or float(column_gaps.max()) > typical_column_gap * 2.75
        or float(row_gaps.max()) > typical_row_gap * 2.75
    ):
        return None
    return list(columns), list(rows), image


def rectify_table_image(
    image: np.ndarray,
    *,
    expand_to_document: bool = True,
) -> tuple[np.ndarray, dict[str, Any]]:
    if image is None or image.size == 0:
        raise ValueError("image is empty")

    height, width = image.shape[:2]
    perspective_image, perspective_transform, perspective_corners, used_document_corners = _warp_perspective_table(
        image,
        expand_to_document=expand_to_document,
    )
    rotated, transform, angle = _deskew(perspective_image)
    # A detected paper contour is stronger completeness evidence than an
    # internal grid weakened by shadow or glare, so keep the complete sheet.
    if used_document_corners:
        return _enhance_for_ocr(rotated), {
            "detected": True,
            "corners": [
                [round(float(px), 2), round(float(py), 2)] for px, py in perspective_corners
            ],
            "deskew_angle": round(angle, 3),
            "paper_expanded": True,
        }
    # A perspective warp can already be the complete spreadsheet.  Re-running
    # contour ranking on it may select a strongly bordered title band and throw
    # away dozens of faint but physically continuous rows.  Keep the full warp
    # only when an independently extracted bounded grid reaches all four edges.
    full_frame_grid = (
        _full_frame_perspective_grid(rotated)
        if perspective_corners is not None
        else None
    )
    if full_frame_grid is not None:
        frame_aspect = max(rotated.shape[:2]) / float(max(1, min(rotated.shape[:2])))
        return _enhance_for_ocr(rotated), {
            "detected": True,
            "corners": [
                [round(float(px), 2), round(float(py), 2)]
                for px, py in perspective_corners
            ],
            "deskew_angle": round(angle, 3),
            "full_frame_perspective_grid": True,
            "dense_near_square": bool(frame_aspect < 1.15),
        }
    horizontal, vertical, grid = _grid_maps(rotated)
    rotated_height, rotated_width = rotated.shape[:2]
    contours, _ = cv2.findContours(grid, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    image_area = float(rotated_width * rotated_height)
    best_bounds = None
    best_geometry = None
    best_score = 0.0
    for contour in contours:
        x, y, candidate_width, candidate_height = cv2.boundingRect(contour)
        area = float(candidate_width * candidate_height)
        aspect = max(candidate_width, candidate_height) / max(1.0, min(candidate_width, candidate_height))
        if area < image_area * 0.025 or aspect < 1.0:
            continue
        horizontal_density = cv2.countNonZero(
            horizontal[y : y + candidate_height, x : x + candidate_width]
        ) / area
        vertical_density = cv2.countNonZero(
            vertical[y : y + candidate_height, x : x + candidate_width]
        ) / area
        score = _table_candidate_score(area, aspect, horizontal_density, vertical_density)
        if score > best_score:
            best_score = score
            best_bounds = (x, y, candidate_width, candidate_height)
            best_geometry = (aspect, horizontal_density, vertical_density)

    if best_bounds is None:
        if perspective_corners is not None:
            return _enhance_for_ocr(perspective_image), {
                "detected": True,
                "corners": [
                    [round(float(px), 2), round(float(py), 2)] for px, py in perspective_corners
                ],
                "deskew_angle": angle,
            }
        return _enhance_for_ocr(image), {"detected": False, "corners": [], "deskew_angle": angle}

    best_bounds = _extend_grid_bounds_along_vertical_rules(vertical, best_bounds)
    x, y, candidate_width, candidate_height = best_bounds
    aspect, horizontal_density, vertical_density = best_geometry or (0.0, 0.0, 0.0)
    dense_near_square = bool(
        aspect < 1.15
        and horizontal_density >= 0.055
        and vertical_density >= 0.025
    )
    margin = max(3, min(48, min(candidate_width, candidate_height) // 35))
    left = max(0, x - margin)
    top = max(0, y - margin)
    right = min(rotated_width, x + candidate_width + margin)
    bottom = min(rotated_height, y + candidate_height + margin)
    rectified = rotated[top:bottom, left:right]

    rotated_corners = np.array(
        [[[left, top], [right - 1, top], [right - 1, bottom - 1], [left, bottom - 1]]],
        dtype=np.float32,
    )
    original_corners = cv2.transform(rotated_corners, cv2.invertAffineTransform(transform))
    if perspective_transform is not None:
        original_corners = cv2.perspectiveTransform(
            original_corners,
            np.linalg.inv(perspective_transform),
        )
    original_corners = original_corners[0]
    original_corners[:, 0] = np.clip(original_corners[:, 0], 0, width - 1)
    original_corners[:, 1] = np.clip(original_corners[:, 1], 0, height - 1)
    return _enhance_for_ocr(rectified), {
        "detected": True,
        "corners": [[round(float(px), 2), round(float(py), 2)] for px, py in original_corners],
        "deskew_angle": round(angle, 3),
        "dense_near_square": dense_near_square,
    }


def prepare_image(image: np.ndarray, crop_mode: str = "auto") -> tuple[np.ndarray, dict[str, Any]]:
    if crop_mode == "full":
        return _enhance_for_ocr(image), {"detected": False, "corners": [], "mode": "full"}
    prepared, metadata = rectify_table_image(
        image,
        expand_to_document=crop_mode != "grid",
    )
    metadata["mode"] = crop_mode if crop_mode == "grid" else "auto"
    return prepared, metadata


class _TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[list[tuple[str, int, int]]] = []
        self._row: list[tuple[str, int, int]] | None = None
        self._cell_text: list[str] | None = None
        self._row_span = 1
        self._column_span = 1

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag == "tr":
            self._row = []
        elif tag in {"td", "th"} and self._row is not None:
            values = dict(attrs)
            self._row_span = max(1, int(values.get("rowspan") or 1))
            self._column_span = max(1, int(values.get("colspan") or 1))
            self._cell_text = []
        elif tag == "br" and self._cell_text is not None:
            self._cell_text.append("\n")

    def handle_data(self, data: str) -> None:
        if self._cell_text is not None:
            self._cell_text.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"td", "th"} and self._row is not None and self._cell_text is not None:
            text = " ".join("".join(self._cell_text).split())
            self._row.append((text, self._row_span, self._column_span))
            self._cell_text = None
        elif tag == "tr" and self._row is not None:
            self.rows.append(self._row)
            self._row = None


def _rectangular(rows: Iterable[Iterable[str]]) -> list[list[str]]:
    materialized = [list(row) for row in rows]
    columns = max((len(row) for row in materialized), default=0)
    return [row + [""] * (columns - len(row)) for row in materialized]


def parse_html_table(html: str) -> tuple[list[list[str]], list[dict[str, int]]]:
    parser = _TableParser()
    parser.feed(html)

    grid: list[list[str]] = []
    occupied: set[tuple[int, int]] = set()
    spans: list[dict[str, int]] = []

    for row_index, source_row in enumerate(parser.rows):
        while len(grid) <= row_index:
            grid.append([])
        column = 0
        for text, row_span, column_span in source_row:
            while (row_index, column) in occupied:
                column += 1
            required_rows = row_index + row_span
            required_columns = column + column_span
            while len(grid) < required_rows:
                grid.append([])
            for target_row in range(required_rows):
                if len(grid[target_row]) < required_columns:
                    grid[target_row].extend([""] * (required_columns - len(grid[target_row])))
            grid[row_index][column] = text
            for target_row in range(row_index, required_rows):
                for target_column in range(column, required_columns):
                    if target_row != row_index or target_column != column:
                        occupied.add((target_row, target_column))
            if row_span > 1 or column_span > 1:
                spans.append(
                    {
                        "row": row_index,
                        "column": column,
                        "row_span": row_span,
                        "column_span": column_span,
                    }
                )
            column += column_span

    return _rectangular(grid), spans


def validate_request(request: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(request, dict):
        raise ValueError("request must be a JSON object")
    if request.get("protocol") != PROTOCOL_VERSION:
        raise ValueError(f"unsupported protocol: {request.get('protocol')!r}")
    action = request.get("action")
    if action not in {"recognize", "export_xlsx", "health", "warmup"}:
        raise ValueError(f"unsupported action: {action!r}")
    if "request_id" in request:
        request_id = request["request_id"]
        if (
            isinstance(request_id, bool)
            or not isinstance(request_id, int)
            or not 1 <= request_id <= 2_147_483_647
        ):
            raise ValueError("request_id must be an integer from 1 to 2147483647")
    if action == "recognize":
        image_path = request.get("image_path")
        if not isinstance(image_path, str) or not image_path.strip():
            raise ValueError("image_path is required")
        output_directory = request.get("output_directory")
        if output_directory is not None and (
            not isinstance(output_directory, str) or not output_directory.strip()
        ):
            raise ValueError("output_directory must be a non-empty path")
        options = request.get("options", {})
        if not isinstance(options, dict):
            raise ValueError("options must be an object")
        crop_mode = options.get("crop_mode", "auto")
        if crop_mode not in {"auto", "grid", "full"}:
            raise ValueError(f"unsupported crop_mode: {crop_mode!r}")
        accuracy_mode = options.get("accuracy_mode", "maximum")
        if accuracy_mode not in {"adaptive", "adaptive_high", "maximum"}:
            raise ValueError(f"unsupported accuracy_mode: {accuracy_mode!r}")
        deadline_seconds = options.get("deadline_seconds", 0)
        if isinstance(deadline_seconds, bool) or not isinstance(
            deadline_seconds, (int, float)
        ):
            raise ValueError("deadline_seconds must be a non-negative number")
        if deadline_seconds < 0:
            raise ValueError("deadline_seconds must be a non-negative number")
    elif action == "export_xlsx":
        output_path = request.get("output_path")
        if not isinstance(output_path, str) or not output_path.strip():
            raise ValueError("output_path is required")
        cells = request.get("cells")
        if not isinstance(cells, list) or not cells:
            raise ValueError("cells must be a non-empty rectangular array")
        expected_columns = len(cells[0]) if isinstance(cells[0], list) else -1
        if expected_columns <= 0 or any(
            not isinstance(row, list) or len(row) != expected_columns
            for row in cells
        ):
            raise ValueError("cells must be a non-empty rectangular array")
    return request


def build_result(
    rows: Iterable[Iterable[str]],
    *,
    confidence: float,
    confidences: Iterable[Iterable[float]] | None = None,
    engine: str,
    rectified_image: str = "",
    spans: list[dict[str, int]] | None = None,
) -> dict[str, Any]:
    grid = _rectangular(rows)
    score = max(0.0, min(1.0, float(confidence)))
    confidence_grid = [list(row) for row in confidences] if confidences is not None else []
    cells = []
    for row_index, row in enumerate(grid):
        cell_row = []
        for column_index, value in enumerate(row):
            raw_cell_score = score
            if row_index < len(confidence_grid) and column_index < len(confidence_grid[row_index]):
                raw_cell_score = float(confidence_grid[row_index][column_index])
            # A yellow low-confidence value must have the same export and
            # summary meaning as an explicit disagreement; otherwise the UI
            # can visibly warn about a cell while reporting zero pending items.
            needs_review = raw_cell_score < 0.0 or (
                raw_cell_score != 0.0 and raw_cell_score < 0.78
            )
            cell_score = max(0.0, min(1.0, raw_cell_score))
            cell_row.append(
                {
                    "text": value,
                    "confidence": cell_score,
                    "needs_review": needs_review,
                }
            )
        cells.append(cell_row)
    return {
        "protocol": PROTOCOL_VERSION,
        "status": "ok",
        "rows": len(grid),
        "columns": len(grid[0]) if grid else 0,
        "cells": cells,
        "spans": spans or [],
        "engine": engine,
        "rectified_image": rectified_image,
    }


def write_xlsx(
    output_path: str | Path,
    rows: Iterable[Iterable[Any]],
    spans: list[dict[str, Any]] | None = None,
) -> None:
    raw_grid = _rectangular(rows)
    grid = [
        [
            str(value.get("text", "")) if isinstance(value, dict) else str(value)
            for value in row
        ]
        for row in raw_grid
    ]
    review_grid = [
        [
            bool(value.get("needs_review", False))
            or (
                bool(str(value.get("text", "")).strip())
                and float(value.get("confidence", 1.0)) < 0.78
            )
            if isinstance(value, dict)
            else False
            for value in row
        ]
        for row in raw_grid
    ]
    occupied: set[tuple[int, int]] = set()
    row_count = len(grid)
    column_count = len(grid[0]) if grid else 0
    for raw_span in spans or []:
        if not isinstance(raw_span, dict):
            raise ValueError("invalid XLSX span: object required")
        values = [
            raw_span.get("row"),
            raw_span.get("column"),
            raw_span.get("row_span", 1),
            raw_span.get("column_span", 1),
        ]
        if any(isinstance(value, bool) or not isinstance(value, int) for value in values):
            raise ValueError("invalid XLSX span: integer coordinates required")
        row, column, row_span, column_span = values
        if (
            row < 0
            or column < 0
            or row_span < 1
            or column_span < 1
            or row_span > row_count - row
            or column_span > column_count - column
        ):
            raise ValueError("invalid XLSX span: outside table bounds")
        for covered_row in range(row, row + row_span):
            for covered_column in range(column, column + column_span):
                location = (covered_row, covered_column)
                if location in occupied:
                    raise ValueError("invalid XLSX span: overlapping merged cells")
                occupied.add(location)
                if location != (row, column) and (
                    bool(grid[covered_row][covered_column].strip())
                    or review_grid[covered_row][covered_column]
                ):
                    raise ValueError(
                        "invalid XLSX span: merge would hide cell content or review state"
                    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "识别结果"

    thin_gray = Side(style="thin", color="CBD5E1")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_fill = PatternFill("solid", fgColor="E8F0FE")
    review_fill = PatternFill("solid", fgColor="FFF4D6")
    title_rows = {
        int(span.get("row", 0))
        for span in (spans or [])
        if span.get("role") == "title"
    }
    header_row_index = 0
    if title_rows:
        header_row_index = max(title_rows) + 1
        if header_row_index >= len(grid):
            header_row_index = 0

    for row_index, row in enumerate(grid, start=1):
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_index == header_row_index + 1:
                cell.fill = header_fill
                cell.font = Font(bold=True, color="172033")
            if review_grid[row_index - 1][column_index - 1]:
                cell.fill = review_fill
                cell.comment = Comment(
                    "此单元格由识别组件标记为待人工复核，请对照原图确认。",
                    "图片转表格",
                )

    for span in spans or []:
        row = int(span["row"]) + 1
        column = int(span["column"]) + 1
        row_span = max(1, int(span.get("row_span", 1)))
        column_span = max(1, int(span.get("column_span", 1)))
        if row_span > 1 or column_span > 1:
            sheet.merge_cells(
                start_row=row,
                start_column=column,
                end_row=row + row_span - 1,
                end_column=column + column_span - 1,
            )
        if span.get("role") == "title":
            title_cell = sheet.cell(row=row, column=column)
            title_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            title_cell.font = Font(bold=True, color="172033", size=16)
            sheet.row_dimensions[row].height = max(
                sheet.row_dimensions[row].height or 0,
                28,
            )

    column_widths = []
    for column_index in range(1, (len(grid[0]) if grid else 0) + 1):
        values = [str(row[column_index - 1]) for row in grid]
        width = min(48, max(10, max((_display_width(value) for value in values), default=0) + 2))
        column_widths.append(width)
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    for row_index, row in enumerate(grid, start=1):
        line_count = 1
        for column_index, value in enumerate(row):
            usable_width = max(6, column_widths[column_index] - 2)
            wrapped_lines = sum(
                max(1, math.ceil(_display_width(line) / usable_width))
                for line in str(value).splitlines() or [""]
            )
            line_count = max(line_count, wrapped_lines)
        calculated_height = min(396, max(22, line_count * 18))
        sheet.row_dimensions[row_index].height = max(
            sheet.row_dimensions[row_index].height or 0,
            calculated_height,
        )
    if grid:
        header_row = header_row_index + 1
        sheet.freeze_panes = f"A{header_row + 1}"
        header_values = grid[header_row_index] if header_row_index < len(grid) else []
        if sum(bool(str(value).strip()) for value in header_values) >= 2:
            last_column = get_column_letter(len(grid[0]))
            sheet.auto_filter.ref = f"A{header_row}:{last_column}{len(grid)}"

    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    # 先在目标目录生成完整工作簿，再由系统一次性替换，避免异常中断破坏已有文件。
    temporary_handle = tempfile.NamedTemporaryFile(
        mode="wb",
        prefix=f".{destination.name}.",
        suffix=".tmp.xlsx",
        dir=destination.parent,
        delete=False,
    )
    temporary = Path(temporary_handle.name)
    temporary_handle.close()
    try:
        workbook.save(temporary)
        os.replace(temporary, destination)
    finally:
        # 仅清理本次创建的临时文件；替换失败时保留用户原目标文件。
        if temporary.exists():
            temporary.unlink()
