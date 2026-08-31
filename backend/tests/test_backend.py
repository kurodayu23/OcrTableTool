import copy
import json
import inspect
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import Mock, patch

import cv2
import numpy as np
from openpyxl import load_workbook
from rapidocr import EngineType, ModelType, OCRVersion

import ocr_backend
import table_pipeline
from ocr_backend import handle_request
from table_pipeline import (
    PROTOCOL_VERSION,
    _table_candidate_score,
    build_result,
    rectify_table_image,
    parse_html_table,
    prepare_image,
    validate_request,
    write_xlsx,
)

extract_ruled_grid = getattr(table_pipeline, "extract_ruled_grid", lambda _: None)
assign_ocr_to_grid = getattr(table_pipeline, "assign_ocr_to_grid", lambda *args: ([], []))


def _real_fixture_path(relative: str) -> Path | None:
    root = os.environ.get("OCR_TABLE_REAL_FIXTURES", "").strip()
    return Path(root) / relative if root else None


class TablePipelineTests(unittest.TestCase):
    def test_backend_runtime_blocks_socket_connections(self):
        import socket

        connection = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        try:
            with self.assertRaisesRegex(PermissionError, "禁止建立网络连接"):
                connection.connect(("127.0.0.1", 9))
        finally:
            connection.close()

    def test_column_median_and_upper_quartile_match_numpy(self):
        for row_count in (8, 9):
            values = np.arange(row_count * 7, dtype=np.float32).reshape(row_count, 7)
            values = values[::-1].copy()

            median, upper_quartile = (
                table_pipeline._column_median_and_upper_quartile(values)
            )

            np.testing.assert_array_equal(median, np.median(values, axis=0))
            np.testing.assert_array_equal(
                upper_quartile,
                np.percentile(values, 75, axis=0),
            )

    def test_openvino_source_graph_is_released_after_compile(self):
        source_model = object()
        component = SimpleNamespace(
            session=SimpleNamespace(model=source_model, session=object())
        )

        self.assertTrue(ocr_backend._release_openvino_source_model(component))

        self.assertIsNone(component.session.model)
        self.assertIsNotNone(component.session.session)
        self.assertFalse(ocr_backend._release_openvino_source_model(component))

    def test_released_openvino_component_drops_native_request_and_model(self):
        inference_session = SimpleNamespace(session=object(), model=object())
        component = SimpleNamespace(session=inference_session)

        self.assertTrue(ocr_backend._dispose_openvino_component(component))
        self.assertIsNone(inference_session.session)
        self.assertIsNone(inference_session.model)
        self.assertFalse(ocr_backend._dispose_openvino_component(component))

    def test_hybrid_engine_releases_only_page_layout_models(self):
        page_engine = SimpleNamespace(
            text_det=object(),
            text_cls=object(),
            text_rec=object(),
        )
        verification_recognizer = object()
        hybrid = ocr_backend._HybridOcrEngine(
            page_engine,
            verification_recognizer,
            page_layout_factory=Mock(),
        )
        page_recognizer = page_engine.text_rec

        self.assertTrue(hybrid.release_page_layout_models())

        self.assertIsNone(page_engine.text_det)
        self.assertIsNone(page_engine.text_cls)
        self.assertIs(page_engine.text_rec, page_recognizer)
        self.assertIs(hybrid._verification_recognizer, verification_recognizer)
        self.assertFalse(hybrid.release_page_layout_models())

    def test_hybrid_engine_restores_page_layout_models_lazily(self):
        restored_detector = object()
        restored_classifier = object()
        factory = Mock(return_value=(restored_detector, restored_classifier))
        page_engine = SimpleNamespace(
            text_det=None,
            text_cls=None,
            text_rec=object(),
        )
        hybrid = ocr_backend._HybridOcrEngine(
            page_engine,
            object(),
            page_layout_factory=factory,
        )

        hybrid._ensure_page_layout_models()
        hybrid._ensure_page_layout_models()

        factory.assert_called_once_with()
        self.assertIs(page_engine.text_det, restored_detector)
        self.assertIs(page_engine.text_cls, restored_classifier)

    def test_hybrid_engine_loads_medium_recognizer_only_on_first_cell_review(self):
        recognizer = Mock(return_value=SimpleNamespace())
        factory = Mock(return_value=recognizer)
        page_engine = SimpleNamespace(
            text_det=object(),
            text_cls=object(),
            text_rec=object(),
        )
        hybrid = ocr_backend._HybridOcrEngine(
            page_engine,
            page_layout_factory=Mock(),
            verification_factory=factory,
            page_recognition_factory=Mock(),
        )

        self.assertIsNone(hybrid._verification_recognizer)
        hybrid.text_rec("first")
        hybrid.text_rec("second")

        factory.assert_called_once_with()
        self.assertIsNone(page_engine.text_det)
        self.assertIsNone(page_engine.text_cls)
        self.assertIsNone(page_engine.text_rec)
        self.assertIs(hybrid._verification_recognizer, recognizer)
        self.assertEqual(recognizer.call_count, 2)

    def test_hybrid_engine_keeps_small_and_medium_recognizers_mutually_exclusive(self):
        first_small = Mock(return_value=SimpleNamespace())
        restored_small = Mock(return_value=SimpleNamespace())
        first_medium = Mock(return_value=SimpleNamespace())
        restored_medium = Mock(return_value=SimpleNamespace())
        small_factory = Mock(return_value=restored_small)
        medium_factory = Mock(side_effect=[first_medium, restored_medium])
        page_engine = SimpleNamespace(
            text_det=object(),
            text_cls=object(),
            text_rec=first_small,
        )
        hybrid = ocr_backend._HybridOcrEngine(
            page_engine,
            page_layout_factory=Mock(),
            verification_factory=medium_factory,
            page_recognition_factory=small_factory,
        )

        hybrid.text_rec("medium-first")
        self.assertIsNone(page_engine.text_det)
        self.assertIsNone(page_engine.text_cls)
        self.assertIsNone(page_engine.text_rec)
        self.assertIs(hybrid._verification_recognizer, first_medium)

        hybrid.fast_text_rec("small-second")
        self.assertIsNone(hybrid._verification_recognizer)
        self.assertIs(page_engine.text_rec, restored_small)

        hybrid.server_text_rec("medium-third")
        self.assertIsNone(page_engine.text_rec)
        self.assertIs(hybrid._verification_recognizer, restored_medium)
        self.assertEqual(medium_factory.call_count, 2)
        small_factory.assert_called_once_with()
        first_medium.assert_called_once_with("medium-first")
        restored_small.assert_called_once_with("small-second")
        restored_medium.assert_called_once_with("medium-third")

    def test_hybrid_engine_can_load_medium_before_enhanced_crops_exist(self):
        page_recognizer = object()
        medium_recognizer = object()
        medium_factory = Mock(return_value=medium_recognizer)
        page_engine = SimpleNamespace(
            text_det=object(),
            text_cls=object(),
            text_rec=page_recognizer,
        )
        hybrid = ocr_backend._HybridOcrEngine(
            page_engine,
            page_layout_factory=Mock(),
            verification_factory=medium_factory,
            page_recognition_factory=Mock(),
        )

        with patch.object(ocr_backend.gc, "collect") as collect:
            hybrid.prepare_verification_recognizer()

        self.assertIsNone(page_engine.text_rec)
        self.assertIsNone(page_engine.text_det)
        self.assertIsNone(page_engine.text_cls)
        self.assertIs(hybrid._verification_recognizer, medium_recognizer)
        medium_factory.assert_called_once_with()
        self.assertEqual(collect.call_count, 3)

    def test_hybrid_engine_releases_all_ocr_models_before_table_structure(self):
        medium = object()
        page_engine = SimpleNamespace(
            text_det=object(),
            text_cls=object(),
            text_rec=object(),
        )
        hybrid = ocr_backend._HybridOcrEngine(
            page_engine,
            medium,
            page_layout_factory=Mock(),
            verification_factory=Mock(),
            page_recognition_factory=Mock(),
        )

        self.assertTrue(hybrid.release_for_table_structure())

        self.assertIsNone(page_engine.text_det)
        self.assertIsNone(page_engine.text_cls)
        self.assertIsNone(page_engine.text_rec)
        self.assertIsNone(hybrid._verification_recognizer)

    def test_hybrid_engine_restores_small_recognizer_for_next_persistent_request(self):
        restored_detector = object()
        restored_classifier = object()
        restored_small = object()
        page_factory = Mock(return_value=(restored_detector, restored_classifier))
        small_factory = Mock(return_value=restored_small)
        page_engine = SimpleNamespace(text_det=None, text_cls=None, text_rec=None)
        hybrid = ocr_backend._HybridOcrEngine(
            page_engine,
            Mock(),
            page_layout_factory=page_factory,
            verification_factory=Mock(),
            page_recognition_factory=small_factory,
        )

        hybrid.prepare_for_request()

        self.assertIsNone(hybrid._verification_recognizer)
        self.assertIs(page_engine.text_det, restored_detector)
        self.assertIs(page_engine.text_cls, restored_classifier)
        self.assertIs(page_engine.text_rec, restored_small)
        page_factory.assert_called_once_with()
        small_factory.assert_called_once_with()

    def test_hybrid_engine_persistent_request_releases_medium_before_page_models(self):
        restored_detector = object()
        restored_classifier = object()
        page_factory = Mock(return_value=(restored_detector, restored_classifier))
        medium = object()
        page_engine = SimpleNamespace(
            text_det=None,
            text_cls=None,
            text_rec=object(),
        )
        hybrid = ocr_backend._HybridOcrEngine(
            page_engine,
            medium,
            page_layout_factory=page_factory,
            verification_factory=Mock(),
        )

        hybrid.prepare_for_request()

        self.assertIsNone(hybrid._verification_recognizer)
        self.assertIs(page_engine.text_det, restored_detector)
        self.assertIs(page_engine.text_cls, restored_classifier)
        page_factory.assert_called_once_with()

    def test_recognition_budget_stops_optional_work_before_gui_deadline(self):
        now = [100.0]
        budget = ocr_backend._RecognitionBudget(28.5, lambda: now[0])

        now[0] = 126.5
        self.assertTrue(budget.allow(2.0))
        now[0] = 127.0
        self.assertFalse(budget.allow(2.0))
        self.assertTrue(budget.limited)
        self.assertAlmostEqual(budget.remaining(), 1.5)

    def test_zero_recognition_budget_is_unlimited(self):
        now = [100.0]
        budget = ocr_backend._RecognitionBudget(0, lambda: now[0])

        now[0] = 10000.0
        self.assertTrue(budget.allow(600.0))
        self.assertFalse(budget.reached())
        self.assertFalse(budget.limited)
        self.assertEqual(budget.seconds, 0.0)

    def test_dense_dash_only_leading_border_row_is_trimmed_before_export(self):
        grid = [
            ["-", "--", "-", "—", "-", "—", "-"],
            ["编号", "名称", "型号", "数值", "单位", "负责人", "状态"],
            ["1", "设备", "A1", "10", "V", "张三", "完成"],
        ]
        confidence = [[0.8] * 7 for _ in grid]

        trimmed, trimmed_confidence = ocr_backend._trim_empty_outer_grid(grid, confidence)

        self.assertEqual(len(trimmed), 2)
        self.assertEqual(trimmed[0][0], "编号")
        self.assertEqual(len(trimmed_confidence), 2)

    def test_sparse_leading_title_is_not_trimmed_as_a_blank_photo_margin(self):
        ocr_backend._load_runtime()
        rows = [0, 60, 105, 150, 195, 240, 285, 330, 375]
        columns = [0, 100, 200, 300, 399]

        def make_image(with_title: bool) -> np.ndarray:
            image = np.full((376, 400, 3), 245, dtype=np.uint8)
            for row in rows:
                cv2.line(image, (0, row), (399, row), (25, 25, 25), 2)
            for column in (0, 399):
                cv2.line(image, (column, 0), (column, 375), (25, 25, 25), 2)
            for column in columns[1:-1]:
                cv2.line(image, (column, 60), (column, 375), (25, 25, 25), 2)
            if with_title:
                cv2.putText(
                    image,
                    "MONTHLY REPORT",
                    (75, 42),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.75,
                    (20, 20, 20),
                    2,
                    cv2.LINE_AA,
                )
            for row in range(1, len(rows) - 1):
                cv2.putText(
                    image,
                    f"R{row}",
                    (15, rows[row] + 30),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.55,
                    (30, 30, 30),
                    1,
                    cv2.LINE_AA,
                )
            return image

        titled = make_image(True)
        titled_horizontal, titled_vertical, _ = table_pipeline._grid_maps(titled)
        self.assertTrue(
            table_pipeline._leading_interval_has_text_components(
                titled,
                rows[0],
                rows[1],
                titled_horizontal,
                titled_vertical,
            )
        )
        self.assertEqual(
            table_pipeline._trim_sparse_leading_page_row(
                titled,
                rows,
                titled_horizontal,
                titled_vertical,
            ),
            rows,
        )

        blank = make_image(False)
        blank_horizontal, blank_vertical, _ = table_pipeline._grid_maps(blank)
        self.assertFalse(
            table_pipeline._leading_interval_has_text_components(
                blank,
                rows[0],
                rows[1],
                blank_horizontal,
                blank_vertical,
            )
        )
        self.assertEqual(
            table_pipeline._trim_sparse_leading_page_row(
                blank,
                rows,
                blank_horizontal,
                blank_vertical,
            ),
            rows[1:],
        )

    def test_warmup_request_loads_models_before_first_import(self):
        with (
            patch.object(ocr_backend, "_require_recognition_memory_headroom") as preflight,
            patch.object(ocr_backend, "_engines", return_value=(object(), None)) as engines,
        ):
            response = handle_request({"protocol": PROTOCOL_VERSION, "action": "warmup"})

        preflight.assert_called_once_with()
        engines.assert_called_once_with()
        self.assertEqual(response["status"], "ok")
        self.assertTrue(response["warmed_up"])

    def test_low_memory_preflight_stops_before_model_loading(self):
        gibibyte = 1024**3
        with (
            patch.object(
                ocr_backend,
                "_available_memory_bytes",
                return_value=(3 * gibibyte, 8 * gibibyte),
            ),
            patch.object(ocr_backend, "_engines") as engines,
        ):
            with self.assertRaisesRegex(MemoryError, "可用内存不足"):
                handle_request({"protocol": PROTOCOL_VERSION, "action": "warmup"})

        engines.assert_not_called()
        self.assertEqual(
            ocr_backend._classify_error(MemoryError("可用内存不足")),
            ("INSUFFICIENT_MEMORY", None, True),
        )

    def test_expired_numeric_verification_withholds_unchecked_values(self):
        now = [0.0]
        budget = ocr_backend._RecognitionBudget(28.5, lambda: now[0])
        now[0] = 27.0
        output = SimpleNamespace(
            boxes=np.asarray([[[1, 1], [20, 1], [20, 10], [1, 10]]], dtype=float),
            txts=["10001"],
            scores=[0.99],
        )
        engine = SimpleNamespace(text_rec=lambda _: self.fail("deadline must prevent OCR retry"))

        _, texts, scores, rejected = ocr_backend._verified_numeric_components(
            np.zeros((20, 30, 3), dtype=np.uint8),
            output,
            engine,
            budget,
        )

        self.assertEqual(texts, [""])
        self.assertEqual(scores, [-1.0])
        self.assertEqual(rejected, 1)

    def test_high_accuracy_grid_requires_medium_model_consensus(self):
        ocr_backend._load_runtime()
        image = np.full((50, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "123", (12, 34), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 2)

        def recognize_with(text):
            def recognize(request):
                count = len(request.img)
                return SimpleNamespace(txts=[text] * count, scores=[0.99] * count)
            return recognize

        engine = SimpleNamespace(
            fast_text_rec=recognize_with("123"),
            text_rec=recognize_with("128"),
        )

        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [0, 120],
            [0, 50],
            engine,
            verify=True,
            quality_image=image,
            require_medium_consensus=True,
        )

        self.assertEqual(grid, [[""]])
        self.assertEqual(confidence, [[-1.0]])

    def test_large_grid_bounds_fast_and_medium_review_and_marks_deferred_cells(self):
        ocr_backend._load_runtime()
        row_count = 15
        column_count = 15
        cell_size = 20
        image = np.zeros(
            (row_count * cell_size, column_count * cell_size, 3),
            dtype=np.uint8,
        )
        fast_call_sizes = []
        medium_call_sizes = []

        def fast_recognize(request):
            fast_call_sizes.append(len(request.img))
            text = ("A", "B", "C")[min(len(fast_call_sizes) - 1, 2)]
            return SimpleNamespace(
                txts=[text] * len(request.img),
                scores=[0.99] * len(request.img),
            )

        def medium_recognize(request):
            medium_call_sizes.append(len(request.img))
            return SimpleNamespace(
                txts=["D"] * len(request.img),
                scores=[0.99] * len(request.img),
            )

        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [index * cell_size for index in range(column_count + 1)],
            [index * cell_size for index in range(row_count + 1)],
            SimpleNamespace(
                fast_text_rec=fast_recognize,
                text_rec=medium_recognize,
            ),
            verify=True,
            quality_image=image,
        )

        self.assertEqual(fast_call_sizes[0], row_count * column_count)
        self.assertLessEqual(fast_call_sizes[1], 96)
        self.assertTrue(medium_call_sizes)
        self.assertLessEqual(max(medium_call_sizes), 24)
        self.assertEqual(grid[-1][-1], "A")
        self.assertEqual(confidence[-1][-1], 0.77)

    def test_certified_grid_rechecks_late_spreadsheet_ruler_evidence(self):
        source = inspect.getsource(ocr_backend._recognize)

        self.assertIn("retrimmed_values", source)
        self.assertIn("_align_structure_certificate_after_ui_strip", source)

    def test_structure_review_policy_is_applied_before_publication(self):
        source = inspect.getsource(ocr_backend._recognize)

        self.assertIn("_apply_structure_review_policy(", source)

    def test_zero_cell_geometry_cannot_enable_unbounded_page_verification(self):
        source = inspect.getsource(ocr_backend._recognize)

        self.assertIn("0 < ruled_cell_count <= 320", source)
        self.assertIn("len(output_texts) <= 200", source)

    def test_large_photo_review_cost_uses_visible_safe_rejection(self):
        source = inspect.getsource(ocr_backend._recognize)

        self.assertIn('dense_grid_safe_rejection_reason = "large_review_cost_exceeds_bound"', source)
        self.assertIn("大型照片表格的逐格复核成本超过安全上限", source)

    def test_high_accuracy_grid_does_not_bypass_clear_chinese_text(self):
        ocr_backend._load_runtime()
        image = np.full((50, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "ABC", (12, 34), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 2)

        def recognize_with(text):
            def recognize(request):
                count = len(request.img)
                return SimpleNamespace(txts=[text] * count, scores=[0.99] * count)
            return recognize

        engine = SimpleNamespace(
            fast_text_rec=recognize_with("合各"),
            text_rec=recognize_with("合格"),
        )

        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [0, 120],
            [0, 50],
            engine,
            verify=True,
            quality_image=image,
            require_medium_consensus=True,
        )

        self.assertEqual(grid, [[""]])
        self.assertEqual(confidence, [[-1.0]])

    def test_maximum_accuracy_recovers_from_medium_source_view_majority(self):
        ocr_backend._load_runtime()
        image = np.full((50, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "ABC", (12, 34), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 2)
        medium_calls = [0]

        def fast_recognize(request):
            return SimpleNamespace(txts=["合各"] * len(request.img), scores=[0.99] * len(request.img))

        def medium_recognize(request):
            medium_calls[0] += 1
            if medium_calls[0] == 1:
                return SimpleNamespace(txts=["台格"] * len(request.img), scores=[0.99] * len(request.img))
            return SimpleNamespace(
                txts=["合格", "合格", "合各"],
                scores=[0.99, 0.98, 0.97],
            )

        engine = SimpleNamespace(
            fast_text_rec=fast_recognize,
            text_rec=medium_recognize,
            server_text_rec=lambda request: SimpleNamespace(
                txts=["合格"] * len(request.img),
                scores=[0.99] * len(request.img),
            ),
        )

        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [0, 120],
            [0, 50],
            engine,
            verify=True,
            quality_image=image,
            require_medium_consensus=True,
            unbounded_consensus=True,
        )

        self.assertEqual(grid, [["合格"]])
        self.assertGreater(confidence[0][0], 0.0)

    def test_maximum_accuracy_accepts_fast_and_medium_agreement(self):
        ocr_backend._load_runtime()
        image = np.full((50, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "ABC", (12, 34), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 2)

        medium_calls = [0]

        def fast_recognize(request):
            return SimpleNamespace(txts=["合格"] * len(request.img), scores=[0.99] * len(request.img))

        def medium_recognize(request):
            medium_calls[0] += 1
            return SimpleNamespace(txts=["合格"] * len(request.img), scores=[0.99] * len(request.img))

        engine = SimpleNamespace(fast_text_rec=fast_recognize, text_rec=medium_recognize)
        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [0, 120],
            [0, 50],
            engine,
            verify=True,
            quality_image=image,
            require_medium_consensus=True,
            unbounded_consensus=True,
        )

        self.assertEqual(grid, [["合格"]])
        self.assertGreater(confidence[0][0], 0.0)
        self.assertEqual(medium_calls[0], 1)

    def test_maximum_accuracy_skips_server_when_mobile_and_medium_families_agree(self):
        ocr_backend._load_runtime()
        image = np.full((50, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "ABC", (12, 34), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 2)
        server_calls = [0]
        v4_calls = [0]

        def recognize(request):
            return SimpleNamespace(txts=["合格"] * len(request.img), scores=[0.99] * len(request.img))

        def server_recognize(request):
            server_calls[0] += 1
            return recognize(request)

        def v4_recognize(request):
            v4_calls[0] += 1
            return recognize(request)

        engine = SimpleNamespace(
            fast_text_rec=recognize,
            text_rec=recognize,
            server_text_rec=server_recognize,
            v4_server_text_rec=v4_recognize,
        )
        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [0, 120],
            [0, 50],
            engine,
            verify=True,
            quality_image=image,
            require_medium_consensus=True,
            unbounded_consensus=True,
        )

        self.assertEqual(grid, [["合格"]])
        self.assertGreater(confidence[0][0], 0.0)
        self.assertEqual(server_calls[0], 0)
        self.assertEqual(v4_calls[0], 0)

    def test_maximum_accuracy_keeps_server_for_unit_consensus(self):
        ocr_backend._load_runtime()
        image = np.full((50, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "MHz", (12, 34), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 2)
        server_calls = [0]

        def recognize(request):
            return SimpleNamespace(txts=["MHz"] * len(request.img), scores=[0.99] * len(request.img))

        def server_recognize(request):
            server_calls[0] += 1
            return recognize(request)

        engine = SimpleNamespace(
            fast_text_rec=recognize,
            text_rec=recognize,
            server_text_rec=server_recognize,
        )

        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [0, 120],
            [0, 50],
            engine,
            verify=True,
            quality_image=image,
            require_medium_consensus=True,
            unbounded_consensus=True,
        )

        self.assertEqual(grid, [["MHz"]])
        self.assertGreater(confidence[0][0], 0.0)
        self.assertGreater(server_calls[0], 0)

    def test_any_negative_model_conflict_is_a_publication_blocker(self):
        locations = ocr_backend._unresolved_model_conflict_locations(
            [[0.99, -1.0], [0.0, -2.0]]
        )

        self.assertEqual(locations, [(0, 1), (1, 1)])
        source = inspect.getsource(ocr_backend._recognize)
        self.assertIn("if unresolved_nonblank_conflicts", source)
        self.assertIn("禁止直接导出", source)

    def test_any_pending_review_is_a_publication_blocker(self):
        source = inspect.getsource(ocr_backend._recognize)
        self.assertIn("if pending_review_locations", source)
        self.assertIn("存在尚未人工确认的黄色单元格", source)

    def test_maximum_accuracy_reads_faint_cell_from_untouched_source(self):
        ocr_backend._load_runtime()
        cleaned = np.full((50, 120, 3), 255, dtype=np.uint8)
        source = cleaned.copy()
        cv2.putText(source, "ABC", (12, 34), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (80, 80, 80), 2)

        def recognize(request):
            return SimpleNamespace(txts=["ABC"] * len(request.img), scores=[0.99] * len(request.img))

        engine = SimpleNamespace(
            fast_text_rec=recognize,
            text_rec=recognize,
            server_text_rec=recognize,
        )
        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            cleaned,
            [0, 120],
            [0, 50],
            engine,
            verify=True,
            quality_image=source,
            require_medium_consensus=True,
            unbounded_consensus=True,
        )

        self.assertEqual(grid, [["ABC"]])
        self.assertGreater(confidence[0][0], 0.0)

    def test_dense_grid_recovers_dash_erased_by_line_removal_from_source(self):
        ocr_backend._load_runtime()
        cleaned = np.full((50, 120, 3), 255, dtype=np.uint8)
        source = cleaned.copy()
        cv2.line(source, (55, 25), (65, 25), (20, 20, 20), 2)
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(txts=[], scores=[]),
        )

        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            cleaned,
            [0, 120],
            [0, 50],
            engine,
            quality_image=source,
        )

        self.assertEqual(grid, [["-"]])
        self.assertEqual(confidence, [[0.86]])

    def test_maximum_accuracy_withholds_opposing_model_consensuses(self):
        ocr_backend._load_runtime()
        image = np.full((50, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "72.2", (8, 34), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 2)

        def wrong_recognize(request):
            return SimpleNamespace(txts=["722"] * len(request.img), scores=[0.99] * len(request.img))

        def server_recognize(request):
            return SimpleNamespace(txts=["72.2"] * len(request.img), scores=[0.99] * len(request.img))

        engine = SimpleNamespace(
            fast_text_rec=wrong_recognize,
            text_rec=wrong_recognize,
            server_text_rec=server_recognize,
        )
        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [0, 120],
            [0, 50],
            engine,
            verify=True,
            quality_image=image,
            require_medium_consensus=True,
            unbounded_consensus=True,
        )

        self.assertEqual(grid, [[""]])
        self.assertLess(confidence[0][0], 0.0)

    def test_maximum_accuracy_accepts_server_with_independent_model_support(self):
        ocr_backend._load_runtime()
        image = np.full((50, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "72.2", (8, 34), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 2)

        def fast_recognize(request):
            return SimpleNamespace(txts=["722"] * len(request.img), scores=[0.99] * len(request.img))

        def correct_recognize(request):
            return SimpleNamespace(txts=["72.2"] * len(request.img), scores=[0.99] * len(request.img))

        engine = SimpleNamespace(
            fast_text_rec=fast_recognize,
            text_rec=correct_recognize,
            server_text_rec=correct_recognize,
        )
        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [0, 120],
            [0, 50],
            engine,
            verify=True,
            quality_image=image,
            require_medium_consensus=True,
            unbounded_consensus=True,
        )

        self.assertEqual(grid, [["72.2"]])
        self.assertGreater(confidence[0][0], 0.0)

    def test_recognition_surfaces_excessive_cross_model_conflicts_for_review(self):
        source = inspect.getsource(ocr_backend._recognize)

        self.assertIn(
            "blocking_review_cells >= 8 and blocking_review_ratio > 0.25",
            source,
        )
        self.assertIn("识别结果存在较多跨模型冲突", source)
        self.assertIn("已留空或标黄", source)
        self.assertIn("_blocking_unresolved_visible_blank_locations", source)
        self.assertIn("有可见内容但未能安全确认", source)
        self.assertIn("preserve_ruled_grid_for_maximum", source)

    def test_unresolved_multilevel_header_requires_merged_geometry(self):
        grid = [
            ["", "信息", "", "", "", "判定", ""],
            ["", "班次", "生产线", "", "完成数量", "不良数量", ""],
            ["2026-08-02", "夜班", "B线", "280", "277", "3", "98.2%"],
        ]

        self.assertTrue(ocr_backend._has_unresolved_multilevel_header(grid, []))
        self.assertTrue(
            ocr_backend._has_unresolved_multilevel_header(
                grid,
                [{"row": 0, "column": 0, "row_span": 1, "column_span": 3}],
            )
        )

    def test_grid_geometry_rejects_pathological_false_grid(self):
        self.assertTrue(
            ocr_backend._grid_geometry_is_bounded(
                list(range(8)),
                list(range(16)),
            )
        )
        self.assertFalse(
            ocr_backend._grid_geometry_is_bounded(
                list(range(61)),
                list(range(17)),
            )
        )

    def test_long_weak_horizontal_sequence_requires_credible_columns(self):
        image = np.full((600, 800, 3), 245, dtype=np.uint8)
        horizontal = np.zeros((600, 800), dtype=np.uint8)
        columns = [0, 100, 200, 300, 400, 500, 600, 700, 799]
        complete_rows = [20, 80, 140, 200, 260, 320, 380, 440, 500, 560]
        for row in complete_rows[:4]:
            cv2.line(horizontal, (0, row), (159, row), 255, 2)
        for row in complete_rows[4:]:
            cv2.line(horizontal, (0, row), (799, row), 255, 2)
        observed = complete_rows[4:]

        self.assertEqual(
            table_pipeline._recover_long_weak_horizontal_sequence(
                image,
                columns,
                observed,
                horizontal,
            ),
            complete_rows,
        )

        two_header_rows = complete_rows[2:]
        self.assertEqual(
            table_pipeline._recover_long_weak_horizontal_sequence(
                image,
                columns,
                two_header_rows,
                horizontal,
            ),
            complete_rows,
        )

        false_columns = list(range(0, 210, 10))
        self.assertEqual(
            table_pipeline._recover_long_weak_horizontal_sequence(
                image,
                false_columns,
                observed,
                horizontal,
            ),
            observed,
        )

    def test_maximum_screen_mode_can_keep_one_thousand_verified_cells(self):
        columns = list(range(24))
        rows = list(range(45))

        self.assertFalse(ocr_backend._grid_geometry_is_bounded(columns, rows))
        self.assertTrue(
            ocr_backend._grid_geometry_is_bounded(
                columns,
                rows,
                maximum_cells=1280,
            )
        )

    def test_dark_full_frame_grid_can_skip_photo_rectification(self):
        ocr_backend._load_runtime()
        image = np.full((500, 700, 3), 92, dtype=np.uint8)
        columns = [0, 100, 200, 300, 400, 500, 600, 699]
        rows = [0] + [40 * index for index in range(1, 12)] + [499]
        for column in columns:
            cv2.line(image, (column, 0), (column, 499), (230, 230, 230), 2)
        for row in rows:
            cv2.line(image, (0, row), (699, row), (230, 230, 230), 2)
        candidate = (columns, rows, np.zeros(image.shape[:2], dtype=np.uint8))

        self.assertTrue(
            ocr_backend._full_frame_ruled_grid_can_bypass_photo_rectification(
                image,
                candidate,
                maximum_cells=1280,
            )
        )
        inset = ([80] + columns[2:-1] + [620], rows, candidate[2])
        self.assertFalse(
            ocr_backend._full_frame_ruled_grid_can_bypass_photo_rectification(
                image,
                inset,
                maximum_cells=1280,
            )
        )
        irregular_rows = [0, 22, 67, 95, 142, 178, 229, 263, 310, 352, 401, 438, 499]
        irregular = (columns, irregular_rows, candidate[2])
        self.assertFalse(
            ocr_backend._full_frame_ruled_grid_can_bypass_photo_rectification(
                image,
                irregular,
                maximum_cells=1280,
            )
        )
        edge_slivers = (
            [0, 45, 145, 245, 345, 445, 545, 645, 690],
            rows,
            candidate[2],
        )
        self.assertFalse(
            ocr_backend._full_frame_ruled_grid_can_bypass_photo_rectification(
                image,
                edge_slivers,
                maximum_cells=1280,
            )
        )

    def test_bottom_document_rectification_preserves_top_edge(self):
        ocr_backend._load_runtime()
        image = np.full((160, 220, 3), 245, dtype=np.uint8)
        corners = np.asarray(
            [[20, 20], [200, 20], [200, 130], [20, 130]],
            dtype=np.float32,
        )

        result = ocr_backend._bottom_document_rectification(
            image,
            corners,
            0.05,
        )

        self.assertIsNotNone(result)
        assert result is not None
        _, _, expanded = result
        np.testing.assert_allclose(expanded[:2], corners[:2])
        self.assertGreater(float(expanded[2][1]), float(corners[2][1]))
        self.assertGreater(float(expanded[3][1]), float(corners[3][1]))
        self.assertIsNone(
            ocr_backend._bottom_document_rectification(
                image,
                corners,
                0.09,
            )
        )

    def test_trailing_grid_rows_require_text_and_continuing_rules(self):
        ocr_backend._load_runtime()
        image = np.full((260, 420, 3), 245, dtype=np.uint8)
        columns = [10, 110, 210, 310, 410]
        rows = [10, 60, 110, 160, 210, 250]
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (35, 35, 35), 2)
        for row in rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (35, 35, 35), 2)
        cv2.putText(
            image,
            "DATA",
            (125, 242),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.6,
            (20, 20, 20),
            2,
        )
        grid = (columns, rows, np.zeros(image.shape[:2], dtype=np.uint8))

        self.assertTrue(
            ocr_backend._grid_has_text_supported_trailing_rows(
                image,
                grid,
                4,
            )
        )
        blank = image.copy()
        blank[212:248, 12:408] = 245
        self.assertFalse(
            ocr_backend._grid_has_text_supported_trailing_rows(
                blank,
                grid,
                4,
            )
        )

    def test_photographic_rectification_touching_edge_requires_recapture(self):
        clipped = {
            "detected": True,
            "corners": [[310.0, 180.0], [1599.0, 180.0], [1599.0, 1040.0], [300.0, 1040.0]],
        }
        complete = {
            "detected": True,
            "corners": [[220.0, 160.0], [1370.0, 180.0], [1360.0, 1020.0], [230.0, 1010.0]],
        }

        self.assertTrue(ocr_backend._rectification_touches_image_edge(clipped, 1600, 1200))
        self.assertFalse(ocr_backend._rectification_touches_image_edge(complete, 1600, 1200))
        self.assertFalse(
            ocr_backend._rectification_indicates_clipped_document(clipped, 1600, 1200)
        )
        clipped["paper_expanded"] = True
        self.assertTrue(
            ocr_backend._rectification_indicates_clipped_document(clipped, 1600, 1200)
        )

    def test_rectify_accepts_portrait_document_candidate(self):
        image = np.full((1100, 900, 3), 225, dtype=np.uint8)
        left, top, right, bottom = 240, 120, 660, 980
        cv2.rectangle(image, (left, top), (right, bottom), (25, 25, 25), 4)
        for y in range(top + 70, bottom, 70):
            cv2.line(image, (left, y), (right, y), (45, 45, 45), 2)
        for x in range(left + 70, right, 70):
            cv2.line(image, (x, top), (x, bottom), (45, 45, 45), 2)

        rectified, metadata = rectify_table_image(image)

        self.assertTrue(metadata["detected"])
        self.assertGreater(rectified.shape[0], rectified.shape[1])

    def test_rectify_borderless_text_uses_visible_paper_without_grid_candidate(self):
        image = np.full((720, 1200, 3), (74, 92, 112), dtype=np.uint8)
        paper = np.array([[155, 120], [1060, 155], [1015, 600], [120, 565]], dtype=np.int32)
        cv2.fillConvexPoly(image, paper, (244, 246, 248))
        for row in range(6):
            y = 230 + row * 54
            for column, x in enumerate((235, 470, 705, 910)):
                cv2.putText(
                    image,
                    f"R{row}C{column}",
                    (x, y),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.55,
                    (25, 25, 25),
                    1,
                    cv2.LINE_AA,
                )

        empty_map = np.zeros(image.shape[:2], dtype=np.uint8)
        with patch.object(
            table_pipeline,
            "_grid_maps",
            return_value=(empty_map, empty_map, empty_map),
        ):
            rectified, metadata = rectify_table_image(image)

        self.assertTrue(metadata["detected"])
        self.assertTrue(metadata.get("paper_expanded"))
        self.assertGreater(rectified.shape[1], 800)
        self.assertLess(rectified.shape[0], image.shape[0])

    def test_rectify_prefers_small_complete_sheet_inside_large_false_grid_frame(self):
        image = np.full((900, 1200, 3), (92, 82, 72), dtype=np.uint8)
        cv2.rectangle(image, (40, 35), (1160, 865), (238, 238, 238), 22)
        cv2.rectangle(image, (360, 145), (840, 765), (246, 247, 248), -1)
        for y in range(205, 735, 53):
            cv2.line(image, (395, y), (805, y), (45, 45, 45), 2)
        for x in range(395, 806, 82):
            cv2.line(image, (x, 205), (x, 735), (45, 45, 45), 2)

        rectified, metadata = rectify_table_image(image)

        self.assertTrue(metadata.get("paper_expanded"))
        self.assertLess(rectified.shape[1], 650)
        self.assertGreater(rectified.shape[0], 500)

    def test_dense_outer_grid_is_not_replaced_by_small_inner_document(self):
        image = np.full((900, 1200, 3), (72, 78, 84), dtype=np.uint8)
        outer = np.asarray(
            [[70, 65], [1130, 65], [1130, 835], [70, 835]],
            dtype=np.float32,
        )
        inner = np.asarray(
            [[420, 250], [780, 250], [780, 650], [420, 650]],
            dtype=np.float32,
        )
        cv2.fillConvexPoly(image, outer.astype(np.int32), (228, 230, 232))
        for x in np.linspace(70, 1130, 9).astype(int):
            cv2.line(image, (x, 65), (x, 835), (30, 30, 30), 3)
        for y in np.linspace(65, 835, 15).astype(int):
            cv2.line(image, (70, y), (1130, y), (30, 30, 30), 3)

        self.assertTrue(table_pipeline._quad_has_dense_ruled_extent(image, outer))
        with (
            patch.object(
                table_pipeline,
                "_detect_light_document_corners",
                return_value=inner,
            ),
            patch.object(
                table_pipeline,
                "_detect_low_contrast_ruled_quad",
                return_value=outer,
            ),
            patch.object(
                table_pipeline,
                "_refine_vertical_table_sides",
                side_effect=lambda _image, corners: corners,
            ),
            patch.object(
                table_pipeline,
                "_expand_one_proven_trailing_grid_row",
                return_value=None,
            ),
        ):
            _, _, selected, used_document = table_pipeline._warp_perspective_table(
                image
            )

        self.assertFalse(used_document)
        np.testing.assert_allclose(selected, outer)

    def test_perspective_warp_keeps_near_matching_complete_sheet_extent(self):
        image = np.full((600, 800, 3), (55, 65, 75), dtype=np.uint8)
        table = np.asarray(
            [[100, 100], [700, 100], [700, 500], [100, 500]],
            dtype=np.float32,
        )
        document = np.asarray(
            [[80, 100], [700, 100], [700, 500], [80, 500]],
            dtype=np.float32,
        )
        empty = np.zeros(image.shape[:2], dtype=np.uint8)
        grid = empty.copy()
        cv2.rectangle(grid, (100, 100), (700, 500), 255, 2)

        with (
            patch.object(
                table_pipeline,
                "_grid_maps",
                return_value=(empty, empty, grid),
            ),
            patch.object(
                table_pipeline,
                "_detect_light_document_corners",
                return_value=document,
            ),
            patch.object(
                table_pipeline,
                "_detect_low_contrast_ruled_quad",
                return_value=table,
            ),
            patch.object(
                table_pipeline,
                "_refine_vertical_table_sides",
                side_effect=lambda _image, corners: corners,
            ),
        ):
            _, _, selected, used_document = table_pipeline._warp_perspective_table(
                image
            )

        self.assertTrue(used_document)
        np.testing.assert_allclose(selected, document)

    def test_perspective_warp_retries_trailing_row_before_side_refinement(self):
        image = np.full((600, 800, 3), (55, 65, 75), dtype=np.uint8)
        table = np.asarray(
            [[100, 100], [700, 100], [700, 500], [100, 500]],
            dtype=np.float32,
        )
        refined = table.copy()
        refined[1, 0] -= 1.0
        expanded = table.copy()
        expanded[2:, 1] += 20.0
        document = np.asarray(
            [[99, 99], [701, 99], [701, 507], [99, 507]],
            dtype=np.float32,
        )
        empty = np.zeros(image.shape[:2], dtype=np.uint8)

        with (
            patch.object(
                table_pipeline,
                "_grid_maps",
                return_value=(empty, empty, empty),
            ),
            patch.object(
                table_pipeline,
                "_detect_light_document_corners",
                return_value=document,
            ),
            patch.object(
                table_pipeline,
                "_detect_low_contrast_ruled_quad",
                return_value=table,
            ),
            patch.object(
                table_pipeline,
                "_refine_vertical_table_sides",
                return_value=refined,
            ),
            patch.object(
                table_pipeline,
                "_expand_one_proven_trailing_grid_row",
                side_effect=[None, expanded],
            ) as recover,
        ):
            _, _, selected, used_document = table_pipeline._warp_perspective_table(
                image
            )

        self.assertFalse(used_document)
        self.assertEqual(recover.call_count, 2)
        np.testing.assert_allclose(selected, expanded)

    def test_document_confirmed_grid_is_not_replaced_by_smaller_inner_quad(self):
        image = np.full((600, 800, 3), (55, 65, 75), dtype=np.uint8)
        grid_corners = np.asarray(
            [[100, 100], [700, 100], [700, 500], [100, 500]],
            dtype=np.float32,
        )
        document = np.asarray(
            [[102, 101], [699, 101], [699, 499], [102, 499]],
            dtype=np.float32,
        )
        inner = np.asarray(
            [[103, 102], [697, 102], [697, 498], [103, 498]],
            dtype=np.float32,
        )
        empty = np.zeros(image.shape[:2], dtype=np.uint8)
        grid = empty.copy()
        cv2.rectangle(grid, (100, 100), (700, 500), 255, 2)

        with (
            patch.object(
                table_pipeline,
                "_grid_maps",
                return_value=(empty, empty, grid),
            ),
            patch.object(
                table_pipeline,
                "_detect_light_document_corners",
                return_value=document,
            ),
            patch.object(
                table_pipeline,
                "_detect_low_contrast_ruled_quad",
                return_value=inner,
            ),
            patch.object(
                table_pipeline,
                "_table_candidate_score",
                return_value=1.0,
            ),
            patch.object(
                table_pipeline,
                "_refine_vertical_table_sides",
                side_effect=lambda _image, corners: corners,
            ),
            patch.object(
                table_pipeline,
                "_expand_one_proven_trailing_grid_row",
                return_value=None,
            ),
        ):
            _, _, selected, used_document = table_pipeline._warp_perspective_table(
                image
            )

        self.assertFalse(used_document)
        np.testing.assert_allclose(selected, grid_corners, atol=2.0)

    def test_document_confirmation_keeps_meaningfully_smaller_inner_quad(self):
        image = np.full((600, 800, 3), (55, 65, 75), dtype=np.uint8)
        document = np.asarray(
            [[102, 101], [699, 101], [699, 499], [102, 499]],
            dtype=np.float32,
        )
        inner = np.asarray(
            [[112, 108], [688, 108], [688, 492], [112, 492]],
            dtype=np.float32,
        )
        empty = np.zeros(image.shape[:2], dtype=np.uint8)
        grid = empty.copy()
        cv2.rectangle(grid, (100, 100), (700, 500), 255, 2)

        with (
            patch.object(
                table_pipeline,
                "_grid_maps",
                return_value=(empty, empty, grid),
            ),
            patch.object(
                table_pipeline,
                "_detect_light_document_corners",
                return_value=document,
            ),
            patch.object(
                table_pipeline,
                "_detect_low_contrast_ruled_quad",
                return_value=inner,
            ),
            patch.object(
                table_pipeline,
                "_table_candidate_score",
                return_value=1.0,
            ),
            patch.object(
                table_pipeline,
                "_refine_vertical_table_sides",
                side_effect=lambda _image, corners: corners,
            ),
            patch.object(
                table_pipeline,
                "_expand_one_proven_trailing_grid_row",
                return_value=None,
            ),
        ):
            _, _, selected, used_document = table_pipeline._warp_perspective_table(
                image
            )

        self.assertFalse(used_document)
        np.testing.assert_allclose(selected, inner)

    def test_small_complete_paper_crop_is_usable(self):
        ocr_backend._load_runtime()
        source = np.zeros((1200, 1600, 3), dtype=np.uint8)
        rectified = np.zeros((460, 610, 3), dtype=np.uint8)
        complete_paper = {
            "detected": True,
            "paper_expanded": True,
            "corners": [[580, 290], [1030, 290], [1030, 890], [580, 890]],
        }
        ordinary_small_crop = {
            "detected": True,
            "corners": [[580, 290], [1030, 290], [1030, 890], [580, 890]],
        }

        self.assertTrue(
            ocr_backend._rectified_crop_is_usable(source, rectified, complete_paper)
        )
        self.assertFalse(
            ocr_backend._rectified_crop_is_usable(source, rectified, ordinary_small_crop)
        )

        minimally_outside = {
            "detected": True,
            "paper_expanded": True,
            "corners": [[100, 100], [1500, 100], [1500, 1204], [100, 1204]],
        }
        substantially_outside = {
            "detected": True,
            "paper_expanded": True,
            "corners": [[100, 100], [1500, 100], [1500, 1225], [100, 1225]],
        }
        self.assertTrue(
            ocr_backend._rectified_crop_is_usable(
                source,
                rectified,
                minimally_outside,
            )
        )
        self.assertFalse(
            ocr_backend._rectified_crop_is_usable(
                source,
                rectified,
                substantially_outside,
            )
        )

    def test_inner_rectification_cannot_discard_large_complete_sheet_extent(self):
        ocr_backend._load_runtime()
        paper = {
            "paper_expanded": True,
            "corners": [[100, 100], [900, 100], [900, 900], [100, 900]],
        }
        truncated = {
            "corners": [[110, 110], [890, 110], [890, 620], [110, 620]],
        }
        nearly_complete = {
            "corners": [[110, 110], [890, 110], [890, 880], [110, 880]],
        }
        both_bottom_edges_cut = {
            "corners": [[110, 110], [890, 110], [890, 870], [110, 870]],
        }
        both_top_edges_cut = {
            "corners": [[110, 130], [890, 130], [890, 890], [110, 890]],
        }

        self.assertTrue(
            ocr_backend._inner_rectification_discards_complete_sheet_extent(
                paper,
                truncated,
            )
        )
        self.assertFalse(
            ocr_backend._inner_rectification_discards_complete_sheet_extent(
                paper,
                nearly_complete,
            )
        )
        self.assertTrue(
            ocr_backend._inner_rectification_discards_complete_sheet_extent(
                paper,
                both_bottom_edges_cut,
            )
        )
        self.assertTrue(
            ocr_backend._inner_rectification_discards_complete_sheet_extent(
                paper,
                both_top_edges_cut,
            )
        )

    def test_rectified_crop_rejects_collapsed_perspective_side(self):
        source = np.zeros((1594, 2400, 3), dtype=np.uint8)
        rectified = np.zeros((963, 2393, 3), dtype=np.uint8)
        collapsed = {
            "detected": True,
            "corners": [
                [220.01, 1323.0],
                [2399.0, 333.0],
                [2292.0, 1290.0],
                [165.54, 1323.0],
            ],
        }

        self.assertFalse(
            ocr_backend._rectified_crop_is_usable(source, rectified, collapsed)
        )

    def test_light_document_candidate_rejects_region_touching_two_frame_edges(self):
        image = np.full((600, 1000, 3), (55, 65, 75), dtype=np.uint8)
        cv2.rectangle(image, (300, 0), (999, 520), (244, 246, 248), -1)

        self.assertIsNone(table_pipeline._detect_light_document_corners(image))

        image = np.full((600, 1000, 3), (55, 65, 75), dtype=np.uint8)
        cv2.rectangle(image, (180, 80), (999, 520), (244, 246, 248), -1)
        self.assertIsNotNone(table_pipeline._detect_light_document_corners(image))

    def test_light_document_candidate_accepts_complete_portrait_sheet(self):
        image = np.full((900, 1000, 3), (55, 65, 75), dtype=np.uint8)
        cv2.rectangle(image, (320, 80), (720, 820), (244, 246, 248), -1)

        corners = table_pipeline._detect_light_document_corners(image)

        self.assertIsNotNone(corners)

    def test_perspective_crop_expands_one_proven_trailing_grid_row(self):
        image = np.full((600, 800, 3), (45, 55, 65), dtype=np.uint8)
        cv2.rectangle(image, (80, 80), (720, 520), (245, 246, 247), -1)
        for column in range(100, 701, 100):
            cv2.line(image, (column, 100), (column, 500), (80, 80, 80), 2)
        for row in range(100, 501, 20):
            cv2.line(image, (100, row), (700, row), (80, 80, 80), 2)
        table = np.asarray(
            [[100, 100], [700, 100], [700, 480], [100, 480]],
            dtype=np.float32,
        )
        document = np.asarray(
            [[80, 80], [720, 80], [720, 520], [80, 520]],
            dtype=np.float32,
        )

        expanded = table_pipeline._expand_one_proven_trailing_grid_row(
            image,
            table,
            document,
        )

        self.assertIsNotNone(expanded)
        self.assertAlmostEqual(float(expanded[2, 1]), 500.0, delta=4.0)
        self.assertAlmostEqual(float(expanded[3, 1]), 500.0, delta=4.0)

    def test_perspective_crop_does_not_expand_blank_document_margin(self):
        image = np.full((600, 800, 3), (45, 55, 65), dtype=np.uint8)
        cv2.rectangle(image, (80, 80), (720, 520), (245, 246, 247), -1)
        for column in range(100, 701, 100):
            cv2.line(image, (column, 100), (column, 480), (80, 80, 80), 2)
        for row in range(100, 481, 20):
            cv2.line(image, (100, row), (700, row), (80, 80, 80), 2)
        table = np.asarray(
            [[100, 100], [700, 100], [700, 480], [100, 480]],
            dtype=np.float32,
        )
        document = np.asarray(
            [[80, 80], [720, 80], [720, 520], [80, 520]],
            dtype=np.float32,
        )

        self.assertIsNone(
            table_pipeline._expand_one_proven_trailing_grid_row(
                image,
                table,
                document,
            )
        )
        self.assertEqual(
            ocr_backend._select_extended_title_candidate(
                "稀疏长备",
                [("稀疏长备注宽表—2026-01批次", 0.96)],
                [("稀疏长备注宽表—2026-01批次", 0.95)],
            ),
            ("稀疏长备注宽表 — 2026-01 批次", 0.95),
        )

    def test_light_document_candidate_ignores_hollow_decorative_frame(self):
        image = np.full((900, 1200, 3), (100, 90, 80), dtype=np.uint8)
        cv2.rectangle(image, (55, 45), (1145, 855), (248, 248, 248), 24)
        cv2.rectangle(image, (365, 150), (835, 760), (245, 246, 247), -1)
        cv2.rectangle(image, (405, 205), (795, 710), (30, 30, 30), 3)

        corners = table_pipeline._detect_light_document_corners(image)

        self.assertIsNotNone(corners)
        self.assertGreater(float(np.min(corners[:, 0])), 300.0)
        self.assertLess(float(np.max(corners[:, 0])), 900.0)
        self.assertGreater(float(np.min(corners[:, 1])), 100.0)
        self.assertLess(float(np.max(corners[:, 1])), 800.0)

    def test_light_document_candidate_accepts_complete_far_sheet(self):
        image = np.full((1200, 1600, 3), (95, 85, 75), dtype=np.uint8)
        cv2.rectangle(image, (620, 280), (980, 810), (245, 246, 247), -1)

        corners = table_pipeline._detect_light_document_corners(image)

        self.assertIsNotNone(corners)
        self.assertGreater(float(np.min(corners[:, 0])), 580.0)
        self.assertLess(float(np.max(corners[:, 0])), 1020.0)
        self.assertGreater(float(np.ptp(corners[:, 1])), float(np.ptp(corners[:, 0])))

    def test_partial_grid_bounds_extend_when_vertical_rules_continue(self):
        vertical = np.zeros((500, 600), dtype=np.uint8)
        for column in (100, 180, 260, 340, 420, 500):
            cv2.line(vertical, (column, 80), (column, 430), 255, 2)

        extended = table_pipeline._extend_grid_bounds_along_vertical_rules(
            vertical,
            (90, 75, 430, 170),
        )

        self.assertLessEqual(extended[1], 80)
        self.assertGreaterEqual(extended[1] + extended[3], 430)

    def test_screen_grid_rejects_repeated_collapsed_source_rows(self):
        grid = [
            ["任务编号", "任务名称", "进度"],
            ["TASK-001\nTASK-002", "需求确认\n界面设计", "7%\n14%"],
            ["TASK-003\nTASK-004", "图像预处理\n表格结构识别", "21%\n28%"],
        ]
        self.assertTrue(ocr_backend._screen_grid_has_collapsed_rows(grid))
        self.assertFalse(
            ocr_backend._screen_grid_has_collapsed_rows(
                [["编号", "名称", "进度"], ["1", "需求确认", "7%"], ["2", "界面设计", "14%"]]
            )
        )

    def test_split_unit_header_row_is_merged_without_shifting_body_rows(self):
        grid = [
            ["编号", "设备名称", "型号", "频率", "功率", "状态", ""],
            ["", "", "", "(MHz)", "(dBm)", "备注", ""],
            ["A001", "信号发生器", "SG-2200", "515.472", "-20", "复核", "校准完成"],
        ]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]

        merged, merged_confidence, spans = ocr_backend._merge_split_header_continuation_row(
            grid, confidence, []
        )

        self.assertEqual(len(merged), 2)
        self.assertEqual(
            merged[0],
            ["编号", "设备名称", "型号", "频率(MHz)", "功率(dBm)", "状态", "备注"],
        )
        self.assertEqual(merged[1][0], "A001")
        self.assertEqual(len(merged_confidence), 2)
        self.assertEqual(spans, [])

    def test_equal_column_ranks_realign_one_wrinkle_split_row(self):
        anchors = [20.0, 60.0, 100.0]
        grouped_rows = [
            [
                {"center_x": 20.0, "center_y": 10.0, "text": "编号", "score": 0.99},
                {"center_x": 60.0, "center_y": 12.0, "text": "数量", "score": 0.99},
                {"center_x": 100.0, "center_y": 14.0, "text": "状态", "score": 0.99},
            ],
            [
                {"center_x": 20.0, "center_y": 30.0, "text": "A01", "score": 0.99},
                {"center_x": 60.0, "center_y": 32.0, "text": "10", "score": 0.99},
            ],
            [
                {"center_x": 100.0, "center_y": 35.0, "text": "正常", "score": 0.99},
                {"center_x": 20.0, "center_y": 50.0, "text": "A02", "score": 0.99},
                {"center_x": 60.0, "center_y": 52.0, "text": "20", "score": 0.99},
                {"center_x": 100.0, "center_y": 55.0, "text": "待机", "score": 0.99},
            ],
        ]
        grid = [["编号", "数量", "状态"], ["A01", "10", ""], ["A02", "20", "正常"], ["", "", "待机"]]

        repaired = ocr_backend._realign_split_spatial_rows_by_column_rank(
            grid, grouped_rows, anchors, []
        )

        self.assertIsNotNone(repaired)
        self.assertEqual(
            repaired[0],
            [["编号", "数量", "状态"], ["A01", "10", "正常"], ["A02", "20", "待机"]],
        )

    def test_concatenated_header_labels_are_redistributed_into_empty_columns(self):
        grid = [
            ["任务编号", "任务名称", "优先级", "进度", "负责人", "计划完成", "当前状态"],
            ["TASK-001", "界面设计", "中", "7%", "陈晨", "2026-08-02", "已完成"],
        ]
        confidence = [[0.95] * 7 for _ in grid]
        grid[0][5] = "计划完成当前状态"
        grid[0][6] = ""
        confidence[0][6] = 0.0

        split, split_confidence, _ = ocr_backend._split_collapsed_header_data_row(
            grid, confidence, []
        )

        self.assertEqual(split[0][5:], ["计划完成", "当前状态"])
        self.assertEqual(len(split), 2)
        self.assertEqual(split_confidence[0][6], 0.95)

        self.assertEqual(
            ocr_backend._segment_concatenated_header_labels(
                "生产线 计划数量完成数量不良数量"
            ),
            ["生产线", "计划数量", "完成数量", "不良数量"],
        )

    def test_sparse_group_headers_move_to_merged_top_left_anchors(self):
        grid = [
            ["生产日报", "", "", "", "", "", ""],
            ["", "基础信息", "", "", "", "测量与判定", ""],
            ["日期", "班次", "生产线", "计划数量", "完成数量", "不良数量", "完成率"],
            ["2026-08-02", "夜班", "B线", "280", "277", "3", "98.2%"],
        ]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]
        spans = [{"row": 0, "column": 0, "row_span": 1, "column_span": 7, "role": "title"}]

        adjusted, adjusted_confidence, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            grid, confidence, spans
        )

        self.assertEqual(adjusted[1], ["基础信息", "", "", "测量与判定", "", "", ""])
        self.assertEqual(adjusted_confidence[1][0], 0.95)
        self.assertIn(
            {"row": 1, "column": 0, "row_span": 1, "column_span": 3, "role": "group_header"},
            adjusted_spans,
        )
        self.assertIn(
            {"row": 1, "column": 3, "row_span": 1, "column_span": 4, "role": "group_header"},
            adjusted_spans,
        )

    def test_sparse_group_headers_allow_title_before_shifted_detail_start(self):
        grid = [
            ["标题", "", "", ""],
            ["", "字段1", "字段2", "字段3"],
            ["1", "值1", "值2", "值3"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        adjusted, adjusted_confidence, adjusted_spans = (
            ocr_backend._infer_sparse_group_header_spans(
                grid,
                confidence,
                [],
                columns=[0, 100, 200, 300, 400],
                rows=[0, 30, 60, 90],
            )
        )

        self.assertEqual(adjusted, grid)
        self.assertEqual(adjusted_confidence, confidence)
        self.assertEqual(adjusted_spans, [])

    def test_sparse_group_headers_use_row_physical_dividers(self):
        ocr_backend._load_runtime()
        columns = [5 + 20 * index for index in range(14)]
        rows = [5, 35, 65, 95, 125]
        image = np.full((130, 270, 3), 255, dtype=np.uint8)
        for boundary in columns:
            cv2.line(image, (boundary, rows[2]), (boundary, rows[3]), (0, 0, 0), 2)
        for boundary_index in (0, 3, 6, 9, 13):
            boundary = columns[boundary_index]
            cv2.line(image, (boundary, rows[1]), (boundary, rows[2]), (0, 0, 0), 2)
        grid = [
            ["标题"] + [""] * 12,
            ["", "基础信息", "", "", "目标与测量", "", "", "过程记录", "", "", "质量判定", "", ""],
            [str(index) for index in range(13)],
            [str(index) for index in range(13)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        _, _, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            grid,
            confidence,
            [{"row": 0, "column": 0, "row_span": 1, "column_span": 13, "role": "title"}],
            image=image,
            columns=columns,
            rows=rows,
        )

        self.assertEqual(
            [
                (span["column"], span["column_span"])
                for span in adjusted_spans
                if span.get("role") == "group_header"
            ],
            [(0, 3), (3, 3), (6, 3), (9, 4)],
        )

    def test_sparse_group_headers_split_fused_prefix_from_repeated_suffix(self):
        ocr_backend._load_runtime()
        columns = [5 + 40 * index for index in range(8)]
        rows = [5, 45, 85, 125, 165, 205]
        image = np.full((210, 290, 3), 255, dtype=np.uint8)
        for boundary_index in (0, 3, 7):
            boundary = columns[boundary_index]
            cv2.line(image, (boundary, rows[0]), (boundary, rows[1]), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (boundary, rows[1]), (boundary, rows[-1]), (0, 0, 0), 2)
        grid = [
            ["基础信息业务数据", "", "业务数", "", "", "据", ""],
            [f"字段{column}" for column in range(7)],
            *[[str(row)] * 7 for row in range(3)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        adjusted, _, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            grid,
            confidence,
            [],
            image=image,
            columns=columns,
            rows=rows,
        )

        self.assertEqual(adjusted[0], ["基础信息", "", "", "业务数据", "", "", ""])
        self.assertEqual(
            [
                (span["column"], span["column_span"])
                for span in adjusted_spans
                if span.get("role") == "group_header"
            ],
            [(0, 3), (3, 4)],
        )

    def test_sparse_group_headers_use_page_evidence_for_combined_duplicate_row(self):
        ocr_backend._load_runtime()
        columns = [5 + 40 * index for index in range(8)]
        rows = [5, 45, 85, 125, 165, 205]
        image = np.full((210, 290, 3), 255, dtype=np.uint8)
        for boundary_index in (0, 3, 7):
            boundary = columns[boundary_index]
            cv2.line(image, (boundary, rows[0]), (boundary, rows[1]), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (boundary, rows[1]), (boundary, rows[-1]), (0, 0, 0), 2)
        grid = [
            ["基础信息  业务数据", "基础信息", "", "", "", "业务数据", ""],
            [f"字段{column}" for column in range(7)],
            *[[str(row)] * 7 for row in range(3)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        adjusted, _, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            grid,
            confidence,
            [],
            image=image,
            columns=columns,
            rows=rows,
            page_text_evidence=[
                ("基础信息", 65.0, 25.0, 0.99),
                ("业务数据", 205.0, 25.0, 0.99),
            ],
        )

        self.assertEqual(adjusted[0], ["基础信息", "", "", "业务数据", "", "", ""])
        self.assertEqual(
            [
                (span["column"], span["column_span"])
                for span in adjusted_spans
                if span.get("role") == "group_header"
            ],
            [(0, 3), (3, 4)],
        )

    def test_spurious_checkmark_is_removed_from_repeated_category_column(self):
        grid = [
            ["编号", "类型", "说明"],
            ["1", "常规项目", "甲"],
            ["2", "现场复核", "乙"],
            ["3", "☑常规项目", "丙"],
            ["4", "A区", "丁"],
            ["5", "常规项目", "戊"],
            ["6", "批次-L08", "己"],
        ]
        confidence = [[0.99 for _ in row] for row in grid]

        repaired = ocr_backend._repair_spurious_leading_checkmarks(grid, confidence)

        self.assertEqual(repaired, {(3, 1)})
        self.assertEqual(grid[3][1], "常规项目")
        self.assertEqual(confidence[3][1], 0.77)

    def test_spurious_checkmark_uses_adjacent_duplicate_with_global_repetition(self):
        grid = [
            ["编号", "区域", "批次", "数量"],
            ["1", "常规项目", "A区", "10"],
            ["2", "常规项目", "一车间", "11"],
            ["3", "现场复核", "批次-L08", "12"],
            ["4", "常规项目", "☑常规项目", "13"],
            ["5", "A区", "华东库", "14"],
            ["6", "一车间", "A区", "15"],
        ]
        confidence = [[0.99 for _ in row] for row in grid]

        repaired = ocr_backend._repair_spurious_leading_checkmarks(grid, confidence)

        self.assertEqual(repaired, {(4, 2)})
        self.assertEqual(grid[4][2], "常规项目")

    def test_spurious_checkmark_uses_strong_global_repetition_in_nonboolean_column(self):
        grid = [
            ["编号", "区域", "批次", "数量"],
            ["1", "综合部", "A区", "10"],
            ["2", "一车间", "综合部", "11"],
            ["3", "综合部", "批次-L08", "12"],
            ["4", "A区", "☑综合部", "13"],
            ["5", "综合部", "华东库", "14"],
            ["6", "一车间", "综合部", "15"],
        ]
        confidence = [[0.99 for _ in row] for row in grid]

        repaired = ocr_backend._repair_spurious_leading_checkmarks(grid, confidence)

        self.assertEqual(repaired, {(4, 2)})
        self.assertEqual(grid[4][2], "综合部")

    def test_spurious_checkmark_is_preserved_in_confirmation_column(self):
        grid = [
            ["编号", "裁判确认", "说明"],
            ["1", "常规项目", "甲"],
            ["2", "常规项目", "乙"],
            ["3", "☑常规项目", "丙"],
            ["4", "A区", "丁"],
            ["5", "常规项目", "戊"],
            ["6", "批次-L08", "己"],
        ]
        confidence = [[0.99 for _ in row] for row in grid]

        repaired = ocr_backend._repair_spurious_leading_checkmarks(grid, confidence)

        self.assertEqual(repaired, set())
        self.assertEqual(grid[3][1], "☑常规项目")

    def test_adjacent_duplicate_is_cleared_from_physically_blank_cell(self):
        ocr_backend._load_runtime()
        columns = [0, 100, 200, 300]
        rows = [0, 40, 80, 120, 160, 200, 240, 280]
        image = np.full((280, 300, 3), 255, dtype=np.uint8)
        for offset in (12, 28, 44, 60):
            cv2.rectangle(image, (100 + offset, 129), (100 + offset + 7, 150), (0, 0, 0), -1)
        grid = [
            ["编号", "上班时间", "下班时间"],
            ["1", "06:00", "15:15"],
            ["2", "20:00", "06:45"],
            ["3", "12:00", "12:00"],
            ["4", "18:45", "06:50"],
            ["5", "20:40", "22:50"],
            ["6", "10:10", "09:20"],
        ]
        confidence = [[0.99 for _ in row] for row in grid]
        confidence[3][2] = 0.77

        removed = ocr_backend._remove_weak_adjacent_duplicate_fills(
            image,
            grid,
            confidence,
            columns,
            rows,
        )

        self.assertEqual(removed, {(3, 2)})
        self.assertEqual(grid[3], ["3", "12:00", ""])

    def test_adjacent_duplicate_is_kept_when_both_cells_have_ink(self):
        ocr_backend._load_runtime()
        columns = [0, 100, 200, 300]
        rows = [0, 40, 80, 120, 160, 200, 240, 280]
        image = np.full((280, 300, 3), 255, dtype=np.uint8)
        for cell_left in (100, 200):
            for offset in (12, 28, 44, 60):
                cv2.rectangle(image, (cell_left + offset, 129), (cell_left + offset + 7, 150), (0, 0, 0), -1)
        grid = [
            ["编号", "区域", "批次"],
            ["1", "A区", "一车间"],
            ["2", "常规项目", "批次-L08"],
            ["3", "常规项目", "常规项目"],
            ["4", "华东库", "A区"],
            ["5", "现场复核", "一车间"],
            ["6", "A区", "批次-L08"],
        ]
        confidence = [[0.77 for _ in row] for row in grid]

        removed = ocr_backend._remove_weak_adjacent_duplicate_fills(
            image,
            grid,
            confidence,
            columns,
            rows,
        )

        self.assertEqual(removed, set())
        self.assertEqual(grid[3], ["3", "常规项目", "常规项目"])

    def test_single_ordinal_duplicate_is_cleared_from_blank_neighbour(self):
        ocr_backend._load_runtime()
        columns = [0, 100, 200, 300]
        rows = [0, 40, 80, 120, 160, 200]
        image = np.full((200, 300, 3), 255, dtype=np.uint8)
        cv2.rectangle(image, (60, 129), (67, 150), (0, 0, 0), -1)
        grid = [
            ["序号", "部门", "姓名"],
            ["1", "综合部", "张伟"],
            ["2", "一车间", "李娜"],
            ["5", "5", "赵敏"],
            ["6", "华东库", "陈晨"],
        ]
        confidence = [[0.77 for _ in row] for row in grid]

        removed = ocr_backend._remove_weak_adjacent_duplicate_fills(
            image,
            grid,
            confidence,
            columns,
            rows,
        )

        self.assertEqual(removed, {(3, 1)})
        self.assertEqual(grid[3], ["5", "", "赵敏"])

    def test_single_ordinal_duplicate_is_kept_when_neighbour_has_ink(self):
        ocr_backend._load_runtime()
        columns = [0, 100, 200, 300]
        rows = [0, 40, 80, 120, 160, 200]
        image = np.full((200, 300, 3), 255, dtype=np.uint8)
        for cell_left in (0, 100):
            cv2.rectangle(
                image,
                (cell_left + 60, 129),
                (cell_left + 67, 150),
                (0, 0, 0),
                -1,
            )
        grid = [
            ["序号", "数量", "姓名"],
            ["1", "10", "张伟"],
            ["2", "12", "李娜"],
            ["5", "5", "赵敏"],
            ["6", "8", "陈晨"],
        ]
        confidence = [[0.77 for _ in row] for row in grid]

        removed = ocr_backend._remove_weak_adjacent_duplicate_fills(
            image,
            grid,
            confidence,
            columns,
            rows,
        )

        self.assertEqual(removed, set())
        self.assertEqual(grid[3], ["5", "5", "赵敏"])

    def test_categorical_duplicate_is_cleared_from_blank_neighbour(self):
        ocr_backend._load_runtime()
        columns = [0, 100, 200, 300]
        rows = [0, 40, 80, 120, 160, 200]
        for duplicated_text in ("A区", "中"):
            with self.subTest(duplicated_text=duplicated_text):
                image = np.full((200, 300, 3), 255, dtype=np.uint8)
                for offset in (12, 28, 44, 60):
                    cv2.rectangle(
                        image,
                        (100 + offset, 129),
                        (100 + offset + 7, 150),
                        (0, 0, 0),
                        -1,
                    )
                grid = [
                    ["编号", "区域", "部门"],
                    ["1", "综合部", "张伟"],
                    ["2", "一车间", "李娜"],
                    ["3", duplicated_text, duplicated_text],
                    ["4", "华东库", "陈晨"],
                ]
                confidence = [[0.77 for _ in row] for row in grid]

                removed = ocr_backend._remove_weak_adjacent_duplicate_fills(
                    image,
                    grid,
                    confidence,
                    columns,
                    rows,
                )

                self.assertEqual(removed, {(3, 2)})
                self.assertEqual(grid[3], ["3", duplicated_text, ""])

    def test_incomplete_merged_metadata_is_recovered_by_small_multiview_consensus(self):
        ocr_backend._load_runtime()
        image = np.full((250, 400, 3), 245, dtype=np.uint8)
        cv2.rectangle(image, (230, 12), (320, 30), (0, 0, 0), -1)
        grid = [
            ["采购验收表", "", "", ""],
            ["单位：物业部", "", "批", ""],
            ["编码", "品名", "数量", "结果"],
            ["A01", "项目甲", "12", "完成"],
            ["A02", "项目乙", "13", "完成"],
            ["A03", "项目丙", "14", "复核"],
        ]
        confidence = [[0.77 if value else 0.0 for value in row] for row in grid]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 4, "role": "title"},
            {"row": 1, "column": 0, "row_span": 1, "column_span": 2, "role": "group_header"},
            {"row": 1, "column": 2, "row_span": 1, "column_span": 2, "role": "group_header"},
        ]
        output = SimpleNamespace(
            txts=["批次：B61", "批次：B61", "批次：B61", "批次:B61"],
            scores=[0.98, 0.97, 0.96, 0.95],
            imgs=None,
        )
        engine = SimpleNamespace(fast_text_rec=Mock(return_value=output))

        recovered, scores = ocr_backend._recover_incomplete_merged_metadata_text(
            image,
            grid,
            confidence,
            spans,
            [0, 100, 200, 300, 400],
            [0, 50, 100, 150, 200, 250],
            engine,
            row_offset=1,
        )

        self.assertEqual(recovered, {(1, 2)})
        self.assertEqual(grid[1], ["单位：物业部", "", "批次：B61", ""])
        self.assertEqual(len(scores), 4)

    def test_incomplete_merged_metadata_keeps_text_without_three_view_consensus(self):
        ocr_backend._load_runtime()
        image = np.full((250, 400, 3), 245, dtype=np.uint8)
        grid = [
            ["采购验收表", "", "", ""],
            ["单位：物业部", "", "批", ""],
            ["编码", "品名", "数量", "结果"],
            ["A01", "项目甲", "12", "完成"],
            ["A02", "项目乙", "13", "完成"],
            ["A03", "项目丙", "14", "复核"],
        ]
        confidence = [[0.77 if value else 0.0 for value in row] for row in grid]
        spans = [
            {"row": 1, "column": 2, "row_span": 1, "column_span": 2, "role": "group_header"},
        ]
        output = SimpleNamespace(
            txts=["批次：B61", "批号：B61", "批次：861", "批次：B61"],
            scores=[0.98, 0.97, 0.96, 0.95],
            imgs=None,
        )
        engine = SimpleNamespace(fast_text_rec=Mock(return_value=output))

        recovered, _ = ocr_backend._recover_incomplete_merged_metadata_text(
            image,
            grid,
            confidence,
            spans,
            [0, 100, 200, 300, 400],
            [0, 50, 100, 150, 200, 250],
            engine,
            row_offset=1,
        )

        self.assertEqual(recovered, set())
        self.assertEqual(grid[1][2], "批")

    def test_group_headers_do_not_trigger_metadata_model_reload(self):
        ocr_backend._load_runtime()
        image = np.full((250, 400, 3), 245, dtype=np.uint8)
        grid = [
            ["采购验收表", "", "", ""],
            ["基础信息", "", "业务数据", ""],
            ["编码", "品名", "数量", "结果"],
            ["A01", "项目甲", "12", "完成"],
            ["A02", "项目乙", "13", "完成"],
            ["A03", "项目丙", "14", "复核"],
        ]
        confidence = [[0.77 if value else 0.0 for value in row] for row in grid]
        spans = [
            {"row": 1, "column": 0, "row_span": 1, "column_span": 2, "role": "group_header"},
            {"row": 1, "column": 2, "row_span": 1, "column_span": 2, "role": "group_header"},
        ]
        recognizer = Mock()
        engine = SimpleNamespace(fast_text_rec=recognizer)

        recovered, scores = ocr_backend._recover_incomplete_merged_metadata_text(
            image,
            grid,
            confidence,
            spans,
            [0, 100, 200, 300, 400],
            [0, 50, 100, 150, 200, 250],
            engine,
            row_offset=1,
        )

        self.assertEqual(recovered, set())
        self.assertEqual(scores, [])
        recognizer.assert_not_called()

    def test_sparse_group_headers_use_centered_page_text_evidence(self):
        columns = [1, 117, 273, 434, 550, 667, 783, 900, 1016, 1203, 1320, 1436, 1552, 1668, 1784, 1904]
        rows = [25, 67, 102, 137]
        image = np.full((140, 1910, 3), 245, dtype=np.uint8)
        grid = [
            ["", "基础信息", "", "", "", "", "目标与测量", "", "", "过程记录", "", "质量判定", "", "追溯信息", ""],
            ["区域", "记录编号", "对象名称", "长文本说明", "当前值", "目标值", "单位", "来源系统", "更新时间", "责任人", "一级分类", "二级分类", "三级分类", "风险说明", "处理建议"],
            [str(index) for index in range(15)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("基础信息", 276.0, 46.0, 1.0),
            ("目标与测量", 878.0, 46.0, 0.999),
            ("过程记录", 1319.5, 46.0, 0.999),
            ("质量判定", 1552.5, 46.0, 0.999),
            ("追溯信息", 1784.0, 46.0, 0.999),
        ]

        adjusted, _, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            grid,
            confidence,
            [],
            image=image,
            columns=columns,
            rows=rows,
            page_text_evidence=evidence,
        )

        self.assertEqual(
            [
                (span["column"], span["column_span"])
                for span in adjusted_spans
                if span.get("role") == "group_header"
            ],
            [(0, 4), (4, 5), (9, 2), (11, 2), (13, 2)],
        )
        self.assertEqual(
            [adjusted[0][index] for index in (0, 4, 9, 11, 13)],
            ["基础信息", "目标与测量", "过程记录", "质量判定", "追溯信息"],
        )

    def test_sparse_group_headers_rejoin_page_fragments_and_repeated_labels(self):
        ocr_backend._load_runtime()
        columns = [
            0, 112, 262, 418, 530, 710, 822, 934, 1047,
            1160, 1272, 1508, 1621, 1733, 1846, 1958, 2144,
        ]
        rows = list(range(0, 321, 40))
        image = np.full((320, 2144, 3), 245, dtype=np.uint8)
        grid = [
            [
                "", "基础信息", "", "", "目标与测量", "", "", "过程",
                "记录", "", "质量判定", "", "追溯", "信息", "", "基础信息",
            ],
            [f"字段{column}" for column in range(16)],
            *[[str(row)] * 16 for row in range(6)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("基础信息", 209.0, 20.0, 1.0),
            ("目标与测量", 621.0, 20.0, 0.999),
            ("过程记录", 1048.0, 20.0, 0.999),
            ("质量判定", 1447.0, 20.0, 0.999),
            ("追溯信息", 1733.5, 20.0, 0.999),
            ("基础信息", 1993.5, 20.0, 1.0),
        ]

        adjusted, _, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            grid,
            confidence,
            [],
            image=image,
            columns=columns,
            rows=rows,
            page_text_evidence=evidence,
        )

        self.assertEqual(
            [
                (span["column"], span["column_span"])
                for span in adjusted_spans
                if span.get("role") == "group_header"
            ],
            [(0, 3), (3, 3), (6, 4), (10, 2), (12, 2), (14, 2)],
        )
        self.assertEqual(
            [adjusted[0][index] for index in (0, 3, 6, 10, 12, 14)],
            ["基础信息", "目标与测量", "过程记录", "质量判定", "追溯信息", "基础信息"],
        )

    def test_sparse_group_headers_collapse_adjacent_partial_and_complete_label(self):
        columns = [
            0, 112, 262, 418, 530, 710, 822, 934, 1047,
            1160, 1272, 1508, 1621, 1733, 1846, 1958, 2144,
        ]
        rows = list(range(0, 321, 40))
        image = np.full((320, 2144, 3), 245, dtype=np.uint8)
        grid = [
            [
                "基础信息", "", "", "目标与测量", "", "", "过程记录", "", "", "",
                "质量判定", "", "追溯", "追溯信息", "基础信息", "",
            ],
            [f"字段{column}" for column in range(16)],
            *[[str(row)] * 16 for row in range(6)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("基础信息", 209.0, 20.0, 1.0),
            ("目标与测量", 621.0, 20.0, 0.999),
            ("过程记录", 1048.0, 20.0, 0.999),
            ("质量判定", 1447.0, 20.0, 0.999),
            ("追溯信息", 1733.5, 20.0, 0.999),
            ("基础信息", 1993.5, 20.0, 1.0),
        ]

        adjusted, _, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            grid,
            confidence,
            [],
            image=image,
            columns=columns,
            rows=rows,
            page_text_evidence=evidence,
        )

        self.assertEqual(
            [(span["column"], span["column_span"]) for span in adjusted_spans],
            [(0, 3), (3, 3), (6, 4), (10, 2), (12, 2), (14, 2)],
        )
        self.assertEqual(adjusted[0][12:14], ["追溯信息", ""])
        self.assertLessEqual(confidence[0][12], 0.77)

    def test_sparse_group_headers_ignore_spreadsheet_column_ruler_in_header_band(self):
        columns = [51, 166, 328, 492, 613, 730, 851, 972, 1091, 1215, 1335, 1460, 1582, 1705, 1836]
        rows = list(range(0, 321, 40))
        image = np.full((320, 1836, 3), 245, dtype=np.uint8)
        grid = [
            ["A", "B 基础信息", "C", "目标与测量", "", "", "过程记录", "", "", "", "", "质量判定", "", ""],
            [f"字段{column}" for column in range(14)],
            *[[str(row)] * 14 for row in range(6)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("A", 113.0, 5.0, 0.999),
            ("B", 253.5, 5.0, 0.999),
            ("C", 416.5, 5.0, 0.855),
            ("基础信息", 274.5, 22.0, 1.0),
            ("目标与测量", 618.5, 22.0, 1.0),
            ("过程记录", 977.5, 22.0, 1.0),
            ("质量判定", 1527.5, 22.0, 1.0),
        ]

        adjusted, _, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            grid,
            confidence,
            [],
            image=image,
            columns=columns,
            rows=rows,
            page_text_evidence=evidence,
        )

        self.assertEqual(
            [
                (span["column"], span["column_span"])
                for span in adjusted_spans
                if span.get("role") == "group_header"
            ],
            [(0, 3), (3, 2), (5, 4), (9, 5)],
        )
        self.assertEqual(
            [adjusted[0][index] for index in (0, 3, 5, 9)],
            ["基础信息", "目标与测量", "过程记录", "质量判定"],
        )

    def test_sparse_group_headers_keep_a_trailing_single_column_header_unmerged(self):
        ocr_backend._load_runtime()
        columns = [20, 112, 211, 328, 451, 575, 692, 792, 882, 972, 1063, 1153, 1244]
        rows = [19, 47, 73, 98, 121]
        image = np.full((125, 1260, 3), 245, dtype=np.uint8)
        cv2.line(
            image,
            (columns[-2], rows[0]),
            (columns[-2], rows[1]),
            (0, 0, 0),
            2,
        )
        grid = [
            ["", "基础信息", "", "", "", "目标与测量", "", "", "", "过程记录", "", "质量判定"],
            ["序号", "销售单号", "客户名称", "产品编码", "产品名称", "规格型号", "订单数量", "已发数量", "待发数量", "单位", "销售单价", "折扣率"],
            [str(column) for column in range(12)],
            [str(column) for column in range(12)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("基础信息", 174.0, 32.0, 1.0),
            ("目标与测量", 605.5, 33.0, 1.0),
            ("过程记录", 1017.0, 34.0, 1.0),
            ("质量判定", 1198.0, 33.0, 1.0),
        ]

        adjusted, _, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            grid,
            confidence,
            [],
            image=image,
            columns=columns,
            rows=rows,
            page_text_evidence=evidence,
        )

        self.assertEqual(
            [
                (span["column"], span["column_span"])
                for span in adjusted_spans
                if span.get("role") == "group_header"
            ],
            [(0, 3), (3, 5), (8, 3)],
        )
        self.assertEqual(adjusted[0][11], "质量判定")

    def test_sparse_group_headers_use_half_up_midpoint_with_trailing_status_column(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 1300, 100))
        rows = list(range(0, 361, 40))
        image = np.full((360, 1200, 3), 245, dtype=np.uint8)
        grid = [
            ["标题", *([""] * 11)],
            ["", "", "基础信息", "", "", "数量与金额", "", "", "过程记录", "", "", "状态判定"],
            [f"列{column}" for column in range(12)],
            *[[str(column) for column in range(12)] for _ in range(6)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        title = {"row": 0, "column": 0, "row_span": 1, "column_span": 12, "role": "title"}

        adjusted, _, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            grid,
            confidence,
            [title],
            image=image,
            columns=columns,
            rows=rows,
            page_text_evidence=[("状态判定", 1150.0, 60.0, 0.99)],
        )

        self.assertEqual(
            [
                (span["column"], span["column_span"])
                for span in adjusted_spans
                if span.get("role") == "group_header"
            ],
            [(0, 4), (4, 3), (7, 4)],
        )
        self.assertEqual(adjusted[1][11], "状态判定")

    def test_sparse_group_headers_keep_proven_last_column_unmerged(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 701, 100))
        rows = list(range(0, 181, 30))
        image = np.full((180, 700, 3), 245, dtype=np.uint8)
        for boundary in (columns[4], columns[6]):
            cv2.line(image, (boundary, rows[0]), (boundary, rows[1]), (0, 0, 0), 2)
        grid = [
            ["", "基础信息", "", "", "数量与金额", "", "过程记录"],
            ["序号", "学号", "姓名", "班级", "课程", "任课教师", "星期"],
            *[[str(column) for column in range(7)] for _ in range(4)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("基础信息", 200.0, 15.0, 1.0),
            ("数量与金额", 500.0, 15.0, 1.0),
            ("过程记录", 650.0, 15.0, 1.0),
        ]

        adjusted, _, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            grid,
            confidence,
            [],
            image=image,
            columns=columns,
            rows=rows,
            page_text_evidence=evidence,
        )

        self.assertEqual(
            [(span["column"], span["column_span"]) for span in adjusted_spans],
            [(0, 4), (4, 2)],
        )
        self.assertEqual(adjusted[0], ["基础信息", "", "", "", "数量与金额", "", "过程记录"])

    def test_sparse_group_headers_keep_semantic_trailing_trace_column_unmerged(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 1101, 100))
        rows = list(range(0, 211, 30))
        image = np.full((211, 1101, 3), 245, dtype=np.uint8)
        for boundary in (columns[4], columns[6], columns[10]):
            cv2.line(image, (boundary, rows[0]), (boundary, rows[1]), (0, 0, 0), 2)
        grid = [
            ["", "基础信息", "", "", "", "数量与金额", "过程记录", "", "状态判定", "", "追溯信息"],
            ["序号", "样品编号", "样品名称", "样品类型", "采样地点", "采样时间", "检测项目", "检测方法", "仪器编号", "检出限", "检测结果"],
            *[[str(column) for column in range(11)] for _ in range(5)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("基础信息", 200.0, 15.0, 1.0),
            ("数量与金额", 500.0, 15.0, 1.0),
            ("过程记录", 700.0, 15.0, 1.0),
            ("状态判定", 900.0, 15.0, 1.0),
            ("追溯信息", 1050.0, 15.0, 1.0),
        ]

        adjusted, _, spans = ocr_backend._infer_sparse_group_header_spans(
            grid,
            confidence,
            [],
            image=image,
            columns=columns,
            rows=rows,
            page_text_evidence=evidence,
        )

        self.assertEqual(
            [(span["column"], span["column_span"]) for span in spans],
            [(0, 4), (4, 2), (6, 2), (8, 2)],
        )
        self.assertEqual(
            [adjusted[0][column] for column in (0, 4, 6, 8, 10)],
            ["基础信息", "数量与金额", "过程记录", "状态判定", "追溯信息"],
        )

    def test_sparse_group_headers_ignore_sequential_left_row_ruler(self):
        ocr_backend._load_runtime()
        columns = [0, 107, 210, 306, 415, 525, 620, 715, 810, 904, 997, 1091, 1185, 1280]
        rows = [35, 68, 99, 126, 152, 178, 204, 230, 256, 282]
        image = np.full((285, 1285, 3), 245, dtype=np.uint8)
        for boundary_index in (0, 3, 8, 13):
            boundary = columns[boundary_index]
            cv2.line(image, (boundary, rows[0]), (boundary, rows[1]), (0, 0, 0), 2)
        grid = [
            ["", "基础信息", "", "", "", "目标与测量", "", "", "", "", "过程记录", "", ""],
            ["", "本期", "", "实际", "", "计划", "", "上限", "", "实际", "", "下限", "累计"],
            ["序号", "区域", "部门", "项目", "一月", "二月", "三月", "一季度", "四月", "五月", "六月", "二季度", "上半年"],
            *[[str(column) for column in range(13)] for _ in range(6)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("1", 17.5, 48.0, 0.999),
            ("基础信息", 153.0, 50.5, 1.0),
            ("目标与测量", 558.0, 52.5, 1.0),
            ("过程记录", 1045.0, 54.0, 1.0),
            ("2", 18.0, 82.0, 0.999),
            ("本期", 130.5, 82.5, 1.0),
            ("实际", 340.5, 83.5, 1.0),
            ("计划", 549.0, 83.5, 1.0),
            ("上限", 739.5, 85.0, 1.0),
            ("实际", 928.5, 85.5, 1.0),
            ("下限", 1117.0, 86.0, 1.0),
            ("累计", 1257.0, 86.0, 1.0),
            ("3", 18.0, 111.0, 0.999),
        ]

        adjusted, adjusted_confidence, adjusted_spans = (
            ocr_backend._infer_sparse_group_header_spans(
                grid,
                confidence,
                [],
                image=image,
                columns=columns,
                rows=rows,
                page_text_evidence=evidence,
            )
        )
        adjusted, _, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            adjusted,
            adjusted_confidence,
            adjusted_spans,
            image=image,
            columns=columns,
            rows=rows,
            page_text_evidence=evidence,
        )

        self.assertEqual(
            sorted(
                [
                (span["row"], span["column"], span["column_span"])
                for span in adjusted_spans
                if span.get("role") == "group_header"
                ]
            ),
            [
                (0, 0, 3),
                (0, 3, 5),
                (0, 8, 5),
                (1, 0, 2),
                (1, 2, 2),
                (1, 4, 2),
                (1, 6, 2),
                (1, 8, 2),
                (1, 10, 2),
            ],
        )
        self.assertEqual(adjusted[1][12], "累计")

    def test_certified_physical_group_headers_use_continuous_dividers(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 1100, 100))
        rows = list(range(0, 361, 40))
        image = np.full((360, 1000, 3), 255, dtype=np.uint8)
        for boundary in rows:
            cv2.line(image, (0, boundary), (999, boundary), (0, 0, 0), 2)
        for boundary in (0, 300, 500, 900, 999):
            cv2.line(image, (boundary, 0), (boundary, 40), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (min(boundary, 999), 40), (min(boundary, 999), 359), (0, 0, 0), 2)
        grid = [
            ["基础信息", "", "目标与测量", "", "过程记录", "", "", "", "质量判定", ""],
            [f"列{column}" for column in range(10)],
            *[[str(column) for column in range(10)] for _ in range(7)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("基础信息", 150.0, 20.0, 0.99),
            ("目标与测量", 400.0, 20.0, 0.99),
            ("过程记录", 700.0, 20.0, 0.99),
            ("质量判定", 950.0, 20.0, 0.99),
        ]

        spans = ocr_backend._recover_certified_physical_group_header_spans(
            image, columns, rows, grid, confidence, evidence, []
        )

        self.assertEqual(
            grid[0],
            ["基础信息", "", "", "目标与测量", "", "过程记录", "", "", "", "质量判定"],
        )
        self.assertEqual(
            [(span["column"], span["column_span"]) for span in spans],
            [(0, 3), (3, 2), (5, 4)],
        )
        self.assertTrue(all(confidence[0][column] <= 0.77 for column in (0, 3, 5, 9)))

    def test_certified_group_headers_reread_one_missing_physical_interval(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 801, 100))
        rows = list(range(0, 321, 40))
        image = np.full((320, 800, 3), 255, dtype=np.uint8)
        for boundary in rows:
            cv2.line(image, (0, min(boundary, 319)), (799, min(boundary, 319)), (0, 0, 0), 2)
        for boundary in (0, 799):
            cv2.line(image, (boundary, 0), (boundary, 79), (0, 0, 0), 2)
        cv2.line(image, (400, 40), (400, 79), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (min(boundary, 799), 80), (min(boundary, 799), 319), (0, 0, 0), 2)
        grid = [
            ["施工现场安全检查表", "", "", "", "", "", "", ""],
            ["业务记录", "", "", "", "", "", "", ""],
            ["编号", "检查区域", "检查项目", "风险等级", "责任单位", "整改期限", "复查结果", "检查人"],
            *[[str(index)] * 8 for index in range(1, 6)],
        ]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 8, "role": "title"},
            {"row": 1, "column": 0, "row_span": 1, "column_span": 8, "role": "title"},
        ]
        output = SimpleNamespace(
            txts=["基础资料"] * 4,
            scores=[0.95, 0.94, 0.93, 0.92],
            imgs=None,
        )
        engine = SimpleNamespace(
            text_rec=Mock(return_value=output),
            server_text_rec=Mock(return_value=output),
        )

        observed_spans = ocr_backend._recover_certified_physical_group_header_spans(
            image,
            columns,
            rows,
            grid,
            confidence,
            [("业务记录", 600.0, 60.0, 0.96)],
            spans,
            ocr_engine=engine,
        )

        self.assertEqual(grid[1], ["基础资料", "", "", "", "业务记录", "", "", ""])
        self.assertEqual(
            [
                (span["column"], span["column_span"], span["role"])
                for span in observed_spans
                if span["row"] == 1
            ],
            [(0, 4, "group_header"), (4, 4, "group_header")],
        )

    def test_semantic_detail_row_is_not_displaced_by_stronger_data_rules(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 1100, 100))
        rows = list(range(0, 361, 40))
        image = np.full((360, 1000, 3), 255, dtype=np.uint8)
        for boundary in rows:
            cv2.line(image, (0, boundary), (999, boundary), (0, 0, 0), 2)
        for boundary in (0, 300, 500, 900, 999):
            cv2.line(image, (boundary, 0), (boundary, 40), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (min(boundary, 999), 80), (min(boundary, 999), 359), (0, 0, 0), 2)
        grid = [
            ["基础信息", "", "", "目标与测量", "", "过程记录", "", "", "", "质量判定"],
            ["序号", "采购单号", "供应商", "物料编码", "物料名称", "规格型号", "订购数量", "已交数量", "未交数量", "单位"],
            *[[str(column) for column in range(10)] for _ in range(7)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("基础信息", 150.0, 20.0, 0.99),
            ("目标与测量", 400.0, 20.0, 0.99),
            ("过程记录", 700.0, 20.0, 0.99),
            ("质量判定", 950.0, 20.0, 0.99),
        ]

        spans = ocr_backend._recover_certified_physical_group_header_spans(
            image,
            columns,
            rows,
            grid,
            confidence,
            evidence,
            [],
        )

        self.assertEqual(
            [(span["column"], span["column_span"]) for span in spans],
            [(0, 3), (3, 2), (5, 4)],
        )

    def test_certified_physical_group_headers_use_ruler_filtered_page_centers(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 1500, 100))
        rows = list(range(0, 361, 40))
        image = np.full((360, 1400, 3), 255, dtype=np.uint8)
        for boundary in (0, 500, 900, 1399):
            cv2.line(image, (boundary, 0), (boundary, 40), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(
                image,
                (min(boundary, 1399), 40),
                (min(boundary, 1399), 359),
                (0, 0, 0),
                2,
            )
        grid = [
            [
                "AB基础信息C目标与测量",
                "",
                "",
                "",
                "",
                "过程记录",
                "",
                "",
                "",
                "质量判定",
                "",
                "",
                "",
                "",
            ],
            [f"列{column}" for column in range(14)],
            *[[str(column) for column in range(14)] for _ in range(7)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("A", 50.0, 5.0, 0.99),
            ("B", 150.0, 5.0, 0.99),
            ("C", 250.0, 5.0, 0.99),
            ("基础信息", 150.0, 24.0, 0.99),
            ("目标与测量", 400.0, 24.0, 0.99),
            ("过程记录", 700.0, 24.0, 0.99),
            ("质量判定", 1150.0, 24.0, 0.99),
        ]

        spans = ocr_backend._recover_certified_physical_group_header_spans(
            image, columns, rows, grid, confidence, evidence, []
        )

        self.assertEqual(
            [(span["column"], span["column_span"]) for span in spans],
            [(0, 3), (3, 2), (5, 4), (9, 5)],
        )
        self.assertEqual(
            [grid[0][column] for column in (0, 3, 5, 9)],
            ["基础信息", "目标与测量", "过程记录", "质量判定"],
        )

    def test_certified_physical_group_headers_rejoin_fragments_after_title(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 1400, 100))
        rows = list(range(0, 401, 40))
        image = np.full((400, 1300, 3), 255, dtype=np.uint8)
        for boundary in rows:
            cv2.line(image, (0, boundary), (1299, boundary), (0, 0, 0), 2)
        cv2.line(image, (0, 0), (0, 40), (0, 0, 0), 2)
        cv2.line(image, (1299, 0), (1299, 40), (0, 0, 0), 2)
        for boundary in (0, 200, 700, 1000, 1200, 1299):
            cv2.line(image, (boundary, 40), (boundary, 80), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (min(boundary, 1299), 80), (min(boundary, 1299), 399), (0, 0, 0), 2)
        grid = [
            ["IT资产网络 — 2026-09 批次", *([""] * 12)],
            ["基础信息", "", "目标与测量", "", "", "", "过程记录", "", "", "质量", "判定", "追溯信息", ""],
            [f"列{column}" for column in range(13)],
            *[[str(column) for column in range(13)] for _ in range(7)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        title = {"row": 0, "column": 0, "row_span": 1, "column_span": 13, "role": "title"}
        evidence = [
            ("基础信息", 100.0, 60.0, 0.99),
            ("目标与测量", 450.0, 60.0, 0.99),
            ("过程记录", 850.0, 60.0, 0.99),
            ("质量", 1070.0, 60.0, 0.99),
            ("判定", 1130.0, 60.0, 0.99),
            ("追溯信息", 1250.0, 60.0, 0.99),
        ]

        spans = ocr_backend._recover_certified_physical_group_header_spans(
            image, columns, rows, grid, confidence, evidence, [title]
        )

        self.assertEqual(
            grid[1],
            ["基础信息", "", "目标与测量", "", "", "", "", "过程记录", "", "", "质量判定", "", "追溯信息"],
        )
        self.assertEqual(
            [(span["row"], span["column"], span["column_span"]) for span in spans],
            [(0, 0, 13), (1, 0, 2), (1, 2, 5), (1, 7, 3), (1, 10, 2)],
        )

    def test_certified_physical_group_headers_do_not_duplicate_existing_row(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 1700, 100))
        rows = list(range(0, 401, 40))
        image = np.full((400, 1600, 3), 255, dtype=np.uint8)
        grid = [
            ["库存出入库 — 2026-10 批次", *([""] * 15)],
            ["基础信息", "", "", "", "", "目标与测量", "", "", "过程记录", "", "", "质量判定", "", "", "", ""],
            [f"列{column}" for column in range(16)],
            *[[str(column) for column in range(16)] for _ in range(7)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        existing = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 16, "role": "title"},
            {"row": 1, "column": 0, "row_span": 1, "column_span": 5, "role": "group_header"},
            {"row": 1, "column": 5, "row_span": 1, "column_span": 3, "role": "group_header"},
            {"row": 1, "column": 8, "row_span": 1, "column_span": 3, "role": "group_header"},
            {"row": 1, "column": 11, "row_span": 1, "column_span": 5, "role": "group_header"},
        ]

        spans = ocr_backend._recover_certified_physical_group_header_spans(
            image, columns, rows, grid, confidence, [], existing
        )

        self.assertEqual(spans, existing)

    def test_certified_physical_group_headers_rebuild_partial_existing_row(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 1700, 100))
        rows = list(range(0, 401, 40))
        image = np.full((400, 1600, 3), 255, dtype=np.uint8)
        for boundary in rows:
            cv2.line(image, (0, boundary), (1599, boundary), (0, 0, 0), 2)
        for boundary in (0, 300, 600, 1000, 1200, 1400, 1599):
            cv2.line(image, (boundary, 0), (boundary, 40), (0, 0, 0), 2)
        cv2.line(image, (1300, 0), (1300, 40), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (min(boundary, 1599), 40), (min(boundary, 1599), 399), (0, 0, 0), 2)
        grid = [
            ["基础信息", "", "", "目标与测量", "", "", "过程记录", "", "", "", "质量判定", "", "追溯", "信息", "基础信息", ""],
            [f"列{column}" for column in range(16)],
            *[[str(column) for column in range(16)] for _ in range(8)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("基础信息", 150.0, 20.0, 0.99),
            ("目标与测量", 450.0, 20.0, 0.99),
            ("过程记录", 800.0, 20.0, 0.99),
            ("质量判定", 1100.0, 20.0, 0.99),
            ("追溯信息", 1300.0, 20.0, 0.99),
            ("基础信息", 1500.0, 20.0, 0.99),
        ]
        partial = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 3, "role": "group_header"},
            {"row": 0, "column": 3, "row_span": 1, "column_span": 3, "role": "group_header"},
            {"row": 0, "column": 6, "row_span": 1, "column_span": 4, "role": "group_header"},
            {"row": 0, "column": 10, "row_span": 1, "column_span": 2, "role": "group_header"},
            {"row": 0, "column": 14, "row_span": 1, "column_span": 2, "role": "group_header"},
        ]

        spans = ocr_backend._recover_certified_physical_group_header_spans(
            image, columns, rows, grid, confidence, evidence, partial
        )

        self.assertEqual(
            [(span["column"], span["column_span"]) for span in spans],
            [(0, 3), (3, 3), (6, 4), (10, 2), (12, 2), (14, 2)],
        )
        self.assertEqual(grid[0][12:14], ["追溯信息", ""])

    def test_certified_physical_group_headers_use_unique_centered_evidence(self):
        ocr_backend._load_runtime()
        columns = [0, 106, 248, 396, 502, 610, 716, 822, 930, 1036, 1144, 1250, 1357, 1468]
        rows = list(range(0, 401, 40))
        image = np.full((400, 1468, 3), 255, dtype=np.uint8)
        for boundary in rows:
            cv2.line(image, (0, boundary), (1467, boundary), (0, 0, 0), 2)
        cv2.line(image, (0, 40), (0, 80), (0, 0, 0), 2)
        cv2.line(image, (1467, 40), (1467, 80), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (min(boundary, 1467), 80), (min(boundary, 1467), 399), (0, 0, 0), 2)
        grid = [
            ["IT资产网络 — 2026-09 批次", *([""] * 12)],
            ["基础信息", "", "目标与测量", "", "", "", "过程记录", "", "", "质量判定", "", "追溯信息", ""],
            [f"列{column}" for column in range(13)],
            *[[str(column) for column in range(13)] for _ in range(7)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        title = {"row": 0, "column": 0, "row_span": 1, "column_span": 13, "role": "title"}
        evidence = [
            ("基础信息", 124.0, 60.0, 0.99),
            ("目标与测量", 535.0, 60.0, 0.99),
            ("过程记录", 983.0, 60.0, 0.99),
            ("质量判定", 1250.0, 60.0, 0.99),
            ("追溯信息", 1412.0, 60.0, 0.99),
        ]

        spans = ocr_backend._recover_certified_physical_group_header_spans(
            image, columns, rows, grid, confidence, evidence, [title]
        )

        self.assertEqual(
            grid[1],
            ["基础信息", "", "目标与测量", "", "", "", "", "过程记录", "", "", "质量判定", "", "追溯信息"],
        )
        self.assertEqual(
            [(span["row"], span["column"], span["column_span"]) for span in spans],
            [(0, 0, 13), (1, 0, 2), (1, 2, 5), (1, 7, 3), (1, 10, 2)],
        )

    def test_certified_physical_group_headers_support_amount_and_status_groups(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 1300, 100))
        rows = list(range(0, 401, 40))
        image = np.full((400, 1200, 3), 255, dtype=np.uint8)
        for boundary in rows:
            cv2.line(image, (0, boundary), (1199, boundary), (0, 0, 0), 2)
        for boundary in (0, 300, 700, 1000, 1199):
            cv2.line(image, (boundary, 40), (boundary, 80), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (min(boundary, 1199), 80), (min(boundary, 1199), 399), (0, 0, 0), 2)
        grid = [
            ["登记表  2026年11月", *([""] * 11)],
            ["", "基础信息", "", "", "数量与", "", "", "", "过程记录", "", "", "状态判定"],
            [f"列{column}" for column in range(12)],
            *[[str(column) for column in range(12)] for _ in range(7)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        title = {"row": 0, "column": 0, "row_span": 1, "column_span": 12, "role": "title"}
        evidence = [
            ("基础信息", 150.0, 60.0, 0.99),
            ("数量与金额", 500.0, 60.0, 0.99),
            ("过程记录", 850.0, 60.0, 0.99),
            ("状态判定", 1100.0, 60.0, 0.99),
        ]

        spans = ocr_backend._recover_certified_physical_group_header_spans(
            image, columns, rows, grid, confidence, evidence, [title]
        )

        self.assertEqual(
            [(span["row"], span["column"], span["column_span"]) for span in spans],
            [(0, 0, 12), (1, 0, 3), (1, 3, 4), (1, 7, 3), (1, 10, 2)],
        )
        self.assertEqual(
            [grid[1][index] for index in (0, 3, 7, 10)],
            ["基础信息", "数量与金额", "过程记录", "状态判定"],
        )

    def test_certified_physical_group_headers_reject_fully_ruled_header(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 1100, 100))
        rows = list(range(0, 361, 40))
        image = np.full((360, 1000, 3), 255, dtype=np.uint8)
        for boundary in columns:
            cv2.line(image, (min(boundary, 999), 0), (min(boundary, 999), 359), (0, 0, 0), 2)
        grid = [["基础信息", "", "目标与测量", "", "过程记录", "", "", "", "质量判定", ""], [f"列{x}" for x in range(10)], *[[str(x) for x in range(10)] for _ in range(7)]]
        original = [row[:] for row in grid]

        spans = ocr_backend._recover_certified_physical_group_header_spans(
            image,
            columns,
            rows,
            grid,
            [[0.99 if value else 0.0 for value in row] for row in grid],
            [],
            [],
        )

        self.assertEqual(spans, [])
        self.assertEqual(grid, original)

    def test_certified_physical_group_headers_anchor_paired_nested_row(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 1600, 100))
        rows = list(range(0, 401, 40))
        image = np.full((400, 1500, 3), 255, dtype=np.uint8)
        for boundary in rows:
            cv2.line(image, (0, boundary), (1499, boundary), (0, 0, 0), 2)
        for boundary in (0, 200, 700, 1000, 1499):
            cv2.line(image, (boundary, 0), (boundary, 40), (0, 0, 0), 2)
        for boundary in range(0, 1501, 200):
            cv2.line(image, (min(boundary, 1499), 40), (min(boundary, 1499), 80), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (min(boundary, 1499), 80), (min(boundary, 1499), 399), (0, 0, 0), 2)
        grid = [
            ["基础信息", "", "目标与测量", "", "", "", "", "过程记录", "", "", "质量判定", "", "", "", ""],
            ["", "本期", "下限", "", "", "上限", "实际", "", "上限", "", "上限", "", "累计", "", "累计"],
            [f"列{column}" for column in range(15)],
            *[[str(column) for column in range(15)] for _ in range(7)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        spans = ocr_backend._recover_certified_physical_group_header_spans(
            image, columns, rows, grid, confidence, [], []
        )

        self.assertEqual(
            grid[1],
            ["本期", "", "下限", "", "上限", "", "实际", "", "上限", "", "上限", "", "累计", "", "累计"],
        )
        self.assertIn(
            {"row": 1, "column": 0, "row_span": 1, "column_span": 2, "role": "group_header"},
            spans,
        )
        self.assertIn(
            {"row": 1, "column": 12, "row_span": 1, "column_span": 2, "role": "group_header"},
            spans,
        )

    def test_certified_physical_group_headers_recover_dense_fragmented_nested_row(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 1400, 100))
        rows = list(range(0, 401, 40))
        image = np.full((400, 1300, 3), 255, dtype=np.uint8)
        for boundary in rows:
            cv2.line(image, (0, boundary), (1299, boundary), (0, 0, 0), 2)
        for boundary in range(0, 1301, 200):
            cv2.line(image, (min(boundary, 1299), 40), (min(boundary, 1299), 80), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (min(boundary, 1299), 80), (min(boundary, 1299), 399), (0, 0, 0), 2)
        grid = [
            ["基础信息", "", "", "", "目标与测量", "", "过程记录", "", "", "", "", "质量判定", ""],
            ["本", "期", "", "下限", "下限", "限", "累", "计", "累", "计", "计", "划", "本期"],
            [f"列{column}" for column in range(13)],
            *[[str(column) for column in range(13)] for _ in range(7)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        existing = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 4, "role": "group_header"},
            {"row": 0, "column": 4, "row_span": 1, "column_span": 2, "role": "group_header"},
            {"row": 0, "column": 6, "row_span": 1, "column_span": 5, "role": "group_header"},
            {"row": 0, "column": 11, "row_span": 1, "column_span": 2, "role": "group_header"},
        ]

        spans = ocr_backend._recover_certified_physical_group_header_spans(
            image, columns, rows, grid, confidence, [], existing
        )

        self.assertEqual(
            grid[1],
            ["本期", "", "下限", "", "下限", "", "累计", "", "累计", "", "计划", "", "本期"],
        )
        self.assertEqual(
            [
                (span["row"], span["column"], span["column_span"])
                for span in spans
                if span["row"] == 1
            ],
            [(1, 0, 2), (1, 2, 2), (1, 4, 2), (1, 6, 2), (1, 8, 2), (1, 10, 2)],
        )

    def test_certified_four_level_headers_recover_title_groups_and_nested_spans(self):
        ocr_backend._load_runtime()
        columns = list(range(0, 1000, 100))
        rows = list(range(0, 361, 40))
        image = np.full((360, 900, 3), 255, dtype=np.uint8)
        for boundary in rows:
            cv2.line(image, (0, min(boundary, 359)), (899, min(boundary, 359)), (0, 0, 0), 2)
        for boundary in (0, 899):
            cv2.line(image, (boundary, 0), (boundary, 80), (0, 0, 0), 2)
        for boundary in (0, 200, 400, 600, 800, 899):
            cv2.line(image, (boundary, 80), (boundary, 120), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (min(boundary, 899), 120), (min(boundary, 899), 359), (0, 0, 0), 2)
        grid = [
            ["", "", "无框浅线登记", "-2026-02批", "次", "", "", "", ""],
            ["", "基础信息", "", "", "", "目标与测量", "", "", "过程记录"],
            ["", "计划", "计划", "", "累计", "", "计划", "", "实际"],
            ["登记号", "日期", "名称", "类别", "来源", "数量", "单位", "状态", "经办人"],
            *[[str(column) for column in range(9)] for _ in range(5)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("无框", 150.0, 20.0, 0.99),
            ("浅线登记", 350.0, 20.0, 0.99),
            ("-2026-02批", 550.0, 20.0, 0.99),
            ("次", 650.0, 20.0, 0.99),
            ("基础信息", 200.0, 60.0, 0.99),
            ("目标与测量", 600.0, 60.0, 0.99),
            ("过程记录", 850.0, 60.0, 0.99),
        ]

        spans = ocr_backend._recover_certified_page_header_spans(
            image, columns, rows, grid, confidence, evidence
        )
        spans = ocr_backend._recover_certified_physical_group_header_spans(
            image, columns, rows, grid, confidence, evidence, spans
        )

        self.assertEqual(grid[0], ["无框浅线登记-2026-02批次"] + [""] * 8)
        self.assertEqual(
            grid[1],
            ["基础信息", "", "", "", "目标与测量", "", "", "", "过程记录"],
        )
        self.assertEqual(
            grid[2],
            ["计划", "", "计划", "", "累计", "", "计划", "", "实际"],
        )
        self.assertEqual(
            [(span["row"], span["column"], span["column_span"]) for span in spans],
            [
                (0, 0, 9),
                (1, 0, 4),
                (1, 4, 4),
                (2, 0, 2),
                (2, 2, 2),
                (2, 4, 2),
                (2, 6, 2),
            ],
        )

    def test_review_only_spatial_outliers_select_decimal_and_identifier_losses(self):
        ocr_backend._load_runtime()
        grid = [
            ["产品编码", "销售单价", "订单数量"],
            *[
                [
                    "AP-X-84" if index == 5 else f"AP-AX{index:03d}-{index + 10:02d}",
                    "3903" if index == 5 else f"{300 + index}.{index:02d}",
                    str(1000 + index),
                ]
                for index in range(10)
            ],
        ]

        selected = ocr_backend._review_only_spatial_outlier_cells(grid)

        self.assertEqual(selected, {(6, 0), (6, 1)})

        class ConsensusEngine:
            @staticmethod
            def text_rec(_payload):
                return SimpleNamespace(
                    txts=["AP-AX005-15", "AP-AX005-15", "305.05", "305.05"],
                    scores=[0.99] * 4,
                )

            @staticmethod
            def server_text_rec(_payload):
                return SimpleNamespace(
                    txts=["AP-AX005-15", "AP-AX005-15", "305.05", "305.05"],
                    scores=[0.99] * 4,
                )

        confidence = [[0.77 if value else 0.0 for value in row] for row in grid]
        scores = ocr_backend._repair_review_only_spatial_outliers(
            np.full((330, 300, 3), 255, dtype=np.uint8),
            grid,
            confidence,
            [0, 100, 200, 300],
            list(range(0, 331, 30)),
            selected,
            ConsensusEngine(),
        )

        self.assertEqual(grid[6][:2], ["AP-AX005-15", "305.05"])
        self.assertEqual(scores, [0.99] * 8)

    def test_review_only_spatial_outliers_restore_unanimous_blank_numeric(self):
        ocr_backend._load_runtime()
        grid = [["序号", "测量值"]] + [
            [str(index), "" if index == 5 else f"{300 + index}.{index:03d}"]
            for index in range(1, 11)
        ]
        confidence = [[0.77 if value else 0.0 for value in row] for row in grid]

        class ConsensusEngine:
            @staticmethod
            def text_rec(_payload):
                return SimpleNamespace(txts=["305.005", "305.005"], scores=[0.99, 0.99])

            @staticmethod
            def server_text_rec(_payload):
                return SimpleNamespace(txts=["305.005", "305.005"], scores=[0.99, 0.99])

        scores = ocr_backend._repair_review_only_spatial_outliers(
            np.full((330, 200, 3), 255, dtype=np.uint8),
            grid,
            confidence,
            [0, 100, 200],
            list(range(0, 331, 30)),
            {(5, 1)},
            ConsensusEngine(),
        )

        self.assertEqual(grid[5][1], "305.005")
        self.assertEqual(confidence[5][1], 0.77)
        self.assertEqual(scores, [0.99] * 4)

    def test_review_only_spatial_outliers_accept_cross_model_blank_text_at_085(self):
        ocr_backend._load_runtime()
        grid = [["序号", "位置"]] + [
            [str(index), "" if index == 5 else ("A区" if index % 2 else "B区")]
            for index in range(1, 11)
        ]
        confidence = [[0.77 if value else 0.0 for value in row] for row in grid]

        class ConsensusEngine:
            @staticmethod
            def text_rec(_payload):
                return SimpleNamespace(txts=["B区", "B区"], scores=[0.87, 0.88])

            @staticmethod
            def server_text_rec(_payload):
                return SimpleNamespace(txts=["B区", "B区"], scores=[0.86, 0.85])

        ocr_backend._repair_review_only_spatial_outliers(
            np.full((330, 200, 3), 255, dtype=np.uint8),
            grid,
            confidence,
            [0, 100, 200],
            list(range(0, 331, 30)),
            {(5, 1)},
            ConsensusEngine(),
        )

        self.assertEqual(grid[5][1], "B区")
        self.assertEqual(confidence[5][1], 0.77)

    def test_visible_leading_b_before_cjk_requires_conservative_lobes(self):
        ocr_backend._load_runtime()
        b_crop = np.full((30, 100, 3), 255, dtype=np.uint8)
        a_crop = b_crop.copy()
        cv2.putText(b_crop, "B", (32, 24), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 2)
        cv2.putText(a_crop, "A", (32, 24), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 0), 2)

        self.assertTrue(ocr_backend._has_visible_leading_b_before_cjk(b_crop))
        self.assertFalse(ocr_backend._has_visible_leading_b_before_cjk(a_crop))

    def test_review_only_spatial_outliers_restore_physically_proven_b_prefix(self):
        ocr_backend._load_runtime()
        grid = [["序号", "位置"]] + [
            [str(index), "区" if index == 5 else "B区"]
            for index in range(1, 11)
        ]
        confidence = [[0.77 for _ in row] for row in grid]

        class IncompleteEngine:
            @staticmethod
            def text_rec(_payload):
                return SimpleNamespace(txts=["区", "区"], scores=[0.78, 0.79])

            @staticmethod
            def server_text_rec(_payload):
                return SimpleNamespace(txts=["区", "区"], scores=[0.64, 0.65])

        with patch.object(
            ocr_backend,
            "_has_visible_leading_b_before_cjk",
            return_value=True,
        ):
            ocr_backend._repair_review_only_spatial_outliers(
                np.full((330, 200, 3), 255, dtype=np.uint8),
                grid,
                confidence,
                [0, 100, 200],
                list(range(0, 331, 30)),
                {(5, 1)},
                IncompleteEngine(),
            )

        self.assertEqual(grid[5][1], "B区")
        self.assertEqual(confidence[5][1], 0.77)

    def test_spatial_row_ruler_repair_uses_only_right_cell_consensus(self):
        ocr_backend._load_runtime()
        image = np.full((120, 200, 3), 255, dtype=np.uint8)
        cv2.putText(
            image,
            "9",
            (65, 112),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.7,
            (0, 0, 0),
            2,
        )
        grid = [
            ["基础信息", ""],
            ["本期", ""],
            ["序号", "名称"],
            ["4", "设备A"],
        ]
        confidence = [[0.77 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("1", 10.0, 15.0, 0.99),
            ("2", 10.0, 45.0, 0.99),
            ("3", 10.0, 75.0, 0.99),
            ("4", 10.0, 105.0, 0.99),
        ]

        class ConsensusEngine:
            @staticmethod
            def text_rec(_payload):
                return SimpleNamespace(txts=["9", "9"], scores=[0.99, 0.99])

            @staticmethod
            def server_text_rec(_payload):
                return SimpleNamespace(txts=["9", "9"], scores=[0.99, 0.99])

        scores = ocr_backend._repair_spatial_spreadsheet_row_ruler_first_column(
            image,
            grid,
            confidence,
            [0, 100, 200],
            [0, 30, 60, 90, 120],
            evidence,
            ConsensusEngine(),
        )

        self.assertEqual(grid[3][0], "9")
        self.assertEqual(confidence[3][0], 0.77)
        self.assertEqual(scores, [0.99, 0.99, 0.99, 0.99])

    def test_review_only_visible_units_remain_yellow_and_need_pixel_evidence(self):
        ocr_backend._load_runtime()
        grid = [["单位"], [""], [""]]
        confidence = [[0.77], [0.0], [0.0]]

        class UnitEngine:
            @staticmethod
            def text_rec(_payload):
                return SimpleNamespace(
                    txts=["C", "C", "v", "v"],
                    scores=[0.75, 0.70, 0.70, 0.70],
                )

            @staticmethod
            def server_text_rec(_payload):
                return SimpleNamespace(
                    txts=["C", "C", "V", "VP"],
                    scores=[0.75, 0.70, 0.58, 0.45],
                )

        with (
            patch.object(
                ocr_backend,
                "_has_visible_celsius_degree_ring",
                side_effect=[True, False],
            ),
            patch.object(
                ocr_backend,
                "_has_visible_uppercase_v_glyph",
                side_effect=[True, True],
            ),
        ):
            scores = ocr_backend._recover_review_only_visible_units(
                np.full((90, 100, 3), 255, dtype=np.uint8),
                grid,
                confidence,
                [0, 100],
                [0, 30, 60, 90],
                [("V", 50.0, 75.0, 0.50)],
                UnitEngine(),
            )

        self.assertEqual(grid, [["单位"], ["℃"], ["V"]])
        self.assertEqual(confidence[1:], [[0.77], [0.77]])
        self.assertEqual(len(scores), 8)

    def test_visible_uppercase_l_after_hyphen_requires_lower_foot(self):
        ocr_backend._load_runtime()
        uppercase_l = np.full((30, 100, 3), 255, dtype=np.uint8)
        cv2.putText(
            uppercase_l, "AB", (3, 23), cv2.FONT_HERSHEY_SIMPLEX, 0.48, (0, 0, 0), 1
        )
        cv2.line(uppercase_l, (30, 15), (35, 15), (0, 0, 0), 2)
        cv2.line(uppercase_l, (41, 7), (41, 20), (0, 0, 0), 2)
        cv2.line(uppercase_l, (45, 22), (48, 22), (0, 0, 0), 1)
        cv2.putText(
            uppercase_l, "08", (52, 23), cv2.FONT_HERSHEY_SIMPLEX, 0.48, (0, 0, 0), 1
        )
        digit_one = uppercase_l.copy()
        digit_one[:, 39:49] = 255
        cv2.line(digit_one, (43, 7), (43, 22), (0, 0, 0), 2)

        self.assertTrue(
            ocr_backend._has_visible_uppercase_l_after_hyphen(uppercase_l)
        )
        self.assertFalse(ocr_backend._has_visible_uppercase_l_after_hyphen(digit_one))

    def test_repeated_visible_uppercase_l_recovery_remains_yellow(self):
        grid = [["名称"], *[["批次-L08"] for _ in range(6)], ["批次-108"], ["批次-1.08"]]
        confidence = [[0.95] for _ in grid]
        rows = list(range(0, len(grid) * 30 + 1, 30))

        with patch.object(
            ocr_backend,
            "_has_visible_uppercase_l_after_hyphen",
            side_effect=[True, True],
        ):
            recovered = ocr_backend._recover_repeated_visible_uppercase_l_tokens(
                np.full((len(grid) * 30, 120, 3), 255, dtype=np.uint8),
                grid,
                confidence,
                [0, 120],
                rows,
            )

        self.assertEqual(recovered, {(7, 0), (8, 0)})
        self.assertEqual(grid[7:], [["批次-L08"], ["批次-L08"]])
        self.assertEqual(confidence[7:], [[0.77], [0.77]])

    def test_repeated_visible_uppercase_l_recovery_requires_repetition_and_pixels(self):
        grid = [["名称"], ["批次-L08"], ["批次-108"]]
        confidence = [[0.95] for _ in grid]

        with patch.object(
            ocr_backend, "_has_visible_uppercase_l_after_hyphen", return_value=True
        ):
            recovered = ocr_backend._recover_repeated_visible_uppercase_l_tokens(
                np.full((90, 120, 3), 255, dtype=np.uint8),
                grid,
                confidence,
                [0, 120],
                [0, 30, 60, 90],
            )

        self.assertEqual(recovered, set())
        self.assertEqual(grid[2][0], "批次-108")

    def test_datetime_spacing_is_limited_to_valid_datetime_columns(self):
        grid = [
            ["序号", "采集时间", "备注"],
            ["1", "2026-05-2208:16", "2026-05-2208:16"],
            ["2", "2026-13-4008:16", "保持原文"],
        ]
        confidence = [[0.95] * 3 for _ in grid]

        restored = ocr_backend._normalize_visible_datetime_spacing(grid, confidence)

        self.assertEqual(restored, {(1, 1)})
        self.assertEqual(grid[1][1], "2026-05-22 08:16")
        self.assertEqual(grid[1][2], "2026-05-2208:16")
        self.assertEqual(grid[2][1], "2026-13-4008:16")
        self.assertEqual(confidence[1][1], 0.77)

        four_level = [
            ["月报", "", ""],
            ["基础信息", "", ""],
            ["本期", "计划", "累计"],
            ["序号", "更新时间", "状态"],
            ["1", "2026-01-0516:38", "完成"],
        ]
        four_level_confidence = [[0.95] * 3 for _ in four_level]
        self.assertEqual(
            ocr_backend._normalize_visible_datetime_spacing(
                four_level,
                four_level_confidence,
            ),
            {(4, 1)},
        )
        self.assertEqual(four_level[4][1], "2026-01-05 16:38")

    def test_overlapping_decimal_fragments_require_a_dominant_column_format(self):
        grid = [
            ["频率(MHz)", "备注"],
            ["581.450", "正常"],
            ["574.414", "正常"],
            ["588.415", "正常"],
            ["563.513", "正常"],
            ["568.31 317", "正常"],
            ["587.650", "正常"],
            ["571.175", "正常"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        repaired = ocr_backend._repair_overlapping_decimal_fragments(
            grid,
            confidence,
        )

        self.assertEqual(repaired, {(5, 0)})
        self.assertEqual(grid[5][0], "568.317")
        self.assertEqual(confidence[5][0], 0.77)
        sparse_grid = [["数值"], ["12.340"], ["12.34 345"], ["13.210"]]
        sparse_confidence = [[0.99] for _ in sparse_grid]
        self.assertEqual(
            ocr_backend._repair_overlapping_decimal_fragments(
                sparse_grid,
                sparse_confidence,
            ),
            set(),
        )
        self.assertEqual(sparse_grid[2][0], "12.34 345")
        trailing_grid = [["金额"]] + [
            ["19491. 1.39" if index == 7 else f"{19000 + index}.39"]
            for index in range(1, 9)
        ]
        trailing_confidence = [[0.99] for _ in trailing_grid]
        self.assertEqual(
            ocr_backend._repair_overlapping_decimal_fragments(
                trailing_grid,
                trailing_confidence,
            ),
            {(7, 0)},
        )
        self.assertEqual(trailing_grid[7][0], "19491.39")

    def test_ipv4_spacing_normalization_requires_valid_ip_column(self):
        grid = [
            ["设备", "IP地址", "备注"],
            ["A", "192.168.11.2 0", "192.168.11.2 0"],
            ["B", "999.168.1. 1", "保持原文"],
        ]
        confidence = [[0.95] * 3 for _ in grid]

        restored = ocr_backend._normalize_visible_ipv4_spacing(grid, confidence)

        self.assertEqual(restored, {(1, 1)})
        self.assertEqual(grid[1][1], "192.168.11.20")
        self.assertEqual(grid[1][2], "192.168.11.2 0")
        self.assertEqual(grid[2][1], "999.168.1. 1")
        self.assertEqual(confidence[1][1], 0.77)

    def test_sparse_group_headers_preserve_leading_two_row_headers(self):
        grid = [
            ["2026年第三季度设备运行统计表", "", "", "", "", "", "", "", "", ""],
            ["序号", "设备名称", "", "运行数据", "", "", "质量指标", "", "维护信息", ""],
            ["", "", "运行时长(h)", "告警次数", "停机时长(h)", "合格率", "平均温度(℃)", "信噪比(dB)", "上次维护", "负责人"],
            ["1", "测试设备-01", "154.5", "1", "0.35", "99.92%", "31.6", "22.6", "2026-07-02", "钱工"],
            ["2", "测试设备-02", "161", "2", "0.7", "99.89%", "32", "22.3", "2026-07-03", "孙工"],
            ["3", "测试设备-03", "167.5", "3", "1.05", "99.86%", "32.4", "22.1", "2026-07-04", "赵工"],
            ["4", "测试设备-04", "174", "0", "1.4", "99.83%", "32.8", "21.8", "2026-07-05", "钱工"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        spans = [{"row": 0, "column": 0, "row_span": 1, "column_span": 10, "role": "title"}]

        adjusted, _, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            grid, confidence, spans
        )

        self.assertEqual(
            adjusted[1],
            ["序号", "设备名称", "运行数据", "", "", "质量指标", "", "", "维护信息", ""],
        )
        for expected in (
            {"row": 1, "column": 0, "row_span": 2, "column_span": 1, "role": "row_header"},
            {"row": 1, "column": 1, "row_span": 2, "column_span": 1, "role": "row_header"},
            {"row": 1, "column": 2, "row_span": 1, "column_span": 3, "role": "group_header"},
            {"row": 1, "column": 5, "row_span": 1, "column_span": 3, "role": "group_header"},
            {"row": 1, "column": 8, "row_span": 1, "column_span": 2, "role": "group_header"},
        ):
            self.assertIn(expected, adjusted_spans)

    def test_sparse_multilevel_header_recovers_known_two_row_labels(self):
        image = np.full((240, 1000, 3), 255, dtype=np.uint8)
        columns = list(range(0, 1001, 100))
        rows = list(range(0, 241, 30))
        grid = [
            ["标题", "", "", "", "", "", "", "", "", ""],
            ["", "", "", "运行数据", "", "", "质量指标", "", "维护信", ""],
            ["", "", "运行时长", "告警次数", "停机时长", "合格率", "平均温度", "信噪比", "上次维护", "负责人"],
        ] + [[str(index)] * 10 for index in range(1, 6)]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        class HeaderEngine:
            def __call__(self, _crop):
                centers = ((50, 30), (150, 30), (350, 15), (650, 15), (850, 15))
                boxes = [
                    np.asarray(
                        [[x - 20, y - 8], [x + 20, y - 8], [x + 20, y + 8], [x - 20, y + 8]],
                        dtype=np.float32,
                    )
                    for x, y in centers
                ]
                return SimpleNamespace(
                    boxes=boxes,
                    txts=["序号", "设备名称", "运行数据", "质量指标", "维护信息"],
                    scores=[0.998, 0.997, 0.999, 0.999, 0.996],
                )

        scores = ocr_backend._recover_sparse_multilevel_row_headers(
            image, columns, rows, grid, confidence, HeaderEngine()
        )

        self.assertEqual(grid[1][:2], ["序号", "设备名称"])
        self.assertEqual(confidence[1][:2], [0.998, 0.997])
        self.assertEqual(grid[1][8], "维护信息")
        self.assertEqual(scores, [0.998, 0.997, 0.999, 0.999, 0.996])

    def test_sparse_group_header_recovers_a_missing_known_measurement_group(self):
        ocr_backend._load_runtime()
        image = np.full((150, 350, 3), 245, dtype=np.uint8)
        columns = list(range(0, 351, 50))
        rows = list(range(0, 151, 10))
        grid = [
            ["生产日报", "", "", "", "", "", ""],
            ["基础信息", "", "", "", "", "", ""],
            ["日期", "班次", "生产线", "计划数量", "完成数量", "不良数量", "完成率"],
        ] + [[str(index)] * 7 for index in range(1, 13)]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]

        class HeaderEngine:
            def __call__(self, _crop):
                return SimpleNamespace(boxes=[], txts=[], scores=[])

            def text_rec(self, request):
                return SimpleNamespace(
                    txts=["基础信息测觉与判定"] * len(request.img),
                    scores=[0.895] * len(request.img),
                )

            def server_text_rec(self, request):
                return SimpleNamespace(
                    txts=["基础信息测与判定"] * len(request.img),
                    scores=[0.91] * len(request.img),
                )

        scores = ocr_backend._recover_sparse_multilevel_row_headers(
            image, columns, rows, grid, confidence, HeaderEngine()
        )

        self.assertEqual(grid[1][3], "测量与判定")
        self.assertEqual(confidence[1][3], 0.77)
        self.assertIn(0.895, scores)

    def test_sparse_group_header_uses_existing_second_anchor_as_boundary(self):
        ocr_backend._load_runtime()
        image = np.full((150, 350, 3), 245, dtype=np.uint8)
        columns = list(range(0, 351, 50))
        rows = list(range(0, 151, 10))
        grid = [
            ["综合测试数据表", "", "", "", "", "", ""],
            ["基", "", "", "测量与判定", "", "", ""],
            ["项目", "类别", "数值1", "数值2", "数值3", "单位", "备注"],
        ] + [[str(index)] * 7 for index in range(1, 13)]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]

        class HeaderEngine:
            def text_rec(self, request):
                return SimpleNamespace(
                    txts=["基础信息测量与判定"] * len(request.img),
                    scores=[0.92] * len(request.img),
                )

            def server_text_rec(self, request):
                return SimpleNamespace(
                    txts=["基础信息测与判定"] * len(request.img),
                    scores=[0.94] * len(request.img),
                )

        ocr_backend._recover_sparse_multilevel_row_headers(
            image, columns, rows, grid, confidence, HeaderEngine()
        )

        self.assertEqual(grid[1], ["基础信息", "", "", "测量与判定", "", "", ""])
        self.assertEqual(confidence[1][0], 0.86)

    def test_sparse_group_header_skips_dual_recognition_when_existing_text_cannot_match(self):
        image = np.full((150, 400, 3), 245, dtype=np.uint8)
        columns = list(range(0, 401, 50))
        rows = list(range(0, 151, 10))
        grid = [
            ["施工现场安全检查表", "", "", "", "", "", "", ""],
            ["项目：客户服务", "", "", "", "", "", "", ""],
            ["编号", "检查区域", "检查项目", "风险等级", "责任单位", "整改期限", "复查结果", "检查人"],
        ] + [[str(index)] * 8 for index in range(1, 13)]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]

        def must_not_recognize(_request):
            self.fail("已有标题不可能满足恢复条件时不得调用双视图识别")

        engine = SimpleNamespace(
            text_rec=must_not_recognize,
            server_text_rec=must_not_recognize,
        )
        page_output = SimpleNamespace(boxes=[], txts=[], scores=[])

        scores = ocr_backend._recover_sparse_multilevel_row_headers(
            image,
            columns,
            rows,
            grid,
            confidence,
            engine,
            page_output=page_output,
        )

        self.assertEqual(scores, [])
        self.assertEqual(grid[1][0], "项目：客户服务")

    def test_sparse_group_header_repairs_one_unique_dictionary_near_miss(self):
        image = np.full((150, 350, 3), 245, dtype=np.uint8)
        grid = [
            ["生产日报", "", "", "", "", "", ""],
            ["基础信息", "", "", "测呈与判定", "", "", ""],
            ["日期", "班次", "生产线", "计划数量", "完成数量", "不良数量", "完成率"],
        ] + [[str(index)] * 7 for index in range(1, 13)]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]

        ocr_backend._recover_sparse_multilevel_row_headers(
            image,
            list(range(0, 351, 50)),
            list(range(0, 151, 10)),
            grid,
            confidence,
            SimpleNamespace(),
        )

        self.assertEqual(grid[1][3], "测量与判定")
        self.assertEqual(confidence[1][3], 0.77)

    def test_periodic_category_repair_fixes_one_confusable_outlier_per_column(self):
        grid = [
            ["日期", "班次", "生产线"],
            *[
                [f"2026-08-{index + 1:02d}", shift, line]
                for index, (shift, line) in enumerate(
                    [
                        ("夜班", "B线"), ("日班", "C线"),
                        ("夜班", "D线"), ("白班", "A线"),
                        ("夜班", "8线"), ("白班", "C线"),
                        ("夜班", "D线"), ("白班", "A线"),
                        ("夜班", "B线"), ("白班", "C线"),
                        ("夜班", "D线"), ("白班", "A线"),
                    ]
                )
            ],
        ]
        confidence = [[0.95] * 3 for _ in grid]

        repaired = ocr_backend._repair_periodic_categorical_confusions(
            grid, confidence
        )

        self.assertEqual(repaired, [(2, 1, "白班"), (5, 2, "B线")])
        self.assertEqual(grid[2][1], "白班")
        self.assertEqual(grid[5][2], "B线")
        self.assertEqual(confidence[2][1], 0.77)
        self.assertEqual(confidence[5][2], 0.77)

    def test_periodic_category_repair_restores_detector_miss(self):
        grid = [
            ["编号", "类别"],
            *[
                [str(index + 1), value]
                for index, value in enumerate(
                    ["类别1", "类别2", "类别3", "类别4", "", "类别2", "类别3", "类别4", "类别1", "类别2"]
                )
            ],
        ]
        confidence = [[0.99, 0.99], *[[0.99, 0.99] for _ in range(10)]]
        confidence[5][1] = -1.0

        repaired = ocr_backend._repair_periodic_categorical_confusions(
            grid,
            confidence,
        )

        self.assertEqual(repaired, [(5, 1, "类别1")])
        self.assertEqual(grid[5][1], "类别1")
        self.assertEqual(confidence[5][1], 0.77)

    def test_periodic_responsible_column_uses_repeated_names_and_one_strong_extension(self):
        names = [
            "陈晨", "王强", "赵敏", "李", "陈具", "王强",
            "赵敏", "", "陈晨", "王强", "赵敏", "李",
        ]
        grid = [["任务编号", "负责人"]] + [
            [f"TASK-{index + 1:03d}", name]
            for index, name in enumerate(names)
        ]
        confidence = [[0.95, 0.95] for _ in grid]
        candidates = {
            (4, 1): [("v5-server", "李娜", 0.885)],
            (12, 1): [("v4-server", "李题", 0.84)],
        }

        repaired = ocr_backend._repair_periodic_responsible_labels(
            grid,
            confidence,
            candidates,
        )

        self.assertEqual(repaired, 4)
        self.assertEqual(
            [row[1] for row in grid[1:]],
            ["陈晨", "王强", "赵敏", "李娜"] * 3,
        )
        self.assertEqual(confidence[4][1], 0.77)

    def test_periodic_category_repair_fills_one_blank_status(self):
        grid = [["任务编号", "当前状态"]] + [
            [f"TASK-{index + 1:03d}", status]
            for index, status in enumerate(
                [
                    "已完成", "", "进行中", "已完成", "等待", "进行中",
                    "已完成", "等待", "进行中", "已完成", "等待", "进行中",
                ]
            )
        ]
        confidence = [[0.95, 0.95] for _ in grid]

        repaired = ocr_backend._repair_periodic_categorical_confusions(
            grid, confidence
        )

        self.assertEqual(repaired, [(2, 1, "等待")])
        self.assertEqual(grid[2][1], "等待")
        self.assertEqual(confidence[2][1], 0.77)

    def test_periodic_category_repair_fills_two_supported_line_blanks(self):
        values = ["B线", "C线", "D线", "A线"] * 3
        values[6] = ""
        values[7] = ""
        grid = [["任务编号", "生产线"]] + [
            [f"TASK-{index + 1:03d}", value]
            for index, value in enumerate(values)
        ]
        confidence = [[0.95, 0.95] for _ in grid]

        repaired = ocr_backend._repair_periodic_categorical_confusions(
            grid, confidence
        )

        self.assertEqual(repaired, [(7, 1, "D线"), (8, 1, "A线")])
        self.assertEqual([row[1] for row in grid[1:]], ["B线", "C线", "D线", "A线"] * 3)

    def test_sparse_measurement_group_uses_detail_header_boundary(self):
        grid = [
            ["生产日报", "", "", "", "", "", ""],
            ["基础信息", "", "", "测量与判定", "", "", ""],
            ["日期", "班次", "生产线", "计划数量", "完成数量", "不良数量", "完成率"],
            ["2026-08-02", "夜班", "B线", "280", "277", "3", "98.2%"],
        ]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]
        spans = [{"row": 0, "column": 0, "row_span": 1, "column_span": 7, "role": "title"}]

        adjusted, _, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            grid, confidence, spans
        )

        self.assertEqual(adjusted[1], ["基础信息", "", "", "测量与判定", "", "", ""])
        self.assertIn(
            {"row": 1, "column": 3, "row_span": 1, "column_span": 4, "role": "group_header"},
            adjusted_spans,
        )

    def test_measurement_group_alignment_survives_a_minor_first_group_ocr_error(self):
        grid = [
            ["综合测试数据表", "", "", "", "", "", ""],
            ["基础信自", "", "", "测量与判定", "", "", ""],
            ["项目", "类别", "数值1", "数值2", "数值3", "单位", "备注"],
            ["项目01", "类别2", "15", "0.375", "7e-05", "mA", "—"],
        ]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]
        spans = [{"row": 0, "column": 0, "row_span": 1, "column_span": 7, "role": "title"}]

        adjusted, _, adjusted_spans = ocr_backend._infer_sparse_group_header_spans(
            grid, confidence, spans
        )

        self.assertEqual(adjusted[1][3], "测量与判定")
        self.assertIn(
            {"row": 1, "column": 3, "row_span": 1, "column_span": 4, "role": "group_header"},
            adjusted_spans,
        )

    def test_concatenated_headers_use_a_single_blank_immediately_to_the_left(self):
        grid = [
            ["日期", "班次", "生产线", "计划数量", "", "完成数量不良数量", "完成率"],
            ["2026-08-02", "夜班", "B线", "280", "277", "3", "98.2%"],
        ]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]

        split, split_confidence, _ = ocr_backend._split_collapsed_header_data_row(
            grid, confidence, []
        )

        self.assertEqual(
            split[0],
            ["日期", "班次", "生产线", "计划数量", "完成数量", "不良数量", "完成率"],
        )
        self.assertEqual(split_confidence[0][4:6], [0.95, 0.95])

    def test_concatenated_header_drops_only_the_adjacent_duplicate_suffix(self):
        grid = [
            ["编号", "设备名称", "型号", "频率", "功率", "状态备注", "备注"],
            ["A001", "信号发生器", "SG-2200", "515.472", "-20", "复核", "校准完成"],
        ]
        confidence = [[0.95] * 7 for _ in grid]

        split, _, _ = ocr_backend._split_collapsed_header_data_row(grid, confidence, [])

        self.assertEqual(split[0][5:], ["状态", "备注"])

    def test_concatenated_header_drops_only_the_adjacent_duplicate_prefix(self):
        grid = [
            ["日期", "班次", "生产线", "计划数量", "完成数量", "完成数量不良数量", "完成率"],
            ["2026-08-02", "夜班", "B线", "280", "277", "3", "98.2%"],
        ]
        confidence = [[0.95] * 7 for _ in grid]

        split, _, _ = ocr_backend._split_collapsed_header_data_row(grid, confidence, [])

        self.assertEqual(split[0][4:6], ["完成数量", "不良数量"])

    def test_adjacent_duplicate_header_is_removed_even_when_merged_token_is_blurred(self):
        grid = [
            ["编号", "设备名称", "型号", "频率", "功率", "状态备注", "备注"],
            ["A001", "信号发生器", "SG-2200", "515.472", "-20", "复核", "校准完成"],
        ]
        confidence = [[0.95] * 7 for _ in grid]
        confidence[0][5] = 0.71

        split, _, _ = ocr_backend._split_collapsed_header_data_row(grid, confidence, [])

        self.assertEqual(split[0][5:], ["状态", "备注"])

    def test_low_confidence_exact_header_pair_uses_immediate_blank_cell(self):
        grid = [
            ["工单号 报修位置", "", "维修项目", "维修人 状态", ""],
            ["WX-001", "A区", "检修", "张伟", "完成"],
        ]
        confidence = [[0.77, 0.0, 0.95, 0.77, 0.0], [0.95] * 5]

        split, split_confidence, _ = ocr_backend._split_collapsed_header_data_row(
            grid, confidence, []
        )

        self.assertEqual(split[0], ["工单号", "报修位置", "维修项目", "维修人", "状态"])
        self.assertEqual(split_confidence[0], [0.77, 0.77, 0.95, 0.77, 0.77])

    def test_numbered_header_suffixes_are_not_split_into_a_data_row(self):
        grid = [
            ["项目", "编号", "确认", "备注1", "备注2", "备注3"],
            ["", "", "", "", "", ""],
            ["", "", "", "", "", ""],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        original = [list(row) for row in grid]

        observed, _, _ = ocr_backend._split_collapsed_header_data_row(
            grid, confidence, []
        )

        self.assertEqual(observed, original)

    def test_adjacent_duplicate_header_uses_visible_suffix_without_dictionary_pair(self):
        grid = [
            ["编号", "姓名", "组别", "项目", "联系电话报名状态", "报名状态"],
            ["1", "赵敏", "华东库", "现场复核", "147****2538", "正常"],
        ]
        confidence = [[0.95] * 6 for _ in grid]

        split, split_confidence, _ = ocr_backend._split_collapsed_header_data_row(
            grid,
            confidence,
            [],
        )

        self.assertEqual(split[0][4:], ["联系电话", "报名状态"])
        self.assertEqual(split_confidence[0][4], 0.77)

    def test_header_checkmark_is_removed_only_from_known_non_boolean_label(self):
        grid = [
            ["编号", "姓名", "组别", "☑项目", "联系电话", "报名状态"],
            ["1", "赵敏", "华东库", "现场复核", "147****2538", "正常"],
        ]
        confidence = [[0.95] * 6 for _ in grid]

        split, split_confidence, _ = ocr_backend._split_collapsed_header_data_row(
            grid,
            confidence,
            [],
        )

        self.assertEqual(split[0][3], "项目")
        self.assertEqual(split_confidence[0][3], 0.77)

    def test_late_dense_header_cleanup_handles_mark_and_adjacent_suffix(self):
        grid = [
            ["赛事报名信息表", "", "", "", "", ""],
            ["单位：服务中心", "", "", "批次：B32", "", ""],
            ["编号", "姓名", "组别", "☑项目", "联系电话报名状态", "报名状态"],
            ["1", "赵敏", "华东库", "现场复核", "147****2538", "正常"],
        ]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]

        repaired = ocr_backend._repair_late_dense_header_tokens(grid, confidence)

        self.assertEqual(repaired, {(2, 3), (2, 4)})
        self.assertEqual(
            grid[2],
            ["编号", "姓名", "组别", "项目", "联系电话", "报名状态"],
        )

        prefix_grid = [
            ["快递交接记录表", "", ""],
            ["运单号", "运单号收件人", "签收状态"],
            ["NO-1", "张伟", "完成"],
        ]
        prefix_confidence = [[0.95 if value else 0.0 for value in row] for row in prefix_grid]
        self.assertEqual(
            ocr_backend._repair_late_dense_header_tokens(
                prefix_grid,
                prefix_confidence,
            ),
            {(1, 1)},
        )
        self.assertEqual(prefix_grid[1][1], "收件人")

    def test_common_header_multiview_repairs_unique_one_glyph_variant(self):
        ocr_backend._load_runtime()
        grid = [
            ["餐饮采购验收表", "", ""],
            ["物料编码", "品名", "金收结果"],
            ["A01", "项目甲", "完成"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        image = np.full((90, 300, 3), 245, dtype=np.uint8)
        cv2.putText(image, "RESULT", (210, 52), cv2.FONT_HERSHEY_SIMPLEX, 0.35, (20, 20, 20), 1)
        output = SimpleNamespace(
            txts=["验收结果", "验收结果", "验收结果", "金收结果"],
            scores=[0.93, 0.92, 0.91, 0.85],
            imgs=None,
        )
        engine = SimpleNamespace(
            text_rec=Mock(return_value=output),
            server_text_rec=Mock(return_value=output),
        )

        repaired, scores = ocr_backend._repair_common_header_multiview(
            image,
            grid,
            confidence,
            [0, 100, 200, 300],
            [0, 30, 60, 90],
            engine,
        )

        self.assertEqual(repaired, {(1, 2)})
        self.assertEqual(grid[1][2], "验收结果")
        self.assertEqual(len(scores), 8)

    def test_common_header_multiview_recovers_dense_header_numeric_outlier(self):
        ocr_backend._load_runtime()
        grid = [
            ["工程材料领用表", "", "", ""],
            ["材料编码", "材料名称", "申请数量", "1207:"],
            ["A001-30", "一车间", "6721.05", "1207.33"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        image = np.full((90, 400, 3), 245, dtype=np.uint8)
        output = SimpleNamespace(
            txts=["实发数量", "实发数量", "实发数量", "实发数量"],
            scores=[0.94, 0.93, 0.92, 0.91],
            imgs=None,
        )
        engine = SimpleNamespace(
            text_rec=Mock(return_value=output),
            server_text_rec=Mock(return_value=output),
        )

        repaired, _ = ocr_backend._repair_common_header_multiview(
            image,
            grid,
            confidence,
            [0, 100, 200, 300, 400],
            [0, 30, 60, 90],
            engine,
        )

        self.assertEqual(repaired, {(1, 3)})
        self.assertEqual(grid[1][3], "实发数量")

    def test_common_header_multiview_removes_unverified_mark_from_nonboolean_header(self):
        ocr_backend._load_runtime()
        grid = [
            ["报名表", "", ""],
            ["☑项目", "姓名", "状态"],
            ["", "", ""],
        ]
        confidence = [[0.77 if value else 0.0 for value in row] for row in grid]
        image = np.full((90, 300, 3), 245, dtype=np.uint8)
        output = SimpleNamespace(
            txts=["项目"] * 4,
            scores=[0.96, 0.95, 0.94, 0.93],
            imgs=None,
        )
        engine = SimpleNamespace(
            text_rec=Mock(return_value=output),
            server_text_rec=Mock(return_value=output),
        )

        repaired, _ = ocr_backend._repair_common_header_multiview(
            image,
            grid,
            confidence,
            [0, 100, 200, 300],
            [0, 30, 60, 90],
            engine,
        )

        self.assertEqual(repaired, {(1, 0)})
        self.assertEqual(grid[1][0], "项目")

    def test_leading_glyph_omission_multiview_restores_repeated_peer(self):
        ocr_backend._load_runtime()
        grid = [
            ["编号", "区域", "状态"],
            ["1", "一车间", "完成"],
            ["2", "A区", "完成"],
            ["3", "车间", "复核"],
            ["4", "现场复核", "完成"],
            ["5", "标准件", "完成"],
            ["6", "批次-L08", "复核"],
            ["7", "常规项目", "完成"],
        ]
        confidence = [[0.99] * 3 for _ in grid]
        image = np.full((240, 300, 3), 245, dtype=np.uint8)
        output = SimpleNamespace(
            txts=["一车间", "一车间", "一车间", "车间"],
            scores=[0.95, 0.94, 0.93, 0.89],
            imgs=None,
        )
        engine = SimpleNamespace(
            text_rec=Mock(return_value=output),
            server_text_rec=Mock(return_value=output),
        )

        repaired, scores = ocr_backend._repair_leading_glyph_omission_multiview(
            image,
            grid,
            confidence,
            [0, 100, 200, 300],
            [row * 30 for row in range(9)],
            engine,
        )

        self.assertEqual(repaired, {(3, 1)})
        self.assertEqual(grid[3][1], "一车间")
        self.assertEqual(len(scores), 8)
        request = engine.text_rec.call_args.args[0]
        self.assertEqual(request.img[0].shape[:2], (26, 92))

    def test_status_checkmark_multiview_removes_only_supported_mark(self):
        ocr_backend._load_runtime()
        grid = [
            ["编号", "状态"],
            ["1", "正常"],
            ["2", "完成"],
            ["3", "☑正常"],
            ["4", "正常"],
            ["5", "完成"],
        ]
        confidence = [[0.99] * 2 for _ in grid]
        image = np.full((180, 200, 3), 245, dtype=np.uint8)
        output = SimpleNamespace(
            txts=["正常", "正常", "正常", "☑正常"],
            scores=[0.95, 0.94, 0.93, 0.89],
            imgs=None,
        )
        engine = SimpleNamespace(
            text_rec=Mock(return_value=output),
            server_text_rec=Mock(return_value=output),
        )

        repaired, scores = ocr_backend._repair_status_checkmark_multiview(
            image,
            grid,
            confidence,
            [0, 100, 200],
            [row * 30 for row in range(7)],
            engine,
        )

        self.assertEqual(repaired, {(3, 1)})
        self.assertEqual(grid[3][1], "正常")
        self.assertEqual(len(scores), 8)

    def test_repeated_categorical_blank_multiview_restores_unique_peer(self):
        ocr_backend._load_runtime()
        grid = [
            ["编号", "区域"],
            ["1", "A区"],
            ["2", "一车间"],
            ["3", "A区"],
            ["4", ""],
            ["5", "标准件"],
            ["6", "A区"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        image = np.full((210, 200, 3), 245, dtype=np.uint8)
        cv2.putText(
            image,
            "A",
            (120, 145),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.8,
            (20, 20, 20),
            2,
            cv2.LINE_AA,
        )
        output = SimpleNamespace(
            txts=["A区", "A区", "A区", "A区"],
            scores=[0.95, 0.94, 0.93, 0.92],
            imgs=None,
        )
        engine = SimpleNamespace(
            text_rec=Mock(return_value=output),
            server_text_rec=Mock(return_value=output),
        )

        repaired, scores = ocr_backend._repair_repeated_categorical_blank_multiview(
            image,
            grid,
            confidence,
            [0, 100, 200],
            [row * 30 for row in range(8)],
            engine,
        )

        self.assertEqual(repaired, {(4, 1)})
        self.assertEqual(grid[4][1], "A区")
        self.assertEqual(len(scores), 4)
        engine.server_text_rec.assert_not_called()

    def test_repeated_categorical_blank_completes_unique_repeated_prefix(self):
        ocr_backend._load_runtime()
        grid = [
            ["编号", "姓名"],
            ["1", "李娜"],
            ["2", "王强"],
            ["3", "李娜"],
            ["4", ""],
            ["5", "陈晨"],
            ["6", "李娜"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        image = np.full((210, 200, 3), 245, dtype=np.uint8)
        cv2.putText(
            image,
            "L",
            (120, 145),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.8,
            (20, 20, 20),
            2,
            cv2.LINE_AA,
        )
        medium_output = SimpleNamespace(
            txts=["李", "李", "李", "李"],
            scores=[0.79, 0.78, 0.70, 0.68],
            imgs=None,
        )
        alternate_output = SimpleNamespace(
            txts=["", "", "", ""],
            scores=[0.0, 0.0, 0.0, 0.0],
            imgs=None,
        )
        engine = SimpleNamespace(
            text_rec=Mock(return_value=medium_output),
            server_text_rec=Mock(return_value=alternate_output),
        )

        repaired, _ = ocr_backend._repair_repeated_categorical_blank_multiview(
            image,
            grid,
            confidence,
            [0, 100, 200],
            [row * 30 for row in range(8)],
            engine,
        )

        self.assertEqual(repaired, {(4, 1)})
        self.assertEqual(grid[4][1], "李娜")
        engine.server_text_rec.assert_not_called()

    def test_repeated_categorical_blank_skips_alternate_without_primary_candidate(self):
        ocr_backend._load_runtime()
        grid = [
            ["编号", "区域"],
            ["1", "A区"],
            ["2", "一车间"],
            ["3", "A区"],
            ["4", ""],
            ["5", "标准件"],
            ["6", "A区"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        image = np.full((210, 200, 3), 245, dtype=np.uint8)
        cv2.putText(image, "X", (120, 145), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (20, 20, 20), 2)
        medium_output = SimpleNamespace(
            txts=["未知"] * 4,
            scores=[0.99] * 4,
            imgs=None,
        )
        engine = SimpleNamespace(
            text_rec=Mock(return_value=medium_output),
            server_text_rec=Mock(side_effect=AssertionError("主视图无候选时不应调用备用模型")),
        )

        repaired, scores = ocr_backend._repair_repeated_categorical_blank_multiview(
            image,
            grid,
            confidence,
            [0, 100, 200],
            [row * 30 for row in range(8)],
            engine,
        )

        self.assertEqual(repaired, set())
        self.assertEqual(len(scores), 4)
        engine.server_text_rec.assert_not_called()

    def test_adjacent_duplicate_header_carries_verified_confidence_after_cleanup(self):
        grid = [
            ["编号", "设备名称", "型号", "频率", "功率", "状态备注", "备注"],
            ["A001", "信号发生器", "SG-2200", "515.472", "-20", "复核", "校准完成"],
        ]
        confidence = [[0.95] * 7 for _ in grid]
        confidence[0][5] = -1.0

        split, split_confidence, _ = ocr_backend._split_collapsed_header_data_row(
            grid,
            confidence,
            [],
        )

        self.assertEqual(split[0][5:], ["状态", "备注"])
        self.assertEqual(split_confidence[0][5], 0.78)

    def test_ruled_grid_geometry_uses_cell_centres_for_missing_cell_recovery(self):
        geometry = ocr_backend._ruled_grid_recovery_geometry(
            [10, 50, 90],
            [20, 60, 100, 140],
        )

        self.assertEqual(geometry["anchors"], [30.0, 70.0])
        self.assertEqual(geometry["row_centers"], [40.0, 80.0, 120.0])
        self.assertEqual(len(geometry["grouped_rows"]), 3)

    def test_screen_grid_keeps_legitimate_multiline_remarks(self):
        grid = [
            ["编号", "名称", "备注"],
            ["1", "频谱仪", "第一行说明\n第二行说明"],
            ["2", "信号源", "正常"],
            ["3", "示波器", "复核"],
        ]
        self.assertFalse(ocr_backend._screen_grid_has_collapsed_rows(grid))

    def test_screen_grid_keeps_bilingual_and_explanatory_multiline_cells(self):
        grid = [
            ["条目", "检测项目", "技术要求", "实测结果", "判定", "说明"],
            ["1", "供电电压\nPower Input", "额定：DC 12 V\n允许范围：10.8～13.2 V", "12.06 V\n纹波：18 mV", "合格", "室温25°C；湿度46%RH"],
            ["2", "频率精度\nFrequency Accuracy", "误差≤±0.5 ppm\n预热时间≥15 min", "+0.18 ppm\n预热20 min", "合格", "参考源：10 MHz铷钟"],
        ]

        self.assertFalse(ocr_backend._screen_grid_has_collapsed_rows(grid))

    def test_grid_rejects_row_with_multiple_physical_rows_concatenated_per_cell(self):
        grid = [
            ["编号", "名称", "数量", "状态"],
            [
                "A01 A02 A03 A04 A05",
                "阀门 电机 泵体 仪表 传感器",
                "10 20 30 40 50",
                "正常 正常 停机 正常 检修",
            ],
        ]
        self.assertTrue(ocr_backend._grid_has_concatenated_physical_rows(grid))
        self.assertFalse(
            ocr_backend._grid_has_concatenated_physical_rows(
                [["编号", "名称", "备注"], ["A01", "阀门", "等待现场复核"]]
            )
        )

    def test_grid_rejects_two_physical_rows_concatenated_across_many_cells(self):
        grid = [
            ["编号", "中心频率", "带宽", "调制方式", "信号名称"],
            [
                "1 2",
                "515.128 MHz 516.347 MHz",
                "9 kHz 35.4 kHz",
                "AM BPSK",
                "fm_9k_Test#01 bpsk_35.4k_28k_0.1",
            ],
        ]

        self.assertTrue(ocr_backend._grid_has_concatenated_physical_rows(grid))

    def test_grid_rejects_multiple_header_columns_concatenated_into_one_cell(self):
        collapsed = [
            ["项目任务跟踪表", ""],
            ["任务编号", "任务名称 优先级 进度 负责人 计划完成 当前状态"],
            ["TASK-001", "界面设计 中 7% 陈晨 2026-08-02 已完成"],
        ]

        self.assertTrue(ocr_backend._grid_has_concatenated_header_columns(collapsed))
        self.assertFalse(
            ocr_backend._grid_has_concatenated_header_columns(
                [["编号", "说明"], ["A01", "等待现场复核"]]
            )
        )

    def test_grid_rejects_multilevel_measurement_headers_concatenated_by_glare(self):
        collapsed = [
            ["", "", ""],
            [
                "",
                "综合测试数据表 基础信息 项目 类别 数值1 数值2 数值3 单位 备注",
                "",
            ],
            ["", "", ""],
        ]

        self.assertTrue(ocr_backend._grid_has_concatenated_header_columns(collapsed))

    def test_wide_screen_grid_rejects_multiple_fused_header_cells(self):
        collapsed = [
            [
                "序号项目组类别",
                "名称参数",
                "T(ms)",
                "TuTs",
                "Tg采样率带宽",
                "系数1比值1",
                "比值2峰值",
                "解调方式备注",
                "创建日期责任人",
                "版本状态",
                "校验码",
            ],
            ["1G01模拟信号", "AM_Signal_001默认"] + [""] * 9,
        ]

        self.assertTrue(ocr_backend._grid_has_concatenated_header_columns(collapsed))

    def test_wide_screen_grid_keeps_independent_headers(self):
        complete = [
            ["序号", "项目组", "类别", "名称", "参数", "T(ms)", "Tu", "Ts"],
            ["1", "G01", "模拟信号", "AM_Signal_001", "默认", "0.00012", "0.0019", "0.017"],
        ]

        self.assertFalse(ocr_backend._grid_has_concatenated_header_columns(complete))

    def test_grid_rejects_multiple_physical_columns_fused_into_one_column(self):
        numeric_columns_fused = [
            ["日期", "编号", "设备名称", "功率(dBm) 频率(MHz) 温度(°C)", "状态", "备注"],
            ["2026-08-05", "A001", "基站收发信机1", "36.4 515.221 -10", "正常", "—"],
            ["2026-08-05", "A002", "基站收发信机2", "37.1 520.125 -10", "正常", "—"],
        ]
        status_and_risk_fused = [
            ["任务编号", "任务名称", "优先级", "进度", "负责人", "计划完成", "当前状态 风险说明"],
            ["TASK-001", "OCR场景验证1", "中", "7%", "李娜", "2026-08-06", "进行中 无"],
            ["TASK-004", "OCR场景验证4", "低", "28%", "张伟", "2026-08-09", "已完成 阴影较重，需人工核对"],
            ["TASK-008", "OCR场景验证8", "低", "56%", "张伟", "2026-08-13", "异常 阴影较重，需人工核对"],
        ]
        progress_and_owner_fused = [
            ["任务编号", "任务名称", "优先级", "进度 负责人", "计划完成", "当前状态"],
            ["TASK-001", "界面设计", "中", "7% 陈晨", "2026-08-02", "已完成"],
            ["TASK-002", "图像预处理", "低", "14% 王强", "2026-08-03", "等待"],
            ["TASK-003", "表格结构识别", "高", "21% 赵敏", "2026-08-04", "进行中"],
        ]
        identifier_and_name_fused = [
            ["设备名称 编号", "型号", "频率(MHz)", "状态"],
            ["信号发生器 A001", "SG-2200", "515.472", "复核"],
            ["数字示波器 A002", "DSO-X3104T", "515.819", "待机"],
            ["直流稳压源 A003", "DP832A", "516.166", "正常"],
        ]
        complete = [
            ["日期", "编号", "设备名称", "频率(MHz)", "功率(dBm)", "温度(°C)", "状态", "备注"],
            ["2026-08-05", "A001", "基站收发信机1", "515.221", "-10", "36.4", "正常", "—"],
        ]

        self.assertTrue(ocr_backend._grid_has_fused_physical_columns(numeric_columns_fused))
        self.assertTrue(ocr_backend._grid_has_fused_physical_columns(status_and_risk_fused))
        self.assertTrue(ocr_backend._grid_has_fused_physical_columns(progress_and_owner_fused))
        self.assertTrue(ocr_backend._grid_has_fused_physical_columns(identifier_and_name_fused))
        self.assertFalse(ocr_backend._grid_has_fused_physical_columns(complete))

    def test_grid_rejects_many_fields_collapsed_into_one_output_column(self):
        collapsed = [
            ["编号设备名称型号频率(MHz)功率(dBm)状态备注"],
            ["A001信号发生器SG-2200515.472-20复核校准完成"],
            ["A002数字示波器DSO-X3104T515.819-30待机接口检查"],
            ["A003直流稳压源DP832A516.166-10正常—"],
        ]
        legitimate_single_column = [
            ["检查项目"],
            ["设备外观无明显破损"],
            ["电源线连接牢固"],
            ["风扇运转正常"],
        ]

        self.assertTrue(ocr_backend._grid_has_fused_physical_columns(collapsed))
        self.assertFalse(ocr_backend._grid_has_fused_physical_columns(legitimate_single_column))

    def test_borderless_grid_rejects_split_row_between_consecutive_identifiers(self):
        shifted = [
            ["样品编号", "检测项目", "结果值", "单位", "结论"],
            ["S-001", "含水率", "3.25", "%", "合格"],
            ["S-002", "密度", "1.08", "g/cm³", "合格"],
            ["S-003", "硬度", "72", "HA", "复测"],
            ["", "", "", "ΔE", "待测"],
            ["S-004", "色差", "28.6", "MPa", "合格"],
            ["S-005", "拉伸强度", "145", "%", "合格"],
        ]
        complete = [
            ["样品编号", "检测项目", "结果值"],
            ["S-001", "含水率", "3.25"],
            ["S-002", "密度", "1.08"],
            ["S-003", "硬度", "72"],
            ["S-004", "色差", ""],
            ["S-005", "拉伸强度", "28.6"],
        ]

        self.assertTrue(ocr_backend._grid_has_misaligned_identifier_rows(shifted))
        self.assertFalse(ocr_backend._grid_has_misaligned_identifier_rows(complete))

    def test_grid_rejects_dense_blank_row_between_consecutive_month_anchors(self):
        shifted = [
            ["月份", "华东地区", "华南地区", "合计"],
            ["9月", "1,598.7", "1,452.3", "6,758.0"],
            ["10月", "1,678.9", "1,529.4", "7,093.3"],
            ["", "1,705.6", "1,553.2", "7,212.7"],
            ["11月", "1,820.4", "1,648.7", "7,668.2"],
            ["12月", "18,046.8", "17,423.0", "78,026.4"],
            ["全年合计", "18,046.8", "17,423.0", "78,026.4"],
        ]
        complete = [
            ["月份", "华东地区", "华南地区", "合计"],
            ["9月", "1,598.7", "1,452.3", "6,758.0"],
            ["10月", "1,678.9", "1,529.4", "7,093.3"],
            ["11月", "1,705.6", "1,553.2", "7,212.7"],
            ["12月", "1,820.4", "1,648.7", "7,668.2"],
            ["全年合计", "18,046.8", "17,423.0", "78,026.4"],
        ]

        self.assertTrue(ocr_backend._grid_has_misaligned_anchor_rows(shifted))
        self.assertFalse(ocr_backend._grid_has_misaligned_anchor_rows(complete))

    def test_spatial_header_population_skips_merged_title_row(self):
        grid = [
            ["设备运行统计表", "", ""],
            ["编号", "名称", "状态"],
            ["A01", "阀门", "正常"],
        ]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 3, "role": "title"}
        ]
        self.assertEqual(ocr_backend._spatial_header_population(grid, spans), (3, 3))
        self.assertEqual(ocr_backend._spatial_header_population([["", "", ""]], []), (0, 3))

    def test_spatial_multilevel_header_requires_sparse_groups_then_dense_labels(self):
        grid = [
            ["生产日报", "", "", "", "", "", ""],
            ["基础信息", "", "", "测量与判定", "", "", ""],
            ["日期", "班次", "生产线", "计划数量", "完成数量", "不良数量", "完成率"],
        ]
        spans = [{"row": 0, "column": 0, "column_span": 7, "role": "title"}]

        self.assertTrue(ocr_backend._spatial_has_supported_multilevel_header(grid, spans))
        grid[2][4:] = ["", "", ""]
        self.assertFalse(ocr_backend._spatial_has_supported_multilevel_header(grid, spans))

    def test_spatial_multilevel_header_accepts_two_missing_labels_when_body_is_repeated(self):
        grid = [
            ["生产日报", "", "", "", "", "", ""],
            ["基础信息", "", "", "测量与判定", "", "", ""],
            ["日期", "班次", "生产线", "计划数量", "", "", "完成率"],
        ] + [
            [f"2026-08-{day:02d}", "白班", "A线", "300", "296", "4", "98.4%"]
            for day in range(2, 10)
        ]
        spans = [{"row": 0, "column": 0, "column_span": 7, "role": "title"}]

        self.assertTrue(ocr_backend._spatial_has_supported_multilevel_header(grid, spans))

        del grid[4:]
        self.assertFalse(ocr_backend._spatial_has_supported_multilevel_header(grid, spans))

    def test_spatial_multilevel_header_accepts_implicit_sparse_title_after_ruler_rebuild(self):
        grid = [
            ["", "", "", "生产日报（2026年8月）", "", "", ""],
            ["基础信息", "", "", "测量与判定", "", "质量记录", ""],
            ["日期", "班次", "生产线", "计划数量", "完成数量", "不良数量", "完成率"],
        ] + [
            [f"2026-08-{day:02d}", "白班", "A线", "300", "296", "4", "98.4%"]
            for day in range(2, 8)
        ]

        self.assertTrue(ocr_backend._spatial_has_implicit_title_multilevel_header(grid))
        grid[0] = ["标题", "副标题", "版本", "", "", "", ""]
        self.assertFalse(ocr_backend._spatial_has_implicit_title_multilevel_header(grid))

    def test_spreadsheet_ruler_can_confirm_rows_swallowed_by_page_grid(self):
        page = [["x"] * 7 for _ in range(12)]
        spatial = [
            ["", "", "", "生产日报（2026年8月）", "", "", ""],
            ["基础信息", "", "", "测量与判定", "", "质量记录", ""],
            ["日期", "班次", "生产线", "计划数量", "完成数量", "不良数量", "完成率"],
        ] + [
            [f"2026-08-{day:02d}", "白班", "A线", "300", "296", "4", "98.4%"]
            for day in range(2, 18)
        ]

        self.assertTrue(
            ocr_backend._spreadsheet_ruler_confirms_collapsed_page_rows(
                page, spatial, [], True
            )
        )
        self.assertTrue(
            ocr_backend._spreadsheet_ruler_spatial_grid_is_complete(
                spatial, [], True
            )
        )
        self.assertFalse(
            ocr_backend._spreadsheet_ruler_confirms_collapsed_page_rows(
                page, [row[:-1] for row in spatial], [], True
            )
        )
        sparse = [row[:] for row in spatial]
        for row in sparse[3:]:
            row[2:] = [""] * (len(row) - 2)
        self.assertFalse(
            ocr_backend._spreadsheet_ruler_spatial_grid_is_complete(
                sparse, [], True
            )
        )

    def test_spreadsheet_rulers_confirm_dense_table_without_semantic_header(self):
        spatial = [["", "教学课表成绩 2026年01月", "", ""], ["序号", "学号", "姓名", "班级"]] + [
            [str(index), "B区", "周工", "一号线"] for index in range(1, 15)
        ]

        self.assertFalse(ocr_backend._spatial_has_credible_header(spatial, []))
        self.assertTrue(
            ocr_backend._spreadsheet_ruler_spatial_grid_is_complete(
                spatial, [], True
            )
        )

    def test_spreadsheet_rulers_accept_dense_three_column_table(self):
        spatial = [["序号", "物料编码", "物料名称"]] + [
            [str(index), f"MAT-{index:03d}", "温度传感器"]
            for index in range(1, 17)
        ]

        self.assertTrue(
            ocr_backend._spreadsheet_ruler_spatial_grid_is_complete(
                spatial, [], True
            )
        )
        sparse = [row[:] for row in spatial]
        for row in sparse[2:]:
            row[1:] = ["", ""]
        self.assertFalse(
            ocr_backend._spreadsheet_ruler_spatial_grid_is_complete(
                sparse, [], True
            )
        )

    def test_spatial_ruler_recovery_may_expand_but_never_shrink_page_extent(self):
        self.assertTrue(
            ocr_backend._spatial_ruler_recovery_expands_page_extent(
                38, 15, 40, 15
            )
        )
        self.assertTrue(
            ocr_backend._spatial_ruler_recovery_expands_page_extent(
                35, 14, 35, 15
            )
        )
        self.assertFalse(
            ocr_backend._spatial_ruler_recovery_expands_page_extent(
                28, 12, 28, 11
            )
        )
        self.assertFalse(
            ocr_backend._spatial_ruler_recovery_expands_page_extent(
                40, 15, 40, 15
            )
        )

    def test_spatial_layout_rejects_body_values_misclassified_as_header(self):
        grid = [
            ["综合测试数据表", "", ""],
            ["15", "0.375", "mA"],
            ["18", "0.625", "N/A"],
            ["21", "0.875", "复测"],
        ]
        spans = [{"row": 0, "column": 0, "column_span": 3, "role": "title"}]

        self.assertFalse(ocr_backend._spatial_has_credible_header(grid, spans))
        self.assertTrue(
            ocr_backend._spatial_has_credible_header(
                [["项目", "数值", "单位"], ["A", "15", "mA"]],
                [],
            )
        )

    def test_spatial_fallback_uses_raw_header_instead_of_false_multilevel_header(self):
        raw = [["", "", "", "", "", ""]] + [
            ["A001", "信号发生器 SG-2200", "515.472", "-20", "复核", "校准完成"],
            ["A002", "DSO-X3104T", "515.819", "-30", "待机", "接口检查"],
        ]
        verified = [
            ["", "", "频率(MHz)", "功率(dBm)", "", "备注"],
            ["A001", "信号发生器 SG-2200", "515.472", "-20", "复核", "校准完成"],
            ["A002", "DSO-X3104T", "515.819", "-30", "待机", "接口检查"],
        ]

        self.assertTrue(ocr_backend._spatial_has_credible_header(verified, []))
        self.assertFalse(
            ocr_backend._spatial_fallback_is_reliable(
                verified,
                [],
                True,
                raw,
                [],
                False,
            )
        )

    def test_spatial_fallback_accepts_complete_layout_when_line_grid_is_wrong(self):
        grid = [
            ["任务编号", "任务名称", "优先级", "进度", "负责人", "计划完成", "当前状态"],
            ["TASK-001", "界面设计", "中", "7%", "陈晨", "2026-08-02", "已完成"],
            ["TASK-002", "图像预处理", "低", "14%", "王强", "2026-08-03", "等待"],
        ]

        self.assertTrue(
            ocr_backend._spatial_fallback_is_reliable(
                grid,
                [],
                True,
                grid,
                [],
                False,
            )
        )
        self.assertFalse(
            ocr_backend._spatial_fallback_is_reliable(
                [row[:-1] for row in grid],
                [],
                True,
                grid,
                [],
                False,
            )
        )

    def test_spatial_fallback_accepts_dense_detail_header_when_one_group_label_is_missed(self):
        raw_grid = [
            ["项目任务跟踪表", "", "", "", "", "", ""],
            ["", "基础信息", "", "", "", "", ""],
            ["任务编号", "任务名称", "优先级", "进度", "负责人", "计划完成", "当前状态"],
        ]
        for index in range(6):
            raw_grid.append(
                [
                    f"TASK-{index + 1:03d}",
                    "图像预处理",
                    "高",
                    "42%",
                    "王强",
                    "2026-08-07",
                    "进行中",
                ]
            )
        verified_grid = [row[:] for row in raw_grid]
        verified_grid[1][5] = "测量与判定"
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 7, "role": "title"}
        ]

        self.assertTrue(
            ocr_backend._spatial_fallback_is_reliable(
                verified_grid,
                spans,
                True,
                raw_grid,
                spans,
                True,
            )
        )

    def test_page_grid_must_preserve_detected_outer_rows_and_columns(self):
        columns = list(range(8))
        rows = list(range(17))
        complete = [["x"] * 7 for _ in range(16)]
        missing_outer_rows = [["x"] * 7 for _ in range(13)]

        self.assertTrue(
            ocr_backend._grid_preserves_detected_extent(complete, columns, rows)
        )
        self.assertFalse(
            ocr_backend._grid_preserves_detected_extent(
                missing_outer_rows,
                columns,
                rows,
            )
        )

    def test_clear_credible_ruled_grid_can_use_bounded_cell_fallback(self):
        clear = {"sharpness": 320.0, "needs_recapture": False}
        blurred = {"sharpness": 6.0, "needs_recapture": True}

        self.assertTrue(
            ocr_backend._clear_image_allows_ruled_cell_fallback(True, clear, clear)
        )
        self.assertFalse(
            ocr_backend._clear_image_allows_ruled_cell_fallback(False, clear, clear)
        )
        self.assertFalse(
            ocr_backend._clear_image_allows_ruled_cell_fallback(True, clear, blurred)
        )

    def test_page_grid_extent_accepts_matching_verified_spatial_layout(self):
        columns = list(range(8))
        rows_with_one_false_boundary = list(range(18))
        grid = [
            ["日期", "班次", "生产线", "计划数量", "完成数量", "不良数量", "完成率"],
            ["2026-08-02", "夜班", "B线", "280", "277", "3", "98.2%"],
        ] + [["x"] * 7 for _ in range(14)]

        self.assertTrue(
            ocr_backend._grid_extent_is_independently_confirmed(
                grid,
                columns,
                rows_with_one_false_boundary,
                grid,
                [],
                True,
            )
        )
        self.assertFalse(
            ocr_backend._grid_extent_is_independently_confirmed(
                [["x"] * 7 for _ in range(13)],
                list(range(9)),
                list(range(14)),
                [["x"] * 5 for _ in range(16)],
                [],
                True,
            )
        )

    def test_page_grid_extent_accepts_strictly_stripped_spreadsheet_rulers(self):
        grid = [
            ["序号", "检查编号", "项目名称", "区域"],
            ["1", "AP-001", "主控模块", "A区"],
            ["2", "AP-002", "传感器组", "B区"],
            ["3", "AP-003", "校准组", "C区"],
        ]
        spatial_grid = [
            ["", "A", "B", "C", "D"],
            ["1", *grid[0]],
            ["2", *grid[1]],
            ["3", *grid[2]],
            ["4", *grid[3]],
        ]

        self.assertTrue(
            ocr_backend._grid_extent_is_independently_confirmed(
                grid,
                list(range(6)),
                list(range(6)),
                spatial_grid,
                [],
                True,
            )
        )

    def test_page_grid_extent_does_not_strip_spatial_grid_with_spans(self):
        grid = [
            ["序号", "检查编号", "项目名称", "区域"],
            ["1", "AP-001", "主控模块", "A区"],
            ["2", "AP-002", "传感器组", "B区"],
            ["3", "AP-003", "校准组", "C区"],
        ]
        spatial_grid = [
            ["", "A", "B", "C", "D"],
            ["1", *grid[0]],
            ["2", *grid[1]],
            ["3", *grid[2]],
            ["4", *grid[3]],
        ]

        self.assertFalse(
            ocr_backend._grid_extent_is_independently_confirmed(
                grid,
                list(range(6)),
                list(range(6)),
                spatial_grid,
                [{"row": 0, "column": 1, "rowSpan": 1, "columnSpan": 4}],
                True,
            )
        )

    def test_page_grid_extent_accepts_stripped_three_level_header(self):
        grid = [
            ["基础信息", "", "", "", "目标与测量", "", "过程记录"],
            ["上限", "", "下限", "", "本期", "实际", "下限"],
            ["序号", "检查编号", "项目名称", "区域", "日期", "检查项", "结果"],
            ["1", "AP-001", "主控模块", "A区", "2026-08-17", "接地", "合格"],
            ["2", "AP-002", "传感器组", "B区", "2026-08-17", "防护", "合格"],
            ["3", "AP-003", "校准组", "C区", "2026-08-17", "围栏", "合格"],
            ["4", "AP-004", "主控模块", "D区", "2026-08-17", "标识", "合格"],
        ]
        spatial_grid = [
            ["", "A", "B", "C", "D", "E", "F", "G"],
            *[[str(index + 1), *row] for index, row in enumerate(grid)],
        ]

        self.assertTrue(
            ocr_backend._grid_extent_is_independently_confirmed(
                grid,
                list(range(9)),
                list(range(9)),
                spatial_grid,
                [],
                True,
            )
        )

    def test_page_grid_extent_rejects_sparse_headers_without_dense_detail_or_numbered_body(self):
        grid = [
            ["基础信息", "", "", "", "目标与测量", "", "过程记录"],
            ["上限", "", "下限", "", "本期", "实际", "下限"],
            ["项目", "", "", "区域", "", "", ""],
            ["", "AP-001", "主控模块", "A区", "2026-08-17", "接地", "合格"],
            ["", "AP-002", "传感器组", "B区", "2026-08-17", "防护", "合格"],
            ["", "AP-003", "校准组", "C区", "2026-08-17", "围栏", "合格"],
        ]
        spatial_grid = [
            ["", "A", "B", "C", "D", "E", "F", "G"],
            *[[str(index + 1), *row] for index, row in enumerate(grid)],
        ]

        self.assertFalse(
            ocr_backend._grid_extent_is_independently_confirmed(
                grid,
                list(range(9)),
                list(range(8)),
                spatial_grid,
                [],
                True,
            )
        )

    def test_processed_structure_certificate_confirms_stripped_grid_extent(self):
        grid = [["x"] * 8 for _ in range(25)]
        certificate = {
            "verified": True,
            "ui_headers_processed": True,
            "rows": 25,
            "columns": 8,
            "row_offset": 1,
            "column_offset": 0,
        }

        self.assertTrue(
            ocr_backend._processed_structure_certificate_confirms_grid_extent(
                grid,
                certificate,
            )
        )

    def test_unprocessed_or_unverified_certificate_cannot_confirm_grid_extent(self):
        grid = [["x"] * 8 for _ in range(25)]
        unprocessed = {
            "verified": True,
            "ui_headers_processed": False,
            "rows": 25,
            "columns": 8,
        }
        unverified = {
            "verified": False,
            "ui_headers_processed": True,
            "rows": 25,
            "columns": 8,
        }

        self.assertFalse(
            ocr_backend._processed_structure_certificate_confirms_grid_extent(
                grid,
                unprocessed,
            )
        )
        self.assertFalse(
            ocr_backend._processed_structure_certificate_confirms_grid_extent(
                grid,
                unverified,
            )
        )

    def test_sparse_group_header_reanchors_overlapping_ocr_fragments(self):
        grid = [
            ["报表标题", "", "", "", "", "", "", ""],
            ["", "基础信息", "", "", "目标与", "与测量", "过程", "记录"],
            ["区域", "编号", "名称", "说明", "当前值", "目标值", "单位", "来源"],
            *[[str(index), "A", "设备", "说明", "1.000", "2.000", "V", "系统"] for index in range(1, 6)],
        ]
        confidence = [[0.95] * 8 for _ in grid]
        columns = list(range(0, 801, 100))
        rows = list(range(0, 401, 50))
        evidence = [
            ("基础信息", 200.0, 75.0, 0.98),
            ("目标与测量", 500.0, 75.0, 0.98),
            ("过程记录", 700.0, 75.0, 0.98),
        ]

        rebuilt, _, spans = ocr_backend._infer_sparse_group_header_spans(
            grid,
            confidence,
            [{"row": 0, "column": 0, "row_span": 1, "column_span": 8, "role": "title"}],
            image=np.full((400, 800, 3), 255, dtype=np.uint8),
            columns=columns,
            rows=rows,
            page_text_evidence=evidence,
        )

        self.assertEqual(rebuilt[1], ["基础信息", "", "", "", "目标与测量", "", "过程记录", ""])
        self.assertEqual(
            [(span["row"], span["column"], span["column_span"]) for span in spans if span.get("role") == "group_header"],
            [(1, 0, 4), (1, 4, 2), (1, 6, 2)],
        )

    def test_title_span_locks_subordinates_and_expands_anchor_crop(self):
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 8, "role": "title"},
            {"row": 1, "column": 0, "row_span": 1, "column_span": 4, "role": "group_header"},
        ]

        locks, widths = ocr_backend._title_span_consensus_geometry(spans, 25, 8)

        self.assertEqual(locks, {(0, column) for column in range(1, 8)})
        self.assertEqual(widths, {(0, 0): 8})

    def test_extended_title_requires_matching_high_confidence_views(self):
        selected = ocr_backend._select_extended_title_candidate(
            "稀疏长备",
            [("稀疏长备注宽表 — 2026-01 批次", 0.96)],
            [("稀疏长备注宽表 — 2026-01 批次", 0.95)],
        )

        self.assertEqual(selected, ("稀疏长备注宽表 — 2026-01 批次", 0.95))
        self.assertIsNone(
            ocr_backend._select_extended_title_candidate(
                "稀疏长备",
                [("稀疏长备注宽表 — 2026-01 批次", 0.96)],
                [("稀疏长备注宽表 — 2026-02 批次", 0.97)],
            )
        )

    def test_photographic_background_bypasses_direct_screen_grid_path(self):
        screenshot = np.full((600, 900, 3), 250, dtype=np.uint8)
        photographed = np.full((600, 900, 3), (155, 175, 195), dtype=np.uint8)

        self.assertFalse(ocr_backend._has_photographic_background(screenshot))
        self.assertTrue(ocr_backend._has_photographic_background(photographed))

    def test_small_clear_screen_grid_uses_page_primary_only_for_bounded_regular_tables(self):
        ocr_backend._load_runtime()
        image = np.full((1144, 800, 3), 255, dtype=np.uint8)
        columns = [0, 200, 400, 600, 799]
        rows = [44 * index for index in range(26)] + [1143]
        for row in rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (30, 30, 30), 2)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (30, 30, 30), 2)

        self.assertTrue(
            ocr_backend._small_clear_screen_grid_uses_page_primary(
                image,
                columns,
                rows,
            )
        )
        self.assertFalse(
            ocr_backend._small_clear_screen_grid_uses_page_primary(
                image,
                list(range(0, 800, 20)) + [799],
                rows,
            )
        )

    def test_missing_full_width_title_recovery_requires_matching_table_body(self):
        page_grid = [
            ["编号", "名称", "结果"],
            ["A-001", "设备一", "正常"],
            ["A-002", "设备二", "复核"],
        ]
        spatial_grid = [
            ["设备巡检记录表", "", ""],
            *page_grid,
        ]
        page_confidence = [[0.99] * 3 for _ in page_grid]
        spatial_confidence = [[0.99, 0.0, 0.0], *page_confidence]
        spans = [
            {
                "row": 0,
                "column": 0,
                "row_span": 1,
                "column_span": 3,
                "role": "title",
            }
        ]

        recovered = ocr_backend._recover_missing_full_width_title_from_spatial_layout(
            page_grid,
            page_confidence,
            spatial_grid,
            spatial_confidence,
            spans,
            True,
        )

        self.assertIsNotNone(recovered)
        self.assertEqual(recovered[0], spatial_grid)
        self.assertTrue(
            ocr_backend._grids_have_matching_cell_ownership(page_grid, page_grid)
        )
        mismatched_spatial = [list(row) for row in spatial_grid]
        mismatched_spatial[1:] = [
            [f"错误-{row}-{column}" for column in range(3)]
            for row in range(3)
        ]
        self.assertIsNone(
            ocr_backend._recover_missing_full_width_title_from_spatial_layout(
                page_grid,
                page_confidence,
                mismatched_spatial,
                spatial_confidence,
                spans,
                True,
            )
        )
        self.assertFalse(
            ocr_backend._grids_have_matching_cell_ownership(
                page_grid,
                mismatched_spatial[1:],
            )
        )

    def test_spatial_header_stack_aligns_with_matching_page_body_and_trims_margins(self):
        spatial_grid = [
            ["月度统计表", "", "", ""],
            ["部门：运营部  期间：2026-08", "", "", ""],
            ["编号", "名称", "数值", "状态"],
            ["1", "设备一", "10.25", "正常"],
            ["2", "设备二", "11.50", "复核"],
            ["3", "设备三", "12.75", "正常"],
            ["4", "设备四", "13.00", "完成"],
        ]
        page_grid = [
            *[list(row) for row in spatial_grid[1:]],
            ["", "", "", ""],
            ["", "", "", ""],
        ]
        spatial_confidence = [[0.99 if value else 0.0 for value in row] for row in spatial_grid]
        page_confidence = [[0.99 if value else 0.0 for value in row] for row in page_grid]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 4, "role": "title"},
            {"row": 1, "column": 0, "row_span": 1, "column_span": 4, "role": "subtitle"},
        ]

        recovered = ocr_backend._recover_spatial_headers_with_aligned_page_body(
            page_grid,
            page_confidence,
            spatial_grid,
            spatial_confidence,
            spans,
            True,
        )

        self.assertIsNotNone(recovered)
        self.assertEqual(recovered[0], spatial_grid)
        self.assertEqual(recovered[2], spans)
        self.assertEqual(recovered[3]["page_start"], 0)
        self.assertEqual(recovered[3]["spatial_start"], 1)
        mismatched = [list(row) for row in page_grid]
        for row in range(4):
            mismatched[row] = [f"错误{row}-{column}" for column in range(4)]
        self.assertIsNone(
            ocr_backend._recover_spatial_headers_with_aligned_page_body(
                mismatched,
                page_confidence,
                spatial_grid,
                spatial_confidence,
                spans,
                True,
            )
        )

    def test_aligned_page_metadata_row_is_not_dropped_by_spatial_header(self):
        page_grid = [
            ["售后服务回访表", "", "", ""],
            ["部门：综合管理部", "", "批次：B84", ""],
            ["编号", "客户", "日期", "状态"],
            *[[str(row), f"客户{row}", f"2026-08-{row:02d}", "完成"] for row in range(1, 7)],
        ]
        spatial_grid = [page_grid[0], *page_grid[2:]]
        page_confidence = [[0.99 if value else 0.0 for value in row] for row in page_grid]
        spatial_confidence = [
            [0.99 if value else 0.0 for value in row]
            for row in spatial_grid
        ]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 4, "role": "title"}
        ]

        recovered = ocr_backend._recover_spatial_headers_with_aligned_page_body(
            page_grid,
            page_confidence,
            spatial_grid,
            spatial_confidence,
            spans,
            True,
        )

        self.assertIsNotNone(recovered)
        assert recovered is not None
        self.assertEqual(recovered[0], page_grid)
        self.assertEqual(recovered[3]["inserted_metadata_rows"], 1)

    def test_title_only_alignment_can_reuse_verified_physical_certificate(self):
        page_grid = [
            ["", "", "", ""],
            ["编号", "名称", "数值", "状态"],
            *[[str(row), f"设备{row}", f"{row}.25", "正常"] for row in range(1, 7)],
        ]
        recovered_grid = [
            ["设备巡检记录表", "", "", ""],
            *[list(row) for row in page_grid[1:]],
        ]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 4, "role": "title"}
        ]
        metrics = {
            "page_start": 1,
            "spatial_start": 1,
            "overlap_rows": len(recovered_grid) - 1,
            "inserted_metadata_rows": 0,
        }
        certificate = {
            "verified": True,
            "rows": len(recovered_grid),
            "columns": 4,
        }

        self.assertTrue(
            ocr_backend._aligned_header_can_keep_physical_certificate(
                page_grid,
                recovered_grid,
                spans,
                metrics,
                certificate,
            )
        )

        shifted_metrics = dict(metrics, page_start=2)
        self.assertFalse(
            ocr_backend._aligned_header_can_keep_physical_certificate(
                page_grid,
                recovered_grid,
                spans,
                shifted_metrics,
                certificate,
            )
        )
        mismatched = [list(row) for row in recovered_grid]
        for row in range(3, 7):
            mismatched[row] = [f"错位{row}-{column}" for column in range(4)]
        self.assertFalse(
            ocr_backend._aligned_header_can_keep_physical_certificate(
                page_grid,
                mismatched,
                spans,
                metrics,
                certificate,
            )
        )
        sparse_metadata = [list(row) for row in recovered_grid]
        sparse_metadata[1] = ["制表：运营中心", "", "", "第2页"]
        self.assertFalse(
            ocr_backend._aligned_header_can_keep_physical_certificate(
                page_grid,
                sparse_metadata,
                spans,
                metrics,
                certificate,
            )
        )

    def test_page_rows_are_ordered_subset_of_expanded_spatial_grid(self):
        spatial = [
            ["快递交接记录表", "", ""],
            ["部门：服务中心", "", ""],
            ["运单号", "收件人", "状态"],
            *[[f"NO-{row}", f"姓名{row}", "完成"] for row in range(1, 9)],
        ]
        page = [
            spatial[2],
            spatial[3],
            spatial[4],
            spatial[6],
            spatial[8],
            spatial[10],
        ]

        self.assertTrue(
            ocr_backend._page_rows_are_ordered_subset_of_spatial_grid(
                page,
                spatial,
            )
        )
        shuffled = [list(row) for row in page]
        shuffled[2], shuffled[3] = shuffled[3], shuffled[2]
        self.assertFalse(
            ocr_backend._page_rows_are_ordered_subset_of_spatial_grid(
                shuffled,
                spatial,
            )
        )

    def test_dense_axis_aligned_grid_overrides_dark_screenshot_chrome(self):
        ocr_backend._load_runtime()
        image = np.full((1068, 1857, 3), 245, dtype=np.uint8)
        image[:, :10] = 20
        image[:, 1816:] = 20
        image[:15] = 20
        image[1054:] = 20
        columns = [10, 47, 92, 154, 228, 364, 430, 504, 580, 655, 730, 827,
                   917, 991, 1066, 1142, 1230, 1322, 1428, 1532, 1599, 1658,
                   1726, 1815]
        rows = [15, 34] + [56 + 24 * index for index in range(42)]
        rows = [row for row in rows if row < 1054] + [1054]
        for row in rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (120, 120, 120), 1)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (120, 120, 120), 1)

        self.assertTrue(ocr_backend._has_photographic_background(image))
        self.assertTrue(
            ocr_backend._dense_grid_is_axis_aligned_screen_capture(
                image,
                columns,
                rows,
            )
        )

    def test_dense_grid_does_not_override_photo_with_perspective_row_spacing(self):
        ocr_backend._load_runtime()
        image = np.full((1068, 1857, 3), (155, 175, 195), dtype=np.uint8)
        columns = list(range(10, 1816, 90))
        rows = [15]
        gap = 12
        while rows[-1] < 1054:
            rows.append(min(1054, rows[-1] + gap))
            gap += 2

        self.assertFalse(
            ocr_backend._dense_grid_is_axis_aligned_screen_capture(
                image,
                columns,
                rows,
            )
        )

    def test_motion_deblurred_geometry_recovers_faint_vertical_rules(self):
        ocr_backend._load_runtime()
        image = np.full((720, 1000, 3), 244, dtype=np.uint8)
        columns = np.linspace(0, 999, 10).round().astype(int).tolist()
        rows = np.linspace(0, 719, 25).round().astype(int).tolist()
        for boundary in rows:
            cv2.line(image, (0, boundary), (999, boundary), (35, 35, 35), 1)
        for boundary in columns:
            cv2.line(image, (boundary, 0), (boundary, 719), (85, 85, 85), 1)
        kernel = np.zeros((3, 7), dtype=np.float32)
        kernel[1, :] = 1.0 / 7.0
        blurred = cv2.filter2D(image, -1, kernel)

        recovered = ocr_backend._recover_motion_blurred_photo_grid(
            blurred,
            None,
            maximum_cells=720,
        )

        self.assertIsNotNone(recovered)
        assert recovered is not None
        self.assertEqual(
            (len(recovered[1]) - 1, len(recovered[0]) - 1),
            (24, 9),
        )

    def test_dense_axis_aligned_grid_accepts_a_bounded_top_frame_margin(self):
        ocr_backend._load_runtime()
        image = np.full((1350, 2400, 3), 203, dtype=np.uint8)
        columns = [58 + 135 * index for index in range(18)] + [2342]
        rows = [56 + 29 * index for index in range(43)] + [1297]
        for row in rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (80, 80, 80), 1)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (80, 80, 80), 1)

        self.assertTrue(
            ocr_backend._dense_grid_is_axis_aligned_screen_capture(
                image,
                columns,
                rows,
            )
        )

    def test_screen_grid_preserves_single_top_title_as_full_width_span(self):
        grid = [["设备运行统计表", "", "", ""], ["编号", "名称", "频率", "状态"]]
        self.assertEqual(
            ocr_backend._title_spans_from_screen_grid(grid),
            [{"row": 0, "column": 0, "row_span": 1, "column_span": 4, "role": "title"}],
        )

    def test_strict_multilevel_layout_moves_title_to_full_width_anchor(self):
        grid = [
            ["", "", "项目任务跟踪表", "", "", "", ""],
            ["", "基础信息", "", "", "", "测量与判定", ""],
            ["任务编号", "任务名称", "优先级", "进度", "负责人", "计划完成", "当前状态"],
        ] + [
            [f"TASK-{index:03d}", "界面设计", "中", "7%", "王强", "2026-08-02", "已完成"]
            for index in range(1, 6)
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        spans = ocr_backend._strict_multilevel_title_span(grid, confidence)

        self.assertEqual(grid[0], ["项目任务跟踪表", "", "", "", "", "", ""])
        self.assertEqual(
            spans,
            [{"row": 0, "column": 0, "row_span": 1, "column_span": 7, "role": "title"}],
        )

    def test_strict_multilevel_layout_allows_period_in_title(self):
        grid = [
            ["", "", "BOM物料清单—2026-07批次", "", "", "", ""],
            ["", "基础信息", "", "", "", "测量与判定", ""],
            ["任务编号", "任务名称", "优先级", "进度", "负责人", "计划完成", "当前状态"],
        ] + [
            [f"TASK-{index:03d}", "界面设计", "中", "7%", "王强", "2026-08-02", "已完成"]
            for index in range(1, 6)
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        spans = ocr_backend._strict_multilevel_title_span(grid, confidence)

        self.assertEqual(grid[0], ["BOM物料清单 — 2026-07 批次", "", "", "", "", "", ""])
        self.assertEqual(
            spans,
            [{"row": 0, "column": 0, "row_span": 1, "column_span": 7, "role": "title"}],
        )

    def test_multilevel_header_requires_title_and_complete_group_spans(self):
        grid = [
            ["项目任务跟踪表", "", "", "", "", "", ""],
            ["基础信息", "", "", "测量与判定", "", "", ""],
            ["任务编号", "任务名称", "优先级", "进度", "负责人", "计划完成", "当前状态"],
        ]
        group_spans = [
            {"row": 1, "column": 0, "row_span": 1, "column_span": 3},
            {"row": 1, "column": 3, "row_span": 1, "column_span": 4},
        ]
        self.assertTrue(ocr_backend._has_unresolved_multilevel_header(grid, group_spans))
        complete = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 7},
            *group_spans,
        ]
        self.assertFalse(ocr_backend._has_unresolved_multilevel_header(grid, complete))

    def test_missing_recovery_locks_verified_merged_title_subordinate_cells(self):
        ocr_backend._load_runtime()
        image = np.full((90, 480, 3), 255, dtype=np.uint8)
        columns = [0, 80, 160, 240, 320, 400, 479]
        rows = [0, 30, 60, 89]
        for column in columns:
            cv2.line(image, (column, rows[1]), (column, rows[-1]), (30, 30, 30), 2)
        for row in rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (30, 30, 30), 2)
        grid = [
            ["产品测试记录", "", "", "", "", ""],
            ["序号", "产品编号", "测试项目", "技术要求", "实测值", "单位"],
            ["1", "SP-260801", "工作频率", "515.000±0.005", "515.002", "MHz"],
        ]

        locked = ocr_backend._locked_merged_blank_locations(
            image,
            columns,
            rows,
            grid,
        )

        self.assertEqual(locked, {(0, column) for column in range(1, 6)})

    def test_missing_recovery_does_not_lock_sparse_ordinary_first_row(self):
        ocr_backend._load_runtime()
        image = np.full((90, 480, 3), 255, dtype=np.uint8)
        columns = [0, 80, 160, 240, 320, 400, 479]
        rows = [0, 30, 60, 89]
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (30, 30, 30), 2)
        for row in rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (30, 30, 30), 2)
        grid = [
            ["批次A", "", "", "", "", ""],
            ["序号", "产品编号", "测试项目", "技术要求", "实测值", "单位"],
            ["1", "SP-260801", "工作频率", "515.000±0.005", "515.002", "MHz"],
        ]

        self.assertEqual(
            ocr_backend._locked_merged_blank_locations(image, columns, rows, grid),
            set(),
        )

    def test_missing_recovery_locks_each_physical_multilevel_header_span(self):
        ocr_backend._load_runtime()
        image = np.full((120, 560, 3), 255, dtype=np.uint8)
        columns = list(range(0, 561, 80))
        rows = [0, 30, 60, 90, 119]
        for row in rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (30, 30, 30), 2)
        for column in (columns[0], columns[3], columns[-1]):
            cv2.line(image, (column, rows[1]), (column, rows[2]), (30, 30, 30), 2)
        for column in columns:
            cv2.line(image, (column, rows[2]), (column, rows[-1]), (30, 30, 30), 2)
        grid = [
            ["生产日报", "", "", "", "", "", ""],
            ["", "基础信息", "", "", "", "测量与判定", ""],
            ["日期", "班次", "生产线", "计划数量", "完成数量", "不良数量", "完成率"],
            ["2026-08-02", "夜班", "B线", "280", "277", "3", "98.2%"],
        ]

        locked = ocr_backend._locked_merged_blank_locations(
            image,
            columns,
            rows,
            grid,
        )

        self.assertTrue({(1, 0), (1, 2)}.issubset(locked))
        self.assertTrue({(1, 3), (1, 4), (1, 6)}.issubset(locked))

    def test_missing_recovery_locks_merged_footer_subordinate_cells(self):
        ocr_backend._load_runtime()
        image = np.full((90, 400, 3), 255, dtype=np.uint8)
        columns = [0, 80, 160, 240, 320, 399]
        rows = [0, 30, 60, 89]
        for row in rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (30, 30, 30), 2)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[2]), (30, 30, 30), 2)
        cv2.line(image, (columns[0], rows[2]), (columns[0], rows[-1]), (30, 30, 30), 2)
        cv2.line(image, (columns[3], rows[2]), (columns[3], rows[-1]), (30, 30, 30), 2)
        cv2.line(image, (columns[-1], rows[2]), (columns[-1], rows[-1]), (30, 30, 30), 2)
        grid = [
            ["A", "B", "C", "D", "E"],
            ["1", "2", "3", "4", "5"],
            ["TOTAL INVENTORY VALUE:", "", "", "$9,442.41", ""],
        ]

        locked = ocr_backend._locked_merged_blank_locations(image, columns, rows, grid)

        self.assertTrue({(2, 1), (2, 2), (2, 4)}.issubset(locked))

    def test_screen_grid_moves_centered_title_text_to_merged_cell_anchor(self):
        grid = [["", "", "设备运行统计表", ""], ["编号", "名称", "频率", "状态"]]

        spans = ocr_backend._title_spans_from_screen_grid(grid)

        self.assertEqual(grid[0], ["设备运行统计表", "", "", ""])
        self.assertEqual(
            spans,
            [{"row": 0, "column": 0, "row_span": 1, "column_span": 4, "role": "title"}],
        )

    def test_title_anchor_survives_one_sparse_group_row_and_moves_confidence(self):
        grid = [
            ["", "", "生产日报", "", "", "", ""],
            ["基础信息", "", "", "", "", "", ""],
            ["日期", "班次", "生产线", "计划数量", "完成数量", "不良数量", "完成率"],
        ]
        confidence = [
            [0.0, 0.0, 0.93, 0.0, 0.0, 0.0, 0.0],
            [0.91, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0],
            [0.95] * 7,
        ]

        spans = ocr_backend._title_spans_from_screen_grid(grid, confidence)

        self.assertEqual(grid[0], ["生产日报", "", "", "", "", "", ""])
        self.assertEqual(confidence[0], [0.93, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0])
        self.assertEqual(spans[0]["column_span"], 7)

    def test_strict_simple_title_span_accepts_title_header_and_dense_body(self):
        grid = [
            ["库存出入库 2026年10月", "", "", ""],
            ["序号", "名称", "日期", "单位"],
            *[[str(row), "设备", f"2026-08-{row:02d}", "台"] for row in range(1, 7)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        spans = ocr_backend._strict_simple_title_span(grid, confidence)

        self.assertEqual(
            spans,
            [{"row": 0, "column": 0, "row_span": 1, "column_span": 4, "role": "title"}],
        )

    def test_strict_simple_title_span_rejects_nonsemantic_second_row(self):
        grid = [
            ["说明", "", "", ""],
            ["华东仓", "夜班", "A区", "待确认"],
            *[[str(row), f"AP-{row:03d}", "设备", str(100 + row)] for row in range(1, 7)],
        ]

        self.assertEqual(ocr_backend._strict_simple_title_span(grid, None), [])

    def test_table_engine_is_requested_only_at_structure_fallback(self):
        source = inspect.getsource(ocr_backend._recognize)

        self.assertIn("_engines()", source)
        self.assertIn("_engines(load_table=True)", source)
        self.assertLess(source.index("_engines()"), source.index("_engines(load_table=True)"))

    def test_spreadsheet_rulers_are_not_exported_as_table_data(self):
        grid = [
            ["", "A", "B", "C"],
            ["1", "编号", "名称", "状态"],
            ["2", "TASK-001", "需求确认", "进行中"],
            ["3", "TASK-002", "界面设计", "等待"],
        ]
        confidence = [[1.0] * 4 for _ in grid]
        trimmed, trimmed_confidence = ocr_backend._strip_spreadsheet_ui_headers(grid, confidence)
        self.assertEqual(trimmed, [["编号", "名称", "状态"], ["TASK-001", "需求确认", "进行中"], ["TASK-002", "界面设计", "等待"]])
        self.assertEqual(trimmed_confidence, [[1.0] * 3 for _ in trimmed])

    def test_complete_spreadsheet_spatial_grid_can_be_selected_after_both_rulers_are_removed(self):
        grid = [["", "A", "B", "C"], ["1", "编号", "名称", "状态"]] + [
            [str(row + 2), f"TASK-{row + 1:03d}", "设备巡检", "正常"]
            for row in range(8)
        ]
        confidence = [[1.0] * 4 for _ in grid]
        geometry = {
            "row_centers": [20.0 + row * 16.0 for row in range(len(grid))],
            "grouped_rows": [[] for _ in grid],
            "anchors": [20.0, 100.0, 200.0, 300.0],
            "first_structured_row": 0,
        }

        stripped_grid, _, stripped_spans, stripped_geometry = (
            ocr_backend._strip_spatial_ui_headers_for_recovery(
                grid,
                confidence,
                [],
                geometry,
            )
        )

        self.assertEqual((len(stripped_grid), len(stripped_grid[0])), (9, 3))
        self.assertTrue(
            ocr_backend._spreadsheet_ruler_spatial_grid_is_complete(
                stripped_grid,
                stripped_spans,
                True,
            )
        )
        self.assertEqual(
            tuple(
                len(values)
                for values in ocr_backend._spatial_geometry_cell_boundaries(
                    stripped_geometry,
                    (200, 400),
                    len(stripped_grid),
                    len(stripped_grid[0]),
                )
            ),
            (4, 10),
        )

    def test_complete_overlaid_column_ruler_preserves_content_columns_for_recovery(self):
        grid = [["A", "B", "C"], ["编号", "名称", "状态"]] + [
            [f"TASK-{row + 1:03d}", "设备巡检", "正常"]
            for row in range(8)
        ]
        confidence = [[1.0] * 3 for _ in grid]
        geometry = {
            "row_centers": [20.0 + row * 16.0 for row in range(len(grid))],
            "grouped_rows": [[] for _ in grid],
            "anchors": [40.0, 160.0, 280.0],
            "first_structured_row": 0,
        }

        stripped_grid, _, stripped_spans, stripped_geometry = (
            ocr_backend._strip_spatial_ui_headers_for_recovery(
                grid,
                confidence,
                [],
                geometry,
            )
        )

        self.assertEqual((len(stripped_grid), len(stripped_grid[0])), (9, 3))
        self.assertEqual(stripped_grid[0], ["编号", "名称", "状态"])
        self.assertTrue(
            ocr_backend._spreadsheet_ruler_spatial_grid_is_complete(
                stripped_grid,
                stripped_spans,
                True,
            )
        )
        self.assertIsNotNone(
            ocr_backend._spatial_geometry_cell_boundaries(
                stripped_geometry,
                (200, 400),
                len(stripped_grid),
                len(stripped_grid[0]),
            )
        )

    def test_spreadsheet_ruler_detection_ignores_unicode_superscript_digit(self):
        grid = [["²", "名称"], ["1", "设备A"], ["2", "设备B"]]
        confidence = [[1.0, 1.0] for _ in grid]

        trimmed, trimmed_confidence = ocr_backend._strip_spreadsheet_ui_headers(
            grid, confidence
        )

        self.assertEqual(trimmed, grid)
        self.assertEqual(trimmed_confidence, confidence)

    def test_spreadsheet_rulers_are_removed_when_last_letters_share_one_cell(self):
        grid = [
            ["", "A", "B", "C", "D", "E", "F", "G H"],
            ["1", "任务编号", "任务名称", "优先级", "进度", "负责人", "计划完成", "当前状态 风险说明"],
            ["2", "TASK-001", "OCR场景验证1", "中", "7%", "李娜", "2026-08-06", "进行中 无"],
            ["3", "TASK-002", "OCR场景验证2", "高", "14%", "王强", "2026-08-07", "等待 无"],
        ]
        confidence = [[1.0] * len(grid[0]) for _ in grid]

        trimmed, trimmed_confidence = ocr_backend._strip_spreadsheet_ui_headers(
            grid, confidence
        )

        self.assertEqual(trimmed[0][0], "任务编号")
        self.assertEqual(len(trimmed), 3)
        self.assertEqual(len(trimmed[0]), 7)
        self.assertEqual(len(trimmed_confidence), 3)

    def test_spreadsheet_rulers_expand_trailing_fused_letters_with_one_gap(self):
        grid = [
            ["A", "BC", "", "D", "", "FG", ""],
            ["序号", "名称", "型号", "编号", "数量", "单位", "状态"],
            ["1", "设备A", "M1", "ID1", "2", "件", "完成"],
            ["2", "设备B", "M2", "ID2", "3", "件", "完成"],
        ]

        trimmed, _ = ocr_backend._strip_spreadsheet_ui_headers(grid, None)

        self.assertEqual(trimmed, grid[1:])

    def test_spreadsheet_rulers_expand_single_fused_abc_corner(self):
        grid = [
            ["ABC", "", "", ""],
            ["1", "标题", "", ""],
            ["2", "序号", "编号", "名称"],
            ["3", "1", "ID1", "设备A"],
            ["4", "2", "ID2", "设备B"],
        ]

        trimmed, _ = ocr_backend._strip_spreadsheet_ui_headers(grid, None)

        self.assertEqual(trimmed, [["标题", "", ""], ["序号", "编号", "名称"], ["1", "ID1", "设备A"], ["2", "ID2", "设备B"]])

    def test_spreadsheet_rulers_strip_fused_ab_from_row_gutter(self):
        grid = [
            ["AB", "", "C", "D", "E", "F", ""],
            ["1", "", "", "校准检验记录 2026年09月", "", "", ""],
            ["2", "序号", "仪器编号", "校准项目", "标准值", "单位", "复核员"],
            ["3", "1", "AP-001", "温度", "10.0", "℃", "张三"],
            ["4", "2", "AP-002", "压力", "20.0", "kPa", "李四"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        trimmed, trimmed_confidence = ocr_backend._strip_spreadsheet_ui_headers(
            grid,
            confidence,
        )

        self.assertEqual((len(trimmed), len(trimmed[0])), (4, 6))
        self.assertEqual(trimmed[0][2], "校准检验记录 2026年09月")
        self.assertEqual(trimmed[1][0], "序号")
        self.assertEqual(len(trimmed_confidence), 4)

    def test_spreadsheet_rulers_are_removed_when_some_letters_are_unreadable(self):
        grid = [
            ["", "A", "B", "", "D", "", "F", "G", ""],
            ["1", "编号", "中心频率", "带宽", "调制方式", "信号名称", "功率", "采样率", "备注"],
            ["2", "1", "515.128 MHz", "9 kHz", "AM", "sig-1", "-10", "2.4 MS/s", "正常"],
            ["3", "2", "516.347 MHz", "35.4 kHz", "BPSK", "sig-2", "-30", "4.8 MS/s", "复核"],
        ]
        confidence = [[1.0] * len(grid[0]) for _ in grid]

        trimmed, trimmed_confidence = ocr_backend._strip_spreadsheet_ui_headers(
            grid, confidence
        )

        self.assertEqual(trimmed[0][0], "编号")
        self.assertEqual(len(trimmed), 3)
        self.assertEqual(len(trimmed[0]), 8)
        self.assertEqual(len(trimmed_confidence), 3)

    def test_spreadsheet_rulers_allow_one_missing_letter_after_a_fused_label_cell(self):
        grid = [
            ["", "A", "B", "C D", "E", "F", "G", "H", "J", "K", "L", "M"],
            ["1", "", "", "", "", "", "宽表 — 2026-05 批次", "", "", "", "", ""],
            ["2", "基础信息", "", "目标与测量", "", "", "", "过程记录", "", "", "质量判定", ""],
            ["3", "区域", "记录编号", "对象名称 长文本说明", "当前值", "目标值", "单位", "来源系统", "更新时间 责任人", "一级分类", "二级分类", "三级分类"],
            ["4", "A区", "R-001", "设备 说明", "1", "2", "V", "系统", "08:00 张三", "一级", "二级", "三级"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        trimmed, trimmed_confidence = ocr_backend._strip_spreadsheet_ui_headers(
            grid, confidence
        )

        self.assertEqual(len(trimmed), len(grid) - 1)
        self.assertEqual(len(trimmed[0]), len(grid[0]) - 1)
        self.assertEqual(trimmed[0][5], "宽表 — 2026-05 批次")
        self.assertEqual(len(trimmed_confidence), len(trimmed))

    def test_spreadsheet_row_ruler_keeps_title_fused_with_column_letter_band(self):
        grid = [
            ["", "A", "B", "C", "D", "E 审批勾选表单 — 2026-11 批次", "F", "G", "H", "", ""],
            ["2", "基础信息", "", "", "目标与测量", "", "", "过程记录", "", "质量判定", ""],
            ["3", "本期", "", "上限", "", "下限", "", "累计", "", "下限", ""],
            ["4", "序号", "申请编号", "申请事项", "申请部门", "申请人", "申请日期", "选项A", "选项B", "选项C", "是否紧急"],
            ["5", "1", "AP-001", "事项", "部门", "申请人", "日期", "是", "否", "否", "否"],
            ["6", "2", "AP-002", "事项", "部门", "申请人", "日期", "否", "是", "否", "是"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        trimmed, trimmed_confidence = ocr_backend._strip_spreadsheet_ui_headers(
            grid, confidence
        )

        self.assertEqual(len(trimmed), len(grid))
        self.assertEqual(len(trimmed[0]), len(grid[0]) - 1)
        self.assertEqual(trimmed[0][0], "审批勾选表单 — 2026-11 批次")
        self.assertTrue(all(not value for value in trimmed[0][1:]))
        self.assertEqual(trimmed_confidence[0][0], 0.77)

    def test_review_recovery_removes_fused_ruler_column_and_preserves_title(self):
        grid = [
            ["1", "A", "B", "C 审批勾选表单一2026-09批次", "D", "E", "F", ""],
            ["2", "", "基础信息", "", "目标与测量", "", "过程记录", ""],
            ["3", "序号", "申请编号", "申请事项", "申请部门", "申请人", "申请日期", "预算金额"],
            *[
                [str(row + 1), str(row - 2), f"AP-{row:03d}", "事项", "部门", "申请人", "日期", "100"]
                for row in range(3, 12)
            ],
        ]
        grid[10][0] = ""
        grid[10][1] = "11 8"
        confidence = [[0.99 if value else -1.0 for value in row] for row in grid]
        geometry = {
            "anchors": [float(index * 100 + 50) for index in range(8)],
            "row_centers": [float(index * 30 + 15) for index in range(12)],
        }

        recovered = ocr_backend._repair_fused_spreadsheet_ruler_title_for_review(
            grid,
            confidence,
            geometry,
        )

        self.assertIsNotNone(recovered)
        values, scores, adjusted_geometry = recovered
        self.assertEqual((len(values), len(values[0])), (12, 7))
        self.assertEqual(values[0][0], "审批勾选表单一2026-09批次")
        self.assertTrue(all(not value for value in values[0][1:]))
        self.assertEqual(values[10][0], "8")
        self.assertEqual(scores[10][0], 0.77)
        self.assertEqual(len(adjusted_geometry["anchors"]), 7)

    def test_standalone_row_ruler_requires_independent_business_ordinal(self):
        grid = [
            ["1", "序号", "名称", "数量"],
            *[
                [str(row + 1), str(row), f"设备{row}", str(row * 10)]
                for row in range(1, 12)
            ],
        ]
        confidence = [[0.99 for _ in row] for row in grid]

        self.assertTrue(
            ocr_backend._strip_standalone_spreadsheet_row_ruler_column(
                grid,
                confidence,
                [],
            )
        )
        self.assertEqual(grid[0], ["序号", "名称", "数量"])
        self.assertEqual(grid[-1][0], "11")
        self.assertTrue(all(value == 0.77 for row in confidence for value in row))

        rejected = [
            ["1", "序号", "名称", "数量"],
            *[
                [str(row + 1), str(row + 5), f"设备{row}", str(row * 10)]
                for row in range(1, 12)
            ],
        ]
        self.assertFalse(
            ocr_backend._strip_standalone_spreadsheet_row_ruler_column(
                rejected,
                [[0.99 for _ in row] for row in rejected],
                [],
            )
        )

        multilevel = [
            ["1", "", "基础信息", "", "目标与测量"],
            ["2", "下限", "", "实际", ""],
            ["3", "序号", "编号", "数值", "单位"],
            *[
                [str(row + 3), str(row), f"AP-{row:03d}", str(row * 10), "V"]
                for row in range(1, 11)
            ],
        ]
        multilevel_confidence = [[0.99 for _ in row] for row in multilevel]
        self.assertTrue(
            ocr_backend._strip_standalone_spreadsheet_row_ruler_column(
                multilevel,
                multilevel_confidence,
                [],
            )
        )
        self.assertEqual((len(multilevel), len(multilevel[0])), (13, 4))
        self.assertEqual(multilevel[2][0], "序号")

        absolute_multilevel = [
            ["1", "基础信息", "", "", "", "目标与测量", "", "", "", "过程记录", "", ""],
            ["2", "下限", "", "实际", "", "本期", "", "累计", "", "本期", "", "实际"],
            ["3", "登记号", "日期", "名称", "类别", "来源", "数量", "单位", "状态", "经办人", "联系方式", "地址"],
            *[
                [
                    str(row + 3),
                    f"AP-{row:03d}",
                    "2026-08-01",
                    f"设备{row}",
                    "类型A",
                    "系统A",
                    str(row * 10),
                    "V",
                    "正常",
                    "张三",
                    "13800000000",
                    "A区",
                ]
                for row in range(1, 11)
            ],
        ]
        absolute_confidence = [[0.99 for _ in row] for row in absolute_multilevel]
        self.assertTrue(
            ocr_backend._strip_standalone_spreadsheet_row_ruler_column(
                absolute_multilevel,
                absolute_confidence,
                [],
            )
        )
        self.assertEqual((len(absolute_multilevel), len(absolute_multilevel[0])), (13, 11))
        self.assertEqual(absolute_multilevel[2][0], "登记号")

        noisy_gutter = [
            ["", "序号", "名称", "数量"],
            *[
                [
                    "" if row % 3 == 0 else str(row + 1)[-1],
                    str(row),
                    f"设备{row}",
                    str(row * 10),
                ]
                for row in range(1, 12)
            ],
        ]
        noisy_confidence = [[0.99 for _ in row] for row in noisy_gutter]
        self.assertTrue(
            ocr_backend._strip_standalone_spreadsheet_row_ruler_column(
                noisy_gutter,
                noisy_confidence,
                [],
            )
        )
        self.assertEqual(noisy_gutter[0], ["序号", "名称", "数量"])
        self.assertEqual(noisy_gutter[-1][0], "11")

    def test_spreadsheet_ruler_after_blank_frame_tolerates_unknown_zero(self):
        grid = [
            ["", "", "", "", "", "", "", "", ""],
            ["A", "B", "", "0", "E", "F", "", "H", ""],
            ["", "", "工程参数配置 — 2026-12 批次", "", "", "", "", "", ""],
            ["", "", "基础信息", "", "", "目标与测量", "", "过程记录", ""],
            ["序号", "通道编号", "参数名称", "英文标识", "目标值", "下限", "上限", "单位", "采样率"],
            ["1", "AP-001", "电压", "CH-01", "1.000", "0.900", "1.100", "V", "100.00%"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        trimmed, trimmed_confidence = ocr_backend._strip_spreadsheet_ui_headers(
            grid, confidence
        )

        self.assertEqual(len(trimmed), len(grid) - 2)
        self.assertEqual(trimmed[0][2], "工程参数配置 — 2026-12 批次")
        self.assertEqual(len(trimmed_confidence), len(trimmed))

    def test_spreadsheet_rulers_allow_unknown_zero_in_corner_cell(self):
        grid = [
            ["0", "A", "B", "C", "D"],
            ["1", "基础信息", "", "目标与测量", ""],
            ["2", "序号", "名称", "目标值", "单位"],
            ["3", "1", "设备A", "10.0", "V"],
            ["4", "2", "设备B", "20.0", "V"],
        ]

        trimmed, _ = ocr_backend._strip_spreadsheet_ui_headers(grid, None)

        self.assertEqual(trimmed[0], ["基础信息", "", "目标与测量", ""])
        self.assertEqual(trimmed[-1], ["2", "设备B", "20.0", "V"])

    def test_spreadsheet_ruler_strips_empty_gutter_when_row_numbers_are_unreadable(self):
        grid = [
            ["", "A", "8", "C", "D"],
            ["", "基础信息", "", "目标与测量", ""],
            ["", "序号", "名称", "目标值", "单位"],
            ["", "1", "设备A", "10.0", "V"],
            ["", "2", "设备B", "20.0", "V"],
        ]

        trimmed, _ = ocr_backend._strip_spreadsheet_ui_headers(grid, None)

        self.assertEqual(trimmed[0], ["基础信息", "", "目标与测量", ""])
        self.assertEqual(trimmed[-1], ["2", "设备B", "20.0", "V"])

    def test_spreadsheet_row_ruler_accepts_missing_first_label_for_title(self):
        grid = [
            ["", "", "", "多级表头统计 — 2026-12 批次", "", ""],
            ["2", "基础信息", "", "目标与测量", "", ""],
            ["3", "计划", "", "实际", "", ""],
            ["4", "序号", "名称", "一月", "二月", "状态"],
            ["5", "1", "设备A", "10.0", "9.9", "正常"],
            ["6", "2", "设备B", "20.0", "20.1", "正常"],
        ]

        trimmed, _ = ocr_backend._strip_spreadsheet_ui_headers(grid, None)

        self.assertEqual(len(trimmed[0]), 5)
        self.assertEqual(trimmed[0][2], "多级表头统计 — 2026-12 批次")
        self.assertEqual(trimmed[3][0], "序号")

    def test_blank_first_row_without_spreadsheet_ruler_is_preserved(self):
        grid = [
            ["", "", ""],
            ["序号", "名称", "状态"],
            ["1", "设备A", "正常"],
            ["2", "设备B", "正常"],
            ["3", "设备C", "正常"],
        ]

        kept, _ = ocr_backend._strip_spreadsheet_ui_headers(grid, None)

        self.assertEqual(kept, grid)

    def test_certificate_alignment_accepts_blank_frame_plus_column_ruler(self):
        certificate = {"rows": 44, "columns": 9, "row_offset": 0, "column_offset": 0}

        aligned = ocr_backend._align_structure_certificate_after_ui_strip(
            certificate, (44, 9), (42, 9)
        )

        self.assertTrue(aligned)
        self.assertEqual(certificate["rows"], 42)
        self.assertEqual(certificate["row_offset"], 2)

    def test_spreadsheet_column_ruler_recovers_i_read_as_digit_one(self):
        grid = [
            ["A", "B", "C", "D", "E", "F", "G", "H", "1", "J", "K"],
            ["基础信息", "", "", "目标与测量", "", "", "过程记录", "", "质量判定", "", ""],
            ["区域", "记录编号", "对象名称", "当前值", "目标值", "单位", "更新时间", "责任人", "判定", "风险", "建议"],
            ["A区", "R-001", "设备", "1", "2", "V", "08:00", "张三", "合格", "无", "放行"],
        ]
        confidence = [[0.99] * len(grid[0]) for _ in grid]

        trimmed, trimmed_confidence = ocr_backend._strip_spreadsheet_ui_headers(
            grid, confidence
        )

        self.assertEqual(len(trimmed), 3)
        self.assertEqual(trimmed[0][0], "基础信息")
        self.assertEqual(len(trimmed_confidence), 3)

    def test_spreadsheet_column_ruler_recovers_i_read_as_chinese_one(self):
        grid = [
            ["A", "B", "C", "D", "E", "F", "G", "H", "一", "J", "K"],
            ["基础信息", "", "", "目标与测量", "", "", "过程记录", "", "质量判定", "", ""],
            ["区域", "记录编号", "对象名称", "当前值", "目标值", "单位", "更新时间", "责任人", "判定", "风险", "建议"],
            ["A区", "R-001", "设备", "1", "2", "V", "08:00", "张三", "合格", "无", "放行"],
        ]
        confidence = [[0.99] * len(grid[0]) for _ in grid]

        trimmed, trimmed_confidence = ocr_backend._strip_spreadsheet_ui_headers(
            grid, confidence
        )

        self.assertEqual(len(trimmed), 3)
        self.assertEqual(trimmed[0][0], "基础信息")
        self.assertEqual(len(trimmed_confidence), 3)

    def test_spreadsheet_rulers_recover_b_read_as_eight_with_fused_row_labels(self):
        grid = [
            ["", "A", "8", "C", "D", "E", "F", "G", "H", "1", "J", "K", "L", "M", "N", "o"],
            ["1", "标题", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["2 3", "基础信息 序号", "编号", "发生时间", "恢复时间", "等级", "设备", "模块", "故障描述", "原因分类", "处理措施", "停机/min", "负责人", "复核人", "备件编号", "关闭状态"],
            *[[str(row), *[f"值{row}-{column}" for column in range(15)]] for row in range(4, 10)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        trimmed, trimmed_confidence = ocr_backend._strip_spreadsheet_ui_headers(
            grid, confidence
        )

        self.assertEqual(len(trimmed), len(grid) - 1)
        self.assertTrue(all(len(row) == 15 for row in trimmed))
        self.assertEqual(trimmed[0][0], "标题")
        self.assertEqual(trimmed[1][0], "基础信息 序号")
        self.assertEqual(len(trimmed_confidence), len(trimmed))

    def test_spreadsheet_page_rulers_recover_one_faint_row_boundary(self):
        image = np.full((320, 900, 3), 245, dtype=np.uint8)
        columns = [0, 40, 140, 240, 340, 440, 540, 640, 740, 840]
        rows = [0, 24, 54, 114, 144, 174, 204, 234, 264, 294]
        cv2.line(image, (columns[1], 84), (columns[-1], 84), (210, 210, 210), 1)

        def box(center_x, center_y, width=12, height=14):
            return np.asarray(
                [
                    [center_x - width / 2, center_y - height / 2],
                    [center_x + width / 2, center_y - height / 2],
                    [center_x + width / 2, center_y + height / 2],
                    [center_x - width / 2, center_y + height / 2],
                ],
                dtype=np.float32,
            )

        texts = [chr(ord("A") + index) for index in range(8)]
        boxes = [
            box((columns[index + 1] + columns[index + 2]) / 2, 12)
            for index in range(8)
        ]
        row_centers = [39, 69, 99, 129, 159, 189, 219, 249]
        texts.extend(str(index) for index in range(1, 9))
        boxes.extend(box(20, center) for center in row_centers)

        with (
            patch.object(ocr_backend, "np", np),
            patch.object(ocr_backend, "cv2", cv2),
        ):
            recovered = ocr_backend._recover_missing_spreadsheet_row_boundary_from_page_rulers(
                image,
                columns,
                rows,
                boxes,
                texts,
                [0.99] * len(texts),
            )

        self.assertIsNotNone(recovered)
        self.assertEqual(len(recovered), len(rows) + 1)
        self.assertTrue(any(82 <= row <= 86 for row in recovered))

    def test_spreadsheet_row_boundary_recovery_requires_faint_physical_line(self):
        image = np.full((320, 900, 3), 245, dtype=np.uint8)
        columns = [0, 40, 140, 240, 340, 440, 540, 640, 740, 840]
        rows = [0, 24, 54, 114, 144, 174, 204, 234, 264, 294]

        def box(center_x, center_y):
            return np.asarray(
                [
                    [center_x - 6, center_y - 7],
                    [center_x + 6, center_y - 7],
                    [center_x + 6, center_y + 7],
                    [center_x - 6, center_y + 7],
                ],
                dtype=np.float32,
            )

        texts = [chr(ord("A") + index) for index in range(8)]
        boxes = [
            box((columns[index + 1] + columns[index + 2]) / 2, 12)
            for index in range(8)
        ]
        texts.extend(str(index) for index in range(1, 9))
        boxes.extend(box(20, center) for center in [39, 69, 99, 129, 159, 189, 219, 249])

        with (
            patch.object(ocr_backend, "np", np),
            patch.object(ocr_backend, "cv2", cv2),
        ):
            self.assertIsNone(
                ocr_backend._recover_missing_spreadsheet_row_boundary_from_page_rulers(
                    image,
                    columns,
                    rows,
                    boxes,
                    texts,
                    [0.99] * len(texts),
                )
            )

    def test_spreadsheet_row_boundary_recovery_requires_column_ruler(self):
        image = np.full((320, 900, 3), 245, dtype=np.uint8)
        columns = [0, 40, 140, 240, 340, 440, 540, 640, 740, 840]
        rows = [0, 24, 54, 114, 144, 174, 204, 234, 264, 294]
        cv2.line(image, (columns[1], 84), (columns[-1], 84), (210, 210, 210), 1)

        def box(center_y):
            return np.asarray(
                [[14, center_y - 7], [26, center_y - 7], [26, center_y + 7], [14, center_y + 7]],
                dtype=np.float32,
            )

        texts = [str(index) for index in range(1, 9)]
        boxes = [box(center) for center in [39, 69, 99, 129, 159, 189, 219, 249]]

        with (
            patch.object(ocr_backend, "np", np),
            patch.object(ocr_backend, "cv2", cv2),
        ):
            self.assertIsNone(
                ocr_backend._recover_missing_spreadsheet_row_boundary_from_page_rulers(
                    image,
                    columns,
                    rows,
                    boxes,
                    texts,
                    [0.99] * len(texts),
                )
            )

    def test_spreadsheet_ruler_count_expands_shared_last_letter_cell(self):
        grid = [
            ["", "A", "B", "C", "D", "E", "F", "G H"],
            ["1", "任务编号", "任务名称", "优先级", "进度", "负责人", "计划完成", "当前状态 风险说明"],
            ["2", "T-001", "验证", "中", "7%", "李娜", "2026-08-06", "进行中 无"],
            ["3", "T-002", "复核", "高", "14%", "王强", "2026-08-07", "等待 无"],
        ]

        self.assertEqual(ocr_backend._spreadsheet_ruler_expected_columns(grid), 8)
        grid[0] = ["", "A", "B", "C D", "E", "F", "G", "H", "J", "K", "L", "M"]
        self.assertEqual(ocr_backend._spreadsheet_ruler_expected_columns(grid), 13)
        grid[0][2] = "项目"
        self.assertEqual(ocr_backend._spreadsheet_ruler_expected_columns(grid), 0)

    def test_spatial_ruler_centers_recover_one_missing_photographed_column(self):
        centers = {
            "A": 111.5,
            "B": 267.5,
            "C": 448.5,
            "D": 607.5,
            "E": 740.0,
            "F": 872.5,
            "G": 1006.5,
            "H": 1139.0,
            "J": 1481.0,
            "K": 1612.5,
            "L": 1743.5,
            "M": 1875.5,
        }
        first_row = [
            {
                "text": label,
                "center_x": center,
                "score": 0.99,
            }
            for label, center in centers.items()
        ]
        geometry = {
            "anchors": [29.5, 118.5, 362.25, 532.25, 741.25, 876.75, 1011.0, 1143.25, 1397.75, 1613.5, 1747.75, 1880.75],
            "row_centers": [29.0, 67.0, 115.0, 159.0],
            "grouped_rows": [first_row, [], [], []],
        }
        grid = [
            ["", "A", "B", "C D", "E", "F", "G", "H", "J", "K", "L", "M"],
            ["1", "", "", "", "", "", "标题", "", "", "", "", ""],
            ["2", "基础信息", "", "目标与测量", "", "", "", "过程记录", "", "", "质量判定", ""],
            ["3", "区域", "编号", "对象 说明", "当前值", "目标值", "单位", "来源", "时间 责任人", "一级", "二级", "三级"],
        ]

        columns = ocr_backend._spreadsheet_columns_from_spatial_ruler(
            (1140, 1984),
            grid,
            geometry,
            expected_columns=13,
        )

        self.assertEqual(len(columns), 14)
        self.assertTrue(all(right > left for left, right in zip(columns, columns[1:])))
        self.assertLess(columns[0], 100)
        self.assertEqual(columns[-1], 1983)

    def test_spreadsheet_rulers_are_stripped_before_fused_rows_are_rejected(self):
        grid = [
            ["", "A", "B", "C", "D", "E", "F", "G H"],
            ["1", "编号", "中心频率", "带宽", "调制方式", "信号名称", "功率", "采样率 备注"],
            ["2 3", "1 2", "515.128 516.347", "9 35.4", "AM BPSK", "sig1 sig2", "-10 -30", "2.4 正常 4.8 相位π/4"],
            ["4", "3", "517.374", "57", "LFM", "lora_52.3k_62.5k", "-10", "8 S/N=18.6dB"],
        ]

        trimmed, _ = ocr_backend._strip_spreadsheet_ui_headers(grid, None)

        self.assertEqual(trimmed[0][0], "编号")
        self.assertTrue(ocr_backend._grid_has_concatenated_physical_rows(trimmed))

    def test_spreadsheet_row_ruler_is_removed_when_column_ruler_was_cropped(self):
        grid = [
            ["1", "项目里程碑计划", "", ""],
            ["2", "阶段", "计划开始", "状态"],
            ["3", "P1", "2026-08-02", "进行中"],
            ["4", "P2", "2026-08-03", "已完成"],
        ]
        confidence = [[0.99] * 4 for _ in grid]

        trimmed, trimmed_confidence = ocr_backend._strip_spreadsheet_ui_headers(
            grid, confidence
        )

        self.assertEqual(
            trimmed,
            [
                ["项目里程碑计划", "", ""],
                ["阶段", "计划开始", "状态"],
                ["P1", "2026-08-02", "进行中"],
                ["P2", "2026-08-03", "已完成"],
            ],
        )
        self.assertEqual(len(trimmed_confidence[0]), 3)

        missing_title = [list(row) for row in grid]
        missing_title[0][1] = ""
        missing_title[1][0] = ""
        trimmed_missing_title, _ = ocr_backend._strip_spreadsheet_ui_headers(
            missing_title, confidence
        )
        self.assertEqual(len(trimmed_missing_title[0]), 3)

    def test_spatial_ruler_keeps_group_headers_overlaid_by_column_letters(self):
        grid = [
            ["1", "A", "基础信息 B", "C", "D 数量与金额", "E", "F"],
            ["2", "序号", "对账单号", "客户", "费用类型", "摘要", "数量"],
            *[
                [str(index + 2), str(index), f"AP-{index:03d}", "客户", "费用", "摘要", "10"]
                for index in range(1, 8)
            ],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        recovered = ocr_backend._strip_overlaid_spreadsheet_rulers(
            grid,
            confidence,
        )

        self.assertIsNotNone(recovered)
        trimmed, trimmed_confidence = recovered
        self.assertEqual((len(trimmed), len(trimmed[0])), (9, 6))
        self.assertEqual(trimmed[0], ["", "基础信息", "", "数量与金额", "", ""])
        self.assertEqual(trimmed[1][0], "序号")
        self.assertEqual(trimmed_confidence[0][0], 0.0)

        invalid = [list(row) for row in grid]
        invalid[0][3] = "G"
        self.assertIsNone(
            ocr_backend._strip_overlaid_spreadsheet_rulers(invalid, confidence)
        )

    def test_review_only_spans_never_hide_non_anchor_text(self):
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 2},
            {"row": 0, "column": 2, "row_span": 1, "column_span": 4},
        ]
        safe = [
            ["基础信息", "", "数量与金额", "", "", ""],
            ["序号", "对账单号", "客户", "费用类型", "摘要", "数量"],
        ]
        unsafe = [list(row) for row in safe]
        unsafe[0][1] = "不可隐藏"

        self.assertFalse(ocr_backend._spans_hide_non_anchor_text(safe, spans))
        self.assertTrue(ocr_backend._spans_hide_non_anchor_text(unsafe, spans))

    def test_spreadsheet_row_ruler_is_removed_before_dense_header(self):
        grid = [
            ["1", "序号", "仪器编号", "校准项目", "标准值", "单位"],
            ["2", "1", "AP-001", "电压", "220", "V"],
            ["3", "2", "AP-002", "电流", "5", "A"],
            ["4", "3", "AP-003", "温度", "20", "℃"],
        ]
        trimmed, _ = ocr_backend._strip_spreadsheet_ui_headers(grid, None)
        self.assertEqual(trimmed[0], ["序号", "仪器编号", "校准项目", "标准值", "单位"])

    def test_spreadsheet_row_ruler_is_removed_before_sparse_group_header(self):
        grid = [
            ["1", "基础信息", "", "目标与测量", "", "质量判定", ""],
            ["2", "序号", "名称", "目标值", "实测值", "判定", "备注"],
            ["3", "1", "设备A", "10", "9.9", "合格", ""],
            ["4", "2", "设备B", "20", "20.1", "合格", ""],
        ]
        trimmed, _ = ocr_backend._strip_spreadsheet_ui_headers(grid, None)
        self.assertEqual(trimmed[0][0], "基础信息")
        self.assertEqual(len(trimmed[0]), 6)

    def test_spreadsheet_row_ruler_is_removed_before_education_headers(self):
        grid = [
            ["1", "基础信息", "", "", "", "数量与金额", "", "过程记录"],
            ["2", "序号", "学号", "姓名", "班级", "课程", "任课教师", "星期"],
            *[
                [str(index + 2), str(index), "B区", "周工", "二车间", "课程", "教师", "星期一"]
                for index in range(1, 7)
            ],
        ]

        trimmed, _ = ocr_backend._strip_spreadsheet_ui_headers(grid, None)

        self.assertEqual((len(trimmed), len(trimmed[0])), (8, 7))
        self.assertEqual(trimmed[0][0], "基础信息")
        self.assertEqual(trimmed[1], ["序号", "学号", "姓名", "班级", "课程", "任课教师", "星期"])

    def test_legitimate_serial_number_column_is_not_removed(self):
        grid = [
            ["设备清单", "", ""],
            ["序号", "名称", "状态"],
            ["1", "频谱仪", "正常"],
            ["2", "信号源", "正常"],
        ]
        confidence = [[0.99] * 3 for _ in grid]

        kept, _ = ocr_backend._strip_spreadsheet_ui_headers(grid, confidence)

        self.assertEqual(kept, grid)

    def test_empty_page_margin_cells_are_trimmed_before_ruler_detection(self):
        grid = [
            ["", "", "", "", ""],
            ["", "1", "设备清单", "", ""],
            ["", "2", "编号", "名称", ""],
            ["", "3", "A001", "频谱仪", ""],
            ["", "", "", "", ""],
        ]
        confidence = [[0.0] * 5 for _ in grid]

        trimmed, trimmed_confidence = ocr_backend._trim_empty_outer_grid(
            grid, confidence
        )

        self.assertEqual(
            trimmed,
            [["1", "设备清单", ""], ["2", "编号", "名称"], ["3", "A001", "频谱仪"]],
        )
        self.assertEqual(len(trimmed_confidence), 3)
        self.assertEqual(len(trimmed_confidence[0]), 3)

    def test_certified_empty_paper_frame_updates_boundary_offsets(self):
        grid = [
            ["", "", "", ""],
            ["", "编号", "名称", ""],
            ["", "1", "设备", ""],
            ["", "", "", ""],
        ]
        confidence = [[0.0] * 4 for _ in grid]
        image = np.full((40, 40, 3), 255, dtype=np.uint8)
        certificate = ocr_backend._new_structure_certificate(
            image,
            [0, 10, 20, 30, 39],
            [0, 10, 20, 30, 39],
            "photographic_ruled_grid",
        )

        trimmed, trimmed_confidence, aligned = (
            ocr_backend._trim_certified_empty_outer_grid(
                grid, confidence, certificate
            )
        )

        self.assertTrue(aligned)
        self.assertEqual(trimmed, [["编号", "名称"], ["1", "设备"]])
        self.assertEqual(len(trimmed_confidence), 2)
        self.assertEqual(certificate["row_offset"], 0)
        self.assertEqual(certificate["column_offset"], 0)
        self.assertEqual(certificate["rows"], 2)
        self.assertEqual(certificate["columns"], 2)
        self.assertEqual(certificate["row_boundaries"], [10, 20, 30])
        self.assertEqual(certificate["column_boundaries"], [10, 20, 30])
        self.assertTrue(
            ocr_backend._structure_certificate_matches(
                certificate, trimmed, []
            )
        )

    def test_edge_recovered_certificate_preserves_empty_outer_cells(self):
        grid = [["", ""], ["编号", "名称"], ["1", "设备"], ["", ""]]
        confidence = [[0.0] * 2 for _ in grid]
        image = np.full((40, 20, 3), 255, dtype=np.uint8)
        certificate = ocr_backend._new_structure_certificate(
            image,
            [0, 10, 19],
            [0, 10, 20, 30, 39],
            "photographic_ruled_grid",
        )
        certificate["preserve_empty_outer_grid"] = True

        kept, kept_confidence, aligned = (
            ocr_backend._trim_certified_empty_outer_grid(
                grid, confidence, certificate
            )
        )

        self.assertTrue(aligned)
        self.assertEqual(kept, grid)
        self.assertEqual(kept_confidence, confidence)
        self.assertEqual(certificate["rows"], 4)

    def test_edge_recovered_certificate_trims_only_a_proven_frame_before_ruler(self):
        grid = [
            ["", "", ""],
            ["A", "B", "C"],
            ["设备清单", "", ""],
            ["编号", "名称", "状态"],
            ["1", "设备", "正常"],
        ]
        confidence = [[0.0] * 3 for _ in grid]
        image = np.full((50, 30, 3), 255, dtype=np.uint8)
        certificate = ocr_backend._new_structure_certificate(
            image,
            [0, 10, 20, 29],
            [0, 10, 20, 30, 40, 49],
            "photographic_ruled_grid",
        )
        certificate["preserve_empty_outer_grid"] = True

        trimmed, _, aligned = ocr_backend._trim_certified_empty_outer_grid(
            grid, confidence, certificate
        )

        self.assertTrue(aligned)
        self.assertEqual(trimmed, grid[1:])
        self.assertEqual(certificate["rows"], 4)
        self.assertEqual(certificate["row_boundaries"], [10, 20, 30, 40, 49])

    def test_edge_recovered_certificate_keeps_page_title_in_blank_cell_row(self):
        grid = [
            ["", "", ""],
            ["A", "B", "C"],
            ["基础信息", "", "业务数据"],
            ["编号", "名称", "状态"],
            ["1", "设备", "正常"],
        ]
        confidence = [[0.0] * 3 for _ in grid]
        image = np.full((50, 30, 3), 255, dtype=np.uint8)
        certificate = ocr_backend._new_structure_certificate(
            image,
            [0, 10, 20, 29],
            [0, 10, 20, 30, 40, 49],
            "photographic_ruled_grid",
        )
        certificate["preserve_empty_outer_grid"] = True

        kept, kept_confidence, aligned = (
            ocr_backend._trim_certified_empty_outer_grid(
                grid,
                confidence,
                certificate,
                [("仓库出入库记录表", 15.0, 5.0, 0.99)],
            )
        )

        self.assertTrue(aligned)
        self.assertEqual(kept, grid)
        self.assertEqual(kept_confidence, confidence)
        self.assertEqual(certificate["rows"], 5)
        self.assertEqual(certificate["row_boundaries"], [0, 10, 20, 30, 40, 49])

    def test_adjacent_physical_spill_candidates_include_fused_and_visible_blank(self):
        image = np.full((60, 120, 3), 245, dtype=np.uint8)
        cv2.putText(
            image,
            "AB",
            (64, 43),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.65,
            (20, 20, 20),
            2,
            cv2.LINE_AA,
        )
        grid = [
            ["学", "学号姓名", ""],
            ["AP-01 陈晨", "", "正常"],
        ]

        with patch.object(ocr_backend, "cv2", cv2), patch.object(
            ocr_backend, "np", np
        ):
            selected = ocr_backend._adjacent_physical_spill_candidates(
                image,
                grid,
                [0, 40, 80, 120],
                [0, 30, 60],
            )

        self.assertTrue({(0, 0), (0, 1)}.issubset(selected))
        self.assertTrue({(1, 0), (1, 1)}.issubset(selected))

    def test_adjacent_physical_spill_candidates_ignore_clean_blank_neighbor(self):
        image = np.full((30, 120, 3), 245, dtype=np.uint8)
        grid = [["备注内容", "", "正常"]]

        with patch.object(ocr_backend, "cv2", cv2), patch.object(
            ocr_backend, "np", np
        ):
            selected = ocr_backend._adjacent_physical_spill_candidates(
                image,
                grid,
                [0, 40, 80, 120],
                [0, 30],
            )

        self.assertEqual(selected, set())

    def test_adjacent_duplicate_suffix_spill_keeps_numeric_prefix(self):
        grid = [["1273李娜", "李娜", "完成"]]
        confidence = [[0.98, 0.99, 0.99]]

        repaired = ocr_backend._trim_adjacent_duplicate_suffix_spills(
            grid,
            confidence,
        )

        self.assertEqual(repaired, {(0, 0)})
        self.assertEqual(grid, [["1273", "李娜", "完成"]])
        self.assertEqual(confidence[0][0], 0.77)

        spaced_grid = [["146****2001 无", "无"]]
        spaced_confidence = [[0.99, 0.99]]
        self.assertEqual(
            ocr_backend._trim_adjacent_duplicate_suffix_spills(
                spaced_grid,
                spaced_confidence,
            ),
            {(0, 0)},
        )
        self.assertEqual(spaced_grid[0][0], "146****2001")

    def test_patterned_adjacent_spills_split_headers_names_and_status(self):
        grid = [
            ["学", "学号姓名", "上期读数本期读数", ""],
            ["", "赵敏 3547", "155 无", ""],
            ["AP-M2017-69 陈晨", "", "完成", ""],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        repaired = ocr_backend._split_patterned_adjacent_spills(
            grid,
            confidence,
        )

        self.assertEqual(grid[0], ["学号", "姓名", "上期读数", "本期读数"])
        self.assertEqual(grid[1], ["赵敏", "3547", "155", "无"])
        self.assertEqual(grid[2], ["AP-M2017-69", "陈晨", "完成", ""])
        self.assertEqual(
            repaired,
            {
                (0, 0), (0, 1), (0, 2), (0, 3),
                (1, 0), (1, 1), (1, 2), (1, 3),
                (2, 0), (2, 1),
            },
        )

        trailing_blank_grid = [["", "单价金额", ""]]
        trailing_blank_confidence = [[0.0, 0.99, 0.0]]
        ocr_backend._split_patterned_adjacent_spills(
            trailing_blank_grid,
            trailing_blank_confidence,
        )
        self.assertEqual(trailing_blank_grid[0], ["", "单价", "金额"])

        repeated_header_grid = [["检查项目标准值", "检查项目标准值"]]
        repeated_header_confidence = [[0.99, 0.99]]
        self.assertEqual(
            ocr_backend._split_patterned_adjacent_spills(
                repeated_header_grid,
                repeated_header_confidence,
            ),
            {(0, 0), (0, 1)},
        )
        self.assertEqual(repeated_header_grid[0], ["检查项目", "标准值"])

        ocr_backend._load_runtime()
        physical_grid = [
            ["姓名", "位置"],
            ["李娜", "A区"],
            ["王强", "一车间"],
            ["李娜 A区", ""],
            ["李娜", "A区"],
        ]
        physical_confidence = [
            [0.99 if value else 0.0 for value in row]
            for row in physical_grid
        ]
        physical_image = np.full((250, 300, 3), 245, dtype=np.uint8)
        cv2.putText(
            physical_image,
            "L",
            (45, 190),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.8,
            (20, 20, 20),
            2,
        )
        cv2.putText(
            physical_image,
            "R",
            (195, 190),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.8,
            (20, 20, 20),
            2,
        )
        self.assertEqual(
            ocr_backend._split_physically_separated_adjacent_tokens(
                physical_image,
                physical_grid,
                physical_confidence,
                [0, 150, 300],
                [0, 50, 100, 150, 200, 250],
            ),
            {(3, 0), (3, 1)},
        )
        self.assertEqual(physical_grid[3], ["李娜", "A区"])

        phone_grid = [
            ["编号", "联系电话"],
            ["1", "138****1000"],
            ["2", "139****1001"],
            ["3", "137****1002"],
            ["4", "180***8661"],
        ]
        phone_confidence = [[0.99, 0.99] for _ in phone_grid]
        self.assertEqual(
            ocr_backend._normalize_masked_phone_columns(
                phone_grid,
                phone_confidence,
            ),
            {(4, 1)},
        )
        self.assertEqual(phone_grid[4][1], "180****8661")

        plate_grid = [["渝A•44539", "普通•项目"]]
        plate_confidence = [[0.99, 0.99]]
        self.assertEqual(
            ocr_backend._normalize_license_plate_separators(
                plate_grid,
                plate_confidence,
            ),
            {(0, 0)},
        )
        self.assertEqual(plate_grid, [["渝A·44539", "普通•项目"]])

        hyphen_plate_grid = [
            ["车牌号", "任务编号"],
            ["渝A-44539", "渝A-44539"],
            ["渝B–10293", "普通-项目"],
        ]
        hyphen_plate_confidence = [[0.99, 0.99] for _ in hyphen_plate_grid]
        self.assertEqual(
            ocr_backend._normalize_license_plate_separators(
                hyphen_plate_grid,
                hyphen_plate_confidence,
            ),
            {(1, 0), (2, 0)},
        )
        self.assertEqual(hyphen_plate_grid[1], ["渝A·44539", "渝A-44539"])
        self.assertEqual(hyphen_plate_grid[2], ["渝B·10293", "普通-项目"])

        group_grid = [["基础信息  业务数据", "基础信息", "", "", "业务数据", "", ""]]
        group_confidence = [[0.77 if value else 0.0 for value in group_grid[0]]]
        group_spans = []
        self.assertEqual(
            ocr_backend._repair_duplicate_combined_group_headers(
                group_grid,
                group_confidence,
                group_spans,
            ),
            {(0, 0), (0, 3)},
        )
        self.assertEqual(group_grid[0], ["基础信息", "", "", "业务数据", "", "", ""])

        flow_grid = [
            ["编号", "姓名", "分类"],
            ["8", "赵敏", "A区"],
            ["9 10", "李娜", "常规项目"],
            ["11", "陈晨 张伟", "华东库 批次-L08"],
            ["12", "王强", "华东库"],
            ["13", "", ""],
            ["", "陈晨", "现场复核"],
            ["14", "周林", "A区"],
            ["15", "王强", "一车间"],
            ["16", "李娜", "批次-L08"],
            ["17", "陈晨", "华东库"],
            ["18", "张伟", "批次-L08"],
        ]
        flow_confidence = [[0.99 if value else 0.0 for value in row] for row in flow_grid]
        repaired_flow = ocr_backend._repair_sequential_row_token_flow(
            flow_grid,
            flow_confidence,
        )
        self.assertTrue(repaired_flow)
        self.assertEqual(
            [flow_grid[row][0] for row in range(2, 7)],
            ["9", "10", "11", "12", "13"],
        )
        self.assertEqual(
            [flow_grid[row][1] for row in range(2, 7)],
            ["李娜", "陈晨", "张伟", "王强", "陈晨"],
        )

    def test_adjacent_duplicate_suffix_spill_does_not_trim_textual_prefix(self):
        grid = [["北京市", "市政", "完成"]]
        confidence = [[0.98, 0.99, 0.99]]

        repaired = ocr_backend._trim_adjacent_duplicate_suffix_spills(
            grid,
            confidence,
        )

        self.assertEqual(repaired, set())
        self.assertEqual(grid[0][0], "北京市")

    def test_edge_recovered_certificate_trims_short_empty_band_after_ui_strip(self):
        grid = [
            ["标题", "", ""],
            ["编号", "名称", "状态"],
            ["1", "设备A", "正常"],
            ["2", "设备B", "正常"],
            ["3", "设备C", "正常"],
            ["4", "设备D", "正常"],
            ["5", "设备E", "正常"],
            ["", "", ""],
        ]
        confidence = [[0.99] * 3 for _ in grid]
        image = np.full((240, 30, 3), 255, dtype=np.uint8)
        certificate = ocr_backend._new_structure_certificate(
            image,
            [0, 10, 20, 29],
            [0, 10, 20, 50, 80, 110, 140, 170, 200, 230, 239],
            "photographic_ruled_grid",
        )
        certificate["preserve_empty_outer_grid"] = True
        certificate["row_offset"] = 2
        certificate["rows"] = len(grid)
        certificate["ui_headers_processed"] = True

        trimmed, trimmed_confidence, aligned = (
            ocr_backend._trim_certified_empty_outer_grid(
                grid, confidence, certificate
            )
        )

        self.assertTrue(aligned)
        self.assertEqual(trimmed, grid[:-1])
        self.assertEqual(trimmed_confidence, confidence[:-1])
        self.assertEqual(certificate["rows"], 7)
        self.assertEqual(
            certificate["row_boundaries"],
            [0, 10, 20, 50, 80, 110, 140, 170, 200, 230],
        )
        self.assertTrue(
            ocr_backend._structure_certificate_matches(
                certificate, trimmed, []
            )
        )

    def test_edge_recovered_certificate_keeps_normal_height_empty_band_after_ui_strip(self):
        grid = [
            ["标题", "", ""],
            ["编号", "名称", "状态"],
            ["1", "设备A", "正常"],
            ["2", "设备B", "正常"],
            ["3", "设备C", "正常"],
            ["4", "设备D", "正常"],
            ["5", "设备E", "正常"],
            ["", "", ""],
        ]
        confidence = [[0.99] * 3 for _ in grid]
        image = np.full((261, 30, 3), 255, dtype=np.uint8)
        certificate = ocr_backend._new_structure_certificate(
            image,
            [0, 10, 20, 29],
            [0, 10, 20, 50, 80, 110, 140, 170, 200, 230, 260],
            "photographic_ruled_grid",
        )
        certificate["preserve_empty_outer_grid"] = True
        certificate["row_offset"] = 2
        certificate["rows"] = len(grid)
        certificate["ui_headers_processed"] = True

        kept, kept_confidence, aligned = (
            ocr_backend._trim_certified_empty_outer_grid(
                grid, confidence, certificate
            )
        )

        self.assertTrue(aligned)
        self.assertEqual(kept, grid)
        self.assertEqual(kept_confidence, confidence)
        self.assertEqual(certificate["rows"], 8)

    def test_visible_double_row_boundary_requires_a_wide_real_transition(self):
        image = np.full((130, 240, 3), 245, dtype=np.uint8)
        rows = [0, 10, 20, 30, 50, 60, 70, 80, 90, 100, 110, 120]
        cv2.line(image, (0, 40), (239, 40), (225, 225, 225), 2)

        recovered = table_pipeline._recover_visible_double_row_boundary(
            image, rows
        )

        self.assertEqual(len(recovered), len(rows) + 1)
        self.assertLessEqual(abs(recovered[4] - 40), 2)
        self.assertEqual(
            table_pipeline._recover_visible_double_row_boundary(
                np.full_like(image, 245), rows
            ),
            rows,
        )

    def test_weak_duplicate_column_boundary_is_collapsed(self):
        vertical = np.zeros((220, 465), dtype=np.uint8)
        for column in [0, 100, 241, 364, 464]:
            vertical[:, max(0, column - 1) : min(465, column + 2)] = 255
        columns = [0, 100, 227, 241, 364, 464]

        collapsed = table_pipeline._remove_weak_split_columns(columns, vertical)

        self.assertEqual(collapsed, [0, 100, 241, 364, 464])

    def test_narrow_column_with_two_supported_rules_is_preserved(self):
        vertical = np.zeros((220, 465), dtype=np.uint8)
        columns = [0, 100, 227, 241, 364, 464]
        for column in columns:
            vertical[:, max(0, column - 1) : min(465, column + 2)] = 255

        kept = table_pipeline._remove_weak_split_columns(columns, vertical)

        self.assertEqual(kept, columns)

    def test_partial_multi_level_header_rules_are_recovered(self):
        horizontal = np.zeros((430, 600), dtype=np.uint8)
        rows = [0, 20, 60, 180, *range(210, 421, 30)]
        for row in rows:
            horizontal[max(0, row - 1) : min(430, row + 2), :] = 255
        horizontal[99:102, 100:280] = 255
        horizontal[139:142, 260:560] = 255

        recovered = table_pipeline._recover_partial_header_row_boundaries(
            rows, horizontal, 600
        )

        self.assertEqual(recovered[:6], [0, 20, 60, 100, 140, 180])

    def test_short_header_text_strokes_do_not_create_row_boundaries(self):
        horizontal = np.zeros((430, 600), dtype=np.uint8)
        rows = [0, 20, 60, 180, *range(210, 421, 30)]
        for row in rows:
            horizontal[max(0, row - 1) : min(430, row + 2), :] = 255
        horizontal[99:102, 100:140] = 255
        horizontal[139:142, 300:340] = 255

        kept = table_pipeline._recover_partial_header_row_boundaries(
            rows, horizontal, 600
        )

        self.assertEqual(kept, rows)

    def test_faint_broad_tone_header_divider_is_recovered(self):
        horizontal = np.zeros((430, 600), dtype=np.uint8)
        image = np.full((430, 600, 3), 230, dtype=np.uint8)
        rows = [0, 20, 60, 134, *range(164, 405, 30)]
        for row in rows:
            horizontal[max(0, row - 1) : min(430, row + 2), :] = 255
        image[97:100, :, :] = 224

        recovered = table_pipeline._recover_partial_header_row_boundaries(
            rows, horizontal, 600, image
        )

        self.assertEqual(recovered[:5], [0, 20, 60, 98, 134])

    def test_faint_spreadsheet_row_number_gutter_boundary_is_recovered(self):
        vertical = np.zeros((1320, 1155), dtype=np.uint8)
        columns = [0, 153, 309, 466, 579, 693, 807, 921, 1034, 1152]
        rows = [1, 24, 64, 104, 139, 176, *range(207, 1315, 31)]
        for column in columns:
            vertical[:, max(0, column - 1) : min(1155, column + 2)] = 255
        vertical[70:1315, 39:42] = 255

        recovered = table_pipeline._recover_leading_spreadsheet_gutter_boundary(
            columns, rows, vertical
        )

        self.assertEqual(recovered[:3], [0, 40, 153])

    def test_weak_leading_text_stroke_does_not_create_spreadsheet_gutter(self):
        vertical = np.zeros((1320, 1155), dtype=np.uint8)
        columns = [0, 153, 309, 466, 579, 693, 807, 921, 1034, 1152]
        rows = [1, 24, 64, 104, 139, 176, *range(207, 1315, 31)]
        for column in columns:
            vertical[:, max(0, column - 1) : min(1155, column + 2)] = 255
        vertical[300:390, 39:42] = 255

        kept = table_pipeline._recover_leading_spreadsheet_gutter_boundary(
            columns, rows, vertical
        )

        self.assertEqual(kept, columns)

    def test_spreadsheet_gutter_and_regular_leading_columns_are_recovered(self):
        vertical = np.zeros((1310, 1077), dtype=np.uint8)
        columns = [0, 344, 447, 550, 654, 757, 861, 965, 1076]
        rows = [2, 26, 61, 134, *range(164, 1307, 29)]
        for column in columns:
            vertical[:, max(0, column - 1) : min(1077, column + 2)] = 255
        vertical[240:1307, 33:36] = 255
        vertical[100:1307, 136:139] = 255
        vertical[100:1307, 239:242] = 255

        recovered = table_pipeline._recover_leading_spreadsheet_gutter_boundary(
            columns, rows, vertical
        )

        self.assertEqual(recovered[:5], [0, 34, 137, 240, 344])

    def test_certified_screen_grid_trims_a_fully_empty_leading_row(self):
        grid = [
            ["", "", ""],
            ["编号", "名称", "状态"],
            ["1", "设备", "正常"],
        ]
        confidence = [[0.0] * 3 for _ in grid]
        image = np.full((30, 30, 3), 255, dtype=np.uint8)
        certificate = ocr_backend._new_structure_certificate(
            image,
            [0, 10, 20, 29],
            [0, 10, 20, 29],
            "screen_ruled_grid",
        )

        trimmed, trimmed_confidence, aligned = (
            ocr_backend._trim_certified_empty_outer_grid(
                grid, confidence, certificate
            )
        )

        self.assertTrue(aligned)
        self.assertEqual(
            trimmed,
            [["编号", "名称", "状态"], ["1", "设备", "正常"]],
        )
        self.assertEqual(len(trimmed_confidence), 2)
        self.assertEqual(certificate["rows"], 2)
        self.assertEqual(certificate["row_boundaries"], [10, 20, 29])
        self.assertTrue(
            ocr_backend._structure_certificate_matches(
                certificate, trimmed, []
            )
        )

    def test_unruled_empty_trailing_margin_requires_rules_to_end(self):
        ocr_backend._load_runtime()
        columns = [0, 50, 100, 150, 199]
        rows = [0, 20, 40, 60, 80, 99]
        grid = [
            ["标题", "", "", ""],
            ["序号", "名称", "数量", "状态"],
            ["1", "设备1", "10", "正常"],
            ["2", "设备2", "20", "正常"],
            ["", "", "", ""],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        image = np.full((100, 200, 3), 255, dtype=np.uint8)
        for boundary in rows:
            cv2.line(image, (0, boundary), (199, boundary), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (boundary, 0), (boundary, 80), (0, 0, 0), 2)

        self.assertTrue(
            ocr_backend._trim_unruled_empty_trailing_margin_row(
                image,
                grid,
                confidence,
                columns,
                rows,
                [],
            )
        )
        self.assertEqual((len(grid), len(rows)), (4, 5))

        ruled_image = image.copy()
        for boundary in columns:
            cv2.line(ruled_image, (boundary, 80), (boundary, 99), (0, 0, 0), 2)
        ruled_grid = [row[:] for row in grid] + [["", "", "", ""]]
        ruled_confidence = [[0.99 if value else 0.0 for value in row] for row in ruled_grid]
        ruled_rows = [0, 20, 40, 60, 80, 99]
        self.assertFalse(
            ocr_backend._trim_unruled_empty_trailing_margin_row(
                ruled_image,
                ruled_grid,
                ruled_confidence,
                columns,
                ruled_rows,
                [],
            )
        )

    def test_sparse_rule_artifacts_in_page_margin_are_trimmed(self):
        grid = [
            ["-", "", "", "", "-"],
            ["", "编号", "名称", "状态", ""],
            ["", "1", "频谱仪", "正常", "-"],
            ["-", "2", "信号源", "正常", ""],
            ["", "3", "功率计", "正常", ""],
            ["-", "", "", "", ""],
        ]
        confidence = [[0.86] * 5 for _ in grid]

        trimmed, _ = ocr_backend._trim_empty_outer_grid(grid, confidence)

        self.assertEqual(
            trimmed,
            [
                ["编号", "名称", "状态"],
                ["1", "频谱仪", "正常"],
                ["2", "信号源", "正常"],
                ["3", "功率计", "正常"],
            ],
        )

    def test_resize_for_processing_caps_large_images_without_upscaling(self):
        ocr_backend._load_runtime()
        image = np.full((3000, 4000, 3), 255, dtype=np.uint8)

        resized, scale = ocr_backend._resize_for_processing(image)
        small, small_scale = ocr_backend._resize_for_processing(image[:600, :800])

        self.assertEqual(resized.shape[:2], (1800, 2400))
        self.assertAlmostEqual(scale, 0.6)
        self.assertEqual(small.shape[:2], (600, 800))
        self.assertEqual(small_scale, 1.0)

    def test_wide_detector_limit_caps_dynamic_long_side_only_for_wide_images(self):
        wide = np.zeros((550, 1900, 3), dtype=np.uint8)
        ordinary = np.zeros((1200, 1600, 3), dtype=np.uint8)

        wide_limit = ocr_backend._adaptive_wide_detector_limit(wide, 1280)

        self.assertEqual(wide_limit, int(2880 / (1900 / 550)))
        self.assertEqual(
            ocr_backend._adaptive_wide_detector_limit(ordinary, 1280),
            1280,
        )

    def test_embedded_row_ruler_group_header_removes_only_empty_margin_row(self):
        grid = [
            ["1", "基础信息", "", "", "", "目标与测量", "", "过程记录", ""],
            ["序号", "资产编号", "设备名称", "设备类型", "主机名", "IP地址", "MAC地址", "操作系统", "CPU"],
            *[[str(row), *[f"值{column}" for column in range(1, 9)]] for row in range(1, 5)],
            ["", "末行", "设备", "类型", "主机", "IP", "MAC", "系统", "CPU"],
            ["", "", "", "", "", "", "", "", ""],
        ]
        confidence = [[0.99 if value else 0.77 for value in row] for row in grid]
        spans: list[dict[str, object]] = []

        changed = ocr_backend._repair_embedded_row_ruler_group_header(
            grid,
            confidence,
            spans,
        )

        self.assertTrue(changed)
        self.assertEqual(len(grid), 7)
        self.assertEqual(
            grid[0],
            ["基础信息", "", "", "", "", "目标与测量", "", "过程记录", ""],
        )
        self.assertEqual(
            [(span["column"], span["column_span"]) for span in spans],
            [(0, 5), (5, 2), (7, 2)],
        )

    def test_fused_spreadsheet_ruler_rejoins_complementary_last_row(self):
        grid = [
            ["B 1 基础信息", "H", "", "", "", "目标与测量", "", "", "", "J K L 过程记录", "", "", ""],
            ["2", "序号", "员工编号", "姓名", "部门", "岗位", "日期", "班次", "计划上班", "实际上班", "计划下班", "实际下班", "工时/h"],
            ["3", "1", "AP-001", "甲", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
            ["4", "2", "AP-002", "乙", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
            ["5", "3", "AP-003", "丙", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
            ["6", "4", "AP-004", "丁", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
            ["7", "5", "AP-005", "戊", "", "", "", "", "", "", "旧值", "旧值", "旧值"],
            ["", "", "", "", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
        ]
        confidence = [[0.99 if value else 0.77 for value in row] for row in grid]
        spans: list[dict[str, object]] = []

        changed = ocr_backend._repair_fused_spreadsheet_ruler_table(
            grid,
            confidence,
            spans,
        )

        self.assertTrue(changed)
        self.assertEqual((len(grid), len(grid[0])), (7, 12))
        self.assertEqual(
            grid[0],
            ["基础信息", "", "", "", "目标与测量", "", "", "", "过程记录", "", "", ""],
        )
        self.assertEqual(
            grid[-1],
            ["5", "AP-005", "戊", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
        )
        self.assertEqual(
            [(span["column"], span["column_span"]) for span in spans],
            [(0, 4), (4, 4), (8, 4)],
        )

    def test_fused_spreadsheet_ruler_repairs_after_fused_column_ruler_strip(self):
        grid = [
            ["A", "C D E F G", "", "", "", "", "", "", "", "", "", "", ""],
            ["B 1 基础信息", "H", "", "", "", "目标与测量", "", "", "", "J K L 过程记录", "", "", ""],
            ["2", "序号", "员工编号", "姓名", "部门", "岗位", "日期", "班次", "计划上班", "实际上班", "计划下班", "实际下班", "工时/h"],
            ["3", "1", "AP-001", "甲", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
            ["4", "2", "AP-002", "乙", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
            ["5", "3", "AP-003", "丙", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
            ["6", "4", "AP-004", "丁", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
            ["7", "5", "AP-005", "戊", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
            ["8", "6", "AP-006", "己", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
            ["9", "7", "AP-007", "庚", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
            ["10", "8", "AP-008", "辛", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
            ["11", "9", "AP-009", "壬", "", "", "", "", "", "", "17:00", "17:01", "8"],
            ["", "", "", "", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"],
        ]
        confidence = [[0.99 if value else 0.77 for value in row] for row in grid]
        spans: list[dict[str, object]] = []

        grid, confidence = ocr_backend._strip_spreadsheet_ui_headers(
            grid,
            confidence,
        )
        changed = ocr_backend._repair_fused_spreadsheet_ruler_table(
            grid,
            confidence,
            spans,
        )

        self.assertTrue(changed)
        self.assertEqual((len(grid), len(grid[0])), (11, 12))
        self.assertEqual(grid[0][0], "基础信息")
        self.assertEqual(grid[-1], ["9", "AP-009", "壬", "部门", "岗位", "日期", "白班", "08:00", "08:01", "17:00", "17:01", "8"])
        self.assertEqual(
            [(span["column"], span["column_span"]) for span in spans],
            [(0, 4), (4, 4), (8, 4)],
        )

    def test_image_quality_reports_dark_blurred_capture_without_altering_it(self):
        image = np.full((600, 900, 3), 28, dtype=np.uint8)
        cv2.putText(image, "A100 72.2", (180, 320), cv2.FONT_HERSHEY_SIMPLEX, 2.0, (48, 48, 48), 3)
        blurred = cv2.GaussianBlur(image, (31, 31), 9.0)

        report = table_pipeline.assess_image_quality(blurred)

        self.assertIn("dark", report["issues"])
        self.assertIn("blur", report["issues"])
        self.assertTrue(report["needs_recapture"])
        self.assertEqual(report["width"], 900)
        self.assertEqual(report["height"], 600)

    def test_image_quality_keeps_clear_document_out_of_recapture_state(self):
        image = np.full((600, 900, 3), 242, dtype=np.uint8)
        for row in range(80, 560, 48):
            cv2.line(image, (70, row), (830, row), (70, 70, 70), 2)
        for column in range(70, 840, 126):
            cv2.line(image, (column, 80), (column, 560), (70, 70, 70), 2)
        cv2.putText(image, "A100 72.2", (180, 320), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (20, 20, 20), 2)

        report = table_pipeline.assess_image_quality(image)

        self.assertFalse(report["needs_recapture"])
        self.assertNotIn("blur", report["issues"])

    def test_dark_capture_enhancement_lifts_exposure_without_inventing_geometry(self):
        image = np.full((480, 720, 3), 62, dtype=np.uint8)
        cv2.rectangle(image, (70, 60), (650, 420), (86, 86, 86), -1)
        cv2.putText(image, "A100 72.2", (150, 255), cv2.FONT_HERSHEY_SIMPLEX, 1.4, (18, 18, 18), 3)

        enhanced = table_pipeline._enhance_for_ocr(image)

        self.assertEqual(enhanced.shape, image.shape)
        self.assertGreater(float(np.median(enhanced)), float(np.median(image)) + 35.0)
        self.assertLess(float(np.mean(enhanced >= 253)), 0.03)

    def test_recognition_keeps_severe_blur_as_a_result_warning(self):
        source = inspect.getsource(ocr_backend._recognize)

        self.assertIn("图片较模糊，部分文字可能无法确认", source)

    def test_recognition_keeps_edge_touching_crop_as_a_result_warning(self):
        source = inspect.getsource(ocr_backend._recognize)

        self.assertIn("表格或纸张边缘贴近画面边界", source)

    def test_implausibly_sparse_grid_is_rejected_before_cell_verification(self):
        self.assertTrue(
            ocr_backend._grid_is_implausibly_sparse(
                [["标题", ""]] + [["", ""] for _ in range(11)]
            )
        )
        self.assertTrue(
            ocr_backend._grid_is_implausibly_sparse(
                [["标题", ""]] + [["", ""] for _ in range(8)]
            )
        )
        self.assertFalse(
            ocr_backend._grid_is_implausibly_sparse(
                [["编号", "名称"], ["1", "频谱仪"], ["2", "信号源"], ["3", "示波器"]]
            )
        )

    def test_stronger_spatial_columns_override_incomplete_narrow_grid(self):
        ruled = [["A", "B", "C", "D"] for _ in range(10)]
        spatial = [["A", "B", "C", "D", "E", "F"] for _ in range(10)]

        self.assertTrue(ocr_backend._prefer_spatial_over_grid(ruled, spatial, True))
        self.assertFalse(ocr_backend._prefer_spatial_over_grid(ruled, spatial, False))
        self.assertFalse(
            ocr_backend._prefer_spatial_over_grid(spatial, ruled, True)
        )

    def test_asymmetric_grid_gap_allows_complete_spatial_edge_column_recovery(self):
        image = np.full((600, 440, 3), 255, dtype=np.uint8)
        self.assertTrue(
            ocr_backend._ruled_grid_has_asymmetric_side_gap(
                image,
                [143, 195, 247, 299, 351, 396],
            )
        )
        ruled = [["A", "B", "C", "D", "E", "F"] for _ in range(15)]
        spatial = [["A", "B", "C", "D", "E", "F", "G"] for _ in range(15)]
        self.assertTrue(
            ocr_backend._prefer_spatial_over_grid(
                ruled,
                spatial,
                True,
                edge_loss_suspected=True,
            )
        )
        self.assertFalse(
            ocr_backend._prefer_spatial_over_grid(
                ruled,
                spatial,
                True,
                edge_loss_suspected=False,
            )
        )

    def test_centered_grid_margin_does_not_claim_missing_edge_column(self):
        image = np.full((600, 670, 3), 255, dtype=np.uint8)
        self.assertFalse(
            ocr_backend._ruled_grid_has_asymmetric_side_gap(
                image,
                [55, 135, 215, 295, 375, 455, 535, 615],
            )
        )

    def test_expanded_grid_requires_complete_regular_supported_columns(self):
        ocr_backend._load_runtime()
        base = np.full((500, 320, 3), 255, dtype=np.uint8)
        candidate = np.full((560, 365, 3), 255, dtype=np.uint8)
        base_columns = [75, 117, 159, 201, 243, 286]
        base_rows = list(range(7, 488, 32))
        candidate_columns = [11, 99, 138, 177, 227, 268, 309, 352]
        candidate_rows = list(range(59, 559, 33))
        for x in candidate_columns:
            cv2.line(candidate, (x, 0), (x, 559), (0, 0, 0), 1)
        for y in candidate_rows:
            cv2.line(candidate, (0, y), (364, y), (0, 0, 0), 1)

        with patch.object(ocr_backend, "_screen_grid_is_credible", return_value=True):
            self.assertTrue(
                ocr_backend._expanded_grid_is_strictly_more_complete(
                    base,
                    (base_columns, base_rows, base.copy()),
                    candidate,
                    (candidate_columns, candidate_rows, candidate.copy()),
                )
            )

            irregular_columns = [0, 32, 39, 207, 337, 377, 426, 464, 492]
            self.assertFalse(
                ocr_backend._expanded_grid_is_strictly_more_complete(
                    base,
                    (base_columns, base_rows, base.copy()),
                    candidate,
                    (irregular_columns, candidate_rows, candidate.copy()),
                )
            )

    def test_sparse_wide_grid_is_not_replaced_by_page_text_layout(self):
        ruled = [["A", "", "", "", "", ""] for _ in range(10)]
        spatial = [["A", "B"] for _ in range(10)]

        self.assertFalse(ocr_backend._prefer_spatial_over_grid(ruled, spatial, True))

    def test_screen_grid_requires_supported_horizontal_lines(self):
        image = np.full((600, 800, 3), 255, dtype=np.uint8)
        columns = [0, 200, 400, 600, 799]
        rows = [20 + 30 * index for index in range(18)]
        self.assertFalse(ocr_backend._screen_grid_is_credible(image, columns, rows))

        for row in rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (205, 205, 205), 1)
        self.assertFalse(ocr_backend._screen_grid_is_credible(image, columns, rows))
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (205, 205, 205), 1)
        self.assertTrue(ocr_backend._screen_grid_is_credible(image, columns, rows))

    def test_spurious_column_collapse_requires_one_line_free_split(self):
        ocr_backend._load_runtime()
        image = np.full((600, 1200, 3), 255, dtype=np.uint8)
        rows = [20 + 35 * index for index in range(17)]
        columns = [20, 120, 220, 320, 420, 520, 620, 660, 720, 820, 920, 1020, 1120, 1180]
        weak_boundary = 660
        for row in rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (0, 0, 0), 1)
        for column in columns:
            if column != weak_boundary:
                cv2.line(image, (column, rows[0]), (column, rows[-1]), (0, 0, 0), 1)

        collapsed = ocr_backend._collapse_one_unruled_spurious_column_boundary(
            image,
            columns,
            rows,
        )

        self.assertEqual(collapsed, [value for value in columns if value != weak_boundary])

        second_weak = image.copy()
        second_weak[:, 920 - 3 : 920 + 4] = 255
        self.assertIsNone(
            ocr_backend._collapse_one_unruled_spurious_column_boundary(
                second_weak,
                columns,
                rows,
            )
        )

    def test_screen_grid_accepts_ruler_conflict_only_with_full_physical_support(self):
        ocr_backend._load_runtime()
        image = np.full((600, 900, 3), 255, dtype=np.uint8)
        columns = [20 + 100 * index for index in range(10)]
        rows = [20 + 18 * index for index in range(32)]
        for row in rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (160, 160, 160), 1)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (160, 160, 160), 1)
        mismatched_recovery = (columns[:-1], rows)

        with (
            patch.object(
                table_pipeline,
                "spreadsheet_ruler_confirms_columns",
                return_value=False,
            ),
            patch.object(
                table_pipeline,
                "_recover_spreadsheet_ruler_grid",
                return_value=mismatched_recovery,
            ) as recover_grid,
        ):
            self.assertTrue(ocr_backend._screen_grid_is_credible(image, columns, rows))

            weak = image.copy()
            weak[:, columns[5] - 4 : columns[5] + 5] = 255
            self.assertFalse(ocr_backend._screen_grid_is_credible(weak, columns, rows))

            recover_grid.return_value = ([10] + columns, rows)
            self.assertFalse(ocr_backend._screen_grid_is_credible(image, columns, rows))

    def test_structure_certificate_requires_exact_shape_and_spans(self):
        image = np.full((120, 240, 3), 255, dtype=np.uint8)
        certificate = ocr_backend._new_structure_certificate(
            image,
            [0, 80, 160, 239],
            [0, 40, 80, 119],
            "test",
        )
        grid = [["" for _ in range(3)] for _ in range(3)]
        self.assertTrue(ocr_backend._structure_certificate_matches(certificate, grid, []))
        self.assertFalse(
            ocr_backend._structure_certificate_matches(
                certificate,
                [row + [""] for row in grid],
                [],
            )
        )
        self.assertFalse(
            ocr_backend._structure_certificate_matches(
                certificate,
                grid,
                [{"row": 0, "column": 0, "row_span": 1, "column_span": 2}],
            )
        )

    def test_review_only_certificate_keeps_only_physically_matching_spans(self):
        image = np.full((120, 240, 3), 255, dtype=np.uint8)
        certificate = ocr_backend._new_structure_certificate(
            image,
            [0, 80, 160, 239],
            [0, 40, 80, 119],
            "test",
        )
        grid = [["分组", "", ""], ["A", "B", "C"], ["1", "2", "3"]]
        spans = [{"row": 0, "column": 0, "row_span": 1, "column_span": 2}]
        certificate["spans"] = [dict(span) for span in spans]
        certificate["verified"] = False

        self.assertTrue(
            ocr_backend._review_only_certificate_spans_match_physical_grid(
                certificate,
                grid,
                spans,
            )
        )
        conflicting = [row[:] for row in grid]
        conflicting[0][1] = "不应被合并"
        self.assertFalse(
            ocr_backend._review_only_certificate_spans_match_physical_grid(
                certificate,
                conflicting,
                spans,
            )
        )
        self.assertFalse(
            ocr_backend._review_only_certificate_spans_match_physical_grid(
                certificate,
                grid,
                [],
            )
        )

    def test_assign_ocr_to_grid_preserves_certified_physical_shape(self):
        columns = [0, 100, 200, 300, 400]
        rows = [0, 40, 80]
        boxes = np.asarray(
            [
                [[5, 5], [95, 5], [95, 35], [5, 35]],
                [[5, 45], [95, 45], [95, 75], [5, 75]],
            ],
            dtype=np.float32,
        )
        grid, confidence = assign_ocr_to_grid(
            columns,
            rows,
            boxes,
            ["编号 1", "频率 515.221"],
            [0.99, 0.99],
            preserve_geometry=True,
        )
        self.assertEqual((len(grid), len(grid[0])), (2, 4))
        self.assertEqual((len(confidence), len(confidence[0])), (2, 4))

    def test_certified_title_span_rejects_any_real_internal_divider(self):
        ocr_backend._load_runtime()
        image = np.full((120, 400, 3), 255, dtype=np.uint8)
        columns = [0, 80, 160, 240, 320, 399]
        rows = [0, 40, 80, 119]
        for boundary in columns:
            cv2.line(image, (boundary, rows[1]), (boundary, rows[2]), (0, 0, 0), 2)
        cv2.line(image, (columns[2], rows[0]), (columns[2], rows[1]), (0, 0, 0), 2)
        certificate = ocr_backend._new_structure_certificate(
            image,
            columns,
            rows,
            "test",
        )
        grid = [["标题", "", "", "", ""], ["A", "B", "C", "D", "E"], ["", "", "", "", ""]]
        confidence = [[0.99, 0.0, 0.0, 0.0, 0.0], [0.99] * 5, [0.0] * 5]
        self.assertEqual(
            ocr_backend._certified_title_span(image, certificate, grid, confidence),
            [],
        )

    def test_certified_title_span_accepts_group_row_before_dense_detail_row(self):
        ocr_backend._load_runtime()
        image = np.full((160, 400, 3), 255, dtype=np.uint8)
        columns = [0, 80, 160, 240, 320, 399]
        rows = [0, 40, 80, 120, 159]
        for boundary in (columns[0], columns[2], columns[-1]):
            cv2.line(image, (boundary, rows[1]), (boundary, rows[2]), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (boundary, rows[2]), (boundary, rows[3]), (0, 0, 0), 2)
        certificate = ocr_backend._new_structure_certificate(
            image, columns, rows, "test"
        )
        grid = [
            ["完整标题", "", "", "", ""],
            ["基础信息", "", "目标与测量", "", ""],
            ["A", "B", "C", "D", "E"],
            ["1", "2", "3", "4", "5"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        spans = ocr_backend._certified_title_span(
            image, certificate, grid, confidence
        )

        self.assertEqual(spans[0]["column_span"], 5)

    def test_certified_title_span_accepts_dense_layout_when_perspective_hides_following_rules(self):
        ocr_backend._load_runtime()
        image = np.full((320, 700, 3), 255, dtype=np.uint8)
        columns = list(range(0, 701, 100))
        rows = list(range(0, 321, 40))
        certificate = ocr_backend._new_structure_certificate(
            image, columns, rows, "test"
        )
        grid = [
            ["", "", "BOM物料清单 — 2026-07 批次", "", "", "", ""],
            ["基础信息", "", "", "目标与测量", "", "", ""],
            ["序号", "编码", "名称", "型号", "数量", "单位", "状态"],
            *[[str(row)] * 7 for row in range(5)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        spans = ocr_backend._certified_title_span(
            image, certificate, grid, confidence
        )

        self.assertEqual(grid[0][0], "BOM物料清单 — 2026-07 批次")
        self.assertEqual(spans[0]["column_span"], 7)

    def test_certified_page_headers_use_page_text_and_physical_group_rules(self):
        ocr_backend._load_runtime()
        image = np.full((320, 900, 3), 255, dtype=np.uint8)
        columns = list(range(0, 901, 100))
        rows = list(range(40, 321, 40))
        for boundary in (columns[0], columns[4], columns[8], columns[-1]):
            cv2.line(image, (boundary, rows[1]), (boundary, rows[2]), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (boundary, rows[2]), (boundary, rows[-1]), (0, 0, 0), 2)
        grid = [
            ["", "", "旧标题碎片", "", "", "", "", "", ""],
            ["", "旧分组", "", "", "", "旧分组2", "", "", "过程记录"],
            ["登记号", "日期", "名称", "类别", "来源", "数量", "单位", "状态", "经办人"],
            *[[str(row)] * 9 for row in range(4)],
        ]
        confidence = [[0.77 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("完整标题 — 2026-06 批次", 450.0, 60.0, 0.99),
            ("基础信息", 200.0, 100.0, 0.99),
            ("目标与测量", 600.0, 100.0, 0.99),
            ("过程记录", 850.0, 100.0, 0.99),
        ]

        spans = ocr_backend._recover_certified_page_header_spans(
            image,
            columns,
            rows,
            grid,
            confidence,
            evidence,
        )

        self.assertEqual(grid[0], ["完整标题 — 2026-06 批次"] + [""] * 8)
        self.assertEqual(
            grid[1],
            ["基础信息", "", "", "", "目标与测量", "", "", "", "过程记录"],
        )
        self.assertEqual(
            [(span["row"], span["column"], span["column_span"]) for span in spans],
            [(0, 0, 9), (1, 0, 4), (1, 4, 4)],
        )
        self.assertFalse(
            ocr_backend._has_unresolved_multilevel_header(grid, spans)
        )
        grid[1][8] = ""
        self.assertTrue(
            ocr_backend._has_unresolved_multilevel_header(grid, spans)
        )

    def test_certified_page_headers_join_only_split_bounded_batch_title(self):
        ocr_backend._load_runtime()
        image = np.full((320, 900, 3), 255, dtype=np.uint8)
        columns = list(range(0, 901, 100))
        rows = list(range(40, 321, 40))
        for boundary in (columns[0], columns[4], columns[8], columns[-1]):
            cv2.line(image, (boundary, rows[1]), (boundary, rows[2]), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (boundary, rows[2]), (boundary, rows[-1]), (0, 0, 0), 2)
        grid = [
            ["教学课表", "", "", "成绩-2026", "", "", "-07 批次", "", ""],
            ["基础信息", "", "", "", "目标与测量", "", "", "", "过程记录"],
            ["登记号", "日期", "名称", "类别", "来源", "数量", "单位", "状态", "经办人"],
            *[[str(row)] * 9 for row in range(4)],
        ]
        confidence = [[0.77 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("教学课表", 200.0, 60.0, 0.99),
            ("成绩-2026", 450.0, 60.0, 0.98),
            ("-07 批次", 650.0, 60.0, 0.97),
            ("基础信息", 200.0, 100.0, 0.99),
            ("目标与测量", 600.0, 100.0, 0.99),
            ("过程记录", 850.0, 100.0, 0.99),
        ]

        spans = ocr_backend._recover_certified_page_header_spans(
            image, columns, rows, grid, confidence, evidence
        )

        self.assertEqual(grid[0], ["教学课表成绩-2026-07 批次"] + [""] * 8)
        self.assertEqual(confidence[0][0], 0.97)
        self.assertEqual(spans[0]["role"], "title")
        self.assertEqual(spans[0]["column_span"], 9)

    def test_strict_simple_title_span_accepts_unlisted_textual_headers(self):
        grid = [
            ["采购询价对比表", "", "", "", "", "", "", ""],
            ["物料名称", "规格", "供应商A", "供应商B", "供应商C", "最低价", "交期", "选择结果"],
            *[["项目", "标准", "甲", "乙", "丙", "低", "本周", "选择"] for _ in range(5)],
        ]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]

        spans = ocr_backend._strict_simple_title_span(grid, confidence)

        self.assertEqual(
            spans,
            [{"row": 0, "column": 0, "row_span": 1, "column_span": 8, "role": "title"}],
        )

    def test_certified_page_headers_join_split_chinese_month_title(self):
        ocr_backend._load_runtime()
        image = np.full((320, 900, 3), 255, dtype=np.uint8)
        columns = list(range(0, 901, 100))
        rows = list(range(40, 321, 40))
        for boundary in (columns[0], columns[4], columns[8], columns[-1]):
            cv2.line(image, (boundary, rows[1]), (boundary, rows[2]), (0, 0, 0), 2)
        for boundary in columns:
            cv2.line(image, (boundary, rows[2]), (boundary, rows[-1]), (0, 0, 0), 2)
        grid = [
            ["", "", "仪器测", "", "式结果 2026", "", "年07月", "", ""],
            ["基础信息", "", "", "", "目标与测量", "", "", "", "过程记录"],
            ["登记号", "日期", "名称", "类别", "来源", "数量", "单位", "状态", "经办人"],
            *[[str(row)] * 9 for row in range(4)],
        ]
        confidence = [[0.77 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("仪器测", 250.0, 60.0, 0.99),
            ("试结果 2026", 450.0, 60.0, 0.98),
            ("年07月", 650.0, 60.0, 0.97),
            ("基础信息", 200.0, 100.0, 0.99),
            ("目标与测量", 600.0, 100.0, 0.99),
            ("过程记录", 850.0, 100.0, 0.99),
        ]

        spans = ocr_backend._recover_certified_page_header_spans(
            image, columns, rows, grid, confidence, evidence
        )

        self.assertEqual(grid[0], ["仪器测试结果  2026年07月"] + [""] * 8)
        self.assertEqual(spans[0]["role"], "title")
        self.assertEqual(spans[0]["column_span"], 9)

    def test_certified_page_headers_recover_simple_split_title_above_detail_row(self):
        ocr_backend._load_runtime()
        image = np.full((320, 500, 3), 255, dtype=np.uint8)
        columns = list(range(0, 501, 100))
        rows = list(range(0, 321, 40))
        for boundary in columns:
            cv2.line(image, (boundary, rows[1]), (boundary, rows[-1]), (0, 0, 0), 2)
        grid = [
            ["校准检", "验记录 2026年11", "月", "", ""],
            ["序号", "仪器编号", "项目", "标准值", "状态"],
            *[[str(row)] * 5 for row in range(6)],
        ]
        confidence = [[0.77 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("校准检", 100.0, 20.0, 0.99),
            ("验记录 2026年11", 250.0, 20.0, 0.98),
            ("月", 400.0, 20.0, 0.97),
        ]

        spans = ocr_backend._recover_certified_page_header_spans(
            image, columns, rows, grid, confidence, evidence
        )

        self.assertEqual(grid[0], ["校准检验记录  2026年11月", "", "", "", ""])
        self.assertEqual(spans[0]["column_span"], 5)

    def test_unresolved_simple_title_requires_full_width_span(self):
        grid = [
            ["", "校准检", "验记录 2026年11月", "", ""],
            ["序号", "仪器编号", "项目", "标准值", "状态"],
            *[[str(row)] * 5 for row in range(6)],
        ]

        self.assertTrue(ocr_backend._has_unresolved_simple_title(grid, []))
        self.assertFalse(
            ocr_backend._has_unresolved_simple_title(
                grid,
                [{"row": 0, "column": 0, "column_span": 5, "role": "title"}],
            )
        )

    def test_unresolved_body_sequence_detects_missing_header(self):
        grid = [[str(index), f"ID-{index:03d}", "设备"] for index in range(1, 9)]

        self.assertTrue(ocr_backend._grid_starts_with_unresolved_body_sequence(grid))

    def test_unresolved_body_sequence_keeps_explicit_header(self):
        grid = [["序号", "编号", "名称"]] + [
            [str(index), f"ID-{index:03d}", "设备"] for index in range(1, 9)
        ]

        self.assertFalse(ocr_backend._grid_starts_with_unresolved_body_sequence(grid))

    def test_unique_page_evidence_restores_only_one_interior_body_cell(self):
        columns = [0, 100, 200, 300]
        rows = [0, 40, 80, 120, 160]
        grid = [
            ["分组", "", ""],
            ["字段A", "字段B", "字段C"],
            ["", "", ""],
            ["1", "2", ""],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        evidence = [
            ("应跳过表头", 150.0, 20.0, 1.0),
            ("二号线", 50.0, 100.0, 0.99995),
            ("边界文本", 100.0, 100.0, 1.0),
            ("重复A", 250.0, 100.0, 1.0),
            ("重复B", 250.0, 100.0, 1.0),
        ]

        scores = ocr_backend._recover_unique_blank_body_page_evidence(
            grid,
            confidence,
            columns,
            rows,
            evidence,
        )

        self.assertEqual(grid[2], ["二号线", "", ""])
        self.assertEqual(confidence[2][0], 0.77)
        self.assertEqual(scores, [0.99995])

    def test_page_evidence_restores_one_explicit_summary_footer_row(self):
        grid = [
            ["序号", "名称", "数量", "单位"],
            ["1", "设备A", "10", "台"],
            ["2", "设备B", "20", "台"],
        ]
        confidence = [[0.99 for _ in row] for row in grid]
        columns = [0, 100, 200, 300, 400]
        rows = [0, 40, 80, 120]
        candidate_rows = [0, 40, 80, 120, 160]
        evidence = [
            ("合计", 50.0, 140.0, 0.99),
            ("共2条", 150.0, 140.0, 0.98),
        ]

        recovered = ocr_backend._recover_explicit_summary_footer_row(
            grid,
            confidence,
            columns,
            rows,
            columns,
            candidate_rows,
            evidence,
        )

        self.assertTrue(recovered)
        self.assertEqual(grid[-1], ["合计", "共2条", "", ""])
        self.assertEqual(rows, candidate_rows)
        self.assertTrue(all(value <= 0.77 for value in confidence[-1] if value))

    def test_page_evidence_does_not_restore_unproven_sparse_metadata_row(self):
        grid = [["序号", "名称"], ["1", "设备A"], ["2", "设备B"]]
        confidence = [[0.99, 0.99] for _ in grid]
        rows = [0, 40, 80, 120]

        recovered = ocr_backend._recover_explicit_summary_footer_row(
            grid,
            confidence,
            [0, 100, 200],
            rows,
            [0, 100, 200],
            [0, 40, 80, 120, 160],
            [("记录条数", 50.0, 140.0, 0.99)],
        )

        self.assertFalse(recovered)
        self.assertEqual(len(grid), 3)

    def test_summary_footer_drops_dash_without_a_visible_mark(self):
        ocr_backend._load_runtime()
        image = np.full((160, 400, 3), 255, dtype=np.uint8)
        grid = [
            ["序号", "名称", "数量", "备注"],
            ["1", "设备A", "10", "正常"],
            ["2", "设备B", "20", "待确认"],
        ]
        confidence = [[0.99 for _ in row] for row in grid]
        columns = [0, 100, 200, 300, 400]
        rows = [0, 40, 80, 120]
        candidate_rows = [0, 40, 80, 120, 160]

        recovered = ocr_backend._recover_explicit_summary_footer_row(
            grid,
            confidence,
            columns,
            rows,
            columns,
            candidate_rows,
            [
                ("合计", 50.0, 140.0, 0.99),
                ("共2条", 150.0, 140.0, 0.98),
                ("—", 350.0, 140.0, 0.97),
            ],
            image,
        )

        self.assertTrue(recovered)
        self.assertEqual(grid[-1], ["合计", "共2条", "", ""])

    def test_summary_footer_keeps_a_physically_visible_dash(self):
        ocr_backend._load_runtime()
        image = np.full((160, 400, 3), 255, dtype=np.uint8)
        cv2.line(image, (330, 140), (370, 140), (0, 0, 0), 2)
        grid = [
            ["序号", "名称", "数量", "备注"],
            ["1", "设备A", "10", "正常"],
            ["2", "设备B", "20", "待确认"],
        ]
        confidence = [[0.99 for _ in row] for row in grid]
        columns = [0, 100, 200, 300, 400]
        rows = [0, 40, 80, 120]
        candidate_rows = [0, 40, 80, 120, 160]

        recovered = ocr_backend._recover_explicit_summary_footer_row(
            grid,
            confidence,
            columns,
            rows,
            columns,
            candidate_rows,
            [
                ("合计", 50.0, 140.0, 0.99),
                ("共2条", 150.0, 140.0, 0.98),
                ("—", 350.0, 140.0, 0.97),
            ],
            image,
        )

        self.assertTrue(recovered)
        self.assertEqual(grid[-1], ["合计", "共2条", "", "—"])

    def test_existing_summary_footer_drops_only_unseen_dash_artifact(self):
        ocr_backend._load_runtime()
        image = np.full((160, 400, 3), 255, dtype=np.uint8)
        columns = [0, 100, 200, 300, 400]
        rows = [0, 40, 80, 120, 160]
        grid = [
            ["序号", "名称", "数量", "备注"],
            ["1", "设备A", "10", "正常"],
            ["2", "设备B", "20", "待确认"],
            ["合计", "共2条", "", "—"],
        ]
        confidence = [[0.99 for _ in row] for row in grid]

        removed = ocr_backend._remove_unseen_summary_footer_marks(
            image,
            grid,
            confidence,
            columns,
            rows,
        )

        self.assertEqual(removed, [(3, 3)])
        self.assertEqual(grid[-1], ["合计", "共2条", "", ""])
        cv2.line(image, (330, 140), (370, 140), (0, 0, 0), 2)
        grid[-1][-1] = "—"
        self.assertEqual(
            ocr_backend._remove_unseen_summary_footer_marks(
                image,
                grid,
                confidence,
                columns,
                rows,
            ),
            [],
        )

    def test_redundant_simple_title_fragment_collapses_to_full_width_title(self):
        grid = [
            ["", "", "BOM物料清单 2026年09月", "", "09月", "", "", ""],
            ["序号", "物料编码", "物料名称", "规格型号", "品牌", "封装", "单位", "单机用量"],
            *[
                [str(index), f"M-{index:03d}", "物料", "A20", "品牌", "盒", "件", "10"]
                for index in range(1, 6)
            ],
        ]
        confidence = [[0.99 if value else -1.0 for value in row] for row in grid]

        spans = ocr_backend._repair_redundantly_split_simple_title(
            grid,
            confidence,
        )

        self.assertEqual(grid[0][0], "BOM物料清单  2026年09月")
        self.assertEqual(sum(bool(value) for value in grid[0]), 1)
        self.assertEqual(
            spans,
            [{"row": 0, "column": 0, "row_span": 1, "column_span": 8, "role": "title"}],
        )

        iso_batch = [list(row) for row in grid]
        iso_batch[0] = ["", "", "", "安全检查 — 2026-05 批次", "", "", "", ""]
        iso_confidence = [
            [0.99 if value else -1.0 for value in row] for row in iso_batch
        ]
        iso_spans = ocr_backend._repair_redundantly_split_simple_title(
            iso_batch,
            iso_confidence,
        )
        self.assertEqual(iso_batch[0][0], "安全检查 — 2026-05 批次")
        self.assertEqual(iso_spans[0]["column_span"], 8)

        first_column_iso = [list(row) for row in grid]
        first_column_iso[0] = ["安全检查 — 2026-05 批次", "", "", "", "", "", "", ""]
        first_column_confidence = [
            [0.99 if value else -1.0 for value in row]
            for row in first_column_iso
        ]
        first_column_spans = ocr_backend._repair_redundantly_split_simple_title(
            first_column_iso,
            first_column_confidence,
        )
        self.assertEqual(first_column_iso[0][0], "安全检查 — 2026-05 批次")
        self.assertEqual(first_column_spans[0]["column_span"], 8)

        confusable_dash = [list(row) for row in grid]
        confusable_dash[0] = ["", "", "", "安全检查一2026-05批次", "", "", "", ""]
        confusable_confidence = [
            [0.99 if value else -1.0 for value in row] for row in confusable_dash
        ]
        confusable_spans = ocr_backend._repair_redundantly_split_simple_title(
            confusable_dash,
            confusable_confidence,
        )
        self.assertEqual(confusable_dash[0][0], "安全检查一2026-05批次")
        self.assertEqual(confusable_spans[0]["column_span"], 8)

        split_date = [list(row) for row in grid]
        split_date[0] = ["", "", "项目人员任务", "", "2026年08月", "月", "", ""]
        split_confidence = [
            [0.99 if value else -1.0 for value in row] for row in split_date
        ]
        split_spans = ocr_backend._repair_redundantly_split_simple_title(
            split_date,
            split_confidence,
        )
        self.assertEqual(split_date[0][0], "项目人员任务  2026年08月")
        self.assertEqual(split_spans[0]["column_span"], 8)

        single_suffix = [list(row) for row in grid]
        single_suffix[0] = ["", "", "", "", "无框浅线登记2026年11月", "", "", "月"]
        single_suffix_confidence = [
            [0.99 if value else -1.0 for value in row] for row in single_suffix
        ]
        single_suffix_spans = ocr_backend._repair_redundantly_split_simple_title(
            single_suffix,
            single_suffix_confidence,
        )
        self.assertEqual(single_suffix[0][0], "无框浅线登记  2026年11月")
        self.assertEqual(single_suffix_spans[0]["column_span"], 8)

        deep_nested = [
            ["", "", "物流运输跟踪 2026年06月", "", "", ""],
            ["", "基础信息", "", "数量与金额", "", "过程记录"],
            ["", "计划", "", "实际", "计划", ""],
            ["序号", "运输单号", "订单号", "承运商", "车牌号", "司机"],
            *[[str(index)] * 6 for index in range(5)],
        ]
        deep_nested_confidence = [
            [0.99 if value else -1.0 for value in row] for row in deep_nested
        ]
        deep_nested_spans = ocr_backend._repair_redundantly_split_simple_title(
            deep_nested,
            deep_nested_confidence,
        )
        self.assertEqual(deep_nested[0][0], "物流运输跟踪  2026年06月")
        self.assertEqual(deep_nested_spans[0]["column_span"], 6)

        education = [
            ["", "教学课表成绩", "2026年01月", ""],
            ["序号", "学号", "姓名", "班级"],
            *[[str(index), "A区", "赵敏", "A区"] for index in range(1, 6)],
        ]
        education_confidence = [
            [0.99 if value else -1.0 for value in row] for row in education
        ]
        education_spans = ocr_backend._repair_redundantly_split_simple_title(
            education,
            education_confidence,
        )
        self.assertEqual(education[0][0], "教学课表成绩  2026年01月")
        self.assertEqual(education_spans[0]["column_span"], 4)

        wide_logistics = [
            ["", "", "", "", "", "", "", "物流运输跟踪", "2026年04月", "月", "", "", "", "", "", ""],
            [
                "序号", "运输单号", "订单号", "承运商", "车牌号", "司机",
                "联系电话", "始发地", "目的地", "计划发车", "实际发车",
                "预计到达", "实际到达", "货物名称", "件数", "重量/kg",
            ],
            *[[str(index)] * 16 for index in range(1, 6)],
        ]
        wide_confidence = [
            [0.99 if value else -1.0 for value in row]
            for row in wide_logistics
        ]
        wide_spans = ocr_backend._repair_redundantly_split_simple_title(
            wide_logistics,
            wide_confidence,
        )
        self.assertEqual(wide_logistics[0][0], "物流运输跟踪  2026年04月")
        self.assertEqual(wide_spans[0]["column_span"], 16)

        three_column = [
            ["校准检", "验记录 2026年11", "月"],
            ["序号", "仪器编号", "校准项目"],
            *[[str(index), f"AP-{index:03d}", "外观"] for index in range(1, 6)],
        ]
        three_column_confidence = [
            [0.99 if value else -1.0 for value in row]
            for row in three_column
        ]
        three_column_spans = ocr_backend._repair_redundantly_split_simple_title(
            three_column,
            three_column_confidence,
        )
        self.assertEqual(three_column[0][0], "校准检验记录  2026年11月")
        self.assertEqual(three_column_spans[0]["column_span"], 3)

    def test_nested_parent_header_moves_left_only_when_center_proves_extension(self):
        columns = list(range(0, 601, 100))
        rows = list(range(0, 501, 50))
        grid = [
            ["标题", "", "", "", "", ""],
            ["", "基础信息", "", "数量与金额", "", "过程记录"],
            ["计划", "", "实际", "", "计划", ""],
            *[[str(column) for column in range(6)] for _ in range(7)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 6, "role": "title"},
            {"row": 1, "column": 1, "row_span": 1, "column_span": 2, "role": "group_header"},
            {"row": 1, "column": 3, "row_span": 1, "column_span": 2, "role": "group_header"},
            {"row": 2, "column": 0, "row_span": 1, "column_span": 2, "role": "group_header"},
            {"row": 2, "column": 2, "row_span": 1, "column_span": 2, "role": "group_header"},
            {"row": 2, "column": 4, "row_span": 1, "column_span": 2, "role": "group_header"},
        ]

        changed = ocr_backend._extend_leading_nested_group_header(
            grid,
            confidence,
            spans,
            columns,
            rows,
            [("基础信息", 150.0, 75.0, 0.99)],
        )

        self.assertTrue(changed)
        self.assertEqual(grid[1][:2], ["基础信息", ""])
        self.assertEqual((spans[1]["column"], spans[1]["column_span"]), (0, 3))

        rejected_grid = [row[:] for row in grid]
        rejected_grid[1][0], rejected_grid[1][1] = "", "基础信息"
        rejected_spans = [dict(span) for span in spans]
        rejected_spans[1]["column"], rejected_spans[1]["column_span"] = 1, 2
        self.assertFalse(
            ocr_backend._extend_leading_nested_group_header(
                rejected_grid,
                confidence,
                rejected_spans,
                columns,
                rows,
                [("基础信息", 200.0, 75.0, 0.99)],
            )
        )

    def test_shifted_paired_nested_header_requires_matching_row_ruler(self):
        grid = [
            ["标题", "", "", "", "", ""],
            ["基础信息", "", "", "", "数量与金额", ""],
            ["3", "本期", "", "累计", "", "上限"],
            ["序号", "仪器编号", "项目", "标准器", "标准值", "示值"],
            *[[str(column) for column in range(6)] for _ in range(5)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 6, "role": "title"},
            {"row": 1, "column": 0, "row_span": 1, "column_span": 4, "role": "group_header"},
            {"row": 1, "column": 4, "row_span": 1, "column_span": 2, "role": "group_header"},
        ]

        self.assertTrue(
            ocr_backend._repair_shifted_paired_nested_header(
                grid,
                confidence,
                spans,
            )
        )
        self.assertEqual(grid[2], ["本期", "", "累计", "", "上限", ""])
        self.assertEqual(
            [
                (span["column"], span["column_span"])
                for span in spans
                if span["row"] == 2
            ],
            [(0, 2), (2, 2), (4, 2)],
        )

        rejected = [row[:] for row in grid]
        rejected[2] = ["4", "本期", "", "累计", "", "上限"]
        self.assertFalse(
            ocr_backend._repair_shifted_paired_nested_header(
                rejected,
                confidence,
                [span for span in spans if span["row"] != 2],
            )
        )

        english_grid = [
            ["", "", "", "中英编码混排", "2026年05月", "", "", "", ""],
            [
                "No.",
                "中文名称",
                "English Name",
                "Model",
                "Serial No.",
                "Code",
                "Value",
                "Lower",
                "Upper",
            ],
            *[
                [str(index), "名称", "Name", "M16", "S-001", "C-001", "1", "0", "2"]
                for index in range(1, 6)
            ],
        ]
        english_confidence = [
            [0.99 if value else -1.0 for value in row] for row in english_grid
        ]

        english_spans = ocr_backend._repair_redundantly_split_simple_title(
            english_grid,
            english_confidence,
        )

        self.assertEqual(english_grid[0][0], "中英编码混排  2026年05月")
        self.assertEqual(english_spans[0]["column_span"], 9)

        nested_grid = [
            ["", "", "", "", "实验样", "实验样品检测 2026年09月", "", "", "", "", ""],
            ["", "基础信息", "", "", "", "数量与金额", "过程记录", "", "状态判定", "", "追溯信息"],
            ["序号", "样品编号", "样品名称", "样品类型", "采样地点", "采样时间", "检测项目", "检测方法", "仪器编号", "检出限", "检测结果"],
            *[[str(index)] * 11 for index in range(5)],
        ]
        nested_confidence = [
            [0.99 if value else -1.0 for value in row] for row in nested_grid
        ]

        nested_spans = ocr_backend._repair_redundantly_split_simple_title(
            nested_grid,
            nested_confidence,
        )

        self.assertEqual(nested_grid[0][0], "实验样品检测  2026年09月")
        self.assertEqual(nested_spans[0]["column_span"], 11)

    def test_fused_paired_nested_header_splits_only_matching_four_column_span(self):
        grid = [
            ["标题", *("" for _ in range(15))],
            ["下限", "", "计划", "", "下限", "", "上限", "", "实际", "", "实际计划", "", "", "", "实际", ""],
            [f"字段{column}" for column in range(16)],
            *[[str(column) for column in range(16)] for _ in range(5)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 16, "role": "title"},
            *[
                {"row": 1, "column": column, "row_span": 1, "column_span": width, "role": "group_header"}
                for column, width in ((0, 2), (2, 2), (4, 2), (6, 2), (8, 2), (10, 4), (14, 2))
            ],
        ]
        original_grid = [row[:] for row in grid]
        original_confidence = [row[:] for row in confidence]
        original_spans = [dict(span) for span in spans]

        self.assertTrue(
            ocr_backend._repair_shifted_paired_nested_header(
                grid,
                confidence,
                spans,
            )
        )
        self.assertEqual(grid[1][10:14], ["实际", "", "计划", ""])
        self.assertEqual(
            [(span["column"], span["column_span"]) for span in spans if span["row"] == 1],
            [(0, 2), (2, 2), (4, 2), (6, 2), (8, 2), (10, 2), (12, 2), (14, 2)],
        )

        missing_span = [
            span
            for span in original_spans
            if not (
                span["row"] == 1
                and span["column"] == 10
                and span["column_span"] == 4
            )
        ]
        self.assertFalse(
            ocr_backend._repair_shifted_paired_nested_header(
                original_grid,
                original_confidence,
                missing_span,
            )
        )

    def test_bounded_medium_recovers_ordinal_and_repeated_last_row_category(self):
        ocr_backend._load_runtime()
        image = np.full((360, 240, 3), 245, dtype=np.uint8)
        grid = [["序号", "区域"]]
        grid.extend([[str(index), "B区"] for index in range(1, 8)])
        grid.append(["I", ""])
        confidence = [[0.77 if value else -1.0 for value in row] for row in grid]

        class FakeEngine:
            @staticmethod
            def text_rec(_request):
                return SimpleNamespace(
                    txts=["8", "B区"],
                    scores=[0.999, 0.88],
                )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((20, 40, 3), 180, dtype=np.uint8),
        ):
            recovered = ocr_backend._recover_bounded_ordinal_and_vocabulary_cells(
                image,
                grid,
                confidence,
                [0, 120, 239],
                list(range(0, 361, 40)),
                FakeEngine(),
            )

        self.assertEqual(grid[-1], ["8", "B区"])
        self.assertEqual(
            recovered,
            [(8, 0, "8"), (8, 1, "B区")],
        )

        numeric_grid = [["序号", "名称"]] + [
            ["6" if index == 9 else str(index), f"设备{index}"]
            for index in range(1, 11)
        ]
        numeric_confidence = [[0.99, 0.99] for _ in numeric_grid]

        class NumericEngine:
            @staticmethod
            def text_rec(_request):
                return SimpleNamespace(txts=["9"], scores=[0.999])

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((20, 40, 3), 180, dtype=np.uint8),
        ):
            numeric_recovered = (
                ocr_backend._recover_bounded_ordinal_and_vocabulary_cells(
                    image,
                    numeric_grid,
                    numeric_confidence,
                    [0, 120, 239],
                    list(range(0, 353, 32)),
                    NumericEngine(),
                )
            )

        self.assertEqual(numeric_grid[9][0], "9")
        self.assertEqual(numeric_recovered, [(9, 0, "9")])

        grid[-1][1] = "p区"

        class FakeNearEngine:
            @staticmethod
            def text_rec(_request):
                return SimpleNamespace(txts=["B区"], scores=[0.91])

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((20, 40, 3), 180, dtype=np.uint8),
        ):
            near = ocr_backend._recover_bounded_ordinal_and_vocabulary_cells(
                image,
                grid,
                confidence,
                [0, 120, 239],
                list(range(0, 361, 40)),
                FakeNearEngine(),
            )
        self.assertEqual(near, [(8, 1, "B区")])

        for row in range(1, len(grid) - 1):
            grid[row][1] = "待确认"
        grid[-1][1] = ""

        class FakeTypoEngine:
            @staticmethod
            def text_rec(_request):
                return SimpleNamespace(txts=["结确认"], scores=[0.90])

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((20, 40, 3), 180, dtype=np.uint8),
        ):
            typo = ocr_backend._recover_bounded_ordinal_and_vocabulary_cells(
                image,
                grid,
                confidence,
                [0, 120, 239],
                list(range(0, 361, 40)),
                FakeTypoEngine(),
            )
        self.assertEqual(typo, [(8, 1, "待确认")])

        identifier_grid = [["序号", "项目编号"]]
        identifier_grid.extend(
            [[str(index), f"APX-{index:03d}"] for index in range(1, 8)]
        )
        identifier_grid.append(["8", "AP-008"])
        identifier_confidence = [
            [0.99 for _ in row] for row in identifier_grid
        ]

        class FakeIdentifierEngine:
            @staticmethod
            def text_rec(_request):
                return SimpleNamespace(txts=["APX-008"], scores=[0.98])

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((20, 40, 3), 180, dtype=np.uint8),
        ):
            identifier = ocr_backend._recover_bounded_ordinal_and_vocabulary_cells(
                image,
                identifier_grid,
                identifier_confidence,
                [0, 120, 239],
                list(range(0, 361, 40)),
                FakeIdentifierEngine(),
            )
        self.assertEqual(identifier, [(8, 1, "APX-008")])

    def test_certified_batch_title_normalizes_only_bounded_batch_pattern(self):
        self.assertEqual(
            ocr_backend._normalize_certified_batch_title(
                "教学课表成绩— 2026-07批次"
            ),
            "教学课表成绩 — 2026-07 批次",
        )
        self.assertEqual(
            ocr_backend._normalize_certified_batch_title("设备台账—第一批"),
            "设备台账—第一批",
        )
        self.assertEqual(
            ocr_backend._normalize_certified_batch_title("设备运行记录2026年05月"),
            "设备运行记录  2026年05月",
        )

    def test_visible_batch_title_dash_requires_isolated_horizontal_component(self):
        ocr_backend._load_runtime()
        image = np.full((60, 200, 3), (80, 95, 70), dtype=np.uint8)
        cv2.rectangle(image, (40, 5), (70, 20), (240, 240, 240), -1)
        cv2.rectangle(image, (80, 12), (95, 13), (240, 240, 240), -1)
        cv2.rectangle(image, (105, 5), (145, 20), (240, 240, 240), -1)
        grid = [["库存出入库 一 2026-10 批次", ""], ["序号", "名称"]]
        confidence = [[0.93, 0.0], [0.99, 0.99]]

        recovered = ocr_backend._recover_visible_batch_title_dash(
            image,
            grid,
            confidence,
            [0, 100, 200],
            [0, 30, 60],
        )

        self.assertTrue(recovered)
        self.assertEqual(grid[0][0], "库存出入库 — 2026-10 批次")
        self.assertEqual(confidence[0][0], 0.77)

        grid[0][0] = "库存出入库-2026-10 批次"
        self.assertTrue(
            ocr_backend._recover_visible_batch_title_dash(
                image,
                grid,
                confidence,
                [0, 100, 200],
                [0, 30, 60],
            )
        )
        self.assertEqual(grid[0][0], "库存出入库 — 2026-10 批次")

        image[12:14, 80:96] = (80, 95, 70)
        grid[0][0] = "库存出入库 一 2026-10 批次"
        self.assertFalse(
            ocr_backend._recover_visible_batch_title_dash(
                image,
                grid,
                confidence,
                [0, 100, 200],
                [0, 30, 60],
            )
        )

    def test_only_top_spreadsheet_ruler_may_explain_an_excluded_rule(self):
        columns = list(range(0, 901, 100))
        rows = list(range(50, 291, 30))
        evidence = [
            (label, (left + right) * 0.5, 35.0, 0.99)
            for label, left, right in zip(
                ("A", "8", "C", "D", "E", "F", "G", "H"),
                columns,
                columns[1:],
            )
        ]
        horizontal = np.zeros((320, 901), dtype=np.uint8)
        with (
            patch.object(
                ocr_backend.pipeline,
                "_grid_maps",
                return_value=(horizontal, horizontal, horizontal),
            ),
            patch.object(
                ocr_backend.pipeline,
                "_line_centers",
                return_value=[20],
            ),
        ):
            self.assertTrue(
                ocr_backend._excluded_supported_rows_are_only_spreadsheet_ruler(
                    np.zeros((320, 901, 3), dtype=np.uint8),
                    columns,
                    rows,
                    evidence,
                )
            )
        with (
            patch.object(
                ocr_backend.pipeline,
                "_grid_maps",
                return_value=(horizontal, horizontal, horizontal),
            ),
            patch.object(
                ocr_backend.pipeline,
                "_line_centers",
                return_value=[20, 320],
            ),
        ):
            self.assertFalse(
                ocr_backend._excluded_supported_rows_are_only_spreadsheet_ruler(
                    np.zeros((350, 901, 3), dtype=np.uint8),
                    columns,
                    rows,
                    evidence,
                )
            )

    def test_primary_report_range_removes_excel_gutter_and_strong_chart_axes(self):
        ocr_backend._load_runtime()
        image = np.full((900, 1500, 3), 255, dtype=np.uint8)
        columns = [31, 45, 155, 271, 386, 500, 612, 717, 821, 939, 1041, 1420, 1431]
        rows = list(range(300, 721, 35))
        for row in rows:
            cv2.line(image, (31, row), (1041, row), (40, 40, 40), 2)
        for column in columns[:11]:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (40, 40, 40), 2)
        cv2.line(image, (1420, rows[0]), (1420, rows[-1]), (40, 40, 40), 2)
        cv2.line(image, (1431, rows[0]), (1431, rows[-1]), (40, 40, 40), 2)

        selected = ocr_backend._select_primary_report_column_range(image, columns, rows)

        self.assertEqual(selected, (1, 10))

    def test_primary_report_range_keeps_a_legitimate_wide_notes_column(self):
        ocr_backend._load_runtime()
        image = np.full((700, 1200, 3), 255, dtype=np.uint8)
        columns = [40, 150, 260, 370, 900, 1010, 1160]
        rows = list(range(120, 581, 46))
        for row in rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (35, 35, 35), 2)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (35, 35, 35), 2)

        self.assertIsNone(
            ocr_backend._select_primary_report_column_range(image, columns, rows)
        )

    def test_primary_report_crop_translates_variable_grid_coordinates_exactly(self):
        ocr_backend._load_runtime()
        image = np.full((900, 1500, 3), 255, dtype=np.uint8)
        columns = [31, 45, 155, 271, 386, 500, 612, 717, 821, 939, 1041, 1420, 1431]
        rows = list(range(300, 721, 35))
        for row in rows:
            cv2.line(image, (31, row), (1041, row), (40, 40, 40), 2)
        for column in columns[:11]:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (40, 40, 40), 2)
        cv2.line(image, (1420, rows[0]), (1420, rows[0] + 70), (40, 40, 40), 2)
        ruled = (columns, rows, image.copy())

        primary = ocr_backend._primary_report_crop_from_ruled_grid(image, ruled)

        self.assertIsNotNone(primary)
        _, (shifted_columns, shifted_rows, _), metadata = primary
        left, top, _, _ = metadata["crop_bounds"]
        self.assertEqual([value + left for value in shifted_columns], columns[1:11])
        self.assertEqual([value + top for value in shifted_rows], rows)
        self.assertNotEqual(np.diff(shifted_columns).min(), np.diff(shifted_columns).max())

    def test_suspicious_ruled_semantic_cells_are_repaired_from_bounded_crops(self):
        ocr_backend._load_runtime()
        grid = [
            ["编号", "中心频率", "功率(dBm)", "采样率"],
            ["1", "523.471 N MHz", "-10.008", "MS/s"],
        ]
        confidence = [[0.99] * 4, [0.99] * 4]
        image = np.full((100, 400, 3), 245, dtype=np.uint8)
        columns = [0, 80, 220, 310, 400]
        rows = [0, 50, 100]

        def recognizer(texts):
            def recognize(request):
                self.assertEqual(len(request.img), 3)
                return SimpleNamespace(txts=texts, scores=[0.99] * 3)

            return recognize

        engine = SimpleNamespace(
            fast_text_rec=recognizer(["523.471MHz", "-10.00", "8MS/s"]),
            text_rec=recognizer(["523.471 MHz", "-10.00", "8 MS/s"]),
            server_text_rec=recognizer(["523.471 MHz", "-10.00", "8 MS/s"]),
        )

        scores = ocr_backend._repair_suspicious_ruled_semantic_cells(
            image,
            grid,
            confidence,
            columns,
            rows,
            engine,
        )

        self.assertEqual(
            grid[1],
            ["1", "523.471 MHz", "-10.00", "8 MS/s"],
        )
        self.assertEqual(len(scores), 9)
        self.assertTrue(all(value >= 0.86 for value in confidence[1][1:]))

    def test_valid_ruled_semantic_cells_do_not_trigger_extra_models(self):
        ocr_backend._load_runtime()
        grid = [
            ["中心频率", "带宽", "功率(dBm)", "采样率"],
            ["515.128 MHz", "9 kHz", "-10.00", "2.4 MS/s"],
        ]

        self.assertEqual(
            ocr_backend._suspicious_ruled_semantic_locations(grid),
            [],
        )

    def test_broken_iso_date_is_repaired_only_from_bounded_model_consensus(self):
        ocr_backend._load_runtime()
        grid = [["日期", "状态"], ["2026- 5-03- -26", "完成"]]
        confidence = [[0.99, 0.99], [0.77, 0.99]]
        image = np.full((100, 240, 3), 245, dtype=np.uint8)

        def recognizer(request):
            self.assertEqual(len(request.img), 1)
            return SimpleNamespace(txts=["2026-03-26"], scores=[0.99])

        engine = SimpleNamespace(
            fast_text_rec=recognizer,
            text_rec=recognizer,
            server_text_rec=recognizer,
        )

        scores = ocr_backend._repair_suspicious_ruled_semantic_cells(
            image,
            grid,
            confidence,
            [0, 120, 240],
            [0, 50, 100],
            engine,
        )

        self.assertEqual(grid[1][0], "2026-03-26")
        self.assertEqual(confidence[1][0], 0.90)
        self.assertEqual(len(scores), 3)

    def test_sample_rate_accepts_frequency_units_without_triggering_boundary_pair(self):
        ocr_backend._load_runtime()
        grid = [
            ["中心频率", "采样率", "峰值功率"],
            ["561.158MHz", "768kHz", "0dBm"],
            ["570.373MHz", "192kHz", "10dBm"],
            ["579.515MHz", "96kHz", "-20dBm"],
        ]

        self.assertEqual(
            ocr_backend._suspicious_ruled_semantic_locations(grid),
            [],
        )

    def test_sample_rate_percentage_column_does_not_trigger_full_column_reread(self):
        ocr_backend._load_runtime()
        grid = [
            ["序号", "采样率"],
            ["1", "91.25%"],
            ["2", "92.50%"],
            ["3", "93.75%"],
            ["4", "95.00%"],
            ["5", "96.25%"],
            ["6", "97.50%"],
        ]

        self.assertEqual(
            ocr_backend._suspicious_ruled_semantic_locations(grid),
            [],
        )

    def test_semantic_headers_with_inline_units_accept_numeric_body_values(self):
        grid = [
            ["频率(MHz)", "带宽(kHz)", "功率(dBm)", "采样率(MS/s)"],
            ["515.472", "9", "-20", "2.4"],
        ]

        self.assertEqual(
            ocr_backend._suspicious_ruled_semantic_locations(grid),
            [],
        )

    def test_power_column_outliers_and_missing_signs_trigger_bounded_recheck(self):
        ocr_backend._load_runtime()
        grid = [
            ["编号", "功率"],
            ["1", "-10"],
            ["2", "-20"],
            ["3", "-30"],
            ["4", "520"],
            ["5", "2530"],
            ["6", "20"],
        ]

        locations = ocr_backend._suspicious_ruled_semantic_locations(grid)

        self.assertIn((4, 1, "power", "power_magnitude_outlier"), locations)
        self.assertIn((5, 1, "power", "power_magnitude_outlier"), locations)
        self.assertIn((6, 1, "power", "power_missing_sign"), locations)

    def test_adjacent_quantity_columns_are_rechecked_as_a_pair(self):
        grid = [
            ["采购数量", "到货数量"],
            ["175", "45"],
            ["66", "ε8"],
        ]

        locations = ocr_backend._suspicious_ruled_semantic_locations(grid)

        self.assertIn((2, 0, "quantity", "boundary_pair"), locations)
        self.assertIn((2, 1, "quantity", "invalid_format"), locations)

    def test_power_outlier_without_independent_consensus_is_withheld(self):
        ocr_backend._load_runtime()
        grid = [
            ["编号", "功率"],
            ["1", "-10"],
            ["2", "-20"],
            ["3", "-30"],
            ["4", "520"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]
        image = np.full((250, 200, 3), 245, dtype=np.uint8)
        columns = [0, 100, 200]
        rows = [0, 50, 100, 150, 200, 250]

        def recognizer(value):
            return lambda request: SimpleNamespace(txts=[value], scores=[0.99])

        engine = SimpleNamespace(
            fast_text_rec=recognizer("520"),
            text_rec=recognizer("20"),
            server_text_rec=recognizer("-205"),
        )

        ocr_backend._repair_suspicious_ruled_semantic_cells(
            image, grid, confidence, columns, rows, engine
        )

        self.assertEqual(grid[4][1], "")
        self.assertLess(confidence[4][1], 0.0)

    def test_missing_power_sign_uses_two_independent_family_votes(self):
        ocr_backend._load_runtime()
        grid = [
            ["编号", "功率"],
            ["1", "-10"],
            ["2", "-20"],
            ["3", "-30"],
            ["4", "20"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]
        image = np.full((250, 200, 3), 245, dtype=np.uint8)
        columns = [0, 100, 200]
        rows = [0, 50, 100, 150, 200, 250]

        def recognizer(value):
            return lambda request: SimpleNamespace(txts=[value], scores=[0.99])

        engine = SimpleNamespace(
            fast_text_rec=recognizer("-20"),
            text_rec=recognizer("20"),
            server_text_rec=recognizer("-20"),
        )

        ocr_backend._repair_suspicious_ruled_semantic_cells(
            image, grid, confidence, columns, rows, engine
        )

        self.assertEqual(grid[4][1], "-20")
        self.assertGreaterEqual(confidence[4][1], 0.86)

    def test_unsigned_positive_power_is_preserved_and_marked_for_review(self):
        ocr_backend._load_runtime()
        grid = [
            ["编号", "功率"],
            ["1", "-10"],
            ["2", "-20"],
            ["3", "-30"],
            ["4", "20"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]
        image = np.full((250, 200, 3), 245, dtype=np.uint8)
        columns = [0, 100, 200]
        rows = [0, 50, 100, 150, 200, 250]

        def recognizer(value):
            return lambda request: SimpleNamespace(txts=[value], scores=[0.99])

        engine = SimpleNamespace(
            fast_text_rec=recognizer("20"),
            text_rec=recognizer("20"),
            server_text_rec=recognizer("20"),
        )

        ocr_backend._repair_suspicious_ruled_semantic_cells(
            image, grid, confidence, columns, rows, engine
        )

        self.assertEqual(grid[4][1], "20")
        self.assertEqual(confidence[4][1], 0.77)

    def test_preverified_unsigned_power_is_not_reread(self):
        ocr_backend._load_runtime()
        grid = [
            ["编号", "功率"],
            ["1", "-10"],
            ["2", "-20"],
            ["3", "-30"],
            ["4", "20"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]
        confidence[4][1] = 0.99
        image = np.full((250, 200, 3), 245, dtype=np.uint8)
        engine = SimpleNamespace(
            fast_text_rec=lambda _: self.fail("preverified cell must not be reread"),
            text_rec=lambda _: self.fail("preverified cell must not be reread"),
            server_text_rec=lambda _: self.fail("preverified cell must not be reread"),
        )

        scores = ocr_backend._repair_suspicious_ruled_semantic_cells(
            image,
            grid,
            confidence,
            [0, 100, 200],
            [0, 50, 100, 150, 200, 250],
            engine,
            preverified_cells={(4, 1)},
        )

        self.assertEqual(scores, [])
        self.assertEqual(grid[4][1], "20")
        self.assertEqual(confidence[4][1], 0.77)

    def test_repeated_status_singleton_is_repaired_by_two_matching_views(self):
        ocr_backend._load_runtime()
        grid = [
            ["编号", "审核状态"],
            ["1", "已审"],
            ["2", "已审"],
            ["3", "单日"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]
        image = np.full((200, 200, 3), 245, dtype=np.uint8)

        def medium_recognizer(request):
            self.assertEqual(len(request.img), 1)
            return SimpleNamespace(txts=["已审"], scores=[0.99])

        def alternate_recognizer(request):
            self.assertEqual(len(request.img), 1)
            return SimpleNamespace(txts=["☑已审"], scores=[0.90])

        engine = SimpleNamespace(
            text_rec=medium_recognizer,
            server_text_rec=alternate_recognizer,
        )
        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 40, 3), 220, dtype=np.uint8),
        ):
            scores = ocr_backend._repair_repeated_semantic_singletons(
                image,
                grid,
                confidence,
                [0, 100, 200],
                [0, 50, 100, 150, 200],
                engine,
            )

        self.assertEqual(grid[3][1], "已审")
        self.assertEqual(confidence[3][1], 0.77)
        self.assertEqual(scores, [0.99, 0.90])

    def test_source_grid_must_span_most_of_frame_before_skipping_rectification(self):
        image = np.full((600, 800, 3), 255, dtype=np.uint8)

        self.assertTrue(
            ocr_backend._ruled_grid_spans_source_frame(
                image,
                [20, 300, 600, 790],
                [5, 200, 400, 590],
            )
        )
        self.assertFalse(
            ocr_backend._ruled_grid_spans_source_frame(
                image,
                [200, 300, 400, 500],
                [150, 250, 350, 450],
            )
        )

    def test_repeated_row_endpoints_restore_a_missing_outer_column(self):
        image = np.full((220, 340, 3), 245, dtype=np.uint8)
        horizontal = np.zeros((220, 340), dtype=np.uint8)
        rows = list(range(20, 201, 20))
        for row in rows:
            cv2.line(horizontal, (30, row), (300, row), 255, 2)
            cv2.line(image, (30, row), (300, row), (45, 45, 45), 2)
        for column in (30, 120, 165, 210, 255, 300):
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (45, 45, 45), 2)

        recovered = table_pipeline._recover_outer_columns_from_horizontal_endpoints(
            image,
            horizontal,
            [120, 165, 210, 255, 300],
            rows,
        )

        self.assertLessEqual(abs(recovered[0] - 30), 2)
        self.assertEqual(recovered[1:], [120, 165, 210, 255, 300])

    def test_thin_auto_crop_is_rejected_to_preserve_complete_page(self):
        source = np.full((1200, 800, 3), 255, dtype=np.uint8)
        header_band = source[:300, :].copy()
        edge_strip = source[:, :250].copy()
        metadata = {"detected": True}

        self.assertFalse(ocr_backend._rectified_crop_is_usable(source, header_band, metadata))
        self.assertFalse(ocr_backend._rectified_crop_is_usable(source, edge_strip, metadata))
        self.assertTrue(
            ocr_backend._rectified_crop_is_usable(source, source[200:1000, 100:700], metadata)
        )

    def test_complementary_rectified_photo_grid_recovers_rows_and_wide_edge_column(self):
        ocr_backend._load_runtime()
        image = np.full((1241, 1792, 3), 245, dtype=np.uint8)
        standard_columns = [18, 136, 271, 412, 512, 614, 715, 816, 918, 1019, 1120, 1222, 1323, 1424, 1526, 1628]
        adaptive_rows = list(range(8, 1212, 27))
        standard_rows = adaptive_rows[1:41]
        adaptive_columns = list(range(35, 1760, 72))
        standard_grid = (standard_columns, standard_rows, image)
        adaptive_grid = (adaptive_columns, adaptive_rows, image)

        with (
            patch.object(ocr_backend, "_photographic_ruled_grid_is_credible", return_value=True),
            patch.object(ocr_backend, "_screen_grid_is_credible", return_value=True),
            patch.object(ocr_backend, "_dense_grid_is_axis_aligned_screen_capture", return_value=True),
            patch.object(ocr_backend, "_ruled_grid_spans_source_frame", return_value=True),
            patch.object(table_pipeline, "_edge_rule_continuation", return_value=0.95),
        ):
            recovered = ocr_backend._recover_complementary_rectified_photo_grid(
                image,
                standard_grid,
                adaptive_grid,
                maximum_cells=1280,
            )

        self.assertIsNotNone(recovered)
        grid, metrics = recovered
        self.assertEqual(len(grid[0]) - 1, 16)
        self.assertEqual(grid[0][-1], image.shape[1] - 1)
        self.assertEqual(metrics["trailing_rule_continuation"], 0.95)

    def test_complementary_rectified_photo_grid_rejects_unsupported_edge(self):
        ocr_backend._load_runtime()
        image = np.full((1241, 1792, 3), 245, dtype=np.uint8)
        standard_grid = (
            [18, 136, 271, 412, 512, 614, 715, 816, 918, 1019, 1120, 1222, 1323, 1424, 1526, 1628],
            list(range(35, 1115, 27)),
            image,
        )
        adaptive_grid = (
            list(range(35, 1760, 72)),
            list(range(8, 1212, 27)),
            image,
        )

        with (
            patch.object(ocr_backend, "_photographic_ruled_grid_is_credible", return_value=True),
            patch.object(ocr_backend, "_screen_grid_is_credible", return_value=True),
            patch.object(ocr_backend, "_dense_grid_is_axis_aligned_screen_capture", return_value=True),
            patch.object(ocr_backend, "_ruled_grid_spans_source_frame", return_value=True),
            patch.object(table_pipeline, "_edge_rule_continuation", return_value=0.40),
        ):
            recovered = ocr_backend._recover_complementary_rectified_photo_grid(
                image,
                standard_grid,
                adaptive_grid,
                maximum_cells=1280,
            )

        self.assertIsNone(recovered)

    def test_disagreeing_descriptive_numeric_is_preserved_for_cell_consensus(self):
        ocr_backend._load_runtime()
        image = np.full((60, 180, 3), 255, dtype=np.uint8)
        boxes = np.array([[[5, 10], [120, 10], [120, 40], [5, 40]]], dtype=float)
        output = SimpleNamespace(boxes=boxes, txts=["频率 515.221"], scores=[0.99])
        alternate = SimpleNamespace(txts=["频率 515.227"], scores=[0.98])
        engine = SimpleNamespace(text_rec=lambda _: alternate)

        _, texts, scores, rejected = ocr_backend._verified_numeric_components(
            image,
            output,
            engine,
        )

        self.assertEqual(texts, ["频率 515.221"])
        self.assertEqual(scores, [0.99])
        self.assertEqual(rejected, 0)

    def test_temperature_description_is_not_erased_by_glyph_disagreement(self):
        ocr_backend._load_runtime()
        image = np.full((60, 260, 3), 255, dtype=np.uint8)
        boxes = np.array([[[5, 10], [220, 10], [220, 40], [5, 40]]], dtype=float)
        output = SimpleNamespace(
            boxes=boxes,
            txts=["室温 25C；湿度 46%RH"],
            scores=[0.91],
        )
        engine = SimpleNamespace(
            text_rec=lambda _: SimpleNamespace(
                txts=["室温25℃；湿度46%RH"],
                scores=[0.96],
            )
        )

        _, texts, scores, rejected = ocr_backend._verified_numeric_components(
            image,
            output,
            engine,
            verify_all=True,
            unbounded=True,
        )

        self.assertEqual(texts, ["室温 25C；湿度 46%RH"])
        self.assertEqual(scores, [0.77])
        self.assertEqual(rejected, 1)

    def test_agreeing_numeric_segment_is_kept(self):
        ocr_backend._load_runtime()
        image = np.full((60, 180, 3), 255, dtype=np.uint8)
        boxes = np.array([[[5, 10], [120, 10], [120, 40], [5, 40]]], dtype=float)
        output = SimpleNamespace(boxes=boxes, txts=["频率 515.221"], scores=[0.99])
        alternate = SimpleNamespace(txts=["频率 515.221"], scores=[0.98])
        engine = SimpleNamespace(text_rec=lambda _: alternate)

        _, texts, scores, rejected = ocr_backend._verified_numeric_components(
            image,
            output,
            engine,
        )

        self.assertEqual(texts, ["频率 515.221"])
        self.assertAlmostEqual(scores[0], 0.99)
        self.assertEqual(rejected, 0)

    def test_numeric_verification_never_uses_motion_views_to_replace_existing_text(self):
        ocr_backend._load_runtime()
        image = np.full((60, 180, 3), 255, dtype=np.uint8)
        boxes = np.array([[[5, 10], [120, 10], [120, 40], [5, 40]]], dtype=float)
        output = SimpleNamespace(boxes=boxes, txts=["30"], scores=[0.99])
        engine = SimpleNamespace(
            text_rec=lambda _: SimpleNamespace(txts=["30"], scores=[0.98]),
            server_text_rec=lambda _: SimpleNamespace(txts=["-30"], scores=[0.99]),
            v4_server_text_rec=lambda _: SimpleNamespace(txts=["-30"], scores=[0.99]),
        )

        with patch.object(
            ocr_backend,
            "_wiener_motion_deblur_view",
            side_effect=AssertionError("numeric verification must not use motion views"),
        ):
            _, texts, _, rejected = ocr_backend._verified_numeric_components(
                image,
                output,
                engine,
                verify_all=True,
                unbounded=True,
            )

        self.assertEqual(texts, ["30"])
        self.assertEqual(rejected, 0)

    def test_known_concatenated_headers_use_high_confidence_dictionary_consensus(self):
        ocr_backend._load_runtime()
        image = np.full((60, 240, 3), 255, dtype=np.uint8)
        boxes = np.array([[[5, 10], [220, 10], [220, 45], [5, 45]]], dtype=float)
        output = SimpleNamespace(boxes=boxes, txts=["完成数量不良数"], scores=[0.94])
        engine = SimpleNamespace(
            text_rec=lambda _: SimpleNamespace(txts=["完成数量不良数量"], scores=[0.998]),
            server_text_rec=lambda _: SimpleNamespace(txts=["完成数量 不良数星"], scores=[0.84]),
            v4_server_text_rec=lambda _: SimpleNamespace(txts=["完成教量不良数"], scores=[0.96]),
        )

        _, texts, scores, rejected = ocr_backend._verified_numeric_components(
            image,
            output,
            engine,
            verify_all=True,
            unbounded=True,
        )

        self.assertEqual(texts, ["完成数量不良数量"])
        self.assertGreaterEqual(scores[0], 0.78)
        self.assertEqual(rejected, 0)

    def test_arbitrary_nearby_text_disagreement_is_not_dictionary_corrected(self):
        ocr_backend._load_runtime()
        image = np.full((60, 240, 3), 255, dtype=np.uint8)
        boxes = np.array([[[5, 10], [220, 10], [220, 45], [5, 45]]], dtype=float)
        output = SimpleNamespace(boxes=boxes, txts=["安全装置"], scores=[0.96])
        engine = SimpleNamespace(
            text_rec=lambda _: SimpleNamespace(txts=["安全装直"], scores=[0.99]),
            server_text_rec=lambda _: SimpleNamespace(txts=["安全装制"], scores=[0.92]),
            v4_server_text_rec=lambda _: SimpleNamespace(txts=["安全装至"], scores=[0.93]),
        )

        _, texts, _, rejected = ocr_backend._verified_numeric_components(
            image,
            output,
            engine,
            verify_all=True,
            unbounded=True,
        )

        self.assertEqual(texts, [""])
        self.assertEqual(rejected, 1)

    def test_maximum_verification_preserves_high_confidence_stacked_row_text(self):
        ocr_backend._load_runtime()
        image = np.full((100, 200, 3), 255, dtype=np.uint8)
        boxes = np.asarray([
            [[5, 5], [50, 5], [50, 25], [5, 25]],
            [[55, 5], [100, 5], [100, 25], [55, 25]],
            [[105, 5], [150, 5], [150, 25], [105, 25]],
            [[155, 5], [180, 5], [180, 77], [155, 77]],
        ], dtype=float)
        output = SimpleNamespace(
            boxes=boxes,
            txts=["状态", "正常", "风险", "无无"],
            scores=[0.99, 0.99, 0.99, 0.997],
        )
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["状态", "正常", "风险"],
                scores=[0.99, 0.99, 0.99],
            )
        )

        _, texts, scores, rejected = ocr_backend._verified_numeric_components(
            image,
            output,
            engine,
            verify_all=True,
            unbounded=True,
        )

        self.assertEqual(texts[-1], "无无")
        self.assertEqual(scores[-1], 0.997)
        self.assertEqual(rejected, 0)

    def test_maximum_component_verification_preserves_visible_compound_unit(self):
        ocr_backend._load_runtime()
        image = np.full((60, 180, 3), 255, dtype=np.uint8)
        boxes = np.array([[[5, 10], [120, 10], [120, 40], [5, 40]]], dtype=float)
        output = SimpleNamespace(boxes=boxes, txts=["V/A/C"], scores=[0.95])
        engine = SimpleNamespace(
            text_rec=lambda _: SimpleNamespace(txts=["V/A/°C"], scores=[0.94])
        )

        _, texts, scores, rejected = ocr_backend._verified_numeric_components(
            image,
            output,
            engine,
            verify_all=True,
            unbounded=True,
        )

        self.assertEqual(texts, ["V/A/C"])
        self.assertGreaterEqual(scores[0], 0.78)
        self.assertEqual(rejected, 0)

    def test_maximum_component_verification_uses_server_tie_break_for_text(self):
        ocr_backend._load_runtime()
        image = np.full((60, 180, 3), 255, dtype=np.uint8)
        boxes = np.array([[[5, 10], [120, 10], [120, 40], [5, 40]]], dtype=float)
        output = SimpleNamespace(boxes=boxes, txts=["数据採集卡"], scores=[0.93])
        engine = SimpleNamespace(
            text_rec=lambda _: SimpleNamespace(txts=["数据采集卡"], scores=[0.94]),
            server_text_rec=lambda _: SimpleNamespace(txts=["数据采集卡"], scores=[0.96]),
        )

        _, texts, scores, rejected = ocr_backend._verified_numeric_components(
            image,
            output,
            engine,
            verify_all=True,
            unbounded=True,
        )

        self.assertEqual(texts, ["数据采集卡"])
        self.assertGreaterEqual(scores[0], 0.78)
        self.assertEqual(rejected, 0)

    def test_maximum_component_verification_accepts_low_confidence_delta_consensus(self):
        ocr_backend._load_runtime()
        image = np.full((60, 180, 3), 255, dtype=np.uint8)
        boxes = np.array([[[5, 10], [80, 10], [80, 40], [5, 40]]], dtype=float)
        output = SimpleNamespace(boxes=boxes, txts=["ΔE"], scores=[0.70])
        engine = SimpleNamespace(
            text_rec=lambda _: SimpleNamespace(txts=["ΔE"], scores=[0.88])
        )

        _, texts, scores, rejected = ocr_backend._verified_numeric_components(
            image,
            output,
            engine,
            verify_all=True,
            unbounded=True,
        )

        self.assertEqual(texts, ["ΔE"])
        self.assertGreaterEqual(scores[0], 0.78)
        self.assertEqual(rejected, 0)

    def test_numeric_verification_checks_all_numeric_tokens_in_small_tables(self):
        ocr_backend._load_runtime()
        image = np.full((90, 1200, 3), 255, dtype=np.uint8)
        boxes = np.array(
            [
                [[index * 25, 10], [index * 25 + 20, 10], [index * 25 + 20, 40], [index * 25, 40]]
                for index in range(40)
            ],
            dtype=float,
        )
        texts = [str(1000 + index) for index in range(40)]
        output = SimpleNamespace(boxes=boxes, txts=texts, scores=[0.99] * 40)
        observed = {}

        def recognize(request):
            observed["count"] = len(request.img)
            return SimpleNamespace(
                txts=texts[: len(request.img)],
                scores=[0.99] * len(request.img),
            )

        _, verified, _, rejected = ocr_backend._verified_numeric_components(
            image,
            output,
            SimpleNamespace(text_rec=recognize),
        )

        self.assertEqual(observed["count"], 40)
        self.assertEqual(verified, texts)
        self.assertEqual(rejected, 0)

    def test_numeric_verification_is_bounded_only_for_dense_tables(self):
        ocr_backend._load_runtime()
        image = np.full((180, 2400, 3), 255, dtype=np.uint8)
        boxes = np.array(
            [
                [[index * 12, 10], [index * 12 + 10, 10], [index * 12 + 10, 40], [index * 12, 40]]
                for index in range(180)
            ],
            dtype=float,
        )
        texts = [str(1000 + index) for index in range(180)]
        output = SimpleNamespace(boxes=boxes, txts=texts, scores=[0.99] * 180)
        observed = {}

        def recognize(request):
            observed["count"] = len(request.img)
            return SimpleNamespace(
                txts=texts[: len(request.img)],
                scores=[0.99] * len(request.img),
            )

        ocr_backend._verified_numeric_components(
            image,
            output,
            SimpleNamespace(text_rec=recognize),
        )

        self.assertEqual(observed["count"], ocr_backend._MAX_VERIFICATION_CROPS)

    def test_maximum_accuracy_verifies_every_dense_table_token(self):
        ocr_backend._load_runtime()
        image = np.full((180, 2400, 3), 255, dtype=np.uint8)
        boxes = np.array(
            [
                [[index * 12, 10], [index * 12 + 10, 10], [index * 12 + 10, 40], [index * 12, 40]]
                for index in range(180)
            ],
            dtype=float,
        )
        texts = [str(1000 + index) for index in range(180)]
        output = SimpleNamespace(boxes=boxes, txts=texts, scores=[0.99] * 180)
        observed = {}

        def recognize(request):
            observed["count"] = len(request.img)
            return SimpleNamespace(txts=texts[: len(request.img)], scores=[0.99] * len(request.img))

        ocr_backend._verified_numeric_components(
            image,
            output,
            SimpleNamespace(text_rec=recognize),
            verify_all=True,
            unbounded=True,
        )

        self.assertEqual(observed["count"], 180)

    def test_low_confidence_short_unit_is_not_discarded_by_numeric_verifier(self):
        ocr_backend._load_runtime()
        image = np.full((60, 180, 3), 255, dtype=np.uint8)
        boxes = np.array([[[5, 10], [40, 10], [40, 40], [5, 40]]], dtype=float)
        output = SimpleNamespace(boxes=boxes, txts=["V"], scores=[0.62])
        engine = SimpleNamespace(
            text_rec=lambda _: self.fail("可信短单位不应触发低收益二次识别")
        )

        _, verified, scores, rejected = ocr_backend._verified_numeric_components(
            image,
            output,
            engine,
        )

        self.assertEqual(verified, ["V"])
        self.assertEqual(scores, [0.62])
        self.assertEqual(rejected, 0)

    def test_collapsed_header_and_first_data_row_are_split(self):
        grid = [
            ["编号 1", "频率 515.128MHz", "信号类型 模拟", "占用带宽 9kHz"],
            ["2", "516.347MHz", "数字", "35.4kHz"],
        ]
        confidence = [[0.95] * 4, [0.96] * 4]

        observed, observed_confidence, spans = ocr_backend._split_collapsed_header_data_row(
            grid,
            confidence,
            [],
        )

        self.assertEqual(observed[0], ["编号", "频率", "信号类型", "占用带宽"])
        self.assertEqual(observed[1], ["1", "515.128MHz", "模拟", "9kHz"])

    def test_repeated_title_before_collapsed_header_is_removed_then_split(self):
        grid = [
            ["工程材料领用表", "", "", "", "", "", "", ""],
            ["工程材料领用表", "", "", "", "", "", "", ""],
            [
                "领料单号 CQ-2026020",
                "材料编码 A001-30",
                "材料名称 一车间",
                "规格型号 标准件",
                "单位 件",
                "申请数量 6721.05",
                "实发数量 1207.33",
                "领用人 周林",
            ],
            ["QF-2026043", "B002-84", "现场复核", "常规项目", "件", "6136.84", "7134.28", "陈晨"],
        ]
        confidence = [[0.96 if value else 0.0 for value in row] for row in grid]
        spans = [
            {
                "row": 0,
                "column": 0,
                "row_span": 1,
                "column_span": 8,
                "role": "title",
            }
        ]

        observed, observed_confidence, observed_spans = (
            ocr_backend._split_collapsed_header_data_row(grid, confidence, spans)
        )

        self.assertEqual(len(observed), 4)
        self.assertEqual(observed[0][0], "工程材料领用表")
        self.assertEqual(
            observed[1],
            ["领料单号", "材料编码", "材料名称", "规格型号", "单位", "申请数量", "实发数量", "领用人"],
        )
        self.assertEqual(
            observed[2],
            ["CQ-2026020", "A001-30", "一车间", "标准件", "件", "6721.05", "1207.33", "周林"],
        )
        self.assertEqual(observed_spans, spans)
        self.assertEqual(len(observed_confidence), len(observed))

    def test_repeated_title_is_retained_without_collapsed_header_evidence(self):
        grid = [
            ["月度汇总表", "", "", ""],
            ["月度汇总表", "", "", ""],
            ["说明", "本页为补充记录", "", ""],
            ["编号", "名称", "数量", "备注"],
        ]
        confidence = [[0.96 if value else 0.0 for value in row] for row in grid]

        observed, _, _ = ocr_backend._split_collapsed_header_data_row(
            grid, confidence, []
        )

        self.assertEqual(observed, grid)

    def test_single_detector_box_is_redistributed_across_known_header_cells(self):
        grid = [
            ["项目 类别 数值1 数值2 数值3 单位 备注", "", "", "", "", "", ""],
            ["项目1", "类別2", "12", "28", "36", "V", "_"],
        ]
        confidence = [[0.96] + [0.0] * 6, [0.95] * 7]

        observed, observed_confidence, spans = ocr_backend._split_collapsed_header_data_row(
            grid,
            confidence,
            [{"row": 0, "column": 0, "row_span": 1, "column_span": 7, "role": "title"}],
        )
        ocr_backend._apply_consistency_checks(observed, observed_confidence)

        self.assertEqual(
            observed[0],
            ["项目", "类别", "数值1", "数值2", "数值3", "单位", "备注"],
        )
        self.assertEqual(observed[1][1], "类别2")
        self.assertEqual(observed[1][6], "—")
        self.assertEqual(spans, [])

    def test_remark_column_normalizes_visible_horizontal_mark_variants(self):
        grid = [
            ["名称", "备注"],
            ["设备1", "一"],
            ["设备2", "_"],
            ["设备3", "−"],
            ["设备4", "-"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual([row[1] for row in grid[1:]], ["—", "—", "—", "—"])

    def test_angle_column_normalizes_ocr_ordinal_ring_to_degree_sign(self):
        grid = [["序号", "方位角"], ["1", "1º"], ["2", "312°"]]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual([row[1] for row in grid[1:]], ["1°", "312°"])

    def test_consistency_checks_ignore_unicode_superscript_in_ordinal_column(self):
        grid = [["序号", "单位"]] + [
            [str(index), "²" if index == 4 else "m"] for index in range(1, 9)
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[4][1], "²")

    def test_consistency_checks_strip_proven_excel_row_prefix_from_ordinals(self):
        grid = [["标题", ""], ["2序号", "名称"]] + [
            [f"{row + 1}{row - 1}" if row < 4 else f"{row + 1} {row - 1}", f"设备{row}"]
            for row in range(2, 9)
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[1][0], "序号")
        self.assertEqual([grid[row][0] for row in range(2, 9)], ["1", "2", "3", "4", "5", "6", "7"])
        self.assertTrue(all(confidence[row][0] == 0.77 for row in range(1, 9)))

    def test_consistency_checks_keep_unproved_two_number_values(self):
        grid = [["序号", "名称"]] + [
            [f"{row + 4} {row}", f"设备{row}"] for row in range(1, 8)
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[1][0], "5 1")

    def test_absolute_spreadsheet_row_prefixes_require_column_wide_proof(self):
        grid = [
            [f"{row} {'A区' if row % 2 else '待确认'}", f"业务{row}"]
            for row in range(1, 13)
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        changed = ocr_backend._strip_absolute_spreadsheet_row_prefixes(
            grid,
            confidence,
        )

        self.assertEqual(len(changed), 12)
        self.assertEqual(grid[0][0], "A区")
        self.assertEqual(grid[1][0], "待确认")
        self.assertEqual(confidence[0][0], 0.77)

        sparse = [["1 A区"], ["2 B区"], *[["正文"] for _ in range(8)]]
        sparse_confidence = [[0.99] for _ in sparse]
        self.assertEqual(
            ocr_backend._strip_absolute_spreadsheet_row_prefixes(
                sparse,
                sparse_confidence,
            ),
            set(),
        )
        self.assertEqual(sparse[0][0], "1 A区")

    def test_consistency_checks_review_title_when_visible_date_spacing_is_lost(self):
        grid = [["交接验收整改2026年11月", "", ""], ["序号", "项目", "状态"]]
        confidence = [[0.99, 0.0, 0.0], [0.99, 0.99, 0.99]]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[0][0], "交接验收整改2026年11月")
        self.assertEqual(confidence[0][0], 0.77)

    def test_consistency_checks_keep_compact_date_in_dense_body(self):
        grid = [["项目", "日期", "状态"], ["巡检2026年11月", "2026-11-01", "完成"]]
        confidence = [[0.99, 0.99, 0.99], [0.99, 0.99, 0.99]]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(confidence[1][0], 0.99)

    def test_consistency_checks_restore_proven_measurement_trailing_zero(self):
        grid = [["序号", "电流/A"]] + [
            [str(index), "8.42" if index == 8 else f"{index}.125"]
            for index in range(1, 10)
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[8][1], "8.420")
        self.assertEqual(confidence[8][1], 0.77)

    def test_consistency_checks_do_not_pad_nonmeasurement_decimal(self):
        grid = [["序号", "备注"]] + [
            [str(index), "8.42" if index == 8 else f"{index}.125"]
            for index in range(1, 10)
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[8][1], "8.42")

    def test_consistency_checks_restore_proven_financial_trailing_zero(self):
        grid = [["序号", "费用类型"]] + [
            [str(index), "2712.1" if index == 8 else f"{index * 100}.25"]
            for index in range(1, 11)
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[8][1], "2712.10")
        self.assertEqual(confidence[8][1], 0.77)

    def test_consistency_checks_restore_proven_detection_limit_trailing_zero(self):
        grid = [["序号", "检出限"]] + [
            [str(index), "970.55" if index == 8 else f"{index}.125"]
            for index in range(1, 11)
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[8][1], "970.550")
        self.assertEqual(confidence[8][1], 0.77)

    def test_consistency_checks_restore_proven_current_value_trailing_zero(self):
        grid = [["序号", "当前值"]] + [
            [str(index), "486.72" if index == 8 else f"{index}.125"]
            for index in range(1, 11)
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[8][1], "486.720")
        self.assertEqual(confidence[8][1], 0.77)

    def test_consistency_checks_restore_proven_test_result_trailing_zero(self):
        for header in ("实测值", "偏差", "峰值", "复测值"):
            with self.subTest(header=header):
                grid = [["序号", header]] + [
                    [str(index), "486.72" if index == 8 else f"{index}.125"]
                    for index in range(1, 11)
                ]
                confidence = [[0.99, 0.99] for _ in grid]

                ocr_backend._apply_consistency_checks(grid, confidence)

                self.assertEqual(grid[8][1], "486.720")
                self.assertEqual(confidence[8][1], 0.77)

    def test_consistency_checks_find_four_level_detail_header(self):
        grid = [
            ["月报  2026年07月", "", "", ""],
            ["基础信息", "", "数值信息", ""],
            ["累计", "计划", "本期", "目标"],
            ["序号", "当前值", "目标值", "单位"],
            *[
                [
                    str(index),
                    "486.72" if index == 8 else f"{index}.125",
                    f"{index + 1}.125",
                    "V",
                ]
                for index in range(1, 11)
            ],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[11][1], "486.720")
        self.assertEqual(confidence[11][1], 0.77)

    def test_consistency_checks_normalize_unique_structured_insertions(self):
        grid = [["序号", "设备编号", "区域"]] + [
            [str(index), value, "A区" if index < 5 else region]
            for index, (value, region) in enumerate(
                [
                    ("XM-A7001-01", "A区"),
                    ("XM-A7002-02", "A区"),
                    ("XM-A7003-03", "A区"),
                    ("CN-B7004-04", "A区"),
                    ("XMI-A7005-05", "AI区"),
                    ("CN-B7006-06", "A区区"),
                ],
                start=1,
            )
        ]
        confidence = [[0.99, 0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[5][1], "XM-A7005-05")
        self.assertEqual(grid[5][2], "A区")
        self.assertEqual(grid[6][2], "A区")

    def test_consistency_checks_keep_unproved_structured_insertions(self):
        grid = [["序号", "设备编号"]] + [
            [str(index), value]
            for index, value in enumerate(
                ["APX-A001", "MOD-A002", "WIC-A003", "XMI-A004"],
                start=1,
            )
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[4][1], "XMI-A004")

    def test_consistency_checks_normalize_numeric_confusables_from_column_format(self):
        grid = [
            ["编号", "日期"],
            ["CQ-2026011", "2026-02-11"],
            ["AP-2026022", "2026-03-12"],
            ["QF-2026033", "2026-04-13"],
            ["MR-2026044", "2026-05-14"],
            ["CQ-20Z6055", "20Z6-D6-15"],
            ["AP-2026066", "2026-07-16"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[5], ["CQ-2026055", "2026-06-15"])
        self.assertEqual(confidence[5], [0.77, 0.77])

    def test_consistency_checks_remove_spaces_from_valid_iso_date_only(self):
        grid = [
            ["日期", "备注"],
            ["2026-02- 21", "2026-02- 99"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[1], ["2026-02-21", "2026-02- 99"])
        self.assertEqual(confidence[1], [0.77, 0.99])

    def test_consistency_checks_keep_alphanumeric_identifier_without_numeric_column_proof(self):
        grid = [
            ["编号", "名称"],
            ["CQ-A0Z6055", "设备甲"],
            ["AP-B126066", "设备乙"],
            ["QF-C126077", "设备丙"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]
        original = [list(row) for row in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid, original)

    def test_consistency_checks_restore_reversed_percent_only_from_repeated_value(self):
        grid = [
            ["编号", "结果"],
            ["1", "95%"],
            ["2", "正常"],
            ["3", "95%"],
            ["4", "%56"],
            ["5", "95%"],
            ["6", "%S6"],
            ["7", "完成"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[4][1], "95%")
        self.assertEqual(grid[6][1], "95%")
        self.assertEqual(confidence[4][1], 0.77)
        self.assertEqual(confidence[6][1], 0.77)

    def test_consistency_checks_restore_repeated_reversed_percent_without_clean_peer(self):
        grid = [
            ["编号", "结果"],
            ["1", "%56"],
            ["2", "正常"],
            ["3", "%56"],
            ["4", "完成"],
            ["5", "%56"],
            ["6", "合格"],
            ["7", "待确认"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual([grid[index][1] for index in (1, 3, 5)], ["95%"] * 3)

    def test_consistency_checks_remove_column_proven_code_whitespace(self):
        grid = [["序号", "Code"]] + [
            [str(index), value]
            for index, value in enumerate(
                [
                    "DAQ-M16-001",
                    "APX-A20-002",
                    "SEN-Q8-003",
                    "WIC-1000B-00 4",
                    "MOD-1000B-00 5",
                    "DAQ-A20-006",
                ],
                start=1,
            )
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[4][1], "WIC-1000B-004")
        self.assertEqual(grid[5][1], "MOD-1000B-005")
        self.assertEqual(confidence[4][1], 0.77)
        self.assertEqual(confidence[5][1], 0.77)

    def test_consistency_checks_remove_overlapping_identifier_tail_prefix(self):
        grid = [["编号"]] + [
            [
                "AP-B7003-1 -19"
                if index == 7
                else f"AP-B{7000 + index:04d}-{10 + index:02d}"
            ]
            for index in range(1, 9)
        ]
        confidence = [[0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[7][0], "AP-B7003-19")
        self.assertEqual(confidence[7][0], 0.77)

    def test_consistency_checks_keep_unproved_code_whitespace(self):
        grid = [
            ["序号", "说明"],
            ["1", "MOD-1000B-00 1"],
            ["2", "SEN-1000B-00 2"],
            ["3", "批次 A 03"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[1][1], "MOD-1000B-00 1")
        self.assertEqual(grid[2][1], "SEN-1000B-00 2")

    def test_consistency_checks_restore_unique_leading_line_label(self):
        grid = [["区域A", "区域B"]] + [
            [left, right]
            for left, right in [
                ("一号线", "号线"),
                ("号线", "号线"),
                ("号线", "一号线"),
                ("号线", "号线"),
            ]
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertTrue(
            all(value == "一号线" for row in grid[1:] for value in row)
        )
        self.assertEqual(confidence[2][0], 0.77)

    def test_consistency_checks_keep_ambiguous_leading_line_label(self):
        grid = [
            ["区域"],
            ["一号线"],
            ["一号线"],
            ["1号线"],
            ["1号线"],
            ["号线"],
            ["号线"],
            ["号线"],
        ]
        confidence = [[0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[5][0], "号线")
        self.assertEqual(grid[6][0], "号线")
        self.assertEqual(grid[7][0], "号线")

    def test_consistency_checks_never_borrow_identifier_prefix_from_other_column(self):
        grid = [
            ["序号", "型号规格", "外部编号"],
            ["1", "SEN-A001", "AP-X001"],
            ["2", "MOD-A002", "AP-X002"],
            ["3", "DAQ-A003", "AP-X003"],
            ["4", "APX-A004", "AP-X004"],
        ]
        confidence = [[0.99] * 3 for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[4][1], "APX-A004")

    def test_consistency_checks_remove_proven_spreadsheet_row_ruler_cells(self):
        grid = [["生产计划"], ["序号"]] + [
            [str(index)] for index in range(1, 26)
        ]
        grid[12][0] = "13"
        grid[23][0] = "24"
        confidence = [[0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[12][0], "")
        self.assertEqual(grid[23][0], "")
        self.assertLess(confidence[12][0], 0.0)
        self.assertLess(confidence[23][0], 0.0)

    def test_consistency_checks_keep_non_ruler_sequence_outlier(self):
        grid = [["生产计划"], ["序号"]] + [
            [str(index)] for index in range(1, 12)
        ]
        grid[7][0] = "99"
        confidence = [[0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[7][0], "99")

    def test_consistency_checks_normalize_column_proven_b_eight_formats(self):
        grid = [["序号", "区域", "批次", "设备编号"]] + [
            [str(index), region, batch, identifier]
            for index, (region, batch, identifier) in enumerate(
                [
                    ("B区", "批次-L08", "WO-B2001-01"),
                    ("B区", "批次-L08", "QD-B2002-02"),
                    ("B区", "批次-L08", "CN-B2003-03"),
                    ("88", "批次-1.08", "WO-82004-04"),
                ],
                start=1,
            )
        ]
        confidence = [[0.99] * 4 for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[4][1:], ["B区", "批次-L08", "WO-B2004-04"])

    def test_consistency_checks_use_table_wide_batch_and_location_enums(self):
        grid = [
            ["字段1", "字段2", "字段3"],
            ["批次-L08", "B区", "批次-LOB"],
            ["B区", "批次-L08", "A区"],
            ["批次-L08", "B区", "8区"],
        ]
        confidence = [[0.99] * 3 for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[1][2], "批次-L08")
        self.assertEqual(grid[3][2], "B区")

    def test_review_only_outliers_include_categorical_singletons_and_four_digit_counts(self):
        grid = [["序号", "操作员", "产量"]] + [
            [
                str(index),
                "爱王" if index == 8 else ("王强" if index % 2 else "赵敏"),
                str(1000 + index),
            ]
            for index in range(1, 10)
        ]

        selected = ocr_backend._review_only_spatial_outlier_cells(
            grid,
            include_dense_counts=True,
        )

        self.assertIn((8, 1), selected)
        self.assertTrue(all((row, 2) in selected for row in range(1, 10)))

    def test_consistency_checks_use_detail_row_after_split_title_and_groups(self):
        grid = [
            ["", "设备运行记录", "2026年05月", ""],
            ["基础信息", "", "数量与金额", ""],
            ["序号", "设备编号", "电流/A", "状态"],
            *[
                [str(index), f"ID-{index:03d}", "8.42" if index == 8 else f"{index}.125", "完成"]
                for index in range(1, 10)
            ],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[10][2], "8.420")

    def test_dense_identifier_case_risks_select_lowercase_prefix_only(self):
        grid = [["序号", "设备编号"]] + [
            [str(index), value]
            for index, value in enumerate(
                ["WO-A001", "XM-B002", "QD-C003", "CN-D004", "Wo-E005"],
                start=1,
            )
        ]

        self.assertEqual(ocr_backend._dense_identifier_case_risks(grid), {(5, 1)})

    def test_dense_percentage_format_risks_select_reversed_percent(self):
        grid = [["序号", "完成率"]] + [
            [str(index), "%60.66" if index == 7 else f"{90 + index / 10:.1f}%"]
            for index in range(1, 9)
        ]

        self.assertEqual(ocr_backend._dense_percentage_format_risks(grid), {(7, 1)})

    def test_dense_percentage_format_risks_accepts_three_valid_peers(self):
        grid = [
            ["编号", "状态"],
            ["1", "95%"],
            ["2", "95%"],
            ["3", "95%"],
            ["4", "%56"],
            ["5", "正常"],
            ["6", "完成"],
            ["7", "待确认"],
        ]

        self.assertEqual(
            ocr_backend._dense_percentage_format_risks(grid),
            {(4, 1)},
        )

    def test_dense_b_eight_confusion_risks_use_table_formats(self):
        grid = [["序号", "区域", "设备编号"]] + [
            [str(index), "B区" if index < 5 else "8区", value]
            for index, value in enumerate(
                [
                    "WO-B2001-31",
                    "QD-A2002-32",
                    "CN-B2003-33",
                    "XM-C2004-34",
                    "WO-82005-35",
                ],
                start=1,
            )
        ]

        self.assertEqual(
            ocr_backend._dense_b_eight_confusion_risks(grid),
            {(5, 1), (5, 2)},
        )

    def test_dense_b_eight_confusion_risks_require_b_evidence(self):
        grid = [["序号", "数值"]] + [[str(index), "88"] for index in range(1, 8)]

        self.assertEqual(ocr_backend._dense_b_eight_confusion_risks(grid), set())

    def test_physical_consensus_selects_invalid_angle_text_for_review(self):
        ocr_backend._load_runtime()
        image = np.full((90, 200, 3), 245, dtype=np.uint8)
        grid = [["序号", "方位角"], ["1", "99°"], ["2", "。66"]]
        confidence = [[0.99, 0.99] for _ in grid]

        selected = ocr_backend._standard_physical_consensus_locations(
            image,
            grid,
            confidence,
            [0, 100, 200],
            [0, 30, 60, 90],
            set(),
            preserve_recognition_batches=False,
        )

        self.assertIn((2, 1), selected)

    def test_multilevel_detail_header_drives_unit_and_remark_normalization(self):
        grid = [
            ["综合测试数据表", "", "", "", "", "", ""],
            ["基础信息", "", "", "测量与判定", "", "", ""],
            ["项目", "类别", "数值1", "数值2", "数值3", "单位", "备注"],
            ["项目01", "类别2", "15", "0.375", "7e-05", "℃", "一"],
        ]
        confidence = [[0.99] * 7 for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[3][5], "℃")
        self.assertEqual(grid[3][6], "—")

    def test_normal_header_row_is_not_split(self):
        grid = [
            ["任务编号", "任务名称", "负责人", "计划完成"],
            ["TASK-001", "界面设计", "陈晨", "2026-08-05"],
        ]

        observed, _, _ = ocr_backend._split_collapsed_header_data_row(
            grid,
            None,
            [],
        )

        self.assertEqual(observed, grid)

    @staticmethod
    def _draw_screen_table(
        rows: list[int],
        columns: list[int],
        *,
        height: int,
        width: int,
    ) -> np.ndarray:
        image = np.full((height, width, 3), 255, dtype=np.uint8)
        for row in rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (225, 225, 225), 1)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (225, 225, 225), 1)
        return image

    def test_screen_grid_recovers_alternating_light_spreadsheet_boundaries(self):
        image = np.full((620, 1000, 3), 255, dtype=np.uint8)
        expected_columns = [30, 150, 300, 420, 560, 700, 850, 970]
        expected_rows = [30 + 32 * index for index in range(18)]

        for row_index in range(len(expected_rows) - 1):
            top = expected_rows[row_index]
            bottom = expected_rows[row_index + 1]
            fill = 246 if row_index % 2 else 255
            image[top + 1 : bottom, expected_columns[0] + 1 : expected_columns[-1]] = fill
            if row_index % 4 == 2:
                image[top + 1 : bottom, expected_columns[4] + 1 : expected_columns[5]] = (225, 238, 250)

        for row_index, row in enumerate(expected_rows):
            line = 225 if row_index % 2 == 0 else 245
            cv2.line(image, (expected_columns[0], row), (expected_columns[-1], row), (line, line, line), 1)
        for column in expected_columns:
            cv2.line(image, (column, expected_rows[0]), (column, expected_rows[-1]), (245, 245, 245), 1)

        for row_index in range(len(expected_rows) - 1):
            baseline = expected_rows[row_index] + 21
            for column_index in range(len(expected_columns) - 1):
                cv2.putText(
                    image,
                    f"{row_index + 1}-{column_index + 1}",
                    (expected_columns[column_index] + 5, baseline),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.32,
                    (30, 30, 30),
                    1,
                    cv2.LINE_AA,
                )

        grid = table_pipeline.extract_screen_grid(image)

        self.assertIsNotNone(grid)
        columns, rows, _ = grid
        self.assertEqual(columns, expected_columns)
        self.assertEqual(rows, expected_rows)

    def test_embedded_spreadsheet_grid_ignores_application_chrome(self):
        height, width = 900, 1200
        image = np.full((height, width, 3), 250, dtype=np.uint8)
        for row in (40, 120, 205):
            cv2.line(image, (0, row), (width - 1, row), (90, 90, 90), 2)
        for column in range(70, 1130, 85):
            cv2.line(image, (column, 45), (column, 190), (120, 120, 120), 2)

        expected_columns = [40, 150, 280, 420, 540, 680, 820, 980, 1150]
        expected_rows = [280, 330] + [355 + 25 * index for index in range(20)]
        for row in expected_rows:
            cv2.line(
                image,
                (expected_columns[0], row),
                (expected_columns[-1], row),
                (180, 180, 180),
                1,
            )
        for column in expected_columns:
            cv2.line(
                image,
                (column, expected_rows[1]),
                (column, expected_rows[-1]),
                (180, 180, 180),
                1,
            )
        cv2.line(
            image,
            (expected_columns[0], expected_rows[0]),
            (expected_columns[0], expected_rows[1]),
            (180, 180, 180),
            1,
        )
        cv2.line(
            image,
            (expected_columns[-1], expected_rows[0]),
            (expected_columns[-1], expected_rows[1]),
            (180, 180, 180),
            1,
        )

        grid = table_pipeline.extract_embedded_spreadsheet_grid(image)

        self.assertIsNotNone(grid)
        columns, rows, _ = grid
        self.assertEqual(columns, expected_columns)
        self.assertEqual(len(rows), len(expected_rows))
        self.assertTrue(
            all(abs(actual - expected) <= 1 for actual, expected in zip(rows, expected_rows))
        )

    def test_dense_spreadsheet_recovers_low_contrast_variable_vertical_rules(self):
        height, width = 1068, 1857
        expected_columns = [
            10, 47, 92, 154, 228, 364, 430, 504, 580, 655, 730, 827,
            917, 991, 1066, 1142, 1230, 1322, 1428, 1532, 1599, 1658,
            1726, 1815,
        ]
        rows = [15, 34] + [56 + 24 * index for index in range(42)]
        rows = [row for row in rows if row < height - 10]
        if rows[-1] != 1054:
            rows.append(1054)
        image = np.full((height, width, 3), 248, dtype=np.uint8)
        for row_index in range(len(rows) - 1):
            top, bottom = rows[row_index : row_index + 2]
            fill = 247 if row_index % 2 else 252
            image[top + 1 : bottom, expected_columns[0] + 1 : expected_columns[-1]] = fill
        for row in rows:
            cv2.line(
                image,
                (expected_columns[0], row),
                (expected_columns[-1], row),
                (220, 220, 220),
                1,
            )
        for column_index, column in enumerate(expected_columns):
            for row_index in range(len(rows) - 1):
                top, bottom = rows[row_index : row_index + 2]
                fill = 247 if row_index % 2 else 252
                strength = 13 + ((row_index * 17 + column_index * 11) % 27)
                value = max(0, fill - strength)
                image[top : bottom + 1, column] = (value, value, value)

        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        recovered = table_pipeline._recover_dense_spreadsheet_columns(
            gray,
            [10, 47, 228, 430, 504, 655, 917, 1066, 1142, 1428, 1658, 1726, 1815],
            rows,
        )

        self.assertEqual(len(recovered), len(expected_columns))
        self.assertTrue(
            all(abs(actual - expected) <= 3 for actual, expected in zip(recovered, expected_columns))
        )

    def test_screen_grid_keeps_narrow_last_column_at_image_edge(self):
        expected_rows = [30 + 32 * index for index in range(18)]
        expected_columns = [30, 170, 340, 520, 720, 930, 998]
        image = self._draw_screen_table(expected_rows, expected_columns, height=620, width=1000)

        grid = table_pipeline.extract_screen_grid(image)

        self.assertIsNotNone(grid)
        columns, rows, _ = grid
        self.assertEqual(len(columns), len(expected_columns))
        self.assertTrue(all(abs(actual - expected) <= 3 for actual, expected in zip(columns, expected_columns)))
        self.assertEqual(rows, expected_rows)

    def test_screen_grid_recovers_partial_merged_header_boundary(self):
        expected_columns = [30, 150, 300, 460, 620, 780, 970]
        expected_rows = [30, 102, 166] + [230 + 32 * index for index in range(12)]
        image = self._draw_screen_table(expected_rows, expected_columns, height=620, width=1000)
        image[165:168, expected_columns[0] : expected_columns[3]] = 255

        grid = table_pipeline.extract_screen_grid(image)

        self.assertIsNotNone(grid)
        columns, rows, _ = grid
        self.assertEqual(columns, expected_columns)
        self.assertEqual(rows, expected_rows)

    def test_screen_grid_recovers_lines_under_smooth_shadow(self):
        expected_rows = [30 + 32 * index for index in range(18)]
        expected_columns = [30, 160, 310, 470, 640, 810, 970]
        image = self._draw_screen_table(expected_rows, expected_columns, height=620, width=1000)
        for y in range(expected_rows[0], expected_rows[-1] + 1):
            shade = 160 + (y % 67)
            image[y, expected_columns[2]] = (shade, shade, shade)

        grid = table_pipeline.extract_screen_grid(image)

        self.assertIsNotNone(grid)
        columns, rows, _ = grid
        self.assertEqual(len(columns), len(expected_columns))
        self.assertTrue(all(abs(actual - expected) <= 3 for actual, expected in zip(columns, expected_columns)))
        self.assertEqual(rows, expected_rows)

    def test_screen_grid_regularizes_jpeg_transition_duplicates(self):
        expected_rows = [22 + 29 * index for index in range(44)]
        expected_columns = [30, 95, 175, 270, 390, 520, 660, 820, 1000, 1190, 1390, 1600, 1810, 2020, 2215]
        image = self._draw_screen_table(expected_rows, expected_columns, height=1277, width=2217)
        for row_index in range(len(expected_rows) - 1):
            baseline = expected_rows[row_index] + 20
            for column_index in range(len(expected_columns) - 1):
                cv2.putText(
                    image,
                    f"{row_index + 1}-{column_index + 1}",
                    (expected_columns[column_index] + 3, baseline),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.3,
                    (25, 25, 25),
                    1,
                    cv2.LINE_AA,
                )
        encoded = cv2.imencode(".jpg", image, [cv2.IMWRITE_JPEG_QUALITY, 58])[1]
        compressed = cv2.imdecode(encoded, cv2.IMREAD_COLOR)

        grid = table_pipeline.extract_screen_grid(compressed)

        self.assertIsNotNone(grid)
        columns, rows, _ = grid
        self.assertEqual(len(columns), len(expected_columns))
        self.assertEqual(len(rows), len(expected_rows))
        self.assertTrue(all(abs(actual - expected) <= 3 for actual, expected in zip(columns, expected_columns)))
        self.assertTrue(all(abs(actual - expected) <= 3 for actual, expected in zip(rows, expected_rows)))

    def test_screen_grid_rejects_text_edges_that_split_real_columns(self):
        transition_columns = [
            60, 77, 240, 456, 672, 733, 840, 1056, 1236, 1380, 1407, 1425, 1522
        ]
        consistent_columns = [59, 236, 452, 668, 838, 1052, 1232, 1376]

        columns = table_pipeline._select_screen_columns(
            transition_columns,
            consistent_columns,
            width=1524,
        )

        self.assertEqual(columns, [60, 236, 452, 668, 838, 1052, 1232, 1376, 1522])

    def test_screen_grid_filters_short_repeated_text_edges_by_vertical_support(self):
        projection = np.zeros(1575, dtype=np.int32)
        real_columns = [60, 246, 512, 672, 830, 962, 1310, 1442, 1573]
        for column in real_columns:
            projection[max(0, column - 1) : min(len(projection), column + 2)] = 490
        projection[166] = 20
        projection[288] = 190

        columns = table_pipeline._filter_transition_columns_by_vertical_support(
            [60, 166, 246, 288, 512, 672, 830, 962, 1310, 1442, 1573],
            projection,
            image_height=579,
            image_width=1575,
        )

        self.assertEqual(columns, real_columns)

    def test_screen_grid_recovers_one_missing_boundary_in_dense_table(self):
        centers = [60, 161, 326, 432, 536, 642, 746, 852, 1076, 1186, 1291, 1396, 1501, 1604]
        projection = np.zeros(1660, dtype=np.int32)
        projection[956:958] = 661
        projection[60] = 750

        recovered = table_pipeline._recover_single_missing_boundary(centers, projection)

        self.assertEqual(len(recovered), 15)
        self.assertEqual(recovered[8], 956)

    def test_screen_grid_trims_decorative_frame_after_missing_row_recovery(self):
        height, width = 1350, 2400
        columns = [58, 266, 558, 812, 1128, 1482, 1732, 1976, 2182, 2342]
        rows = [14, 56, 82, 113, 139] + [168 + 29 * index for index in range(40)]
        image = np.full((height, width, 3), 203, dtype=np.uint8)
        image[15:56, columns[0] : columns[-1]] = 194
        image[57:82, columns[0] : columns[-1]] = 94
        for column in range(columns[0] + 30, columns[-1] - 20, 120):
            cv2.line(image, (column, 61), (column + 15, 77), (225, 225, 225), 3)

        trimmed = table_pipeline._trim_decorative_screen_top_frame(
            image,
            columns,
            rows,
        )

        self.assertEqual(trimmed, rows[1:])

    def test_screen_grid_keeps_regular_top_row_without_decorative_frame(self):
        height, width = 620, 1000
        columns = [30, 150, 300, 460, 620, 780, 970]
        rows = [30 + 32 * index for index in range(18)]
        image = self._draw_screen_table(rows, columns, height=height, width=width)

        self.assertEqual(
            table_pipeline._trim_decorative_screen_top_frame(image, columns, rows),
            rows,
        )

    def test_ruled_grid_request_cache_reuses_only_exact_mode_and_pixels(self):
        image = np.full((80, 120, 3), 245, dtype=np.uint8)
        first_result = ([0, 60, 119], [0, 40, 79], image.copy())
        with patch.object(
            table_pipeline,
            "_extract_ruled_grid_uncached",
            return_value=first_result,
        ) as extractor:
            table_pipeline.begin_ruled_grid_request_cache()
            try:
                first = table_pipeline.extract_ruled_grid(image, prefer_adaptive=True)
                second = table_pipeline.extract_ruled_grid(image.copy(), prefer_adaptive=True)
                table_pipeline.extract_ruled_grid(image, prefer_adaptive=False)
                changed = image.copy()
                changed[10, 10] = 0
                table_pipeline.extract_ruled_grid(changed, prefer_adaptive=True)
            finally:
                table_pipeline.end_ruled_grid_request_cache()

        self.assertEqual(extractor.call_count, 3)
        self.assertEqual(first[0], second[0])
        second[2][0, 0] = 0
        self.assertEqual(first[2][0, 0].tolist(), [245, 245, 245])

    def test_screen_grid_does_not_split_wide_column_without_a_supported_rule(self):
        centers = [60, 152, 251, 371, 565, 661, 781, 904, 1021, 1141, 1261]
        projection = np.zeros(1320, dtype=np.int32)
        projection[60] = 650
        projection[383] = 95

        recovered = table_pipeline._recover_single_missing_boundary(centers, projection)

        self.assertEqual(recovered, centers)

    def test_screen_grid_removes_bold_header_text_edge_next_to_real_rule(self):
        horizontal = np.zeros((220, 600), dtype=np.uint8)
        cv2.line(horizontal, (0, 80), (599, 80), 255, 1)
        cv2.line(horizontal, (0, 122), (599, 122), 255, 1)
        cv2.line(horizontal, (0, 164), (599, 164), 255, 1)

        rows = table_pipeline._remove_false_transition_rows(
            [30, 80, 114, 122, 164, 206],
            horizontal,
            image_width=600,
        )

        self.assertEqual(rows, [30, 80, 122, 164, 206])

    def test_screen_grid_removes_near_half_height_false_header_edge(self):
        horizontal = np.zeros((500, 1400), dtype=np.uint8)
        for row in [30, 84, 128, 163, 198, 232, 266, 301, 336, 370, 404, 439, 472]:
            cv2.line(horizontal, (0, row), (1399, row), 255, 1)

        rows = table_pipeline._remove_false_transition_rows(
            [30, 84, 113, 128, 163, 198, 232, 266, 301, 336, 370, 404, 439, 472],
            horizontal,
            image_width=1400,
        )

        self.assertEqual(rows, [30, 84, 128, 163, 198, 232, 266, 301, 336, 370, 404, 439, 472])

    def test_extract_ruled_grid_preserves_faint_spreadsheet_screenshot_columns(self):
        image = np.full((1216, 2559, 3), 255, dtype=np.uint8)
        expected_columns = [
            36, 112, 188, 264, 340, 460, 565, 674, 794, 965, 1251, 1362,
            1438, 1514, 1590, 1666, 1742, 1818, 1894, 1970, 2046, 2122,
            2198, 2274, 2350, 2426, 2502,
        ]
        expected_rows = [34, 54, 93] + list(range(113, 1174, 20))
        for column in expected_columns:
            bottom = expected_rows[-1] if column <= 2198 else 147
            cv2.line(image, (column, 2), (column, bottom), (230, 230, 230), 1)
        for row in expected_rows:
            cv2.line(image, (expected_columns[0], row), (expected_columns[-1], row), (230, 230, 230), 1)
        for row_index in range(len(expected_rows) - 1):
            baseline = (expected_rows[row_index] + expected_rows[row_index + 1]) // 2 + 4
            for column_index in range(len(expected_columns) - 1):
                if column_index % 3 != 1:
                    cv2.putText(
                        image,
                        str((row_index + 1) * (column_index + 1)),
                        (expected_columns[column_index] + 3, baseline),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        0.25,
                        (0, 0, 0),
                        1,
                        cv2.LINE_AA,
                    )

        grid = extract_ruled_grid(image)

        self.assertIsNotNone(grid)
        columns, rows, _ = grid
        self.assertEqual(columns, expected_columns)
        self.assertEqual(rows, expected_rows)

    def test_ruled_grid_prefers_adaptive_lines_over_motion_text_strokes(self):
        height, width = 420, 800
        expected_columns = [60, 160, 270, 380, 490, 600, 700, 760]
        expected_rows = [30 + 22 * index for index in range(17)]
        image = np.full((height, width, 3), 255, dtype=np.uint8)
        horizontal = np.zeros((height, width), dtype=np.uint8)
        vertical = np.zeros((height, width), dtype=np.uint8)
        for column in expected_columns:
            cv2.line(
                image,
                (column, expected_rows[0]),
                (column, expected_rows[-1]),
                (0, 0, 0),
                1,
            )
            cv2.line(
                vertical,
                (column, expected_rows[0]),
                (column, expected_rows[-1]),
                255,
                1,
            )
        for row in expected_rows:
            cv2.line(
                image,
                (expected_columns[0], row),
                (expected_columns[-1], row),
                (0, 0, 0),
                1,
            )
            cv2.line(
                horizontal,
                (expected_columns[0], row),
                (expected_columns[-1], row),
                255,
                1,
            )

        false_screen_columns = sorted(
            expected_columns
            + [205, 215, 225, 235, 245, 255, 285, 295, 305, 315, 325, 335, 345, 355]
        )
        false_screen_rows = sorted(expected_rows + [expected_rows[1] + 2])
        with (
            patch.object(
                table_pipeline,
                "extract_screen_grid",
                return_value=(false_screen_columns, false_screen_rows, image.copy()),
            ),
            patch.object(
                table_pipeline,
                "_grid_maps",
                return_value=(
                    horizontal,
                    vertical,
                    cv2.bitwise_or(horizontal, vertical),
                ),
            ),
        ):
            grid = table_pipeline.extract_ruled_grid(image, prefer_adaptive=True)

        self.assertIsNotNone(grid)
        columns, rows, _ = grid
        self.assertEqual(columns, expected_columns)
        self.assertEqual(rows, expected_rows)

    def test_spreadsheet_ruler_and_body_jointly_recover_faint_columns(self):
        image = np.full((460, 900), 248, dtype=np.uint8)
        rows = [0, 28] + list(range(78, 429, 50))
        columns = [5, 60, 170, 310, 470, 650, 780, 870]
        for row_index, (top, bottom) in enumerate(zip(rows[:-1], rows[1:])):
            image[top:bottom, :] = 244 if row_index % 2 else 252
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), 220, 1)
        for row in rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), 220, 1)

        recovered = table_pipeline._confirmed_spreadsheet_ruler_columns(image, rows)

        self.assertEqual(len(recovered), len(columns))
        self.assertTrue(
            all(abs(actual - expected) <= 2 for actual, expected in zip(recovered, columns))
        )

    def test_spreadsheet_ruler_columns_use_cropped_right_edge_as_last_boundary(self):
        image = np.full((120, 1258, 3), 248, dtype=np.uint8)
        starts = [50, 190, 409, 509, 609, 709, 859, 979]
        values = [244, 250]
        for index, (left, right) in enumerate(zip(starts, starts[1:] + [1258])):
            image[:72, left:right] = values[index % 2]

        recovered = table_pipeline.extract_spreadsheet_ruler_columns(image, 8, 72)

        self.assertEqual(len(recovered), 9)
        self.assertTrue(
            all(
                abs(actual - expected) <= 1
                for actual, expected in zip(recovered, starts + [1257])
            )
        )

    def test_spreadsheet_rulers_recover_faint_rotated_sheet_rows_and_columns(self):
        height, width = 630, 1680
        image = np.full((height, width), 245, dtype=np.uint8)
        expected_columns = [70, 173, 378, 557, 749, 1107, 1261, 1440, 1670]
        expected_rows = [40, 92, 146, 200, 253, 306, 359, 413, 466, 519, 572, 626]

        for index in range(len(expected_columns) - 1):
            left, right = expected_columns[index : index + 2]
            image[: expected_rows[0], left:right] = 238 + (index % 2) * 5
        for column in expected_columns:
            cv2.line(image, (column, 0), (column, expected_rows[0]), 185, 1)

        for row_index in range(len(expected_rows) - 1):
            top, bottom = expected_rows[row_index : row_index + 2]
            base = 229 + (row_index % 2) * 10
            for column_index in range(len(expected_columns) - 1):
                left, right = expected_columns[column_index : column_index + 2]
                image[top:bottom, left:right] = base + (column_index % 2) * 4
        for row in expected_rows:
            cv2.line(image, (0, row), (expected_columns[0], row), 175, 1)

        recovered = table_pipeline._recover_spreadsheet_ruler_grid(
            image,
            [8, 70, 173, 749, 1440, 1670],
            [43, 88, 200, 253, 306, 359, 413, 466, 519, 572, 626],
        )

        self.assertIsNotNone(recovered)
        columns, rows = recovered
        self.assertEqual(columns, expected_columns)
        self.assertEqual(len(rows), len(expected_rows))
        self.assertTrue(
            all(abs(actual - expected) <= 1 for actual, expected in zip(rows, expected_rows))
        )

    def test_spreadsheet_ruler_removes_body_text_false_columns_without_equal_width_guess(self):
        height, width = 1180, 2035
        image = np.full((height, width), 248, dtype=np.uint8)
        expected_columns = [
            45, 94, 165, 247, 401, 474, 560, 644, 731, 814, 924, 1027,
            1110, 1196, 1280, 1381, 1485, 1603, 1722, 1800, 1864, 1941, 2034,
        ]
        expected_rows = [22]
        while expected_rows[-1] < 1175:
            gap = (26, 27, 28)[(len(expected_rows) - 1) % 3]
            expected_rows.append(min(1175, expected_rows[-1] + gap))

        for index, (left, right) in enumerate(
            zip(expected_columns[:-1], expected_columns[1:])
        ):
            image[: expected_rows[0], left:right] = 235 + (index % 2) * 12
        for column in expected_columns[:-1]:
            cv2.line(image, (column, 0), (column, expected_rows[0] - 1), 170, 1)
        for row_index, (top, bottom) in enumerate(zip(expected_rows[:-1], expected_rows[1:])):
            for column_index, (left, right) in enumerate(
                zip(expected_columns[:-1], expected_columns[1:])
            ):
                image[top:bottom, left:right] = (
                    238 + (row_index % 2) * 6 + (column_index % 2) * 2
                )
        for row in expected_rows:
            cv2.line(image, (expected_columns[0], row), (expected_columns[-1], row), 180, 1)
        for column in expected_columns:
            cv2.line(image, (column, expected_rows[0]), (column, expected_rows[-1]), 220, 1)

        # 这些竖线只出现在正文，模拟密集小字笔画。顶部 A-V 标尺并不支持它们。
        false_columns = [5, 1335, 1514, 1678]
        for column in false_columns:
            cv2.line(image, (column, expected_rows[0]), (column, expected_rows[-1]), 100, 1)
        over_split_columns = sorted(expected_columns[:-1] + false_columns)

        recovered = table_pipeline._recover_spreadsheet_ruler_grid(
            image,
            over_split_columns,
            expected_rows,
        )

        self.assertIsNotNone(recovered)
        columns, rows = recovered
        self.assertEqual(columns, expected_columns)
        self.assertEqual(len(rows), len(expected_rows))
        self.assertTrue(
            all(abs(actual - expected) <= 2 for actual, expected in zip(rows, expected_rows))
        )

        two_false_columns = sorted(expected_columns + false_columns[1:3])
        recovered = table_pipeline._recover_spreadsheet_ruler_grid(
            image,
            two_false_columns,
            expected_rows,
        )

        self.assertIsNotNone(recovered)
        columns, rows = recovered
        self.assertEqual(columns, expected_columns)
        self.assertEqual(len(rows), len(expected_rows))

    def test_real_dense_low_resolution_spreadsheet_keeps_all_22_columns(self):
        image_path = _real_fixture_path(r"excel\base_03.png")
        if image_path is None or not image_path.exists():
            self.skipTest("base_03.png real regression fixture is unavailable")

        image = cv2.imdecode(np.fromfile(str(image_path), dtype=np.uint8), cv2.IMREAD_COLOR)
        recovered = table_pipeline.extract_screen_grid(image)

        self.assertIsNotNone(recovered)
        columns, rows, _ = recovered
        self.assertEqual(len(columns) - 1, 22)
        self.assertEqual(len(rows) - 1, 43)

    def test_page_rulers_and_headers_recover_oversplit_dense_spreadsheet(self):
        ocr_backend._load_runtime()
        image = np.full((799, 1388, 3), 248, dtype=np.uint8)
        candidate_columns = [
            27, 61, 74, 112, 128, 167, 200, 223, 271, 286, 322, 355,
            380, 436, 494, 552, 628, 667, 696, 754, 812, 846, 869, 937,
            954, 1007, 1088, 1117, 1132, 1172, 1220, 1266, 1283, 1319,
            1386,
        ]
        expected_columns = [
            27, 61, 112, 167, 271, 322, 380, 436, 494, 552, 628, 696,
            754, 812, 869, 937, 1007, 1088, 1172, 1220, 1266, 1319, 1386,
        ]
        rows = [14 + index * 18 for index in range(44)]
        ruler_centers = [
            45.5, 86.0, 138.0, 219.0, 296.5, 351.0, 409.0, 466.5,
            524.0, 590.5, 662.0, 725.5, 783.0, 840.5, 904.0, 973.0,
            1048.5, 1129.5, 1195.5, 1243.5, 1293.0, 1353.5,
        ]
        labels = [chr(ord("A") + index) for index in range(len(ruler_centers))]
        labels[8] = "1"   # PP-OCRv6 Small can read the I ruler as one.
        labels[14] = "0"  # The O ruler can be read as zero.
        labels[16] = "a"  # The Q ruler can be read as a lower-case a.
        boxes = []
        texts = []
        scores = []
        for index, center in enumerate(ruler_centers):
            boxes.append(
                np.asarray(
                    [[center - 3, 0], [center + 3, 0], [center + 3, 14], [center - 3, 14]],
                    dtype=np.float32,
                )
            )
            texts.append(labels[index])
            scores.append(0.98)
            boxes.append(
                np.asarray(
                    [[center - 8, 17], [center + 8, 17], [center + 8, 30], [center - 8, 30]],
                    dtype=np.float32,
                )
            )
            texts.append(f"H{index + 1}")
            scores.append(0.99)

        recovered = ocr_backend._recover_dense_spreadsheet_grid_from_page_rulers(
            image,
            candidate_columns,
            rows,
            boxes,
            texts,
            scores,
        )

        self.assertIsNotNone(recovered)
        columns, recovered_rows = recovered
        self.assertEqual(columns, expected_columns)
        self.assertEqual(recovered_rows, rows)

        # The ruler alone is not sufficient: missing independent header-row
        # support must keep the uncertain geometry on the conservative path.
        ruler_with_sparse_headers = [
            index
            for index in range(len(boxes))
            if index % 2 == 0 or index < 10
        ]
        rejected = ocr_backend._recover_dense_spreadsheet_grid_from_page_rulers(
            image,
            candidate_columns,
            rows,
            [boxes[index] for index in ruler_with_sparse_headers],
            [texts[index] for index in ruler_with_sparse_headers],
            [scores[index] for index in ruler_with_sparse_headers],
        )
        self.assertIsNone(rejected)

    def test_embedded_spreadsheet_ruler_recovers_nine_faint_columns(self):
        height, width = 1000, 1200
        image = np.full((height, width), 18, dtype=np.uint8)
        expected_columns = [40, 135, 250, 380, 520, 670, 810, 940, 1060, 1199]
        ruler_top, ruler_bottom = 180, 210
        expected_rows = list(range(ruler_bottom, 811, 60))

        for index, (left, right) in enumerate(
            zip(expected_columns[:-1], expected_columns[1:])
        ):
            image[ruler_top:ruler_bottom, left:right] = 235 + (index % 2) * 14
            cv2.line(image, (left, ruler_top), (left, ruler_bottom), 170, 1)
        for row_index, (top, bottom) in enumerate(zip(expected_rows[:-1], expected_rows[1:])):
            for column_index, (left, right) in enumerate(
                zip(expected_columns[:-1], expected_columns[1:])
            ):
                image[top:bottom, left:right] = (
                    220 + (row_index % 2) * 16 + (column_index % 2) * 6
                )
        for row in [ruler_top] + expected_rows:
            cv2.line(image, (expected_columns[0], row), (expected_columns[-1], row), 175, 1)
        cv2.line(
            image,
            (expected_columns[0], ruler_top),
            (expected_columns[0], expected_rows[-1]),
            175,
            1,
        )

        recovered = table_pipeline.extract_ruled_grid(
            cv2.cvtColor(image, cv2.COLOR_GRAY2BGR),
            prefer_adaptive=True,
        )

        self.assertIsNotNone(recovered)
        columns, rows, _ = recovered
        self.assertEqual(len(columns) - 1, 9)
        self.assertEqual(len(rows) - 1, len(expected_rows))
        self.assertTrue(
            all(abs(actual - expected) <= 2 for actual, expected in zip(columns, expected_columns))
        )

    def test_row_centers_become_exact_cell_boundaries(self):
        self.assertEqual(
            ocr_backend._row_center_boundaries([12.0, 48.0, 88.0, 128.0], 150),
            [0, 30, 68, 108, 148],
        )

    def test_spreadsheet_ruler_rebuild_splits_fused_last_columns(self):
        columns = [50, 190, 409, 509, 609, 709, 859, 979, 1257]
        row_centers = [12.0, 48.0, 88.0, 128.0]
        rows = ocr_backend._row_center_boundaries(row_centers, 150)
        boxes = []
        texts = []
        scores = []

        def add(column: int, row: int, value: str, column_span: int = 1):
            left = columns[column] + 4
            right = columns[column + column_span] - 4
            top = rows[row] + 4
            bottom = rows[row + 1] - 4
            boxes.append([[left, top], [right, top], [right, bottom], [left, bottom]])
            texts.append(value)
            scores.append(0.99)

        for column, label in enumerate("ABCDEFGH"):
            add(column, 0, label)
        headers = ["任务编号", "任务名称", "优先级", "进度", "负责人", "计划完成"]
        for column, value in enumerate(headers):
            add(column, 1, value)
        add(6, 1, "当前状态 风险说明", 2)
        for row, prefix in ((2, "001"), (3, "002")):
            values = [prefix, "OCR验证", "高", "50%", "李娜", "2026-08-09"]
            for column, value in enumerate(values):
                add(column, row, value)
            add(6, row, "进行中 无", 2)

        raw_spatial = [
            ["A", "B", "C", "D", "E", "F", "G H", ""],
            ["任务编号", "任务名称", "优先级", "进度", "负责人", "计划完成", "当前状态 风险说明", ""],
            ["001", "OCR验证", "高", "50%", "李娜", "2026-08-09", "进行中 无", ""],
            ["002", "OCR验证", "高", "50%", "李娜", "2026-08-09", "进行中 无", ""],
        ]
        geometry = {"row_centers": row_centers}
        with patch.object(
            table_pipeline,
            "extract_spreadsheet_ruler_columns",
            return_value=columns,
        ):
            rebuilt = ocr_backend._rebuild_spreadsheet_grid_from_ruler(
                np.full((150, 1258, 3), 255, dtype=np.uint8),
                raw_spatial,
                geometry,
                np.asarray(boxes, dtype=float),
                texts,
                scores,
            )

        self.assertIsNotNone(rebuilt)
        rebuilt_grid, _, rebuilt_geometry = rebuilt
        trimmed, _ = ocr_backend._strip_spreadsheet_ui_headers(rebuilt_grid, None)
        self.assertEqual(len(trimmed[0]), 8)
        self.assertEqual(trimmed[0][-2:], ["当前状态", "风险说明"])
        self.assertEqual(trimmed[1][-2:], ["进行中", "无"])
        self.assertEqual(rebuilt_geometry["first_structured_row"], 1)

    def test_spreadsheet_ruler_rebuild_may_use_near_identity_source_before_warp(self):
        spatial = [
            ["A", "B", "C", "D"],
            ["编号", "名称", "状态 备注", ""],
            ["1", "设备", "正常 无", ""],
            ["2", "仪器", "复核 无", ""],
        ]
        geometry = {"row_centers": [10.0, 30.0, 50.0, 70.0]}
        target = np.full((80, 399, 3), 255, dtype=np.uint8)
        source = np.full((81, 400, 3), 255, dtype=np.uint8)
        columns = [0, 100, 200, 300, 399]
        rows = [0, 20, 40, 60, 79]
        boxes = []
        texts = []
        for row, values in enumerate((
            ["A", "B", "C", "D"],
            ["编号", "名称", "状态", "备注"],
            ["1", "设备", "正常", "无"],
            ["2", "仪器", "复核", "无"],
        )):
            for column, value in enumerate(values):
                boxes.append([
                    [columns[column] + 3, rows[row] + 3],
                    [columns[column + 1] - 3, rows[row] + 3],
                    [columns[column + 1] - 3, rows[row + 1] - 3],
                    [columns[column] + 3, rows[row + 1] - 3],
                ])
                texts.append(value)
        with patch.object(
            table_pipeline,
            "extract_spreadsheet_ruler_columns",
            side_effect=lambda image, expected, bottom: [0, 100, 200, 300, 399]
            if image.shape[1] == 400
            else [],
        ):
            rebuilt = ocr_backend._rebuild_spreadsheet_grid_from_ruler(
                target,
                spatial,
                geometry,
                np.asarray(boxes, dtype=float),
                texts,
                [0.99] * len(texts),
                ruler_image=source,
            )

        self.assertIsNotNone(rebuilt)

    def test_screen_grid_cleanup_enhances_small_low_contrast_text(self):
        image = np.full((500, 900, 3), 240, dtype=np.uint8)
        for row in range(40, 461, 20):
            cv2.line(image, (30, row), (870, row), (210, 210, 210), 1)
        for column in range(30, 871, 70):
            cv2.line(image, (column, 40), (column, 460), (210, 210, 210), 1)
        cv2.putText(
            image,
            "123",
            (40, 70),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.5,
            (180, 180, 180),
            1,
            cv2.LINE_AA,
        )

        grid = extract_ruled_grid(image)

        self.assertIsNotNone(grid)
        _, _, cleaned = grid
        source_gray = cv2.cvtColor(image[64:76, 35:75], cv2.COLOR_BGR2GRAY)
        cleaned_gray = cv2.cvtColor(cleaned[64:76, 35:75], cv2.COLOR_BGR2GRAY)
        self.assertGreater(
            int(cleaned_gray.max()) - int(cleaned_gray.min()),
            int(source_gray.max()) - int(source_gray.min()),
        )
        self.assertGreater(int(cleaned[70, 104].mean()), 250)

    def test_extract_ruled_grid_finds_cell_boundaries_and_removes_lines(self):
        image = np.full((180, 300, 3), 255, dtype=np.uint8)
        expected_columns = [8, 92, 196, 292]
        expected_rows = [8, 58, 112, 172]
        for x in expected_columns:
            cv2.line(image, (x, 8), (x, 172), (0, 0, 0), 2)
        for y in expected_rows:
            cv2.line(image, (8, y), (292, y), (0, 0, 0), 2)

        grid = extract_ruled_grid(image)

        self.assertIsNotNone(grid)
        columns, rows, cleaned = grid
        self.assertEqual(len(columns), len(expected_columns))
        self.assertEqual(len(rows), len(expected_rows))
        self.assertTrue(all(abs(actual - expected) <= 2 for actual, expected in zip(columns, expected_columns)))
        self.assertTrue(all(abs(actual - expected) <= 2 for actual, expected in zip(rows, expected_rows)))
        self.assertGreater(int(cleaned[58, 120].mean()), 245)

    def test_extract_ruled_grid_recovers_shadowed_header_rule_and_cropped_bottom(self):
        height, width = 242, 620
        image = np.full((height, width, 3), 245, dtype=np.uint8)
        columns = [4, 72, 180, 330, 470, 615]
        expected_rows = [3 + 22 * index for index in range(11)] + [height - 1]
        for column in columns:
            cv2.line(image, (column, expected_rows[0]), (column, height - 1), (20, 20, 20), 2)
        for index, row in enumerate(expected_rows[:-1]):
            if index == 1:
                cv2.line(image, (4, row), (260, row), (35, 35, 35), 2)
            else:
                cv2.line(image, (4, row), (615, row), (20, 20, 20), 2)

        grid = extract_ruled_grid(image)

        self.assertIsNotNone(grid)
        _, actual_rows, _ = grid
        self.assertEqual(len(actual_rows), len(expected_rows))
        self.assertTrue(
            all(abs(actual - expected) <= 2 for actual, expected in zip(actual_rows, expected_rows))
        )

    def test_extract_ruled_grid_recovers_very_faint_short_header_rule(self):
        height, width = 220, 620
        image = np.full((height, width, 3), 245, dtype=np.uint8)
        columns = [4, 92, 210, 350, 480, 615]
        expected_rows = [3, 25, 47, 69, 91, 113, 135, 157, 179, 201]
        for column in columns:
            cv2.line(image, (column, expected_rows[0]), (column, expected_rows[-1]), (20, 20, 20), 2)
        for index, row in enumerate(expected_rows):
            if index == 2:
                cv2.line(image, (4, row), (104, row), (35, 35, 35), 2)
            else:
                cv2.line(image, (4, row), (615, row), (20, 20, 20), 2)

        grid = extract_ruled_grid(image)

        self.assertIsNotNone(grid)
        _, actual_rows, _ = grid
        self.assertEqual(len(actual_rows), len(expected_rows))
        self.assertTrue(
            all(abs(actual - expected) <= 2 for actual, expected in zip(actual_rows, expected_rows))
        )

    def test_regular_missing_row_recovery_handles_a_long_fading_run(self):
        projection = np.zeros(910, dtype=np.int32)
        projection[327] = 913
        for row, support in ((516, 291), (580, 218), (643, 184), (706, 159), (771, 133)):
            projection[row] = support

        recovered = table_pipeline._recover_regular_missing_boundaries(
            [7, 141, 198, 264, 327, 390, 454, 902],
            projection,
            1142,
            maximum_multiple=8,
            minimum_support_ratio=0.10,
            minimum_global_peak_ratio=0.10,
        )

        for expected in (516, 580, 643, 706, 771):
            self.assertTrue(any(abs(actual - expected) <= 4 for actual in recovered))
        self.assertFalse(any(820 <= actual <= 860 for actual in recovered))

    def test_regular_missing_row_recovery_handles_thirteen_row_shadow_run(self):
        projection = np.zeros(1600, dtype=np.int32)
        existing = [496, 544, 592, 640, 688, 1315, 1364, 1412, 1462, 1513]
        for row in existing:
            projection[row] = 900
        expected = [int(round(688 + (1315 - 688) * part / 13)) for part in range(1, 13)]
        for row in expected:
            projection[row] = 180

        recovered = table_pipeline._recover_regular_missing_boundaries(
            existing,
            projection,
            1600,
            maximum_multiple=16,
            minimum_support_ratio=0.10,
            minimum_global_peak_ratio=0.10,
        )

        self.assertEqual(len(recovered), len(existing) + len(expected))
        for row in expected:
            self.assertTrue(any(abs(actual - row) <= 4 for actual in recovered))

    def test_regular_missing_row_recovery_requires_line_evidence_in_large_gap(self):
        projection = np.zeros(1600, dtype=np.int32)
        existing = [496, 544, 592, 640, 688, 1315, 1364, 1412, 1462, 1513]
        for row in existing:
            projection[row] = 900

        recovered = table_pipeline._recover_regular_missing_boundaries(
            existing,
            projection,
            1600,
            maximum_multiple=16,
            minimum_support_ratio=0.10,
            minimum_global_peak_ratio=0.10,
        )

        self.assertEqual(recovered, existing)

    def test_body_consensus_can_repair_a_seven_column_shadow_candidate(self):
        candidate = [0, 204, 478, 666, 848, 1096, 1614, 1791]
        body_consensus = [0, 204, 478, 666, 848, 1096, 1300, 1456, 1614, 1791]

        selected = table_pipeline._prefer_body_consensus_columns(candidate, body_consensus)

        self.assertEqual(selected, body_consensus)

    def test_body_consensus_does_not_replace_an_unrelated_grid(self):
        candidate = [0, 204, 478, 666, 848, 1096, 1614, 1791]
        unrelated = [15, 125, 285, 455, 635, 825, 1015, 1225, 1455, 1780]

        selected = table_pipeline._prefer_body_consensus_columns(candidate, unrelated)

        self.assertEqual(selected, candidate)

    def test_vertical_rule_window_coverage_matches_scalar_reference(self):
        rng = np.random.default_rng(20260816)
        vertical = rng.integers(0, 2, size=(17, 29), dtype=np.uint8) * 255
        expected = np.asarray(
            [
                np.mean(
                    np.any(
                        vertical[:, max(0, position - 3) : min(29, position + 4)]
                        > 0,
                        axis=1,
                    )
                )
                for position in range(29)
            ]
        )

        actual = table_pipeline._vertical_rule_window_coverage(vertical)

        np.testing.assert_array_equal(actual, expected)

    def test_regular_row_extension_recovers_fading_table_tail_with_line_evidence(self):
        projection = np.zeros(848, dtype=np.int32)
        rows = [140, 192, 254, 314, 373, 432, 492, 551, 610]
        for row in rows:
            projection[row] = 512
        projection[670] = 413
        projection[730] = 335

        recovered = table_pipeline._extend_regular_boundaries_with_evidence(
            rows,
            projection,
            1183,
            0,
            780,
        )

        self.assertTrue(any(abs(row - 670) <= 2 for row in recovered))
        self.assertTrue(any(abs(row - 730) <= 2 for row in recovered))
        self.assertFalse(any(row > 740 for row in recovered))

    def test_regular_row_extension_accepts_a_supported_boundary_at_the_limit(self):
        projection = np.zeros(800, dtype=np.int32)
        rows = [542, 585, 628, 671, 714]
        for row in rows:
            projection[row] = 400
        projection[756] = 167

        recovered = table_pipeline._extend_regular_boundaries_with_evidence(
            rows,
            projection,
            1187,
            95,
            757,
        )

        self.assertTrue(any(abs(row - 756) <= 1 for row in recovered))

    def test_regular_row_extension_follows_warped_rules_beyond_vertical_span(self):
        projection = np.zeros(510, dtype=np.int32)
        rows = [76, 102, 137, 170, 202, 235, 268, 301, 334, 367, 400, 434]
        for row in [10, 36] + rows + [467, 501]:
            projection[row] = 420

        recovered = table_pipeline._extend_regular_boundaries_with_evidence(
            rows,
            projection,
            549,
            0,
            509,
        )

        self.assertTrue(all(any(abs(actual - expected) <= 1 for actual in recovered) for expected in (10, 36, 467, 501)))

    def test_grid_coverage_detects_adjacent_supported_row_outside_certificate(self):
        image = np.full((360, 600, 3), 245, dtype=np.uint8)
        columns = [50, 150, 250, 350, 450, 550]
        all_rows = [40, 80, 120, 160, 200, 240, 280, 320]
        for row in all_rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (20, 20, 20), 2)
        for column in columns:
            cv2.line(image, (column, all_rows[0]), (column, all_rows[-1]), (20, 20, 20), 2)

        self.assertTrue(
            table_pipeline.grid_has_excluded_supported_rows(
                image,
                columns,
                all_rows[:-1],
            )
        )
        self.assertFalse(
            table_pipeline.grid_has_excluded_supported_rows(
                image,
                columns,
                all_rows,
            )
        )

    def test_grid_coverage_ignores_disconnected_empty_bottom_frame_band(self):
        image = np.full((360, 600, 3), 245, dtype=np.uint8)
        columns = [50, 150, 250, 350, 450, 550]
        rows = [40, 80, 120, 160, 200, 240, 280]
        for row in rows + [320]:
            cv2.line(image, (columns[0], row), (columns[-1], row), (20, 20, 20), 2)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (20, 20, 20), 2)

        self.assertFalse(
            table_pipeline.grid_has_excluded_supported_rows(image, columns, rows)
        )

    def test_grid_coverage_keeps_disconnected_bottom_band_with_visible_text(self):
        image = np.full((360, 600, 3), 245, dtype=np.uint8)
        columns = [50, 150, 250, 350, 450, 550]
        rows = [40, 80, 120, 160, 200, 240, 280]
        for row in rows + [320]:
            cv2.line(image, (columns[0], row), (columns[-1], row), (20, 20, 20), 2)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (20, 20, 20), 2)
        cv2.putText(image, "VISIBLE", (180, 307), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (20, 20, 20), 2)

        self.assertTrue(
            table_pipeline.grid_has_excluded_supported_rows(image, columns, rows)
        )

    def test_grid_coverage_ignores_disconnected_empty_top_frame_band(self):
        image = np.full((360, 600, 3), 245, dtype=np.uint8)
        columns = [50, 150, 250, 350, 450, 550]
        rows = [80, 120, 160, 200, 240, 280, 320]
        for row in [40] + rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (20, 20, 20), 2)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (20, 20, 20), 2)
        cv2.line(image, (80, 43), (230, 77), (210, 210, 210), 1)

        self.assertFalse(
            table_pipeline.grid_has_excluded_supported_rows(image, columns, rows)
        )

    def test_grid_coverage_ignores_gradient_empty_frame_band(self):
        image = np.full((360, 600, 3), 245, dtype=np.uint8)
        columns = [50, 150, 250, 350, 450, 550]
        rows = [40, 80, 120, 160, 200, 240, 280]
        for row in rows + [320]:
            cv2.line(image, (columns[0], row), (columns[-1], row), (20, 20, 20), 2)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (20, 20, 20), 2)
        gradient = np.linspace(0, -28, 40, dtype=np.int16)[:, None, None]
        image[280:320] = np.clip(
            image[280:320].astype(np.int16) + gradient,
            0,
            255,
        ).astype(np.uint8)
        cv2.line(image, (columns[0], 320), (columns[-1], 320), (20, 20, 20), 2)

        self.assertFalse(
            table_pipeline.grid_has_excluded_supported_rows(image, columns, rows)
        )

    def test_grid_coverage_ignores_horizontal_texture_in_empty_frame_band(self):
        image = np.full((360, 600, 3), 245, dtype=np.uint8)
        columns = [50, 150, 250, 350, 450, 550]
        rows = [40, 80, 120, 160, 200, 240, 280]
        for row in rows + [320]:
            cv2.line(image, (columns[0], row), (columns[-1], row), (20, 20, 20), 2)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (20, 20, 20), 2)
        for y, tone in ((290, 205), (299, 195), (308, 210)):
            cv2.line(image, (85, y), (515, y), (tone, tone, tone), 1)

        self.assertFalse(
            table_pipeline.grid_has_excluded_supported_rows(image, columns, rows)
        )

    def test_grid_coverage_ignores_low_contrast_decorative_slivers(self):
        image = np.full((360, 600, 3), 245, dtype=np.uint8)
        columns = [50, 150, 250, 350, 450, 550]
        rows = [40, 80, 120, 160, 200, 240, 280]
        for row in rows + [320]:
            cv2.line(image, (columns[0], row), (columns[-1], row), (20, 20, 20), 2)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (20, 20, 20), 2)
        cv2.line(image, (412, 287), (409, 307), (234, 234, 234), 2, cv2.LINE_AA)
        cv2.line(image, (438, 292), (435, 308), (233, 233, 233), 2, cv2.LINE_AA)

        self.assertFalse(
            table_pipeline.grid_has_excluded_supported_rows(image, columns, rows)
        )

    def test_grid_coverage_keeps_faint_text_on_gradient_frame_band(self):
        image = np.full((360, 600, 3), 245, dtype=np.uint8)
        columns = [50, 150, 250, 350, 450, 550]
        rows = [40, 80, 120, 160, 200, 240, 280]
        for row in rows + [320]:
            cv2.line(image, (columns[0], row), (columns[-1], row), (20, 20, 20), 2)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (20, 20, 20), 2)
        gradient = np.linspace(0, -28, 40, dtype=np.int16)[:, None, None]
        image[280:320] = np.clip(
            image[280:320].astype(np.int16) + gradient,
            0,
            255,
        ).astype(np.uint8)
        cv2.line(image, (columns[0], 320), (columns[-1], 320), (20, 20, 20), 2)
        cv2.putText(
            image,
            "A1",
            (210, 307),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.55,
            (155, 155, 155),
            1,
        )

        self.assertTrue(
            table_pipeline.grid_has_excluded_supported_rows(image, columns, rows)
        )

    def test_grid_coverage_keeps_barely_visible_text_on_frame_band(self):
        image = np.full((360, 600, 3), 245, dtype=np.uint8)
        columns = [50, 150, 250, 350, 450, 550]
        rows = [40, 80, 120, 160, 200, 240, 280]
        for row in rows + [320]:
            cv2.line(image, (columns[0], row), (columns[-1], row), (20, 20, 20), 2)
        for column in columns:
            cv2.line(image, (column, rows[0]), (column, rows[-1]), (20, 20, 20), 2)
        cv2.putText(
            image,
            "7",
            (260, 307),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.55,
            (220, 220, 220),
            1,
        )

        self.assertTrue(
            table_pipeline.grid_has_excluded_supported_rows(image, columns, rows)
        )

    def test_real_strong_perspective_grid_keeps_the_last_data_row(self):
        image_path = _real_fixture_path(
            r"excel\生活场景表格测试-20260808\strong_perspective\L029_strong_perspective.jpg"
        )
        if image_path is None or not image_path.exists():
            self.skipTest("L029 real regression fixture is unavailable")

        image = cv2.imdecode(
            np.fromfile(str(image_path), dtype=np.uint8),
            cv2.IMREAD_COLOR,
        )
        rectified, _ = rectify_table_image(image)
        recovered = table_pipeline.extract_ruled_grid(rectified, prefer_adaptive=True)

        self.assertIsNotNone(recovered)
        columns, rows, _ = recovered
        self.assertEqual(len(columns) - 1, 7)
        self.assertEqual(len(rows) - 1, 15)

    def test_irregular_column_recovery_requires_independent_full_height_support(self):
        projection = np.zeros(1153, dtype=np.int32)
        projection[125] = 421
        projection[308] = 303
        projection[359] = 38
        recovered = table_pipeline._recover_supported_irregular_boundaries(
            [0, 125, 449, 542, 634, 726, 941, 1037, 1152],
            [125, 308, 359, 449, 542, 634, 726, 941, 1037],
            projection,
            886,
        )

        self.assertIn(308, recovered)
        self.assertNotIn(359, recovered)

    def test_irregular_column_recovery_accepts_an_exceptionally_strong_motion_rule(self):
        projection = np.zeros(1152, dtype=np.int32)
        for position, support in {
            128: 386,
            309: 325,
            448: 288,
            542: 325,
            632: 288,
            723: 286,
            937: 283,
            1034: 349,
        }.items():
            projection[position] = support

        recovered = table_pipeline._recover_supported_irregular_boundaries(
            [128, 309, 542, 1034],
            [128, 309, 448, 542, 632, 723, 937, 1034],
            projection,
            881,
        )

        self.assertIn(448, recovered)

    def test_irregular_column_recovery_follows_a_perspective_rule_per_row(self):
        height, width = 600, 800
        vertical = np.zeros((height, width), dtype=np.uint8)
        rows = [40, 90, 140, 190, 240, 290, 340, 390, 440, 490, 540]
        for row_index, (top, bottom) in enumerate(zip(rows, rows[1:])):
            x = 230 - row_index
            cv2.line(vertical, (x, top + 3), (x, bottom - 3), 255, 2)
        projection = np.count_nonzero(vertical, axis=0)

        recovered = table_pipeline._recover_supported_irregular_boundaries(
            [120, 340, 500, 650, 760],
            [120, 225, 340, 500, 650, 760],
            projection,
            height,
            vertical=vertical,
            rows=rows,
        )

        self.assertIn(225, recovered)

    def test_real_strong_perspective_grid_keeps_all_seven_columns(self):
        image_path = _real_fixture_path(
            r"excel\生活场景表格测试-20260808\strong_perspective\L012_strong_perspective.jpg"
        )
        if image_path is None or not image_path.exists():
            self.skipTest("L012 real regression fixture is unavailable")

        image = cv2.imdecode(
            np.fromfile(str(image_path), dtype=np.uint8),
            cv2.IMREAD_COLOR,
        )
        rectified, _ = rectify_table_image(image)
        recovered = table_pipeline.extract_ruled_grid(rectified, prefer_adaptive=True)

        self.assertIsNotNone(recovered)
        columns, rows, _ = recovered
        self.assertEqual(len(columns) - 1, 7)
        self.assertEqual(len(rows) - 1, 10)

    def test_page_frame_columns_are_trimmed_by_missing_horizontal_grid_support(self):
        image = np.full((600, 800, 3), 245, dtype=np.uint8)
        horizontal = np.zeros((600, 800), dtype=np.uint8)
        columns = [0, 100, 300, 500, 700, 799]
        rows = [100, 180, 260, 340, 420, 500]
        for row in rows:
            cv2.line(horizontal, (100, row), (700, row), 255, 2)
        for row in rows[:-1]:
            for left in (100, 300, 500):
                cv2.putText(
                    image,
                    "DATA",
                    (left + 30, row + 45),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.6,
                    (20, 20, 20),
                    2,
                )

        recovered = table_pipeline._trim_sparse_page_edge_columns(
            image,
            columns,
            horizontal,
        )

        self.assertEqual(recovered, [100, 300, 500, 700])

    def test_trailing_page_frame_row_is_trimmed_by_missing_vertical_grid_support(self):
        image = np.full((881, 800, 3), 245, dtype=np.uint8)
        vertical = np.zeros((881, 800), dtype=np.uint8)
        rows = [100, 150, 200, 250, 300, 350, 400, 450, 500, 550, 600, 650, 700, 850]
        for column in [100, 300, 500, 700]:
            cv2.line(vertical, (column, 100), (column, 700), 255, 2)
        for row in rows[:-2]:
            cv2.putText(
                image,
                "DATA",
                (130, row + 35),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.5,
                (20, 20, 20),
                1,
            )

        recovered = table_pipeline._trim_sparse_trailing_page_row(
            image,
            rows,
            vertical,
        )

        self.assertEqual(recovered, rows[:-1])

    def test_dark_trailing_page_margin_is_trimmed_by_line_support(self):
        image = np.full((852, 1169, 3), 238, dtype=np.uint8)
        image[755:850] = 185
        vertical = np.zeros((852, 1169), dtype=np.uint8)
        rows = [100, 148, 200, 250, 300, 351, 402, 452, 502, 553, 604, 654, 704, 755, 850]
        for column in [120, 303, 445, 539, 633, 728, 950, 1050]:
            cv2.line(vertical, (column, 100), (column, 755), 255, 2)

        recovered = table_pipeline._trim_sparse_trailing_page_row(
            image,
            rows,
            vertical,
        )

        self.assertEqual(recovered, rows[:-1])

    def test_oversized_last_row_keeps_grid_when_vertical_rules_continue(self):
        image = np.full((852, 1169, 3), 238, dtype=np.uint8)
        vertical = np.zeros((852, 1169), dtype=np.uint8)
        rows = [100, 148, 200, 250, 300, 351, 402, 452, 502, 553, 604, 654, 704, 755, 850]
        for column in [120, 303, 445, 539, 633, 728, 950, 1050]:
            cv2.line(vertical, (column, 100), (column, 850), 255, 2)

        recovered = table_pipeline._trim_sparse_trailing_page_row(
            image,
            rows,
            vertical,
        )

        self.assertEqual(recovered, rows)

    def test_narrow_empty_trailing_frame_is_trimmed(self):
        image = np.full((466, 800, 3), 240, dtype=np.uint8)
        vertical = np.zeros((466, 800), dtype=np.uint8)
        rows = [50, 100, 150, 200, 250, 300, 350, 400, 450, 465]
        for column in [50, 250, 450, 650, 750]:
            cv2.line(vertical, (column, rows[0]), (column, rows[-1]), 255, 2)

        recovered = table_pipeline._trim_sparse_trailing_page_row(
            image,
            rows,
            vertical,
        )

        self.assertEqual(recovered, rows[:-1])

    def test_disconnected_outer_frame_cells_are_trimmed_without_dropping_header(self):
        image = np.full((520, 920, 3), 235, dtype=np.uint8)
        horizontal = np.zeros((520, 920), dtype=np.uint8)
        vertical = np.zeros((520, 920), dtype=np.uint8)
        columns = [5, 25, 125, 325, 525, 725, 895, 915]
        rows = [2, 24, 52, 92, 132, 172, 212, 252, 292, 332, 372, 412, 452, 490, 517]

        cv2.rectangle(image, (5, 2), (915, 517), (35, 35, 35), 2)
        for row in rows[2:-2]:
            cv2.line(horizontal, (25, row), (895, row), 255, 2)
        for column in columns[1:-1]:
            cv2.line(vertical, (column, rows[2]), (column, rows[-3]), 255, 2)
        vertical[rows[2] : rows[3], :] = 0

        actual_columns, actual_rows = table_pipeline._trim_disconnected_outer_frame_cells(
            image,
            horizontal,
            vertical,
            columns,
            rows,
        )

        self.assertEqual(actual_columns, columns[1:-1])
        self.assertEqual(actual_rows, rows[2:-2])

    def test_narrow_real_edge_column_keeps_grid_when_row_rules_cross_it(self):
        image = np.full((420, 760, 3), 245, dtype=np.uint8)
        horizontal = np.zeros((420, 760), dtype=np.uint8)
        vertical = np.zeros((420, 760), dtype=np.uint8)
        columns = [2, 24, 124, 284, 444, 604, 757]
        rows = [10, 50, 90, 130, 170, 210, 250, 290, 330, 370, 410]
        for row in rows:
            cv2.line(horizontal, (columns[0], row), (columns[-1], row), 255, 2)
        for column in columns:
            cv2.line(vertical, (column, rows[0]), (column, rows[-1]), 255, 2)

        actual_columns, actual_rows = table_pipeline._trim_disconnected_outer_frame_cells(
            image,
            horizontal,
            vertical,
            columns,
            rows,
        )

        self.assertEqual(actual_columns, columns)
        self.assertEqual(actual_rows, rows)

    def test_paired_disconnected_top_and_bottom_frame_rows_are_trimmed(self):
        image = np.full((520, 920, 3), 235, dtype=np.uint8)
        horizontal = np.zeros((520, 920), dtype=np.uint8)
        vertical = np.zeros((520, 920), dtype=np.uint8)
        columns = [25, 125, 325, 525, 725, 895]
        rows = [2, 24, 52, 92, 132, 172, 212, 252, 292, 332, 372, 412, 452, 490, 517]

        for row in rows[2:-2]:
            cv2.line(horizontal, (columns[0], row), (columns[-1], row), 255, 2)
        for column in columns:
            cv2.line(vertical, (column, rows[2]), (column, rows[-3]), 255, 2)
        vertical[rows[2] : rows[3], :] = 0

        actual_columns, actual_rows = table_pipeline._trim_disconnected_outer_frame_cells(
            image,
            horizontal,
            vertical,
            columns,
            rows,
        )

        self.assertEqual(actual_columns, columns)
        self.assertEqual(actual_rows, rows[2:-2])

    def test_flat_trailing_frame_row_is_trimmed_without_dropping_dark_header(self):
        image = np.full((520, 920, 3), 235, dtype=np.uint8)
        horizontal = np.zeros((520, 920), dtype=np.uint8)
        vertical = np.zeros((520, 920), dtype=np.uint8)
        columns = [25, 125, 325, 525, 725, 895]
        rows = [2, 38, 79, 119, 159, 199, 239, 279, 319, 359, 399, 439, 479, 517]
        midpoint = (columns[0] + columns[-1]) // 2
        image[rows[0] : rows[1], columns[0] : midpoint] = 70
        image[rows[0] : rows[1], midpoint : columns[-1]] = 115
        image[rows[-3] : rows[-2], columns[0] : midpoint] = 190
        image[rows[-3] : rows[-2], midpoint : columns[-1]] = 230
        for column in columns:
            cv2.line(vertical, (column, rows[0]), (column, rows[-1]), 255, 2)

        actual_columns, actual_rows = table_pipeline._trim_disconnected_outer_frame_cells(
            image,
            horizontal,
            vertical,
            columns,
            rows,
        )

        self.assertEqual(actual_columns, columns)
        self.assertEqual(actual_rows, rows[:-1])

    def test_flat_leading_frame_row_is_trimmed_without_dropping_last_data_row(self):
        image = np.full((520, 920, 3), 235, dtype=np.uint8)
        horizontal = np.zeros((520, 920), dtype=np.uint8)
        vertical = np.zeros((520, 920), dtype=np.uint8)
        columns = [25, 125, 325, 525, 725, 895]
        rows = [2, 38, 79, 119, 159, 199, 239, 279, 319, 359, 399, 439, 479, 517]
        midpoint = (columns[0] + columns[-1]) // 2
        image[rows[1] : rows[2], columns[0] : midpoint] = 70
        image[rows[1] : rows[2], midpoint : columns[-1]] = 115
        image[rows[-2] : rows[-1], columns[0] : midpoint] = 190
        image[rows[-2] : rows[-1], midpoint : columns[-1]] = 230
        for column in columns:
            cv2.line(vertical, (column, rows[0]), (column, rows[-1]), 255, 2)

        actual_columns, actual_rows = table_pipeline._trim_disconnected_outer_frame_cells(
            image,
            horizontal,
            vertical,
            columns,
            rows,
        )

        self.assertEqual(actual_columns, columns)
        self.assertEqual(actual_rows, rows[1:])

    def test_clipped_leading_column_is_recovered_from_crossing_row_rules(self):
        height, width = 460, 1366
        image = np.full((height, width, 3), 245, dtype=np.uint8)
        horizontal = np.zeros((height, width), dtype=np.uint8)
        vertical = np.zeros((height, width), dtype=np.uint8)
        columns = [112, 345, 572, 799, 1026, 1253]
        rows = [10, 50, 90, 130, 170, 210, 250, 290, 330, 370, 410, 450]
        for row in rows:
            cv2.line(horizontal, (0, row), (columns[-1], row), 255, 2)

        actual_columns, actual_rows = table_pipeline._recover_clipped_frame_boundaries(
            image,
            horizontal,
            vertical,
            columns,
            rows,
        )

        self.assertEqual(actual_columns, [0] + columns)
        self.assertEqual(actual_rows, rows)

        no_evidence_columns, _ = table_pipeline._recover_clipped_frame_boundaries(
            image,
            np.zeros_like(horizontal),
            vertical,
            columns,
            rows,
        )
        self.assertEqual(no_evidence_columns, columns)

    def test_clipped_wide_trailing_column_uses_image_edge_when_all_row_rules_continue(self):
        height, width = 678, 934
        image = np.full((height, width, 3), 245, dtype=np.uint8)
        horizontal = np.zeros((height, width), dtype=np.uint8)
        vertical = np.zeros((height, width), dtype=np.uint8)
        columns = [2, 94, 300, 394, 590]
        rows = [6, 62, 98, 134, 172, 210, 250, 288, 328, 366, 405, 444, 482, 522, 560, 599, 636, 672]
        for row in rows:
            cv2.line(horizontal, (columns[0], row), (width - 1, row), 255, 2)

        actual_columns, actual_rows = table_pipeline._recover_clipped_frame_boundaries(
            image,
            horizontal,
            vertical,
            columns,
            rows,
        )

        self.assertEqual(actual_columns, columns + [width - 1])
        self.assertEqual(actual_rows, rows)

        no_evidence_columns, _ = table_pipeline._recover_clipped_frame_boundaries(
            image,
            np.zeros_like(horizontal),
            vertical,
            columns,
            rows,
        )
        self.assertEqual(no_evidence_columns, columns)

    def test_sparse_low_contrast_grid_recovers_supported_rows_and_columns(self):
        height, width = 850, 1184
        image = np.full((height, width, 3), 245, dtype=np.uint8)
        expected_columns = [129, 222, 334, 546, 700, 856, 948, 1046]
        expected_rows = [101, 143, 183, 224, 265, 307, 349, 389, 430, 472, 514, 555, 596, 638, 679, 720, 762]
        vertical = np.zeros((height, width), dtype=np.uint8)
        horizontal = np.zeros((height, width), dtype=np.uint8)

        for column in expected_columns:
            cv2.line(image, (column, expected_rows[0]), (column, expected_rows[-1]), (35, 35, 35), 2)
            cv2.line(vertical, (column, expected_rows[0]), (column, expected_rows[-1]), 255, 2)
        for row in expected_rows:
            cv2.line(horizontal, (expected_columns[0], row), (expected_columns[-1], row), 255, 2)
        for row in expected_rows[:-2]:
            cv2.line(image, (expected_columns[0], row), (expected_columns[-1], row), (35, 35, 35), 2)

        columns, rows = table_pipeline._recover_sparse_low_contrast_grid(
            image,
            horizontal,
            vertical,
            [11, 19, 1174],
            [16] + expected_rows[:9],
        )

        self.assertEqual(len(columns), len(expected_columns))
        self.assertTrue(all(abs(actual - expected) <= 3 for actual, expected in zip(columns, expected_columns)))
        self.assertEqual(len(rows), len(expected_rows))
        self.assertTrue(all(abs(actual - expected) <= 5 for actual, expected in zip(rows, expected_rows)))

    def test_sparse_motion_grid_recovers_weak_left_columns_from_line_map_support(self):
        height, width = 918, 1206
        image = np.full((height, width, 3), 245, dtype=np.uint8)
        expected_columns = [162, 260, 378, 599, 754, 904, 995, 1086]
        expected_rows = [153, 206, 270, 332, 394, 458, 520, 584, 649, 712]
        vertical = np.zeros((height, width), dtype=np.uint8)
        horizontal = np.zeros((height, width), dtype=np.uint8)

        for row in expected_rows:
            cv2.line(horizontal, (expected_columns[0], row), (expected_columns[-1], row), 255, 2)
            cv2.line(image, (expected_columns[0], row), (expected_columns[-1], row), (45, 45, 45), 2)
        for column in expected_columns[3:]:
            cv2.line(vertical, (column, expected_rows[0]), (column, expected_rows[-1]), 255, 2)
            cv2.line(image, (column, expected_rows[0]), (column, expected_rows[-1]), (45, 45, 45), 2)
        for column in expected_columns[:3]:
            for start in range(expected_rows[0], expected_rows[-1], 58):
                cv2.line(vertical, (column, start), (column, min(start + 12, expected_rows[-1])), 255, 2)

        columns, rows = table_pipeline._recover_sparse_low_contrast_grid(
            image,
            horizontal,
            vertical,
            expected_columns[3:],
            expected_rows,
        )

        self.assertEqual(len(columns), len(expected_columns))
        self.assertTrue(all(abs(actual - expected) <= 8 for actual, expected in zip(columns, expected_columns)))
        self.assertEqual(len(rows), len(expected_rows))
        self.assertTrue(all(abs(actual - expected) <= 2 for actual, expected in zip(rows, expected_rows)))

    def test_low_contrast_grid_rejects_text_stroke_screen_over_split(self):
        candidate_columns = list(range(20, 1521, 100))
        candidate_rows = list(range(10, 1241, 30))
        screen_columns = sorted(
            candidate_columns
            + list(range(205, 706, 10))
        )
        screen_rows = candidate_rows[2:-2]

        self.assertTrue(
            table_pipeline._low_contrast_grid_supersedes_screen_grid(
                candidate_columns,
                candidate_rows,
                screen_columns,
                screen_rows,
                width=1600,
            )
        )

    def test_low_contrast_grid_restores_one_supported_left_outer_column(self):
        candidate_columns = [22, 132, 242, 352, 528, 638, 748, 858, 968, 1078, 1188, 1302]
        candidate_rows = list(range(1, 1209, 30))
        screen_columns = [132, 242, 352, 528, 638, 748, 858, 968, 1079, 1188, 1316]
        screen_rows = list(candidate_rows)

        self.assertTrue(
            table_pipeline._low_contrast_grid_supersedes_screen_grid(
                candidate_columns,
                candidate_rows,
                screen_columns,
                screen_rows,
                width=1335,
            )
        )

    def test_low_contrast_grid_does_not_replace_similarly_supported_screen_grid(self):
        candidate_columns = [20, 120, 220, 320, 420, 520]
        candidate_rows = list(range(10, 431, 30))
        screen_columns = [22, 121, 221, 321, 421, 521]
        screen_rows = list(candidate_rows)

        self.assertFalse(
            table_pipeline._low_contrast_grid_supersedes_screen_grid(
                candidate_columns,
                candidate_rows,
                screen_columns,
                screen_rows,
                width=540,
            )
        )

    def test_sparse_group_header_without_merge_map_is_unresolved(self):
        grid = [
            ["", "基础信息", "", "", "目标与测量", "", "", "过程记录"],
            ["序号", "编号", "名称", "区域", "日期", "类型", "结果", "备注"],
            ["1", "AP-001", "主控模块", "A区", "2026-08-01", "巡检", "正常", "通过"],
            ["2", "AP-002", "传感器", "B区", "2026-08-02", "复检", "正常", "通过"],
        ]

        self.assertTrue(
            ocr_backend._multilevel_header_spans_are_unresolved(grid, [])
        )
        incomplete = [
            {
                "row": 0,
                "column": 0,
                "row_span": 1,
                "column_span": 3,
            }
        ]
        self.assertTrue(
            ocr_backend._multilevel_header_spans_are_unresolved(
                grid,
                incomplete,
            )
        )
        self.assertFalse(
            ocr_backend._multilevel_header_spans_are_unresolved(
                grid,
                incomplete
                + [
                    {
                        "row": 0,
                        "column": 3,
                        "row_span": 1,
                        "column_span": 3,
                    },
                    {
                        "row": 0,
                        "column": 6,
                        "row_span": 1,
                        "column_span": 2,
                    },
                ],
            )
        )

    def test_each_sparse_multilevel_header_row_requires_its_own_merge_map(self):
        grid = [
            ["基础信息", "", "", "", "目标与测量", "", "", "过程记录"],
            ["本期", "", "实际", "", "计划", "", "备注", ""],
            ["序号", "编号", "名称", "区域", "日期", "类型", "结果", "备注"],
            ["1", "AP-001", "主控模块", "A区", "2026-08-01", "巡检", "正常", "通过"],
            ["2", "AP-002", "传感器", "B区", "2026-08-02", "复检", "正常", "通过"],
        ]
        first_row_only = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 4},
            {"row": 0, "column": 4, "row_span": 1, "column_span": 3},
        ]

        self.assertTrue(
            ocr_backend._multilevel_header_spans_are_unresolved(
                grid,
                first_row_only,
            )
        )

    def test_wide_month_header_is_treated_as_multilevel(self):
        grid = [
            ["", "基础信息", "", "", "", "目标与测量", "", "", "", "", "过程记录", "", ""],
            ["", "本期", "", "实际", "", "计划", "", "上限", "", "实际", "", "下限", "累计"],
            ["序号", "区域", "部门", "项目", "一月", "二月", "三月", "一季度", "四月", "五月", "六月", "二季度", "上半年"],
        ]
        grid.extend(
            [str(index), "A区", "一车间", "传感器", "10", "20", "30", "60", "40", "50", "60", "150", "210"]
            for index in range(1, 7)
        )

        self.assertEqual(
            ocr_backend._multilevel_header_rows_for_review(grid),
            {0, 1, 2},
        )

    def test_collapsed_spatial_recovery_accepts_supported_multilevel_headers(self):
        grid = [
            ["基础信息", "", "", "目标与测量", "", "", "过程记录", ""],
            ["本期", "", "实际", "计划", "", "上限", "实际", "累计"],
            ["序号", "区域", "部门", "项目", "一月", "二月", "三月", "上半年"],
        ]
        grid.extend(
            [str(index), "A区", "一车间", "传感器", "10", "20", "30", "60"]
            for index in range(1, 8)
        )
        confidence = [[0.95 for _ in row] for row in grid]
        anchors = [20.0 + column * 80.0 for column in range(8)]
        row_centers = [5.0 + row * 10.0 for row in range(10)]
        grouped_rows = [
            [
                {"center_x": anchors[column]}
                for column, value in enumerate(row)
                if value
            ]
            for row in grid
        ]
        geometry = {
            "anchors": anchors,
            "row_centers": row_centers,
            "grouped_rows": grouped_rows,
        }

        recovered = ocr_backend._recover_collapsed_page_spatial_rows(
            grid,
            confidence,
            [],
            True,
            geometry,
            list(range(0, 101, 10)),
        )

        self.assertIsNotNone(recovered)
        self.assertEqual((len(recovered[0]), len(recovered[0][0])), (10, 8))
        self.assertEqual(recovered[2]["multilevel_header_rows"], [0, 1, 2])

    def test_collapsed_spatial_recovery_allows_header_only_physical_misalignment(self):
        grid = [
            ["基础信息", "", "", "目标与测量", "", "", "过程记录", ""],
            ["本期", "", "实际", "计划", "", "上限", "实际", "累计"],
            ["序号", "区域", "部门", "项目", "一月", "二月", "三月", "上半年"],
        ]
        grid.extend(
            [str(index), "A区", "一车间", "传感器", "10", "20", "30", "60"]
            for index in range(1, 8)
        )
        confidence = [[0.95 for _ in row] for row in grid]
        anchors = [20.0 + column * 80.0 for column in range(8)]
        row_centers = [15.0, 23.0, 35.0, 45.0, 55.0, 65.0, 75.0, 85.0, 95.0, 105.0]
        grouped_rows = [
            [
                {"center_x": anchors[column]}
                for column, value in enumerate(row)
                if value
            ]
            for row in grid
        ]
        geometry = {
            "anchors": anchors,
            "row_centers": row_centers,
            "grouped_rows": grouped_rows,
        }
        physical_rows = [0, 10, 20, 40, 50, 60, 70, 80, 90, 100, 110]

        recovered = ocr_backend._recover_collapsed_page_spatial_rows(
            grid,
            confidence,
            [],
            True,
            geometry,
            physical_rows,
        )

        self.assertIsNotNone(recovered)
        self.assertTrue(recovered[2]["multilevel_physical_alignment"])
        self.assertEqual((len(recovered[0]), len(recovered[0][0])), (10, 8))

        misaligned_geometry = dict(geometry)
        misaligned_geometry["row_centers"] = list(row_centers)
        misaligned_geometry["row_centers"][6] = 85.0
        self.assertIsNone(
            ocr_backend._recover_collapsed_page_spatial_rows(
                grid,
                confidence,
                [],
                True,
                misaligned_geometry,
                physical_rows,
            )
        )

    def test_spatial_rows_recover_only_a_collapsed_multilevel_header_band(self):
        grid = [
            ["基础信息", "", "", "目标与测量", "", "", "过程记录", ""],
            ["序号", "区域", "部门", "项目", "一月", "二月", "三月", "上半年"],
        ]
        grid.extend(
            [str(index), "A区", "一车间", "传感器", "10", "20", "30", "60"]
            for index in range(1, 7)
        )
        confidence = [[0.95 for _ in row] for row in grid]
        anchors = [20.0 + column * 80.0 for column in range(8)]
        row_centers = [15.0, 23.0, 35.0, 45.0, 55.0, 65.0, 75.0, 85.0]
        geometry = {
            "anchors": anchors,
            "row_centers": row_centers,
            "grouped_rows": [
                [
                    {"center_x": anchors[column]}
                    for column, value in enumerate(row)
                    if value
                ]
                for row in grid
            ],
        }
        physical_rows = [0, 10, 20, 50, 60, 70, 80, 90]
        page_grid = [["" for _ in range(8)] for _ in range(6)]

        recovered = ocr_backend._recover_multilevel_header_spatial_rows(
            grid,
            confidence,
            [],
            True,
            geometry,
            physical_rows,
            page_grid,
        )

        self.assertIsNotNone(recovered)
        self.assertEqual((len(recovered[0]), len(recovered[0][0])), (8, 8))
        self.assertEqual(recovered[2]["missing_header_boundaries"], 2)

        misaligned_geometry = dict(geometry)
        misaligned_geometry["row_centers"] = list(row_centers)
        misaligned_geometry["row_centers"][5] = 75.0
        self.assertIsNone(
            ocr_backend._recover_multilevel_header_spatial_rows(
                grid,
                confidence,
                [],
                True,
                misaligned_geometry,
                physical_rows,
                page_grid,
            )
        )

    def test_extract_ruled_grid_restores_clipped_edges_and_rejects_repeated_text_stroke(self):
        height, width = 353, 662
        image = np.full((height, width, 3), 255, dtype=np.uint8)
        expected_columns = [0, 108, 184, 298, 453, 543, width - 1]
        expected_rows = [0, 42, 92, 136, 180, 222, 266, 308, height - 1]

        # The top and left outer rules were clipped by perspective
        # rectification. All remaining real rules still span the table.
        for column in expected_columns[1:]:
            cv2.line(image, (column, 0), (column, height - 1), (15, 15, 15), 1)
        for row in expected_rows[1:]:
            cv2.line(image, (0, row), (width - 1, row), (15, 15, 15), 1)

        # Repeated vertical glyph strokes in the last column survive the
        # morphology pass but do not have the continuity of a real rule.
        for top, bottom in zip(expected_rows[:-1], expected_rows[1:]):
            center = (top + bottom) // 2
            cv2.line(image, (604, center - 11), (604, center + 11), (0, 0, 0), 2)

        grid = extract_ruled_grid(image)

        self.assertIsNotNone(grid)
        actual_columns, actual_rows, _ = grid
        self.assertEqual(actual_columns, expected_columns)
        self.assertEqual(actual_rows, expected_rows)

    def test_extract_ruled_grid_does_not_split_merged_tall_row_without_rule_evidence(self):
        image = np.full((220, 520, 3), 255, dtype=np.uint8)
        columns = [3, 100, 250, 400, 516]
        expected_rows = [3, 47, 69, 91, 113, 135, 157, 179, 201]
        for column in columns:
            cv2.line(image, (column, expected_rows[0]), (column, expected_rows[-1]), (0, 0, 0), 2)
        for row in expected_rows:
            cv2.line(image, (columns[0], row), (columns[-1], row), (0, 0, 0), 2)

        grid = extract_ruled_grid(image)

        self.assertIsNotNone(grid)
        _, actual_rows, _ = grid
        self.assertEqual(actual_rows, expected_rows)

    def test_assign_ocr_to_grid_keeps_text_inside_its_cell(self):
        columns = [0, 100, 220]
        rows = [0, 50, 110]
        boxes = np.array(
            [
                [[8, 8], [40, 8], [40, 28], [8, 28]],
                [[110, 8], [150, 8], [150, 28], [110, 28]],
                [[155, 8], [205, 8], [205, 28], [155, 28]],
                [[8, 62], [45, 62], [45, 88], [8, 88]],
            ],
            dtype=float,
        )

        grid, confidence = assign_ocr_to_grid(
            columns,
            rows,
            boxes,
            ["编号", "515.128", "MHz", "1"],
            [0.99, 0.96, 0.94, 0.98],
        )

        self.assertEqual(grid, [["编号", "515.128 MHz"], ["1", ""]])
        self.assertAlmostEqual(confidence[0][1], 0.95)
        self.assertEqual(confidence[1][1], 0.0)

    def test_assign_ocr_to_grid_orders_one_baseline_left_to_right_despite_y_jitter(self):
        columns = [0, 220]
        rows = [0, 60]
        boxes = np.array(
            [
                [[12, 21], [70, 21], [70, 31], [12, 31]],
                [[82, 18], [150, 18], [150, 28], [82, 28]],
            ],
            dtype=float,
        )

        grid, confidence = assign_ocr_to_grid(
            columns,
            rows,
            boxes,
            ["BPSK", "35.4k"],
            [0.98, 0.96],
            preserve_geometry=True,
        )

        self.assertEqual(grid, [["BPSK 35.4k"]])
        self.assertAlmostEqual(confidence[0][0], 0.97)

    def test_assign_ocr_to_grid_rejects_spreadsheet_ruler_outside_grid_frame(self):
        columns = [18, 136, 260]
        rows = [8, 40, 72]
        boxes = np.array(
            [
                [[6, 14], [29, 14], [29, 34], [6, 34]],
                [[14, 46], [58, 46], [58, 66], [14, 66]],
            ],
            dtype=float,
        )

        grid, confidence = assign_ocr_to_grid(
            columns,
            rows,
            boxes,
            ["10", "边缘数据"],
            [0.999, 0.97],
            preserve_geometry=True,
        )

        self.assertEqual(grid, [["", ""], ["边缘数据", ""]])
        self.assertEqual(confidence, [[0.0, 0.0], [0.97, 0.0]])

    def test_reconcile_verified_grid_keeps_only_matching_high_confidence_text(self):
        verified, confidence = ocr_backend._reconcile_verified_grid(
            [["编号", "515.221", "QPSK", ""]],
            [[0.99, 0.98, 0.96, 0.0]],
            [["编号", "515.221", "OPSK", ""]],
            [[0.98, 0.97, 0.95, 0.0]],
        )

        self.assertEqual(verified, [["编号", "515.221", "", ""]])
        self.assertGreater(confidence[0][0], 0)
        self.assertEqual(confidence[0][2], -1.0)
        self.assertEqual(confidence[0][3], 0.0)

    def test_reconcile_verified_grid_rejects_missing_or_low_confidence_text(self):
        verified, confidence = ocr_backend._reconcile_verified_grid(
            [["180K", "-10", "文字"]],
            [[0.96, 0.97, 0.74]],
            [["", "-10", "文字"]],
            [[-1.0, 0.96, 0.95]],
        )

        self.assertEqual(verified, [["", "-10", ""]])
        self.assertEqual(confidence[0], [-1.0, 0.96, -1.0])

    def test_consistency_checks_reject_numeric_unit_outlier_and_relation_digit(self):
        grid = [
            ["功率", "关系"],
            ["-30 dBm", "≤"],
            ["-20 dBm", "范围"],
            [") dBm", "2"],
            ["0 dBm", "≈"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[3], ["", ""])
        self.assertEqual(confidence[3], [-1.0, -1.0])

    def test_consistency_checks_reject_confusable_interface_and_formula_tokens(self):
        grid = [
            ["接口 Interface", "表达式"],
            ["ETHO", "P=10log1o(P/1mW)"],
            ["RF-IN", "U=I×R"],
            ["USB", "ΔT=T₂-T_₁"],
            ["LAN", "fo=10 MHz"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[1], ["", ""])
        self.assertEqual(confidence[1], [-1.0, -1.0])
        self.assertEqual(grid[3], ["USB", ""])
        self.assertEqual(confidence[3], [0.99, -1.0])
        self.assertEqual(grid[4], ["LAN", ""])
        self.assertEqual(confidence[4], [0.99, -1.0])

    def test_consistency_checks_preserve_visible_dash_in_unit_column(self):
        grid = [["量名称", "单位"], ["驻波", "_"]]
        confidence = [[0.99, 0.99], [0.99, 0.91]]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[1], ["驻波", "—"])
        self.assertEqual(confidence[1], [0.99, 0.91])

    def test_dense_header_and_column_blank_are_marked_for_review(self):
        grid = [
            ["编号", "型号", "状态"],
            ["A001", "SG-1", "正常"],
            ["A002", "", "待机"],
            ["A003", "SG-3", "复核"],
            ["A004", "SG-4", "正常"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        grid[0][1] = ""
        confidence[0][1] = 0.0

        ocr_backend._mark_unexpected_blank_cells(grid, confidence, [])

        self.assertEqual(confidence[0][1], -1.0)
        self.assertEqual(confidence[2][1], -1.0)

    def test_unresolved_collapsed_header_and_its_blank_slots_are_marked_for_review(self):
        grid = [
            ["日期", "生产线 计划数量 完成数量 不良数量 班次", "", "", "", "", "完成率"],
            ["2026-08-08", "A线", "100", "98", "2", "白班", "98%"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        confidence[0][1] = 0.0

        grid, confidence, _ = ocr_backend._split_collapsed_header_data_row(
            grid, confidence, []
        )
        ocr_backend._mark_unexpected_blank_cells(grid, confidence, [])

        self.assertEqual(grid[0][1], "生产线 计划数量 完成数量 不良数量 班次")
        self.assertEqual(grid[0][2:6], [""] * 4)
        self.assertEqual(confidence[0][0], 0.99)
        self.assertEqual(confidence[0][1:6], [-1.0] * 5)
        self.assertEqual(confidence[0][6], 0.99)

    def test_sparse_semantic_blank_is_not_guessed_without_ink_evidence(self):
        grid = [
            ["编号", "型号", "状态", "备注"],
            ["A001", "SG-1", "正常", "已确认"],
            ["A002", "SG-2", "", "待确认"],
            ["A003", "SG-3", "复核", "待处理"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        ocr_backend._mark_unexpected_blank_cells(grid, confidence, [])

        self.assertEqual(confidence[2][2], 0.0)

    def test_optional_empty_remark_column_is_not_marked_for_review(self):
        grid = [["编号", "名称", "备注"]] + [
            [f"A{index:03d}", f"项目{index}", ""] for index in range(1, 10)
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        ocr_backend._mark_unexpected_blank_cells(grid, confidence, [])

        self.assertEqual([row[2] for row in confidence[1:]], [0.0] * 9)

    def test_merged_group_header_covered_cells_are_not_marked_for_review(self):
        grid = [
            ["基础信息", "", "状态", "备注"],
            ["A001", "SG-1", "正常", "已确认"],
            ["A002", "SG-2", "待机", "待处理"],
        ]
        confidence = [[0.99 if value else -1.0 for value in row] for row in grid]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 2, "role": "group_header"}
        ]

        ocr_backend._mark_unexpected_blank_cells(grid, confidence, spans)

        self.assertEqual(confidence[0][1], 0.0)

    def test_unresolved_visible_blank_is_a_hard_publishing_blocker(self):
        grid = [["编号", "名称"], ["A001", ""], ["A002", ""]]
        confidence = [[0.99, 0.99], [0.99, -2.0], [0.99, -1.0]]

        self.assertEqual(
            ocr_backend._blocking_unresolved_visible_blank_locations(
                grid, confidence
            ),
            [(1, 1)],
        )

    def test_nonblank_review_and_true_blank_do_not_trigger_visible_blank_blocker(self):
        grid = [["编号", "名称"], ["A001", ""], ["A002", ""]]
        confidence = [[0.99, 0.99], [0.77, 0.0], [0.99, 0.0]]

        self.assertEqual(
            ocr_backend._blocking_unresolved_visible_blank_locations(
                grid, confidence
            ),
            [],
        )

    def test_merged_span_clears_visible_blank_blocker_from_covered_cell(self):
        grid = [["设备资产清单", "", ""], ["编号", "名称", "状态"]]
        confidence = [[0.99, -2.0, -2.0], [0.99, 0.99, 0.99]]
        spans = [
            {
                "row": 0,
                "column": 0,
                "row_span": 1,
                "column_span": 3,
                "role": "title",
            }
        ]

        ocr_backend._mark_unexpected_blank_cells(grid, confidence, spans)

        self.assertEqual(confidence[0][1:], [0.0, 0.0])
        self.assertEqual(
            ocr_backend._blocking_unresolved_visible_blank_locations(
                grid, confidence
            ),
            [],
        )

    def test_missing_second_group_heading_is_marked_at_measurement_boundary(self):
        grid = [
            ["项目任务跟踪表", "", "", "", "", "", ""],
            ["基础信息", "", "", "", "", "", ""],
            ["任务编号", "任务名称", "优先级", "进度", "负责人", "计划完成", "当前状态"],
            ["TASK-001", "界面设计", "中", "7%", "陈晨", "2026-08-02", "已完成"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 7, "role": "title"},
            {"row": 1, "column": 0, "row_span": 1, "column_span": 7, "role": "title"},
        ]

        ocr_backend._mark_unexpected_blank_cells(grid, confidence, spans)

        self.assertEqual(confidence[1][3], -1.0)
        self.assertEqual(confidence[1][1], 0.0)

    def test_consistency_checks_preserve_text_in_mixed_result_column(self):
        grid = [
            ["项目", "结果"],
            ["电压", "12.06"],
            ["频率", "9.9998"],
            ["端口", "USB-C"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[3][1], "USB-C")
        self.assertEqual(confidence[3][1], 0.99)

    def test_consistency_checks_do_not_guess_unique_category_or_lost_decimal(self):
        grid = [
            ["类别", "带宽(kHz)"],
            ["模拟信号", "26.6"],
            ["数字信号", "31.4"],
            ["模拟信号", "48.2"],
            ["数字信号", "72.2"],
            ["模拟信号", "96.4"],
            ["数字信号", "108.8"],
            ["横拟信号", "1292"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[-1], ["横拟信号", "1292"])
        self.assertEqual(confidence[-1], [0.99, 0.77])

    def test_consistency_checks_mark_missing_percent_format_without_guessing(self):
        grid = [
            ["日期", "完成率"],
            ["08-01", "98.2%"],
            ["08-02", "98.4%"],
            ["08-03", "98.5%"],
            ["08-04", "98.7%"],
            ["08-05", "98.8%"],
            ["08-06", "99.0%"],
            ["08-07", "98.19"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[-1][1], "98.19")
        self.assertEqual(confidence[-1][1], 0.77)

    def test_consistency_checks_mark_lost_trailing_zero_without_rewriting(self):
        grid = [["序号", "目标值"]] + [
            [str(index), value]
            for index, value in enumerate(
                ["1.000", "2.100", "3.200", "4.300", "5.400", "6.500", "7.60"],
                start=1,
            )
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[-1][1], "7.60")
        self.assertEqual(confidence[-1][1], 0.77)

    def test_consistency_checks_mark_ordinal_digit_loss_without_guessing(self):
        grid = [["序号", "名称"]] + [
            [str(index), f"设备{index}"] for index in range(1, 13)
        ]
        grid[10][0] = "0"
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[10][0], "0")
        self.assertEqual(confidence[10][0], 0.77)

    def test_consistency_checks_preserve_legitimate_mixed_numeric_precision(self):
        grid = [
            ["序号", "结果"], ["1", "1"], ["2", "2.0"], ["3", "3.00"],
            ["4", "4"], ["5", "5.0"], ["6", "6.00"], ["7", "7"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual([row[1] for row in confidence], [0.99] * len(grid))

    def test_consistency_checks_preserve_high_confidence_business_labels(self):
        grid = [
            ["任务编号", "任务名称", "负责人"],
            ["TASK-001", "图像预处理", "陈晨"],
            ["TASK-002", "图像预处理", "陈晨"],
            ["TASK-003", "面像预处理", "陈"],
            ["TASK-004", "表格结构识别", "李"],
            ["TASK-005", "表格结构识别高", "李"],
            ["TASK-006", "表格结构识别低", "王强"],
            ["TASK-007", "表格结构识别", "李娜"],
            ["TASK-008", "表格结构识别", "李娜"],
        ]
        confidence = [[0.99 for _ in row] for row in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[3][1:], ["面像预处理", "陈"])
        self.assertEqual(grid[4][1:], ["表格结构识别", "李"])
        self.assertEqual(grid[5][1:], ["表格结构识别高", "李"])
        self.assertEqual(grid[6][1:], ["表格结构识别低", "王强"])
        self.assertEqual(confidence[3][1:], [0.99, 0.99])
        self.assertEqual(confidence[4][1:], [0.99, 0.99])
        self.assertEqual(confidence[5][1:], [0.99, 0.99])
        self.assertEqual(confidence[6][1:], [0.99, 0.99])

    def test_consistency_checks_preserve_legitimate_status_and_shift_values(self):
        grid = [
            ["编号", "状态", "班次"],
            ["A001", "正常", "白班"],
            ["A002", "正常", "白班"],
            ["A003", "正常", "白班"],
            ["A004", "异常", "夜班"],
        ]
        confidence = [[0.99 for _ in row] for row in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(confidence[-1], [0.99, 0.99, 0.99])

    def test_consistency_checks_mark_unique_one_character_header_confusion(self):
        grid = [
            ["日期", "生产线", "计划数型", "完成数量"],
            ["2026-08-08", "A线", "100", "98"],
        ]
        confidence = [[0.99 for _ in row] for row in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[0][2], "计划数型")
        self.assertEqual(confidence[0][2], 0.77)

    def test_suspicious_leading_cells_select_a_header_with_numeric_ghost(self):
        grid = [
            ["日期", "班次", "生产线", "计划数量", "完成数量 767"],
            ["2026-08-02", "白班", "A线", "300", "296"],
        ]
        confidence = [[0.99, 0.99, 0.99, 0.99, 0.84], [0.99] * 5]

        selected = ocr_backend._suspicious_leading_ruled_cells(grid, confidence)

        self.assertIn((0, 4), selected)
        self.assertNotIn((0, 3), selected)

    def test_consistency_checks_never_publish_a_header_with_numeric_ghost(self):
        grid = [
            ["日期", "班次", "生产线", "计划数量", "完成数量 767"],
            ["2026-08-02", "白班", "A线", "300", "296"],
        ]
        confidence = [[0.99] * 5 for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[0][4], "完成数量 767")
        self.assertEqual(confidence[0][4], 0.77)

    def test_consistency_checks_find_header_after_single_cell_title(self):
        grid = [
            ["中英文接口清单", "", ""],
            ["模块", "接口 Interface", "方向"],
            ["主控", "ETHO", "双向"],
        ]
        confidence = [[0.99, 0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[2][1], "")
        self.assertEqual(confidence[2][1], -1.0)

    def test_consistency_checks_preserve_second_level_column_headers(self):
        grid = [
            ["基础信息", "基础信息", "技术参数", "技术参数", "允差范围", "允差范围", "结果", "结果"],
            ["样品编号", "测量项目", "标称值", "测量值", "下偏差", "上偏差", "结论", "记录人"],
            ["SMP-001", "质量", "11.25", "11.28", "-0.10", "+0.10", "合格", "刘工"],
            ["SMP-002", "长度", "80.00", "80.03", "-0.05", "+0.05", "合格", "王工"],
        ]
        confidence = [[0.99] * 8 for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(
            grid[1],
            ["样品编号", "测量项目", "标称值", "测量值", "下偏差", "上偏差", "结论", "记录人"],
        )
        self.assertEqual(confidence[1], [0.99] * 8)

    def test_assign_ocr_to_grid_splits_one_box_across_two_cells_at_whitespace(self):
        columns = [0, 100, 220]
        rows = [0, 50]
        boxes = np.array([[[50, 8], [170, 8], [170, 30], [50, 30]]], dtype=float)

        grid, confidence = assign_ocr_to_grid(
            columns,
            rows,
            boxes,
            ["-10.00 2.4 MS/s"],
            [0.97],
        )

        self.assertEqual(grid, [["-10.00", "2.4 MS/s"]])
        self.assertEqual(confidence, [[0.97, 0.97]])

    def test_assign_ocr_to_grid_does_not_silently_shift_unsplittable_crossing_text(self):
        columns = [0, 100, 220]
        rows = [0, 50, 100]
        boxes = np.array([[[70, 58], [150, 58], [150, 82], [70, 82]]], dtype=float)

        grid, confidence = assign_ocr_to_grid(
            columns,
            rows,
            boxes,
            ["无法可靠拆分"],
            [0.96],
        )

        self.assertEqual(grid, [["", ""], ["无法可靠拆分", ""]])
        self.assertEqual(confidence, [[0.0, 0.0], [-1.0, 0.0]])

    def test_assign_ocr_to_grid_keeps_full_width_title_at_left_edge(self):
        columns = [0, 100, 220, 340]
        rows = [0, 50, 100]
        boxes = np.array([[[65, 8], [285, 8], [285, 32], [65, 32]]], dtype=float)

        grid, confidence = assign_ocr_to_grid(
            columns,
            rows,
            boxes,
            ["设备运行统计表"],
            [0.97],
        )

        self.assertEqual(grid[0], ["设备运行统计表", "", ""])
        self.assertEqual(confidence[0], [0.97, 0.0, 0.0])

    def test_assign_ocr_to_grid_splits_fused_header_and_first_data_row(self):
        columns = [0, 120, 260, 400, 540, 680]
        rows = [0, 60, 120]
        boxes = np.array(
            [
                [[8, 8], [112, 8], [112, 42], [8, 42]],
                [[128, 8], [252, 8], [252, 42], [128, 42]],
                [[268, 8], [392, 8], [392, 42], [268, 42]],
                [[408, 8], [532, 8], [532, 42], [408, 42]],
                [[548, 8], [672, 8], [672, 42], [548, 42]],
            ],
            dtype=float,
        )

        grid, confidence = assign_ocr_to_grid(
            columns,
            rows,
            boxes,
            [
                "编号 1",
                "频率 515.128MHz",
                "信号类型 模拟",
                "占用带宽 9kHz",
                "调制方式 OFDM",
            ],
            [0.98, 0.97, 0.96, 0.95, 0.94],
        )

        self.assertEqual(
            grid[:2],
            [
                ["编号", "频率", "信号类型", "占用带宽", "调制方式"],
                ["1", "515.128MHz", "模拟", "9kHz", "OFDM"],
            ],
        )
        self.assertEqual(confidence[0], [0.98, 0.97, 0.96, 0.95, 0.94])
        self.assertEqual(confidence[1], [0.98, 0.97, 0.96, 0.95, 0.94])

    def test_assign_ocr_to_grid_splits_short_text_box_spanning_two_rows(self):
        columns = [0, 100, 200]
        rows = [0, 40, 80, 120]
        boxes = np.asarray([
            [[120, 5], [160, 5], [160, 32], [120, 32]],
            [[120, 44], [150, 44], [150, 114], [120, 114]],
        ], dtype=float)

        grid, confidence = assign_ocr_to_grid(
            columns,
            rows,
            boxes,
            ["状态", "中高"],
            [0.99, 0.98],
        )

        self.assertEqual(grid[1][1], "中")
        self.assertEqual(grid[2][1], "高")
        self.assertEqual(confidence[1][1], 0.98)
        self.assertEqual(confidence[2][1], 0.98)

    def test_assign_ocr_to_grid_prefers_horizontal_split_near_row_edge(self):
        grid, confidence = assign_ocr_to_grid(
            [0, 50, 100],
            [0, 30, 54, 84],
            np.asarray([[[25, 28], [75, 28], [75, 58], [25, 58]]], dtype=float),
            ["25 G04"],
            [0.99],
            preserve_geometry=True,
        )

        self.assertEqual(grid, [["", ""], ["25", "G04"], ["", ""]])
        self.assertEqual(confidence, [[0.0, 0.0], [0.99, 0.99], [0.0, 0.0]])

    def test_assign_ocr_to_grid_splits_two_multichar_statuses_spanning_rows(self):
        grid, _ = assign_ocr_to_grid(
            [0, 100],
            [0, 40, 80],
            np.asarray([[[20, 4], [70, 4], [70, 76], [20, 76]]], dtype=float),
            ["正常复核"],
            [0.99],
        )

        self.assertEqual(grid, [["正常"], ["复核"]])

    def test_assign_ocr_to_grid_does_not_split_cjk_tail_at_cell_edge(self):
        grid, confidence = assign_ocr_to_grid(
            [0, 100],
            [0, 40, 80],
            np.asarray(
                [
                    [[20, 2], [70, 2], [70, 46], [20, 46]],
                    [[20, 48], [70, 48], [70, 70], [20, 70]],
                ],
                dtype=float,
            ),
            ["数字", "模拟"],
            [0.99, 0.99],
            preserve_geometry=True,
        )

        self.assertEqual(grid, [["数字"], ["模拟"]])
        self.assertEqual(confidence, [[0.99], [0.99]])

    def test_assign_ocr_to_grid_splits_three_stacked_short_cells(self):
        grid, _ = assign_ocr_to_grid(
            [0, 100],
            [0, 40, 80, 120],
            np.asarray([[[20, 4], [50, 4], [50, 116], [20, 116]]], dtype=float),
            ["无无无"],
            [0.99],
        )

        self.assertEqual(grid, [["无"], ["无"], ["无"]])

    def test_assign_ocr_to_grid_ignores_minor_boundary_jitter(self):
        grid, confidence = assign_ocr_to_grid(
            [0, 100, 200],
            [0, 40],
            np.asarray([[[94, 5], [128, 5], [128, 30], [94, 30]]], dtype=float),
            ["无"],
            [0.99],
        )

        self.assertEqual(grid, [["", "无"]])
        self.assertEqual(confidence, [[0.0, 0.99]])

    def test_assign_ocr_to_grid_keeps_normal_height_text_crossing_row_line(self):
        grid, _ = assign_ocr_to_grid(
            [0, 100],
            [0, 40, 80],
            np.asarray([[[20, 31], [70, 31], [70, 49], [20, 49]]], dtype=float),
            ["正常"],
            [0.99],
        )

        self.assertEqual(grid, [[""], ["正常"]])

    def test_assign_ocr_to_grid_keeps_ordinary_multiword_cells_in_one_row(self):
        columns = [0, 180, 360, 540]
        rows = [0, 60]
        boxes = np.array(
            [
                [[8, 8], [172, 8], [172, 40], [8, 40]],
                [[188, 8], [352, 8], [352, 40], [188, 40]],
                [[368, 8], [532, 8], [532, 40], [368, 40]],
            ],
            dtype=float,
        )

        grid, _ = assign_ocr_to_grid(
            columns,
            rows,
            boxes,
            ["Signal Name", "Test Result", "Operator Name"],
            [0.98, 0.97, 0.96],
        )

        self.assertEqual(grid, [["Signal Name", "Test Result", "Operator Name"]])

    def test_assign_ocr_to_grid_splits_concatenated_numeric_unit_pair(self):
        columns = [0, 100, 220]
        rows = [0, 50]
        boxes = np.array([[[110, 8], [205, 8], [205, 30], [110, 30]]], dtype=float)

        grid, _ = assign_ocr_to_grid(
            columns,
            rows,
            boxes,
            ["-30.004.8 MS/s"],
            [0.96],
        )

        self.assertEqual(grid, [["-30.00", "4.8 MS/s"]])

    def test_build_result_preserves_per_cell_confidence(self):
        if "confidences" not in inspect.signature(build_result).parameters:
            self.fail("build_result does not accept per-cell confidence")

        result = build_result(
            [["编号", "频率"], ["1", "515.128"]],
            confidence=0.5,
            confidences=[[0.99, 0.97], [0.98, 0.91]],
            engine="grid-test",
        )

        self.assertEqual(result["cells"][0][0]["confidence"], 0.99)
        self.assertEqual(result["cells"][1][1]["confidence"], 0.91)

    def test_build_result_marks_negative_internal_confidence_for_manual_review(self):
        result = build_result(
            [[""]],
            confidence=0.9,
            confidences=[[-1.0]],
            engine="test",
        )

        self.assertEqual(result["cells"][0][0]["confidence"], 0.0)
        self.assertTrue(result["cells"][0][0]["needs_review"])

    def test_build_result_marks_explicit_uncertain_blank_for_manual_review(self):
        result = build_result(
            [["", ""]],
            confidence=0.9,
            confidences=[[0.77, 0.0]],
            engine="test",
        )

        self.assertTrue(result["cells"][0][0]["needs_review"])
        self.assertFalse(result["cells"][0][1]["needs_review"])

    def test_recognition_confirms_conflicting_ruled_and_spatial_structures(self):
        source = inspect.getsource(ocr_backend._recognize)

        self.assertIn("extract_ruled_grid", source)
        self.assertIn("assign_ocr_to_grid", source)
        self.assertIn("structure_confirmed", source)
        self.assertIn("线框、文字布局与结构模型结果不一致", source)

    def test_borderless_spatial_grid_recovers_columns_and_split_header(self):
        ocr_backend._load_runtime()

        def box(left, top, right, bottom):
            return [[left, top], [right, top], [right, bottom], [left, bottom]]

        boxes = np.array(
            [
                box(55, 5, 285, 27),
                box(10, 40, 70, 62),
                box(88, 40, 142, 62),
                box(160, 40, 300, 62),
                box(10, 75, 70, 97),
                box(88, 75, 142, 97),
                box(165, 75, 215, 97),
                box(245, 75, 295, 97),
                box(10, 110, 70, 132),
                box(88, 110, 142, 132),
                box(165, 110, 215, 132),
                box(245, 110, 295, 132),
            ],
            dtype=float,
        )
        texts = [
            "设备记录",
            "日期",
            "编号",
            "功率 温度",
            "08-05",
            "A001",
            "-10",
            "36.5",
            "08-06",
            "A002",
            "-20",
            "37.1",
        ]
        scores = [0.99] * 11 + [0.55]

        grid, confidence, spans, strong = ocr_backend._borderless_spatial_grid(
            boxes,
            texts,
            scores,
        )

        self.assertTrue(strong)
        self.assertEqual(grid[0], ["设备记录", "", "", ""])
        self.assertEqual(grid[1], ["日期", "编号", "功率", "温度"])
        self.assertEqual(grid[2], ["08-05", "A001", "-10", "36.5"])
        self.assertEqual(grid[3], ["08-06", "A002", "-20", ""])
        self.assertEqual(confidence[3][3], -1.0)
        self.assertEqual(
            spans,
            [{"row": 0, "column": 0, "row_span": 1, "column_span": 4, "role": "title"}],
        )

    def test_borderless_spatial_grid_does_not_merge_header_and_first_data_row(self):
        def box(left, top, right, bottom):
            return [[left, top], [right, top], [right, bottom], [left, bottom]]

        # The two physical rows are unusually close due to camera perspective.
        # They must remain separate; a merged result such as "编号 1" cannot be
        # repaired safely after OCR and corrupts CSV/XLSX geometry.
        boxes = np.array(
            [
                box(10, 30, 70, 50),
                box(90, 30, 170, 50),
                box(10, 43, 70, 63),
                box(90, 43, 170, 63),
            ],
            dtype=float,
        )
        grid, confidence, _, strong = ocr_backend._borderless_spatial_grid(
            boxes,
            ["编号", "频率", "1", "515.128MHz"],
            [0.99, 0.99, 0.99, 0.99],
        )

        self.assertTrue(strong)
        self.assertEqual(grid, [["编号", "频率"], ["1", "515.128MHz"]])
        self.assertEqual(confidence, [[0.99, 0.99], [0.99, 0.99]])

    def test_spatial_grid_splits_concatenated_known_header_labels_across_anchors(self):
        ocr_backend._load_runtime()

        def box(left, top, right, bottom):
            return [[left, top], [right, top], [right, bottom], [left, bottom]]

        anchors = [20, 70, 120, 170, 220, 270, 320]
        boxes = [
            box(120, 5, 220, 25),
            box(45, 35, 95, 55),
            box(215, 35, 325, 55),
            box(5, 65, 35, 85),
            box(198, 62, 338, 88),
        ]
        texts = ["生产日报", "基础信息", "测量与判定", "日期", "完成数量不良数量完成率"]
        for row_index in range(5):
            top = 100 + row_index * 30
            for column, anchor in enumerate(anchors):
                boxes.append(box(anchor - 16, top, anchor + 16, top + 18))
                texts.append(f"R{row_index}C{column}")

        grid, _, _, strong = ocr_backend._borderless_spatial_grid(
            np.asarray(boxes, dtype=float),
            texts,
            [0.99] * len(texts),
        )

        self.assertTrue(strong)
        self.assertEqual(
            grid[2],
            ["日期", "", "", "", "完成数量", "不良数量", "完成率"],
        )

    def test_borderless_spatial_grid_uses_complete_header_for_sparse_rows(self):
        ocr_backend._load_runtime()

        def box(left, top, right, bottom):
            return [[left, top], [right, top], [right, bottom], [left, bottom]]

        boxes = []
        texts = []
        anchors = [20, 70, 120, 170, 220, 270]
        rows = [
            ["编号", "项目", "结果", "单位", "结论", "说明"],
            ["S1", "含水率", "3.2", "%", "合格", "常规"],
            ["S2", "密度", "1.1", "g/cm³", "合格", ""],
            ["S3", "硬度", "72", "HA", "复测", "划痕"],
            ["S4", "色差", "", "ΔE", "待测", "未到"],
            ["S5", "强度", "28.6", "MPa", "合格", ""],
            ["S6", "伸长率", "145", "%", "合格", "室温"],
            ["S7", "耐压", "—", "kV", "不适用", ""],
        ]
        for row_index, row in enumerate(rows):
            top = 10 + row_index * 30
            for column_index, value in enumerate(row):
                if not value:
                    continue
                center = anchors[column_index]
                boxes.append(box(center - 15, top, center + 15, top + 18))
                texts.append(value)

        grid, _, _, strong = ocr_backend._borderless_spatial_grid(
            np.asarray(boxes, dtype=float),
            texts,
            [0.99] * len(texts),
        )

        self.assertTrue(strong)
        self.assertEqual(len(grid[0]), 6)
        self.assertEqual(grid[4], ["S4", "色差", "", "ΔE", "待测", "未到"])

    def test_borderless_spatial_grid_merges_one_wrinkle_shifted_cell_into_its_row(self):
        ocr_backend._load_runtime()

        def box(center_x, center_y):
            return [
                [center_x - 16, center_y - 8],
                [center_x + 16, center_y - 8],
                [center_x + 16, center_y + 8],
                [center_x - 16, center_y + 8],
            ]

        boxes = []
        texts = []
        anchors = (30, 90, 150, 210)
        rows = [
            ["编号", "名称", "数值", "状态"],
            ["A001", "设备一", "12", "正常"],
            ["A002", "设备二", "18", "复核"],
            ["A003", "设备三", "24", "正常"],
        ]
        for row_index, row in enumerate(rows):
            center_y = 20 + row_index * 34
            for column_index, value in enumerate(row):
                # A paper ridge moves one cell down far enough to exceed the
                # ordinary same-line threshold, while the other columns stay
                # on their physical row.
                shifted_y = center_y + 13 if row_index == 2 and column_index == 0 else center_y
                boxes.append(box(anchors[column_index], shifted_y))
                texts.append(value)

        grid, _, _, strong = ocr_backend._borderless_spatial_grid(
            np.asarray(boxes, dtype=float),
            texts,
            [0.99] * len(texts),
        )

        self.assertTrue(strong)
        self.assertEqual(len(grid), 4)
        self.assertEqual(grid[2], rows[2])

    def test_stacked_spatial_entry_is_split_across_confirmed_rows(self):
        grid = [["状态"], ["无无"], [""]]
        confidence = [[0.99], [0.98], [0.0]]
        stacked_entry = {
            "text": "无无",
            "top": 44.0,
            "bottom": 116.0,
            "center_x": 50.0,
            "score": 0.98,
        }
        geometry = {
            "anchors": [50.0],
            "row_centers": [20.0, 60.0, 100.0],
            "grouped_rows": [[], [stacked_entry], []],
        }

        ocr_backend._split_stacked_spatial_entries(grid, confidence, geometry)

        self.assertEqual(grid, [["状态"], ["无"], ["无"]])
        self.assertEqual(confidence, [[0.99], [0.98], [0.98]])

    def test_normal_height_spatial_entry_is_not_split(self):
        grid = [["状态"], ["无无"], [""]]
        confidence = [[0.99], [0.98], [0.0]]
        normal_entry = {
            "text": "无无",
            "top": 51.0,
            "bottom": 69.0,
            "center_x": 50.0,
            "score": 0.98,
        }
        geometry = {
            "anchors": [50.0],
            "row_centers": [20.0, 60.0, 100.0],
            "grouped_rows": [[], [normal_entry], []],
        }

        ocr_backend._split_stacked_spatial_entries(grid, confidence, geometry)

        self.assertEqual(grid, [["状态"], ["无无"], [""]])
        self.assertEqual(confidence, [[0.99], [0.98], [0.0]])

    def test_missing_spatial_cell_is_recovered_only_when_variants_agree(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "B", (48, 54), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 0), 2)
        grid = [
            ["A", "B", "C"],
            ["1", "2", "3"],
            ["4", "", "6"],
            ["7", "8", "9"],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[2][1] = 0.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0],
            "grouped_rows": [[{}], [{}], [{}], [{}]],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(txts=["B", "B"], scores=[0.96, 0.94])
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image, grid, confidence, geometry, engine
        )

        self.assertEqual(recovered[2][1], "B")
        self.assertEqual(recovered_confidence[2][1], 0.94)

    def test_missing_spatial_cell_is_not_guessed_when_variants_disagree(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "9", (48, 54), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 0), 2)
        grid = [
            ["A", "B", "C"],
            ["1", "2", "3"],
            ["4", "", "6"],
            ["7", "8", "9"],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[2][1] = -1.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0],
            "grouped_rows": [[{}], [{}], [{}], [{}]],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(txts=["9", "6"], scores=[0.97, 0.97])
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image, grid, confidence, geometry, engine
        )

        self.assertEqual(recovered[2][1], "")
        self.assertEqual(recovered_confidence[2][1], -1.0)

    def test_cross_model_recovery_withholds_when_mobile_model_is_missing(self):
        ocr_backend._load_runtime()
        image = np.full((120, 160, 3), 255, dtype=np.uint8)
        cv2.putText(image, "D", (52, 57), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 0), 2)
        grid = [
            ["日期", "生产线", "数量", "状态"],
            ["08-01", "A线", "10", "正常"],
            ["08-02", "", "12", "正常"],
            ["08-03", "", "14", "正常"],
            ["08-04", "", "16", "正常"],
            ["08-05", "B线", "18", "正常"],
        ]
        confidence = [[0.95] * 4 for _ in grid]
        confidence[2][1] = 0.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0, 140.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0, 90.0, 110.0],
            "grouped_rows": [[{}] for _ in grid],
            "first_structured_row": 0,
        }
        recognizer = lambda request: SimpleNamespace(
            txts=["D线", "D线"], scores=[0.96, 0.94]
        )
        engine = SimpleNamespace(
            text_rec=recognizer,
            server_text_rec=recognizer,
            v4_server_text_rec=recognizer,
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image,
            grid,
            confidence,
            geometry,
            engine,
            max_candidates=1,
            require_cross_model=True,
        )

        self.assertEqual(recovered[2][1], "")
        self.assertEqual(recovered_confidence[2][1], -2.0)

    def test_header_recovery_withholds_when_mobile_model_is_missing(self):
        ocr_backend._load_runtime()
        image = np.full((100, 160, 3), 255, dtype=np.uint8)
        cv2.putText(image, "PLAN", (47, 38), cv2.FONT_HERSHEY_SIMPLEX, 0.35, (0, 0, 0), 1)
        grid = [
            ["日期", "生产线", "数量", "状态"],
            ["08-01", "", "10", "正常"],
            ["08-02", "A线", "12", "正常"],
            ["08-03", "B线", "14", "正常"],
            ["08-04", "C线", "16", "正常"],
        ]
        confidence = [[0.95] * 4 for _ in grid]
        confidence[1][1] = 0.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0, 140.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0, 90.0],
            "grouped_rows": [[{}] for _ in grid],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["划数量", "划数量"], scores=[0.96, 0.95]
            ),
            server_text_rec=lambda request: SimpleNamespace(
                txts=["十划数量", "计划数量"], scores=[0.74, 0.71]
            ),
            v4_server_text_rec=lambda request: SimpleNamespace(
                txts=["划数塑", "划蚊塑"], scores=[0.80, 0.65]
            ),
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image,
            grid,
            confidence,
            geometry,
            engine,
            require_cross_model=True,
        )

        self.assertEqual(recovered[1][1], "")
        self.assertEqual(recovered_confidence[1][1], -2.0)

    def test_wrinkled_text_withholds_incomplete_cross_model_evidence(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "D", (48, 54), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 0), 2)
        grid = [
            ["日期", "生产线", "状态"],
            ["1", "A线", "正常"],
            ["2", "", "正常"],
            ["3", "B线", "正常"],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[2][1] = 0.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0],
            "grouped_rows": [[{}] for _ in grid],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["D线", "D线"], scores=[0.96, 0.72]
            ),
            server_text_rec=lambda request: SimpleNamespace(
                txts=["D线", "D线"], scores=[0.73, 0.64]
            ),
            v4_server_text_rec=lambda request: SimpleNamespace(
                txts=["D哦", "D"], scores=[0.82, 0.86]
            ),
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image,
            grid,
            confidence,
            geometry,
            engine,
            require_cross_model=True,
        )

        self.assertEqual(recovered[2][1], "")
        self.assertEqual(recovered_confidence[2][1], -2.0)

    def test_exact_motion_cell_uses_three_independent_model_families(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 255, dtype=np.uint8)
        cv2.putText(
            image,
            "TEXT",
            (43, 54),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.35,
            (0, 0, 0),
            1,
        )
        grid = [
            ["编号", "班次", "状态"],
            ["1", "", "正常"],
            ["2", "", "复核"],
            ["3", "", "正常"],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[1][1] = 0.0
        confidence[2][1] = -1.0
        confidence[3][1] = 0.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0],
            "column_boundaries": [0.0, 40.0, 80.0, 120.0],
            "row_boundaries": [0.0, 20.0, 40.0, 60.0, 80.0],
            "grouped_rows": [[] for _ in grid],
            "first_structured_row": 0,
        }

        def recognizer(normal_texts, motion_text):
            def recognize(request):
                count = len(request.img)
                if count == 1:
                    return SimpleNamespace(txts=[motion_text], scores=[0.96])
                return SimpleNamespace(txts=normal_texts, scores=[0.95] * count)

            return recognize

        engine = SimpleNamespace(
            fast_text_rec=recognizer([], "白班"),
            text_rec=recognizer(["自班", "自班"], "白班"),
            server_text_rec=recognizer(["自班", "自班"], "白班"),
            v4_server_text_rec=recognizer(["自班", "自班"], "白班"),
        )

        def false_mark_on_text(crop, allow_shorter=False):
            gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY) if crop.ndim == 3 else crop
            return "-" if float(np.mean(gray < 200)) > 0.003 else ""

        with patch.object(
            ocr_backend,
            "_recover_horizontal_mark",
            side_effect=false_mark_on_text,
        ):
            recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
                image,
                grid,
                confidence,
                geometry,
                engine,
                require_cross_model=True,
            )

        self.assertEqual(recovered[2][1], "白班")
        self.assertEqual(recovered_confidence[2][1], 0.77)
        self.assertEqual(recovered[2], ["2", "白班", "复核"])

    def test_blank_suggestion_requires_two_families_in_each_view(self):
        selected = ocr_backend._select_cross_view_blank_suggestion(
            [
                ("original", "medium", "D线", 0.97),
                ("original", "v5-server", "D线", 0.85),
                ("enhanced", "mobile", "D线", 0.92),
                ("enhanced", "medium", "D线", 0.99),
                ("enhanced", "v4-server", "D线", 0.89),
            ]
        )
        self.assertEqual(("D线", 0.77), selected)

    def test_blank_suggestion_keeps_single_cjk_with_cross_view_family_support(self):
        selected = ocr_backend._select_cross_view_blank_suggestion(
            [
                ("original", "medium", "高", 0.96),
                ("original", "v5-server", "高", 0.92),
                ("enhanced", "mobile", "高", 0.91),
                ("enhanced", "medium", "高", 0.94),
            ]
        )

        self.assertEqual(("高", 0.77), selected)

    def test_single_cjk_review_consensus_uses_independent_families(self):
        selected = ocr_backend._select_single_cjk_review_consensus(
            [
                [("高", 0.91), ("高", 0.89)],
                [("离", 0.70)],
                [("高", 0.93)],
                [],
            ]
        )
        competing = ocr_backend._select_single_cjk_review_consensus(
            [
                [("高", 0.91)],
                [("低", 0.92)],
                [("高", 0.93)],
                [("低", 0.94)],
            ]
        )

        self.assertEqual(("高", 0.77), selected)
        self.assertIsNone(competing)

    def test_single_cjk_review_adapter_requires_two_image_views(self):
        selected = ocr_backend._select_single_cjk_review_from_view_candidates(
            [
                ("original", "medium", "高", 0.96),
                ("original", "v5-server", "高", 0.93),
                ("enhanced", "medium", "高", 0.91),
            ]
        )
        one_view = ocr_backend._select_single_cjk_review_from_view_candidates(
            [
                ("original", "medium", "高", 0.96),
                ("original", "v5-server", "高", 0.93),
            ]
        )

        self.assertEqual(("高", 0.77), selected)
        self.assertIsNone(one_view)

    def test_blank_suggestion_rejects_one_view_or_competing_answers(self):
        one_view_only = ocr_backend._select_cross_view_blank_suggestion(
            [
                ("original", "mobile", "日班", 0.82),
                ("original", "v5-server", "日班", 0.93),
                ("enhanced", "medium", "白班", 0.95),
                ("enhanced", "v4-server", "班", 0.92),
            ]
        )
        competing = ocr_backend._select_cross_view_blank_suggestion(
            [
                ("original", "mobile", "白班", 0.91),
                ("original", "medium", "白班", 0.92),
                ("original", "v5-server", "日班", 0.93),
                ("original", "v4-server", "日班", 0.94),
                ("enhanced", "mobile", "白班", 0.91),
                ("enhanced", "medium", "白班", 0.92),
            ]
        )
        numeric = ocr_backend._select_cross_view_blank_suggestion(
            [
                ("original", "mobile", "98.1%", 0.95),
                ("original", "medium", "98.1%", 0.96),
                ("enhanced", "v5-server", "98.1%", 0.97),
                ("enhanced", "v4-server", "98.1%", 0.98),
            ]
        )
        self.assertIsNone(one_view_only)
        self.assertIsNone(competing)
        self.assertIsNone(numeric)

    def test_blank_numeric_suggestion_requires_all_active_families_in_both_views(self):
        complete = [
            (view, family, "-20", 0.96)
            for view in ("original", "enhanced")
            for family in ("mobile", "medium", "v5-server")
        ]
        missing_sign = list(complete)
        missing_sign[-1] = ("enhanced", "v5-server", "20", 0.96)

        self.assertEqual(
            ocr_backend._select_cross_view_blank_suggestion(complete),
            ("-20", 0.77),
        )
        self.assertIsNone(
            ocr_backend._select_cross_view_blank_suggestion(missing_sign)
        )

    def test_blank_integer_suggestion_requires_three_active_families_in_both_views(self):
        candidates = [
            (view, family, "277", 0.96)
            for view in ("original", "enhanced")
            for family in ("mobile", "medium", "v5-server")
        ]

        self.assertEqual(
            ocr_backend._select_cross_view_blank_suggestion(candidates),
            ("277", 0.77),
        )

    def test_server_guarded_unsharp_suggestion_recovers_name_only_as_review(self):
        selected = ocr_backend._select_server_guarded_unsharp_suggestion(
            [
                ("original", "medium", "赵敬", 0.81),
                ("original", "v5-server", "赵敏", 0.85),
                ("original", "v4-server", "赵敏", 0.81),
                ("unsharp", "medium", "赵敏", 0.84),
                ("unsharp", "v4-server", "赵敏", 0.85),
            ]
        )

        self.assertEqual(selected, ("赵敏", 0.77))

    def test_server_guarded_unsharp_suggestion_rejects_original_server_conflict(self):
        selected = ocr_backend._select_server_guarded_unsharp_suggestion(
            [
                ("original", "v5-server", "班", 0.93),
                ("unsharp", "medium", "日班", 0.85),
                ("unsharp", "v5-server", "日班", 0.96),
            ]
        )

        self.assertIsNone(selected)

    def test_stable_local_suggestion_is_review_only(self):
        selected = ocr_backend._select_stable_local_review_suggestion(
            [
                ("original", "mobile", "8线", 0.41),
                ("original", "medium", "B线", 0.98),
                ("original", "v5-server", "B线", 0.70),
                ("enhanced", "medium", "B线", 0.83),
                ("enhanced", "v5-server", "B线", 0.82),
                ("unsharp", "medium", "B线", 0.88),
            ]
        )

        self.assertEqual(selected, ("B线", 0.77))

    def test_stable_local_suggestion_rejects_competing_family_answer(self):
        selected = ocr_backend._select_stable_local_review_suggestion(
            [
                ("original", "medium", "白班", 0.85),
                ("enhanced", "medium", "白班", 0.95),
                ("original", "mobile", "日班", 0.83),
                ("original", "v5-server", "日班", 0.94),
                ("unsharp", "medium", "日班", 0.85),
                ("unsharp", "v4-server", "日班", 0.91),
            ]
        )

        self.assertIsNone(selected)

    def test_stable_local_suggestion_never_handles_numeric_text(self):
        selected = ocr_backend._select_stable_local_review_suggestion(
            [
                ("original", "medium", "20", 0.99),
                ("enhanced", "medium", "20", 0.98),
                ("unsharp", "medium", "20", 0.97),
            ]
        )

        self.assertIsNone(selected)

    def test_low_confidence_existing_text_requires_two_families_per_view(self):
        supported = [
            ("original", "medium", "白班", 0.99),
            ("original", "v5-server", "白班", 0.98),
            ("enhanced", "medium", "白班", 0.99),
            ("enhanced", "mobile", "白班", 0.93),
        ]
        unsupported = [
            ("original", "medium", "日班", 0.98),
            ("original", "v5-server", "日班", 0.70),
            ("enhanced", "medium", "日班", 0.92),
            ("enhanced", "mobile", "日班", 0.69),
        ]

        self.assertTrue(
            ocr_backend._cross_view_text_has_two_family_support(
                "白班", supported
            )
        )
        self.assertFalse(
            ocr_backend._cross_view_text_has_two_family_support(
                "日班", unsupported
            )
        )

    def test_low_confidence_fixed_cell_is_yellowed_but_never_rewritten(self):
        ocr_backend._load_runtime()
        image = np.full((60, 80, 3), 255, dtype=np.uint8)
        cv2.putText(
            image,
            "TEXT",
            (42, 38),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.25,
            (0, 0, 0),
            1,
        )
        grid = [["编号", "班次"], ["1", "日班"], ["2", "夜班"]]
        confidence = [[0.99, 0.99], [0.99, 0.84], [0.99, 0.96]]
        geometry = {
            "anchors": [20.0, 60.0],
            "row_centers": [10.0, 30.0, 50.0],
            "first_structured_row": 0,
        }

        def recognizer(text, score):
            return lambda request: SimpleNamespace(
                txts=[text] * len(request.img),
                scores=[score] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=recognizer("日", 0.91),
            text_rec=recognizer("日班", 0.98),
            server_text_rec=recognizer("日班", 0.70),
            v4_server_text_rec=recognizer("主班", 0.92),
        )

        ocr_backend._mark_low_confidence_fixed_cell_reviews(
            image, grid, confidence, geometry, engine
        )

        self.assertEqual(grid[1][1], "日班")
        self.assertEqual(confidence[1][1], 0.77)
        self.assertEqual(grid[2][1], "夜班")
        self.assertEqual(confidence[2][1], 0.96)

    def test_motion_consensus_rejects_repeated_views_from_one_model(self):
        selected = ocr_backend._select_cross_model_motion_consensus(
            [
                ("medium", "白班", 0.99),
                ("medium", "白班", 0.98),
                ("medium", "白班", 0.97),
            ]
        )

        self.assertIsNone(selected)

    def test_motion_consensus_rejects_a_two_model_competing_value(self):
        selected = ocr_backend._select_cross_model_motion_consensus(
            [
                ("mobile", "白班", 0.97),
                ("medium", "白班", 0.99),
                ("v5-server", "白班", 0.98),
                ("v4-server", "自班", 0.99),
                ("aux-server", "自班", 0.98),
            ]
        )

        self.assertIsNone(selected)

    def test_motion_consensus_uses_a_stricter_numeric_score_floor(self):
        selected = ocr_backend._select_cross_model_motion_consensus(
            [
                ("mobile", "515.472", 0.99),
                ("medium", "515.472", 0.83),
                ("v5-server", "515.472", 0.98),
                ("v4-server", "515.472", 0.97),
            ]
        )

        self.assertIsNone(selected)

    def test_motion_consensus_requires_all_four_models_for_one_cjk_character(self):
        three_models = ocr_backend._select_cross_model_motion_consensus(
            [
                ("mobile", "低", 0.99),
                ("v5-server", "低", 0.98),
                ("v4-server", "低", 0.97),
            ]
        )
        four_models = ocr_backend._select_cross_model_motion_consensus(
            [
                ("mobile", "低", 0.99),
                ("medium", "低", 0.98),
                ("v5-server", "低", 0.97),
                ("v4-server", "低", 0.96),
            ]
        )

        self.assertIsNone(three_models)
        self.assertEqual(four_models[0], "低")

    def test_motion_consensus_requires_medium_and_v5_for_long_text(self):
        selected = ocr_backend._select_cross_model_motion_consensus(
            [
                ("mobile", "表格结构识别", 0.99),
                ("medium", "表格结构识别", 0.98),
                ("v5-server", "表格结构识别", 0.97),
            ]
        )

        self.assertEqual(selected[0], "表格结构识别")

    def test_motion_cell_key_unifies_compatible_punctuation_but_keeps_spaces(self):
        self.assertEqual(
            ocr_backend._motion_cell_consensus_key("频率(MHz）"),
            ocr_backend._motion_cell_consensus_key("频率（mHz)"),
        )
        self.assertNotEqual(
            ocr_backend._motion_cell_consensus_key("Screenshot 2026.png"),
            ocr_backend._motion_cell_consensus_key("Screenshot2026.png"),
        )

    def test_dense_uncertain_page_grid_triggers_motion_cell_consensus(self):
        grid = [[f"R{row}C{column}" for column in range(5)] for row in range(6)]
        uncertain = [[0.72] * 5 for _ in range(6)]
        clear = [[0.96] * 5 for _ in range(6)]

        self.assertTrue(
            ocr_backend._page_grid_needs_motion_cell_consensus(grid, uncertain)
        )
        self.assertFalse(
            ocr_backend._page_grid_needs_motion_cell_consensus(grid, clear)
        )

    def test_structural_blanks_push_a_near_threshold_page_into_motion_consensus(self):
        grid = [[f"R{row}C{column}" for column in range(7)] for row in range(10)]
        confidence = [[0.96] * 7 for _ in range(10)]
        low_cells = [(row, column) for row in range(2, 10) for column in range(7)][:16]
        for row, column in low_cells:
            confidence[row][column] = 0.72
        for row, column in [(0, 1), (0, 5), (1, 4), (1, 5), (1, 6)]:
            grid[row][column] = ""
            confidence[row][column] = 0.0

        self.assertTrue(
            ocr_backend._page_grid_needs_motion_cell_consensus(grid, confidence)
        )

    def test_legitimate_sparse_blanks_do_not_trigger_motion_consensus(self):
        grid = [["", "", "", ""] for _ in range(8)]
        confidence = [[0.0] * 4 for _ in range(8)]
        for row in range(8):
            grid[row][0] = f"ID-{row}"
            confidence[row][0] = 0.96

        self.assertFalse(
            ocr_backend._page_grid_needs_motion_cell_consensus(grid, confidence)
        )

    def test_standard_consensus_selects_risks_without_full_table_medium(self):
        ocr_backend._load_runtime()
        image = np.full((180, 450, 3), 245, dtype=np.uint8)
        columns = [column * 50 for column in range(10)]
        rows = [row * 20 for row in range(10)]
        grid = [
            ["编号", "名称", "类别", "状态", "数值", "备注", "区域", "日期", "人员"],
            [f"二级{column}" for column in range(9)],
        ] + [
            [f"R{row}C{column}" for column in range(9)]
            for row in range(2, 9)
        ]
        confidence = [[0.99] * 9 for _ in grid]
        grid[2][4] = "IP:192.168.3.225"
        confidence[4][3] = 0.70
        grid[5][2] = ""
        confidence[5][2] = 0.0
        cv2.putText(
            image,
            "X",
            (columns[2] + 18, rows[5] + 15),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.45,
            (20, 20, 20),
            1,
        )
        locked = {(3, 1)}

        selected = ocr_backend._standard_physical_consensus_locations(
            image, grid, confidence, columns, rows, locked
        )

        self.assertIn((0, 0), selected)
        self.assertIn((1, 8), selected)
        self.assertIn((2, 1), selected)
        self.assertIn((4, 3), selected)
        self.assertIn((2, 4), selected)
        self.assertIn((5, 2), selected)
        self.assertNotIn((3, 1), selected)
        self.assertLess(len(selected), len(grid) * len(grid[0]))

    def test_standard_consensus_preserves_full_medium_for_small_table(self):
        image = np.full((140, 300, 3), 245, dtype=np.uint8)
        columns = [column * 50 for column in range(7)]
        rows = [row * 20 for row in range(8)]
        grid = [
            [f"R{row}C{column}" for column in range(6)]
            for row in range(7)
        ]
        confidence = [[0.99] * 6 for _ in grid]
        locked = {(3, 1)}

        selected = ocr_backend._standard_physical_consensus_locations(
            image, grid, confidence, columns, rows, locked
        )

        self.assertEqual(len(selected), 41)
        self.assertNotIn((3, 1), selected)
        self.assertIn((6, 5), selected)

    def test_standard_consensus_selects_repeated_leading_glyph_omission(self):
        ocr_backend._load_runtime()
        image = np.full((200, 300, 3), 245, dtype=np.uint8)
        columns = [0, 100, 200, 300]
        rows = [row * 20 for row in range(11)]
        grid = [
            ["编号", "区域", "数量"],
            ["1", "一车间", "10"],
            ["2", "A区", "11"],
            ["3", "车间", "12"],
            ["4", "现场复核", "13"],
            ["5", "一车间", "14"],
            ["6", "A区", "15"],
            ["7", "华东库", "16"],
            ["8", "批次L08", "17"],
            ["9", "常规项目", "18"],
        ]
        confidence = [[0.99] * 3 for _ in grid]

        selected = ocr_backend._standard_physical_consensus_locations(
            image,
            grid,
            confidence,
            columns,
            rows,
            set(),
            preserve_recognition_batches=False,
        )

        self.assertIn((3, 1), selected)

    def test_standard_consensus_trusts_only_high_confidence_valid_time(self):
        ocr_backend._load_runtime()
        image = np.full((180, 450, 3), 245, dtype=np.uint8)
        columns = [column * 50 for column in range(10)]
        rows = [row * 20 for row in range(10)]
        grid = [
            ["编号", "上班时间", "字段2", "字段3", "字段4", "字段5", "字段6", "字段7", "字段8"],
            *[
                [str(row), "12:30", "甲", "乙", "丙", "丁", "戊", "己", "庚"]
                for row in range(1, 9)
            ],
        ]
        confidence = [[0.99] * 9 for _ in grid]
        confidence[3][1] = 0.90
        grid[4][1] = "25:70"

        selected = ocr_backend._standard_physical_consensus_locations(
            image,
            grid,
            confidence,
            columns,
            rows,
            set(),
            preserve_recognition_batches=False,
        )

        self.assertNotIn((2, 1), selected)
        self.assertIn((3, 1), selected)
        self.assertIn((4, 1), selected)

    def test_large_consensus_keeps_all_real_risks_without_batch_companions(self):
        ocr_backend._load_runtime()
        image = np.full((180, 450, 3), 245, dtype=np.uint8)
        columns = [column * 50 for column in range(10)]
        rows = [row * 20 for row in range(10)]
        grid = [
            [f"R{row}C{column}" for column in range(9)]
            for row in range(9)
        ]
        confidence = [[0.99] * 9 for _ in grid]
        confidence[4][3] = 0.70

        expanded = ocr_backend._standard_physical_consensus_locations(
            image, grid, confidence, columns, rows, set()
        )
        exact_risks = ocr_backend._standard_physical_consensus_locations(
            image,
            grid,
            confidence,
            columns,
            rows,
            set(),
            preserve_recognition_batches=False,
        )

        self.assertIn((4, 3), exact_risks)
        self.assertTrue(exact_risks.issubset(expanded))
        self.assertLess(len(exact_risks), len(expanded))

    def test_large_photo_low_confidence_review_is_bounded_and_excludes_headers(self):
        grid = [
            ["标题", "", "分组"],
            ["3069", "℃", "V4.7.7"],
            ["", "6193", "99.85%"],
        ]
        confidence = [
            [-1.0, 0.0, -1.0],
            [0.62, 0.68, 0.72],
            [0.0, 0.65, 0.75],
        ]

        selected = ocr_backend._bounded_large_photo_low_confidence_cells(
            grid,
            confidence,
            {(1, 1)},
            maximum_cells=4,
        )

        self.assertEqual(selected, {(1, 0), (1, 2), (2, 1), (2, 2)})
        self.assertEqual(
            ocr_backend._bounded_large_photo_low_confidence_cells(
                grid,
                confidence,
                set(),
                maximum_cells=4,
            ),
            set(),
        )
        self.assertEqual(
            ocr_backend._bounded_large_photo_low_confidence_cells(
                grid,
                confidence[:-1],
                set(),
            ),
            set(),
        )

    def test_visible_blank_body_review_is_bounded_and_excludes_headers_and_locks(self):
        grid = [
            ["标题", "", ""],
            ["字段A", "字段B", "字段C"],
            ["1", "", "3"],
            ["4", "5", ""],
            ["7", "", "9"],
        ]

        selected = ocr_backend._bounded_visible_blank_body_cells(
            grid,
            {(2, 1)},
            maximum_cells=1,
        )

        self.assertEqual(selected, {(3, 2)})

    def test_non_unit_visible_blank_cells_leave_unit_columns_to_glyph_recovery(self):
        grid = [
            ["序号", "单位", "结果"],
            ["1", "", ""],
            ["2", "", ""],
        ]

        selected = ocr_backend._non_unit_visible_blank_cells(
            grid,
            {(1, 1), (1, 2), (2, 1), (2, 2)},
        )

        self.assertEqual(selected, {(1, 2), (2, 2)})

    def test_certified_selective_review_skips_clear_simple_header_cells(self):
        grid = [[
            "编号", "频率", "信号类型", "占用带宽", "调制方式",
            "信号名称", "功率", "识别编号", "合法/非法",
        ]] + [
            [
                str(row), "537.714MHz", "扩频信号", "200kHz", "BPSK",
                "测控链路", "-30", f"A{row:03d}B11", "合法",
            ]
            for row in range(1, 9)
        ]
        confidence = [[0.99] * 9 for _ in grid]

        self.assertEqual(
            ocr_backend._certified_selective_header_risks(grid, confidence),
            set(),
        )
        confidence[0][3] = 0.70
        self.assertEqual(
            ocr_backend._certified_selective_header_risks(grid, confidence),
            {(0, 3)},
        )

    def test_certified_selective_review_keeps_complete_multilevel_headers(self):
        grid = [
            ["设备验收记录", "", "", ""],
            ["基本信息", "", "测量信息", ""],
            ["编号", "名称", "频率", "单位"],
        ] + [
            [str(row), f"设备{row}", f"{100 + row}MHz", "MHz"]
            for row in range(1, 5)
        ]
        confidence = [[0.99] * 4 for _ in grid]

        selected = ocr_backend._certified_selective_header_risks(
            grid, confidence
        )

        self.assertEqual(
            selected,
            {
                (0, 0),
                (0, 1),
                (0, 2),
                (0, 3),
                (1, 0),
                (1, 1),
                (1, 2),
                (1, 3),
                (2, 0),
                (2, 1),
                (2, 2),
                (2, 3),
            },
        )

    def test_multilevel_header_blank_anchors_and_dense_four_digit_values_are_risks(self):
        grid = [
            ["基础信息", "", "目标与测量", ""],
            ["序号", "名称", "计划数量", "完成数量"],
        ] + [
            [str(row), f"设备{row}", f"{4000 + row}", f"{6000 + row}"]
            for row in range(1, 7)
        ]
        confidence = [[0.99] * 4 for _ in grid]

        self.assertEqual(
            ocr_backend._multilevel_header_rows_for_review(grid), {0, 1}
        )
        self.assertEqual(
            ocr_backend._certified_selective_header_risks(grid, confidence),
            {(row, column) for row in range(2) for column in range(4)},
        )
        self.assertEqual(
            ocr_backend._dense_four_digit_value_risks(grid),
            {
                (row, column)
                for row in range(2, len(grid))
                for column in (2, 3)
            },
        )

    def test_dense_quantity_magnitude_outlier_selects_only_isolated_low_value(self):
        grid = [["序号", "订购数量", "备注"]] + [
            [str(row), value, "正常"]
            for row, value in enumerate(
                ["4738", "6291", "7160", "0006", "8420", "9155", "3688"],
                start=1,
            )
        ]

        self.assertEqual(
            ocr_backend._dense_quantity_magnitude_outlier_risks(grid),
            {(4, 1)},
        )

    def test_dense_quantity_magnitude_outlier_accepts_two_spreadsheet_ruler_rows(self):
        grid = [
            ["", "A", "B"],
            ["1", "", ""],
            ["基础信息", "目标与测量", ""],
            ["序号", "订购数量", "备注"],
        ] + [
            [str(row), value, "正常"]
            for row, value in enumerate(
                ["4738", "6291", "7160", "0006", "8420", "9155", "3688"],
                start=1,
            )
        ]

        self.assertIn(
            (7, 1),
            ocr_backend._dense_quantity_magnitude_outlier_risks(grid),
        )

    def test_dense_digit_shape_confusion_risks_are_bounded(self):
        grid = [["编号", "测量值"]] + [
            [str(row), value]
            for row, value in enumerate(
                ["6066", "1234", "8606", "166", "2966", "5555"],
                start=1,
            )
        ]

        self.assertEqual(
            ocr_backend._dense_digit_shape_confusion_risks(grid),
            {(1, 1), (3, 1), (4, 1), (5, 1)},
        )
        self.assertEqual(
            ocr_backend._dense_digit_shape_confusion_risks(
                grid, maximum_cells=3
            ),
            set(),
        )

    def test_bounded_numeric_and_quantity_glyph_risks_require_column_evidence(self):
        grid = [["序号", "测量值", "订购数量", "备注"]] + [
            [str(row), measured, quantity, quantity]
            for row, (measured, quantity) in enumerate(
                [
                    ("4738", "4738"),
                    ("6291", "6291"),
                    ("7160", "£999"),
                    ("390", "8420"),
                    ("9155", "9155"),
                    ("1666", "3688"),
                    ("5522", "5522"),
                    ("91196", "1234"),
                ],
                start=1,
            )
        ]

        self.assertEqual(
            ocr_backend._dense_numeric_length_review_risks(grid),
            {(4, 1), (8, 1)},
        )
        self.assertEqual(
            ocr_backend._repeated_digit_shape_risks(grid), {(6, 1)}
        )
        self.assertEqual(
            ocr_backend._invalid_quantity_glyph_risks(grid), {(3, 2)}
        )
        self.assertNotIn((3, 3), ocr_backend._invalid_quantity_glyph_risks(grid))

    def test_original_resolution_recovers_bounded_five_digit_insertion(self):
        ocr_backend._load_runtime()
        grid = [["数量"], *[[value] for value in [
            "4738", "6291", "7160", "390", "9155", "5522", "3688", "91196",
        ]]]
        confidence = [[0.99] for _ in grid]

        class OriginalPixelEngine:
            @staticmethod
            def text_rec(request):
                return SimpleNamespace(
                    txts=["9196"] * len(request.img),
                    scores=[0.99] * len(request.img),
                )

            @staticmethod
            def server_text_rec(request):
                return SimpleNamespace(
                    txts=["9196"] * len(request.img),
                    scores=[0.99] * len(request.img),
                )

        with patch.object(
            ocr_backend,
            "_read_image",
            return_value=np.full((720, 200, 3), 245, dtype=np.uint8),
        ):
            scores, recovered = (
                ocr_backend._recover_original_resolution_risk_cells(
                    Path("source.jpg"),
                    0.5,
                    {
                        "corners": [
                            [0.0, 0.0], [99.0, 0.0],
                            [99.0, 359.0], [0.0, 359.0],
                        ],
                    },
                    (360, 100),
                    [0, 100],
                    list(range(0, 361, 40)),
                    grid,
                    confidence,
                    OriginalPixelEngine(),
                )
            )

        self.assertEqual(grid[8][0], "9196")
        self.assertEqual(confidence[8][0], 0.77)
        self.assertEqual(recovered, {(8, 0)})
        self.assertEqual(scores, [0.99] * 2)

        text_grid = [["状态"], *[["正常"] for _ in range(7)], ["崇"]]
        text_confidence = [[0.99] for _ in text_grid]

        class OriginalTextEngine:
            @staticmethod
            def text_rec(request):
                return SimpleNamespace(
                    txts=["正常"] * len(request.img),
                    scores=[0.99] * len(request.img),
                )

            @staticmethod
            def server_text_rec(request):
                return SimpleNamespace(
                    txts=["正常"] * len(request.img),
                    scores=[0.99] * len(request.img),
                )

        with (
            patch.object(
                ocr_backend,
                "_read_image",
                return_value=np.full((720, 200, 3), 245, dtype=np.uint8),
            ),
            patch.object(
                ocr_backend,
                "_dense_repeated_text_residual_risks",
                return_value={(8, 0)},
            ),
        ):
            _, text_recovered = ocr_backend._recover_original_resolution_risk_cells(
                Path("source.jpg"),
                0.5,
                {
                    "corners": [
                        [0.0, 0.0], [99.0, 0.0],
                        [99.0, 359.0], [0.0, 359.0],
                    ],
                },
                (360, 100),
                [0, 100],
                list(range(0, 361, 40)),
                text_grid,
                text_confidence,
                OriginalTextEngine(),
            )

        self.assertEqual(text_grid[8][0], "正常")
        self.assertEqual(text_recovered, {(8, 0)})

        ordinal_grid = [["序号"], *[
            ["6" if index == 9 else str(index)]
            for index in range(1, 11)
        ]]
        ordinal_confidence = [[0.99] for _ in ordinal_grid]

        class OriginalOrdinalEngine:
            @staticmethod
            def text_rec(request):
                return SimpleNamespace(
                    txts=["9"] * len(request.img),
                    scores=[0.999] * len(request.img),
                )

            @staticmethod
            def server_text_rec(request):
                return SimpleNamespace(
                    txts=["9"] * len(request.img),
                    scores=[0.999] * len(request.img),
                )

        with patch.object(
            ocr_backend,
            "_read_image",
            return_value=np.full((880, 200, 3), 245, dtype=np.uint8),
        ):
            _, ordinal_recovered = ocr_backend._recover_original_resolution_risk_cells(
                Path("source.jpg"),
                0.5,
                {
                    "corners": [
                        [0.0, 0.0], [99.0, 0.0],
                        [99.0, 439.0], [0.0, 439.0],
                    ],
                },
                (440, 100),
                [0, 100],
                list(range(0, 441, 40)),
                ordinal_grid,
                ordinal_confidence,
                OriginalOrdinalEngine(),
            )

        self.assertEqual(ordinal_grid[9][0], "9")
        self.assertEqual(ordinal_recovered, {(9, 0)})

        unit_grid = [["单位"], ["V"]]
        unit_confidence = [[0.99], [0.77]]

        class OriginalUnitEngine:
            @staticmethod
            def text_rec(request):
                return SimpleNamespace(
                    txts=["A"] * len(request.img),
                    scores=[0.88] * len(request.img),
                )

            @staticmethod
            def server_text_rec(request):
                return SimpleNamespace(
                    txts=["A"] * len(request.img),
                    scores=[0.87] * len(request.img),
                )

        with (
            patch.object(
                ocr_backend,
                "_read_image",
                return_value=np.full((80, 100, 3), 245, dtype=np.uint8),
            ),
            patch.object(
                ocr_backend,
                "_short_unit_visual_confusion_risks",
                return_value={(1, 0)},
            ),
        ):
            _, unit_recovered = ocr_backend._recover_original_resolution_risk_cells(
                Path("source.jpg"),
                1.0,
                {
                    "corners": [
                        [0.0, 0.0], [99.0, 0.0],
                        [99.0, 79.0], [0.0, 79.0],
                    ],
                },
                (80, 100),
                [0, 100],
                [0, 40, 80],
                unit_grid,
                unit_confidence,
                OriginalUnitEngine(),
            )

        self.assertEqual(unit_grid[1][0], "A")
        self.assertEqual(unit_recovered, {(1, 0)})

    def test_short_unit_and_person_name_risks_require_explicit_headers(self):
        grid = [
            ["项目", "单位", "负责人", "备注"],
            ["电压", "A", "陈吴", "陈吴"],
            ["速度", "m/s", "李明", "V"],
            ["电流", "mA", "王强", "正常"],
        ]

        self.assertEqual(
            ocr_backend._short_unit_visual_confusion_risks(grid),
            {(1, 1), (2, 1)},
        )
        self.assertEqual(
            ocr_backend._person_name_glyph_confusion_risks(grid),
            {(1, 2)},
        )

    def test_clipped_short_unit_requires_enhanced_and_alternate_agreement(self):
        ocr_backend._load_runtime()
        image = np.full((60, 160, 3), 240, dtype=np.uint8)
        grid = [["项目", "单位"], ["速度", "m/s"]]
        confidence = [[0.99, 0.99], [0.99, 0.91]]

        def medium(request):
            return SimpleNamespace(
                txts=["m/s", "mm/s", "mm/s", "mm/s"],
                scores=[0.97, 0.91, 0.995, 0.993],
            )

        def alternate(request):
            return SimpleNamespace(
                txts=["m/s", "mm/s", "m/s", "m/s"],
                scores=[0.99, 0.92, 0.99, 0.99],
            )

        engine = SimpleNamespace(text_rec=medium, server_text_rec=alternate)
        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 30, 3), 220, dtype=np.uint8),
        ):
            scores = ocr_backend._recover_clipped_short_unit_multiview(
                image,
                grid,
                confidence,
                [0, 80, 160],
                [0, 30, 60],
                engine,
            )

        self.assertEqual(grid[1][1], "mm/s")
        self.assertEqual(confidence[1][1], 0.77)
        self.assertEqual(len(scores), 8)

    def test_repeated_digits_require_all_models_and_views_to_agree(self):
        ocr_backend._load_runtime()
        image = np.full((60, 160, 3), 240, dtype=np.uint8)
        grid = [["序号", "数量"], ["1", "1666"]]
        confidence = [[0.99, 0.99], [0.99, 0.42]]

        def recognizer(text):
            return lambda request: SimpleNamespace(
                txts=[text] * len(request.img),
                scores=[0.99] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=recognizer("9991"),
            text_rec=recognizer("9991"),
            server_text_rec=recognizer("9991"),
        )
        with patch.object(
            ocr_backend,
            "_unit_cell_morphology_views",
            return_value=[np.full((12, 30, 3), 220, dtype=np.uint8)],
        ):
            scores, recovered = ocr_backend._recover_repeated_digits_multimodel(
                image,
                grid,
                confidence,
                [0, 80, 160],
                [0, 30, 60],
                engine,
            )

        self.assertEqual(recovered, {(1, 1)})
        self.assertEqual(grid[1][1], "9991")
        self.assertEqual(confidence[1][1], 0.77)
        self.assertEqual(len(scores), 6)

        grid[1][1] = "1666"
        engine.text_rec = recognizer("1666")
        with patch.object(
            ocr_backend,
            "_unit_cell_morphology_views",
            return_value=[np.full((12, 30, 3), 220, dtype=np.uint8)],
        ):
            _, recovered = ocr_backend._recover_repeated_digits_multimodel(
                image,
                grid,
                confidence,
                [0, 80, 160],
                [0, 30, 60],
                engine,
            )
        self.assertEqual(recovered, set())
        self.assertEqual(grid[1][1], "1666")

        calls = []

        def counted_recognizer(request):
            calls.append(len(request.img))
            return SimpleNamespace(
                txts=["9991"] * len(request.img),
                scores=[0.99] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=counted_recognizer,
            text_rec=counted_recognizer,
            server_text_rec=counted_recognizer,
        )
        _, recovered = ocr_backend._recover_repeated_digits_multimodel(
            image,
            grid,
            confidence,
            [0, 80, 160],
            [0, 30, 60],
            engine,
            excluded_cells={(1, 1)},
        )
        self.assertEqual(recovered, set())
        self.assertEqual(calls, [])

    def test_dense_page_risks_can_skip_full_identifier_column_for_certified_review(self):
        ocr_backend._load_runtime()
        image = np.full((200, 200, 3), 255, dtype=np.uint8)
        grid = [["编号", "信号名称"]] + [
            [str(row), f"链路{row}"] for row in range(1, 9)
        ]
        confidence = [[0.99, 0.99] for _ in grid]
        columns = [0, 100, 200]
        rows = [row * 20 for row in range(10)]

        default = ocr_backend._dense_screen_page_risk_locations(
            image, grid, confidence, columns, rows
        )
        selective = ocr_backend._dense_screen_page_risk_locations(
            image,
            grid,
            confidence,
            columns,
            rows,
            include_identifier_columns=False,
        )

        self.assertIn((1, 1), default)
        self.assertNotIn((1, 1), selective)

    def test_large_consensus_isolates_each_risk_cell_recognition(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 240, dtype=np.uint8)
        grid = [["A", "B"]]
        confidence = [[0.70, 0.70]]
        calls: list[int] = []

        def recognizer(request):
            calls.append(len(request.img))
            return SimpleNamespace(
                txts=["X"] * len(request.img),
                scores=[0.95] * len(request.img),
            )

        engine = SimpleNamespace(
            text_rec=recognizer,
            server_text_rec=recognizer,
        )
        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 30, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 60, 120],
                [0, 80],
                engine,
                view_mode="standard",
                selected_cells={(0, 0), (0, 1)},
                isolated_recognition=True,
            )

        self.assertEqual(calls, [1, 1, 1, 1])

    def test_homogeneous_unit_risks_can_use_bounded_recognition_groups(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 240, dtype=np.uint8)
        grid = [["A", "V"]]
        confidence = [[0.70, 0.70]]
        calls: list[int] = []

        def recognizer(request):
            calls.append(len(request.img))
            return SimpleNamespace(
                txts=["A", "V"][: len(request.img)],
                scores=[0.95] * len(request.img),
            )

        engine = SimpleNamespace(
            text_rec=recognizer,
            server_text_rec=recognizer,
        )
        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 30, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 60, 120],
                [0, 80],
                engine,
                view_mode="standard",
                selected_cells={(0, 0), (0, 1)},
                isolated_recognition=True,
                isolated_group_size=4,
            )

        self.assertEqual(calls, [2, 2])

    def test_strict_visible_blank_requires_small_and_both_medium_views_and_stays_review(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 240, dtype=np.uint8)
        grid = [["状态"], [""]]
        confidence = [[0.99], [0.0]]
        calls: list[str] = []

        def recognizer(family):
            def run(request):
                calls.append(family)
                return SimpleNamespace(
                    txts=["通过"] * len(request.img),
                    scores=[0.999] * len(request.img),
                )

            return run

        engine = SimpleNamespace(
            fast_text_rec=recognizer("small"),
            text_rec=recognizer("medium"),
            server_text_rec=recognizer("alternate"),
        )
        verified_cells = set()
        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 30, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 120],
                [0, 40, 80],
                engine,
                view_mode="standard",
                selected_cells={(1, 0)},
                strict_blank_cross_model=True,
                verified_cells_out=verified_cells,
            )

        self.assertEqual(grid[1][0], "通过")
        self.assertEqual(confidence[1][0], 0.77)
        self.assertEqual(calls, ["small", "medium", "alternate"])
        self.assertEqual(verified_cells, {(1, 0)})

    def test_strict_visible_blank_keeps_disagreement_empty(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 240, dtype=np.uint8)
        grid = [["状态"], [""]]
        confidence = [[0.99], [0.0]]

        def recognizer(text):
            return lambda request: SimpleNamespace(
                txts=[text] * len(request.img),
                scores=[0.999] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=recognizer("通过"),
            text_rec=recognizer("通过"),
            server_text_rec=recognizer("通遇"),
        )
        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 30, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 120],
                [0, 40, 80],
                engine,
                view_mode="standard",
                selected_cells={(1, 0)},
                strict_blank_cross_model=True,
            )

        self.assertEqual(grid[1][0], "")
        self.assertLessEqual(confidence[1][0], 0.0)

    def test_quantity_magnitude_outlier_uses_isolated_medium_agreement_for_review(self):
        ocr_backend._load_runtime()
        image = np.full((320, 160, 3), 240, dtype=np.uint8)
        grid = [["序号", "订购数量"]] + [
            [str(row), value]
            for row, value in enumerate(
                ["4738", "6291", "7160", "0006", "8420", "9155", "3688"],
                start=1,
            )
        ]
        confidence = [[0.99, 0.99] for _ in grid]
        columns = [0, 80, 160]
        rows = [row * 40 for row in range(9)]

        def recognizer(request):
            text = "9000" if len(request.img) == 1 else "0006"
            return SimpleNamespace(
                txts=[text] * len(request.img),
                scores=[0.99] * len(request.img),
            )

        engine = SimpleNamespace(
            text_rec=recognizer,
            server_text_rec=recognizer,
        )
        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 30, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                columns,
                rows,
                engine,
                view_mode="standard",
                selected_cells={(1, 0), (4, 1)},
            )

        self.assertEqual(grid[4][1], "9000")
        self.assertEqual(confidence[4][1], 0.77)

    def test_percentage_prefix_outlier_uses_isolated_medium_agreement(self):
        ocr_backend._load_runtime()
        image = np.full((360, 120, 3), 240, dtype=np.uint8)
        grid = [["合格率"]] + [
            ["%60.66" if row == 7 else f"9{row}.09%"]
            for row in range(1, 9)
        ]
        confidence = [[0.99] for _ in grid]
        medium_calls = 0

        def medium_recognizer(request):
            nonlocal medium_calls
            medium_calls += 1
            text = "%60.66" if medium_calls == 1 else "99.09%"
            return SimpleNamespace(
                txts=[text] * len(request.img),
                scores=[0.99] * len(request.img),
            )

        def alternate_recognizer(request):
            return SimpleNamespace(
                txts=["99.09%"] * len(request.img),
                scores=[0.99] * len(request.img),
            )

        engine = SimpleNamespace(
            text_rec=medium_recognizer,
            server_text_rec=alternate_recognizer,
        )
        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 30, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 120],
                [row * 40 for row in range(10)],
                engine,
                view_mode="standard",
                selected_cells={(7, 0)},
            )

        self.assertEqual(grid[7][0], "99.09%")
        self.assertEqual(confidence[7][0], 0.77)


    def test_signal_name_consensus_repairs_page_fragment_order(self):
        ocr_backend._load_runtime()
        image = np.full((80, 160, 3), 240, dtype=np.uint8)
        grid = [["信号名称"], ["28k发3停6 35.4k bpsk"]]
        confidence = [[0.99], [0.99]]

        def recognizer(text):
            return lambda request: SimpleNamespace(
                txts=[text] * len(request.img),
                scores=[0.97] * len(request.img),
            )

        engine = SimpleNamespace(
            text_rec=recognizer("bpsk 35.4k 28k发3停6"),
            server_text_rec=recognizer("bpsk35.4k28k发3停6"),
            v4_server_text_rec=lambda _: self.fail("retired v4 must never be invoked"),
        )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 40, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 160],
                [0, 40, 80],
                engine,
                view_mode="standard",
                selected_cells={(1, 0)},
            )

        self.assertEqual(grid[1][0], "bpsk 35.4k 28k发3停6")

    def test_signal_name_consensus_preserves_visible_page_underscores(self):
        ocr_backend._load_runtime()
        image = np.full((80, 160, 3), 240, dtype=np.uint8)
        grid = [["信号名称"], ["qpsk_30.7k_30k_0.1"]]
        confidence = [[0.99], [0.99]]

        def recognizer(request):
            return SimpleNamespace(
                txts=["qpsk 30.7k 30k 0.1"] * len(request.img),
                scores=[0.97] * len(request.img),
            )

        engine = SimpleNamespace(
            text_rec=recognizer,
            server_text_rec=recognizer,
            v4_server_text_rec=lambda _: self.fail("retired v4 must never be invoked"),
        )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 40, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 160],
                [0, 40, 80],
                engine,
                view_mode="standard",
                selected_cells={(1, 0)},
            )

        self.assertEqual(grid[1][0], "qpsk_30.7k_30k_0.1")

    def test_unique_repeated_semantic_value_repairs_one_glyph_blur(self):
        grid = [
            ["编号", "备注"],
            ["A001", "校准完成"],
            ["A002", "接口检查"],
            ["A003", "校准完成"],
            ["A004", "准完成"],
        ]
        target = ocr_backend._repeated_semantic_column_target(
            grid,
            4,
            1,
            [
                ("medium", "交准完成", 0.88),
                ("medium-alternate", "准完成", 0.91),
            ],
        )

        self.assertEqual(target, "校准完成")

    def test_repeated_semantic_value_requires_a_unique_near_target(self):
        grid = [
            ["编号", "备注"],
            ["A001", "校准完成"],
            ["A002", "接口检查"],
            ["A003", "接口检查"],
            ["A004", "完成"],
        ]
        target = ocr_backend._repeated_semantic_column_target(
            grid,
            4,
            1,
            [
                ("medium", "交付完成", 0.91),
                ("medium-alternate", "准完成", 0.90),
            ],
        )

        self.assertEqual(target, "")

    def test_strong_dense_raw_spatial_layout_survives_before_text_verification(self):
        grid = [
            ["项目任务跟踪表", "", "", "", "", "", ""],
            ["基础信息", "", "", "测量与判定", "", "", ""],
            ["任务编号", "任务名称", "优先级", "进度", "负责人", "计划完成", "当前状态"],
            *[
                [f"TASK-{index:03d}", "界面设计", "中", "7%", "陈晨", "2026-08-02", "已完成"]
                for index in range(1, 7)
            ],
        ]

        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 7, "role": "title"}
        ]
        self.assertTrue(
            ocr_backend._raw_spatial_layout_is_trustworthy(grid, spans, True)
        )
        self.assertFalse(
            ocr_backend._raw_spatial_layout_is_trustworthy(grid, spans, False)
        )

    def test_spatial_centers_convert_to_bounded_physical_cells(self):
        geometry = {
            "anchors": [88.25, 158.0, 207.0],
            "row_centers": [77.5, 108.5, 140.0, 173.0],
        }
        boundaries = ocr_backend._spatial_geometry_cell_boundaries(
            geometry,
            (240, 300),
            4,
            3,
        )

        self.assertEqual(boundaries, ([53, 123, 182, 232], [62, 93, 124, 156, 190]))

    def test_percentage_column_repairs_a_single_blurred_percent_sign(self):
        grid = [["完成率"], *[[value] for value in (
            "98.2%", "98.4%", "98.5%", "98.7%", "98.8%", "99.0%",
            "99.1%", "99.3%", "99.4%", "99.6%", "99.8%", "98.19",
        )]]
        confidence = [[0.95] for _ in grid]

        repaired = ocr_backend._repair_percentage_symbol_confusions(
            grid, confidence
        )

        self.assertEqual(repaired, [(12, 0, "98.1%")])
        self.assertEqual(grid[12][0], "98.1%")
        self.assertEqual(confidence[12][0], 0.77)

    def test_motion_cell_consensus_repairs_precision_and_sign_in_one_pass(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 240, dtype=np.uint8)
        grid = [["", "功率(d8m)"], ["515472", "20"]]
        confidence = [[0.60, 0.60], [0.60, 0.60]]

        def recognizer(texts):
            return lambda request: SimpleNamespace(
                txts=texts[: len(request.img)],
                scores=[0.96] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=recognizer(
                ["频率(MHz）", "功率(dBm)", "515472", "20"]
            ),
            text_rec=recognizer(
                ["频率(mHz)", "功率(dBm)", "515.472", "-20"]
            ),
            server_text_rec=recognizer(
                ["频率（MHz)", "功率(dBm)", "515.472", "-20"]
            ),
            v4_server_text_rec=lambda _: self.fail(
                "three-family consensus must not invoke the v4 tie-breaker"
            ),
        )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ), patch.object(
            ocr_backend,
            "_wiener_motion_deblur_view",
            return_value=np.full((12, 24), 220, dtype=np.uint8),
        ):
            scores = ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 60, 120],
                [0, 40, 80],
                engine,
            )

        self.assertEqual(
            grid,
            [["频率(mHz)", "功率(dBm)"], ["515.472", "-20"]],
        )
        self.assertEqual(len(scores), 8)
        self.assertTrue(all(value == 0.77 for row in confidence for value in row))

    def test_motion_cell_consensus_uses_repeated_local_vocabulary_for_one_blurred_cell(self):
        ocr_backend._load_runtime()
        image = np.full((160, 80, 3), 240, dtype=np.uint8)
        grid = [["设备名称"], ["直流稳压源"], ["直流稳压源"], ["直淋稳压源"]]
        confidence = [[0.96], [0.96], [0.96], [0.60]]

        def recognizer(texts):
            return lambda request: SimpleNamespace(
                txts=texts[: len(request.img)],
                scores=[0.92] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=recognizer(
                ["设备名称", "直流稳压源", "直流稳压源", "直流稳压洲"]
            ),
            text_rec=recognizer(
                ["设备名称", "直流稳压源", "直流稳压源", "直流稳压源"]
            ),
            server_text_rec=recognizer(
                ["设备名称", "直流稳压源", "直流稳压源", "直流稳压测"]
            ),
            v4_server_text_rec=recognizer(["直流稳压洲"]),
        )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ), patch.object(
            ocr_backend,
            "_wiener_motion_deblur_view",
            return_value=np.full((12, 24), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 80],
                [0, 40, 80, 120, 160],
                engine,
            )

        self.assertEqual(grid[3][0], "直流稳压源")
        self.assertEqual(confidence[3][0], 0.86)

    def test_standard_cell_consensus_keeps_page_text_when_medium_confirms_it(self):
        ocr_backend._load_runtime()
        image = np.full((40, 40, 3), 240, dtype=np.uint8)
        grid = [["4"]]
        confidence = [[0.99]]

        def recognizer(text, score):
            return lambda request: SimpleNamespace(
                txts=[text] * len(request.img),
                scores=[score] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=lambda _: self.fail(
                "a strong page/mobile result confirmed by medium must stop early"
            ),
            text_rec=recognizer("4", 0.97),
            server_text_rec=lambda _: self.fail(
                "server recognition is only for unresolved standard cells"
            ),
            v4_server_text_rec=lambda _: self.fail(
                "v4 recognition is only for unresolved standard cells"
            ),
        )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 16, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 40],
                [0, 40],
                engine,
                view_mode="standard",
            )

        self.assertEqual(grid, [["4"]])
        self.assertGreaterEqual(confidence[0][0], 0.90)

    def test_standard_cell_consensus_removes_confirmed_right_neighbor_prefix(self):
        ocr_backend._load_runtime()
        image = np.full((40, 80, 3), 240, dtype=np.uint8)
        grid = [["A014数", "数字示波器"]]
        confidence = [[0.99, 0.99]]

        def recognizer(texts):
            return lambda request: SimpleNamespace(
                txts=texts[: len(request.img)],
                scores=[0.97] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=lambda _: self.fail(
                "non-empty page cells must not repeat the mobile family"
            ),
            text_rec=recognizer(["A014", "数字示波器"]),
            server_text_rec=recognizer(["A014"]),
            v4_server_text_rec=lambda _: self.fail("retired v4 must never be invoked"),
        )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 40, 80],
                [0, 40],
                engine,
                view_mode="standard",
            )

        self.assertEqual(grid, [["A014", "数字示波器"]])

    def test_standard_cell_consensus_rejects_right_neighbor_prefix_expansion(self):
        ocr_backend._load_runtime()
        image = np.full((40, 80, 3), 240, dtype=np.uint8)
        grid = [["A014", "数字示波器"]]
        confidence = [[0.99, 0.99]]

        def recognizer(texts):
            return lambda request: SimpleNamespace(
                txts=texts[: len(request.img)],
                scores=[0.97] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=lambda _: self.fail(
                "non-empty page cells must not repeat the mobile family"
            ),
            text_rec=recognizer(["A014数", "数字示波器"]),
            server_text_rec=recognizer(["A014数"]),
            v4_server_text_rec=lambda _: self.fail("retired v4 must never be invoked"),
        )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 40, 80],
                [0, 40],
                engine,
                view_mode="standard",
            )

        self.assertEqual(grid, [["A014", "数字示波器"]])

    def test_standard_cell_consensus_sends_only_conflicts_to_server(self):
        ocr_backend._load_runtime()
        image = np.full((40, 80, 3), 240, dtype=np.uint8)
        grid = [["515.221", "-20"]]
        confidence = [[0.99, 0.99]]
        calls = {"medium": 0, "server": 0}

        def medium_recognizer(request):
            calls["medium"] += len(request.img)
            return SimpleNamespace(
                txts=["515.221", "20"],
                scores=[0.98, 0.97],
            )

        def server_recognizer(request):
            calls["server"] += len(request.img)
            return SimpleNamespace(txts=["-20"], scores=[0.96])

        engine = SimpleNamespace(
            fast_text_rec=lambda _: self.fail(
                "non-empty page cells already provide the mobile-family evidence"
            ),
            text_rec=medium_recognizer,
            server_text_rec=server_recognizer,
            v4_server_text_rec=lambda _: self.fail(
                "the page/server agreement must settle the only conflict"
            ),
        )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 40, 80],
                [0, 40],
                engine,
                view_mode="standard",
            )

        self.assertEqual(grid, [["515.221", "-20"]])
        self.assertEqual(calls, {"medium": 2, "server": 1})

    def test_standard_cell_consensus_marks_medium_only_blank_recovery_for_review(self):
        ocr_backend._load_runtime()
        image = np.full((40, 40, 3), 240, dtype=np.uint8)
        grid = [[""]]
        confidence = [[-1.0]]
        calls = {"mobile": 0, "medium": 0, "server": 0}

        def recognizer(family):
            def run(request):
                calls[family] += len(request.img)
                return SimpleNamespace(txts=["A001"], scores=[0.96])
            return run

        engine = SimpleNamespace(
            fast_text_rec=recognizer("mobile"),
            text_rec=recognizer("medium"),
            server_text_rec=recognizer("server"),
            v4_server_text_rec=lambda _: self.fail(
                "retired v4 must never be invoked"
            ),
        )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 40],
                [0, 40],
                engine,
                view_mode="standard",
            )

        self.assertEqual(grid, [["A001"]])
        self.assertEqual(confidence, [[0.77]])
        self.assertEqual(calls, {"mobile": 0, "medium": 1, "server": 1})

    def test_standard_cell_consensus_does_not_invent_header_suffix(self):
        ocr_backend._load_runtime()
        image = np.full((40, 80, 3), 240, dtype=np.uint8)
        grid = [["任务编号", "不良数"]]
        confidence = [[0.99, 0.72]]

        engine = SimpleNamespace(
            fast_text_rec=lambda _: self.fail(
                "non-empty page cells must not repeat the mobile family"
            ),
            text_rec=lambda request: SimpleNamespace(
                txts=["任务编号", "不良数"],
                scores=[0.97, 0.97],
            ),
            server_text_rec=lambda request: SimpleNamespace(
                txts=["不良数"] * len(request.img),
                scores=[0.97] * len(request.img),
            ),
            v4_server_text_rec=lambda _: self.fail(
                "three-family header agreement must not invoke v4"
            ),
        )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 40, 80],
                [0, 40],
                engine,
                view_mode="standard",
            )

        self.assertEqual(grid[0][1], "不良数")
        self.assertEqual(confidence[0][1], 0.77)

    def test_standard_cell_consensus_accepts_exact_category_and_unit_agreement(self):
        ocr_backend._load_runtime()
        image = np.full((80, 160, 3), 240, dtype=np.uint8)
        grid = [["类别", "单位"], ["", ""]]
        confidence = [[0.99, 0.99], [-1.0, -1.0]]

        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["类别1", "kHz"][: len(request.img)],
                scores=[0.985, 0.841][: len(request.img)],
            ),
            server_text_rec=lambda request: SimpleNamespace(
                txts=["类别1", "kHz"][: len(request.img)],
                scores=[0.839, 0.644][: len(request.img)],
            ),
            v4_server_text_rec=lambda _: self.fail("retired v4 must never run"),
        )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 80, 160],
                [0, 40, 80],
                engine,
                view_mode="standard",
                first_row=1,
            )

        self.assertEqual(grid[1], ["类别1", "kHz"])
        self.assertEqual(confidence[1], [0.77, 0.77])

    def test_standard_semantic_code_prefers_medium_visible_slash_for_review(self):
        ocr_backend._load_runtime()
        image = np.full((80, 80, 3), 240, dtype=np.uint8)
        grid = [["备注"], ["NA"]]
        confidence = [[0.99], [0.98]]

        def medium(request):
            return SimpleNamespace(txts=["N/A"] * len(request.img), scores=[0.95] * len(request.img))

        def server(request):
            return SimpleNamespace(txts=["NA"] * len(request.img), scores=[0.97] * len(request.img))

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 80],
                [0, 40, 80],
                SimpleNamespace(text_rec=medium, server_text_rec=server),
                view_mode="standard",
                first_row=1,
            )

        self.assertEqual(grid[1][0], "N/A")
        self.assertEqual(confidence[1][0], 0.77)

    def test_standard_blank_cell_retries_unscaled_crop_after_scaled_disagreement(self):
        ocr_backend._load_runtime()
        image = np.full((80, 80, 3), 240, dtype=np.uint8)
        grid = [["类别"], [""]]
        confidence = [[0.99], [-1.0]]
        calls = {"medium": 0, "server": 0}

        def medium(request):
            calls["medium"] += 1
            return SimpleNamespace(txts=["类别1"], scores=[0.99])

        def server(request):
            calls["server"] += 1
            if calls["server"] == 1:
                return SimpleNamespace(txts=["类刷1"], scores=[0.85])
            return SimpleNamespace(txts=["类别1"], scores=[0.89])

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 80],
                [0, 40, 80],
                SimpleNamespace(text_rec=medium, server_text_rec=server),
                view_mode="standard",
                first_row=1,
            )

        self.assertEqual(grid[1][0], "类别1")
        self.assertEqual(confidence[1][0], 0.77)
        self.assertEqual(calls, {"medium": 2, "server": 2})

    def test_raw_blank_retry_skips_alternate_when_primary_cannot_be_accepted(self):
        ocr_backend._load_runtime()
        image = np.full((80, 80, 3), 240, dtype=np.uint8)
        grid = [["类别"], [""]]
        confidence = [[0.99], [-1.0]]
        calls = {"medium": 0, "server": 0}

        def medium(request):
            calls["medium"] += 1
            text = "类刷1" if calls["medium"] == 1 else "1"
            return SimpleNamespace(txts=[text], scores=[0.99])

        def server(_request):
            calls["server"] += 1
            if calls["server"] > 1:
                self.fail("主视图已不满足接受条件时不得调用原始裁格的第二视图")
            return SimpleNamespace(txts=["类别1"], scores=[0.90])

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 80],
                [0, 40, 80],
                SimpleNamespace(text_rec=medium, server_text_rec=server),
                view_mode="standard",
                first_row=1,
            )

        self.assertEqual(grid[1][0], "")
        self.assertLessEqual(confidence[1][0], 0.0)
        self.assertEqual(calls, {"medium": 2, "server": 1})

    def test_raw_blank_retry_skips_degenerate_tight_crop(self):
        ocr_backend._load_runtime()
        image = np.full((80, 80, 3), 240, dtype=np.uint8)
        grid = [["类别"], [""]]
        confidence = [[0.99], [-1.0]]
        calls = {"medium": 0, "server": 0}

        def medium(request):
            calls["medium"] += 1
            return SimpleNamespace(txts=["类刷1"] * len(request.img), scores=[0.99] * len(request.img))

        def server(request):
            calls["server"] += 1
            return SimpleNamespace(txts=["类别1"] * len(request.img), scores=[0.90] * len(request.img))

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((4, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 80],
                [0, 40, 80],
                SimpleNamespace(text_rec=medium, server_text_rec=server),
                view_mode="standard",
                first_row=1,
            )

        self.assertEqual(grid[1][0], "")
        self.assertLessEqual(confidence[1][0], 0.0)
        self.assertEqual(calls, {"medium": 1, "server": 1})

    def test_raw_blank_retry_batches_only_base_ratio_primary_views(self):
        ocr_backend._load_runtime()
        image = np.full((80, 240, 3), 240, dtype=np.uint8)
        grid = [["", "", ""], ["", "", ""]]
        confidence = [[-1.0, -1.0, -1.0], [-1.0, -1.0, -1.0]]
        calls = {"medium": [], "server": []}

        def medium(request):
            calls["medium"].append(len(request.img))
            text = "类刷" if len(calls["medium"]) == 1 else "1"
            return SimpleNamespace(
                txts=[text] * len(request.img),
                scores=[0.99] * len(request.img),
            )

        def server(request):
            calls["server"].append(len(request.img))
            return SimpleNamespace(
                txts=["类别"] * len(request.img),
                scores=[0.90] * len(request.img),
            )

        engine = SimpleNamespace(
            text_rec=medium,
            server_text_rec=server,
            _verification_recognizer=SimpleNamespace(
                rec_image_shape=(3, 48, 320)
            ),
        )
        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 80, 160, 240],
                [0, 40, 80],
                engine,
                view_mode="standard",
            )

        self.assertEqual(calls["medium"], [6, 6])
        self.assertEqual(calls["server"], [6])
        self.assertEqual(grid, [["", "", ""], ["", "", ""]])

    def test_raw_retry_refreshes_semantics_after_recovering_category_header(self):
        ocr_backend._load_runtime()
        image = np.full((80, 80, 3), 240, dtype=np.uint8)
        grid = [[""], [""]]
        confidence = [[-1.0], [-1.0]]
        calls = {"medium": 0, "server": 0}

        def medium(request):
            calls["medium"] += 1
            values = (
                ["类别", "类别1"]
                if len(request.img) == 2
                else ["类别" if calls["medium"] == 2 else "类别1"]
            )
            return SimpleNamespace(txts=values, scores=[0.99] * len(values))

        def server(request):
            calls["server"] += 1
            if calls["server"] == 1:
                return SimpleNamespace(txts=["类别", "类刷1"], scores=[0.74, 0.85])
            text = "类别" if calls["server"] == 2 else "类别1"
            score = 0.75 if calls["server"] == 2 else 0.89
            return SimpleNamespace(txts=[text], scores=[score])

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 80],
                [0, 40, 80],
                SimpleNamespace(text_rec=medium, server_text_rec=server),
                view_mode="standard",
            )

        self.assertEqual(grid, [["类别"], ["类别1"]])
        self.assertEqual(confidence[1][0], 0.77)
        self.assertEqual(calls, {"medium": 3, "server": 3})

    def test_standard_cell_consensus_never_shortens_complete_group_header(self):
        ocr_backend._load_runtime()
        image = np.full((40, 80, 3), 240, dtype=np.uint8)
        grid = [["基础信息"]]
        confidence = [[0.86]]

        def recognizer(request):
            return SimpleNamespace(txts=["基"] * len(request.img), scores=[0.99] * len(request.img))

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 80],
                [0, 40],
                SimpleNamespace(text_rec=recognizer, server_text_rec=recognizer),
                view_mode="standard",
            )

        self.assertEqual(grid, [["基础信息"]])
        self.assertEqual(confidence, [[0.86]])

    def test_standard_cell_consensus_uses_server_only_for_missing_decimal_and_sign(self):
        ocr_backend._load_runtime()
        image = np.full((40, 80, 3), 240, dtype=np.uint8)
        grid = [["515221", "20"]]
        confidence = [[0.96, 0.95]]
        calls = {"medium": 0, "server": 0}

        def medium_recognizer(request):
            calls["medium"] += len(request.img)
            return SimpleNamespace(
                txts=["515.221", "-20"],
                scores=[0.98, 0.98],
            )

        def server_recognizer(request):
            calls["server"] += len(request.img)
            return SimpleNamespace(
                txts=["515.221", "-20"],
                scores=[0.97, 0.97],
            )

        engine = SimpleNamespace(
            fast_text_rec=lambda _: self.fail(
                "non-empty page cells must not repeat the mobile family"
            ),
            text_rec=medium_recognizer,
            server_text_rec=server_recognizer,
            v4_server_text_rec=lambda _: self.fail(
                "medium/server agreement must settle punctuation conflicts"
            ),
        )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 40, 80],
                [0, 40],
                engine,
                view_mode="standard",
            )

        self.assertEqual(grid, [["515.221", "-20"]])
        self.assertEqual(calls, {"medium": 2, "server": 2})

    def test_shadowed_responsible_column_stops_after_medium_server_agreement(self):
        ocr_backend._load_runtime()
        image = np.full((180, 60, 3), 230, dtype=np.uint8)
        grid = [["负责人"], *[[""] for _ in range(8)]]
        confidence = [[0.99], *[[-1.0] for _ in range(8)]]

        def agreeing_recognizer(request):
            return SimpleNamespace(
                txts=["陈晨"] * len(request.img),
                scores=[0.96] * len(request.img),
            )

        engine = SimpleNamespace(
            text_rec=agreeing_recognizer,
            server_text_rec=agreeing_recognizer,
            v4_server_text_rec=lambda _: self.fail(
                "medium/server agreement must not repeat the column with v4"
            ),
        )
        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 20, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._recover_shadowed_responsible_column(
                image,
                grid,
                confidence,
                [0, 60],
                list(range(0, 181, 20)),
                engine,
            )

        self.assertEqual([row[0] for row in grid[1:]], ["陈晨"] * 8)

    def test_shadowed_responsible_column_never_invokes_retired_v4(self):
        ocr_backend._load_runtime()
        image = np.full((180, 60, 3), 230, dtype=np.uint8)
        names = ["陈晨", "王强", "赵敏", "李娜", "陈晨", "王强", "赵敏", "李娜"]
        grid = [["负责人"], *[[""] for _ in names]]
        confidence = [[0.99], *[[-1.0] for _ in names]]
        calls = {"v4": 0}

        def recognizer(values):
            return lambda request: SimpleNamespace(
                txts=[values[index // 3] for index in range(len(request.img))],
                scores=[0.96] * len(request.img),
            )

        conflicting_names = names[:]
        conflicting_names[2] = "赵具"

        def v4_recognizer(request):
            calls["v4"] += len(request.img)
            return SimpleNamespace(
                txts=["赵敏"] * len(request.img),
                scores=[0.95] * len(request.img),
            )

        engine = SimpleNamespace(
            text_rec=recognizer(names),
            server_text_rec=recognizer(conflicting_names),
            v4_server_text_rec=v4_recognizer,
        )
        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 20, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._recover_shadowed_responsible_column(
                image,
                grid,
                confidence,
                [0, 60],
                list(range(0, 181, 20)),
                engine,
            )

        self.assertEqual([row[0] for row in grid[1:]], names)
        self.assertEqual(calls["v4"], 0)

    def test_standard_cell_consensus_uses_repeated_vocabulary_for_a_high_score_variant(self):
        ocr_backend._load_runtime()
        image = np.full((160, 80, 3), 240, dtype=np.uint8)
        grid = [["负责人"], ["李娜"], ["王强"], ["李娜"], ["李"]]
        confidence = [[0.99], [0.96], [0.96], [0.96], [0.97]]

        def recognizer(texts, score):
            return lambda request: SimpleNamespace(
                txts=texts[: len(request.img)],
                scores=[score] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=recognizer(["李娜", "王强", "李娜", "李"], 0.92),
            text_rec=recognizer(["李娜", "王强", "李娜", "李娜"], 0.94),
            server_text_rec=recognizer(["李那", "王强", "李那", "李"], 0.91),
            v4_server_text_rec=recognizer(["李"], 0.92),
        )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 80],
                [0, 32, 64, 96, 128, 160],
                engine,
                view_mode="standard",
                first_row=1,
            )

        self.assertEqual(grid[4][0], "李娜")
        self.assertEqual(confidence[4][0], 0.77)

    def test_standard_cell_consensus_never_shortens_strong_page_text(self):
        ocr_backend._load_runtime()
        image = np.full((40, 80, 3), 240, dtype=np.uint8)
        grid = [["2026-08-02"]]
        confidence = [[0.99]]

        def recognizer(request):
            return SimpleNamespace(
                txts=["026-08-02"] * len(request.img),
                scores=[0.97] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=recognizer,
            text_rec=recognizer,
            server_text_rec=recognizer,
            v4_server_text_rec=lambda _: self.fail("three families already agree"),
        )

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 24, 3), 220, dtype=np.uint8),
        ):
            ocr_backend._verify_motion_blurred_ruled_cells(
                image,
                grid,
                confidence,
                [0, 80],
                [0, 40],
                engine,
                view_mode="standard",
            )

        self.assertEqual(grid, [["2026-08-02"]])
        self.assertEqual(confidence, [[0.99]])

    def test_motion_refinement_fills_only_initial_blanks_as_review(self):
        ocr_backend._load_runtime()
        grid = [["王", "国像预处理"], ["", ""]]
        confidence = [[0.97, 0.96], [-1.0, 0.0]]
        crops = [np.full((12, 24, 3), 255, dtype=np.uint8) for _ in range(4)]
        locations = [(0, 0, 0, 1), (0, 1, 0, 1), (1, 0, 0, 1), (1, 1, 0, 1)]

        def recognizer(texts):
            return lambda request: SimpleNamespace(
                txts=texts[: len(request.img)],
                scores=[0.97] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=recognizer(["低", ""]),
            text_rec=recognizer(["低", ""]),
            server_text_rec=recognizer(["低", "噪音"]),
            v4_server_text_rec=recognizer(["低", "其他"]),
        )

        ocr_backend._refine_motion_blurred_ruled_lines(
            grid,
            confidence,
            crops,
            locations,
            engine,
        )

        self.assertEqual(grid, [["王", "国像预处理"], ["低", ""]])
        self.assertEqual(confidence[0][0], 0.97)
        self.assertEqual(confidence[0][1], 0.96)
        self.assertEqual(confidence[1][0], 0.77)
        self.assertEqual(confidence[1][1], 0.0)

    def test_motion_refinement_does_not_run_for_one_ordinary_blank_cell(self):
        grid = [["编号", "名称"], ["1", ""]]
        confidence = [[0.99, 0.99], [0.99, 0.0]]
        engine = SimpleNamespace(
            fast_text_rec=lambda _: self.fail("one pending cell must not trigger motion pass"),
            text_rec=lambda _: self.fail("one pending cell must not trigger motion pass"),
            server_text_rec=lambda _: self.fail("one pending cell must not trigger motion pass"),
            v4_server_text_rec=lambda _: self.fail("one pending cell must not trigger motion pass"),
        )

        scores = ocr_backend._refine_motion_blurred_ruled_lines(
            grid,
            confidence,
            [np.full((12, 24, 3), 255, dtype=np.uint8) for _ in range(4)],
            [(0, 0, 0, 1), (0, 1, 0, 1), (1, 0, 0, 1), (1, 1, 0, 1)],
            engine,
        )

        self.assertEqual(scores, [])
        self.assertEqual(grid, [["编号", "名称"], ["1", ""]])

    def test_motion_refinement_never_replaces_a_horizontal_dash(self):
        grid = [["—", ""], ["", "正常"]]
        confidence = [[0.86, -1.0], [-1.0, 0.99]]
        crops = [np.full((12, 24, 3), 255, dtype=np.uint8) for _ in range(4)]
        locations = [(0, 0, 0, 1), (0, 1, 0, 1), (1, 0, 0, 1), (1, 1, 0, 1)]

        def recognizer(request):
            return SimpleNamespace(
                txts=["文字"] * len(request.img),
                scores=[0.99] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=recognizer,
            text_rec=recognizer,
            server_text_rec=recognizer,
            v4_server_text_rec=recognizer,
        )

        ocr_backend._refine_motion_blurred_ruled_lines(
            grid,
            confidence,
            crops,
            locations,
            engine,
        )

        self.assertEqual(grid[0][0], "—")

    def test_motion_refinement_runs_for_a_near_text_variant_without_blanks(self):
        grid = [["负责人", "任务名称"], ["王", "国像预处理"], ["王强", "图像预处理"]]
        confidence = [[0.99, 0.99], [0.97, 0.96], [0.99, 0.99]]
        crops = [np.full((12, 24, 3), 255, dtype=np.uint8) for _ in range(6)]
        locations = [(row, column, 0, 1) for row in range(3) for column in range(2)]

        expected = [
            "负责人",
            "任务名称",
            "王强",
            "图像预处理",
            "王强",
            "图像预处理",
        ]

        def recognizer(request):
            return SimpleNamespace(
                txts=expected[: len(request.img)],
                scores=[0.98] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=recognizer,
            text_rec=recognizer,
            server_text_rec=recognizer,
            v4_server_text_rec=recognizer,
        )

        ocr_backend._refine_motion_blurred_ruled_lines(
            grid,
            confidence,
            crops,
            locations,
            engine,
        )

        self.assertEqual(grid[1], ["王", "国像预处理"])

    def test_motion_ruled_grid_keeps_full_bounded_crop_for_initial_blanks(self):
        ocr_backend._load_runtime()
        image = np.full((40, 40, 3), 255, dtype=np.uint8)
        image[3:17, 23:37] = 71
        image[23:37, 3:17] = 93
        grid = [["A", ""], ["", "B"]]
        confidence = [[0.99, -1.0], [-1.0, 0.99]]
        compact = np.full((3, 4, 3), 17, dtype=np.uint8)

        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((5, 6, 3), 29, dtype=np.uint8),
        ) as tight, patch.object(
            ocr_backend,
            "_split_text_line_crops",
            return_value=[compact],
        ), patch.object(
            ocr_backend,
            "_refine_motion_blurred_ruled_lines",
            return_value=[0.77],
        ) as refine:
            scores = ocr_backend._refine_motion_blurred_ruled_grid(
                image,
                grid,
                confidence,
                [0, 20, 40],
                [0, 20, 40],
                SimpleNamespace(),
            )

        self.assertEqual(scores, [0.77])
        self.assertEqual(tight.call_count, 2)
        source_crops = refine.call_args.args[2]
        locations = refine.call_args.args[3]
        self.assertEqual(locations, [(0, 0, 0, 1), (0, 1, 0, 1), (1, 0, 0, 1), (1, 1, 0, 1)])
        self.assertEqual(source_crops[0].shape[:2], (3, 4))
        self.assertEqual(source_crops[1].shape[:2], (16, 16))
        self.assertEqual(source_crops[2].shape[:2], (16, 16))
        self.assertEqual(source_crops[3].shape[:2], (3, 4))
        self.assertEqual(int(source_crops[1][2, 2, 0]), 71)
        self.assertEqual(int(source_crops[2][2, 2, 0]), 93)

    def test_motion_refinement_skips_locked_merged_subordinate_cells(self):
        ocr_backend._load_runtime()
        image = np.full((40, 60, 3), 80, dtype=np.uint8)
        grid = [["标题", "", ""], ["A", "B", "C"]]
        confidence = [[0.99, 0.0, 0.0], [0.99, 0.99, 0.99]]

        with patch.object(
            ocr_backend,
            "_split_text_line_crops",
            side_effect=lambda crop: [crop],
        ), patch.object(
            ocr_backend,
            "_refine_motion_blurred_ruled_lines",
            return_value=[],
        ) as refine:
            ocr_backend._refine_motion_blurred_ruled_grid(
                image,
                grid,
                confidence,
                [0, 20, 40, 60],
                [0, 20, 40],
                SimpleNamespace(),
                excluded_cells={(0, 1), (0, 2)},
            )

        locations = refine.call_args.args[3]
        self.assertNotIn((0, 1, 0, 1), locations)
        self.assertNotIn((0, 2, 0, 1), locations)
        self.assertIn((0, 0, 0, 1), locations)
        self.assertIn((1, 1, 0, 1), locations)

    def test_maximum_missing_cell_recovery_requires_server_support(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "B", (48, 54), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 0), 2)
        grid = [
            ["A", "B", "C"],
            ["1", "2", "3"],
            ["4", "", "6"],
            ["7", "8", "9"],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[2][1] = -1.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0],
            "grouped_rows": [[{}], [{}], [{}], [{}]],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(txts=["B", "B"], scores=[0.96, 0.94]),
            server_text_rec=lambda request: SimpleNamespace(txts=["8", "8"], scores=[0.98, 0.97]),
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image,
            grid,
            confidence,
            geometry,
            engine,
            require_cross_model=True,
        )

        self.assertEqual(recovered[2][1], "")
        self.assertEqual(recovered_confidence[2][1], -2.0)

    def test_maximum_missing_cell_withholds_when_two_model_families_are_missing(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "CC", (45, 54), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 0, 0), 1)
        grid = [
            ["编号", "责任人", "状态"],
            ["1", "张伟", "正常"],
            ["2", "", "复核"],
            ["3", "李娜", "正常"],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[2][1] = -1.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0],
            "grouped_rows": [[{}], [{}], [{}], [{}]],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["陈晨", "陈晨"], scores=[0.864, 0.769]
            ),
            server_text_rec=lambda request: SimpleNamespace(
                txts=["陈晨", "陈晨"], scores=[0.994, 0.994]
            ),
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image,
            grid,
            confidence,
            geometry,
            engine,
            require_cross_model=True,
        )

        self.assertEqual(recovered[2][1], "")
        self.assertEqual(recovered_confidence[2][1], -2.0)

    def test_maximum_missing_header_withholds_without_four_model_evidence(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "QTY", (44, 14), cv2.FONT_HERSHEY_SIMPLEX, 0.35, (0, 0, 0), 1)
        grid = [
            ["编号", "", "状态"],
            ["1", "2", "正常"],
            ["2", "1", "复核"],
            ["3", "0", "正常"],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[0][1] = -1.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0],
            "grouped_rows": [[{}], [{}], [{}], [{}]],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["不良数压", "不良数压"], scores=[0.955, 0.895]
            ),
            server_text_rec=lambda request: SimpleNamespace(
                txts=["不良数", "不良数量"], scores=[0.997, 0.782]
            ),
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image,
            grid,
            confidence,
            geometry,
            engine,
            require_cross_model=True,
        )

        self.assertEqual(recovered[0][1], "")
        self.assertEqual(recovered_confidence[0][1], -2.0)

    def test_maximum_missing_short_header_withholds_without_four_model_evidence(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "TYPE", (43, 14), cv2.FONT_HERSHEY_SIMPLEX, 0.35, (0, 0, 0), 1)
        grid = [
            ["编号", "", "状态"],
            ["1", "SG-2200", "正常"],
            ["2", "DP832A", "复核"],
            ["3", "SA-5000", "正常"],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[0][1] = -1.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0],
            "grouped_rows": [[{}], [{}], [{}], [{}]],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["型号", "型号"], scores=[0.83, 0.98]
            ),
            server_text_rec=lambda request: SimpleNamespace(
                txts=["型号", "登号"], scores=[0.53, 0.91]
            ),
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image,
            grid,
            confidence,
            geometry,
            engine,
            require_cross_model=True,
        )

        self.assertEqual(recovered[0][1], "")
        self.assertEqual(recovered_confidence[0][1], -2.0)

    def test_maximum_missing_status_header_does_not_guess_from_body_semantics(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "STATUS", (42, 14), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (0, 0, 0), 1)
        grid = [
            ["编号", "", "备注"],
            ["1", "正常", "完成"],
            ["2", "复核", "检查"],
            ["3", "待机", "—"],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[0][1] = -1.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0],
            "grouped_rows": [[{}], [{}], [{}], [{}]],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["状绣", "状绣"], scores=[0.49, 0.76]
            ),
            server_text_rec=lambda request: SimpleNamespace(
                txts=["状态", "状态"], scores=[0.93, 0.86]
            ),
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image,
            grid,
            confidence,
            geometry,
            engine,
            require_cross_model=True,
        )

        self.assertEqual(recovered[0][1], "")
        self.assertEqual(recovered_confidence[0][1], -2.0)

    def test_trusted_ruled_geometry_rechecks_a_sparse_header_and_repeated_column_misses(self):
        ocr_backend._load_runtime()
        image = np.full((120, 160, 3), 255, dtype=np.uint8)
        cv2.putText(image, "TXT", (42, 17), cv2.FONT_HERSHEY_SIMPLEX, 0.35, (0, 0, 0), 1)
        cv2.putText(image, "TXT", (42, 57), cv2.FONT_HERSHEY_SIMPLEX, 0.35, (0, 0, 0), 1)
        cv2.putText(image, "TXT", (42, 77), cv2.FONT_HERSHEY_SIMPLEX, 0.35, (0, 0, 0), 1)
        grid = [
            ["编号", "", "状态", "备注"],
            ["1", "设备A", "正常", "完成"],
            ["2", "", "复核", "完成"],
            ["3", "", "正常", "完成"],
            ["4", "设备D", "正常", "完成"],
            ["5", "设备E", "正常", "完成"],
        ]
        confidence = [[0.95] * 4 for _ in grid]
        confidence[0][1] = -1.0
        confidence[2][1] = -1.0
        confidence[3][1] = -1.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0, 140.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0, 90.0, 110.0],
            "grouped_rows": [[] for _ in grid],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["设备名称", "设备B", "设备C"] * 2,
                scores=[0.99] * 6,
            ),
            server_text_rec=lambda request: SimpleNamespace(
                txts=["设备名称", "设备B", "设备C"] * 2,
                scores=[0.98] * 6,
            ),
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image,
            grid,
            confidence,
            geometry,
            engine,
            require_cross_model=True,
        )

        for row in (0, 2, 3):
            self.assertEqual(recovered[row][1], "")
            self.assertEqual(recovered_confidence[row][1], -2.0)

    def test_maximum_missing_text_withholds_medium_only_clear_views(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "TEXT", (43, 54), cv2.FONT_HERSHEY_SIMPLEX, 0.35, (0, 0, 0), 1)
        grid = [
            ["编号", "备注", "状态"],
            ["1", "校准完成", "正常"],
            ["2", "", "复核"],
            ["3", "接口检查", "正常"],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[2][1] = -1.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0],
            "grouped_rows": [[{}] for _ in grid],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["校准完成", "校准完成"], scores=[0.99, 0.98]
            ),
            server_text_rec=lambda request: SimpleNamespace(
                txts=["大城完成", "大盛宗球"], scores=[0.42, 0.28]
            ),
            v4_server_text_rec=lambda request: SimpleNamespace(
                txts=["交减亮成", "交减亮惑"], scores=[0.41, 0.30]
            ),
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image,
            grid,
            confidence,
            geometry,
            engine,
            require_cross_model=True,
        )

        self.assertEqual(recovered[2][1], "")
        self.assertEqual(recovered_confidence[2][1], -2.0)

    def test_maximum_missing_text_withholds_near_only_cross_model_support(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "TEXT", (43, 54), cv2.FONT_HERSHEY_SIMPLEX, 0.35, (0, 0, 0), 1)
        grid = [
            ["编号", "设备名称", "状态"],
            ["1", "信号发生器", "正常"],
            ["2", "", "复核"],
            ["3", "直流稳压源", "正常"],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[2][1] = -1.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0],
            "grouped_rows": [[{}] for _ in grid],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["数字示波器", "数字示波器"], scores=[0.95, 0.94]
            ),
            server_text_rec=lambda request: SimpleNamespace(
                txts=["数字元浓", "数字元"], scores=[0.65, 0.86]
            ),
            v4_server_text_rec=lambda request: SimpleNamespace(
                txts=["数字示波源", "数字示波源"], scores=[0.79, 0.78]
            ),
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image,
            grid,
            confidence,
            geometry,
            engine,
            require_cross_model=True,
        )

        self.assertEqual(recovered[2][1], "")
        self.assertEqual(recovered_confidence[2][1], -2.0)

    def test_motion_sign_consensus_requires_two_deblur_views_and_matching_magnitude(self):
        selected = ocr_backend._select_signed_motion_consensus(
            [("20", 0.999), ("20", 0.998)],
            [("-20", 0.986), ("-20", 0.999)],
        )

        self.assertEqual(selected, ("-20", 0.986))
        self.assertIsNone(
            ocr_backend._select_signed_motion_consensus(
                [("20", 0.999)],
                [("-20", 0.99), ("20", 0.99)],
            )
        )
        self.assertIsNone(
            ocr_backend._select_signed_motion_consensus(
                [("30", 0.999)],
                [("-20", 0.99), ("-20", 0.98)],
            )
        )

    def test_ruled_numeric_sign_recovery_flags_but_does_not_rewrite_existing_value(self):
        ocr_backend._load_runtime()
        image = np.full((120, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "-30", (65, 52), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 0, 0), 1)
        grid = [
            ["编号", "功率(dBm)"],
            ["1", "30"],
            ["2", "-20"],
            ["3", "N/A"],
        ]
        confidence = [[0.95, 0.95] for _ in grid]
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["30", "-30", "-30", "30"],
                scores=[0.99, 0.91, 0.90, 0.98],
            )
        )

        scores = ocr_backend._restore_ruled_numeric_signs(
            image,
            grid,
            confidence,
            [0, 60, 120],
            [0, 30, 60, 90, 120],
            engine,
        )

        self.assertEqual(grid[1][1], "30")
        self.assertEqual(confidence[1][1], 0.77)
        self.assertEqual(scores, [0.77])

    def test_ruled_numeric_sign_recovery_does_not_use_a_single_signed_view(self):
        ocr_backend._load_runtime()
        image = np.full((120, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "30", (65, 52), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 0, 0), 1)
        grid = [
            ["编号", "功率(dBm)"],
            ["1", "30"],
            ["2", "-20"],
            ["3", "N/A"],
        ]
        confidence = [[0.95, 0.95] for _ in grid]
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["30", "-30", "30", "30"],
                scores=[0.99, 0.91, 0.98, 0.97],
            )
        )

        scores = ocr_backend._restore_ruled_numeric_signs(
            image,
            grid,
            confidence,
            [0, 60, 120],
            [0, 30, 60, 90, 120],
            engine,
        )

        self.assertEqual(grid[1][1], "30")
        self.assertEqual(scores, [])

    def test_ruled_numeric_sign_recovery_never_fills_a_blank_cell(self):
        image = np.full((120, 120, 3), 255, dtype=np.uint8)
        grid = [
            ["编号", "功率(dBm)"],
            ["1", ""],
            ["2", "-20"],
            ["3", "N/A"],
        ]
        confidence = [[0.95, 0.95] for _ in grid]
        confidence[1][1] = -1.0
        engine = SimpleNamespace(
            text_rec=lambda _: self.fail("blank signs require strict four-model recovery")
        )

        scores = ocr_backend._restore_ruled_numeric_signs(
            image,
            grid,
            confidence,
            [0, 60, 120],
            [0, 30, 60, 90, 120],
            engine,
        )

        self.assertEqual(grid[1][1], "")
        self.assertEqual(confidence[1][1], -1.0)
        self.assertEqual(scores, [])

    def test_unverified_structure_marks_every_present_cell_for_review(self):
        confidence = [[0.99, 0.0], [-2.0, 0.85]]

        ocr_backend._mark_unverified_structure_for_review(confidence)

        self.assertEqual(confidence, [[0.77, 0.77], [-2.0, 0.77]])

    def test_verified_structure_edge_warning_marks_whole_table_for_review(self):
        confidence = [[0.99, 0.77], [0.0, 0.88]]

        result = ocr_backend._apply_structure_review_policy(
            [["编号", "名称"], ["1", "设备"]],
            confidence,
            structure_verified=True,
            structural_warnings=["照片边缘证据尚未完成闭环。"],
        )

        self.assertIs(result, confidence)
        self.assertEqual(result, [[0.77, 0.77], [0.77, 0.77]])

    def test_unverified_structure_warning_marks_complete_grid_for_review(self):
        confidence = [[0.99, 0.0], [-2.0, 0.88]]

        result = ocr_backend._apply_structure_review_policy(
            [["编号", ""], ["1", "设备"]],
            confidence,
            structure_verified=False,
            structural_warnings=["行列边界不够稳定。"],
        )

        self.assertEqual(result, [[0.77, 0.77], [-2.0, 0.77]])

    def test_header_only_structure_review_preserves_body_confidence(self):
        confidence = [[0.99, 0.0], [0.99, 0.88], [0.76, 0.95]]

        result = ocr_backend._apply_structure_review_policy(
            [["标题", ""], ["编号", "名称"], ["1", "设备"]],
            confidence,
            structure_verified=False,
            structural_warnings=["物理网格遗漏顶部合并标题。"],
            review_rows={0},
        )

        self.assertEqual(result[0], [0.77, 0.77])
        self.assertEqual(result[1], [0.99, 0.88])
        self.assertEqual(result[2], [0.76, 0.95])

    def test_strong_borderless_title_limits_review_to_risk_cells(self):
        grid = [
            ["设备状态表", "", ""],
            ["编号", "设备号", "状态"],
            ["1", "AP-001", "在用"],
            ["2", "AP-002", "正常"],
            ["3", "AP-003", "待复核"],
            ["4", "AP-004", "已确认"],
        ]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 3, "role": "title"}
        ]

        self.assertTrue(
            ocr_backend._spatial_review_can_be_limited(
                grid,
                spans,
                ["未取得可锁定的物理网格，合并关系未自动应用，请核对行列。"],
                document_mode="borderless_columns",
                engine_name="Hybrid OCR + borderless spatial layout",
                rectification_mode="screen",
            )
        )
        self.assertFalse(
            ocr_backend._spatial_review_can_be_limited(
                grid,
                spans,
                ["检测到疑似行融合，结果已保留供人工调整。"],
                document_mode="borderless_columns",
                engine_name="Hybrid OCR + borderless spatial layout",
                rectification_mode="screen",
            )
        )
        self.assertTrue(
            ocr_backend._spatial_review_can_be_limited(
                grid,
                spans,
                ["物理网格遗漏顶部合并标题。"],
                document_mode="borderless_columns",
                engine_name="Hybrid OCR + review-only page spatial recovery",
                rectification_mode="screen",
                photographic_background=False,
            )
        )
        self.assertFalse(
            ocr_backend._spatial_review_can_be_limited(
                grid,
                spans,
                ["物理网格遗漏顶部合并标题。"],
                document_mode="borderless_columns",
                engine_name="Hybrid OCR + review-only page spatial recovery",
                rectification_mode="auto",
                photographic_background=True,
            )
        )

    def test_missing_confidence_with_warning_creates_review_grid(self):
        result = ocr_backend._apply_structure_review_policy(
            [["编号", "名称"], ["1", "设备"]],
            None,
            structure_verified=True,
            structural_warnings=["照片边缘证据尚未完成闭环。"],
        )

        self.assertEqual(result, [[0.77, 0.77], [0.77, 0.77]])

    def test_merged_blank_subordinate_reviews_are_folded_into_anchor(self):
        grid = [["基础信息", "", ""], ["编号", "名称", "状态"]]
        confidence = [[0.99, 0.77, -2.0], [0.99, 0.99, 0.99]]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 3}
        ]

        ocr_backend._fold_merged_subordinate_reviews_into_anchors(
            grid, confidence, spans
        )

        self.assertEqual(confidence[0], [-2.0, 0.0, 0.0])
        result = table_pipeline.build_result(
            grid,
            confidence=0.99,
            confidences=confidence,
            engine="test",
            spans=spans,
        )
        self.assertTrue(result["cells"][0][0]["needs_review"])
        self.assertFalse(result["cells"][0][1]["needs_review"])
        self.assertFalse(result["cells"][0][2]["needs_review"])

    def test_merged_nonanchor_text_keeps_its_review_evidence(self):
        grid = [["基础信息", "独立文字", ""], ["编号", "名称", "状态"]]
        confidence = [[0.99, 0.77, 0.77], [0.99, 0.99, 0.99]]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 3}
        ]

        ocr_backend._fold_merged_subordinate_reviews_into_anchors(
            grid, confidence, spans
        )

        self.assertEqual(confidence[0], [0.77, 0.77, 0.0])

    def test_leading_metadata_fragments_merge_only_without_physical_dividers(self):
        ocr_backend._load_runtime()
        columns = [0, 100, 200, 300, 400, 499]
        rows = [0, 50, 100, 150, 200, 250, 300, 349]
        image = np.full((350, 500, 3), 255, dtype=np.uint8)
        for boundary in rows:
            cv2.line(image, (0, boundary), (499, boundary), (0, 0, 0), 2)
        for boundary in (0, 499):
            cv2.line(image, (boundary, 0), (boundary, 349), (0, 0, 0), 2)
        for boundary in columns[1:-1]:
            cv2.line(image, (boundary, rows[2]), (boundary, 349), (0, 0, 0), 2)
        cv2.putText(image, "TITLE", (180, 34), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0, 0, 0), 2)
        cv2.putText(image, "META", (180, 84), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0, 0, 0), 2)
        grid = [
            ["台站信息表", "", "", "", ""],
            ["代表队：", "", "", "频段：560-59", "OMHz"],
            ["编号", "频率(MHz)", "信号类型", "占用带宽(kHz)", "调制方式"],
            ["1", "581.450", "GSM", "250", "GMSK"],
            ["2", "574.414", "dpmr", "6", "4FSK"],
            ["3", "588.415", "AM", "6", "AM"],
            ["4", "563.513", "ASK", "25", "ASK"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 5, "role": "title"}
        ]
        evidence = [
            ("代表队：", 50.0, 75.0, 0.98),
            ("频段：560-590MHz", 350.0, 75.0, 0.97),
        ]

        recovered_grid, recovered_confidence, recovered_spans = (
            ocr_backend._recover_leading_metadata_spans(
                [list(row) for row in grid],
                [list(row) for row in confidence],
                list(spans),
                image=image,
                columns=columns,
                rows=rows,
                page_text_evidence=evidence,
            )
        )

        self.assertEqual(
            recovered_grid[1],
            ["代表队：  频段：560-590MHz", "", "", "", ""],
        )
        self.assertEqual(recovered_confidence[1][1:], [0.0, 0.0, 0.0, 0.0])
        self.assertIn(
            {"row": 1, "column": 0, "row_span": 1, "column_span": 5, "role": "subtitle"},
            recovered_spans,
        )
        self.assertTrue(
            ocr_backend._grid_has_two_leading_merged_header_rows(
                image,
                columns,
                rows,
            )
        )

        divided = image.copy()
        for boundary in columns[1:-1]:
            cv2.line(divided, (boundary, rows[1]), (boundary, rows[2]), (0, 0, 0), 2)
        divided_grid, _, divided_spans = ocr_backend._recover_leading_metadata_spans(
            [list(row) for row in grid],
            [list(row) for row in confidence],
            list(spans),
            image=divided,
            columns=columns,
            rows=rows,
            page_text_evidence=evidence,
        )
        self.assertEqual(divided_grid[1], grid[1])
        self.assertNotIn("subtitle", {span.get("role") for span in divided_spans})
        self.assertFalse(
            ocr_backend._grid_has_two_leading_merged_header_rows(
                divided,
                columns,
                rows,
            )
        )

        paged_grid = [list(row) for row in grid]
        paged_grid[1] = ["制表：运营中心", "", "", "", "第8页"]
        paged_confidence = [
            [0.99 if value else 0.0 for value in row]
            for row in paged_grid
        ]
        paged_grid, _, paged_spans = ocr_backend._recover_leading_metadata_spans(
            paged_grid,
            paged_confidence,
            list(spans),
            image=image,
            columns=columns,
            rows=rows,
            page_text_evidence=[
                ("制表：运营中心", 90.0, 75.0, 0.97),
                ("第8页", 450.0, 75.0, 0.98),
            ],
        )
        self.assertEqual(
            paged_grid[1],
            ["制表：运营中心", "", "", "", "第8页"],
        )
        self.assertFalse(any(span["row"] == 1 for span in paged_spans))

        split_image = image.copy()
        cv2.line(
            split_image,
            (columns[2], rows[1]),
            (columns[2], rows[2]),
            (0, 0, 0),
            2,
        )
        split_grid = [list(row) for row in grid]
        split_grid[1] = ["", "单位：服务中心", "", "批次：B84", ""]
        split_confidence = [[0.99 if value else 0.0 for value in row] for row in split_grid]
        split_evidence = [
            ("单位：服务中心", 100.0, 75.0, 0.97),
            ("批次：B84", 350.0, 75.0, 0.96),
        ]
        split_grid, _, split_spans = ocr_backend._recover_leading_metadata_spans(
            split_grid,
            split_confidence,
            list(spans),
            image=split_image,
            columns=columns,
            rows=rows,
            page_text_evidence=split_evidence,
        )
        self.assertEqual(split_grid[1], ["单位：服务中心", "", "批次：B84", "", ""])
        self.assertEqual(
            [span for span in split_spans if span["row"] == 1],
            [
                {"row": 1, "column": 0, "row_span": 1, "column_span": 2, "role": "subtitle"},
                {"row": 1, "column": 2, "row_span": 1, "column_span": 3, "role": "subtitle"},
            ],
        )

        wrong_spans = list(spans) + [
            {"row": 1, "column": 0, "row_span": 1, "column_span": 2, "role": "group_header"},
            {"row": 1, "column": 2, "row_span": 1, "column_span": 1, "role": "group_header"},
            {"row": 1, "column": 3, "row_span": 1, "column_span": 2, "role": "group_header"},
        ]
        rebuilt_grid = [list(row) for row in grid]
        rebuilt_grid[1] = ["", "单位：服务中心", "", "批次：B84", ""]
        rebuilt_confidence = [
            [0.99 if value else 0.0 for value in row]
            for row in rebuilt_grid
        ]
        rebuilt_grid, _, rebuilt_spans = ocr_backend._recover_leading_metadata_spans(
            rebuilt_grid,
            rebuilt_confidence,
            wrong_spans,
            image=split_image,
            columns=columns,
            rows=rows,
            page_text_evidence=split_evidence,
        )
        self.assertEqual(
            [span for span in rebuilt_spans if span["row"] == 1],
            [
                {"row": 1, "column": 0, "row_span": 1, "column_span": 2, "role": "subtitle"},
                {"row": 1, "column": 2, "row_span": 1, "column_span": 3, "role": "subtitle"},
            ],
        )

        body_grid, _, body_spans = ocr_backend._recover_leading_metadata_spans(
            [list(row) for row in split_grid[1:]],
            [[0.99 if value else 0.0 for value in row] for row in split_grid[1:]],
            [],
            image=split_image,
            columns=columns,
            rows=rows[1:],
            page_text_evidence=split_evidence,
            allow_first_row_metadata=True,
        )
        self.assertEqual(body_grid[0], ["单位：服务中心", "", "批次：B84", "", ""])
        self.assertEqual(
            body_spans,
            [
                {"row": 0, "column": 0, "row_span": 1, "column_span": 2, "role": "subtitle"},
                {"row": 0, "column": 2, "row_span": 1, "column_span": 3, "role": "subtitle"},
            ],
        )

    def test_stacked_subtitle_review_does_not_merge_plain_group_headers(self):
        grid = [
            ["快递交接记录表", "", "", "", "", ""],
            ["", "", "部门：综合组", "期间：2026-11", "", ""],
            ["运单号", "收件人", "联系电话", "件数", "交接时间", "签收状态"],
            ["NO-1", "张伟", "138****1000", "1", "08:30", "正常"],
            ["NO-2", "李娜", "138****1001", "2", "09:30", "完成"],
            ["NO-3", "王强", "138****1002", "1", "10:30", "正常"],
            ["NO-4", "陈晨", "138****1003", "3", "11:30", "完成"],
        ]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]
        spans = []

        self.assertTrue(
            ocr_backend._recover_stacked_subtitle_for_review(
                grid,
                confidence,
                spans,
            )
        )
        self.assertEqual(
            grid[1],
            ["部门：综合组", "", "", "期间：2026-11", "", ""],
        )
        self.assertEqual(
            [span["role"] for span in spans],
            ["title", "subtitle", "subtitle"],
        )
        mislabeled_grid = [list(row) for row in grid]
        mislabeled_grid[1] = ["部门：综合组  期间：2026-11", "", "", "", "", ""]
        mislabeled_confidence = [
            [0.95 if value else 0.0 for value in row]
            for row in mislabeled_grid
        ]
        mislabeled_spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 6, "role": "title"},
            {"row": 1, "column": 0, "row_span": 1, "column_span": 6, "role": "title"},
        ]
        self.assertTrue(
            ocr_backend._recover_stacked_subtitle_for_review(
                mislabeled_grid,
                mislabeled_confidence,
                mislabeled_spans,
            )
        )
        self.assertEqual(
            [span["role"] for span in mislabeled_spans],
            ["title", "subtitle"],
        )
        page_grid = [list(row) for row in grid]
        page_grid[1] = ["制表：运营中心", "", "", "", "", "第2页"]
        page_confidence = [
            [0.95 if value else 0.0 for value in row]
            for row in page_grid
        ]
        page_spans = []
        self.assertFalse(
            ocr_backend._recover_stacked_subtitle_for_review(
                page_grid,
                page_confidence,
                page_spans,
            )
        )
        self.assertEqual(
            page_grid[1],
            ["制表：运营中心", "", "", "", "", "第2页"],
        )
        self.assertEqual(page_spans, [])
        spacing_grid = [list(row) for row in grid]
        spacing_grid[1] = ["部门：综合组  期间：2026-11", "", "", "", "", ""]
        spacing_confidence = [[0.95 if value else 0.0 for value in row] for row in spacing_grid]
        spacing_spans = [
            {"row": 1, "column": 0, "row_span": 1, "column_span": 6, "role": "subtitle"}
        ]
        self.assertEqual(
            ocr_backend._normalize_subtitle_field_spacing(
                spacing_grid,
                spacing_confidence,
                spacing_spans,
            ),
            set(),
        )
        statistics_grid = [list(row) for row in grid]
        statistics_grid[1] = ["单位：生产一部统计期：2026-04", "", "", "", "", ""]
        statistics_confidence = [
            [0.95 if value else 0.0 for value in row]
            for row in statistics_grid
        ]
        statistics_spans = [
            {"row": 1, "column": 0, "row_span": 1, "column_span": 6, "role": "subtitle"}
        ]
        self.assertEqual(
            ocr_backend._normalize_subtitle_field_spacing(
                statistics_grid,
                statistics_confidence,
                statistics_spans,
            ),
            {(1, 0)},
        )
        self.assertEqual(
            statistics_grid[1][0],
            "单位：生产一部  统计期：2026-04",
        )
        self.assertEqual(spacing_grid[1][0], "部门：综合组  期间：2026-11")

        partial_grid = [list(row) for row in grid]
        partial_grid[1] = ["", "单位：服务中心", "", "", "批", "批次：B74"]
        partial_confidence = [[0.95 if value else 0.0 for value in row] for row in partial_grid]
        partial_spans = []
        self.assertTrue(
            ocr_backend._recover_stacked_subtitle_for_review(
                partial_grid,
                partial_confidence,
                partial_spans,
            )
        )
        self.assertEqual(
            partial_grid[1],
            ["单位：服务中心", "", "", "批次：B74", "", ""],
        )
        self.assertEqual(
            [(span["column"], span["column_span"]) for span in partial_spans if span["role"] == "subtitle"],
            [(0, 3), (3, 3)],
        )

        adjacent_grid = [list(row) for row in grid]
        adjacent_grid.append(["", "", "", "", "", ""])
        adjacent_grid = [row + ([""] if len(row) == 6 else []) for row in adjacent_grid]
        adjacent_grid[1] = ["单位：物业部", "批次：B61", "", "", "", "", ""]
        adjacent_confidence = [
            [0.95 if value else 0.0 for value in row]
            for row in adjacent_grid
        ]
        adjacent_spans = []
        self.assertTrue(
            ocr_backend._recover_stacked_subtitle_for_review(
                adjacent_grid,
                adjacent_confidence,
                adjacent_spans,
            )
        )
        self.assertEqual(
            adjacent_grid[1],
            ["单位：物业部", "", "", "批次：B61", "", "", ""],
        )
        self.assertEqual(
            [
                (span["column"], span["column_span"])
                for span in adjacent_spans
                if span["role"] == "subtitle"
            ],
            [(0, 3), (3, 4)],
        )

        proven_grid = [list(row) for row in adjacent_grid]
        proven_grid[1] = ["单位：物业部", "", "", "批次：B61", "", "", ""]
        proven_confidence = [
            [0.95 if value else 0.0 for value in row]
            for row in proven_grid
        ]
        proven_spans = [
            {"row": 1, "column": 0, "row_span": 1, "column_span": 3, "role": "subtitle"},
            {"row": 1, "column": 3, "row_span": 1, "column_span": 4, "role": "subtitle"},
        ]
        self.assertFalse(
            ocr_backend._recover_stacked_subtitle_for_review(
                proven_grid,
                proven_confidence,
                proven_spans,
            )
        )
        self.assertEqual(
            proven_grid[1],
            ["单位：物业部", "", "", "批次：B61", "", "", ""],
        )

        split_grid = [list(row) for row in grid]
        split_grid[1] = ["单位：物业部  批次：B86", "", "", "", "", ""]
        split_confidence = [[0.95 if value else 0.0 for value in row] for row in split_grid]
        split_spans = [
            {"row": 1, "column": 0, "row_span": 1, "column_span": 3, "role": "subtitle"},
            {"row": 1, "column": 3, "row_span": 1, "column_span": 3, "role": "subtitle"},
        ]
        self.assertEqual(
            ocr_backend._split_combined_metadata_across_physical_spans(
                split_grid,
                split_confidence,
                split_spans,
            ),
            {(1, 0), (1, 3)},
        )
        self.assertEqual(
            split_grid[1],
            ["单位：物业部", "", "", "批次：B86", "", ""],
        )

        group_grid = [list(row) for row in grid]
        group_grid[1] = ["基础信息", "", "", "业务数据", "", ""]
        group_confidence = [[0.95 if value else 0.0 for value in row] for row in group_grid]
        self.assertFalse(
            ocr_backend._recover_stacked_subtitle_for_review(
                group_grid,
                group_confidence,
                [],
            )
        )

    def test_stacked_two_field_subtitle_uses_balanced_export_spans(self):
        grid = [
            ["运输车辆调度表", "", "", "", "", "", "", ""],
            ["部门：综合管理部", "", "批次：B63", "", "", "", "", ""],
            ["任务号", "车牌号", "司机", "出发地", "目的地", "计划时间", "实际时间", "状态"],
            *[[str(index)] * 8 for index in range(1, 6)],
        ]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]
        spans: list[dict[str, object]] = []

        self.assertTrue(
            ocr_backend._recover_stacked_subtitle_for_review(
                grid, confidence, spans
            )
        )

        self.assertEqual(grid[1][0], "部门：综合管理部")
        self.assertEqual(grid[1][4], "批次：B63")
        self.assertEqual(
            [
                (span["column"], span["column_span"])
                for span in spans
                if span["role"] == "subtitle"
            ],
            [(0, 4), (4, 4)],
        )

    def test_source_margin_title_requires_small_medium_consensus(self):
        ocr_backend._load_runtime()
        source = np.full((200, 300, 3), 245, dtype=np.uint8)
        cv2.putText(
            source,
            "TITLE",
            (100, 70),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.7,
            (20, 20, 20),
            2,
            cv2.LINE_AA,
        )
        grid = [
            ["部门：综合组", "", "期间：2026-06", ""],
            ["编号", "名称", "数量", "状态"],
            *[[str(row), "设备", "1", "正常"] for row in range(5)],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 4, "role": "title"}
        ]
        output = SimpleNamespace(
            txts=["设备维护记录表"] * 4,
            scores=[0.99, 0.98, 0.97, 0.96],
            imgs=None,
        )
        engine = SimpleNamespace(
            fast_text_rec=Mock(return_value=output),
            text_rec=Mock(return_value=output),
        )

        recovered, scores = ocr_backend._recover_missing_title_from_source_margin(
            source,
            {"corners": [[10, 100], [290, 100], [290, 190], [10, 190]]},
            grid,
            confidence,
            spans,
            engine,
        )

        self.assertTrue(recovered)
        self.assertEqual(grid[0], ["设备维护记录表", "", "", ""])
        self.assertEqual(len(grid), 8)
        self.assertEqual(len(scores), 8)
        self.assertTrue(
            any(span["row"] == 1 and span["role"] == "subtitle" for span in spans)
        )
        self.assertTrue(
            any(span["row"] == 0 and span["role"] == "title" for span in spans)
        )

    def test_source_margin_consensus_replaces_one_weak_existing_title(self):
        ocr_backend._load_runtime()
        source = np.full((200, 300, 3), 245, dtype=np.uint8)
        cv2.putText(
            source,
            "TITLE",
            (100, 70),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.7,
            (20, 20, 20),
            2,
            cv2.LINE_AA,
        )
        grid = [
            ["错误标题", "", "", ""],
            ["编号", "名称", "数量", "状态"],
            *[[str(row), "设备", "1", "正常"] for row in range(5)],
        ]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]
        confidence[0][0] = 0.77
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 4, "role": "title"}
        ]
        output = SimpleNamespace(
            txts=["设备维护记录表"] * 4,
            scores=[0.99, 0.98, 0.97, 0.96],
            imgs=None,
        )
        engine = SimpleNamespace(
            fast_text_rec=Mock(return_value=output),
            text_rec=Mock(return_value=output),
        )

        recovered, _ = ocr_backend._recover_missing_title_from_source_margin(
            source,
            {"corners": [[10, 100], [290, 100], [290, 190], [10, 190]]},
            grid,
            confidence,
            spans,
            engine,
        )

        self.assertTrue(recovered)
        self.assertEqual(grid[0], ["设备维护记录表", "", "", ""])
        self.assertEqual(len(grid), 7)
        self.assertEqual(spans[0]["role"], "title")

    def test_weak_title_accepts_source_page_and_medium_agreement(self):
        ocr_backend._load_runtime()
        source = np.full((200, 300, 3), 245, dtype=np.uint8)
        grid = [
            ["错误标题", "", "", ""],
            ["编号", "名称", "数量", "状态"],
            *[[str(row), "设备", "1", "正常"] for row in range(5)],
        ]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]
        confidence[0][0] = 0.77
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 4, "role": "title"}
        ]
        small = SimpleNamespace(
            txts=["设备维护表"] * 4,
            scores=[0.99, 0.98, 0.97, 0.96],
            imgs=None,
        )
        medium = SimpleNamespace(
            txts=["设备维护记录表"] * 4,
            scores=[0.99, 0.98, 0.97, 0.96],
            imgs=None,
        )
        source_page = SimpleNamespace(
            boxes=[np.asarray([[80, 20], [220, 20], [220, 45], [80, 45]])],
            txts=["设备维护记录表"],
            scores=[0.995],
            imgs=None,
        )

        class Engine:
            fast_text_rec = Mock(return_value=small)
            text_rec = Mock(return_value=medium)

            def __call__(self, _image):
                return source_page

        recovered, _ = ocr_backend._recover_missing_title_from_source_margin(
            source,
            {"corners": [[10, 100], [290, 100], [290, 190], [10, 190]]},
            grid,
            confidence,
            spans,
            Engine(),
        )

        self.assertTrue(recovered)
        self.assertEqual(grid[0], ["设备维护记录表", "", "", ""])

    def test_short_categorical_token_is_selected_for_visual_rereading(self):
        ocr_backend._load_runtime()
        grid = [
            ["编号", "设备", "状态"],
            ["1", "A", "异"],
            ["2", "B", "待复核"],
            ["3", "C", "已确认"],
            ["4", "D", "停用"],
            ["5", "E", "在用"],
            ["6", "F", "正常"],
            ["7", "G", "异常"],
            ["8", "H", "待复核"],
            ["9", "I", "已确认"],
        ]

        risks = ocr_backend._short_categorical_token_visual_risks(grid)
        all_risks = ocr_backend._short_categorical_token_visual_risks(
            grid,
            include_all=True,
        )

        self.assertEqual(risks, {(1, 2)})
        self.assertEqual(all_risks, {(row, 2) for row in range(1, 10)})

    def test_signal_name_visual_risks_select_decimal_comma_and_leading_line(self):
        grid = [
            ["编号", "信号名称"],
            ["1", "dmr_7,3k 12.5k"],
            ["2", "1ora 52.3k 62.5k"],
            ["3", "qpsk 25.4k"],
        ]

        risks = ocr_backend._signal_name_visual_risks(grid)

        self.assertEqual(risks, {(1, 1), (2, 1)})

    def test_ruled_recovery_geometry_preserves_physical_cell_boundaries(self):
        geometry = ocr_backend._ruled_grid_recovery_geometry(
            [3, 41, 96],
            [5, 27, 72],
        )

        self.assertEqual(geometry["column_boundaries"], [3.0, 41.0, 96.0])
        self.assertEqual(geometry["row_boundaries"], [5.0, 27.0, 72.0])

    def test_sparse_status_column_is_recovered_only_from_visual_agreement(self):
        ocr_backend._load_runtime()
        image = np.full((100, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "OK", (84, 54), cv2.FONT_HERSHEY_SIMPLEX, 0.35, (0, 0, 0), 1)
        grid = [
            ["编号", "名称", "状态"],
            ["1", "设备A", "正常"],
            ["2", "设备B", ""],
            ["3", "设备C", "复测"],
            ["4", "设备D", ""],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[2][2] = -1.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0, 90.0],
            "grouped_rows": [[{}] for _ in grid],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["复测", "复测"],
                scores=[0.94, 0.92],
            )
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image, grid, confidence, geometry, engine
        )

        self.assertEqual(recovered[2][2], "复测")
        self.assertEqual(recovered_confidence[2][2], 0.92)

    def test_sparse_remark_column_below_old_density_threshold_is_recovered(self):
        ocr_backend._load_runtime()
        image = np.full((120, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "OK", (84, 54), cv2.FONT_HERSHEY_SIMPLEX, 0.35, (0, 0, 0), 1)
        grid = [
            ["编号", "名称", "风险说明"],
            ["1", "任务A", "较重阴影"],
            ["2", "任务B", ""],
            ["3", "任务C", ""],
            ["4", "任务D", ""],
            ["5", "任务E", ""],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[2][2] = -1.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0, 90.0, 110.0],
            "grouped_rows": [[{}] for _ in grid],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["无", "无"],
                scores=[0.94, 0.92],
            )
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image, grid, confidence, geometry, engine, max_candidates=1
        )

        self.assertEqual(recovered[2][2], "无")
        self.assertEqual(recovered_confidence[2][2], 0.92)

    def test_missing_unit_cell_accepts_equivalent_typography_from_two_variants(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 255, dtype=np.uint8)
        cv2.putText(image, "C", (88, 54), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 0), 2)
        grid = [
            ["项目", "数值", "单位"],
            ["A", "1", "V"],
            ["B", "2", ""],
            ["C", "3", "mA"],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[2][2] = -1.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0],
            "grouped_rows": [[{}], [{}], [{}], [{}]],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(txts=["v", "V"], scores=[0.49, 0.49])
        )

        recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
            image, grid, confidence, geometry, engine
        )

        self.assertEqual(recovered[2][2], "V")
        self.assertEqual(recovered_confidence[2][2], 0.49)

    def test_horizontal_mark_confusable_accepts_only_dash_like_recognition(self):
        for value in ("", "-", "_", "一", "—", "–", "−"):
            with self.subTest(value=value):
                self.assertTrue(ocr_backend._is_horizontal_mark_confusable(value))

        for value in ("1", "V", "正常", "."):
            with self.subTest(value=value):
                self.assertFalse(ocr_backend._is_horizontal_mark_confusable(value))

    def test_celsius_degree_ring_requires_a_left_superscript_component(self):
        ocr_backend._load_runtime()
        celsius = np.full((56, 92, 3), 255, dtype=np.uint8)
        cv2.rectangle(celsius, (38, 20), (51, 37), (140, 140, 140), -1)
        cv2.rectangle(celsius, (33, 23), (36, 26), (140, 140, 140), -1)
        plain_c = celsius.copy()
        plain_c[23:27, 33:37] = 255

        self.assertTrue(ocr_backend._has_visible_celsius_degree_ring(celsius))
        self.assertFalse(ocr_backend._has_visible_celsius_degree_ring(plain_c))

    def test_celsius_degree_ring_accepts_connected_camera_glyph_but_rejects_plain_shapes(self):
        ocr_backend._load_runtime()
        celsius = np.full((80, 120, 3), 255, dtype=np.uint8)
        cv2.ellipse(
            celsius,
            (67, 45),
            (19, 18),
            0,
            45,
            315,
            (70, 70, 70),
            8,
            cv2.LINE_AA,
        )
        cv2.circle(celsius, (49, 29), 8, (70, 70, 70), 5, cv2.LINE_AA)
        cv2.line(celsius, (53, 33), (53, 32), (70, 70, 70), 5, cv2.LINE_AA)
        celsius = cv2.GaussianBlur(celsius, (0, 0), 1.2)

        self.assertTrue(ocr_backend._has_visible_celsius_degree_ring(celsius))
        for text in ("C", "V", "%", "O", "A"):
            with self.subTest(text=text):
                plain = np.full((80, 120, 3), 255, dtype=np.uint8)
                cv2.putText(
                    plain,
                    text,
                    (40, 62),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    1.5,
                    (70, 70, 70),
                    5,
                    cv2.LINE_AA,
                )
                plain = cv2.GaussianBlur(plain, (0, 0), 1.2)
                self.assertFalse(
                    ocr_backend._has_visible_celsius_degree_ring(plain)
                )

    def test_percent_glyph_requires_two_diagonal_lobes_and_joining_slash(self):
        ocr_backend._load_runtime()
        percent = np.full((24, 93, 3), 245, dtype=np.uint8)
        cv2.circle(percent, (45, 9), 3, (90, 90, 90), -1, cv2.LINE_AA)
        cv2.circle(percent, (50, 16), 3, (90, 90, 90), -1, cv2.LINE_AA)
        cv2.line(percent, (47, 11), (48, 14), (90, 90, 90), 3, cv2.LINE_AA)
        percent = cv2.GaussianBlur(percent, (0, 0), 0.8)

        self.assertTrue(ocr_backend._has_visible_percent_glyph(percent))
        for text in ("V", "A", "C", "O"):
            with self.subTest(text=text):
                plain = np.full((24, 93, 3), 245, dtype=np.uint8)
                cv2.putText(
                    plain,
                    text,
                    (40, 19),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.45,
                    (90, 90, 90),
                    2,
                    cv2.LINE_AA,
                )
                self.assertFalse(ocr_backend._has_visible_percent_glyph(plain))

    def test_uppercase_v_against_a_uses_open_top_and_lower_point(self):
        ocr_backend._load_runtime()
        glyphs = {}
        for text in ("V", "A"):
            image = np.full((40, 80, 3), 245, dtype=np.uint8)
            cv2.putText(
                image,
                text,
                (5, 32),
                cv2.FONT_HERSHEY_SIMPLEX,
                1.0,
                (50, 50, 50),
                2,
                cv2.LINE_AA,
            )
            glyphs[text] = image

        self.assertTrue(ocr_backend._has_visible_uppercase_v_against_a(glyphs["V"]))
        self.assertFalse(ocr_backend._has_visible_uppercase_v_against_a(glyphs["A"]))

    def test_celsius_suggestion_requires_small_medium_and_visible_ring(self):
        crop = np.full((56, 92, 3), 255, dtype=np.uint8)
        candidates = [
            ("original", "mobile", "C", 0.34),
            ("original", "medium", "C", 0.81),
            ("original", "v5-server", "C", 0.79),
            ("enhanced", "mobile", "C", 0.51),
            ("enhanced", "medium", "C", 0.71),
        ]

        with patch.object(
            ocr_backend, "_has_visible_celsius_degree_ring", return_value=True
        ):
            self.assertEqual(
                ocr_backend._select_visible_celsius_review_suggestion(
                    crop, candidates
                ),
                ("℃", 0.77),
            )
            self.assertIsNone(
                ocr_backend._select_visible_celsius_review_suggestion(
                    crop,
                    [candidate for candidate in candidates if candidate[1] != "mobile"],
                )
            )
        with patch.object(
            ocr_backend, "_has_visible_celsius_degree_ring", return_value=False
        ):
            self.assertIsNone(
                ocr_backend._select_visible_celsius_review_suggestion(
                    crop, candidates
                )
            )

    def test_certified_celsius_recovery_uses_two_model_sizes_and_keeps_review(self):
        ocr_backend._load_runtime()
        image = np.full((80, 160, 3), 255, dtype=np.uint8)
        grid = [["项目", "单位"], ["温度", ""]]
        confidence = [[0.99, 0.99], [0.99, -2.0]]

        def recognizer(text: str, score: float):
            return lambda request: SimpleNamespace(
                txts=[text] * len(request.img),
                scores=[score] * len(request.img),
                imgs=request.img,
            )

        engine = SimpleNamespace(
            fast_text_rec=recognizer("C", 0.51),
            text_rec=recognizer("C", 0.81),
            server_text_rec=recognizer("C", 0.79),
        )
        with patch.object(
            ocr_backend, "_has_visible_celsius_degree_ring", return_value=True
        ):
            scores = ocr_backend._recover_certified_visible_celsius_units(
                image,
                grid,
                confidence,
                [0, 80, 160],
                [0, 40, 80],
                engine,
            )

        self.assertEqual(grid[1][1], "℃")
        self.assertEqual(confidence[1][1], 0.77)
        self.assertEqual(len(scores), 9)

    def test_certified_celsius_recovery_withholds_without_small_model_support(self):
        ocr_backend._load_runtime()
        image = np.full((80, 160, 3), 255, dtype=np.uint8)
        grid = [["项目", "单位"], ["温度", ""]]
        confidence = [[0.99, 0.99], [0.99, -2.0]]
        engine = SimpleNamespace(
            fast_text_rec=lambda request: SimpleNamespace(
                txts=[""] * len(request.img), scores=[0.0] * len(request.img)
            ),
            text_rec=lambda request: SimpleNamespace(
                txts=["C"] * len(request.img), scores=[0.90] * len(request.img)
            ),
            server_text_rec=lambda request: SimpleNamespace(
                txts=["C"] * len(request.img), scores=[0.90] * len(request.img)
            ),
        )
        with patch.object(
            ocr_backend, "_has_visible_celsius_degree_ring", return_value=True
        ):
            ocr_backend._recover_certified_visible_celsius_units(
                image,
                grid,
                confidence,
                [0, 80, 160],
                [0, 40, 80],
                engine,
            )

        self.assertEqual(grid[1][1], "")
        self.assertEqual(confidence[1][1], -2.0)

    def test_path_consensus_preserves_meaningful_spaces(self):
        with_space = r"C:\Users\AP\Pictures\Screenshot 2026-08-03.png"
        without_space = r"C:\Users\AP\Pictures\Screenshot2026-08-03.png"

        self.assertNotEqual(
            ocr_backend._comparable_text(with_space),
            ocr_backend._comparable_text(without_space),
        )

    def test_precision_consensus_preserves_fullwidth_punctuation_and_mixed_text_spaces(self):
        self.assertNotEqual(
            ocr_backend._comparable_text("测试报告（修订2）.pdf"),
            ocr_backend._comparable_text("测试报告(修订2).pdf"),
        )
        self.assertNotEqual(
            ocr_backend._comparable_text("CSV 默认使用 UTF-8"),
            ocr_backend._comparable_text("CSV默认使用UTF-8"),
        )

    def test_near_consensus_recovers_only_long_text_with_three_model_families(self):
        correct = r"E:\设备日志\2026\08\SN-A001-00004567.txt"
        selected = ocr_backend._select_near_consensus_long_text(
            [
                [(correct, 0.99), (correct, 0.98)],
                [(r"E\设备日志|20261081SN-A00-0004567.", 0.96)],
                [(r"E：设备日志\2026\08\SN-A001-00004567.txt", 0.95)],
                [(r"E\设备日志\2026\08\SN-A001-00004567.txt", 0.97)],
            ]
        )

        self.assertEqual(selected, (correct, 0.77))
        self.assertIsNone(
            ocr_backend._select_near_consensus_long_text(
                [[("72.2", 0.99), ("72.2", 0.98)], [("722", 0.99)], [("72.2", 0.97)]]
            )
        )

    def test_near_consensus_preserves_spaces_when_model_families_split_two_to_two(self):
        with_space = "Screenshot 2026-08-03 155849.png"
        without_space = "Screenshot2026-08-03155849.png"

        selected = ocr_backend._select_near_consensus_long_text(
            [
                [(with_space, 0.989), (with_space, 0.986)],
                [(with_space, 0.997)],
                [(without_space, 0.990), (without_space, 0.990)],
                [(without_space, 0.997)],
            ]
        )

        self.assertEqual(selected, (with_space, 0.77))

    def test_near_consensus_accepts_only_whitespace_differences_in_technical_values(self):
        selected = ocr_backend._select_near_consensus_long_text(
            [
                [("8 h 12 min", 0.95), ("8 h 12 min", 0.94)],
                [("8 h 12 min", 0.99)],
                [("8h12min", 0.98), ("8h12min", 0.97)],
                [("8h12min", 0.996)],
            ]
        )

        self.assertEqual(selected, ("8h12min", 0.82))

    def test_exact_cross_family_consensus_recovers_short_precision_symbols(self):
        selected = ocr_backend._select_exact_cross_family_consensus(
            [
                [("÷。", 0.781)],
                [("÷。", 0.902), ("÷。", 0.931)],
                [("÷。", 0.867)],
                [("一。", 0.674)],
            ]
        )

        self.assertIsNotNone(selected)
        self.assertEqual(selected[0], "÷。")
        self.assertGreaterEqual(selected[1], 0.78)

    def test_review_identifier_consensus_accepts_exact_local_server_pair_and_near_family(self):
        selected = ocr_backend._select_review_identifier_consensus(
            [
                [("G-5200", 0.91)],
                [("SG-5200", 0.94)],
                [("ca", 0.28)],
                [("SG-5200", 0.95)],
            ]
        )

        self.assertEqual(selected, ("SG-5200", 0.77))

    def test_review_identifier_consensus_rejects_non_identifiers_and_competing_values(self):
        self.assertIsNone(
            ocr_backend._select_review_identifier_consensus(
                [[("98.1", 0.99)], [("98.1", 0.98)], [("98.19", 0.97)], []]
            )
        )

    def test_identifier_review_adapter_keeps_model_families_independent(self):
        selected = ocr_backend._select_identifier_review_from_view_candidates(
            [
                ("original", "mobile", "G-5200", 0.91),
                ("original", "medium", "SG-5200", 0.94),
                ("enhanced", "medium", "SG-5200", 0.93),
                ("original", "v5-server", "SG-5200", 0.95),
            ]
        )
        self.assertEqual(selected, ("SG-5200", 0.77))

    def test_quantity_single_digit_review_requires_two_families_and_views(self):
        selected = ocr_backend._select_quantity_single_digit_review_suggestion(
            [
                ("original", "medium", "1", 0.81),
                ("enhanced", "medium", "11", 0.92),
                ("unsharp", "mobile", "1", 0.80),
            ]
        )
        self.assertEqual(selected, ("1", 0.77))
        self.assertIsNone(
            ocr_backend._select_quantity_single_digit_review_suggestion(
                [
                    ("original", "medium", "1", 0.91),
                    ("unsharp", "medium", "1", 0.93),
                ]
            )
        )

    def test_quantity_multidigit_review_overrides_isolated_glyph_only_with_full_cell_consensus(self):
        candidates = [
            ("original", "mobile", "1408", 1.0),
            ("original", "medium", "1408", 1.0),
            ("unsharp", "mobile", "1408", 0.9999),
            ("unsharp", "medium", "1408", 1.0),
            ("isolated", "mobile", "1", 0.9999),
            ("isolated", "medium", "1", 0.9999),
            ("isolated", "v5-server", "1", 0.9999),
        ]

        self.assertEqual(
            ocr_backend._select_quantity_multidigit_review_suggestion(candidates),
            ("1408", 0.77),
        )

    def test_quantity_multidigit_review_rejects_one_family_or_competing_full_value(self):
        self.assertIsNone(
            ocr_backend._select_quantity_multidigit_review_suggestion(
                [
                    ("original", "medium", "1408", 1.0),
                    ("unsharp", "medium", "1408", 1.0),
                ]
            )
        )
        self.assertIsNone(
            ocr_backend._select_quantity_multidigit_review_suggestion(
                [
                    ("original", "mobile", "1408", 1.0),
                    ("unsharp", "medium", "1408", 1.0),
                    ("original", "mobile", "1409", 0.99),
                    ("unsharp", "medium", "1409", 0.99),
                ]
            )
        )

    def test_isolated_quantity_digit_requires_all_three_active_families(self):
        votes = [
            ("isolated", "mobile", "1", 0.96),
            ("isolated", "medium", "1", 0.99),
            ("isolated", "v5-server", "1", 0.94),
        ]
        self.assertEqual(
            ocr_backend._select_isolated_quantity_digit_consensus(votes),
            ("1", 0.77),
        )
        self.assertIsNone(
            ocr_backend._select_isolated_quantity_digit_consensus(votes[:-1])
        )

    def test_final_quantity_repair_replaces_a_low_confidence_non_digit(self):
        ocr_backend._load_runtime()
        image = np.full((80, 80, 3), 245, dtype=np.uint8)
        cv2.putText(image, "3", (46, 67), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (20, 20, 20), 2)
        grid = [["编号", "不良数量"], ["1", "m"]]
        confidence = [[0.99, 0.99], [0.99, 0.58]]

        def recognizer(request):
            return SimpleNamespace(
                txts=["3"] * len(request.img),
                scores=[0.96] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=recognizer,
            text_rec=recognizer,
            server_text_rec=recognizer,
            v4_server_text_rec=recognizer,
        )
        with patch.object(
            ocr_backend, "_isolated_single_quantity_glyph_crop", return_value=None
        ):
            ocr_backend._recover_final_ruled_quantity_digits(
                image,
                grid,
                confidence,
                [0, 40, 80],
                [0, 40, 80],
                engine,
            )

        self.assertEqual(grid[1][1], "3")
        self.assertEqual(confidence[1][1], 0.77)

    def test_final_quantity_repair_marks_candidates_beyond_budget_for_review(self):
        ocr_backend._load_runtime()
        image = np.full((160, 80, 3), 245, dtype=np.uint8)
        for row, value in enumerate(("1", "2", "3"), start=1):
            cv2.putText(
                image,
                value,
                (50, row * 40 + 30),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.65,
                (20, 20, 20),
                2,
            )
        grid = [["编号", "数量"], ["1", ""], ["2", ""], ["3", ""]]
        confidence = [[0.99, 0.99], [0.99, 0.0], [0.99, 0.0], [0.99, 0.0]]

        def recognizer(request):
            return SimpleNamespace(
                txts=["1"] * len(request.img),
                scores=[0.96] * len(request.img),
            )

        engine = SimpleNamespace(
            fast_text_rec=recognizer,
            text_rec=recognizer,
            server_text_rec=recognizer,
        )
        with patch.object(
            ocr_backend, "_isolated_single_quantity_glyph_crop", return_value=None
        ):
            ocr_backend._recover_final_ruled_quantity_digits(
                image,
                grid,
                confidence,
                [0, 40, 80],
                [0, 40, 80, 120, 160],
                engine,
                max_candidates=1,
            )

        self.assertEqual(confidence[2][1], -2.0)
        self.assertEqual(confidence[3][1], -2.0)

    def test_final_quantity_repair_groups_views_by_model_family(self):
        ocr_backend._load_runtime()
        image = np.full((80, 80, 3), 245, dtype=np.uint8)
        cv2.putText(image, "3", (46, 67), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (20, 20, 20), 2)
        grid = [["编号", "数量"], ["1", ""]]
        confidence = [[0.99, 0.99], [0.99, 0.0]]
        calls = []

        def recognizer(name):
            def run(request):
                calls.append((name, len(request.img)))
                return SimpleNamespace(
                    txts=["3"] * len(request.img),
                    scores=[0.96] * len(request.img),
                )

            return run

        engine = SimpleNamespace(
            fast_text_rec=recognizer("small"),
            text_rec=recognizer("medium"),
            server_text_rec=recognizer("alternate"),
        )
        with patch.object(
            ocr_backend, "_isolated_single_quantity_glyph_crop", return_value=None
        ):
            ocr_backend._recover_final_ruled_quantity_digits(
                image,
                grid,
                confidence,
                [0, 40, 80],
                [0, 40, 80],
                engine,
            )

        self.assertEqual(
            calls,
            [("medium", 1), ("medium", 1), ("small", 1), ("small", 1)],
        )
        self.assertEqual(grid[1][1], "3")

    def test_isolated_quantity_glyph_rejects_dash_and_two_digits(self):
        single = np.full((82, 204, 3), 245, dtype=np.uint8)
        cv2.rectangle(single, (97, 28), (109, 57), (30, 30, 30), -1)
        self.assertIsNotNone(ocr_backend._isolated_single_quantity_glyph_crop(single))

        dash = np.full((82, 204, 3), 245, dtype=np.uint8)
        cv2.rectangle(dash, (80, 39), (124, 43), (30, 30, 30), -1)
        self.assertIsNone(ocr_backend._isolated_single_quantity_glyph_crop(dash))

        two_digits = single.copy()
        cv2.rectangle(two_digits, (120, 28), (132, 57), (30, 30, 30), -1)
        self.assertIsNone(ocr_backend._isolated_single_quantity_glyph_crop(two_digits))

    def test_isolated_quantity_glyph_ignores_off_center_paper_speckle_cluster(self):
        image = np.full((50, 121, 3), 235, dtype=np.uint8)
        cv2.rectangle(image, (53, 14), (63, 34), (35, 35, 35), -1)
        cv2.rectangle(image, (30, 12), (38, 24), (90, 90, 90), -1)

        self.assertIsNotNone(ocr_backend._isolated_single_quantity_glyph_crop(image))
        self.assertIsNone(
            ocr_backend._select_review_identifier_consensus(
                [
                    [("SG-5200", 0.95)],
                    [("SG-5200", 0.94)],
                    [("SG-520O", 0.96)],
                    [("SG-520O", 0.93)],
                ]
            )
        )

    def test_near_consensus_preserves_multiplication_sign_supported_by_two_families(self):
        selected = ocr_backend._select_near_consensus_long_text(
            [
                [("注意全角标点：，。；！？以及≤≥±×", 0.96)],
                [("注意全角标点：，。；！？以及≤≥±x", 0.98)],
                [("注意全角标点：，。；！？以及≤≥×", 0.94)],
                [("注意全角标点：，。；！？以及≤≥±x", 0.95)],
            ]
        )

        self.assertEqual(selected, ("注意全角标点：，。；！？以及≤≥±×", 0.77))
        self.assertEqual(
            ocr_backend._restore_supported_precision_symbols(
                "axis x",
                [[("axis ×", 0.99)], [("axis ×", 0.98)]],
            ),
            "axis x",
        )
        self.assertEqual(
            ocr_backend._restore_supported_precision_symbols(
                "误差±x",
                [[("误差±×", 0.99)], [("误差±x", 0.98)]],
            ),
            "误差±x",
        )

    def test_borderless_dash_recovery_accepts_short_centered_mark_only_in_relaxed_mode(self):
        ocr_backend._load_runtime()
        image = np.full((40, 240, 3), 255, dtype=np.uint8)
        cv2.line(image, (116, 20), (124, 20), (20, 20, 20), 2)

        self.assertEqual(ocr_backend._recover_horizontal_mark(image), "")
        self.assertEqual(
            ocr_backend._recover_horizontal_mark(image, allow_shorter=True),
            "-",
        )

    def test_borderless_dash_recovery_accepts_a_visibly_lower_aligned_mark(self):
        ocr_backend._load_runtime()
        image = np.full((52, 86, 3), 238, dtype=np.uint8)
        cv2.line(image, (30, 40), (54, 40), (55, 55, 55), 2)

        self.assertEqual(
            ocr_backend._recover_horizontal_mark(image, allow_shorter=True),
            "-",
        )

    def test_borderless_dash_recovery_rejects_a_low_residual_cell_rule(self):
        ocr_backend._load_runtime()
        image = np.full((52, 86, 3), 238, dtype=np.uint8)
        cv2.line(image, (4, 44), (81, 44), (55, 55, 55), 2)

        self.assertEqual(
            ocr_backend._recover_speckled_semantic_horizontal_mark(image),
            "",
        )

    def test_borderless_dash_recovery_accepts_a_short_mark_clipped_by_bottom_warp(self):
        ocr_backend._load_runtime()
        image = np.full((37, 84, 3), 238, dtype=np.uint8)
        cv2.rectangle(image, (30, 34), (53, 36), (55, 55, 55), -1)

        self.assertEqual(
            ocr_backend._recover_horizontal_mark(image, allow_shorter=True),
            "-",
        )

    def test_borderless_dash_recovery_rejects_fragmented_bottom_rule(self):
        ocr_backend._load_runtime()
        image = np.full((30, 104, 3), 238, dtype=np.uint8)
        cv2.line(image, (11, 29), (15, 29), (55, 55, 55), 1)
        cv2.line(image, (28, 29), (31, 29), (55, 55, 55), 1)

        self.assertEqual(
            ocr_backend._recover_horizontal_mark(image, allow_shorter=True),
            "",
        )

    def test_borderless_dash_recovery_rejects_a_wide_bottom_border_fragment(self):
        ocr_backend._load_runtime()
        image = np.full((37, 84, 3), 238, dtype=np.uint8)
        cv2.rectangle(image, (10, 34), (73, 36), (55, 55, 55), -1)

        self.assertEqual(
            ocr_backend._recover_horizontal_mark(image, allow_shorter=True),
            "",
        )

    def test_horizontal_stroke_inside_text_is_not_recovered_as_dash(self):
        ocr_backend._load_runtime()
        image = np.full((50, 160, 3), 255, dtype=np.uint8)
        cv2.rectangle(image, (25, 16), (37, 34), (20, 20, 20), 2)
        cv2.line(image, (70, 25), (88, 25), (20, 20, 20), 2)
        cv2.rectangle(image, (118, 15), (132, 35), (20, 20, 20), 2)

        self.assertEqual(
            ocr_backend._recover_horizontal_mark(image, allow_shorter=True),
            "",
        )

    def test_centered_dash_survives_small_dark_background_speckles(self):
        ocr_backend._load_runtime()
        image = np.full((54, 86, 3), 224, dtype=np.uint8)
        cv2.line(image, (30, 38), (54, 38), (70, 70, 70), 2)
        for x, y in ((8, 20), (15, 47), (51, 23), (72, 45), (23, 32)):
            cv2.circle(image, (x, y), 1, (160, 160, 160), -1)

        self.assertEqual(
            ocr_backend._recover_horizontal_mark(image, allow_shorter=True),
            "-",
        )

    def test_centered_dash_survives_many_tiny_low_light_speckles(self):
        ocr_backend._load_runtime()
        image = np.full((54, 86, 3), 218, dtype=np.uint8)
        cv2.line(image, (30, 36), (54, 36), (55, 55, 55), 2)
        for x, y in (
            (7, 9), (12, 16), (18, 24), (25, 46), (34, 12), (43, 20),
            (51, 48), (60, 10), (67, 22), (73, 43), (79, 15), (82, 31),
            (9, 35), (20, 8), (38, 47), (57, 17), (70, 50), (80, 40),
        ):
            cv2.circle(image, (x, y), 1, (135, 135, 135), -1)

        self.assertEqual(
            ocr_backend._recover_horizontal_mark(image, allow_shorter=True),
            "-",
        )

    def test_narrow_remark_dash_survives_dense_paper_speckles(self):
        ocr_backend._load_runtime()
        image = np.full((80, 800, 3), 210, dtype=np.uint8)
        cv2.line(image, (378, 40), (418, 40), (60, 60, 60), 2)
        for index in range(48):
            x = 8 + (index * 73) % 770
            y = 5 + (index * 29) % 68
            cv2.circle(image, (x, y), 1, (145, 145, 145), -1)

        self.assertEqual(ocr_backend._recover_horizontal_mark(image), "")
        self.assertEqual(
            ocr_backend._recover_speckled_semantic_horizontal_mark(image),
            "-",
        )

    def test_distant_phone_photo_tiny_remark_dash_survives_speckles(self):
        ocr_backend._load_runtime()
        image = np.full((46, 130, 3), 215, dtype=np.uint8)
        cv2.line(image, (60, 23), (69, 23), (58, 58, 58), 2)
        for x, y in (
            (8, 8), (19, 35), (31, 13), (45, 38), (78, 9),
            (92, 33), (108, 14), (121, 37),
        ):
            cv2.circle(image, (x, y), 1, (145, 145, 145), -1)

        self.assertEqual(
            ocr_backend._recover_speckled_semantic_horizontal_mark(image),
            "-",
        )

    def test_relaxed_remark_dash_filter_does_not_accept_a_vertical_digit_one(self):
        ocr_backend._load_runtime()
        image = np.full((80, 800, 3), 210, dtype=np.uint8)
        cv2.rectangle(image, (397, 25), (409, 53), (60, 60, 60), -1)

        self.assertEqual(
            ocr_backend._recover_speckled_semantic_horizontal_mark(image),
            "",
        )

    def test_low_light_speckles_without_an_elongated_mark_remain_blank(self):
        ocr_backend._load_runtime()
        image = np.full((54, 86, 3), 218, dtype=np.uint8)
        for x, y in (
            (7, 9), (12, 16), (18, 24), (25, 46), (34, 12), (43, 20),
            (51, 48), (60, 10), (67, 22), (73, 43), (79, 15), (82, 31),
        ):
            cv2.circle(image, (x, y), 1, (90, 90, 90), -1)

        self.assertEqual(
            ocr_backend._recover_horizontal_mark(image, allow_shorter=True),
            "",
        )

    def test_page_grid_recovers_visible_dash_only_from_empty_cell(self):
        ocr_backend._load_runtime()
        image = np.full((80, 160, 3), 255, dtype=np.uint8)
        cv2.line(image, (116, 60), (124, 60), (20, 20, 20), 2)
        grid = [["名称", "备注"], ["设备A", ""]]
        confidence = [[0.99, 0.99], [0.99, 0.0]]

        ocr_backend._recover_visible_marks_in_grid(
            image,
            grid,
            confidence,
            [0, 80, 160],
            [0, 40, 80],
        )

        self.assertEqual(grid, [["名称", "备注"], ["设备A", "—"]])
        self.assertEqual(confidence[1][1], 0.86)

    def test_page_grid_never_invents_a_dash_in_a_unit_column(self):
        ocr_backend._load_runtime()
        image = np.full((80, 160, 3), 255, dtype=np.uint8)
        cv2.line(image, (116, 60), (124, 60), (20, 20, 20), 2)
        grid = [["名称", "单位"], ["温度", ""]]
        confidence = [[0.99, 0.99], [0.99, -1.0]]

        ocr_backend._recover_visible_marks_in_grid(
            image,
            grid,
            confidence,
            [0, 80, 160],
            [0, 40, 80],
        )

        self.assertEqual(grid[1][1], "")
        self.assertEqual(confidence[1][1], -1.0)

    def test_visible_mark_recovery_preserves_a_negative_conflict(self):
        ocr_backend._load_runtime()
        image = np.full((80, 160, 3), 255, dtype=np.uint8)
        cv2.line(image, (116, 60), (124, 60), (20, 20, 20), 2)
        grid = [["名称", "备注"], ["设备A", ""]]
        confidence = [[0.99, 0.99], [0.99, -1.0]]

        ocr_backend._recover_visible_marks_in_grid(
            image,
            grid,
            confidence,
            [0, 80, 160],
            [0, 40, 80],
        )

        self.assertEqual(grid[1][1], "—")
        self.assertEqual(confidence[1][1], -1.0)

    def test_visible_mark_and_existing_dash_release_negative_conflict_for_review(self):
        ocr_backend._load_runtime()
        image = np.full((80, 160, 3), 255, dtype=np.uint8)
        cv2.line(image, (116, 60), (124, 60), (20, 20, 20), 2)
        grid = [["名称", "备注"], ["设备A", "-"]]
        confidence = [[0.99, 0.99], [0.99, -1.0]]

        ocr_backend._recover_visible_marks_in_grid(
            image,
            grid,
            confidence,
            [0, 80, 160],
            [0, 40, 80],
        )

        self.assertEqual(grid[1][1], "—")
        self.assertEqual(confidence[1][1], 0.77)

    def test_summary_row_recovers_short_dash_outside_semantic_columns(self):
        ocr_backend._load_runtime()
        image = np.full((90, 360, 3), 255, dtype=np.uint8)
        cv2.line(image, (293, 75), (299, 75), (20, 20, 20), 2)
        grid = [
            ["产品", "数量", "单价"],
            ["设备A", "2", "10"],
            ["合计", "2", ""],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        ocr_backend._recover_visible_marks_in_grid(
            image,
            grid,
            confidence,
            [0, 120, 240, 360],
            [0, 30, 60, 90],
        )

        self.assertEqual(grid[2], ["合计", "2", "—"])
        self.assertEqual(confidence[2][2], 0.86)

    def test_summary_row_does_not_turn_bottom_rule_fragment_into_dash(self):
        ocr_backend._load_runtime()
        image = np.full((90, 360, 3), 255, dtype=np.uint8)
        cv2.line(image, (293, 86), (299, 86), (20, 20, 20), 2)
        grid = [
            ["产品", "数量", "单价"],
            ["设备A", "2", "10"],
            ["合计", "2", ""],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        ocr_backend._recover_visible_marks_in_grid(
            image,
            grid,
            confidence,
            [0, 120, 240, 360],
            [0, 30, 60, 90],
        )

        self.assertEqual(grid[2], ["合计", "2", ""])
        self.assertEqual(confidence[2][2], 0.0)

    def test_page_grid_does_not_turn_title_or_header_rules_into_dashes(self):
        ocr_backend._load_runtime()
        image = np.full((120, 160, 3), 255, dtype=np.uint8)
        cv2.line(image, (116, 20), (124, 20), (20, 20, 20), 2)
        cv2.line(image, (116, 60), (124, 60), (20, 20, 20), 2)
        cv2.line(image, (116, 100), (124, 100), (20, 20, 20), 2)
        grid = [["统计表", ""], ["名称", "备注"], ["设备A", ""]]
        confidence = [[0.99, 0.0], [0.99, 0.99], [0.99, 0.0]]

        ocr_backend._recover_visible_marks_in_grid(
            image,
            grid,
            confidence,
            [0, 80, 160],
            [0, 40, 80, 120],
        )

        self.assertEqual(grid[0], ["统计表", ""])
        self.assertEqual(grid[1], ["名称", "备注"])
        self.assertEqual(grid[2], ["设备A", "—"])

    def test_visible_dash_is_recovered_when_recognizer_reads_chinese_one(self):
        ocr_backend._load_runtime()
        image = np.full((80, 120, 3), 255, dtype=np.uint8)
        grid = [
            ["名称", "备注", "状态"],
            ["设备A", "-", "正常"],
            ["设备B", "", "正常"],
            ["设备C", "-", "正常"],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[2][1] = -1.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0],
            "grouped_rows": [[{}], [{}], [{}], [{}]],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(txts=["一", "一"], scores=[0.92, 0.90])
        )

        with patch.object(ocr_backend, "_recover_horizontal_mark", return_value="-"):
            recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
                image, grid, confidence, geometry, engine
            )

        self.assertEqual(recovered[2][1], "—")
        self.assertEqual(recovered_confidence[2][1], 0.86)

    def test_repeated_visible_marks_in_remark_column_override_ocr_hallucination(self):
        ocr_backend._load_runtime()
        image = np.full((100, 120, 3), 230, dtype=np.uint8)
        grid = [
            ["名称", "备注", "状态"],
            ["设备A", "", "正常"],
            ["设备B", "", "正常"],
            ["设备C", "完成", "正常"],
            ["设备D", "完成", "正常"],
        ]
        confidence = [[0.95] * 3 for _ in grid]
        confidence[1][1] = -1.0
        confidence[2][1] = -1.0
        geometry = {
            "anchors": [20.0, 60.0, 100.0],
            "row_centers": [10.0, 30.0, 50.0, 70.0, 90.0],
            "grouped_rows": [[{}], [{}], [{}], [{}], [{}]],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(
                txts=["7", "7", "7", "7"], scores=[0.94, 0.93, 0.92, 0.91]
            )
        )

        with patch.object(ocr_backend, "_recover_horizontal_mark", return_value="-"):
            recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
                image, grid, confidence, geometry, engine
            )

        self.assertEqual(recovered[1][1], "—")
        self.assertEqual(recovered[2][1], "—")
        self.assertEqual(recovered_confidence[1][1], 0.86)
        self.assertEqual(recovered_confidence[2][1], 0.86)

    def test_visible_dash_keeps_priority_beyond_missing_cell_ocr_budget(self):
        ocr_backend._load_runtime()
        image = np.full((210, 250, 3), 255, dtype=np.uint8)
        grid = [[f"R{row}C{column}" for column in range(5)] for row in range(20)]
        confidence = [[0.95] * 5 for _ in grid]
        for row in range(1, 16):
            column = (row - 1) % 5
            grid[row][column] = ""
            confidence[row][column] = -1.0

        # The visually present mark is in the earliest blank.  The previous
        # reverse tie ordering dropped it before crop inspection when the OCR
        # budget was smaller than the number of blanks.
        cv2.line(image, (5, 15), (45, 15), (0, 0, 0), 2)
        geometry = {
            "anchors": [25.0, 75.0, 125.0, 175.0, 225.0],
            "row_centers": [5.0 + row * 10.0 for row in range(20)],
            "grouped_rows": [[{}] for _ in grid],
            "first_structured_row": 0,
        }
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(txts=["一", "一"], scores=[0.92, 0.90])
        )

        def visible_mark(crop, allow_shorter=False):
            gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY) if crop.ndim == 3 else crop
            return "-" if float(np.mean(gray < 100)) > 0.01 else ""

        with patch.object(ocr_backend, "_recover_horizontal_mark", side_effect=visible_mark):
            recovered, recovered_confidence, _ = ocr_backend._recover_missing_spatial_cells(
                image,
                grid,
                confidence,
                geometry,
                engine,
                max_candidates=3,
            )

        self.assertEqual(recovered[1][0], "—")
        self.assertEqual(recovered_confidence[1][0], 0.86)

    def test_repeated_visible_marks_only_normalize_dash_confusable_semantic_text(self):
        grid = [
            ["编号", "名称", "备注"],
            ["A001", "横杠测试", "人"],
            ["A002", "保留文字", "一"],
        ]
        confidence = [[0.99] * 3 for _ in grid]

        ocr_backend._restore_repeated_semantic_visible_marks(
            grid,
            confidence,
            {(1, 1): "-", (2, 1): "-", (1, 2): "-", (2, 2): "-"},
        )

        self.assertEqual(grid[1][1], "横杠测试")
        self.assertEqual(grid[2][1], "保留文字")
        self.assertEqual(grid[1][2], "人")
        self.assertEqual(grid[2][2], "—")
        self.assertEqual(confidence[1][2], 0.99)

    def test_repeated_visible_marks_never_replace_confident_status_text(self):
        grid = [
            ["文件名称", "结果"],
            ["报告.xlsx", "已完成"],
            ["config.json", "已配置"],
        ]
        confidence = [[0.99] * 2 for _ in grid]

        ocr_backend._restore_repeated_semantic_visible_marks(
            grid,
            confidence,
            {(1, 1): "-", (2, 1): "-"},
        )

        self.assertEqual(grid[1][1], "已完成")
        self.assertEqual(grid[2][1], "已配置")
        self.assertEqual(confidence[1][1], 0.99)
        self.assertEqual(confidence[2][1], 0.99)

    def test_consistency_checks_preserve_visible_unit_spelling(self):
        grid = [
            ["名称", "单位", "备注"],
            ["供电", "v", "v"],
            ["温度", "℃", "保持原样"],
        ]
        confidence = [[0.99] * 3 for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[1], ["供电", "v", "v"])
        self.assertEqual(grid[2][1], "℃")

    def test_noncanonical_unit_case_risks_only_mark_unit_columns(self):
        grid = [
            ["名称", "单位", "备注"],
            ["电压", "v", "v"],
            ["电流", "mA", "ma"],
            ["频率", "khz", "khz"],
            ["温度", "℃", "保持原样"],
            ["标准", "V", "V"],
        ]

        self.assertEqual(
            ocr_backend._noncanonical_unit_case_risks(grid),
            {(1, 1), (3, 1)},
        )

    def test_noncanonical_unit_case_risks_require_rectangular_unit_header(self):
        self.assertEqual(
            ocr_backend._noncanonical_unit_case_risks([["名称", "备注"], ["电压", "v"]]),
            set(),
        )
        self.assertEqual(
            ocr_backend._noncanonical_unit_case_risks([["名称", "单位"], ["电压"]]),
            set(),
        )

    def test_repeated_unit_prefix_recovery_requires_two_high_medium_views(self):
        ocr_backend._load_runtime()
        grid = [
            ["序号", "单位"],
            ["1", "mm/s"],
            ["2", "mm/s"],
            ["3", "mm/s"],
            ["4", "m/s"],
        ]
        confidence = [[0.99, 0.99] for _ in grid]
        image = np.full((100, 100, 3), 245, dtype=np.uint8)
        engine = SimpleNamespace(
            text_rec=lambda _: SimpleNamespace(txts=["mm/s"], scores=[0.97]),
            server_text_rec=lambda _: SimpleNamespace(txts=["mm/s"], scores=[0.98]),
        )
        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 30, 3), 245, dtype=np.uint8),
        ):
            scores = ocr_backend._recover_repeated_unit_prefix_confusions(
                image,
                grid,
                confidence,
                [0, 50, 100],
                [0, 20, 40, 60, 80, 100],
                engine,
            )
        self.assertEqual(grid[4][1], "mm/s")
        self.assertEqual(confidence[4][1], 0.77)
        self.assertEqual(scores, [0.97, 0.98])

        rejected_grid = [list(row) for row in grid]
        rejected_grid[4][1] = "m/s"
        rejected_confidence = [[0.99, 0.99] for _ in rejected_grid]
        rejected_engine = SimpleNamespace(
            text_rec=lambda _: SimpleNamespace(txts=["mm/s"], scores=[0.94]),
            server_text_rec=lambda _: SimpleNamespace(txts=["mm/s"], scores=[0.99]),
        )
        with patch.object(
            ocr_backend,
            "_tight_text_crop",
            return_value=np.full((12, 30, 3), 245, dtype=np.uint8),
        ):
            ocr_backend._recover_repeated_unit_prefix_confusions(
                image,
                rejected_grid,
                rejected_confidence,
                [0, 50, 100],
                [0, 20, 40, 60, 80, 100],
                rejected_engine,
            )
        self.assertEqual(rejected_grid[4][1], "m/s")

    def test_ambiguous_identifier_case_risks_do_not_rewrite_model_output(self):
        grid = [
            ["序号", "英文标识", "备注"],
            ["1", "DAQ-1Il-001", "normal"],
            ["2", "DAQ-111-002", "keep"],
            ["3", "DAQ-A1IlB-003", "embedded"],
            ["4", "DAQ-8B-234", "unchanged"],
        ]
        confidence = [[0.99, 0.99, 0.99] for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[1][1], "DAQ-1Il-001")
        self.assertEqual(grid[2][1], "DAQ-111-002")
        self.assertEqual(grid[3][1], "DAQ-A1IlB-003")
        self.assertEqual(
            ocr_backend._ambiguous_identifier_case_risks(grid),
            {(1, 1), (2, 1), (3, 1)},
        )

    def test_ambiguous_identifier_case_risks_cover_model_and_serial_columns(self):
        grid = [
            ["Model", "Serial No.", "普通备注"],
            ["IO-11I-001", "MOD-1II-002", "IO-11I-003"],
        ]

        self.assertEqual(
            ocr_backend._ambiguous_identifier_case_risks(grid),
            {(1, 0), (1, 1)},
        )

    def test_percentage_decimal_pattern_risks_selects_only_format_outlier(self):
        grid = [["序号", "采样率"]] + [
            [str(index), value]
            for index, value in enumerate(
                ["97.49%", "89.26%", "99.27%", "93.65%", "9339%", "88.51%", "96.00%"],
                start=1,
            )
        ]

        self.assertEqual(
            ocr_backend._percentage_decimal_pattern_risks(grid),
            {(5, 1)},
        )

    def test_percentage_decimal_pattern_risks_keeps_integer_percent_column(self):
        grid = [["序号", "完成率"]] + [
            [str(index), value]
            for index, value in enumerate(
                ["80%", "85%", "90%", "95%", "100%", "88%", "92%"],
                start=1,
            )
        ]

        self.assertEqual(ocr_backend._percentage_decimal_pattern_risks(grid), set())

    def test_ipv4_sequence_truncation_risks_selects_short_final_octets(self):
        grid = [["序号", "IP地址"]] + [
            [str(index), value]
            for index, value in enumerate(
                [
                    "192.168.8.20",
                    "192.168.7.21",
                    "192.168.4.22",
                    "192.168.6.23",
                    "192.168.9.24",
                    "192.168.11.2",
                    "192.168.10.2",
                    "192.168.3.27",
                ],
                start=1,
            )
        ]

        self.assertEqual(
            ocr_backend._ipv4_sequence_truncation_risks(grid),
            {(6, 1), (7, 1)},
        )

    def test_ipv4_sequence_truncation_risks_keeps_short_ip_series(self):
        grid = [["序号", "IP地址"]] + [
            [str(index), f"10.0.0.{index}"] for index in range(1, 9)
        ]

        self.assertEqual(ocr_backend._ipv4_sequence_truncation_risks(grid), set())

    def test_visible_uppercase_v_glyph_rejects_lower_x_height_v(self):
        ocr_backend._load_runtime()
        uppercase = np.full((34, 60, 3), 255, dtype=np.uint8)
        lowercase = np.full((34, 60, 3), 255, dtype=np.uint8)
        cv2.putText(
            uppercase,
            "V",
            (22, 26),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.75,
            (20, 20, 20),
            2,
            cv2.LINE_AA,
        )
        cv2.putText(
            lowercase,
            "v",
            (22, 26),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.75,
            (20, 20, 20),
            2,
            cv2.LINE_AA,
        )

        self.assertTrue(ocr_backend._has_visible_uppercase_v_glyph(uppercase))
        self.assertFalse(ocr_backend._has_visible_uppercase_v_glyph(lowercase))

    def test_visible_leading_horizontal_cjk_stroke_requires_three_glyph_width(self):
        ocr_backend._load_runtime()
        positive = np.full((24, 100), 255, dtype=np.uint8)
        cv2.line(positive, (8, 12), (23, 12), 0, 2)
        cv2.rectangle(positive, (23, 5), (37, 19), 0, 2)
        cv2.line(positive, (23, 12), (37, 12), 0, 2)
        cv2.rectangle(positive, (42, 5), (56, 19), 0, 2)
        cv2.line(positive, (42, 12), (56, 12), 0, 2)

        negative = np.full((24, 100), 255, dtype=np.uint8)
        cv2.rectangle(negative, (9, 5), (23, 19), 0, 2)
        cv2.line(negative, (9, 12), (23, 12), 0, 2)
        cv2.rectangle(negative, (28, 5), (42, 19), 0, 2)
        cv2.line(negative, (28, 12), (42, 12), 0, 2)

        self.assertTrue(
            ocr_backend._has_visible_leading_horizontal_cjk_stroke(positive)
        )
        self.assertFalse(
            ocr_backend._has_visible_leading_horizontal_cjk_stroke(negative)
        )

    def test_visible_uppercase_v_glyph_supports_left_aligned_cells(self):
        ocr_backend._load_runtime()
        uppercase = np.full((34, 100, 3), 255, dtype=np.uint8)
        lowercase = np.full((34, 100, 3), 255, dtype=np.uint8)
        cv2.putText(
            uppercase,
            "V",
            (4, 26),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.75,
            (20, 20, 20),
            2,
            cv2.LINE_AA,
        )
        cv2.putText(
            lowercase,
            "v",
            (4, 26),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.75,
            (20, 20, 20),
            2,
            cv2.LINE_AA,
        )

        self.assertTrue(ocr_backend._has_visible_uppercase_v_glyph(uppercase))
        self.assertFalse(ocr_backend._has_visible_uppercase_v_glyph(lowercase))

    def test_visible_uppercase_v_glyph_accepts_clipped_left_origin(self):
        ocr_backend._load_runtime()
        uppercase = np.full((34, 60, 3), 255, dtype=np.uint8)
        lowercase = np.full((34, 60, 3), 255, dtype=np.uint8)
        cv2.putText(
            uppercase,
            "V",
            (-2, 26),
            cv2.FONT_HERSHEY_SIMPLEX,
            1.0,
            (20, 20, 20),
            2,
            cv2.LINE_AA,
        )
        cv2.putText(
            lowercase,
            "v",
            (-2, 26),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.75,
            (20, 20, 20),
            2,
            cv2.LINE_AA,
        )

        self.assertTrue(ocr_backend._has_visible_uppercase_v_glyph(uppercase))
        self.assertFalse(ocr_backend._has_visible_uppercase_v_glyph(lowercase))

    def test_visible_uppercase_v_suffix_uses_relative_glyph_height(self):
        ocr_backend._load_runtime()
        uppercase = np.full((34, 80, 3), 255, dtype=np.uint8)
        lowercase = np.full((34, 80, 3), 255, dtype=np.uint8)
        cv2.putText(
            uppercase,
            "mV",
            (4, 27),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.75,
            (20, 20, 20),
            2,
            cv2.LINE_AA,
        )
        cv2.putText(
            lowercase,
            "mv",
            (4, 27),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.75,
            (20, 20, 20),
            2,
            cv2.LINE_AA,
        )

        self.assertTrue(ocr_backend._has_visible_uppercase_v_suffix(uppercase))
        self.assertFalse(ocr_backend._has_visible_uppercase_v_suffix(lowercase))

    def test_active_certificate_boundaries_apply_ui_offsets(self):
        certificate = {
            "column_boundaries": [0, 20, 70, 120],
            "row_boundaries": [0, 15, 45, 75, 105],
            "column_offset": 1,
            "row_offset": 1,
        }
        grid = [["名称", "单位"], ["电压", "V"], ["电流", "A"]]

        self.assertEqual(
            ocr_backend._active_structure_certificate_boundaries(certificate, grid),
            ([20, 70, 120], [15, 45, 75, 105]),
        )

    def test_active_certificate_boundaries_reject_shape_mismatch(self):
        certificate = {
            "column_boundaries": [0, 50, 100],
            "row_boundaries": [0, 30, 60],
            "column_offset": 0,
            "row_offset": 0,
        }
        grid = [["名称", "单位"], ["电压", "V"], ["电流", "A"]]

        self.assertIsNone(
            ocr_backend._active_structure_certificate_boundaries(certificate, grid)
        )

    def test_visible_uppercase_v_recovery_changes_only_unit_column_for_review(self):
        grid = [
            ["名称", "单位", "备注"],
            ["电压", "v", "v"],
            ["说明", "mA", "保持原样"],
        ]
        confidence = [[0.99 for _ in row] for row in grid]
        image = np.full((90, 240, 3), 255, dtype=np.uint8)
        columns = [0, 80, 150, 240]
        rows = [0, 30, 60, 90]

        with patch.object(
            ocr_backend,
            "_has_visible_uppercase_v_glyph",
            return_value=True,
        ):
            recovered = ocr_backend._recover_visible_uppercase_v_units(
                image,
                grid,
                confidence,
                columns,
                rows,
            )

        self.assertEqual(recovered, {(1, 1)})
        self.assertEqual(grid[1], ["电压", "V", "v"])
        self.assertEqual(confidence[1][1], 0.77)

    def test_visible_uppercase_v_recovery_preserves_prefix_and_marks_review(self):
        grid = [["名称", "单位"], ["电压", "mv"]]
        confidence = [[0.99, 0.99], [0.99, 0.91]]
        image = np.full((60, 160, 3), 255, dtype=np.uint8)

        with patch.object(
            ocr_backend,
            "_has_visible_uppercase_v_suffix",
            return_value=True,
        ):
            recovered = ocr_backend._recover_visible_uppercase_v_units(
                image,
                grid,
                confidence,
                [0, 80, 160],
                [0, 30, 60],
            )

        self.assertEqual(recovered, {(1, 1)})
        self.assertEqual(grid[1][1], "mV")
        self.assertEqual(confidence[1][1], 0.77)

    def test_visible_uppercase_v_recovery_accepts_visible_blank_unit(self):
        grid = [["名称", "单位"], ["电压", ""]]
        confidence = [[0.99, 0.99], [0.99, 0.0]]
        image = np.full((60, 160, 3), 255, dtype=np.uint8)
        with patch.object(
            ocr_backend,
            "_has_visible_uppercase_v_glyph",
            return_value=True,
        ):
            recovered = ocr_backend._recover_visible_uppercase_v_units(
                image,
                grid,
                confidence,
                [0, 80, 160],
                [0, 30, 60],
            )

        self.assertEqual(recovered, {(1, 1)})
        self.assertEqual(grid[1][1], "V")
        self.assertEqual(confidence[1][1], 0.77)

        review_grid = [["名称", "单位"], ["电压", ""]]
        review_confidence = [[0.99, 0.99], [0.99, -1.0]]
        with patch.object(
            ocr_backend,
            "_has_visible_uppercase_v_glyph",
            return_value=True,
        ):
            review_recovered = ocr_backend._recover_visible_uppercase_v_units(
                image,
                review_grid,
                review_confidence,
                [0, 80, 160],
                [0, 30, 60],
                allow_blank=False,
            )
        self.assertEqual(review_recovered, set())
        self.assertEqual(review_grid[1][1], "")

    def test_visible_celsius_recovery_requires_unit_header_and_stays_review(self):
        grid = [["名称", "单位", "备注"], ["温度", "V", "V"], ["温度", "°C", "V"]]
        confidence = [[0.99, 0.99, 0.99], [0.99, 0.93, 0.93], [0.99, 0.93, 0.93]]
        image = np.full((90, 240, 3), 255, dtype=np.uint8)
        with patch.object(
            ocr_backend,
            "_has_visible_celsius_degree_ring",
            return_value=True,
        ):
            recovered = ocr_backend._recover_visible_celsius_unit_glyphs(
                image,
                grid,
                confidence,
                [0, 80, 160, 240],
                [0, 30, 60, 90],
            )

        self.assertEqual(recovered, {(1, 1), (2, 1)})
        self.assertEqual(grid[1], ["温度", "℃", "V"])
        self.assertEqual(grid[2], ["温度", "℃", "V"])
        self.assertEqual(confidence[1][1], 0.77)
        self.assertEqual(confidence[2][1], 0.77)

    def test_visible_celsius_header_recovery_restores_degree_ring(self):
        grid = [
            ["实验测量数据表", "", ""],
            ["样本编号", "温度(C)", "结论"],
            ["A01", "21.59", "正常"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        image = np.full((90, 240, 3), 255, dtype=np.uint8)
        with patch.object(
            ocr_backend,
            "_has_visible_celsius_degree_ring",
            return_value=True,
        ):
            recovered = ocr_backend._recover_visible_celsius_header_glyphs(
                image,
                grid,
                confidence,
                [0, 80, 160, 240],
                [0, 30, 60, 90],
            )

        self.assertEqual(recovered, {(1, 1)})
        self.assertEqual(grid[1][1], "温度(℃)")
        self.assertEqual(confidence[1][1], 0.77)

    def test_celsius_header_multiview_requires_unanimous_medium_text(self):
        ocr_backend._load_runtime()
        grid = [
            ["实验测量数据表", "", ""],
            ["样本编号", "温度(C)", "结论"],
            ["A01", "21.59", "正常"],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        image = np.full((90, 240, 3), 245, dtype=np.uint8)
        cv2.putText(image, "TEMP(C)", (85, 52), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (20, 20, 20), 1)
        output = SimpleNamespace(
            txts=["温度(℃)"] * 4,
            scores=[0.96, 0.95, 0.97, 0.99],
            imgs=None,
        )
        engine = SimpleNamespace(
            text_rec=Mock(return_value=output),
            server_text_rec=Mock(return_value=output),
        )

        recovered, scores = ocr_backend._recover_celsius_header_multiview(
            image,
            grid,
            confidence,
            [0, 80, 160, 240],
            [0, 30, 60, 90],
            engine,
        )

        self.assertEqual(recovered, {(1, 1)})
        self.assertEqual(grid[1][1], "温度(℃)")
        self.assertEqual(len(scores), 8)

    def test_visible_percent_recovery_requires_unit_header_and_stays_review(self):
        grid = [["名称", "单位", "备注"], ["偏差", "V", "V"]]
        confidence = [[0.99, 0.99, 0.99], [0.99, 0.93, 0.93]]
        image = np.full((60, 240, 3), 255, dtype=np.uint8)
        with patch.object(
            ocr_backend,
            "_has_visible_percent_glyph",
            return_value=True,
        ):
            recovered = ocr_backend._recover_visible_percent_unit_glyphs(
                image,
                grid,
                confidence,
                [0, 80, 160, 240],
                [0, 30, 60],
            )

        self.assertEqual(recovered, {(1, 1)})
        self.assertEqual(grid[1], ["偏差", "%", "V"])
        self.assertEqual(confidence[1][1], 0.77)

    def test_visible_v_from_a_recovery_requires_two_unit_crop_views(self):
        grid = [["名称", "单位", "备注"], ["电压", "A", "A"]]
        confidence = [[0.99, 0.99, 0.99], [0.99, 0.93, 0.93]]
        image = np.full((60, 240, 3), 255, dtype=np.uint8)
        with patch.object(
            ocr_backend,
            "_has_visible_uppercase_v_against_a",
            side_effect=[True, True, False, False],
        ):
            recovered = ocr_backend._recover_visible_v_from_a_units(
                image,
                grid,
                confidence,
                [0, 80, 160, 240],
                [0, 30, 60],
            )

        self.assertEqual(recovered, {(1, 1)})
        self.assertEqual(grid[1], ["电压", "V", "A"])
        self.assertEqual(confidence[1][1], 0.77)

    def test_dominant_celsius_spelling_repairs_only_bounded_unit_variant(self):
        grid = [
            ["名称", "单位", "备注"],
            ["温度1", "℃", "°C"],
            ["温度2", "℃", "°C"],
            ["温度3", "℃", "保持"],
            ["温度4", "°C", "保持"],
        ]
        confidence = [[0.99 for _ in row] for row in grid]

        repaired = ocr_backend._restore_dominant_celsius_spelling(
            grid, confidence
        )

        self.assertEqual(repaired, {(4, 1)})
        self.assertEqual(grid[4], ["温度4", "℃", "保持"])
        self.assertEqual(grid[1][2], "°C")

    def test_visible_uppercase_v_recovery_keeps_left_aligned_first_stroke(self):
        ocr_backend._load_runtime()
        grid = [["名称", "单位"], ["电压", "v"]]
        confidence = [[0.99, 0.99], [0.99, 0.91]]
        image = np.full((60, 160, 3), 255, dtype=np.uint8)
        cv2.putText(
            image,
            "V",
            (84, 55),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.75,
            (20, 20, 20),
            2,
            cv2.LINE_AA,
        )

        recovered = ocr_backend._recover_visible_uppercase_v_units(
            image,
            grid,
            confidence,
            [0, 80, 160],
            [0, 30, 60],
        )

        self.assertEqual(recovered, {(1, 1)})
        self.assertEqual(grid[1][1], "V")
        self.assertEqual(confidence[1][1], 0.77)

    def test_periodic_units_repair_one_missing_canonical_value_for_review(self):
        grid = [
            ["项目", "单位"],
            ["项目01", "mA"],
            ["项目02", "kHz"],
            ["项目03", "°C"],
            ["项目04", "V"],
            ["项目05", "mA"],
            ["项目06", "kHz"],
            ["项目07", ""],
            ["项目08", "V"],
            ["项目09", "mA"],
            ["项目10", "kHz"],
        ]
        confidence = [[0.99] * 2 for _ in grid]
        confidence[7][1] = 0.0

        repaired = ocr_backend._repair_periodic_unit_blanks(grid, confidence)

        self.assertEqual(repaired, [(7, 1, "°C")])
        self.assertEqual(grid[7][1], "°C")
        self.assertEqual(confidence[7][1], 0.77)

    def test_periodic_units_repair_one_visible_case_outlier_for_review(self):
        grid = [
            ["项目", "单位"],
            ["项目01", "mA"],
            ["项目02", "kHz"],
            ["项目03", "℃"],
            ["项目04", "v"],
            ["项目05", "mA"],
            ["项目06", "kHz"],
            ["项目07", "℃"],
            ["项目08", "V"],
            ["项目09", "mA"],
            ["项目10", "kHz"],
            ["项目11", "℃"],
            ["项目12", "V"],
        ]
        confidence = [[0.99] * 2 for _ in grid]

        repaired = ocr_backend._repair_periodic_unit_blanks(grid, confidence)

        self.assertEqual(repaired, [(4, 1, "V")])
        self.assertEqual(grid[4][1], "V")
        self.assertEqual(confidence[4][1], 0.77)

    def test_consistency_checks_preserve_visible_ascii_case_in_identifier_columns(self):
        grid = [
            ["设备编号", "型号", "名称", "单位"],
            ["a001", "sa-5000", "signal_a", "dbm"],
            ["task-002", "dso-x3104t", "mixedCase", "mhz"],
        ]
        confidence = [[0.99] * 4 for _ in grid]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[1], ["a001", "sa-5000", "signal_a", "dbm"])
        self.assertEqual(grid[2], ["task-002", "dso-x3104t", "mixedCase", "mhz"])

    def test_rectified_crop_quality_rejects_false_paper_edge_strip(self):
        ocr_backend._load_runtime()
        source = np.full((1000, 1400, 3), 255, dtype=np.uint8)
        false_strip = np.full((100, 1450, 3), 255, dtype=np.uint8)
        valid_crop = np.full((500, 1000, 3), 255, dtype=np.uint8)

        self.assertFalse(
            ocr_backend._rectified_crop_is_usable(
                source,
                false_strip,
                {"detected": True},
            )
        )
        self.assertTrue(
            ocr_backend._rectified_crop_is_usable(
                source,
                valid_crop,
                {"detected": True},
            )
        )

    def test_recognition_falls_back_to_full_image_and_borderless_layout(self):
        ocr_backend._load_runtime()
        image = np.full((100, 300, 3), 255, dtype=np.uint8)
        false_strip = np.full((10, 300, 3), 255, dtype=np.uint8)
        boxes = np.array(
            [
                [[10, 8], [80, 8], [80, 25], [10, 25]],
                [[160, 8], [230, 8], [230, 25], [160, 25]],
                [[10, 38], [80, 38], [80, 55], [10, 55]],
                [[160, 38], [230, 38], [230, 55], [160, 55]],
                [[10, 68], [80, 68], [80, 85], [10, 85]],
                [[160, 68], [230, 68], [230, 85], [160, 85]],
            ],
            dtype=float,
        )
        ocr_result = SimpleNamespace(
            boxes=boxes,
            txts=["编号", "频率", "A001", "515.221", "A002", "516.347"],
            scores=[0.99] * 6,
        )
        seen_shapes = []

        def engine(candidate):
            seen_shapes.append(candidate.shape)
            return ocr_result

        with tempfile.TemporaryDirectory() as directory:
            image_path = Path(directory) / "borderless.png"
            cv2.imencode(".png", image)[1].tofile(str(image_path))
            with (
                patch.object(table_pipeline, "extract_screen_grid", return_value=None),
                patch.object(
                    table_pipeline,
                    "prepare_image",
                    return_value=(false_strip, {"detected": True, "mode": "auto", "corners": []}),
                ),
                patch.object(table_pipeline, "extract_ruled_grid", return_value=None),
                patch.object(ocr_backend, "_engines", return_value=(engine, None)),
            ):
                result = ocr_backend._recognize(
                    {
                        "protocol": PROTOCOL_VERSION,
                        "action": "recognize",
                        "image_path": str(image_path),
                        "output_directory": directory,
                    }
                )

        self.assertEqual(seen_shapes, [image.shape])
        self.assertEqual(result["rectification"]["mode"], "full_fallback")
        self.assertEqual(result["rows"], 3)
        self.assertEqual(result["columns"], 2)
        self.assertIn("borderless spatial layout", result["engine"])
        self.assertEqual(result["cells"][2][1]["text"], "516.347")

    def test_recognition_keeps_screen_grid_in_original_image_coordinates(self):
        ocr_backend._load_runtime()
        image = np.full((60, 80, 3), 255, dtype=np.uint8)
        boxes = np.array([[[5, 5], [25, 5], [25, 20], [5, 20]]], dtype=float)
        ocr_result = SimpleNamespace(boxes=boxes, txts=["A1"], scores=[0.99])

        with tempfile.TemporaryDirectory() as directory:
            image_path = Path(directory) / "screen.png"
            cv2.imencode(".png", image)[1].tofile(str(image_path))
            with (
                patch.object(
                    table_pipeline,
                    "extract_screen_grid",
                    return_value=([0, 40, 80], [0, 30, 60], image.copy()),
                ),
                patch.object(
                    table_pipeline,
                    "prepare_image",
                    return_value=(image.copy(), {"detected": False, "mode": "auto"}),
                ) as prepare_image_mock,
                patch.object(ocr_backend, "_screen_grid_is_credible", return_value=True),
                patch.object(ocr_backend, "_engines", return_value=(lambda _: ocr_result, None)),
            ):
                result = ocr_backend._recognize(
                    {
                        "protocol": PROTOCOL_VERSION,
                        "action": "recognize",
                        "image_path": str(image_path),
                        "output_directory": directory,
                    }
                )

        prepare_image_mock.assert_not_called()
        self.assertEqual(result["rows"], 2)
        self.assertEqual(result["columns"], 2)
        self.assertEqual(result["rectification"]["mode"], "screen")

    def test_maximum_screen_recognition_uses_physical_cell_crops_below_320_cells(self):
        ocr_backend._load_runtime()
        image = np.full((60, 80, 3), 255, dtype=np.uint8)
        columns = [0, 20, 40, 60, 80]
        rows = [0, 20, 40, 60]
        recognized = [[f"R{row}C{column}" for column in range(4)] for row in range(3)]
        confidence = [[0.99 for _ in range(4)] for _ in range(3)]
        page_engine = Mock(side_effect=AssertionError("screen mode must not use page assignment"))

        with tempfile.TemporaryDirectory() as directory:
            image_path = Path(directory) / "screen-physical-cells.png"
            cv2.imencode(".png", image)[1].tofile(str(image_path))
            with (
                patch.object(
                    table_pipeline,
                    "assess_image_quality",
                    return_value={"sharpness": 1000.0, "issues": [], "issue_labels": []},
                ),
                patch.object(
                    table_pipeline,
                    "extract_screen_grid",
                    return_value=(columns, rows, image.copy()),
                ),
                patch.object(ocr_backend, "_screen_grid_is_credible", return_value=True),
                patch.object(ocr_backend, "_engines", return_value=(page_engine, None)),
                patch.object(
                    ocr_backend,
                    "_recognize_screen_grid_cells",
                    return_value=(recognized, confidence, [0.99]),
                ) as recognize_cells,
            ):
                result = ocr_backend._recognize(
                    {
                        "protocol": PROTOCOL_VERSION,
                        "action": "recognize",
                        "image_path": str(image_path),
                        "output_directory": directory,
                        "options": {"accuracy_mode": "maximum", "deadline_seconds": 0},
                    }
                )

        recognize_cells.assert_called_once()
        page_engine.assert_not_called()
        self.assertEqual(result["rows"], 3)
        self.assertEqual(result["columns"], 4)

    def test_consistency_checks_restore_celsius_in_temperature_context(self):
        grid = [["说明"], ["室温 25C；湿度 46%RH"]]
        confidence = [[0.99], [0.99]]

        ocr_backend._apply_consistency_checks(grid, confidence)

        self.assertEqual(grid[1][0], "室温 25℃；湿度 46%RH")
        self.assertEqual(confidence[1][0], 0.99)

    def test_maximum_recognition_keeps_dense_screen_grid_above_legacy_cell_limit(self):
        ocr_backend._load_runtime()
        image = np.full((60, 80, 3), 255, dtype=np.uint8)
        columns = list(range(24))
        rows = list(range(45))
        recognized = [["A" for _ in range(23)] for _ in range(44)]
        confidence = [[0.99 for _ in range(23)] for _ in range(44)]

        with tempfile.TemporaryDirectory() as directory:
            image_path = Path(directory) / "dense-screen.png"
            cv2.imencode(".png", image)[1].tofile(str(image_path))
            with (
                patch.object(
                    table_pipeline,
                    "assess_image_quality",
                    return_value={"sharpness": 1000.0, "issues": [], "issue_labels": []},
                ),
                patch.object(ocr_backend, "_has_photographic_background", return_value=True),
                patch.object(
                    ocr_backend,
                    "_dense_grid_is_axis_aligned_screen_capture",
                    return_value=True,
                ),
                patch.object(
                    ocr_backend,
                    "_photographic_ruled_grid_is_credible",
                    return_value=True,
                ),
                patch.object(
                    table_pipeline,
                    "extract_screen_grid",
                    return_value=(columns, rows, image.copy()),
                ),
                patch.object(table_pipeline, "prepare_image") as prepare_image,
                patch.object(ocr_backend, "_screen_grid_is_credible", return_value=True),
                patch.object(
                    ocr_backend,
                    "_should_use_dense_screen_page_grid",
                    return_value=False,
                ),
                patch.object(ocr_backend, "_engines", return_value=(object(), None)),
                patch.object(
                    ocr_backend,
                    "_recognize_screen_grid_cells",
                    return_value=(recognized, confidence, [0.99]),
                ) as recognize_cells,
            ):
                result = ocr_backend._recognize(
                    {
                        "protocol": PROTOCOL_VERSION,
                        "action": "recognize",
                        "image_path": str(image_path),
                        "output_directory": directory,
                        "options": {
                            "accuracy_mode": "maximum",
                            "deadline_seconds": 0,
                        },
                    }
                )

        recognize_cells.assert_called_once()
        prepare_image.assert_not_called()
        self.assertEqual(result["rows"], 44)
        self.assertEqual(result["columns"], 23)
        self.assertEqual(result["rectification"]["mode"], "screen")

    def test_screen_grid_recognition_uses_row_aligned_bands(self):
        ocr_backend._load_runtime()
        image = np.full((80, 40, 3), 255, dtype=np.uint8)
        boxes = np.array([[[4, 4], [24, 4], [24, 20], [4, 20]]], dtype=float)
        calls = []

        def engine(crop):
            calls.append(crop.shape)
            text = "上" if len(calls) == 1 else "下"
            return SimpleNamespace(boxes=boxes, txts=[text], scores=[0.98])

        recognize_grid = getattr(ocr_backend, "_recognize_screen_grid", lambda *args, **kwargs: ([], [], []))

        grid, confidence, scores = recognize_grid(
            image,
            [0, 40],
            [0, 20, 40, 60, 80],
            engine,
            rows_per_band=2,
        )

        self.assertEqual(calls, [(40, 40, 3), (40, 40, 3)])
        self.assertEqual(grid, [["上"], [""], ["下"], [""]])
        self.assertEqual(confidence, [[0.98], [0.0], [0.98], [0.0]])
        self.assertEqual(scores, [0.98, 0.98])

    def test_screen_grid_default_band_size_stays_below_dense_page_candidate_limit(self):
        default = inspect.signature(ocr_backend._recognize_screen_grid).parameters[
            "rows_per_band"
        ].default

        self.assertEqual(default, 28)

    def test_dense_single_line_grid_uses_batched_cell_recognition(self):
        ocr_backend._load_runtime()
        image = np.full((40, 80, 3), 255, dtype=np.uint8)
        cv2.putText(image, "A", (5, 15), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 0), 1)
        cv2.putText(image, "B", (45, 35), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 0), 1)
        engine = SimpleNamespace(
            text_rec=lambda request: SimpleNamespace(txts=["A", "B"], scores=[0.99, 0.97])
        )

        grid, confidence, scores = ocr_backend._recognize_dense_screen_grid(
            image, [0, 40, 80], [0, 20, 40], engine
        )

        self.assertEqual(grid, [["A", ""], ["", "B"]])
        self.assertEqual(confidence, [[0.99, 0.0], [0.0, 0.97]])
        self.assertEqual(scores, [0.99, 0.97])

    def test_verified_cell_recognition_blanks_disagreement_instead_of_guessing(self):
        ocr_backend._load_runtime()
        image = np.full((20, 40, 3), 255, dtype=np.uint8)
        cv2.putText(image, "90", (5, 15), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 0), 1)
        outputs = [
            SimpleNamespace(txts=["%06"], scores=[1.0]),
            SimpleNamespace(txts=["90%"], scores=[0.96]),
            SimpleNamespace(txts=["9O%"], scores=[0.95]),
        ]
        engine = SimpleNamespace(text_rec=lambda request: outputs.pop(0))

        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [0, 40],
            [0, 20],
            engine,
            verify=True,
        )

        self.assertEqual(grid, [[""]])
        self.assertEqual(confidence, [[-1.0]])

    def test_verified_cell_recognition_accepts_two_of_three_matching_variants(self):
        ocr_backend._load_runtime()
        image = np.full((20, 60, 3), 255, dtype=np.uint8)
        cv2.putText(image, "P1", (5, 15), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 0), 1)
        outputs = [
            SimpleNamespace(txts=["PI"], scores=[0.91]),
            SimpleNamespace(txts=["P1"], scores=[0.95]),
            SimpleNamespace(txts=["P1"], scores=[0.97]),
        ]
        engine = SimpleNamespace(text_rec=lambda request: outputs.pop(0))

        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [0, 60],
            [0, 20],
            engine,
            verify=True,
        )

        self.assertEqual(grid, [["P1"]])
        self.assertEqual(confidence, [[0.95]])

    def test_verified_cell_recognition_keeps_retry_input_three_channel(self):
        ocr_backend._load_runtime()
        image = np.full((20, 60, 3), 255, dtype=np.uint8)
        cv2.putText(image, "P1", (5, 15), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 0), 1)
        outputs = [
            SimpleNamespace(txts=["PI"], scores=[0.91]),
            SimpleNamespace(txts=["P1"], scores=[0.95]),
            SimpleNamespace(txts=["P1"], scores=[0.97]),
        ]
        input_shapes = []

        def recognize(request):
            input_shapes.extend(item.shape for item in request.img)
            return outputs.pop(0)

        grid, _, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [0, 60],
            [0, 20],
            SimpleNamespace(text_rec=recognize),
            verify=True,
        )

        self.assertEqual(grid, [["P1"]])
        self.assertTrue(all(len(shape) == 3 and shape[2] == 3 for shape in input_shapes))

    def test_weak_fast_consensus_is_corrected_by_original_source_views(self):
        ocr_backend._load_runtime()
        image = np.full((28, 100, 3), 255, dtype=np.uint8)
        cv2.putText(image, "OK", (8, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1)
        fast_outputs = [
            SimpleNamespace(txts=["合各"], scores=[0.80]),
            SimpleNamespace(txts=["合各"], scores=[0.81]),
            SimpleNamespace(
                txts=["合格", "合格", "合格"],
                scores=[0.98, 0.97, 0.99],
            ),
        ]
        medium_outputs = [
            SimpleNamespace(txts=["合各"], scores=[0.79]),
        ]
        engine = SimpleNamespace(
            fast_text_rec=lambda request: fast_outputs.pop(0),
            text_rec=lambda request: medium_outputs.pop(0),
        )

        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [0, 100],
            [0, 28],
            engine,
            verify=True,
            quality_image=image.copy(),
        )

        self.assertEqual(grid, [["合格"]])
        self.assertEqual(confidence, [[0.97]])
        self.assertFalse(fast_outputs)
        self.assertFalse(medium_outputs)

    def test_verified_cell_recognition_blanks_agreement_with_low_score(self):
        ocr_backend._load_runtime()
        image = np.full((20, 40, 3), 255, dtype=np.uint8)
        cv2.putText(image, "A", (5, 15), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 0), 1)
        outputs = [
            SimpleNamespace(txts=["A"], scores=[0.74]),
            SimpleNamespace(txts=["A"], scores=[0.76]),
        ]
        engine = SimpleNamespace(text_rec=lambda request: outputs.pop(0))

        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [0, 40],
            [0, 20],
            engine,
            verify=True,
        )

        self.assertEqual(grid, [[""]])
        self.assertEqual(confidence, [[-1.0]])

    def test_screen_grid_cells_uses_recognition_only_for_first_merged_row(self):
        image = np.full((60, 80, 3), 255, dtype=np.uint8)
        direct_grid = [["标", "题"], ["A", "B"], ["1", "2"]]
        direct_confidence = [[0.8, 0.8], [0.99, 0.99], [0.99, 0.99]]

        with patch.object(
            ocr_backend,
            "_recognize_dense_screen_grid",
            side_effect=[
                (direct_grid, direct_confidence, [0.8, 0.8, 0.99, 0.99]),
                ([["标题"]], [[0.98]], [0.98]),
            ],
        ) as recognizer:
            grid, confidence, scores = ocr_backend._recognize_screen_grid_cells(
                image,
                [0, 40, 80],
                [0, 20, 40, 60],
                object(),
            )

        self.assertEqual(recognizer.call_count, 2)
        self.assertEqual(grid, [["标题", ""], ["A", "B"], ["1", "2"]])
        self.assertEqual(confidence[0], [0.98, 0.0])
        self.assertEqual(scores[-1], 0.98)

    def test_screen_grid_cells_uses_original_layout_lines_for_regular_header(self):
        image = np.full((60, 80, 3), 255, dtype=np.uint8)
        layout_image = image.copy()
        cv2.line(layout_image, (40, 0), (40, 20), (100, 100, 100), 1)
        direct_grid = [["表头1", "表头2"], ["A", "B"], ["1", "2"]]
        direct_confidence = [[0.99, 0.99], [0.99, 0.99], [0.99, 0.99]]

        with patch.object(
            ocr_backend,
            "_recognize_dense_screen_grid",
            return_value=(direct_grid, direct_confidence, [0.99] * 6),
        ) as recognizer:
            grid, confidence, _ = ocr_backend._recognize_screen_grid_cells(
                image,
                [0, 40, 80],
                [0, 20, 40, 60],
                object(),
                layout_image=layout_image,
            )

        recognizer.assert_called_once()
        self.assertEqual(grid, direct_grid)
        self.assertEqual(confidence, direct_confidence)

    def test_screen_grid_cells_recovers_sparse_title_above_dense_header(self):
        ocr_backend._load_runtime()
        image = np.full((90, 280, 3), 255, dtype=np.uint8)
        layout_image = image.copy()
        for x in range(40, 280, 40):
            cv2.line(layout_image, (x, 0), (x, 30), (80, 80, 80), 2)
        direct_grid = [
            ["", "", "设备", "检记录", "", "", ""],
            ["编号", "设备名称", "频率", "功率", "温度", "状态", "备注"],
            ["1", "频谱仪", "515.2", "-10", "36.5", "正常", "—"],
        ]
        direct_confidence = [[0.0, 0.0, 0.91, 0.89, 0.0, 0.0, 0.0]] + [
            [0.99] * 7,
            [0.99] * 7,
        ]

        with patch.object(
            ocr_backend,
            "_recognize_dense_screen_grid",
            side_effect=[
                (direct_grid, direct_confidence, [0.9] * 16),
                ([["设备巡检记录"]], [[0.986]], [0.986]),
            ],
        ) as recognizer:
            grid, confidence, _ = ocr_backend._recognize_screen_grid_cells(
                image,
                list(range(0, 281, 40)),
                [0, 30, 60, 90],
                object(),
                layout_image=layout_image,
                quality_image=image,
            )

        self.assertEqual(recognizer.call_count, 2)
        self.assertEqual(grid[0], ["设备巡检记录", "", "", "", "", "", ""])
        self.assertEqual(confidence[0], [0.986, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0])

    def test_screen_grid_cells_recovers_fragmented_multilevel_title(self):
        ocr_backend._load_runtime()
        image = np.full((120, 400, 3), 255, dtype=np.uint8)
        direct_grid = [
            ["", "", "", "2026", "年第三季度", "设备运行", "充计表", "", "", ""],
            ["", "", "", "运行数据", "", "", "质量指标", "", "维护信息", ""],
            ["", "", "运行时长", "告警次数", "停机时长", "合格率", "平均温度", "信噪比", "上次维护", "负责人"],
            ["1", "测试设备-01", "154.5", "1", "0.35", "99.92%", "31.6", "22.6", "2026-07-02", "钱工"],
        ]
        direct_confidence = [[0.99 if value else 0.0 for value in row] for row in direct_grid]

        with patch.object(
            ocr_backend,
            "_recognize_dense_screen_grid",
            side_effect=[
                (direct_grid, direct_confidence, [0.99] * 25),
                ([["2026 年第三季度设备运行统计表"]], [[0.995]], [0.995]),
            ],
        ) as recognizer:
            grid, confidence, _ = ocr_backend._recognize_screen_grid_cells(
                image,
                list(range(0, 401, 40)),
                [0, 30, 60, 90, 120],
                object(),
                layout_image=image,
            )

        self.assertEqual(recognizer.call_count, 2)
        self.assertEqual(
            grid[0],
            ["2026 年第三季度设备运行统计表", "", "", "", "", "", "", "", "", ""],
        )
        self.assertEqual(confidence[0][0], 0.995)

    def test_merged_screen_cells_recover_title_and_summary_label(self):
        ocr_backend._load_runtime()
        image = np.full((90, 480, 3), 255, dtype=np.uint8)
        columns = [0, 80, 160, 240, 320, 400, 479]
        rows = [0, 30, 60, 89]
        for column in columns[1:-1]:
            cv2.line(image, (column, 30), (column, 60), (80, 80, 80), 1)
        for column in columns[4:-1]:
            cv2.line(image, (column, 60), (column, 89), (80, 80, 80), 1)
        cv2.putText(
            image,
            "TITLE",
            (155, 22),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.55,
            (20, 20, 20),
            1,
            cv2.LINE_AA,
        )
        cv2.putText(
            image,
            "TOTAL",
            (95, 82),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.55,
            (20, 20, 20),
            1,
            cv2.LINE_AA,
        )
        grid = [
            ["2024年", "公司产", "品销售", "据汇总表", "", ""],
            ["日期", "部门", "区域", "产品", "金额", "负责人"],
            ["", "计", "", "", "4,248", "1,139,377"],
        ]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]

        with patch.object(
            ocr_backend,
            "_recognize_dense_screen_grid",
            side_effect=[
                ([['2024年公司产品销售数据汇总表']], [[0.99]], [0.99]),
                ([["合计"]], [[0.98]], [0.98]),
            ],
        ) as recognizer:
            scores = ocr_backend._recover_verified_merged_screen_cells(
                image,
                image,
                columns,
                rows,
                grid,
                confidence,
                object(),
                None,
                True,
                True,
            )

        self.assertEqual(recognizer.call_count, 2)
        self.assertEqual(grid[0], ["2024年公司产品销售数据汇总表", "", "", "", "", ""])
        self.assertEqual(grid[2], ["合计", "", "", "", "4,248", "1,139,377"])
        self.assertEqual(scores, [0.99, 0.98])

    def test_leading_merged_row_reuses_strong_page_text_before_cell_models(self):
        ocr_backend._load_runtime()
        image = np.full((180, 480, 3), 245, dtype=np.uint8)
        columns = [0, 80, 160, 240, 320, 400, 479]
        rows = [0, 30, 60, 90, 120, 150, 179]
        for boundary in columns:
            cv2.line(image, (boundary, rows[1]), (boundary, rows[-1]), (60, 60, 60), 2)
        for boundary in rows:
            cv2.line(image, (columns[0], boundary), (columns[-1], boundary), (60, 60, 60), 2)
        grid = [
            ["设备", "巡检记录", "", "", "", ""],
            ["编号", "名称", "数量", "单位", "状态", "备注"],
            ["1", "设备甲", "12", "台", "完成", "无"],
            ["2", "设备乙", "13", "台", "完成", "无"],
            ["3", "设备丙", "14", "台", "复核", "无"],
            ["4", "设备丁", "15", "台", "完成", "无"],
        ]
        confidence = [[0.96 if value else 0.0 for value in row] for row in grid]
        page_output = SimpleNamespace(
            boxes=np.asarray(
                [[[150, 5], [330, 5], [330, 25], [150, 25]]],
                dtype=np.float32,
            ),
            txts=("设备巡检记录",),
            scores=(0.98,),
        )
        recovered_rows = set()

        with patch.object(
            ocr_backend,
            "_recognize_dense_screen_grid",
            side_effect=AssertionError("strong page evidence must avoid cell model reload"),
        ) as recognizer:
            scores = ocr_backend._recover_leading_merged_rows_from_cell_views(
                image,
                columns,
                rows,
                grid,
                confidence,
                object(),
                ocr_backend._RecognitionBudget(0),
                page_output=page_output,
                recovered_rows_out=recovered_rows,
            )

        recognizer.assert_not_called()
        self.assertEqual(grid[0], ["设备巡检记录", "", "", "", "", ""])
        self.assertEqual(confidence[0][0], 0.77)
        self.assertEqual(scores, [0.98])
        self.assertEqual(recovered_rows, {0})

    def test_leading_merged_row_without_page_proof_keeps_visible_cells(self):
        ocr_backend._load_runtime()
        image = np.full((180, 480, 3), 245, dtype=np.uint8)
        columns = [0, 80, 160, 240, 320, 400, 479]
        rows = [0, 30, 60, 90, 120, 150, 179]
        for boundary in columns:
            cv2.line(image, (boundary, rows[1]), (boundary, rows[-1]), (60, 60, 60), 2)
        for boundary in rows:
            cv2.line(image, (columns[0], boundary), (columns[-1], boundary), (60, 60, 60), 2)
        grid = [
            ["设备", "巡检记录", "", "", "", ""],
            ["编号", "名称", "数量", "单位", "状态", "备注"],
            ["1", "设备甲", "12", "台", "完成", "无"],
            ["2", "设备乙", "13", "台", "完成", "无"],
            ["3", "设备丙", "14", "台", "复核", "无"],
            ["4", "设备丁", "15", "台", "完成", "无"],
        ]
        confidence = [[0.96 if value else 0.0 for value in row] for row in grid]

        with patch.object(
            ocr_backend,
            "_recognize_dense_screen_grid",
            side_effect=AssertionError("无页面证据时不得切换模型猜合并标题"),
        ) as recognizer:
            scores = ocr_backend._recover_leading_merged_rows_from_cell_views(
                image,
                columns,
                rows,
                grid,
                confidence,
                object(),
                ocr_backend._RecognitionBudget(0),
                page_output=None,
            )

        recognizer.assert_not_called()
        self.assertEqual(grid[0], ["设备", "巡检记录", "", "", "", ""])
        self.assertEqual(scores, [])

    def test_merged_screen_cells_collapse_contained_metadata_fragment_without_models(self):
        ocr_backend._load_runtime()
        image = np.full((60, 480, 3), 245, dtype=np.uint8)
        columns = [0, 80, 160, 240, 320, 400, 479]
        rows = [0, 30, 59]
        grid = [
            ["部门：综合组 期间：2026-02", "", "", "26-02", "", ""],
            ["学号", "姓名", "语文", "数学", "总分", "等级"],
        ]
        confidence = [[0.96 if value else 0.0 for value in row] for row in grid]

        with patch.object(
            ocr_backend,
            "_recognize_dense_screen_grid",
            side_effect=AssertionError("contained duplicate must not reload OCR models"),
        ) as recognizer:
            scores = ocr_backend._recover_verified_merged_screen_cells(
                image,
                image,
                columns,
                rows,
                grid,
                confidence,
                object(),
                None,
                True,
                False,
            )

        recognizer.assert_not_called()
        self.assertEqual(
            grid[0],
            ["部门：综合组 期间：2026-02", "", "", "", "", ""],
        )
        self.assertEqual(confidence[0][0], 0.77)
        self.assertEqual(scores, [0.96, 0.96])

    def test_tight_merged_text_crop_removes_border_and_wide_whitespace(self):
        ocr_backend._load_runtime()
        image = np.full((70, 560, 3), (238, 246, 252), dtype=np.uint8)
        cv2.rectangle(image, (0, 0), (559, 69), (75, 75, 75), 2)
        cv2.putText(
            image,
            "TOTAL",
            (238, 47),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.9,
            (15, 15, 15),
            2,
            cv2.LINE_AA,
        )

        cropped = ocr_backend._tight_merged_text_crop(image)

        self.assertIsNotNone(cropped)
        self.assertLess(cropped.shape[1], image.shape[1] // 2)
        self.assertLess(cropped.shape[0], image.shape[0])
        self.assertGreater(cropped.shape[1], 45)

    def test_tight_merged_text_crop_rejects_empty_bordered_cell(self):
        ocr_backend._load_runtime()
        image = np.full((70, 560, 3), 245, dtype=np.uint8)
        cv2.rectangle(image, (0, 0), (559, 69), (75, 75, 75), 2)

        self.assertIsNone(ocr_backend._tight_merged_text_crop(image))

    def test_tight_merged_text_crop_rejects_extreme_noise_span(self):
        ocr_backend._load_runtime()
        image = np.full((70, 1600, 3), 245, dtype=np.uint8)
        cv2.rectangle(image, (30, 25), (45, 45), (20, 20, 20), -1)
        cv2.rectangle(image, (1550, 25), (1565, 45), (20, 20, 20), -1)

        self.assertIsNone(ocr_backend._tight_merged_text_crop(image))

    def test_merged_summary_uses_full_page_context_after_crop_hallucination(self):
        ocr_backend._load_runtime()
        image = np.full((60, 480, 3), 245, dtype=np.uint8)
        columns = [0, 80, 160, 240, 320, 400, 479]
        rows = [0, 30, 59]
        for column in columns[4:-1]:
            cv2.line(image, (column, 30), (column, 59), (80, 80, 80), 1)
        cv2.putText(
            image,
            "TOTAL",
            (95, 52),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.55,
            (20, 20, 20),
            1,
            cv2.LINE_AA,
        )
        grid = [
            ["日期", "部门", "区域", "产品", "金额", "负责人"],
            ["", "计", "", "", "4,248", "1,139,377"],
        ]
        confidence = [[0.95 if value else 0.0 for value in row] for row in grid]
        page_output = SimpleNamespace(
            boxes=np.asarray(
                [[[92, 35], [145, 35], [145, 56], [92, 56]]],
                dtype=np.float32,
            ),
            txts=("合计",),
            scores=(0.993,),
        )
        engine = Mock(return_value=page_output)

        with patch.object(
            ocr_backend,
            "_recognize_dense_screen_grid",
            return_value=([["북셕"]], [[0.999]], [0.999]),
        ):
            scores = ocr_backend._recover_verified_merged_screen_cells(
                image,
                image,
                columns,
                rows,
                grid,
                confidence,
                engine,
                None,
                True,
                True,
            )

        engine.assert_called_once_with(image)
        self.assertEqual(grid[1], ["合计", "", "", "", "4,248", "1,139,377"])
        self.assertEqual(scores, [0.993])

        reused_grid = [
            ["日期", "部门", "区域", "产品", "金额", "负责人"],
            ["", "计", "", "", "4,248", "1,139,377"],
        ]
        reused_confidence = [
            [0.95 if value else 0.0 for value in row] for row in reused_grid
        ]
        reused_engine = Mock(side_effect=AssertionError("page OCR must be reused"))
        with patch.object(
            ocr_backend,
            "_recognize_dense_screen_grid",
            return_value=([["북셕"]], [[0.999]], [0.999]),
        ):
            reused_scores = ocr_backend._recover_verified_merged_screen_cells(
                image,
                image,
                columns,
                rows,
                reused_grid,
                reused_confidence,
                reused_engine,
                None,
                True,
                True,
                page_output=page_output,
            )

        reused_engine.assert_not_called()
        self.assertEqual(
            reused_grid[1],
            ["合计", "", "", "", "4,248", "1,139,377"],
        )
        self.assertEqual(reused_scores, [0.993])

        deferred_grid = [
            ["日期", "部门", "区域", "产品", "金额", "负责人"],
            ["", "计", "", "", "4,248", "1,139,377"],
        ]
        deferred_confidence = [
            [0.95 if value else 0.0 for value in row] for row in deferred_grid
        ]
        deferred_engine = Mock(side_effect=AssertionError("fallback must be deferred"))
        with patch.object(
            ocr_backend,
            "_recognize_dense_screen_grid",
            return_value=([["북셕"]], [[0.999]], [0.999]),
        ):
            deferred_scores = ocr_backend._recover_verified_merged_screen_cells(
                image,
                image,
                columns,
                rows,
                deferred_grid,
                deferred_confidence,
                deferred_engine,
                None,
                True,
                True,
                allow_page_fallback=False,
            )

        deferred_engine.assert_not_called()
        self.assertEqual(deferred_grid[1][1], "计")
        self.assertEqual(deferred_scores, [])

    def test_screen_grid_title_uses_verified_cell_confidence_not_low_raw_variant(self):
        ocr_backend._load_runtime()
        image = np.full((90, 280, 3), 255, dtype=np.uint8)
        direct_grid = [
            ["", "", "设备", "巡检记录", "", "", ""],
            ["编号", "名称", "型号", "频率", "功率", "状态", "备注"],
            ["1", "频谱仪", "SA-5000", "515", "-10", "正常", "—"],
        ]
        direct_confidence = [[0.0, 0.0, 0.91, 0.89, 0.0, 0.0, 0.0]] + [
            [0.99] * 7,
            [0.99] * 7,
        ]

        with patch.object(
            ocr_backend,
            "_recognize_dense_screen_grid",
            side_effect=[
                (direct_grid, direct_confidence, [0.9] * 16),
                ([["设备巡检记录"]], [[0.986]], [0.31, 0.986]),
            ],
        ):
            grid, confidence, _ = ocr_backend._recognize_screen_grid_cells(
                image,
                list(range(0, 281, 40)),
                [0, 30, 60, 90],
                object(),
                layout_image=image,
                quality_image=image,
            )

        self.assertEqual(grid[0], ["设备巡检记录", "", "", "", "", "", ""])
        self.assertEqual(confidence[0][0], 0.986)

    def test_screen_grid_cells_never_merges_a_populated_header_row(self):
        image = np.full((60, 120, 3), 255, dtype=np.uint8)
        direct_grid = [["编号", "名称", "状态"], ["1", "设备", "正常"]]
        direct_confidence = [[0.99] * 3, [0.99] * 3]

        with patch.object(
            ocr_backend,
            "_recognize_dense_screen_grid",
            return_value=(direct_grid, direct_confidence, [0.99] * 6),
        ) as recognizer:
            grid, confidence, _ = ocr_backend._recognize_screen_grid_cells(
                image,
                [0, 40, 80, 120],
                [0, 30, 60],
                object(),
            )

        recognizer.assert_called_once()
        self.assertEqual(grid, direct_grid)
        self.assertEqual(confidence, direct_confidence)

    def test_dense_cell_path_only_applies_to_large_short_row_grids(self):
        ocr_backend._load_runtime()
        self.assertTrue(ocr_backend._is_dense_single_line_grid(list(range(23)), list(range(0, 44))))
        self.assertFalse(ocr_backend._is_dense_single_line_grid(list(range(10)), list(range(0, 20))))
        self.assertFalse(ocr_backend._is_dense_single_line_grid(list(range(23)), list(range(0, 1804, 41))))

    def test_non_axis_screen_grid_rejection_only_targets_pathological_dense_rows(self):
        ocr_backend._load_runtime()
        pathological = np.full((940, 420, 3), 255, dtype=np.uint8)
        pathological_columns = list(range(10, 411, 100))
        pathological_rows = list(range(10, 941, 10))
        for column in pathological_columns:
            cv2.line(pathological, (column, 10), (column, 939), (0, 0, 0), 1)
        for row in pathological_rows[::6]:
            cv2.line(pathological, (10, row), (410, row), (0, 0, 0), 1)
        self.assertTrue(
            ocr_backend._looks_like_pathological_dense_screen_grid(
                pathological,
                pathological_columns,
                pathological_rows,
            )
        )

        for row_count, column_count in ((45, 4), (60, 6)):
            ruled = np.full((row_count * 12 + 1, column_count * 80 + 1, 3), 255, dtype=np.uint8)
            columns = list(range(0, ruled.shape[1], 80))
            rows = list(range(0, ruled.shape[0], 12))
            for column in columns:
                cv2.line(ruled, (column, 0), (column, ruled.shape[0] - 1), (0, 0, 0), 2)
            for row in rows:
                cv2.line(ruled, (0, row), (ruled.shape[1] - 1, row), (0, 0, 0), 2)
            self.assertFalse(
                ocr_backend._looks_like_pathological_dense_screen_grid(
                    ruled,
                    columns,
                    rows,
                )
            )

        self.assertFalse(
            ocr_backend._looks_like_pathological_dense_screen_grid(
                np.full((701, 601, 3), 255, dtype=np.uint8),
                list(range(0, 601, 100)),
                list(range(0, 701, 100)),
            )
        )

    def test_near_identity_non_photo_page_bounds_unbounded_text_rechecks(self):
        near_identity = {
            "mode": "auto",
            "deskew_angle": -0.467,
        }
        self.assertTrue(
            ocr_backend._screen_like_page_verification_is_bounded(
                near_identity,
                photographic_background=False,
                source_shape=(851, 1462, 3),
                rectified_shape=(831, 1467, 3),
                output_count=206,
            )
        )
        self.assertFalse(
            ocr_backend._screen_like_page_verification_is_bounded(
                near_identity,
                photographic_background=True,
                source_shape=(851, 1462, 3),
                rectified_shape=(831, 1467, 3),
                output_count=206,
            )
        )
        self.assertTrue(
            ocr_backend._screen_like_page_verification_is_bounded(
                {
                    "mode": "full_fallback",
                    "rejected_non_axis_screen_grid": True,
                },
                photographic_background=False,
                source_shape=(851, 1462, 3),
                rectified_shape=(851, 1462, 3),
                output_count=206,
            )
        )
        self.assertFalse(
            ocr_backend._screen_like_page_verification_is_bounded(
                {"mode": "full_fallback"},
                photographic_background=False,
                source_shape=(851, 1462, 3),
                rectified_shape=(851, 1462, 3),
                output_count=206,
            )
        )
        self.assertFalse(
            ocr_backend._screen_like_page_verification_is_bounded(
                {"mode": "auto", "deskew_angle": 2.5},
                photographic_background=False,
                source_shape=(851, 1462, 3),
                rectified_shape=(831, 1467, 3),
                output_count=206,
            )
        )
        self.assertFalse(
            ocr_backend._screen_like_page_verification_is_bounded(
                near_identity,
                photographic_background=False,
                source_shape=(851, 1462, 3),
                rectified_shape=(600, 1000, 3),
                output_count=206,
            )
        )
    def test_dense_screen_page_route_requires_all_safety_gates(self):
        ocr_backend._load_runtime()
        image = np.full((430, 440, 3), 255, dtype=np.uint8)
        columns = list(range(0, 441, 20))
        rows = list(range(0, 431, 10))
        common = dict(
            maximum_accuracy=True,
            rectification_mode="screen",
            photographic_background=False,
        )
        with (
            patch.object(ocr_backend, "_screen_grid_is_credible", return_value=True),
            patch.object(
                ocr_backend,
                "_dense_grid_is_axis_aligned_screen_capture",
                return_value=True,
            ),
        ):
            self.assertTrue(
                ocr_backend._should_use_dense_screen_page_grid(
                    image, columns, rows, **common
                )
            )
            self.assertFalse(
                ocr_backend._should_use_dense_screen_page_grid(
                    image,
                    columns,
                    rows,
                    **{**common, "photographic_background": True},
                )
            )
            self.assertFalse(
                ocr_backend._should_use_dense_screen_page_grid(
                    image,
                    columns,
                    rows,
                    **{**common, "rectification_mode": "auto"},
                )
            )
        with patch.object(ocr_backend, "_screen_grid_is_credible", return_value=False):
            self.assertFalse(
                ocr_backend._should_use_dense_screen_page_grid(
                    image, columns, rows, **common
                )
            )

    def test_independent_page_text_columns_reject_collapsed_non_screen_grid(self):
        self.assertTrue(
            ocr_backend._page_grid_has_independent_column_collapse(
                3, 12, True, "auto"
            )
        )
        self.assertFalse(
            ocr_backend._page_grid_has_independent_column_collapse(
                3, 3, True, "auto"
            )
        )
        self.assertFalse(
            ocr_backend._page_grid_has_independent_column_collapse(
                3, 12, False, "auto"
            )
        )
        self.assertFalse(
            ocr_backend._page_grid_has_independent_column_collapse(
                3, 12, True, "screen"
            )
        )
        self.assertFalse(
            ocr_backend._page_grid_has_independent_column_collapse(
                6, 8, True, "auto"
            )
        )
        self.assertFalse(
            ocr_backend._page_grid_has_independent_column_collapse(
                6, 7, True, "auto"
            )
        )

    def test_collapsed_page_spatial_rows_restore_one_review_only_edge_row(self):
        ocr_backend._load_runtime()
        columns = 6
        anchors = [float(20 + index * 40) for index in range(columns)]
        table_rows = [[f"R{row}C{column}" for column in range(columns)] for row in range(10)]
        grid = [
            ["统计期间", "批次", "标题", "", "", ""],
            *table_rows,
            ["本页记录10条", "", "", "", "", ""],
        ]
        confidence = [
            [0.99 if value else 0.0 for value in row]
            for row in grid
        ]
        row_centers = [70.0] + [110.0 + row * 20.0 for row in range(10)] + [340.0]
        grouped_rows = [
            [
                {"center_x": anchors[column]}
                for column, value in enumerate(row)
                if value
            ]
            for row in grid
        ]
        recovered = ocr_backend._recover_collapsed_page_spatial_rows(
            grid,
            confidence,
            [],
            True,
            {
                "anchors": anchors,
                "row_centers": row_centers,
                "grouped_rows": grouped_rows,
            },
            list(range(100, 281, 20)),
        )
        self.assertIsNotNone(recovered)
        values, scores, metrics = recovered
        self.assertEqual((len(values), len(values[0])), (10, 6))
        self.assertEqual(values[-1][0], "R9C0")
        self.assertEqual(metrics["edge_rows_recovered"], 1)
        self.assertTrue(all(score <= 0.77 for row in scores for score in row if score))

    def test_collapsed_page_spatial_rows_reject_ambiguous_physical_row(self):
        ocr_backend._load_runtime()
        columns = 6
        anchors = [float(20 + index * 40) for index in range(columns)]
        table_row = [f"C{column}" for column in range(columns)]
        grid = [list(table_row) for _ in range(9)]
        confidence = [[0.99] * columns for _ in grid]
        geometry = {
            "anchors": anchors,
            "row_centers": [110.0, 111.0, 130.0, 150.0, 170.0, 190.0, 210.0, 230.0, 250.0],
            "grouped_rows": [
                [{"center_x": anchor} for anchor in anchors]
                for _ in grid
            ],
        }
        self.assertIsNone(
            ocr_backend._recover_collapsed_page_spatial_rows(
                grid,
                confidence,
                [],
                True,
                geometry,
                list(range(100, 261, 20)),
            )
        )

    def test_collapsed_page_spatial_rows_restore_regular_tail_before_footer(self):
        ocr_backend._load_runtime()
        columns = 8
        anchors = [float(20 + index * 40) for index in range(columns)]
        table_rows = [
            [f"R{row}C{column}" for column in range(columns)]
            for row in range(14)
        ]
        grid = [
            ["统计期间", "批次", "", "", "", "", "", ""],
            *table_rows,
            ["本页记录11条", "", "", "", "", "", "", ""],
        ]
        confidence = [
            [0.99 if value else 0.0 for value in row]
            for row in grid
        ]
        geometry = {
            "anchors": anchors,
            "row_centers": [70.0] + [110.0 + row * 20.0 for row in range(14)] + [410.0],
            "grouped_rows": [
                [
                    {"center_x": anchors[column]}
                    for column, value in enumerate(row)
                    if value
                ]
                for row in grid
            ],
        }
        recovered = ocr_backend._recover_collapsed_page_spatial_rows(
            grid,
            confidence,
            [],
            True,
            geometry,
            list(range(100, 261, 20)),
        )
        self.assertIsNotNone(recovered)
        values, scores, metrics = recovered
        self.assertEqual((len(values), len(values[0])), (14, 8))
        self.assertEqual(values[-1][0], "R13C0")
        self.assertEqual(metrics["edge_rows_recovered"], 6)
        self.assertTrue(all(score <= 0.77 for row in scores for score in row if score))

    def test_large_spatial_table_body_isolates_unique_dense_run(self):
        ocr_backend._load_runtime()
        columns = 8
        anchors = [float(20 + index * 40) for index in range(columns)]
        body = [
            [f"R{row}C{column}" for column in range(columns)]
            for row in range(26)
        ]
        grid = [
            ["文件编号", "填表日期", "", "", "", "", "", ""],
            ["项目名称", "负责人", "", "", "", "", "", ""],
            ["", "", "", "", "", "", "", ""],
            *body,
            ["记录条数", "", "", "", "", "", "", ""],
        ]
        confidence = [
            [0.99 if value else 0.0 for value in row]
            for row in grid
        ]
        geometry = {
            "anchors": anchors,
            "row_centers": [float(50 + index * 20) for index in range(len(grid))],
            "grouped_rows": [
                [
                    {"center_x": anchors[column]}
                    for column, value in enumerate(row)
                    if value
                ]
                for row in grid
            ],
        }
        recovered = ocr_backend._recover_large_spatial_table_body_for_review(
            grid,
            confidence,
            [],
            True,
            geometry,
        )
        self.assertIsNotNone(recovered)
        values, scores, metrics = recovered
        self.assertEqual((len(values), len(values[0])), (26, 8))
        self.assertEqual(metrics["body_first_row"], 3)
        self.assertEqual(metrics["_selected_rows"], list(range(3, 29)))
        self.assertEqual(metrics["_active_columns"], list(range(8)))
        self.assertTrue(all(score <= 0.77 for row in scores for score in row if score))

    def test_large_spatial_table_body_keeps_explicit_summary_footer(self):
        ocr_backend._load_runtime()
        columns = 8
        anchors = [float(20 + index * 40) for index in range(columns)]
        body = [
            [f"R{row}C{column}" for column in range(columns)]
            for row in range(26)
        ]
        footer = ["合计", "共26条", "", "", "", "", "", ""]
        grid = [
            ["文件编号", "填表日期", "", "", "", "", "", ""],
            ["", "", "", "", "", "", "", ""],
            *body,
            footer,
        ]
        confidence = [
            [0.99 if value else 0.0 for value in row]
            for row in grid
        ]
        geometry = {
            "anchors": anchors,
            "row_centers": [float(50 + index * 20) for index in range(len(grid))],
            "grouped_rows": [
                [
                    {"center_x": anchors[column]}
                    for column, value in enumerate(row)
                    if value
                ]
                for row in grid
            ],
        }

        recovered = ocr_backend._recover_large_spatial_table_body_for_review(
            grid,
            confidence,
            [],
            True,
            geometry,
        )

        self.assertIsNotNone(recovered)
        values, scores, metrics = recovered
        self.assertEqual((len(values), len(values[0])), (27, 8))
        self.assertEqual(values[-1], footer)
        self.assertEqual(metrics["summary_footer_row"], 28)
        self.assertTrue(all(score <= 0.77 for row in scores for score in row if score))

    def test_dense_leading_glyph_omission_risks_select_short_suffix_only(self):
        grid = [["位置"], *([["号线"]] * 5), *([["一号线"]] * 3), ["二车间"]]

        risks = ocr_backend._dense_leading_glyph_omission_risks(grid)

        self.assertEqual(risks, {(row, 0) for row in range(1, 6)})

    def test_dense_leading_glyph_omission_risks_select_truncation_chain(self):
        grid = [
            ["位置"],
            ["一车间"],
            ["车间"],
            ["车"],
            ["车"],
            ["A区"],
            ["B区"],
            ["C区"],
        ]

        risks = ocr_backend._dense_leading_glyph_omission_risks(grid)

        self.assertEqual(risks, {(2, 0), (3, 0), (4, 0)})

    def test_batched_multilevel_review_keeps_dense_blank_cells(self):
        ocr_backend._load_runtime()
        grid = [
            ["序号", "设备编号", "数值", "操作员"],
            ["1", "DEV-001", "10.100", "张伟"],
            ["2", "DEV-002", "20.200", "李娜"],
            ["3", "", "30.300", "王强"],
            ["4", "DEV-004", "", "赵敏"],
            ["5", "DEV-005", "50.500", ""],
        ]
        image = np.full((180, 400, 3), 255, dtype=np.uint8)
        cv2.rectangle(image, (125, 100), (175, 112), (0, 0, 0), -1)
        cv2.rectangle(image, (225, 130), (275, 142), (0, 0, 0), -1)

        selected = ocr_backend._batched_multilevel_review_cells(
            grid,
            image,
            [0, 100, 200, 300, 400],
            [0, 30, 60, 90, 120, 150, 180],
        )

        self.assertIn((3, 1), selected)
        self.assertIn((4, 2), selected)
        self.assertNotIn((5, 3), selected)
        self.assertLessEqual(len(selected), 64)

    def test_dense_leading_glyph_omission_risks_marks_incomplete_line_label(self):
        grid = [["位置"], *([["号线"]] * 5), ["二车间"], ["A区"]]

        risks = ocr_backend._dense_leading_glyph_omission_risks(grid)

        self.assertEqual(risks, {(row, 0) for row in range(1, 6)})

    def test_dense_repeated_token_risks_mark_inserted_thin_glyph(self):
        grid = [["区域"], *([["B区"]] * 3), *([["BI区"]] * 4), ["A区"]]

        risks, _ = ocr_backend._dense_repeated_token_residual_risks(grid)

        self.assertEqual(risks, {(row, 0) for row in range(4, 8)})

    def test_dense_repeated_token_risks_mark_prefixed_row_number(self):
        grid = [
            ["批次"],
            ["批次-L08"],
            ["批次-L08"],
            ["批次-L08"],
            ["13 批次-L08"],
            ["夜班"],
            ["待确认"],
            ["A区"],
        ]

        risks, _ = ocr_backend._dense_repeated_token_residual_risks(grid)

        self.assertEqual(risks, {(4, 0)})

    def test_structure_review_policy_marks_whole_table_for_edge_warning(self):
        grid = [["序号", "名称"], ["1", "设备A"]]
        confidence = [[0.99, 0.99], [0.99, 0.99]]

        reviewed = ocr_backend._apply_structure_review_policy(
            grid,
            confidence,
            structure_verified=True,
            structural_warnings=[
                "照片边缘证据尚未完成闭环，已限制为风险格复核并禁止直接发布。"
            ],
        )

        self.assertTrue(all(value <= 0.77 for row in reviewed for value in row))

    def test_spatial_shape_contradiction_rejects_severely_short_verified_grid(self):
        verified = [["X"] * 15 for _ in range(21)]
        spatial = [["X"] * 15 for _ in range(37)]
        spatial[0] = ["序号", "名称", "日期", "单位", *(["字段"] * 11)]

        self.assertTrue(
            ocr_backend._spatial_shape_strongly_contradicts_verified_grid(
                verified,
                spatial,
                [],
                True,
            )
        )

    def test_spatial_shape_contradiction_allows_one_row_difference(self):
        verified = [["X"] * 8 for _ in range(30)]
        spatial = [["X"] * 8 for _ in range(31)]
        spatial[0] = ["序号", "名称", "日期", "单位", "状态", "人员", "备注", "编号"]

        self.assertFalse(
            ocr_backend._spatial_shape_strongly_contradicts_verified_grid(
                verified,
                spatial,
                [],
                True,
            )
        )

    def test_spatial_shape_contradiction_rejects_combined_row_and_column_loss(self):
        verified = [["X"] * 4 for _ in range(18)]
        spatial = [["X"] * 3 for _ in range(20)]

        self.assertTrue(
            ocr_backend._spatial_shape_strongly_contradicts_verified_grid(
                verified,
                spatial,
                [],
                True,
            )
        )

    def test_dense_leading_zero_measurement_risks_excludes_identifiers(self):
        grid = [
            ["安全库存", "编号"],
            *[[str(1200 + row), f"0{row:02d}"] for row in range(8)],
            ["086", "009"],
        ]

        risks = ocr_backend._dense_leading_zero_measurement_risks(grid)

        self.assertEqual(risks, {(9, 0)})

    def test_dense_ordinal_sequence_risks_mark_missing_and_shifted_values(self):
        grid = [
            ["序号", "名称"],
            ["1", "A"],
            ["2", "B"],
            ["", "C"],
            ["4", "D"],
            ["5", "E"],
            ["6", "F"],
            ["7", "G"],
            ["8", "H"],
            ["0", "I"],
            ["10", "J"],
        ]

        risks = ocr_backend._dense_ordinal_sequence_risks(grid)

        self.assertEqual(risks, {(3, 0), (9, 0)})

    def test_dense_numeric_width_outlier_risks_mark_short_measurements(self):
        grid = [
            ["安全库存"],
            *[[str(1200 + row)] for row in range(8)],
            ["591"],
            ["699"],
        ]

        risks = ocr_backend._dense_numeric_width_outlier_risks(grid)

        self.assertEqual(risks, {(9, 0), (10, 0)})

    def test_misaligned_anchor_check_ignores_superscript_digit(self):
        grid = [
            ["序号", "名称", "状态"],
            ["1", "设备A", "正常"],
            ["²", "设备B", "正常"],
            ["3", "设备C", "正常"],
            ["4", "设备D", "正常"],
        ]

        self.assertFalse(ocr_backend._grid_has_misaligned_anchor_rows(grid))

    def test_dense_structured_token_format_risks_mark_batch_and_location_variants(self):
        grid = [
            ["批次", "区域"],
            ["批次-L08", "A区"],
            ["批次-L08", "B区"],
            ["批次-L08", "A区"],
            ["批次-1.08", "AI区"],
            ["批次-L.08", "AIK"],
            ["夜班", "B区"],
            ["待确认", "A区"],
        ]

        risks = ocr_backend._dense_structured_token_format_risks(grid)

        self.assertEqual(risks, {(4, 0), (5, 0), (4, 1), (5, 1)})

    def test_standalone_horizontal_mark_risks_preserve_but_review_marks(self):
        grid = [["值", "备注"], ["—", "正常"], ["-", ""], ["10", "完成"]]

        risks = ocr_backend._standalone_horizontal_mark_risks(grid)

        self.assertEqual(risks, {(1, 0), (2, 0)})

    def test_large_spatial_table_body_rejects_two_dense_runs(self):
        ocr_backend._load_runtime()
        columns = 8
        anchors = [float(20 + index * 40) for index in range(columns)]
        dense = [["X"] * columns for _ in range(20)]
        grid = [*dense, [""] * columns, *dense]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]
        geometry = {
            "anchors": anchors,
            "row_centers": [float(50 + index * 20) for index in range(len(grid))],
            "grouped_rows": [
                [{"center_x": anchor} for anchor in anchors]
                if any(row)
                else []
                for row in grid
            ],
        }
        self.assertIsNone(
            ocr_backend._recover_large_spatial_table_body_for_review(
                grid,
                confidence,
                [],
                True,
                geometry,
            )
        )

    def test_large_spatial_table_body_ignores_outer_metadata_only_columns(self):
        ocr_backend._load_runtime()
        columns = 13
        anchors = [float(20 + index * 40) for index in range(columns)]
        body = [
            [
                "",
                *[
                    f"R{row}C{column}"
                    if column < 11 or row <= 12
                    else ""
                    for column in range(12)
                ],
            ]
            for row in range(35)
        ]
        grid = [
            ["文件编号", "项目名称", "负责人", *([""] * 10)],
            ["", *([""] * 12)],
            *body,
            ["记录条数", "", "", *([""] * 10)],
        ]
        confidence = [
            [0.99 if value else 0.0 for value in row]
            for row in grid
        ]
        geometry = {
            "anchors": anchors,
            "row_centers": [float(50 + index * 20) for index in range(len(grid))],
            "grouped_rows": [
                [
                    {"center_x": anchors[column]}
                    for column, value in enumerate(row)
                    if value
                ]
                for row in grid
            ],
        }
        recovered = ocr_backend._recover_large_spatial_table_body_for_review(
            grid,
            confidence,
            [],
            True,
            geometry,
        )
        self.assertIsNotNone(recovered)
        values, scores, metrics = recovered
        self.assertEqual((len(values), len(values[0])), (35, 12))
        self.assertEqual(values[0][0], "R0C0")
        self.assertEqual(metrics["body_first_row"], 2)
        self.assertTrue(all(score <= 0.77 for row in scores for score in row if score))

        too_sparse = [list(row) for row in grid]
        for row in range(3, 2 + len(body)):
            too_sparse[row][-1] = ""
        too_sparse_geometry = dict(geometry)
        too_sparse_geometry["grouped_rows"] = [
            [
                {"center_x": anchors[column]}
                for column, value in enumerate(row)
                if value
            ]
            for row in too_sparse
        ]
        self.assertIsNone(
            ocr_backend._recover_large_spatial_table_body_for_review(
                too_sparse,
                [[0.99 if value else 0.0 for value in row] for row in too_sparse],
                [],
                True,
                too_sparse_geometry,
            )
        )

    def test_large_photo_page_consensus_requires_certificate_text_and_edge_gates(self):
        ocr_backend._load_runtime()
        image = np.full((600, 900, 3), 245, dtype=np.uint8)
        columns = list(range(0, 901, 50))
        rows = list(range(0, 601, 20))
        grid = [["A" for _ in range(18)] for _ in range(30)]
        confidence = [[0.99 for _ in range(18)] for _ in range(30)]
        certificate = ocr_backend._new_structure_certificate(
            image,
            columns,
            rows,
            "photographic_ruled_grid",
        )
        rectification = {
            "detected": True,
            "mode": "auto",
            "edge_completeness_checked": True,
            "corners": [[40, 40], [860, 40], [860, 560], [40, 560]],
        }
        common = dict(
            photographic_background=True,
            maximum_accuracy=True,
            source_shape=(700, 1000, 3),
        )
        with (
            patch.object(ocr_backend, "_screen_grid_is_credible", return_value=True),
            patch.object(
                ocr_backend, "_photographic_ruled_grid_is_credible", return_value=True
            ),
            patch.object(
                ocr_backend,
                "_dense_grid_is_axis_aligned_screen_capture",
                return_value=True,
            ),
            patch.object(
                ocr_backend,
                "_dense_ruled_grid_has_strong_physical_support",
                return_value=True,
            ),
            patch.object(
                ocr_backend, "_grid_has_misaligned_anchor_rows", return_value=False
            ),
        ):
            self.assertTrue(
                ocr_backend._certified_large_photo_page_consensus_is_safe(
                    image,
                    columns,
                    rows,
                    grid,
                    confidence,
                    certificate,
                    rectification,
                    **common,
                )
            )
            self.assertFalse(
                ocr_backend._certified_large_photo_page_consensus_is_safe(
                    image,
                    columns,
                    rows,
                    grid,
                    confidence,
                    certificate,
                    {**rectification, "edge_completeness_checked": False},
                    **common,
                )
            )
            low_confidence = [list(row) for row in confidence]
            for index in range(28):
                low_confidence[index // 18][index % 18] = 0.50
            self.assertFalse(
                ocr_backend._certified_large_photo_page_consensus_is_safe(
                    image,
                    columns,
                    rows,
                    grid,
                    low_confidence,
                    certificate,
                    rectification,
                    **common,
                )
            )

    def test_review_only_large_photo_consensus_keeps_edge_risk_but_bounds_review(self):
        ocr_backend._load_runtime()
        image = np.full((600, 900, 3), 245, dtype=np.uint8)
        columns = list(range(0, 901, 50))
        rows = list(range(0, 601, 20))
        grid = [["A" for _ in range(18)] for _ in range(30)]
        confidence = [[0.99 for _ in range(18)] for _ in range(30)]
        certificate = ocr_backend._new_structure_certificate(
            image,
            columns,
            rows,
            "photographic_ruled_grid",
        )
        rectification = {
            "detected": True,
            "mode": "auto",
            "edge_completeness_checked": False,
            "corners": [[40, 40], [860, 40], [860, 560], [40, 560]],
        }
        with (
            patch.object(ocr_backend, "_screen_grid_is_credible", return_value=True),
            patch.object(
                ocr_backend, "_photographic_ruled_grid_is_credible", return_value=True
            ),
            patch.object(
                ocr_backend,
                "_dense_ruled_grid_has_strong_physical_support",
                return_value=True,
            ),
            patch.object(
                ocr_backend, "_grid_preserves_detected_extent", return_value=True
            ),
            patch.object(ocr_backend, "_screen_grid_has_collapsed_rows", return_value=False),
            patch.object(
                ocr_backend, "_grid_has_concatenated_physical_rows", return_value=False
            ),
            patch.object(ocr_backend, "_grid_has_fused_physical_columns", return_value=False),
            patch.object(
                ocr_backend, "_grid_has_misaligned_anchor_rows", return_value=False
            ),
        ):
            self.assertTrue(
                ocr_backend._review_only_large_photo_page_consensus_is_safe(
                    image,
                    columns,
                    rows,
                    grid,
                    confidence,
                    certificate,
                    rectification,
                    photographic_background=True,
                    maximum_accuracy=True,
                )
            )
            weak = [list(row) for row in confidence]
            for index in range(28):
                weak[index // 18][index % 18] = 0.50
            self.assertFalse(
                ocr_backend._review_only_large_photo_page_consensus_is_safe(
                    image,
                    columns,
                    rows,
                    grid,
                    weak,
                    certificate,
                    rectification,
                    photographic_background=True,
                    maximum_accuracy=True,
                )
            )

    def test_review_only_photo_consensus_covers_240_to_499_cell_gap(self):
        ocr_backend._load_runtime()
        image = np.full((720, 900, 3), 245, dtype=np.uint8)
        columns = list(range(0, 901, 100))
        rows = list(range(0, 621, 20))
        grid = [["A" for _ in range(9)] for _ in range(31)]
        confidence = [[0.99 for _ in range(9)] for _ in range(31)]
        certificate = ocr_backend._new_structure_certificate(
            image, columns, rows, "photographic_ruled_grid"
        )
        rectification = {"mode": "auto"}
        with (
            patch.object(ocr_backend, "_screen_grid_is_credible", return_value=True),
            patch.object(
                ocr_backend, "_photographic_ruled_grid_is_credible", return_value=True
            ),
            patch.object(
                ocr_backend,
                "_dense_ruled_grid_has_strong_physical_support",
                return_value=True,
            ),
            patch.object(ocr_backend, "_grid_preserves_detected_extent", return_value=True),
            patch.object(ocr_backend, "_screen_grid_has_collapsed_rows", return_value=False),
            patch.object(
                ocr_backend, "_grid_has_concatenated_physical_rows", return_value=False
            ),
            patch.object(ocr_backend, "_grid_has_fused_physical_columns", return_value=False),
            patch.object(
                ocr_backend, "_grid_has_misaligned_anchor_rows", return_value=False
            ),
        ):
            self.assertTrue(
                ocr_backend._review_only_large_photo_page_consensus_is_safe(
                    image,
                    columns,
                    rows,
                    grid,
                    confidence,
                    certificate,
                    rectification,
                    photographic_background=True,
                    maximum_accuracy=True,
                )
            )
            self.assertFalse(
                ocr_backend._review_only_large_photo_page_consensus_is_safe(
                    image,
                    columns,
                    rows[:27],
                    grid[:26],
                    confidence[:26],
                    ocr_backend._new_structure_certificate(
                        image, columns, rows[:27], "photographic_ruled_grid"
                    ),
                    rectification,
                    photographic_background=True,
                    maximum_accuracy=True,
                )
            )

    def test_strong_ruled_support_accepts_eight_columns_not_seven(self):
        image = np.full((441, 801, 3), 245, dtype=np.uint8)
        columns = list(range(0, 801, 100))
        rows = list(range(0, 441, 20))
        line_map = np.ones(image.shape[:2], dtype=np.uint8)
        with (
            patch.object(
                ocr_backend.pipeline,
                "_grid_maps",
                return_value=(line_map, line_map, line_map),
            ),
            patch.object(
                ocr_backend,
                "_vertical_rule_supports",
                return_value=[1.0] * len(columns),
            ),
        ):
            self.assertTrue(
                ocr_backend._dense_ruled_grid_has_strong_physical_support(
                    image, columns, rows
                )
            )
            self.assertFalse(
                ocr_backend._dense_ruled_grid_has_strong_physical_support(
                    image, columns[:-1], rows
                )
            )

    def test_photo_grid_accepts_only_a_faint_narrow_leading_ruler_gutter(self):
        ocr_backend._load_runtime()
        image = np.full((1308, 1076, 3), 245, dtype=np.uint8)
        columns = [0, 34, 137, 240, 344, 447, 550, 654, 757, 861, 965, 1075]
        rows = list(range(0, 1308, 29))
        supports = [1.0, 0.074, 0.42, 0.39, 0.44, 0.36, 0.41, 0.38, 0.43, 0.40, 0.37, 1.0]
        with (
            patch.object(ocr_backend, "_screen_grid_is_credible", return_value=True),
            patch.object(
                ocr_backend,
                "_vertical_rule_supports",
                return_value=supports,
            ),
        ):
            self.assertTrue(
                ocr_backend._photographic_ruled_grid_is_credible(
                    image, columns, rows
                )
            )
            supports[1] = 0.05
            self.assertFalse(
                ocr_backend._photographic_ruled_grid_is_credible(
                    image, columns, rows
                )
            )
            supports[1] = 0.40
            supports[4] = 0.05
            self.assertFalse(
                ocr_backend._photographic_ruled_grid_is_credible(
                    image, columns, rows
                )
            )

    def test_dense_grid_text_resolution_rejects_only_extreme_large_grid(self):
        ocr_backend._load_runtime()
        columns = [index * 63 for index in range(23)]
        extreme_rows = [index * 18 for index in range(44)]
        readable_rows = [index * 29 for index in range(44)]

        self.assertTrue(
            ocr_backend._dense_grid_text_resolution_requires_safe_rejection(
                columns,
                extreme_rows,
                sharpness=703.74,
            )
        )
        self.assertFalse(
            ocr_backend._dense_grid_text_resolution_requires_safe_rejection(
                columns,
                readable_rows,
                sharpness=703.74,
            )
        )
        self.assertFalse(
            ocr_backend._dense_grid_text_resolution_requires_safe_rejection(
                columns[:13],
                extreme_rows[:31],
                sharpness=703.74,
            )
        )
        self.assertFalse(
            ocr_backend._dense_grid_text_resolution_requires_safe_rejection(
                [0, 63, 126],
                [0, 18, 17],
                sharpness=703.74,
            )
        )
        self.assertFalse(
            ocr_backend._dense_grid_text_resolution_requires_safe_rejection(
                columns,
                extreme_rows,
                sharpness=1500.0,
            )
        )
        self.assertFalse(
            ocr_backend._dense_grid_text_resolution_requires_safe_rejection(
                list(range(24)),
                list(range(45)),
                sharpness=703.74,
            )
        )

    def test_vertical_crop_recovery_requires_adjacent_stable_strong_expansions(self):
        ocr_backend._load_runtime()
        source = np.full((800, 1200, 3), 245, dtype=np.uint8)
        rectified = np.full((390, 900, 3), 245, dtype=np.uint8)
        columns = list(range(0, 901, 50))
        base_rows = list(range(0, 391, 10))
        recovered_rows = list(range(0, 431, 10))
        base_grid = (columns, base_rows, rectified)
        recovered_grid = (columns, recovered_rows, rectified)
        baseline = {
            "columns": columns,
            "rows": base_rows,
            "height": rectified.shape[0],
            "line_support": [1.0] * len(base_rows),
            "intersection_support": [1.0] * len(base_rows),
        }
        expanded_evidence = {
            **baseline,
            "rows": recovered_rows,
            "height": 430,
            "line_support": [1.0] * len(recovered_rows),
            "intersection_support": [1.0] * len(recovered_rows),
        }
        metadata = {
            "detected": True,
            "mode": "auto",
            "corners": [[100, 150], [1100, 150], [1100, 650], [100, 650]],
        }
        expansion = (
            np.full((430, 900, 3), 245, dtype=np.uint8),
            np.eye(3, dtype=np.float32),
            np.asarray(metadata["corners"], dtype=np.float32),
        )
        with (
            patch.object(ocr_backend, "_crop_has_truncated_edge", return_value=True),
            patch.object(
                ocr_backend,
                "_vertical_document_rectification",
                side_effect=[expansion] * 8,
            ),
            patch.object(
                ocr_backend.pipeline,
                "extract_ruled_grid",
                side_effect=[recovered_grid] * 6,
            ),
            patch.object(
                ocr_backend,
                "_ruled_grid_edge_evidence",
                side_effect=[baseline] + [expanded_evidence] * 6,
            ),
            patch.object(
                ocr_backend,
                "_expanded_grid_preserves_base_and_adds_strong_rows",
                return_value=(True, True),
            ),
            patch.object(ocr_backend, "_grid_geometry_is_bounded", return_value=True),
            patch.object(
                ocr_backend, "_photographic_ruled_grid_is_credible", return_value=True
            ),
        ):
            recovered_image, grid, result_metadata, inconclusive = (
                ocr_backend._recover_vertically_truncated_photo_grid(
                    source,
                    rectified,
                    metadata,
                    base_grid,
                    maximum_cells=1280,
                )
            )
        self.assertIsNot(recovered_image, rectified)
        self.assertEqual((len(grid[1]) - 1, len(grid[0]) - 1), (43, 18))
        self.assertEqual(result_metadata["vertical_expansion_ratio"], 0.06)
        self.assertEqual(result_metadata["edge_rows_recovered"], 4)
        self.assertTrue(result_metadata["edge_completeness_checked"])
        self.assertFalse(inconclusive)

    def test_top_only_recovery_accepts_three_stable_single_row_expansions(self):
        ocr_backend._load_runtime()
        source = np.full((800, 1200, 3), 245, dtype=np.uint8)
        rectified = np.full((600, 900, 3), 245, dtype=np.uint8)
        columns = list(range(0, 901, 150))
        base_rows = list(range(0, 601, 30))
        recovered_rows = list(range(0, 631, 30))
        base_grid = (columns, base_rows, rectified)
        recovered_grid = (columns, recovered_rows, rectified)
        baseline = {
            "columns": columns,
            "rows": base_rows,
            "height": rectified.shape[0],
            "line_support": [1.0] * len(base_rows),
            "intersection_support": [1.0] * len(base_rows),
        }
        expanded_evidence = {
            **baseline,
            "rows": recovered_rows,
            "height": 630,
            "line_support": [1.0] * len(recovered_rows),
            "intersection_support": [1.0] * len(recovered_rows),
        }
        metadata = {
            "detected": True,
            "mode": "auto",
            "corners": [[100, 150], [1100, 150], [1100, 790], [100, 790]],
        }
        expansion = (
            np.full((630, 900, 3), 245, dtype=np.uint8),
            np.eye(3, dtype=np.float32),
            np.asarray(metadata["corners"], dtype=np.float32),
        )
        with (
            patch.object(ocr_backend, "_crop_has_truncated_edge", return_value=False),
            patch.object(
                ocr_backend,
                "_top_document_rectification",
                side_effect=[expansion] * 7,
            ),
            patch.object(
                ocr_backend.pipeline,
                "extract_ruled_grid",
                side_effect=[recovered_grid] * 7,
            ),
            patch.object(
                ocr_backend,
                "_ruled_grid_edge_evidence",
                side_effect=[baseline] + [expanded_evidence] * 7,
            ),
            patch.object(
                ocr_backend,
                "_expanded_grid_preserves_base_and_adds_strong_rows",
                return_value=(True, False),
            ),
            patch.object(ocr_backend, "_grid_geometry_is_bounded", return_value=True),
            patch.object(
                ocr_backend, "_photographic_ruled_grid_is_credible", return_value=True
            ),
        ):
            recovered_image, grid, result_metadata, inconclusive = (
                ocr_backend._recover_vertically_truncated_photo_grid(
                    source,
                    rectified,
                    metadata,
                    base_grid,
                    maximum_cells=1280,
                )
            )

        self.assertIsNot(recovered_image, rectified)
        self.assertEqual((len(grid[1]) - 1, len(grid[0]) - 1), (21, 6))
        self.assertEqual(result_metadata["vertical_expansion_ratio"], 0.03)
        self.assertTrue(result_metadata["top_only_expansion"])
        self.assertFalse(inconclusive)

    def test_bottom_source_edge_skips_unrecoverable_vertical_expansions(self):
        ocr_backend._load_runtime()
        source = np.full((800, 1200, 3), 245, dtype=np.uint8)
        rectified = np.full((600, 900, 3), 245, dtype=np.uint8)
        columns = list(range(0, 901, 150))
        rows = list(range(0, 601, 30))
        base_grid = (columns, rows, rectified)
        metadata = {
            "detected": True,
            "mode": "auto",
            "corners": [[100, 50], [1100, 50], [1100, 799], [100, 799]],
        }

        with (
            patch.object(ocr_backend, "_crop_has_truncated_edge", return_value=True),
            patch.object(
                ocr_backend,
                "_vertical_document_rectification",
                side_effect=AssertionError("pixels beyond the source frame cannot be recovered"),
            ) as expansion,
        ):
            recovered_image, grid, result_metadata, inconclusive = (
                ocr_backend._recover_vertically_truncated_photo_grid(
                    source,
                    rectified,
                    metadata,
                    base_grid,
                    maximum_cells=1280,
                )
            )

        expansion.assert_not_called()
        self.assertIs(recovered_image, rectified)
        self.assertIs(grid, base_grid)
        self.assertFalse(result_metadata["edge_completeness_checked"])
        self.assertTrue(result_metadata["unrecoverable_source_edge"])
        self.assertTrue(inconclusive)

    def test_visible_top_heading_restores_clipped_outer_boundary(self):
        ocr_backend._load_runtime()
        image = np.full((240, 400, 3), 245, dtype=np.uint8)
        cv2.putText(
            image,
            "TITLE",
            (150, 40),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.8,
            (20, 20, 20),
            2,
        )
        cv2.line(image, (0, 60), (399, 60), (30, 30, 30), 2)
        grid = (
            [0, 100, 200, 300, 399],
            [60, 100, 140, 180, 239],
            image.copy(),
        )

        restored = ocr_backend._restore_visible_top_interval_boundary(image, grid)

        self.assertEqual(restored[1], [0, 60, 100, 140, 180, 239])

    def test_empty_top_margin_does_not_create_a_heading_row(self):
        ocr_backend._load_runtime()
        image = np.full((240, 400, 3), 245, dtype=np.uint8)
        cv2.line(image, (0, 60), (399, 60), (30, 30, 30), 2)
        grid = (
            [0, 100, 200, 300, 399],
            [60, 100, 140, 180, 239],
            image.copy(),
        )

        restored = ocr_backend._restore_visible_top_interval_boundary(image, grid)

        self.assertIs(restored, grid)

    def test_top_only_recovery_accepts_stable_two_row_pair(self):
        ocr_backend._load_runtime()
        source = np.full((800, 1200, 3), 245, dtype=np.uint8)
        rectified = np.full((600, 900, 3), 245, dtype=np.uint8)
        columns = list(range(0, 901, 150))
        base_rows = list(range(0, 601, 30))
        recovered_rows = list(range(0, 661, 30))
        base_grid = (columns, base_rows, rectified)
        recovered_grid = (columns, recovered_rows, rectified)
        baseline = {
            "columns": columns,
            "rows": base_rows,
            "height": rectified.shape[0],
            "line_support": [1.0] * len(base_rows),
            "intersection_support": [1.0] * len(base_rows),
        }
        expanded_evidence = {
            **baseline,
            "rows": recovered_rows,
            "height": 660,
            "line_support": [1.0] * len(recovered_rows),
            "intersection_support": [1.0] * len(recovered_rows),
        }
        metadata = {
            "detected": True,
            "mode": "auto",
            "corners": [[100, 150], [1100, 150], [1100, 790], [100, 790]],
        }
        expansion = (
            np.full((660, 900, 3), 245, dtype=np.uint8),
            np.eye(3, dtype=np.float32),
            np.asarray(metadata["corners"], dtype=np.float32),
        )
        with (
            patch.object(ocr_backend, "_crop_has_truncated_edge", return_value=False),
            patch.object(
                ocr_backend,
                "_top_document_rectification",
                side_effect=[None, None, None, expansion, expansion],
            ),
            patch.object(
                ocr_backend.pipeline,
                "extract_ruled_grid",
                side_effect=[recovered_grid, recovered_grid],
            ),
            patch.object(
                ocr_backend,
                "_ruled_grid_edge_evidence",
                side_effect=[baseline, expanded_evidence, expanded_evidence],
            ),
            patch.object(
                ocr_backend,
                "_expanded_grid_preserves_base_and_adds_strong_rows",
                return_value=(True, False),
            ),
            patch.object(ocr_backend, "_grid_geometry_is_bounded", return_value=True),
            patch.object(
                ocr_backend, "_photographic_ruled_grid_is_credible", return_value=True
            ),
        ):
            recovered_image, grid, result_metadata, inconclusive = (
                ocr_backend._recover_vertically_truncated_photo_grid(
                    source,
                    rectified,
                    metadata,
                    base_grid,
                    maximum_cells=1280,
                )
            )

        self.assertIsNot(recovered_image, rectified)
        self.assertEqual((len(grid[1]) - 1, len(grid[0]) - 1), (22, 6))
        self.assertEqual(result_metadata["vertical_expansion_ratio"], 0.10)
        self.assertTrue(result_metadata["stable_two_row_top_expansion"])
        self.assertFalse(inconclusive)

    def test_vertical_crop_recovery_accepts_a_stable_clipped_edge_pair(self):
        ocr_backend._load_runtime()
        source = np.full((800, 1200, 3), 245, dtype=np.uint8)
        rectified = np.full((390, 900, 3), 245, dtype=np.uint8)
        columns = list(range(0, 901, 50))
        base_rows = list(range(0, 391, 10))
        recovered_rows = list(range(0, 411, 10))
        base_grid = (columns, base_rows, rectified)
        recovered_grid = (columns, recovered_rows, rectified)
        baseline = {
            "columns": columns,
            "rows": base_rows,
            "height": rectified.shape[0],
            "line_support": [1.0] * len(base_rows),
            "intersection_support": [1.0] * len(base_rows),
        }
        expanded_evidence = {
            **baseline,
            "rows": recovered_rows,
            "height": 410,
            "line_support": [1.0] * len(recovered_rows),
            "intersection_support": [1.0] * len(recovered_rows),
        }
        metadata = {
            "detected": True,
            "mode": "auto",
            "corners": [[100, 150], [1100, 150], [1100, 650], [100, 650]],
        }
        expansion = (
            np.full((410, 900, 3), 245, dtype=np.uint8),
            np.eye(3, dtype=np.float32),
            np.asarray(metadata["corners"], dtype=np.float32),
        )
        with (
            patch.object(ocr_backend, "_crop_has_truncated_edge", return_value=True),
            patch.object(
                ocr_backend,
                "_vertical_document_rectification",
                side_effect=[expansion] * 8,
            ),
            patch.object(
                ocr_backend.pipeline,
                "extract_ruled_grid",
                side_effect=[recovered_grid] * 8,
            ),
            patch.object(
                ocr_backend,
                "_ruled_grid_edge_evidence",
                side_effect=[baseline] + [expanded_evidence] * 8,
            ),
            patch.object(
                ocr_backend,
                "_expanded_grid_preserves_base_and_adds_strong_rows",
                return_value=(True, False),
            ),
            patch.object(ocr_backend, "_grid_geometry_is_bounded", return_value=True),
            patch.object(
                ocr_backend, "_photographic_ruled_grid_is_credible", return_value=True
            ),
        ):
            recovered_image, grid, result_metadata, inconclusive = (
                ocr_backend._recover_vertically_truncated_photo_grid(
                    source,
                    rectified,
                    metadata,
                    base_grid,
                    maximum_cells=1280,
                )
            )

        self.assertIsNot(recovered_image, rectified)
        self.assertEqual((len(grid[1]) - 1, len(grid[0]) - 1), (41, 18))
        self.assertEqual(result_metadata["vertical_expansion_ratio"], 0.02)
        self.assertTrue(result_metadata["stable_clipped_edge_pair"])
        self.assertTrue(result_metadata["edge_completeness_checked"])
        self.assertFalse(inconclusive)

    def test_vertical_expansion_can_complete_stable_missing_columns(self):
        ocr_backend._load_runtime()
        baseline = {
            "columns": [0, 20, 40],
            "rows": [0, 10, 20, 30],
            "height": 31,
            "line_support": [1.0] * 4,
            "intersection_support": [1.0] * 4,
        }
        candidate = {
            "columns": [0, 10, 20, 30, 40],
            "rows": [0, 10, 20, 30, 40],
            "height": 41,
            "line_support": [1.0] * 5,
            "intersection_support": [1.0] * 5,
        }
        corners = np.asarray(
            [[0, 0], [40, 0], [40, 30], [0, 30]], dtype=np.float32
        )

        preserved, strong = (
            ocr_backend._expanded_grid_preserves_base_and_adds_strong_rows(
                baseline,
                candidate,
                corners,
                np.eye(3, dtype=np.float32),
            )
        )

        self.assertTrue(preserved)
        self.assertTrue(strong)

    def test_vertical_expansion_allows_only_unmatched_outer_base_rows(self):
        ocr_backend._load_runtime()
        baseline_rows = list(range(0, 401, 10))
        baseline = {
            "columns": [0, 20, 40],
            "rows": baseline_rows,
            "height": 401,
            "line_support": [1.0] * len(baseline_rows),
            "intersection_support": [1.0] * len(baseline_rows),
        }
        candidate_rows = [0] + list(range(20, 411, 10)) + [420]
        candidate = {
            **baseline,
            "rows": candidate_rows,
            "height": 421,
            "line_support": [1.0] * len(candidate_rows),
            "intersection_support": [1.0] * len(candidate_rows),
        }
        corners = np.asarray(
            [[0, 10], [40, 10], [40, 410], [0, 410]], dtype=np.float32
        )

        preserved, strong = (
            ocr_backend._expanded_grid_preserves_base_and_adds_strong_rows(
                baseline, candidate, corners, np.eye(3, dtype=np.float32)
            )
        )
        self.assertTrue(preserved)
        self.assertTrue(strong)

        candidate["rows"] = [row for row in candidate_rows if row != 210]
        candidate["line_support"] = [1.0] * len(candidate["rows"])
        candidate["intersection_support"] = [1.0] * len(candidate["rows"])
        preserved, _ = ocr_backend._expanded_grid_preserves_base_and_adds_strong_rows(
            baseline, candidate, corners, np.eye(3, dtype=np.float32)
        )
        self.assertFalse(preserved)

    def test_edge_recovery_prefers_stronger_bounded_morphology_grid(self):
        ocr_backend._load_runtime()
        image = np.full((120, 180, 3), 255, dtype=np.uint8)
        current = ([0, 90, 179], [0, 40, 80, 119], image.copy())
        horizontal = np.zeros((120, 180), dtype=np.uint8)
        vertical = np.zeros((120, 180), dtype=np.uint8)

        with (
            patch.object(
                ocr_backend,
                "_photographic_ruled_grid_is_credible",
                side_effect=[False, True],
            ),
            patch.object(
                ocr_backend.pipeline,
                "_grid_maps",
                return_value=(horizontal, vertical, np.zeros_like(horizontal)),
            ),
            patch.object(
                ocr_backend.pipeline,
                "_line_centers",
                side_effect=[
                    [0, 45, 90, 135, 179],
                    [0, 40, 80, 119],
                ],
            ),
            patch.object(ocr_backend, "_grid_geometry_is_bounded", return_value=True),
        ):
            recovered = ocr_backend._stronger_morphology_grid_for_edge_recovery(
                image, current, maximum_cells=1280
            )

        self.assertEqual(recovered[0], [0, 45, 90, 135, 179])
        self.assertEqual(recovered[1], [0, 40, 80, 119])

    def test_column_completed_edge_recovery_checks_the_stable_eight_percent_pair(self):
        source = inspect.getsource(
            ocr_backend._recover_vertically_truncated_photo_grid
        )

        self.assertIn("morphology_columns_recovered", source)
        self.assertIn("(0.08, 0.085)", source)
        self.assertIn("(0.095, 0.10)", source)
        self.assertIn("morphology_columns_recovered and recovered_rows >= 2", source)

    def test_partial_screen_grid_yields_to_more_complete_perspective_geometry(self):
        ocr_backend._load_runtime()
        image = np.full((1000, 900, 3), 255, dtype=np.uint8)
        screen_grid = (
            [100, 260, 420, 580, 740, 899],
            [450, 490, 530, 570, 610, 650],
            image.copy(),
        )
        candidate_image = np.full((800, 900, 3), 255, dtype=np.uint8)
        candidate_grid = (
            [0, 150, 300, 450, 600, 750, 899],
            list(range(0, 801, 40)),
            candidate_image.copy(),
        )
        metadata = {
            "detected": True,
            "full_frame_perspective_grid": True,
            "corners": [[10, 100], [890, 90], [899, 900], [0, 910]],
        }

        with (
            patch.object(
                ocr_backend.pipeline,
                "prepare_image",
                return_value=(candidate_image, metadata),
            ),
            patch.object(
                ocr_backend.pipeline,
                "extract_ruled_grid",
                return_value=candidate_grid,
            ),
            patch.object(ocr_backend, "_grid_geometry_is_bounded", return_value=True),
            patch.object(
                ocr_backend,
                "_photographic_ruled_grid_is_credible",
                return_value=True,
            ),
        ):
            selected = (
                ocr_backend._more_complete_perspective_grid_than_partial_screen(
                    image,
                    screen_grid,
                    maximum_cells=1280,
                )
            )

        self.assertIsNotNone(selected)
        self.assertIs(selected[0], candidate_image)
        self.assertIs(selected[1], candidate_grid)

    def test_expanded_perspective_restores_symmetric_columns_and_trailing_row(self):
        ocr_backend._load_runtime()
        image = np.full((1002, 913, 3), 255, dtype=np.uint8)
        columns = [84, 256, 428, 600, 762]
        rows = [
            29,
            87,
            128,
            169,
            210,
            251,
            292,
            334,
            375,
            416,
            458,
            500,
            542,
            584,
            626,
            668,
            710,
            752,
            794,
            836,
            879,
            922,
            964,
        ]
        grid = (columns, rows, image.copy())

        restored = ocr_backend._restore_expanded_perspective_crop_edges(
            image,
            grid,
            expected_columns=6,
            base_rows=20,
        )

        self.assertEqual(restored[0], [0, *columns, 912])
        self.assertEqual(restored[1], [*rows, 1001])

    def test_vertical_crop_recovery_accepts_one_stable_supported_edge_row(self):
        ocr_backend._load_runtime()
        source = np.full((800, 1200, 3), 245, dtype=np.uint8)
        rectified = np.full((390, 900, 3), 245, dtype=np.uint8)
        columns = list(range(0, 901, 50))
        base_rows = list(range(0, 391, 10))
        recovered_rows = list(range(0, 401, 10))
        base_grid = (columns, base_rows, rectified)
        recovered_grid = (columns, recovered_rows, rectified)
        baseline = {
            "columns": columns,
            "rows": base_rows,
            "height": rectified.shape[0],
            "line_support": [1.0] * len(base_rows),
            "intersection_support": [1.0] * len(base_rows),
        }
        expanded_evidence = {
            **baseline,
            "rows": recovered_rows,
            "height": 400,
            "line_support": [1.0] * len(recovered_rows),
            "intersection_support": [1.0] * len(recovered_rows),
        }
        metadata = {
            "detected": True,
            "mode": "auto",
            "corners": [[100, 150], [1100, 150], [1100, 650], [100, 650]],
        }
        expansion = (
            np.full((400, 900, 3), 245, dtype=np.uint8),
            np.eye(3, dtype=np.float32),
            np.asarray(metadata["corners"], dtype=np.float32),
        )
        with (
            patch.object(ocr_backend, "_crop_has_truncated_edge", return_value=True),
            patch.object(
                ocr_backend,
                "_vertical_document_rectification",
                side_effect=[expansion] * 8,
            ),
            patch.object(
                ocr_backend.pipeline,
                "extract_ruled_grid",
                side_effect=[recovered_grid] * 6,
            ),
            patch.object(
                ocr_backend,
                "_ruled_grid_edge_evidence",
                side_effect=[baseline] + [expanded_evidence] * 6,
            ),
            patch.object(
                ocr_backend,
                "_expanded_grid_preserves_base_and_adds_strong_rows",
                return_value=(True, True),
            ),
            patch.object(ocr_backend, "_grid_geometry_is_bounded", return_value=True),
            patch.object(
                ocr_backend, "_photographic_ruled_grid_is_credible", return_value=True
            ),
        ):
            recovered_image, grid, result_metadata, inconclusive = (
                ocr_backend._recover_vertically_truncated_photo_grid(
                    source,
                    rectified,
                    metadata,
                    base_grid,
                    maximum_cells=1280,
                )
            )

        self.assertIsNot(recovered_image, rectified)
        self.assertEqual((len(grid[1]) - 1, len(grid[0]) - 1), (40, 18))
        self.assertEqual(result_metadata["vertical_expansion_ratio"], 0.06)
        self.assertEqual(result_metadata["edge_rows_recovered"], 1)
        self.assertTrue(result_metadata["edge_completeness_checked"])
        self.assertFalse(inconclusive)

    def test_vertical_crop_recovery_prefers_the_most_complete_stable_expansion(self):
        ocr_backend._load_runtime()
        source = np.full((800, 1200, 3), 245, dtype=np.uint8)
        rectified = np.full((390, 900, 3), 245, dtype=np.uint8)
        columns = list(range(0, 651, 50))
        base_rows = list(range(0, 391, 10))
        candidate_rows = [
            list(range(0, 411, 10)),
            list(range(0, 411, 10)),
            list(range(0, 431, 10)),
            list(range(0, 431, 10)),
            list(range(0, 421, 10)),
            list(range(0, 421, 10)),
        ]
        base_grid = (columns, base_rows, rectified)
        candidate_grids = [(columns, rows, rectified) for rows in candidate_rows]
        baseline = {
            "columns": columns,
            "rows": base_rows,
            "height": rectified.shape[0],
            "line_support": [1.0] * len(base_rows),
            "intersection_support": [1.0] * len(base_rows),
        }
        candidate_evidence = [
            {
                **baseline,
                "rows": rows,
                "height": rows[-1],
                "line_support": [1.0] * len(rows),
                "intersection_support": [1.0] * len(rows),
            }
            for rows in candidate_rows
        ]
        metadata = {
            "detected": True,
            "mode": "auto",
            "corners": [[100, 150], [1100, 150], [1100, 650], [100, 650]],
        }
        expansion = (
            rectified.copy(),
            np.eye(3, dtype=np.float32),
            np.asarray(metadata["corners"], dtype=np.float32),
        )
        with (
            patch.object(ocr_backend, "_crop_has_truncated_edge", return_value=True),
            patch.object(
                ocr_backend,
                "_vertical_document_rectification",
                side_effect=[expansion] * 7,
            ),
            patch.object(
                ocr_backend.pipeline,
                "extract_ruled_grid",
                side_effect=candidate_grids,
            ),
            patch.object(
                ocr_backend,
                "_ruled_grid_edge_evidence",
                side_effect=[baseline] + candidate_evidence,
            ),
            patch.object(
                ocr_backend,
                "_expanded_grid_preserves_base_and_adds_strong_rows",
                return_value=(True, True),
            ),
            patch.object(ocr_backend, "_grid_geometry_is_bounded", return_value=True),
            patch.object(
                ocr_backend, "_photographic_ruled_grid_is_credible", return_value=True
            ),
        ):
            recovered_image, grid, result_metadata, inconclusive = (
                ocr_backend._recover_vertically_truncated_photo_grid(
                    source,
                    rectified,
                    metadata,
                    base_grid,
                    maximum_cells=1280,
                )
            )

        self.assertIsNot(recovered_image, rectified)
        self.assertEqual((len(grid[1]) - 1, len(grid[0]) - 1), (43, 13))
        self.assertEqual(result_metadata["vertical_expansion_ratio"], 0.06)
        self.assertEqual(result_metadata["edge_rows_recovered"], 4)
        self.assertTrue(result_metadata["edge_completeness_checked"])
        self.assertFalse(inconclusive)

    def test_vertical_crop_recovery_keeps_base_when_expanded_shapes_are_unstable(self):
        ocr_backend._load_runtime()
        source = np.full((800, 1200, 3), 245, dtype=np.uint8)
        rectified = np.full((390, 900, 3), 245, dtype=np.uint8)
        columns = list(range(0, 901, 50))
        base_rows = list(range(0, 391, 10))
        base_grid = (columns, base_rows, rectified)
        candidate_rows = [
            list(range(0, 421, 10)),
            list(range(0, 431, 10)),
            list(range(0, 411, 10)),
            list(range(0, 441, 10)),
            list(range(0, 401, 10)),
            list(range(0, 451, 10)),
        ]
        candidate_grids = [(columns, rows, rectified) for rows in candidate_rows]
        baseline = {
            "columns": columns,
            "rows": base_rows,
            "height": rectified.shape[0],
            "line_support": [1.0] * len(base_rows),
            "intersection_support": [1.0] * len(base_rows),
        }
        candidate_evidence = [
            {
                **baseline,
                "rows": rows,
                "line_support": [1.0] * len(rows),
                "intersection_support": [1.0] * len(rows),
            }
            for rows in candidate_rows
        ]
        metadata = {
            "detected": True,
            "mode": "auto",
            "corners": [[100, 150], [1100, 150], [1100, 650], [100, 650]],
        }
        expansion = (rectified.copy(), np.eye(3, dtype=np.float32), np.asarray(metadata["corners"]))
        with (
            patch.object(ocr_backend, "_crop_has_truncated_edge", return_value=True),
            patch.object(ocr_backend, "_vertical_document_rectification", side_effect=[expansion] * 8),
            patch.object(
                ocr_backend.pipeline,
                "extract_ruled_grid",
                side_effect=candidate_grids,
            ),
            patch.object(
                ocr_backend,
                "_ruled_grid_edge_evidence",
                side_effect=[baseline] + candidate_evidence,
            ),
            patch.object(
                ocr_backend,
                "_expanded_grid_preserves_base_and_adds_strong_rows",
                return_value=(True, True),
            ),
            patch.object(ocr_backend, "_grid_geometry_is_bounded", return_value=True),
            patch.object(
                ocr_backend, "_photographic_ruled_grid_is_credible", return_value=True
            ),
        ):
            recovered_image, grid, result_metadata, inconclusive = (
                ocr_backend._recover_vertically_truncated_photo_grid(
                    source,
                    rectified,
                    metadata,
                    base_grid,
                    maximum_cells=1280,
                )
            )
        self.assertIs(recovered_image, rectified)
        self.assertIs(grid, base_grid)
        self.assertFalse(result_metadata["edge_completeness_checked"])
        self.assertTrue(inconclusive)

    def test_dense_screen_page_suggestions_fill_only_empty_cells_for_review(self):
        grid = [["A", ""], ["", "D"]]
        confidence = [[0.99, 0.0], [0.0, 0.99]]
        suggestions = [["WRONG", "B"], ["C", "WRONG"]]
        suggestion_confidence = [[0.99, 0.97], [0.89, 0.99]]

        applied = ocr_backend._apply_page_suggestions_to_empty_screen_cells(
            grid,
            confidence,
            suggestions,
            suggestion_confidence,
        )

        self.assertEqual(grid, [["A", "B"], ["", "D"]])
        self.assertEqual(confidence, [[0.99, 0.77], [0.0, 0.99]])
        self.assertEqual(applied, {(0, 1)})

    def test_dense_screen_page_risk_locations_select_only_weak_or_visible_misses(self):
        ocr_backend._load_runtime()
        image = np.full((60, 120, 3), 255, dtype=np.uint8)
        cv2.putText(
            image,
            "42",
            (84, 51),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.45,
            (0, 0, 0),
            1,
        )
        grid = [["A", "B", "C"], ["1", "", ""]]
        confidence = [[0.99, 0.81, 0.99], [0.99, 0.0, 0.0]]

        selected = ocr_backend._dense_screen_page_risk_locations(
            image,
            grid,
            confidence,
            [0, 40, 80, 120],
            [0, 30, 60],
        )

        self.assertEqual(selected, {(0, 1), (1, 2)})

    def test_dense_screen_page_risk_locations_respects_locked_merged_cells(self):
        ocr_backend._load_runtime()
        image = np.full((20, 40, 3), 255, dtype=np.uint8)
        cv2.putText(image, "X", (24, 15), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 0), 1)

        selected = ocr_backend._dense_screen_page_risk_locations(
            image,
            [["标题", ""]],
            [[0.99, 0.0]],
            [0, 20, 40],
            [0, 20],
            locked_cells={(0, 1)},
        )

        self.assertEqual(selected, set())

    def test_dense_numeric_sequence_risks_select_decimal_shift_and_runaway_tail(self):
        grid = [["序号", "T(ms)", "采样率"]]
        for number in range(1, 43):
            grid.append(
                [
                    str(number),
                    str(round(0.00012 * number, 8)).rstrip("0").rstrip("."),
                    str(round(82031.25 + 195.3125 * number, 4)).rstrip("0").rstrip("."),
                ]
            )
        grid[41][1] = "0.0492"
        grid[41][2] = "900025"

        selected = ocr_backend._dense_numeric_sequence_risks(grid)

        self.assertEqual(selected, {(41, 1), (41, 2)})

    def test_dense_numeric_sequence_risks_select_misread_record_number(self):
        grid = [
            ["告警维护记录", ""],
            ["基础信息", ""],
            ["序号", "告警编号"],
        ]
        for number in range(1, 23):
            value = "" if number == 9 else str(number)
            grid.append([value, f"AP-{number:03d}"])
        grid[21][0] = "11"

        self.assertEqual(ocr_backend._dense_numeric_sequence_risks(grid), {(21, 0)})

    def test_large_photo_safe_rejection_routes_numeric_sequence_risks(self):
        source = inspect.getsource(ocr_backend._recognize)

        self.assertIn(
            "numeric_sequence_cells = _dense_numeric_sequence_risks(page_grid)",
            source,
        )
        self.assertIn("(numeric_sequence_cells, False, False)", source)

    def test_dense_numeric_sequence_risks_ignore_legitimate_periodic_columns(self):
        grid = [["序号", "Tu", "Ts"]]
        for number in range(1, 43):
            grid.append(
                [
                    str(number),
                    str(round(0.0019 * (number % 9), 8)).rstrip("0").rstrip("."),
                    str(round(0.017 * (number % 7), 8)).rstrip("0").rstrip("."),
                ]
            )

        self.assertEqual(ocr_backend._dense_numeric_sequence_risks(grid), set())

    def test_numeric_unit_row_context_marks_only_decimal_place_conflict(self):
        grid = [
            ["编号", "占用带宽", "名称"],
            ["3", "57kHz", "lora 52.3k 62.5k"],
            ["4", "2526kHz", "gsmr 252.6k 200k"],
            ["5", "20.9kHz", "dqpsk 20.9k 25k"],
        ]

        self.assertEqual(
            ocr_backend._numeric_unit_row_context_risks(grid),
            {(2, 1), (2, 2)},
        )

    def test_dense_numeric_sequence_risks_tolerate_several_ocr_outliers(self):
        grid = [["序号", "T(ms)"]]
        for number in range(1, 43):
            grid.append(
                [
                    str(number),
                    str(round(0.00012 * number, 8)).rstrip("0").rstrip("."),
                ]
            )
        replacements = {
            3: "0.0006",
            9: "0.008",
            19: "02028",
            25: "003",
            41: "0.0492",
        }
        for row, value in replacements.items():
            grid[row][1] = value

        selected = ocr_backend._dense_numeric_sequence_risks(grid)

        self.assertEqual(selected, {(row, 1) for row in replacements})

    def test_spatial_numeric_sequence_outliers_verify_only_selected_cells(self):
        grid = [["序号", "名称"]]
        for number in range(1, 12):
            grid.append([str(number), f"项目{number}"])
        grid[9][0] = "6"
        confidence = [[0.77, 0.77] for _ in grid]
        columns = [0, 40, 120]
        rows = list(range(0, 20 * (len(grid) + 1), 20))

        with patch.object(
            ocr_backend,
            "_verify_motion_blurred_ruled_cells",
            return_value=[0.99],
        ) as verify:
            scores = ocr_backend._review_spatial_numeric_sequence_outliers(
                np.zeros((rows[-1], columns[-1], 3), dtype=np.uint8),
                grid,
                confidence,
                columns,
                rows,
                Mock(),
            )

        self.assertEqual(scores, [0.99])
        self.assertEqual(verify.call_args.kwargs["selected_cells"], {(9, 0)})
        self.assertEqual(verify.call_args.kwargs["view_mode"], "standard")
        self.assertFalse(verify.call_args.kwargs["isolated_recognition"])

    def test_dense_repeated_token_residuals_select_only_rare_near_variants(self):
        grid = [["字段一", "字段二"]]
        grid.extend(
            [
                ["批次-L08", "A区"],
                ["批次-L08", "A区"],
                ["批次-L08", "A区"],
                ["批次-L08", "A区"],
                ["批次-L.08", "区"],
                ["型号-X19", "B区"],
                ["型号-X20", "B区"],
            ]
        )

        selected, counts = ocr_backend._dense_repeated_token_residual_risks(grid)

        self.assertEqual(selected, {(5, 0), (5, 1)})
        self.assertEqual(counts["批次-L08"], 4)
        self.assertEqual(counts["A区"], 4)

    def test_repeated_token_raw_consensus_requires_common_exact_dual_view(self):
        counts = {"批次-L08": 8, "批次-L.08": 1}
        self.assertEqual(
            ocr_backend._select_repeated_token_raw_consensus(
                "批次-L.08", "批次-L08", 0.96, "批次-L08", 0.95, counts
            ),
            "批次-L08",
        )
        self.assertIsNone(
            ocr_backend._select_repeated_token_raw_consensus(
                "批次-L.08", "批次-L08", 0.96, "批次-L.08", 0.99, counts
            )
        )
        self.assertIsNone(
            ocr_backend._select_repeated_token_raw_consensus(
                "批次-L.08", "批次-L08", 0.93, "批次-L08", 0.99, counts
            )
        )

    def test_dense_repeated_text_residual_risks_select_rare_chinese_variants(self):
        grid = [["序号", "类别", "状态", "自由文本"]]
        categories = ["模拟信号", "数字信号", "雷达信号", "其他"]
        for number in range(1, 43):
            grid.append(
                [
                    str(number),
                    categories[(number - 1) % len(categories)],
                    "正常",
                    f"唯一说明{number}",
                ]
            )
        grid[4][1] = "富达信号"
        for row in range(7, 14):
            grid[row][2] = "正用"

        selected = ocr_backend._dense_repeated_text_residual_risks(grid)

        self.assertEqual(
            selected,
            {(4, 1)} | {(row, 2) for row in range(7, 14)},
        )

    def test_mark_dense_residual_reviews_preserves_text_and_marks_header(self):
        grid = [["序号", "新注"], ["1", "正常"], ["2", "正常"]]
        for number in range(3, 10):
            grid.append([str(number), "正常"])
        confidence = [[0.99, 0.99] for _ in grid]

        selected = ocr_backend._mark_dense_residual_reviews(
            grid,
            confidence,
            review_header=True,
        )

        self.assertEqual(grid[0], ["序号", "新注"])
        self.assertEqual(selected, {(0, 0), (0, 1)})
        self.assertEqual(confidence[0], [0.77, 0.77])

    def test_dense_screen_page_candidate_allows_bounded_misses_for_cell_recovery(self):
        columns = list(range(11))
        rows = list(range(11))
        accepted = [["A" for _ in range(10)] for _ in range(10)]
        rejected = [["A" for _ in range(10)] for _ in range(10)]
        for index in range(6):
            accepted[index // 10][index % 10] = ""
        for index in range(8):
            rejected[index // 10][index % 10] = ""

        self.assertTrue(
            ocr_backend._dense_screen_page_candidate_is_trusted(
                accepted,
                columns,
                rows,
            )
        )
        self.assertFalse(
            ocr_backend._dense_screen_page_candidate_is_trusted(
                rejected,
                columns,
                rows,
            )
        )

    def test_dense_screen_page_primary_skips_full_cell_recognition(self):
        ocr_backend._load_runtime()
        image = np.full((90, 120, 3), 255, dtype=np.uint8)
        columns = list(range(24))
        rows = list(range(45))
        suggestions = [["PAGE" for _ in range(23)] for _ in range(44)]
        suggestion_confidence = [[0.99 for _ in range(23)] for _ in range(44)]

        with tempfile.TemporaryDirectory() as directory:
            image_path = Path(directory) / "dense-screen.png"
            cv2.imencode(".png", image)[1].tofile(str(image_path))
            page_engine = Mock(
                return_value=SimpleNamespace(boxes=[], txts=[], scores=[])
            )
            with (
                patch.object(
                    table_pipeline,
                    "assess_image_quality",
                    return_value={"sharpness": 1000.0, "issues": [], "issue_labels": []},
                ),
                patch.object(ocr_backend, "_has_photographic_background", return_value=False),
                patch.object(
                    table_pipeline,
                    "extract_screen_grid",
                    return_value=(columns, rows, image.copy()),
                ),
                patch.object(ocr_backend, "_screen_grid_is_credible", return_value=True),
                patch.object(
                    ocr_backend,
                    "_should_use_dense_screen_page_grid",
                    return_value=True,
                ),
                patch.object(ocr_backend, "_engines", return_value=(page_engine, None)),
                patch.object(
                    table_pipeline,
                    "assign_ocr_to_grid",
                    return_value=(suggestions, suggestion_confidence),
                ),
                patch.object(
                    ocr_backend,
                    "_recognize_screen_grid_cells",
                    side_effect=AssertionError("trusted dense page must skip all-cell OCR"),
                ) as recognize_cells,
            ):
                result = ocr_backend._recognize(
                    {
                        "protocol": PROTOCOL_VERSION,
                        "action": "recognize",
                        "image_path": str(image_path),
                        "output_directory": directory,
                        "options": {"accuracy_mode": "maximum", "deadline_seconds": 0},
                    }
                )

        recognize_cells.assert_not_called()
        self.assertEqual(result["rows"], 44)
        self.assertEqual(result["columns"], 23)
        self.assertEqual(result["cells"][4][9]["text"], "PAGE")
        self.assertEqual(result["cells"][4][9]["confidence"], 0.99)
        self.assertFalse(result["cells"][4][9]["needs_review"])
        self.assertIn("certified dense screen page grid", result["engine"])

    def test_tight_text_crop_removes_blank_cell_margins(self):
        ocr_backend._load_runtime()
        image = np.full((40, 240, 3), 255, dtype=np.uint8)
        cv2.putText(image, "AP-0001", (60, 27), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (20, 20, 20), 1)

        cropped = ocr_backend._tight_text_crop(image)

        self.assertLess(cropped.shape[1], image.shape[1] * 0.7)
        self.assertGreater(cropped.shape[1], 40)

    def test_split_text_line_crops_finds_two_visible_lines(self):
        image = np.full((60, 180, 3), 255, dtype=np.uint8)
        cv2.putText(image, "LINE1", (5, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (20, 20, 20), 1)
        cv2.putText(image, "LINE2", (5, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (20, 20, 20), 1)

        lines = ocr_backend._split_text_line_crops(ocr_backend._tight_text_crop(image))

        self.assertEqual(len(lines), 2)

    def test_split_text_line_crops_keeps_compact_wrapped_lines_separate(self):
        ocr_backend._load_runtime()
        image = np.full((60, 220, 3), 255, dtype=np.uint8)
        for x in range(8, 190, 22):
            cv2.rectangle(image, (x, 5), (x + 7, 22), (20, 20, 20), -1)
            cv2.rectangle(image, (x, 26), (x + 7, 43), (20, 20, 20), -1)

        lines = ocr_backend._split_text_line_crops(ocr_backend._tight_text_crop(image))

        self.assertEqual(len(lines), 2)

    def test_split_text_line_crops_keeps_disconnected_symbol_parts_together(self):
        ocr_backend._load_runtime()
        image = np.full((60, 220, 3), 255, dtype=np.uint8)
        for x in range(8, 190, 22):
            cv2.rectangle(image, (x, 5), (x + 7, 22), (20, 20, 20), -1)
        cv2.circle(image, (12, 34), 1, (20, 20, 20), -1)
        cv2.line(image, (8, 38), (16, 38), (20, 20, 20), 2)
        cv2.circle(image, (12, 43), 1, (20, 20, 20), -1)
        cv2.circle(image, (28, 41), 3, (20, 20, 20), 1)

        lines = ocr_backend._split_text_line_crops(ocr_backend._tight_text_crop(image))

        self.assertEqual(len(lines), 2)
        self.assertGreaterEqual(lines[1].shape[0], 12)

    def test_source_line_layout_wins_when_grid_removal_adds_one_wide_fill_band(self):
        cleaned_lines = [
            np.full((22, 242, 3), 255, dtype=np.uint8),
            np.full((22, 81, 3), 255, dtype=np.uint8),
            np.full((22, 115, 3), 255, dtype=np.uint8),
        ]
        source_lines = [
            np.full((22, 81, 3), 255, dtype=np.uint8),
            np.full((22, 115, 3), 255, dtype=np.uint8),
        ]

        self.assertTrue(
            ocr_backend._source_line_layout_is_safer(
                cleaned_lines, source_lines, 278, True
            )
        )
        self.assertFalse(
            ocr_backend._source_line_layout_is_safer(
                cleaned_lines, source_lines, 278, False
            )
        )

    def test_verified_date_cell_ignores_only_isolated_border_rule_fragment(self):
        ocr_backend._load_runtime()
        image = np.full((70, 220, 3), 255, dtype=np.uint8)
        cv2.putText(image, "-", (5, 16), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (20, 20, 20), 1)
        cv2.putText(
            image,
            "2026-08-08",
            (5, 58),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.5,
            (20, 20, 20),
            1,
        )
        outputs = [
            SimpleNamespace(txts=["-", "2026-08-08"], scores=[0.91, 0.99]),
            SimpleNamespace(txts=["-", "2026-08-08"], scores=[0.90, 0.98]),
        ]
        engine = SimpleNamespace(text_rec=lambda request: outputs.pop(0))

        line_crops = [
            np.full((12, 40, 3), 255, dtype=np.uint8),
            np.full((18, 150, 3), 255, dtype=np.uint8),
        ]
        with patch.object(
            ocr_backend,
            "_split_text_line_crops",
            return_value=line_crops,
        ):
            grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
                image,
                [0, 220],
                [0, 70],
                engine,
                verify=True,
            )

        self.assertEqual(grid, [["2026-08-08"]])
        self.assertGreater(confidence[0][0], 0.0)

    def test_verified_cell_recognition_joins_independently_verified_lines(self):
        ocr_backend._load_runtime()
        image = np.full((60, 180, 3), 255, dtype=np.uint8)
        cv2.putText(image, "LINE1", (5, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (20, 20, 20), 1)
        cv2.putText(image, "LINE2", (5, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (20, 20, 20), 1)
        outputs = [
            SimpleNamespace(txts=["第一行", "第二行"], scores=[0.96, 0.95]),
            SimpleNamespace(txts=["第一行", "第二行"], scores=[0.94, 0.93]),
        ]
        engine = SimpleNamespace(text_rec=lambda request: outputs.pop(0))

        grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
            image,
            [0, 180],
            [0, 60],
            engine,
            verify=True,
        )

        self.assertEqual(grid, [["第一行\n第二行"]])
        self.assertEqual(confidence, [[0.93]])

    def test_maximum_accuracy_withholds_unresolved_narrow_line(self):
        ocr_backend._load_runtime()
        image = np.full((60, 220, 3), 255, dtype=np.uint8)
        cv2.putText(image, "X", (8, 34), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (20, 20, 20), 2)
        long_line = np.full((20, 180, 3), 255, dtype=np.uint8)
        narrow_line = np.full((14, 22, 3), 255, dtype=np.uint8)

        def recognizer(family, correct_when_isolated=True):
            wrong = {"mobile": ".", "medium": "0", "server": "-", "v4": "="}[family]

            def recognize(request):
                isolated = len(request.img) == 1
                texts = []
                for crop in request.img:
                    narrow = crop.shape[1] <= 100
                    texts.append("÷。" if isolated and narrow and correct_when_isolated else wrong if narrow else "处理说明")
                return SimpleNamespace(txts=texts, scores=[0.99] * len(texts))

            return recognize

        engine = SimpleNamespace(
            fast_text_rec=recognizer("mobile"),
            text_rec=recognizer("medium"),
            server_text_rec=recognizer("server"),
            v4_server_text_rec=recognizer("v4"),
        )
        with patch.object(
            ocr_backend,
            "_split_text_line_crops",
            return_value=[long_line, narrow_line],
        ), patch.object(ocr_backend, "_recover_horizontal_mark", return_value=""):
            grid, confidence, _ = ocr_backend._recognize_dense_screen_grid(
                image,
                [0, 220],
                [0, 60],
                engine,
                verify=True,
                quality_image=image,
                require_medium_consensus=True,
                unbounded_consensus=True,
            )

        self.assertEqual(grid, [[""]])
        self.assertLess(confidence[0][0], 0.0)

    def test_target_device_uses_small_page_and_medium_cell_cpu_models(self):
        with tempfile.TemporaryDirectory() as directory:
            with patch.dict(os.environ, {"OCR_TABLE_MODEL_DIR": directory}):
                model_directory = getattr(ocr_backend, "model_directory", lambda: Path())()
                build_parameters = getattr(ocr_backend, "build_ocr_parameters", lambda _: {})
                parameters = build_parameters(model_directory)
                fast_parameters = ocr_backend.build_fast_ocr_parameters(model_directory)
                server_parameters = ocr_backend.build_server_ocr_parameters(model_directory)

        self.assertEqual(model_directory, Path(directory).resolve())
        self.assertEqual(parameters["Det.engine_type"], EngineType.OPENVINO)
        self.assertEqual(parameters["Cls.engine_type"], EngineType.OPENVINO)
        self.assertEqual(parameters["Rec.engine_type"], EngineType.OPENVINO)
        self.assertEqual(parameters["Det.model_type"], ModelType.MEDIUM)
        self.assertEqual(parameters["Rec.model_type"], ModelType.MEDIUM)
        self.assertEqual(fast_parameters["Rec.model_type"], ModelType.SMALL)
        self.assertEqual(parameters["Det.ocr_version"], OCRVersion.PPOCRV6)
        self.assertEqual(parameters["Rec.ocr_version"], OCRVersion.PPOCRV6)
        self.assertEqual(parameters["Global.max_side_len"], 2400)
        self.assertEqual(parameters["Det.limit_side_len"], 1280)
        self.assertEqual(parameters["Cls.cls_batch_num"], 8)
        # Crop batches are memory-bounded for a 16 GB tablet.  This does not
        # remove any model or recognition view; it only amortizes OpenVINO
        # scheduling while keeping width-sorted batches bounded.
        self.assertEqual(parameters["Rec.rec_batch_num"], 8)
        self.assertEqual(server_parameters["Rec.rec_batch_num"], 8)
        self.assertEqual(parameters["EngineConfig.openvino.inference_num_threads"], 3)
        self.assertEqual(parameters["EngineConfig.openvino.performance_hint"], "LATENCY")
        self.assertEqual(parameters["EngineConfig.openvino.num_streams"], 1)
        self.assertEqual(parameters["Global.model_root_dir"], str(Path(directory).resolve()))
        self.assertEqual(parameters["Det.model_path"], str(Path(directory).resolve() / "PP-OCRv6_det_medium.xml"))
        self.assertEqual(parameters["Rec.model_path"], str(Path(directory).resolve() / "PP-OCRv6_rec_medium.xml"))
        self.assertEqual(
            fast_parameters["Rec.model_path"],
            str(Path(directory).resolve() / "PP-OCRv6_rec_small.xml"),
        )
        self.assertEqual(
            parameters["Cls.model_path"],
            str(Path(directory).resolve() / "ch_ppocr_mobile_v2.0_cls_mobile.xml"),
        )

    def test_model_directory_finds_project_models_from_copied_backend(self):
        with tempfile.TemporaryDirectory() as directory:
            project = Path(directory) / "ocr-table-tool"
            models = project / "runtime" / "models"
            models.mkdir(parents=True)
            copied_backend = project / "build" / "kit" / "src" / "gui" / "bin" / "backend" / "ocr_backend.py"
            python = project / ".venv" / "Scripts" / "python.exe"

            with patch.object(ocr_backend, "__file__", str(copied_backend)), patch.object(
                sys, "executable", str(python)
            ), patch.dict(os.environ, {"OCR_TABLE_MODEL_DIR": ""}):
                located = ocr_backend.model_directory()

        self.assertEqual(located, models.resolve())

    def test_missing_portable_model_fails_before_engine_start(self):
        validator = getattr(ocr_backend, "validate_model_files", None)
        self.assertTrue(callable(validator), "model validator is missing")
        with tempfile.TemporaryDirectory() as directory:
            with self.assertRaisesRegex(FileNotFoundError, "识别模型文件缺失.*PP-OCRv6_det_medium.xml"):
                validator(Path(directory))

    def test_setup_script_reuses_environment_and_propagates_native_failures(self):
        setup = (Path(__file__).resolve().parents[2] / "scripts" / "setup_backend.ps1").read_text(
            encoding="utf-8-sig"
        )

        self.assertIn("function Invoke-Checked", setup)
        self.assertIn("if (-not (Test-Path -LiteralPath $pythonExe))", setup)

    def test_package_script_collects_openvino_runtime(self):
        package_script = (Path(__file__).resolve().parents[2] / "scripts" / "package_backend.ps1").read_text(
            encoding="utf-8-sig"
        )

        self.assertIn("--collect-all openvino", package_script)
        self.assertIn("validate_model_files", package_script)
        self.assertIn('"PP-OCRv6_rec_small.xml"', package_script)
        self.assertIn('"PP-OCRv6_rec_small.bin"', package_script)
        self.assertNotIn('"ch_PP-OCRv5_rec_server.onnx"', package_script)

    def test_warmup_reuses_local_table_model_without_downloading(self):
        import warmup

        with tempfile.TemporaryDirectory() as directory:
            models = Path(directory)
            table_model = models / "slanet-plus.onnx"
            table_model.write_bytes(b"local-model")
            with patch.object(sys, "argv", ["warmup.py", "--model-dir", directory]), patch.object(
                warmup, "RapidOCR"
            ), patch.object(warmup, "RapidTable") as rapid_table, patch.object(
                warmup, "validate_model_files"
            ):
                self.assertEqual(warmup.main(), 0)

        rapid_table.assert_called_once()
        table_input = rapid_table.call_args.args[0]
        self.assertEqual(Path(table_input.model_dir_or_path), table_model.resolve())

    def test_backend_reports_startup_dependency_error_as_json(self):
        backend = Path(__file__).resolve().parents[1] / "ocr_backend.py"

        completed = subprocess.run(
            [sys.executable, "-S", str(backend)],
            input='{"protocol":1,"action":"health"}',
            capture_output=True,
            text=True,
            check=False,
        )

        response = json.loads(completed.stdout)
        self.assertEqual(response["protocol"], 1)
        self.assertEqual(response["status"], "error")
        self.assertIn("No module named", response["message"])

    def test_backend_json_is_ascii_safe_under_windows_code_page(self):
        backend = Path(__file__).resolve().parents[1] / "ocr_backend.py"
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "结果.xlsx"
            request = {
                "protocol": 1,
                "action": "export_xlsx",
                "request_id": 2718,
                "output_path": str(output),
                "cells": [[{"text": "中文", "confidence": 1.0}]],
                "spans": [],
            }
            environment = os.environ.copy()
            environment["PYTHONIOENCODING"] = "cp936"

            completed = subprocess.run(
                [sys.executable, str(backend)],
                input=json.dumps(request).encode("ascii", errors="backslashreplace"),
                capture_output=True,
                env=environment,
                check=False,
            )

            response = json.loads(completed.stdout.decode("ascii"))
            self.assertEqual(response["status"], "ok")
            self.assertEqual(response["request_id"], 2718)
            self.assertEqual(Path(response["output_path"]).name, "结果.xlsx")

    def test_backend_one_shot_echoes_request_id_on_error(self):
        backend = Path(__file__).resolve().parents[1] / "ocr_backend.py"
        request = {
            "protocol": 1,
            "action": "unsupported",
            "request_id": 314,
        }

        completed = subprocess.run(
            [sys.executable, str(backend)],
            input=json.dumps(request).encode("ascii"),
            capture_output=True,
            check=False,
        )

        response = json.loads(completed.stdout.decode("ascii"))
        self.assertNotEqual(completed.returncode, 0)
        self.assertEqual(response["status"], "error")
        self.assertEqual(response["request_id"], 314)

    def test_backend_reads_utf8_stdin_and_exports_complete_chinese_grid(self):
        backend = Path(__file__).resolve().parents[1] / "ocr_backend.py"
        rows = [
            ["编号", "频率", "信号类型", "备注"],
            ["1", "515.128MHz", "数字", "正常"],
            ["2", "516.347MHz", "模拟", "需复核"],
        ]
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "多行结果.xlsx"
            request = {
                "protocol": 1,
                "action": "export_xlsx",
                "output_path": str(output),
                "cells": [
                    [{"text": value, "confidence": 1.0} for value in row]
                    for row in rows
                ],
                "spans": [],
            }
            environment = os.environ.copy()
            environment["PYTHONIOENCODING"] = "cp936"
            environment["PYTHONUTF8"] = "0"

            completed = subprocess.run(
                [sys.executable, str(backend)],
                input=json.dumps(request, ensure_ascii=False).encode("utf-8"),
                capture_output=True,
                env=environment,
                check=False,
            )

            self.assertEqual(completed.returncode, 0, completed.stdout.decode("ascii", errors="replace"))
            try:
                workbook = load_workbook(output)
            except Exception as error:
                self.fail(f"生成的 XLSX 无法打开: {error}")
            sheet = workbook.active
            self.assertEqual(sheet.max_row, 3)
            self.assertEqual(sheet.max_column, 4)
            self.assertEqual(
                [[sheet.cell(row, column).value for column in range(1, 5)] for row in range(1, 4)],
                rows,
            )

    def test_parse_html_expands_rowspan_and_colspan(self):
        html = """
        <table>
          <tr><th rowspan="2">编号</th><th colspan="2">参数</th></tr>
          <tr><th>频率</th><th>功率</th></tr>
          <tr><td>1</td><td>515.221</td><td>-10</td></tr>
        </table>
        """

        grid, spans = parse_html_table(html)

        self.assertEqual(
            grid,
            [
                ["编号", "参数", ""],
                ["", "频率", "功率"],
                ["1", "515.221", "-10"],
            ],
        )
        self.assertEqual(
            spans,
            [
                {"row": 0, "column": 0, "row_span": 2, "column_span": 1},
                {"row": 0, "column": 1, "row_span": 1, "column_span": 2},
            ],
        )

    def test_build_result_is_rectangular_and_versioned(self):
        result = build_result(
            [["名称", "数值"], ["带宽"]],
            confidence=0.83,
            engine="test-engine",
            rectified_image="rectified.png",
        )

        self.assertEqual(result["protocol"], PROTOCOL_VERSION)
        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["rows"], 2)
        self.assertEqual(result["columns"], 2)
        self.assertEqual(result["cells"][1][1]["text"], "")
        self.assertEqual(result["cells"][0][0]["confidence"], 0.83)

    def test_validate_request_rejects_unknown_protocol(self):
        with self.assertRaisesRegex(ValueError, "protocol"):
            validate_request({"protocol": 99, "action": "recognize"})

    def test_write_xlsx_preserves_grid_and_merges(self):
        cells = [["编号", "参数", ""], ["", "频率", "功率"], ["1", "515.221", "-10"]]
        spans = [
            {"row": 0, "column": 0, "row_span": 2, "column_span": 1},
            {"row": 0, "column": 1, "row_span": 1, "column_span": 2},
        ]

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "result.xlsx"
            write_xlsx(output, cells, spans)
            workbook = load_workbook(output)
            sheet = workbook.active

            self.assertEqual(sheet["A1"].value, "编号")
            self.assertEqual(sheet["B2"].value, "频率")
            self.assertEqual(sheet["C3"].value, "-10")
            self.assertIn("A1:A2", {str(item) for item in sheet.merged_cells.ranges})
            self.assertIn("B1:C1", {str(item) for item in sheet.merged_cells.ranges})

    def test_write_xlsx_replace_failure_preserves_existing_target(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "existing.xlsx"
            original = b"existing-user-file"
            output.write_bytes(original)

            with patch.object(table_pipeline.os, "replace", side_effect=OSError("replace failed")):
                with self.assertRaisesRegex(OSError, "replace failed"):
                    write_xlsx(output, [["名称"], ["保留原文件"]])

            self.assertEqual(output.read_bytes(), original)
            self.assertEqual(
                list(Path(directory).glob(f".{output.name}.*.tmp.xlsx")),
                [],
            )

    def test_write_xlsx_rejects_merge_that_hides_content_or_review(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "unsafe.xlsx"
            with self.assertRaisesRegex(ValueError, "hide cell content"):
                write_xlsx(
                    output,
                    [["标题", "不能隐藏"]],
                    [{"row": 0, "column": 0, "row_span": 1, "column_span": 2}],
                )
            with self.assertRaisesRegex(ValueError, "review state"):
                write_xlsx(
                    output,
                    [[{"text": "标题"}, {"text": "", "needs_review": True}]],
                    [{"row": 0, "column": 0, "row_span": 1, "column_span": 2}],
                )

    def test_write_xlsx_sizes_chinese_and_wrapped_text_for_readability(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "readable.xlsx"
            write_xlsx(output, [["设备名称"], ["便携式频谱分析仪\n第二行说明"]])
            sheet = load_workbook(output).active

            self.assertGreaterEqual(sheet.column_dimensions["A"].width, 18)
            self.assertGreaterEqual(sheet.row_dimensions[2].height, 36)

    def test_write_xlsx_preserves_review_cells_as_yellow_with_comment(self):
        cells = [
            [
                {"text": "编号", "confidence": 0.99, "needs_review": False},
                {"text": "计划数量", "confidence": 0.77, "needs_review": True},
            ],
            [
                {"text": "A001", "confidence": 0.99, "needs_review": False},
                {"text": "722", "confidence": 0.77, "needs_review": True},
            ],
        ]

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "review.xlsx"
            write_xlsx(output, cells)
            sheet = load_workbook(output).active

            self.assertEqual(sheet["B2"].value, "722")
            self.assertEqual(sheet["B2"].fill.fill_type, "solid")
            self.assertEqual(sheet["B2"].fill.fgColor.rgb[-6:], "FFF4D6")
            self.assertIsNotNone(sheet["B2"].comment)
            self.assertIn("人工复核", sheet["B2"].comment.text)

    def test_write_xlsx_centers_and_emphasizes_merged_title(self):
        cells = [["设备巡检记录", "", ""], ["日期", "编号", "状态"]]
        spans = [
            {
                "row": 0,
                "column": 0,
                "row_span": 1,
                "column_span": 3,
                "role": "title",
            }
        ]

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "title.xlsx"
            write_xlsx(output, cells, spans)
            sheet = load_workbook(output).active

            self.assertIn("A1:C1", {str(item) for item in sheet.merged_cells.ranges})
            self.assertEqual(sheet["A1"].alignment.horizontal, "center")
            self.assertEqual(sheet["A1"].alignment.vertical, "center")
            self.assertTrue(sheet["A1"].font.bold)
            self.assertGreaterEqual(sheet["A1"].font.sz, 16)
            self.assertGreaterEqual(sheet.row_dimensions[1].height, 28)

    def test_write_xlsx_does_not_clip_long_wrapped_text_at_four_lines(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "long-note.xlsx"
            note = "需要人工复核阴影反光倾斜模糊褶皱干扰下的文字数字单位和空白单元格" * 8
            write_xlsx(output, [["说明"], [note]])
            sheet = load_workbook(output).active

            self.assertGreater(sheet.row_dimensions[2].height, 72)
            self.assertLessEqual(sheet.row_dimensions[2].height, 396)

    def test_result_can_be_encoded_as_utf8_json(self):
        payload = build_result([["中文", "QPSK"]], confidence=0.9, engine="test")
        encoded = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.assertIn("中文".encode("utf-8"), encoded)

    def test_rectify_table_image_finds_perspective_grid(self):
        image = np.full((720, 1000, 3), 235, dtype=np.uint8)
        top_left = np.array([170, 160])
        top_right = np.array([850, 205])
        bottom_right = np.array([790, 560])
        bottom_left = np.array([220, 525])
        corners = [top_left, top_right, bottom_right, bottom_left]
        for index in range(4):
            cv2.line(image, corners[index], corners[(index + 1) % 4], (25, 25, 25), 4)
        for ratio in (0.2, 0.4, 0.6, 0.8):
            left = (top_left * (1 - ratio) + bottom_left * ratio).astype(int)
            right = (top_right * (1 - ratio) + bottom_right * ratio).astype(int)
            cv2.line(image, left, right, (50, 50, 50), 2)
        for ratio in (0.16, 0.33, 0.5, 0.67, 0.84):
            top = (top_left * (1 - ratio) + top_right * ratio).astype(int)
            bottom = (bottom_left * (1 - ratio) + bottom_right * ratio).astype(int)
            cv2.line(image, top, bottom, (50, 50, 50), 2)

        rectified, metadata = rectify_table_image(image)

        self.assertTrue(metadata["detected"])
        self.assertGreater(rectified.shape[1], rectified.shape[0])
        self.assertGreater(rectified.shape[1], 500)
        grid = extract_ruled_grid(rectified)
        self.assertIsNotNone(grid)
        columns, rows, _ = grid
        self.assertEqual(len(columns) - 1, 6)
        self.assertEqual(len(rows) - 1, 5)

    def test_rectify_keeps_complete_full_frame_perspective_grid(self):
        warped = np.full((720, 960, 3), 245, dtype=np.uint8)
        for x in np.linspace(0, warped.shape[1] - 1, 11).round().astype(int):
            cv2.line(warped, (int(x), 0), (int(x), 719), (35, 35, 35), 2)
        for y in np.linspace(0, warped.shape[0] - 1, 21).round().astype(int):
            cv2.line(warped, (0, int(y)), (959, int(y)), (35, 35, 35), 2)
        source = np.full((900, 1200, 3), 230, dtype=np.uint8)
        corners = np.array(
            [[120.0, 90.0], [1080.0, 110.0], [1070.0, 830.0], [130.0, 820.0]],
            dtype=np.float32,
        )
        affine = np.array([[1.0, 0.0, 0.0], [0.0, 1.0, 0.0]], dtype=np.float32)
        with patch.object(
            table_pipeline,
            "_warp_perspective_table",
            return_value=(warped, np.eye(3, dtype=np.float32), corners, False),
        ), patch.object(
            table_pipeline,
            "_deskew",
            return_value=(warped, affine, 0.0),
        ):
            rectified, metadata = rectify_table_image(source)

        self.assertEqual(rectified.shape[:2], warped.shape[:2])
        self.assertTrue(metadata["full_frame_perspective_grid"])
        grid = extract_ruled_grid(rectified, prefer_adaptive=True)
        self.assertIsNotNone(grid)
        columns, rows, _ = grid
        self.assertEqual(len(columns) - 1, 10)
        self.assertEqual(len(rows) - 1, 20)

    def test_full_frame_perspective_grid_rejects_inset_grid(self):
        image = np.full((720, 960, 3), 245, dtype=np.uint8)
        for x in np.linspace(120, 840, 9).round().astype(int):
            cv2.line(image, (int(x), 100), (int(x), 620), (35, 35, 35), 2)
        for y in np.linspace(100, 620, 17).round().astype(int):
            cv2.line(image, (120, int(y)), (840, int(y)), (35, 35, 35), 2)

        self.assertIsNone(table_pipeline._full_frame_perspective_grid(image))

    def test_rectification_keeps_full_source_when_incomplete_table_frame_touches_edges(self):
        image = np.full((716, 1515, 3), 255, dtype=np.uint8)
        columns = [0, 101, 286, 409, 532, 706, 1062, 1180, 1329, 1514]
        rows = list(range(2, 683, 34))
        for column in columns:
            cv2.line(image, (column, 0), (column, image.shape[0] - 1), (30, 30, 30), 2)
        for row in rows:
            cv2.line(image, (0, row), (image.shape[1] - 1, row), (30, 30, 30), 2)

        rectified, metadata = rectify_table_image(image)

        self.assertEqual(rectified.shape[:2], image.shape[:2])
        grid = extract_ruled_grid(rectified, prefer_adaptive=True)
        self.assertIsNotNone(grid)
        actual_columns, actual_rows, _ = grid
        self.assertEqual(len(actual_columns) - 1, 9)
        self.assertEqual(len(actual_rows) - 1, 21)

    def test_rectify_prefers_grid_over_large_rotated_border(self):
        image = np.full((800, 1200, 3), 240, dtype=np.uint8)
        distractor = cv2.boxPoints(((120, 420), (900, 280), 58)).astype(int)
        cv2.polylines(image, [distractor], True, (20, 20, 20), 7)
        cv2.rectangle(image, (390, 210), (1100, 620), (25, 25, 25), 4)
        for y in range(278, 620, 68):
            cv2.line(image, (390, y), (1100, y), (45, 45, 45), 2)
        for x in range(508, 1100, 118):
            cv2.line(image, (x, 210), (x, 620), (45, 45, 45), 2)

        rectified, metadata = rectify_table_image(image)

        self.assertTrue(metadata["detected"])
        self.assertGreater(rectified.shape[1], 650)
        self.assertLess(rectified.shape[0], 500)
        self.assertTrue(all(0 <= x < 1200 and 0 <= y < 800 for x, y in metadata["corners"]))

    def test_vertical_side_refinement_does_not_drop_a_narrow_last_column(self):
        image = np.full((900, 1600, 3), 255, dtype=np.uint8)
        corners = np.array(
            [[500.0, 120.0], [1200.0, 160.0], [1190.0, 780.0], [480.0, 820.0]],
            dtype=np.float32,
        )
        # A strong internal separator sits close enough to the right edge to
        # satisfy the old image-wide tolerance, but moving to it would crop a
        # genuine narrow final column.
        internal_line = np.array([[[1135, 156, 1125, 784]]], dtype=np.int32)
        with patch.object(cv2, "HoughLinesP", return_value=internal_line):
            refined = table_pipeline._refine_vertical_table_sides(image, corners)

        np.testing.assert_allclose(refined, corners)

    def test_vertical_side_refinement_rejects_large_inward_jump_to_internal_rule(self):
        image = np.full((1350, 2400, 3), 255, dtype=np.uint8)
        corners = np.array(
            [[45.0, 35.0], [2355.0, 35.0], [2355.0, 1310.0], [45.0, 1310.0]],
            dtype=np.float32,
        )
        internal_sides = np.array(
            [
                [[124, 40, 129, 1305]],
                [[2324, 40, 2320, 1305]],
            ],
            dtype=np.int32,
        )
        with patch.object(cv2, "HoughLinesP", return_value=internal_sides):
            refined = table_pipeline._refine_vertical_table_sides(image, corners)

        np.testing.assert_allclose(refined[[0, 3], 0], corners[[0, 3], 0])

    def test_table_candidate_score_prefers_dense_wide_grid_over_broad_frame(self):
        broad_frame = _table_candidate_score(535_000, 1.45, 0.065, 0.031)
        table_grid = _table_candidate_score(160_000, 2.47, 0.155, 0.031)

        self.assertGreater(table_grid, broad_frame)

    def test_table_candidate_score_accepts_dense_near_square_photo_table(self):
        dense_square = _table_candidate_score(680_000, 1.01, 0.081, 0.040)
        weak_square_frame = _table_candidate_score(680_000, 1.01, 0.018, 0.012)

        self.assertGreater(dense_square, 0.0)
        self.assertEqual(weak_square_frame, 0.0)

    def test_rectify_marks_dense_near_square_photo_table_for_spatial_ocr(self):
        image = np.full((900, 1200, 3), 245, dtype=np.uint8)
        left, top, right, bottom = 250, 40, 1050, 840
        cv2.rectangle(image, (left, top), (right, bottom), (25, 25, 25), 3)
        for row in range(1, 25):
            y = top + round((bottom - top) * row / 25)
            cv2.line(image, (left, y), (right, y), (45, 45, 45), 2)
        for column in range(1, 12):
            x = left + round((right - left) * column / 12)
            cv2.line(image, (x, top), (x, bottom), (45, 45, 45), 2)

        rectified, metadata = rectify_table_image(image)

        self.assertTrue(metadata["detected"])
        self.assertTrue(metadata["dense_near_square"])
        self.assertLess(abs(rectified.shape[1] - rectified.shape[0]), 100)

    def test_full_image_mode_keeps_original_extent(self):
        image = np.full((120, 200, 3), 180, dtype=np.uint8)

        prepared, metadata = prepare_image(image, "full")

        self.assertEqual(prepared.shape, image.shape)
        self.assertFalse(metadata["detected"])
        self.assertEqual(metadata["mode"], "full")

    def test_grid_crop_mode_ignores_large_light_decorative_frame(self):
        image = np.full((120, 200, 3), 180, dtype=np.uint8)
        expected = np.full((80, 120, 3), 220, dtype=np.uint8)
        with patch.object(
            table_pipeline,
            "rectify_table_image",
            return_value=(expected, {"detected": True, "corners": []}),
        ) as rectify:
            inner, metadata = prepare_image(image, "grid")

        rectify.assert_called_once_with(image, expand_to_document=False)
        self.assertIs(inner, expected)
        self.assertEqual(metadata["mode"], "grid")

    def test_empty_single_cell_grid_is_implausibly_sparse(self):
        self.assertTrue(ocr_backend._grid_is_implausibly_sparse([[""]]))

    def test_spatial_header_recovery_requires_dense_stable_body(self):
        grid = [
            ["生产日报", "", "", "", "", "", ""],
            ["基础信息", "", "", "", "测量与判定", "", ""],
            ["日期", "", "", "", "完成数量", "不良数量", "完成率"],
        ]
        for index in range(8):
            grid.append(
                [
                    f"2026-08-{index + 1:02d}",
                    "白班",
                    "A线",
                    "280",
                    "277",
                    "3",
                    "98.2%",
                ]
            )
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 7, "role": "title"}
        ]

        self.assertTrue(ocr_backend._spatial_header_is_safely_recoverable(grid, spans))
        grid[6][0] = ""
        grid[7][0] = ""
        grid[8][0] = ""
        self.assertFalse(ocr_backend._spatial_header_is_safely_recoverable(grid, spans))

    def test_two_column_false_ruled_grid_yields_to_complete_spatial_table(self):
        spatial = [
            ["项目任务跟踪表", "", "", "", "", "", ""],
            ["基础信息", "", "", "", "测量与判定", "", ""],
            ["任务编号", "任务名称", "优先级", "进度", "负责人", "计划完成", "当前状态"],
        ]
        for index in range(12):
            spatial.append(
                [
                    f"TASK-{index + 1:03d}",
                    "图像预处理",
                    "高",
                    "42%",
                    "王强",
                    "2026-08-07",
                    "进行中",
                ]
            )
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 7, "role": "title"}
        ]

        self.assertTrue(
            ocr_backend._collapsed_ruled_grid_can_yield_to_spatial(
                [0, 40, 480],
                list(range(18)),
                spatial,
                spans,
                True,
            )
        )
        self.assertFalse(
            ocr_backend._collapsed_ruled_grid_can_yield_to_spatial(
                [0, 40, 480],
                list(range(18)),
                spatial[:-5],
                spans,
                True,
            )
        )
        self.assertTrue(
            ocr_backend._collapsed_ruled_grid_can_yield_to_spatial(
                list(range(6)),
                list(range(16)),
                spatial,
                spans,
                True,
            )
        )
        self.assertFalse(
            ocr_backend._collapsed_ruled_grid_can_yield_to_spatial(
                list(range(7)),
                list(range(16)),
                spatial,
                spans,
                True,
            )
        )

    def test_health_request_reports_offline_engine(self):
        response = handle_request({"protocol": PROTOCOL_VERSION, "action": "health"})

        self.assertEqual(response["status"], "ok")
        self.assertEqual(response["protocol"], PROTOCOL_VERSION)
        self.assertTrue(response["offline"])
        self.assertEqual(response["action"], "health")
        self.assertTrue(response["models_valid"])
        self.assertTrue(response["capabilities"]["persistent"])

    def test_request_validation_rejects_invalid_recognition_options(self):
        with self.assertRaisesRegex(ValueError, "unsupported accuracy_mode"):
            validate_request(
                {
                    "protocol": PROTOCOL_VERSION,
                    "action": "recognize",
                    "image_path": "C:/input.png",
                    "options": {"accuracy_mode": "fast-and-risky"},
                }
            )

        with self.assertRaisesRegex(ValueError, "selected_table_region"):
            validate_request(
                {
                    "protocol": PROTOCOL_VERSION,
                    "action": "recognize",
                    "image_path": "C:/input.png",
                    "options": {"selected_table_region": "yes"},
                }
            )

    def test_request_validation_rejects_non_rectangular_export_grid(self):
        with self.assertRaisesRegex(ValueError, "rectangular"):
            validate_request(
                {
                    "protocol": PROTOCOL_VERSION,
                    "action": "export_xlsx",
                    "output_path": "C:/output.xlsx",
                    "cells": [["A", "B"], ["C"]],
                }
            )

    def test_error_response_has_stable_code_field_and_retryability(self):
        response = ocr_backend._error_response(
            ValueError("image does not exist: C:/missing.png"),
            {"action": "recognize", "image_path": "C:/missing.png"},
        )

        self.assertEqual(response["status"], "error")
        self.assertEqual(response["action"], "recognize")
        self.assertEqual(response["error_code"], "IMAGE_NOT_FOUND")
        self.assertEqual(response["field"], "image_path")
        self.assertFalse(response["retryable"])

    def test_export_request_writes_xlsx_from_cell_objects(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "from_request.xlsx"
            response = handle_request(
                {
                    "protocol": PROTOCOL_VERSION,
                    "action": "export_xlsx",
                    "output_path": str(output),
                    "cells": [[{"text": "编号", "confidence": 0.9}], [{"text": "1", "confidence": 0.8}]],
                    "spans": [],
                }
            )

            self.assertEqual(response["status"], "ok")
            self.assertTrue(output.exists())
            self.assertEqual(load_workbook(output).active["A2"].value, "1")

    def test_truncated_dense_crop_is_retained_only_with_complete_row_rules(self):
        ocr_backend._load_runtime()
        image = np.full((250, 700, 3), 245, dtype=np.uint8)
        horizontal = np.zeros((250, 700), dtype=np.uint8)
        for row in range(5, 246, 20):
            horizontal[row, :] = 255
        empty = np.zeros_like(horizontal)

        with patch.object(
            table_pipeline,
            "_grid_maps",
            return_value=(horizontal, empty, horizontal),
        ):
            self.assertTrue(
                ocr_backend._truncated_crop_has_complete_dense_row_boundaries(
                    image
                )
            )

        horizontal[245, :] = 0
        with patch.object(
            table_pipeline,
            "_grid_maps",
            return_value=(horizontal, empty, horizontal),
        ):
            self.assertFalse(
                ocr_backend._truncated_crop_has_complete_dense_row_boundaries(
                    image
                )
            )

    def test_blank_form_keeps_physical_empty_rows_and_recovers_merged_spans(self):
        ocr_backend._load_runtime()
        margin_x = 12
        margin_y = 12
        columns = [margin_x + value for value in (0, 100, 200, 300)]
        rows = [margin_y + value for value in ([0, 36, 128] + [152 + index * 24 for index in range(14)])]
        height, width = rows[-1] + margin_y + 1, columns[-1] + margin_x + 1
        image = np.full((height, width, 3), 245, dtype=np.uint8)
        horizontal = np.zeros((height, width), dtype=np.uint8)
        vertical = np.zeros((height, width), dtype=np.uint8)
        for boundary in rows:
            cv2.line(horizontal, (columns[0], boundary), (columns[-1], boundary), 255, 2)
        cv2.line(vertical, (columns[0], rows[0]), (columns[0], rows[-1]), 255, 2)
        cv2.line(vertical, (columns[-1], rows[0]), (columns[-1], rows[-1]), 255, 2)
        cv2.line(
            vertical,
            (columns[1], rows[0] + 4),
            (columns[1], rows[0] + int((rows[1] - rows[0]) * 0.65)),
            255,
            2,
        )
        for row in range(1, 10):
            for boundary in columns[1:-1]:
                cv2.line(
                    vertical,
                    (boundary, rows[row]),
                    (boundary, rows[row + 1]),
                    255,
                    2,
                )
        for row in range(10, 16):
            cv2.line(
                vertical,
                (columns[1], rows[row]),
                (columns[1], rows[row + 1]),
                255,
                2,
            )
        grid = [["", "徒步定位记录表", ""], ["频率", "编号", "裁判确认"]]
        grid.extend([["", "", ""] for _ in range(8)])
        grid.extend(
            [[label, "", ""] for label in ("出发时间", "返回时间", "比赛用时", "队员签字", "裁判员签字")]
        )
        grid.append(["", "", ""])
        confidence = [[0.99 if value else -1.0 for value in row] for row in grid]

        with patch.object(
            table_pipeline,
            "_grid_maps",
            return_value=(horizontal, vertical, cv2.bitwise_or(horizontal, vertical)),
        ):
            self.assertTrue(
                ocr_backend._blank_form_physical_grid_is_safe(
                    image,
                    columns,
                    rows,
                    grid,
                    photographic_background=False,
                )
            )
            trimmed = ocr_backend._trim_blank_form_outer_closure_rows(
                rows,
                grid,
                confidence,
            )
            self.assertIsNotNone(trimmed)
            rows, grid, confidence = trimmed
            self.assertEqual(len(grid), 15)
            self.assertTrue(
                ocr_backend._blank_form_physical_grid_is_safe(
                    image,
                    columns,
                    rows,
                    grid,
                    photographic_background=False,
                )
            )
            spans = ocr_backend._recover_blank_form_physical_spans(
                image,
                columns,
                rows,
                grid,
                confidence,
            )
            confidence = [
                [-1.0 if not value else 0.77 for value in values]
                for values in grid
            ]
            ink_row = 4
            ink_column = 1
            cv2.rectangle(
                image,
                (columns[ink_column] + 20, rows[ink_row] + 8),
                (columns[ink_column] + 60, rows[ink_row] + 24),
                (10, 10, 10),
                -1,
            )
            released = ocr_backend._release_confirmed_blank_form_empty_reviews(
                image,
                columns,
                rows,
                grid,
                confidence,
            )

        self.assertEqual((len(grid), len(grid[0])), (15, 3))
        self.assertEqual(grid[0], ["徒步定位记录表", "", ""])
        self.assertIn(
            {"row": 0, "column": 0, "row_span": 1, "column_span": 3, "role": "title"},
            spans,
        )
        for row in range(10, 15):
            self.assertIn(
                {"row": row, "column": 1, "row_span": 1, "column_span": 2, "role": "merged"},
                spans,
            )
        self.assertNotIn((ink_row, ink_column), released)
        self.assertLess(confidence[ink_row][ink_column], 0.0)
        self.assertIn((3, 0), released)
        self.assertEqual(confidence[3][0], 0.0)

        broken_horizontal = horizontal.copy()
        broken_horizontal[rows[6] - 2 : rows[6] + 3, :] = 0
        with patch.object(
            table_pipeline,
            "_grid_maps",
            return_value=(broken_horizontal, vertical, cv2.bitwise_or(broken_horizontal, vertical)),
        ):
            self.assertFalse(
                ocr_backend._blank_form_physical_grid_is_safe(
                    image,
                    columns,
                    rows,
                    grid,
                    photographic_background=False,
                )
            )

    def test_blank_form_collapses_only_weak_split_inside_tall_header(self):
        ocr_backend._load_runtime()
        columns = [8, 108, 208, 308]
        rows = [8, 64, 110, 152, 188, 216, 252, 288, 326, 358, 395, 433, 471, 509]
        image = np.full((518, 317, 3), 245, dtype=np.uint8)
        horizontal = np.zeros(image.shape[:2], dtype=np.uint8)
        vertical = np.zeros(image.shape[:2], dtype=np.uint8)
        for index, boundary in enumerate(rows):
            right = columns[0] + 60 if index == 2 else columns[-1]
            cv2.line(horizontal, (columns[0], boundary), (right, boundary), 255, 2)
        for boundary in columns:
            cv2.line(vertical, (boundary, rows[0]), (boundary, rows[-1]), 255, 2)
        grid = [
            ["维修验收记录表", "", ""],
            ["项目", "编号", "确认"],
            ["", "", ""],
            *[["", "", ""] for _ in range(6)],
            ["开始时间", "", ""],
            ["结束时间", "", ""],
            ["经办人签字", "", ""],
            ["负责人签字", "", ""],
        ]
        confidence = [[0.99 if value else 0.0 for value in row] for row in grid]

        with patch.object(
            table_pipeline,
            "_grid_maps",
            return_value=(horizontal, vertical, cv2.bitwise_or(horizontal, vertical)),
        ):
            collapsed = ocr_backend._collapse_blank_form_split_header_row(
                image, columns, rows, grid, confidence
            )

        self.assertIsNotNone(collapsed)
        collapsed_rows, collapsed_grid, collapsed_confidence = collapsed
        self.assertEqual(len(collapsed_grid), 12)
        self.assertEqual(collapsed_grid[1], ["项目", "编号", "确认"])
        self.assertEqual(collapsed_rows, [*rows[:2], *rows[3:]])
        self.assertEqual(len(collapsed_confidence), 12)

        strong_horizontal = horizontal.copy()
        cv2.line(
            strong_horizontal,
            (columns[0], rows[2]),
            (columns[-1], rows[2]),
            255,
            2,
        )
        with patch.object(
            table_pipeline,
            "_grid_maps",
            return_value=(strong_horizontal, vertical, cv2.bitwise_or(strong_horizontal, vertical)),
        ):
            self.assertIsNone(
                ocr_backend._collapse_blank_form_split_header_row(
                    image, columns, rows, grid, confidence
                )
            )

    def test_raw_blank_form_grid_recovers_body_only_column_and_ignores_dense_table(self):
        ocr_backend._load_runtime()
        columns = [8, 208, 424, 604, 851]
        rows = [8, 64, 152, 190, 228, 264, 292, 332, 369, 407, 445, 483, 521]
        image = np.full((529, 859, 3), 245, dtype=np.uint8)
        horizontal = np.zeros(image.shape[:2], dtype=np.uint8)
        vertical = np.zeros(image.shape[:2], dtype=np.uint8)
        for boundary in rows:
            cv2.line(horizontal, (columns[0], boundary), (columns[-1], boundary), 255, 2)
        for boundary in (columns[0], columns[-1]):
            cv2.line(vertical, (boundary, rows[0]), (boundary, rows[-1]), 255, 2)
        for row in range(1, 8):
            for boundary in columns[1:-1]:
                cv2.line(vertical, (boundary, rows[row]), (boundary, rows[row + 1]), 255, 2)
        for row in range(8, len(rows) - 1):
            cv2.line(vertical, (columns[1], rows[row]), (columns[1], rows[row + 1]), 255, 2)
        current = ([8, 208, 424, 851], rows, image.copy())

        with patch.object(
            table_pipeline,
            "_grid_maps",
            return_value=(horizontal, vertical, cv2.bitwise_or(horizontal, vertical)),
        ):
            recovered = ocr_backend._raw_blank_form_grid_candidate(
                image, current, maximum_cells=320
            )

        self.assertIsNotNone(recovered)
        self.assertEqual(recovered[0], columns)
        self.assertEqual(recovered[1], rows)

        dense_vertical = vertical.copy()
        for row in range(8, len(rows) - 1):
            for boundary in columns[2:-1]:
                cv2.line(
                    dense_vertical,
                    (boundary, rows[row]),
                    (boundary, rows[row + 1]),
                    255,
                    2,
                )
        with patch.object(
            table_pipeline,
            "_grid_maps",
            return_value=(horizontal, dense_vertical, cv2.bitwise_or(horizontal, dense_vertical)),
        ):
            self.assertIsNone(
                ocr_backend._raw_blank_form_grid_candidate(
                    image, current, maximum_cells=320
                )
            )

    def test_blank_form_title_uses_full_span_three_model_consensus(self):
        ocr_backend._load_runtime()
        image = np.full((120, 300, 3), 245, dtype=np.uint8)
        cv2.putText(image, "TITLE", (80, 42), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (20, 20, 20), 2)
        grid = [["误读标题", "", ""], ["项目", "编号", "确认"], ["", "", ""]]
        confidence = [[0.77, 0.0, 0.0], [0.99] * 3, [0.0] * 3]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 3, "role": "title"}
        ]
        output = SimpleNamespace(
            txts=["维修验收记录表"] * 3,
            scores=[0.99, 0.98, 0.97],
            imgs=None,
        )
        engine = SimpleNamespace(
            fast_text_rec=Mock(return_value=output),
            text_rec=Mock(return_value=output),
            server_text_rec=Mock(return_value=output),
        )

        scores = ocr_backend._recover_blank_form_title_multiview(
            image,
            [0, 100, 200, 300],
            [0, 60, 90, 120],
            grid,
            confidence,
            spans,
            engine,
        )

        self.assertEqual(grid[0][0], "维修验收记录表")
        self.assertEqual(confidence[0][0], 0.97)
        self.assertEqual(len(scores), 9)

    def test_blank_form_clears_model_fills_from_merged_subordinates(self):
        grid = [
            ["维修验收记录表", "误填标题", ""],
            ["项目", "编号", "确认"],
            ["开始时间", "误填内容", ""],
        ]
        confidence = [[0.77 if value else 0.0 for value in row] for row in grid]
        spans = [
            {"row": 0, "column": 0, "row_span": 1, "column_span": 3, "role": "title"},
            {"row": 2, "column": 0, "row_span": 1, "column_span": 3, "role": "merged"},
        ]

        cleared = ocr_backend._clear_blank_form_span_subordinates(
            grid, confidence, spans
        )

        self.assertEqual(cleared, {(0, 1), (2, 1)})
        self.assertEqual(grid[0], ["维修验收记录表", "", ""])
        self.assertEqual(grid[2], ["开始时间", "", ""])
        self.assertEqual(confidence[0][1:], [0.0, 0.0])
        self.assertEqual(confidence[2][1:], [0.0, 0.0])

    def test_spatial_title_requires_large_centered_text(self):
        ocr_backend._load_runtime()
        grid = [
            ["设备状态表", "", "", ""],
            ["编号", "设备号", "频率", "状态"],
            ["1", "AP-001", "562.990", "在用"],
            ["2", "AP-002", "563.360", "正常"],
            ["3", "AP-003", "563.730", "待复核"],
            ["4", "AP-004", "564.100", "已确认"],
        ]

        def entry(center_x: float, height: float, score: float = 0.99):
            return {"center_x": center_x, "height": height, "score": score}

        ordinary_row = [entry(value, 30.0) for value in (100, 300, 500, 700)]
        geometry = {
            "anchors": [100.0, 300.0, 500.0, 700.0],
            "grouped_rows": [
                [entry(400.0, 40.0)],
                ordinary_row,
                ordinary_row,
                ordinary_row,
                ordinary_row,
                ordinary_row,
            ],
        }
        self.assertTrue(ocr_backend._spatial_title_is_prominent(grid, geometry))

        small = copy.deepcopy(geometry)
        small["grouped_rows"][0][0]["height"] = 32.0
        self.assertFalse(ocr_backend._spatial_title_is_prominent(grid, small))

        off_center = copy.deepcopy(geometry)
        off_center["grouped_rows"][0][0]["center_x"] = 180.0
        self.assertFalse(ocr_backend._spatial_title_is_prominent(grid, off_center))

    def test_runtime_trace_converts_numpy_scalars_without_failing_recognition(self):
        ocr_backend._load_runtime()
        buffer = []
        with patch.object(ocr_backend, "_runtime_trace_active", True), patch.object(
            ocr_backend, "_runtime_trace_started", 0.0
        ), patch.object(ocr_backend, "_runtime_trace_buffer", buffer), patch.object(
            ocr_backend, "_runtime_trace_file", None
        ):
            ocr_backend._write_runtime_trace(
                "numpy_scalar",
                flag=np.bool_(True),
                count=np.int64(3),
            )

        payload = json.loads(buffer[0])
        self.assertIs(payload["flag"], True)
        self.assertEqual(payload["count"], 3)

    def test_third_grid_row_distinguishes_simple_and_multilevel_headers(self):
        ocr_backend._load_runtime()
        image = np.full((160, 700, 3), 245, dtype=np.uint8)
        columns = list(range(0, 701, 100))
        rows = [0, 40, 80, 120, 159]
        for column in columns[1:-1]:
            cv2.line(image, (column, 80), (column, 120), (30, 30, 30), 2)

        self.assertTrue(
            ocr_backend._third_grid_row_has_complete_internal_boundaries(
                image,
                columns,
                rows,
            )
        )
        image[:, 195:306] = 245
        self.assertFalse(
            ocr_backend._third_grid_row_has_complete_internal_boundaries(
                image,
                columns,
                rows,
            )
        )

    def test_perspective_spreadsheet_recovery_requires_rulers_and_marks_all_cells(self):
        ocr_backend._load_runtime()
        image = np.full((250, 700, 3), 245, dtype=np.uint8)
        cv2.line(image, (350, 0), (350, 249), (20, 20, 20), 2)
        horizontal = np.zeros((250, 700), dtype=np.uint8)
        row_boundaries = list(range(5, 246, 20))
        for row in row_boundaries:
            horizontal[row, :] = 255
        vertical = np.zeros_like(horizontal)
        raw_boundaries = [0, 50, 150, 250, 350, 450, 550, 650]
        for column in raw_boundaries:
            vertical[:, column] = 255

        spatial_grid = []
        boxes = []
        texts = []
        scores = []
        for row in range(12):
            if row == 0:
                values = ["1", "", "基础信息", "", "", "质量判定", ""]
            elif row == 1:
                values = ["2", "序号", "客户", "数量", "单价", "日期", "状态"]
            else:
                values = [str(row + 1)] + [f"R{row}C{column}" for column in range(6)]
            spatial_grid.append(values)
            top = row_boundaries[row]
            bottom = row_boundaries[row + 1]
            if row == 0:
                content = [(1, "基础信息"), (4, "质量判定")]
            else:
                content = list(enumerate(values[1:]))
            for column, value in content:
                left = raw_boundaries[column + 1]
                right = raw_boundaries[column + 2]
                boxes.append(
                    np.asarray(
                        [
                            [left + 8, top + 5],
                            [right - 8, top + 5],
                            [right - 8, bottom - 5],
                            [left + 8, bottom - 5],
                        ],
                        dtype=float,
                    )
                )
                texts.append(value)
                scores.append(0.99)

        geometry = {
            "anchors": [25, 100, 200, 300, 400, 500, 600],
            "row_centers": [15 + row * 20 for row in range(12)],
        }
        with (
            patch.object(
                table_pipeline,
                "_grid_maps",
                return_value=(horizontal, vertical, cv2.bitwise_or(horizontal, vertical)),
            ),
            patch.object(
                ocr_backend,
                "_spatial_geometry_cell_boundaries",
                return_value=(raw_boundaries, row_boundaries),
            ),
        ):
            recovered = ocr_backend._recover_perspective_spreadsheet_from_rulers(
                image,
                spatial_grid,
                geometry,
                boxes,
                texts,
                scores,
            )

        self.assertIsNotNone(recovered)
        grid, confidence, spans, metrics = recovered
        self.assertEqual((len(grid), len(grid[0])), (12, 6))
        self.assertEqual(
            [(span["column"], span["column_span"]) for span in spans],
            [(0, 3), (3, 3)],
        )
        self.assertEqual(metrics["ruler_rows"], 12)
        self.assertTrue(
            all(float(value) < 0.78 for row in confidence for value in row)
        )

    def test_strong_header_group_boundaries_choose_unique_best_physical_set(self):
        selected = ocr_backend._select_strong_header_group_boundaries(
            [1, 3, 6, 10],
            {1: 0.6875, 2: 1.0, 5: 1.0, 7: 0.625, 9: 1.0, 10: 0.75},
            11,
        )

        self.assertEqual(selected, [0, 2, 5, 9, 11])

        self.assertIsNone(
            ocr_backend._select_strong_header_group_boundaries(
                [1, 3],
                {2: 0.80, 3: 0.75},
                5,
            )
        )

    def test_spreadsheet_ruler_row_count_allows_one_bracketed_gap(self):
        grid = [["", "A"]]
        grid.extend([[str(value), f"行{value}"] for value in range(1, 34)])
        grid.append(["", "行34"])
        grid.extend([[str(value), f"行{value}"] for value in range(35, 43)])
        grid.append(["44 43", "尾部两行"])

        self.assertEqual(
            ocr_backend._spreadsheet_ruler_row_count_with_one_bracketed_gap(grid),
            44,
        )

        grid[20][0] = ""
        self.assertIsNone(
            ocr_backend._spreadsheet_ruler_row_count_with_one_bracketed_gap(grid)
        )

    def test_adjacent_monotonic_row_track_skips_local_outliers(self):
        ocr_backend._load_runtime()
        previous = np.asarray([20.0 + index * 30.0 for index in range(12)])
        observed = previous + np.linspace(0.5, 3.0, previous.size)
        observed = np.sort(np.r_[observed, 77.0, 201.0, 288.0])

        selected = ocr_backend._select_adjacent_monotonic_row_track(
            previous,
            observed,
        )

        self.assertIsNotNone(selected)
        self.assertEqual(selected.size, previous.size)
        self.assertLessEqual(float(np.max(np.abs(selected - previous))), 3.0)

    def test_group_header_boundary_requires_full_band_support_when_edges_are_faint(self):
        ocr_backend._load_runtime()
        gradient = np.zeros((24, 80), dtype=np.float32)
        gradient[:, 38:40] = 100.0
        gradient[5:-5, 40] = 100.0
        gradient[0:5, 58:60] = 100.0
        gradient[-5:, 58:60] = 100.0

        self.assertTrue(
            ocr_backend._group_header_boundary_has_edge_evidence(gradient, 40)
        )
        self.assertFalse(
            ocr_backend._group_header_boundary_has_edge_evidence(gradient, 60)
        )

    def test_slanted_dual_rulers_recover_full_photographed_grid(self):
        ocr_backend._load_runtime()
        image = np.full((360, 720, 3), 245, dtype=np.uint8)
        letter_centers = [100, 190, 280, 370, 460, 550]
        row_centers = [60 + row * 22 for row in range(12)]
        column_boundaries = [55, 145, 235, 325, 415, 505, 595]
        row_boundaries = [49 + row * 22 for row in range(13)]
        for row in row_boundaries:
            cv2.line(
                image,
                (40, row),
                (680, int(round(row + 0.01 * 640))),
                (45, 45, 45),
                2,
            )
        for column in column_boundaries:
            cv2.line(
                image,
                (column, 45),
                (int(round(column + 0.03 * 280)), 325),
                (45, 45, 45),
                2,
            )

        boxes = []
        texts = []
        scores = []

        def add_entry(center_x, center_y, text):
            boxes.append(
                np.asarray(
                    [
                        [center_x - 6, center_y - 7],
                        [center_x + 6, center_y - 7],
                        [center_x + 6, center_y + 7],
                        [center_x - 6, center_y + 7],
                    ],
                    dtype=float,
                )
            )
            texts.append(text)
            scores.append(0.99)

        for index, center_x in enumerate(letter_centers):
            add_entry(center_x, 25, chr(ord("A") + index))
        for index, center_y in enumerate(row_centers):
            add_entry(15, center_y, str(index + 1))
        for index, center_x in enumerate(letter_centers):
            add_entry(center_x, row_centers[0], f"字段{index + 1}")

        recovered = ocr_backend._recover_slanted_spreadsheet_grid_from_rulers(
            image,
            boxes,
            texts,
            scores,
        )

        self.assertIsNotNone(recovered)
        grid, confidence, spans, metrics = recovered
        self.assertEqual((len(grid), len(grid[0])), (12, 6))
        self.assertEqual(grid[0], [f"字段{index + 1}" for index in range(6)])
        self.assertEqual(spans, [])
        self.assertTrue(metrics["slanted_dual_ruler_recovery"])
        self.assertTrue(
            all(value <= 0.77 for row in confidence for value in row if value >= 0.0)
        )

        drifted_boxes = [np.asarray(box, dtype=float).copy() for box in boxes]
        for row_index in range(len(row_centers)):
            shift = 16.0 * row_index / float(len(row_centers) - 1)
            drifted_boxes[len(letter_centers) + row_index][:, 0] += shift
        drifted = ocr_backend._recover_slanted_spreadsheet_grid_from_rulers(
            image,
            drifted_boxes,
            texts,
            scores,
        )
        self.assertIsNotNone(drifted)
        self.assertEqual((len(drifted[0]), len(drifted[0][0])), (12, 6))

        retained = [
            index
            for index in range(len(texts))
            if index not in {5, 6}
        ]
        inferred = ocr_backend._recover_slanted_spreadsheet_grid_from_rulers(
            image,
            [boxes[index] for index in retained],
            [texts[index] for index in retained],
            [scores[index] for index in retained],
            expected_columns=6,
        )
        self.assertIsNotNone(inferred)
        self.assertEqual((len(inferred[0]), len(inferred[0][0])), (12, 6))

        texts[8] = "15"
        self.assertIsNone(
            ocr_backend._recover_slanted_spreadsheet_grid_from_rulers(
                image,
                boxes,
                texts,
                scores,
            )
        )

    def test_table_structure_cell_geometry_preserves_exact_cell_bounds(self):
        boxes = np.asarray(
            [
                [0, 0, 200, 0, 200, 30, 0, 30],
                [0, 30, 100, 30, 100, 60, 0, 60],
                [100, 30, 200, 30, 200, 60, 100, 60],
                [0, 60, 100, 60, 100, 90, 0, 90],
                [100, 60, 200, 60, 200, 90, 100, 90],
            ],
            dtype=float,
        )
        logic = np.asarray(
            [
                [0, 0, 0, 1],
                [1, 1, 0, 0],
                [1, 1, 1, 1],
                [2, 2, 0, 0],
                [2, 2, 1, 1],
            ],
            dtype=int,
        )

        with patch.object(ocr_backend, "np", np):
            geometry = ocr_backend._table_structure_cell_recovery_geometry(
                boxes,
                logic,
                3,
                2,
            )

        self.assertIsNotNone(geometry)
        self.assertEqual(geometry["first_structured_row"], 1)
        self.assertEqual(geometry["cell_bounds"][(1, 0)], (0, 100, 30, 60))
        self.assertEqual(geometry["cell_bounds"][(2, 1)], (100, 200, 60, 90))
        self.assertEqual(geometry["anchors"], [50.0, 150.0])
        self.assertEqual(geometry["row_centers"], [15.0, 45.0, 75.0])

    def test_table_structure_cell_geometry_rejects_shape_drift(self):
        boxes = np.asarray(
            [[0, 0, 100, 0, 100, 30, 0, 30]],
            dtype=float,
        )
        logic = np.asarray([[0, 0, 0, 2]], dtype=int)

        with patch.object(ocr_backend, "np", np):
            self.assertIsNone(
                ocr_backend._table_structure_cell_recovery_geometry(
                    boxes,
                    logic,
                    3,
                    2,
                )
            )

    def test_aligned_table_structure_geometry_accepts_unique_split_tail(self):
        final_grid = [
            ["标题", ""],
            ["编号", "数值"],
        ] + [[f"A{row}", f"V{row}"] for row in range(2, 10)]
        table_grid = [list(row) for row in final_grid[:-1]] + [
            ["A9", ""],
            ["", "V9"],
        ]
        boxes = []
        logic = []
        for row in range(len(table_grid)):
            for column in range(2):
                left = column * 100
                top = row * 30
                boxes.append(
                    [left, top, left + 100, top, left + 100, top + 30, left, top + 30]
                )
                logic.append([row, row, column, column])

        with patch.object(ocr_backend, "np", np):
            geometry = ocr_backend._aligned_table_structure_cell_recovery_geometry(
                table_grid,
                np.asarray(boxes, dtype=float),
                np.asarray(logic, dtype=int),
                final_grid,
            )

        self.assertIsNotNone(geometry)
        self.assertTrue(geometry["restrict_to_cell_bounds"])
        self.assertEqual(geometry["aligned_prefix_rows"], 9)
        self.assertIn((8, 1), geometry["cell_bounds"])
        self.assertNotIn((9, 0), geometry["cell_bounds"])

    def test_aligned_table_structure_geometry_rejects_ambiguous_rows(self):
        final_grid = [["重复", "值"] for _ in range(8)]
        table_grid = [["重复", "值"] for _ in range(9)]
        boxes = []
        logic = []
        for row in range(9):
            for column in range(2):
                left = column * 100
                top = row * 30
                boxes.append(
                    [left, top, left + 100, top, left + 100, top + 30, left, top + 30]
                )
                logic.append([row, row, column, column])

        with patch.object(ocr_backend, "np", np):
            self.assertIsNone(
                ocr_backend._aligned_table_structure_cell_recovery_geometry(
                    table_grid,
                    np.asarray(boxes, dtype=float),
                    np.asarray(logic, dtype=int),
                    final_grid,
                )
            )


if __name__ == "__main__":
    unittest.main()
