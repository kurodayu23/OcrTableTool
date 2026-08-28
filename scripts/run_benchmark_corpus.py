from __future__ import annotations

import argparse
import csv
import gc
import json
import os
import re
import sys
import threading
import time
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "backend"))

import ocr_backend  # noqa: E402


def working_set_bytes() -> int:
    try:
        import psutil

        return int(psutil.Process().memory_info().rss)
    except ImportError:
        pass
    if os.name != "nt":
        return 0
    import ctypes
    from ctypes import wintypes

    class ProcessMemoryCounters(ctypes.Structure):
        _fields_ = [
            ("cb", wintypes.DWORD),
            ("PageFaultCount", wintypes.DWORD),
            ("PeakWorkingSetSize", ctypes.c_size_t),
            ("WorkingSetSize", ctypes.c_size_t),
            ("QuotaPeakPagedPoolUsage", ctypes.c_size_t),
            ("QuotaPagedPoolUsage", ctypes.c_size_t),
            ("QuotaPeakNonPagedPoolUsage", ctypes.c_size_t),
            ("QuotaNonPagedPoolUsage", ctypes.c_size_t),
            ("PagefileUsage", ctypes.c_size_t),
            ("PeakPagefileUsage", ctypes.c_size_t),
        ]

    counters = ProcessMemoryCounters()
    counters.cb = ctypes.sizeof(counters)
    handle = ctypes.windll.kernel32.GetCurrentProcess()
    if not ctypes.windll.psapi.GetProcessMemoryInfo(handle, ctypes.byref(counters), counters.cb):
        return 0
    return int(counters.WorkingSetSize)


class MemorySampler:
    def __init__(self, interval: float = 0.05) -> None:
        self.interval = interval
        self.peak = 0
        self._stop = threading.Event()
        self._thread = threading.Thread(target=self._sample, daemon=True)

    def _sample(self) -> None:
        while not self._stop.wait(self.interval):
            self.peak = max(self.peak, working_set_bytes())

    def __enter__(self) -> "MemorySampler":
        self.peak = working_set_bytes()
        self._thread.start()
        return self

    def __exit__(self, exc_type, exc, traceback) -> None:
        self.peak = max(self.peak, working_set_bytes())
        self._stop.set()
        self._thread.join(timeout=1.0)


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", "", value).casefold()


def collect_strings(value: Any) -> list[str]:
    if isinstance(value, str):
        return [value] if value.strip() else []
    if isinstance(value, dict):
        for key in ("text", "transcription"):
            direct = value.get(key)
            if isinstance(direct, str) and direct.strip():
                return [direct]
        strings: list[str] = []
        for key, item in value.items():
            if key == "value":
                strings.extend(collect_strings(item))
            elif key in {"form", "document", "words", "cells", "gt_parse", "valid_line", "ground_truth"}:
                strings.extend(collect_strings(item))
        return strings
    if isinstance(value, list):
        strings: list[str] = []
        for item in value:
            strings.extend(collect_strings(item))
        return strings
    return []


def collect_semantic_strings(value: Any) -> list[str]:
    if isinstance(value, str):
        return [value] if value.strip() else []
    if isinstance(value, dict):
        return [
            text
            for item in value.values()
            for text in collect_semantic_strings(item)
        ]
    if isinstance(value, list):
        return [text for item in value for text in collect_semantic_strings(item)]
    return []


def numeric_tokens(strings: list[str]) -> set[str]:
    pattern = re.compile(r"(?<![\w.])[+\-]?\d+(?:[.,]\d+)*(?:%|[A-Za-z]+)?")
    return {match.group(0) for value in strings for match in pattern.finditer(value)}


def score_text(ground_truth: Any, recognized: list[str], metric_scope: str) -> dict[str, Any]:
    if metric_scope in {"table_presence+stability", "layout_presence+stability"}:
        return {
            "ground_truth_text_items": 0,
            "text_item_coverage": None,
            "sequence_similarity": None,
            "numeric_exact_rate": None,
        }
    scoring_truth = ground_truth
    if metric_scope == "ocr_text+semantic_fields" and isinstance(ground_truth, dict):
        payload = ground_truth.get("ground_truth")
        if isinstance(payload, dict) and isinstance(payload.get("gt_parse"), dict):
            scoring_truth = payload["gt_parse"]
    expected = (
        collect_semantic_strings(scoring_truth)
        if metric_scope == "ocr_text+semantic_fields"
        else collect_strings(scoring_truth)
    )
    recognized_joined = normalize_text(" ".join(recognized))
    normalized_expected = [normalize_text(value) for value in expected if normalize_text(value)]
    matched = sum(value in recognized_joined for value in normalized_expected)
    expected_joined = normalize_text(" ".join(expected))
    expected_numbers = numeric_tokens(expected)
    recognized_numbers = numeric_tokens(recognized)
    return {
        "ground_truth_text_items": len(normalized_expected),
        "text_item_coverage": matched / len(normalized_expected) if normalized_expected else None,
        "sequence_similarity": SequenceMatcher(None, expected_joined, recognized_joined).ratio()
        if expected_joined
        else None,
        "numeric_exact_rate": len(expected_numbers & recognized_numbers) / len(expected_numbers)
        if expected_numbers
        else None,
        "numeric_precision": len(expected_numbers & recognized_numbers) / len(recognized_numbers)
        if recognized_numbers
        else None,
        "numeric_set_exact": expected_numbers == recognized_numbers,
        "expected_numeric_count": len(expected_numbers),
        "recognized_numeric_count": len(recognized_numbers),
    }


def cell_text_grid(result: dict[str, Any]) -> list[list[str]]:
    return [
        [str(cell.get("text", "")) if isinstance(cell, dict) else str(cell) for cell in row]
        for row in result.get("cells", [])
    ]


def export_and_verify(case_output: Path, grid: list[list[str]], spans: list[dict[str, int]]) -> dict[str, Any]:
    from openpyxl import load_workbook

    csv_path = case_output / "result.csv"
    xlsx_path = case_output / "result.xlsx"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as stream:
        csv.writer(stream).writerows(grid)
    ocr_backend.pipeline.write_xlsx(xlsx_path, grid, spans)

    with csv_path.open("r", encoding="utf-8-sig", newline="") as stream:
        csv_roundtrip = list(csv.reader(stream))
    workbook = load_workbook(xlsx_path, read_only=False, data_only=False)
    sheet = workbook.active
    row_count = len(grid)
    column_count = max((len(row) for row in grid), default=0)
    xlsx_roundtrip = [
        ["" if sheet.cell(row, column).value is None else str(sheet.cell(row, column).value) for column in range(1, column_count + 1)]
        for row in range(1, row_count + 1)
    ]
    rectangular_grid = [row + [""] * (column_count - len(row)) for row in grid]
    return {
        "csv_ok": csv_roundtrip == rectangular_grid,
        "xlsx_ok": xlsx_roundtrip == rectangular_grid,
        "csv_path": str(csv_path),
        "xlsx_path": str(xlsx_path),
    }


def load_cases(corpus: Path, args: argparse.Namespace) -> list[dict[str, str]]:
    with (corpus / "manifest.csv").open("r", encoding="utf-8-sig", newline="") as stream:
        cases = list(csv.DictReader(stream))
    requested_ids = {value.strip() for value in args.ids.split(",") if value.strip()} if args.ids else set()
    if requested_ids:
        cases = [case for case in cases if case["id"] in requested_ids]
    if args.source:
        cases = [case for case in cases if case.get("source", "local") == args.source]
    if args.split:
        cases = [case for case in cases if case.get("benchmark_split", "development") == args.split]
    if args.limit:
        cases = cases[: args.limit]
    return cases


def write_summary(output: Path, records: list[dict[str, Any]]) -> None:
    output.mkdir(parents=True, exist_ok=True)
    (output / "summary.json").write_text(
        json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    scalar_keys = sorted(
        {
            key
            for record in records
            for key, value in record.items()
            if not isinstance(value, (dict, list))
        }
    )
    with (output / "summary.csv").open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=scalar_keys, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="顺序运行 OCR 120 张本地回归集")
    parser.add_argument("--corpus", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--model-dir", type=Path)
    parser.add_argument("--ids", default="")
    parser.add_argument("--source", default="")
    parser.add_argument("--split", choices=("development", "holdout"))
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--skip-export", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    corpus = args.corpus.resolve()
    output = args.output.resolve()
    if output.exists() and any(output.iterdir()):
        raise RuntimeError(f"输出目录非空，为避免覆盖历史结果已停止：{output}")
    output.mkdir(parents=True, exist_ok=True)
    if args.model_dir:
        os.environ["OCR_TABLE_MODEL_DIR"] = str(args.model_dir.resolve())

    cases = load_cases(corpus, args)
    if not cases:
        raise RuntimeError("没有符合筛选条件的基准样本")
    ocr_backend._load_runtime()
    records: list[dict[str, Any]] = []
    for position, case in enumerate(cases, start=1):
        case_output = output / case["id"]
        case_output.mkdir(parents=True, exist_ok=True)
        image_path = corpus / case["image_path"]
        ground_truth_wrapper = json.loads((corpus / case["ground_truth_path"]).read_text(encoding="utf-8"))
        started = time.perf_counter()
        metric_scope = case.get("metric_scope", "ocr_text")
        source = case.get("source", "local")
        if source == "DocLayNet-dev":
            metric_scope = "layout_presence+stability"
        record: dict[str, Any] = {
            "id": case["id"],
            "source": source,
            "benchmark_split": case.get("benchmark_split", "development"),
            "expected_mode": case.get("expected_mode", "grid_table"),
            "metric_scope": metric_scope,
            "status": "error",
        }
        try:
            with MemorySampler() as memory:
                result = ocr_backend.handle_request(
                    {
                        "protocol": 1,
                        "action": "recognize",
                        "image_path": str(image_path),
                        "output_directory": str(case_output),
                        "options": {"crop_mode": "auto"},
                    }
                )
                grid = cell_text_grid(result)
                recognized = [value for row in grid for value in row if value.strip()]
                export = (
                    {"csv_ok": None, "xlsx_ok": None}
                    if args.skip_export
                    else export_and_verify(case_output, grid, result.get("spans", []))
                )
            (case_output / "result.json").write_text(
                json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            record.update(
                {
                    "status": "ok",
                    "engine": result.get("engine", ""),
                    "document_mode": result.get("document_mode", ""),
                    "rows": result.get("rows", 0),
                    "columns": result.get("columns", 0),
                    "nonempty_cells": len(recognized),
                    "review_cells": sum(
                        bool(cell.get("needs_review", False))
                        for row in result.get("cells", [])
                        for cell in row
                        if isinstance(cell, dict)
                    ),
                    "withheld_numeric_segments": result.get("withheld_numeric_segments", 0),
                    "peak_working_set_mb": round(memory.peak / 1024 / 1024, 1),
                    **score_text(ground_truth_wrapper, recognized, metric_scope),
                    **export,
                }
            )
        except Exception as error:
            record["error"] = str(error)
            record["peak_working_set_mb"] = round(working_set_bytes() / 1024 / 1024, 1)
        record["elapsed_seconds"] = round(time.perf_counter() - started, 3)
        records.append(record)
        write_summary(output, records)
        print(
            f"[{position}/{len(cases)}] {case['id']} {record['status']} "
            f"{record['elapsed_seconds']:.3f}s {record.get('rows', 0)}x{record.get('columns', 0)} "
            f"{record['peak_working_set_mb']:.1f}MB",
            flush=True,
        )
        gc.collect()

    succeeded = sum(record["status"] == "ok" for record in records)
    print(json.dumps({"status": "complete", "cases": len(records), "succeeded": succeeded}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
